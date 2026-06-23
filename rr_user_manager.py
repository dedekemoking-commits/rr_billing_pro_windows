"""
rr_user_manager.py — Tool Kelola & Cek Username RR BILLING PRO
===============================================================
Double-click via BUKA_USER_MANAGER.bat untuk membuka.
Hanya untuk developer/admin.

Fitur:
- Lihat semua user terdaftar
- Deteksi username duplikat
- Hapus / rename username
- Reset password user
- Lihat detail profil rental
- Export daftar user ke Excel
"""

import os
import json
import hashlib
import tkinter as tk
from tkinter import ttk, messagebox
import customtkinter as ctk
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─── TEMA ────────────────────────────────────────────────────────────────────
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

C_BG     = "#001A2E"  # Neon Sky Blue background
C_PANEL  = "#003d5c"
C_CARD   = "#00527a"
C_ACCENT = "#00E5FF"  # Neon Cyan
C_ACCENT2= "#00BFFF"  # Deep Sky Blue
C_RED    = "#FF3366"
C_GREEN  = "#39FF14"
C_YELLOW = "#FFD700"
C_TEXT   = "#E0FFFF"
C_MUTED  = "#7EB3D6"
C_BTN    = "#0A2A42"
C_BORDER = "#1A5F7A"
C_ORANGE = "#FF8C00"

FONT_TITLE = ("Russo One", 15, "bold")
FONT_SUB   = ("Russo One", 10, "bold")
FONT_BODY  = ("Courier New", 10)
FONT_SMALL = ("Courier New", 9)
FONT_LABEL = ("Consolas", 10)

CONFIG_FILE = "rr_billing_config.json"
KEYGEN_LOG  = "rr_keygen_log.json"

_DEV_PASSWORD_HASH = hashlib.sha256("rrcctv2026".encode()).hexdigest()

DEFAULT_USERS = {"admin", "kasir"}  # user bawaan sistem


# ═══════════════════════════════════════════════════════════════════════════════
#  HELPER
# ═══════════════════════════════════════════════════════════════════════════════
def load_config() -> dict:
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def save_config(data: dict):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

def load_keygen_log() -> list:
    if os.path.exists(KEYGEN_LOG):
        try:
            with open(KEYGEN_LOG, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return []

def save_keygen_log(data: list):
    with open(KEYGEN_LOG, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


# ═══════════════════════════════════════════════════════════════════════════════
#  LOGIN DEVELOPER
# ═══════════════════════════════════════════════════════════════════════════════
class LoginDeveloper(ctk.CTkToplevel):
    def __init__(self, master, on_success):
        super().__init__(master)
        self.title("🔐 Developer Access")
        self.geometry("400x300")
        self.configure(fg_color=C_BG)
        self.grab_set()
        self.resizable(False, False)
        self.on_success = on_success
        self._attempt  = 0
        self._build()

    def _build(self):
        ctk.CTkLabel(self, text="🔐  DEVELOPER ACCESS",
                     font=FONT_TITLE, text_color=C_RED).pack(pady=(28, 4))
        ctk.CTkLabel(self, text="Password developer diperlukan untuk melanjutkan",
                     font=FONT_SMALL, text_color=C_MUTED).pack(pady=(0, 18))

        panel = ctk.CTkFrame(self, fg_color=C_PANEL, corner_radius=14)
        panel.pack(padx=30, fill="x")

        ctk.CTkLabel(panel, text="Password:", font=FONT_LABEL,
                     text_color=C_MUTED, anchor="w").pack(anchor="w", padx=20, pady=(16, 4))
        self.entry = ctk.CTkEntry(panel, show="●", height=40,
                                   fg_color=C_BTN, text_color=C_ACCENT,
                                   border_color=C_BORDER,
                                   font=("Consolas", 13, "bold"))
        self.entry.pack(fill="x", padx=20, pady=(0, 8))
        self.entry.bind("<Return>", lambda e: self._cek())
        self.entry.focus()

        self.lbl_err = ctk.CTkLabel(panel, text="", font=FONT_LABEL,
                                     text_color=C_RED)
        self.lbl_err.pack(pady=(0, 6))

        ctk.CTkButton(panel, text="🔓  Masuk", height=38,
                      fg_color=C_RED, hover_color="#CC0033",
                      font=FONT_SUB, text_color="white",
                      command=self._cek).pack(fill="x", padx=20, pady=(0, 16))

        ctk.CTkLabel(self, text="Default: rrcctv2026",
                     font=FONT_SMALL, text_color=C_MUTED).pack(pady=8)

    def _cek(self):
        pw = self.entry.get().strip()
        if hashlib.sha256(pw.encode()).hexdigest() == _DEV_PASSWORD_HASH:
            self.destroy()
            self.on_success()
        else:
            self._attempt += 1
            self.lbl_err.configure(
                text=f"✖  Password salah ({self._attempt}/3)")
            self.entry.delete(0, "end")
            if self._attempt >= 3:
                self.destroy()


# ═══════════════════════════════════════════════════════════════════════════════
#  DIALOG DETAIL USER
# ═══════════════════════════════════════════════════════════════════════════════
class DialogDetailUser(ctk.CTkToplevel):
    """
    Menampilkan detail profil rental satu user, dan menyediakan aksi:
    - Ganti username (dengan invalidasi lisensi lama yang terikat username lama)
    - (bisa dikembangkan lagi: edit profil, dsb)
    """
    def __init__(self, master, username, user_data, profil_data, on_update):
        super().__init__(master)
        self.title(f"🔍 Detail User — {username}")
        self.geometry("480x520")
        self.configure(fg_color=C_BG)
        self.grab_set()
        self.resizable(False, False)

        self.username    = username
        self.user_data   = user_data
        self.profil_data = profil_data
        self.on_update    = on_update

        self._build()

    def _build(self):
        ctk.CTkLabel(self, text=f"👤  {self.username}",
                     font=FONT_TITLE, text_color=C_ACCENT).pack(pady=(20, 2))
        ctk.CTkLabel(self, text=f"Role: {self.user_data.get('role', 'kasir').upper()}",
                     font=FONT_SMALL, text_color=C_MUTED).pack(pady=(0, 12))

        scroll = ctk.CTkScrollableFrame(self, fg_color=C_BG, height=300)
        scroll.pack(fill="both", expand=True, padx=20)

        # ── Profil Rental (read-only) ────────────────────────────────────────
        profil_f = ctk.CTkFrame(scroll, fg_color=C_PANEL, corner_radius=10)
        profil_f.pack(fill="x", pady=6)
        ctk.CTkLabel(profil_f, text="🏢  PROFIL RENTAL", font=FONT_SUB,
                     text_color=C_ACCENT2).pack(anchor="w", padx=14, pady=(10, 6))

        for label, key in [
            ("Nama Pemilik", "nama_pemilik"),
            ("Nama Rental",  "nama_rental"),
            ("Email",        "email"),
            ("No HP",        "no_hp"),
            ("Alamat",       "alamat"),
            ("Tgl Daftar",   "tanggal_daftar"),
        ]:
            val = self.profil_data.get(key, "—")
            row = ctk.CTkFrame(profil_f, fg_color="transparent")
            row.pack(fill="x", padx=14, pady=2)
            ctk.CTkLabel(row, text=f"{label}:", font=FONT_LABEL,
                         text_color=C_MUTED, width=110, anchor="w").pack(side="left")
            ctk.CTkLabel(row, text=str(val), font=FONT_BODY,
                         text_color=C_TEXT, wraplength=260, justify="left",
                         anchor="w").pack(side="left", fill="x", expand=True)
        ctk.CTkFrame(profil_f, fg_color="transparent", height=8).pack()

        # ── Ganti Username ────────────────────────────────────────────────────
        gu_f = ctk.CTkFrame(scroll, fg_color=C_PANEL, corner_radius=10,
                             border_width=1, border_color=C_YELLOW)
        gu_f.pack(fill="x", pady=6)
        ctk.CTkLabel(gu_f, text="✏️  GANTI USERNAME", font=FONT_SUB,
                     text_color=C_YELLOW).pack(anchor="w", padx=14, pady=(10, 4))
        ctk.CTkLabel(gu_f,
                     text="⚠ Mengganti username akan membuat kode aktivasi\n"
                          "lama (jika terikat username) tidak valid lagi.",
                     font=FONT_SMALL, text_color=C_MUTED,
                     justify="left").pack(anchor="w", padx=14, pady=(0, 6))

        self.entry_username_baru = ctk.CTkEntry(
            gu_f, placeholder_text="Username baru",
            fg_color=C_BTN, text_color=C_ACCENT,
            border_color=C_BORDER, font=FONT_BODY, height=34)
        self.entry_username_baru.pack(fill="x", padx=14, pady=(0, 8))

        ctk.CTkButton(gu_f, text="✏️  Ganti Username", height=32,
                      fg_color=C_YELLOW, hover_color="#CCAA00",
                      font=FONT_SUB, text_color=C_BG,
                      command=self._ganti_username).pack(fill="x", padx=14, pady=(0, 12))

        self.lbl_status = ctk.CTkLabel(self, text="", font=FONT_SMALL,
                                        text_color=C_GREEN, wraplength=420,
                                        justify="left")
        self.lbl_status.pack(pady=(4, 10), padx=20)

        ctk.CTkButton(self, text="✖  Tutup", height=34, width=140,
                      fg_color=C_BTN, border_width=1, border_color=C_MUTED,
                      font=FONT_SMALL, text_color=C_MUTED,
                      command=self.destroy).pack(pady=(0, 16))

    def _ganti_username(self):
        baru = self.entry_username_baru.get().strip().lower()
        if not baru:
            self.lbl_status.configure(
                text="⚠  Username baru tidak boleh kosong.", text_color=C_YELLOW)
            return
        if " " in baru:
            self.lbl_status.configure(
                text="⚠  Username tidak boleh mengandung spasi.", text_color=C_YELLOW)
            return
        if baru == self.username:
            self.lbl_status.configure(
                text="⚠  Username sama dengan yang lama.", text_color=C_YELLOW)
            return

        cfg   = load_config()
        users = cfg.get("users", {})
        if baru in users:
            self.lbl_status.configure(
                text=f"✖  Username '{baru}' sudah dipakai user lain!",
                text_color=C_RED)
            return

        konfirm = messagebox.askyesno(
            "✏️ Konfirmasi Ganti Username",
            f"Ganti username '{self.username}' → '{baru}'?\n\n"
            f"⚠  PERHATIAN:\n"
            f"Kode aktivasi lama yang terikat username '{self.username}'\n"
            f"akan TIDAK VALID. Setelah ini:\n"
            f"  1. Buka Keygen Tool\n"
            f"  2. Cari username '{baru}' (baru)\n"
            f"  3. Generate kode aktivasi baru\n"
            f"  4. Kirim ke user")
        if not konfirm:
            return

        # Update users dict
        users[baru] = users.pop(self.username)
        cfg["users"] = users

        # Update profil_rental
        profil_semua = cfg.get("profil_rental", {})
        if self.username in profil_semua:
            profil_semua[baru] = profil_semua.pop(self.username)
        cfg["profil_rental"] = profil_semua

        # Update license.json jika binding ke username lama
        import json as _json
        LICENSE_FILE = "rr_billing_license.json"
        if os.path.exists(LICENSE_FILE):
            try:
                with open(LICENSE_FILE, "r", encoding="utf-8") as f:
                    lic = _json.load(f)
                if lic.get("username") == self.username:
                    # Invalidasi lisensi lama — user harus aktivasi ulang
                    lic["aktif"]           = False
                    lic["username"]        = baru
                    lic["kode_aktivasi"]   = ""
                    lic["invalidasi_info"] = (
                        f"Username diganti dari '{self.username}' ke '{baru}' "
                        f"pada {datetime.now().strftime('%Y-%m-%d %H:%M')}. "
                        f"Generate kode aktivasi baru."
                    )
                    with open(LICENSE_FILE, "w", encoding="utf-8") as f:
                        _json.dump(lic, f, indent=2, ensure_ascii=False)
            except Exception:
                pass

        # Update keygen log — tandai kode lama invalid, catat username baru
        log = load_keygen_log()
        for entry in log:
            if entry.get("username") == self.username:
                entry["kode_invalid"]   = True
                entry["username_lama"]  = self.username
                entry["username_baru"]  = baru
                entry["catatan"]        = (
                    f"Username diganti ke '{baru}' — generate kode baru!"
                )
        save_keygen_log(log)

        save_config(cfg)

        self.lbl_status.configure(
            text=f"✅  Username berhasil diganti ke '{baru}'.\n"
                 f"⚠  Buka Keygen Tool → pilih '{baru}' → generate kode baru.",
            text_color=C_GREEN)
        self.username = baru
        self.on_update()


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN APP
# ═══════════════════════════════════════════════════════════════════════════════
class UserManagerApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("👤 RR BILLING PRO — User Manager")
        self.geometry("1050x680")
        self.configure(fg_color=C_BG)
        self.resizable(True, True)
        self._authed = False
        self.after(200, self._minta_login)

    def _minta_login(self):
        LoginDeveloper(self, on_success=self._build_ui)

    # ── BUILD UI ──────────────────────────────────────────────────────────────
    def _build_ui(self):
        self._authed = True

        # Header
        hdr = ctk.CTkFrame(self, fg_color=C_PANEL, height=58, corner_radius=0)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        ctk.CTkLabel(hdr, text="👤  USER MANAGER — RR BILLING PRO",
                     font=FONT_TITLE, text_color=C_ACCENT).pack(side="left", padx=20, pady=14)
        ctk.CTkLabel(hdr, text="⚠  Hanya untuk Developer/Admin",
                     font=FONT_SMALL, text_color=C_YELLOW).pack(side="right", padx=20)

        # Stat bar
        self.stat_bar = ctk.CTkFrame(self, fg_color=C_CARD, height=52, corner_radius=0)
        self.stat_bar.pack(fill="x")
        self.stat_bar.pack_propagate(False)
        self._build_stat_bar()

        # Body
        body = ctk.CTkFrame(self, fg_color="transparent")
        body.pack(fill="both", expand=True, padx=12, pady=10)
        body.columnconfigure(0, weight=3)
        body.columnconfigure(1, weight=2)
        body.rowconfigure(0, weight=1)

        self._build_tabel(body)
        self._build_panel_kanan(body)

        self._load_users()

    def _build_stat_bar(self):
        for w in self.stat_bar.winfo_children():
            w.destroy()

        cfg   = load_config()
        users = cfg.get("users", {})
        profil_semua = cfg.get("profil_rental", {})

        # Deteksi duplikat: username yang sama dihitung lebih dari 1
        # (seharusnya tidak mungkin di dict, tapi cek nama rental duplikat)
        nama_rental_list = [
            p.get("nama_rental", "").strip().lower()
            for p in profil_semua.values()
            if p.get("nama_rental")
        ]
        duplikat_rental = {n for n in nama_rental_list if nama_rental_list.count(n) > 1}

        total   = len(users)
        admin_c = sum(1 for u in users.values() if u.get("role") == "admin")
        kasir_c = sum(1 for u in users.values() if u.get("role") == "kasir")
        rental_c= len(profil_semua)
        dup_c   = len(duplikat_rental)

        stats = [
            ("Total User",       str(total),    C_ACCENT),
            ("Admin",            str(admin_c),  C_GREEN),
            ("Kasir",            str(kasir_c),  C_YELLOW),
            ("Punya Profil",     str(rental_c), C_ACCENT2),
            ("⚠ Nama Rental Sama", str(dup_c) if dup_c else "✅ 0", C_RED if dup_c else C_GREEN),
        ]
        for label, val, color in stats:
            f = ctk.CTkFrame(self.stat_bar, fg_color="transparent")
            f.pack(side="left", padx=18, pady=6)
            ctk.CTkLabel(f, text=val, font=("Russo One", 18, "bold"),
                         text_color=color).pack()
            ctk.CTkLabel(f, text=label, font=FONT_SMALL,
                         text_color=C_MUTED).pack()

        sep = ctk.CTkFrame(self.stat_bar, width=1, fg_color=C_BORDER)
        sep.pack(side="left", fill="y", padx=4, pady=8)

        # Tombol di stat bar
        ctk.CTkButton(self.stat_bar, text="🔄 Refresh", width=90, height=32,
                      fg_color=C_BTN, border_width=1, border_color=C_ACCENT,
                      font=FONT_SMALL, text_color=C_ACCENT,
                      command=self._load_users).pack(side="right", padx=8, pady=10)
        ctk.CTkButton(self.stat_bar, text="📥 Export Excel", width=120, height=32,
                      fg_color=C_BTN, border_width=1, border_color=C_GREEN,
                      font=FONT_SMALL, text_color=C_GREEN,
                      command=self._export_excel).pack(side="right", padx=4, pady=10)

    def _build_tabel(self, parent):
        left = ctk.CTkFrame(parent, fg_color=C_BG, corner_radius=0)
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 8))

        # Toolbar pencarian
        toolbar = ctk.CTkFrame(left, fg_color=C_PANEL, corner_radius=10)
        toolbar.pack(fill="x", pady=(0, 8))

        ctk.CTkLabel(toolbar, text="🔍 Cari:", font=FONT_LABEL,
                     text_color=C_MUTED).pack(side="left", padx=(12, 4), pady=10)
        self.var_cari = ctk.StringVar()
        self.var_cari.trace_add("write", lambda *a: self._filter_tabel())
        ctk.CTkEntry(toolbar, textvariable=self.var_cari,
                     placeholder_text="Ketik username / nama rental / email...",
                     fg_color=C_BTN, text_color=C_ACCENT,
                     border_color=C_BORDER, font=FONT_BODY,
                     height=32, width=260).pack(side="left", padx=4)

        ctk.CTkLabel(toolbar, text="Filter:", font=FONT_LABEL,
                     text_color=C_MUTED).pack(side="left", padx=(12, 4))
        self.var_filter = ctk.StringVar(value="Semua")
        ctk.CTkOptionMenu(toolbar,
                          values=["Semua", "Admin", "Kasir",
                                  "Punya Profil", "Tanpa Profil",
                                  "⚠ Nama Rental Sama"],
                          variable=self.var_filter,
                          fg_color=C_BTN, button_color=C_ACCENT2,
                          button_hover_color="#5A0FCC",
                          text_color=C_TEXT, font=FONT_BODY,
                          dropdown_fg_color=C_CARD,
                          dropdown_text_color=C_TEXT,
                          width=170,
                          command=lambda v: self._filter_tabel()
                          ).pack(side="left", padx=4)

        # Treeview
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("UM.Treeview",
                        background=C_CARD, fieldbackground=C_CARD,
                        foreground=C_TEXT, rowheight=30,
                        font=("Consolas", 10))
        style.configure("UM.Treeview.Heading",
                        background=C_PANEL, foreground=C_ACCENT,
                        font=("Russo One", 9, "bold"), relief="flat")
        style.map("UM.Treeview", background=[("selected", C_ACCENT2)])

        cols = ("", "Username", "Role", "Nama Rental", "Email", "No HP", "Tgl Daftar")
        self.tree = ttk.Treeview(left, columns=cols, show="headings",
                                  style="UM.Treeview")
        widths = [28, 120, 70, 170, 170, 120, 110]
        anchors = ["center","w","center","w","w","center","center"]
        for col, w, anc in zip(cols, widths, anchors):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=w, anchor=anc, minwidth=w)

        # Scrollbar
        sb_y = ttk.Scrollbar(left, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb_y.set)

        tree_frame = ctk.CTkFrame(left, fg_color="transparent")
        tree_frame.pack(fill="both", expand=True)
        self.tree.pack(side="left", fill="both", expand=True, in_=tree_frame)
        sb_y.pack(side="right", fill="y", in_=tree_frame)

        self.tree.tag_configure("duplikat",  foreground=C_RED)
        self.tree.tag_configure("default",   foreground=C_MUTED)
        self.tree.tag_configure("admin",     foreground=C_GREEN)
        self.tree.tag_configure("kasir",     foreground=C_YELLOW)
        self.tree.tag_configure("tanpa_profil", foreground=C_ORANGE)

        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.tree.bind("<Double-Button-1>", lambda e: self._buka_detail())

        # Bottom action bar
        bot = ctk.CTkFrame(left, fg_color=C_PANEL, corner_radius=10)
        bot.pack(fill="x", pady=(8, 0))

        for txt, cmd, color in [
            ("🔍 Detail / Edit",    self._buka_detail,   C_ACCENT),
            ("🗑 Hapus User",        self._hapus_user,    C_RED),
            ("🔒 Reset Password",   self._reset_pass_cepat, C_GREEN),
            ("🔑 Buka Keygen",      self._buka_keygen,   C_YELLOW),
        ]:
            ctk.CTkButton(bot, text=txt, height=32,
                          fg_color=C_BTN, border_width=1, border_color=color,
                          font=FONT_SMALL, text_color=color,
                          command=cmd).pack(side="left", padx=6, pady=8)

    def _build_panel_kanan(self, parent):
        right = ctk.CTkFrame(parent, fg_color=C_BG, corner_radius=0)
        right.grid(row=0, column=1, sticky="nsew")

        # ── Cek Duplikat ──────────────────────────────────────────────────────
        dup_f = ctk.CTkFrame(right, fg_color=C_PANEL, corner_radius=12,
                              border_width=1, border_color=C_RED)
        dup_f.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(dup_f, text="⚠  CEK DUPLIKAT & KONFLIK",
                     font=FONT_SUB, text_color=C_RED).pack(anchor="w", padx=14, pady=(12, 6))
        ctk.CTkLabel(dup_f,
                     text="Deteksi otomatis:\n"
                          "• Username ganda (tidak mungkin di dict, tapi dicek)\n"
                          "• Nama rental yang sama (2 user pakai nama rental sama)\n"
                          "• Email yang sama\n"
                          "• No HP yang sama",
                     font=FONT_SMALL, text_color=C_MUTED,
                     justify="left").pack(anchor="w", padx=14, pady=(0, 8))

        ctk.CTkButton(dup_f, text="🔍  Jalankan Cek Duplikat", height=34,
                      fg_color="#3A0000", hover_color="#2A0000",
                      border_width=1, border_color=C_RED,
                      font=FONT_SUB, text_color=C_RED,
                      command=self._cek_duplikat).pack(fill="x", padx=14, pady=(0, 12))

        self.txt_duplikat = ctk.CTkTextbox(dup_f, height=160,
                                            fg_color=C_BTN, text_color=C_TEXT,
                                            border_color=C_BORDER, border_width=1,
                                            font=("Consolas", 9))
        self.txt_duplikat.pack(fill="x", padx=14, pady=(0, 12))
        self.txt_duplikat.configure(state="disabled")

        # ── Tambah User Manual ────────────────────────────────────────────────
        add_f = ctk.CTkFrame(right, fg_color=C_PANEL, corner_radius=12)
        add_f.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(add_f, text="➕  TAMBAH USER BARU",
                     font=FONT_SUB, text_color=C_ACCENT2).pack(anchor="w", padx=14, pady=(12, 6))

        for attr, label, ph, show in [
            ("_e_add_user", "Username",   "mis. rentalku01",  None),
            ("_e_add_pass", "Password",   "min 6 karakter",   "●"),
        ]:
            ctk.CTkLabel(add_f, text=label, font=FONT_LABEL,
                         text_color=C_MUTED, anchor="w").pack(anchor="w", padx=14, pady=(2, 1))
            kw = {"show": show} if show else {}
            e = ctk.CTkEntry(add_f, placeholder_text=ph,
                              fg_color=C_BTN, text_color=C_ACCENT,
                              border_color=C_BORDER, font=FONT_BODY, height=32, **kw)
            e.pack(fill="x", padx=14, pady=(0, 4))
            setattr(self, attr, e)

        ctk.CTkLabel(add_f, text="Role:", font=FONT_LABEL,
                     text_color=C_MUTED, anchor="w").pack(anchor="w", padx=14)
        self._var_add_role = ctk.StringVar(value="kasir")
        role_row = ctk.CTkFrame(add_f, fg_color="transparent")
        role_row.pack(anchor="w", padx=14, pady=(2, 8))
        for val, label in [("kasir", "Kasir"), ("admin", "Admin")]:
            ctk.CTkRadioButton(role_row, text=label,
                                variable=self._var_add_role, value=val,
                                font=FONT_LABEL, text_color=C_TEXT,
                                fg_color=C_ACCENT2).pack(side="left", padx=8)

        self.lbl_add_status = ctk.CTkLabel(add_f, text="", font=FONT_SMALL,
                                            text_color=C_GREEN)
        self.lbl_add_status.pack(pady=(0, 4))

        ctk.CTkButton(add_f, text="➕  Tambah User", height=32,
                      fg_color=C_ACCENT2, hover_color="#5A0FCC",
                      font=FONT_SUB, text_color="white",
                      command=self._tambah_user).pack(fill="x", padx=14, pady=(0, 12))

    # ── LOAD & FILTER ─────────────────────────────────────────────────────────
    def _load_users(self):
        self._build_stat_bar()
        self._semua_rows = self._buat_semua_rows()
        self._filter_tabel()

    def _buat_semua_rows(self):
        cfg          = load_config()
        users        = cfg.get("users", {})
        profil_semua = cfg.get("profil_rental", {})

        # Kumpulkan nilai yang muncul lebih dari sekali
        nama_rental_list = [p.get("nama_rental", "").strip().lower()
                            for p in profil_semua.values()]
        email_list       = [p.get("email", "").strip().lower()
                            for p in profil_semua.values()]
        hp_list          = [p.get("no_hp", "").strip()
                            for p in profil_semua.values()]

        dup_rental = {n for n in nama_rental_list if n and nama_rental_list.count(n) > 1}
        dup_email  = {e for e in email_list       if e and email_list.count(e) > 1}
        dup_hp     = {h for h in hp_list          if h and hp_list.count(h) > 1}

        rows = []
        for uname, udata in users.items():
            profil     = profil_semua.get(uname, {})
            nama_rental= profil.get("nama_rental", "—")
            email      = profil.get("email", "—")
            no_hp      = profil.get("no_hp", "—")
            tgl        = profil.get("tanggal_daftar", "—")[:10]
            role       = udata.get("role", "kasir")
            punya_profil = uname in profil_semua

            # Deteksi duplikat
            is_dup = (
                nama_rental.strip().lower() in dup_rental or
                email.strip().lower() in dup_email or
                no_hp.strip() in dup_hp
            )

            # Tag warna
            if is_dup:
                tag = "duplikat"
            elif uname in DEFAULT_USERS:
                tag = "default"
            elif not punya_profil:
                tag = "tanpa_profil"
            elif role == "admin":
                tag = "admin"
            else:
                tag = "kasir"

            ikon = "⚠" if is_dup else ("👑" if role == "admin" else "👤")

            rows.append({
                "uname":       uname,
                "role":        role,
                "nama_rental": nama_rental,
                "email":       email,
                "no_hp":       no_hp,
                "tgl":         tgl,
                "tag":         tag,
                "ikon":        ikon,
                "punya_profil": punya_profil,
                "is_dup":      is_dup,
            })

        # Urutkan: duplikat dulu, lalu admin, lalu kasir
        rows.sort(key=lambda r: (0 if r["is_dup"] else 1, r["uname"]))
        return rows

    def _filter_tabel(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        kata  = self.var_cari.get().strip().lower()
        fil   = self.var_filter.get()

        for r in self._semua_rows:
            # Filter dropdown
            if fil == "Admin"          and r["role"] != "admin":           continue
            if fil == "Kasir"          and r["role"] != "kasir":           continue
            if fil == "Punya Profil"   and not r["punya_profil"]:          continue
            if fil == "Tanpa Profil"   and r["punya_profil"]:              continue
            if fil == "⚠ Nama Rental Sama" and not r["is_dup"]:            continue

            # Pencarian teks
            if kata:
                gabung = f"{r['uname']} {r['nama_rental']} {r['email']} {r['no_hp']}".lower()
                if kata not in gabung:
                    continue

            self.tree.insert("", "end", values=(
                r["ikon"], r["uname"], r["role"].upper(),
                r["nama_rental"], r["email"], r["no_hp"], r["tgl"]
            ), tags=(r["tag"],))

    def _on_select(self, event):
        sel = self.tree.selection()
        if not sel:
            return

    def _get_username_terpilih(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("⚠", "Pilih user dari tabel dulu.")
            return None
        vals = self.tree.item(sel[0], "values")
        return vals[1] if vals else None  # index 1 = Username

    # ── AKSI ──────────────────────────────────────────────────────────────────
    def _buka_detail(self):
        uname = self._get_username_terpilih()
        if not uname:
            return
        cfg   = load_config()
        users = cfg.get("users", {})
        profil_semua = cfg.get("profil_rental", {})
        if uname not in users:
            messagebox.showerror("✖", "User tidak ditemukan di config.")
            return
        DialogDetailUser(self, uname, users[uname], profil_semua.get(uname, {}),
                         on_update=self._load_users)

    def _hapus_user(self):
        uname = self._get_username_terpilih()
        if not uname:
            return
        if uname in DEFAULT_USERS:
            messagebox.showwarning("⚠ Tidak Bisa Dihapus",
                                    f"User '{uname}' adalah user bawaan sistem dan tidak bisa dihapus.")
            return

        if not messagebox.askyesno(
                "🗑 Hapus User",
                f"Hapus user '{uname}' secara permanen?\n\n"
                f"Data profil rental dan semua kode aktivasi yang\n"
                f"terikat username ini juga akan dihapus dari log."):
            return

        cfg   = load_config()
        users = cfg.get("users", {})
        profil_semua = cfg.get("profil_rental", {})

        users.pop(uname, None)
        profil_semua.pop(uname, None)
        cfg["users"]         = users
        cfg["profil_rental"] = profil_semua
        save_config(cfg)

        # Hapus dari keygen log
        log = load_keygen_log()
        log = [e for e in log if e.get("username") != uname]
        save_keygen_log(log)

        messagebox.showinfo("✅ Dihapus", f"User '{uname}' berhasil dihapus.")
        self._load_users()

    def _reset_pass_cepat(self):
        uname = self._get_username_terpilih()
        if not uname:
            return

        dlg = ctk.CTkInputDialog(
            text=f"Password baru untuk '{uname}' (min 6 karakter):",
            title=f"🔒 Reset Password — {uname}")
        pw = dlg.get_input()
        if not pw:
            return
        pw = pw.strip()
        if len(pw) < 6:
            messagebox.showwarning("⚠", "Password minimal 6 karakter.")
            return

        cfg   = load_config()
        users = cfg.get("users", {})
        if uname not in users:
            messagebox.showerror("✖", "User tidak ditemukan.")
            return

        users[uname]["password"] = hashlib.sha256(pw.encode()).hexdigest()
        cfg["users"] = users
        save_config(cfg)
        messagebox.showinfo("✅ Berhasil", f"Password '{uname}' berhasil direset.")

    def _tambah_user(self):
        uname = self._e_add_user.get().strip().lower()
        pw    = self._e_add_pass.get().strip()
        role  = self._var_add_role.get()

        if not uname:
            self.lbl_add_status.configure(text="⚠ Username wajib diisi.", text_color=C_YELLOW)
            return
        if " " in uname:
            self.lbl_add_status.configure(text="⚠ Username tidak boleh ada spasi.", text_color=C_YELLOW)
            return
        if len(pw) < 6:
            self.lbl_add_status.configure(text="⚠ Password min 6 karakter.", text_color=C_YELLOW)
            return

        cfg   = load_config()
        users = cfg.get("users", {})
        if uname in users:
            self.lbl_add_status.configure(
                text=f"✖ Username '{uname}' sudah ada!", text_color=C_RED)
            return

        users[uname] = {
            "password": hashlib.sha256(pw.encode()).hexdigest(),
            "role":     role,
        }
        cfg["users"] = users
        save_config(cfg)

        self._e_add_user.delete(0, "end")
        self._e_add_pass.delete(0, "end")
        self.lbl_add_status.configure(
            text=f"✅ User '{uname}' [{role}] berhasil ditambahkan.",
            text_color=C_GREEN)
        self._load_users()

    # ── CEK DUPLIKAT ──────────────────────────────────────────────────────────
    def _cek_duplikat(self):
        cfg          = load_config()
        profil_semua = cfg.get("profil_rental", {})
        users        = cfg.get("users", {})

        # Kumpulkan semua nilai
        nama_rental_map = {}  # nilai → [username, ...]
        email_map       = {}
        hp_map          = {}

        for uname, profil in profil_semua.items():
            nr = profil.get("nama_rental", "").strip()
            em = profil.get("email", "").strip().lower()
            hp = profil.get("no_hp", "").strip()

            if nr:
                nama_rental_map.setdefault(nr.lower(), []).append(uname)
            if em:
                email_map.setdefault(em, []).append(uname)
            if hp:
                hp_map.setdefault(hp, []).append(uname)

        hasil = []
        ada_masalah = False

        # Duplikat nama rental
        for val, list_user in nama_rental_map.items():
            if len(list_user) > 1:
                ada_masalah = True
                hasil.append(f"⚠ NAMA RENTAL SAMA: '{val}'")
                for u in list_user:
                    hasil.append(f"   → {u}")

        # Duplikat email
        for val, list_user in email_map.items():
            if len(list_user) > 1:
                ada_masalah = True
                hasil.append(f"⚠ EMAIL SAMA: '{val}'")
                for u in list_user:
                    hasil.append(f"   → {u}")

        # Duplikat no HP
        for val, list_user in hp_map.items():
            if len(list_user) > 1:
                ada_masalah = True
                hasil.append(f"⚠ NO HP SAMA: '{val}'")
                for u in list_user:
                    hasil.append(f"   → {u}")

        # User tanpa profil
        tanpa_profil = [u for u in users if u not in profil_semua]
        if tanpa_profil:
            hasil.append(f"\nℹ USER TANPA PROFIL RENTAL ({len(tanpa_profil)}):")
            for u in tanpa_profil:
                hasil.append(f"   • {u} [{users[u].get('role','?')}]")

        if not ada_masalah and not tanpa_profil:
            hasil = ["✅ Tidak ada duplikat atau konflik yang ditemukan.",
                     "",
                     f"Total user    : {len(users)}",
                     f"Punya profil  : {len(profil_semua)}",
                     "Semua aman ✅"]

        self.txt_duplikat.configure(state="normal")
        self.txt_duplikat.delete("1.0", "end")
        self.txt_duplikat.insert("end", "\n".join(hasil))
        self.txt_duplikat.configure(state="disabled")

        self._load_users()  # refresh tabel juga

    # ── EXPORT EXCEL ──────────────────────────────────────────────────────────
    def _export_excel(self):
        cfg          = load_config()
        users        = cfg.get("users", {})
        profil_semua = cfg.get("profil_rental", {})

        if not users:
            messagebox.showwarning("⚠", "Belum ada user terdaftar.")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Daftar User"

        hdr_fill  = PatternFill("solid", fgColor="1A1A3A")
        hdr_font  = Font(name="Consolas", bold=True, color="00FFCC", size=11)
        norm_font = Font(name="Consolas", size=10)
        ctr       = Alignment(horizontal="center", vertical="center")
        border    = Border(
            left=Side(style="thin", color="2A2A5A"),
            right=Side(style="thin", color="2A2A5A"),
            top=Side(style="thin", color="2A2A5A"),
            bottom=Side(style="thin", color="2A2A5A"),
        )

        # Judul
        ws.merge_cells("A1:H1")
        ws["A1"] = "DAFTAR USER — RR BILLING PRO"
        ws["A1"].font      = Font(name="Consolas", bold=True, color="00FFCC", size=14)
        ws["A1"].fill      = PatternFill("solid", fgColor="0D0D1A")
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 28

        ws.merge_cells("A2:H2")
        ws["A2"] = f"Export: {datetime.now().strftime('%d %B %Y %H:%M')}"
        ws["A2"].font      = Font(name="Consolas", color="6060A0", size=9)
        ws["A2"].fill      = PatternFill("solid", fgColor="0D0D1A")
        ws["A2"].alignment = Alignment(horizontal="center")

        headers = ["Username", "Role", "Nama Pemilik", "Nama Rental",
                   "Email", "No HP", "Alamat", "Tgl Daftar"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=ci, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = ctr
            cell.border = border
        ws.row_dimensions[3].height = 22

        for ri, (uname, udata) in enumerate(users.items(), 4):
            profil = profil_semua.get(uname, {})
            row_data = [
                uname,
                udata.get("role", "kasir").upper(),
                profil.get("nama_pemilik", "—"),
                profil.get("nama_rental",  "—"),
                profil.get("email",        "—"),
                profil.get("no_hp",        "—"),
                profil.get("alamat",       "—"),
                profil.get("tanggal_daftar", "—")[:10],
            ]
            alt_fill = PatternFill("solid", fgColor="161628" if ri % 2 == 0 else "1A1A3A")
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font      = norm_font
                cell.fill      = alt_fill
                cell.border    = border
                cell.alignment = ctr if ci in (1, 2, 8) else Alignment(vertical="center")

        col_widths = [18, 10, 20, 22, 28, 16, 36, 14]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

        tgl  = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = os.path.abspath(f"daftar_user_rr_billing_{tgl}.xlsx")
        wb.save(path)

        try:
            import subprocess
            if os.name == "nt":
                os.startfile(path)
            else:
                subprocess.Popen(["xdg-open", path])
        except Exception:
            pass

        messagebox.showinfo("✅ Export Berhasil", f"File disimpan:\n{path}")

    def _buka_keygen(self):
        try:
            import subprocess, sys
            subprocess.Popen([sys.executable, "rr_keygen.py"])
        except Exception as e:
            messagebox.showerror("✖ Error", f"Gagal buka Keygen:\n{e}")


# ═══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    app = UserManagerApp()
    app.mainloop()