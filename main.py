import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import subprocess
import threading
import time
import socket
import asyncio
import websockets
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import openpyxl.utils
from datetime import datetime, timedelta, date
import os
import shutil
import json
import hashlib
import bcrypt  # ← Password hashing with salt
import jwt  # ← JWT tokens
import sys
import re
import webbrowser
import smtplib
import ssl
import random
import uuid
import tempfile
import logging
from urllib.parse import quote
from email.message import EmailMessage
_LOGGER = logging.getLogger(__name__)
from PIL import Image, ImageTk  # ← tambah untuk logo
try:
    import fcntl  # ← FIX: File locking untuk thread-safe config save (Unix/Linux)
except ImportError:
    fcntl = None  # Windows tidak support fcntl

from typing import Optional
from firebase_auth import get_firebase_auth
from firestore_sync import FirestoreClient
import tv_mesin
from tv_ws_hub import TvWsHub
from tv_test_api import TvTestApi
from tv_media_server import TvMediaServer
from media_prepare import ffmpeg_path, ffmpeg_target_path, prepare_video, FFMPEG_URL

try:
    import matplotlib
    matplotlib.use('Agg')
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    from matplotlib.figure import Figure
    import matplotlib.pyplot as plt
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

try:
    from escpos.printer import Usb, Network, File
    ESCPOS_AVAILABLE = True
except ImportError:
    ESCPOS_AVAILABLE = False


# ─── Printer Bluetooth BLE (RPP02N dkk — ISSC Transparent UART) ──────────────
# Karakteristik RX menerima byte ESC/POS mentah tanpa framing paket.
BLE_PRINT_RX_UUID = "49535343-8841-43f4-a8d4-ecbe34729bb3"
BLE_PRINT_SVC_HINT = "49535343"          # prefiks service ISSC (Transparent UART)
_BLE_PRINTER_HINTS = ("RPP", "POS58", "58MM", "MTP", "THERMAL", "PRINTER", "TSC")

def _ble_escpos_bytes(text: str) -> bytes:
    """Bangun byte ESC/POS teks: init + baris teks + feed + gunting kertas."""
    data = bytearray()
    data += b"\x1b\x40"                  # ESC @ init
    for line in str(text).split("\n"):
        data += line.encode("cp437", errors="replace") + b"\n"
    data += b"\x1b\x64\x05"              # feed 5 baris
    data += b"\x1d\x56\x42"              # GS V 66 — potong kertas
    return bytes(data)

def _ble_write_printer(addr: str, data: bytes):
    """Tulis byte ESC/POS ke printer BLE (chunk sesuai MTU, tanpa response)."""
    import asyncio
    from bleak import BleakClient

    async def _run():
        async with BleakClient(addr, timeout=20) as client:
            if not client.is_connected:
                raise RuntimeError("BLE tidak terhubung")
            mtu = client.mtu_size
            chunk = max(20, min(500, mtu - 3))
            for i in range(0, len(data), chunk):
                await client.write_gatt_char(
                    BLE_PRINT_RX_UUID, data[i:i + chunk], response=False)
                await asyncio.sleep(0.02)

    asyncio.run(_run())

def _ble_find_printer(timeout: int = 6) -> str:
    """Cari MAC printer BLE: 1) perangkat BLE ter-pair Windows, 2) pindai
    iklan BLE. Return MAC 'AA:BB:...' atau ''."""
    import re
    import subprocess

    # 1) Perangkat BLE ter-pair (Get-PnpDevice BTHLE) — dapat terlihat walau
    #    tidak sedang mengiklankan.
    try:
        ps_cmd = ("powershell -NoProfile -Command "
                  "\"Get-PnpDevice -PresentOnly -Class Bluetooth | "
                  "Where-Object { $_.InstanceId -like 'BTHLE*' } | "
                  "ForEach-Object { \\\"$($_.FriendlyName)`t$($_.InstanceId)\\\" }\"")
        out = subprocess.run(ps_cmd, shell=True, capture_output=True, text=True,
                             timeout=20, **subprocess_no_window_kwargs())
        for line in (out.stdout or "").splitlines():
            parts = line.split("\t")
            if len(parts) < 2:
                continue
            name, iid = (parts[0] or ""), (parts[1] or "")
            if any(h in name.upper() for h in _BLE_PRINTER_HINTS):
                m = re.search(r"BTHLE\\DEV_([0-9A-F]{12})", iid.upper())
                if m:
                    mac = ":".join(m.group(1)[i:i + 2] for i in range(0, 12, 2))
                    return mac
    except Exception:
        pass

    # 2) Pindai iklan BLE (fallback).
    try:
        from bleak import BleakScanner
        import asyncio

        async def _scan():
            devs = await BleakScanner.discover(timeout=timeout)
            for d in devs:
                if not d.address:
                    continue
                nm = (d.name or "").upper()
                if any(h in nm for h in _BLE_PRINTER_HINTS):
                    return d.address
                uuids = d.metadata.get("uuids") or []
                if any(str(u).lower().startswith(BLE_PRINT_SVC_HINT) for u in uuids):
                    return d.address
            return ""

        return asyncio.run(_scan())
    except Exception:
        return ""


try:
    from cryptography.fernet import Fernet
    from cryptography.hazmat.primitives import hashes
    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
    CRYPTO_AVAILABLE = True
except ImportError:
    CRYPTO_AVAILABLE = False

import base64


# ─── ASYNCIO BRIDGE THREAD (untuk WebSocket server) ──────────────────────────
class _WSLoopThread:
    """Singleton thread running asyncio event loop for WebSocket server."""
    _instance = None
    _lock = threading.Lock()

    def __init__(self):
        self.loop = None
        self.thread = None

    @classmethod
    def get_loop(cls):
        if cls._instance is None:
            with cls._lock:
                if cls._instance is None:
                    inst = cls()
                    inst._start()
                    cls._instance = inst
        return cls._instance.loop

    def _start(self):
        self.thread = threading.Thread(target=self._run, daemon=True)
        self.thread.start()
        while self.loop is None:
            time.sleep(0.01)

    def _run(self):
        self.loop = asyncio.new_event_loop()
        asyncio.set_event_loop(self.loop)
        self.loop.run_forever()

# ─── PASSWORD SECURITY HELPERS ──────────────────────────────────────────────
def hash_password(password: str) -> str:
    """Hash password dengan bcrypt (dengan salt). Format: bcrypt$<hash>"""
    salt = bcrypt.gensalt(rounds=12)
    hashed = bcrypt.hashpw(password.encode(), salt)
    return "bcrypt$" + hashed.decode('utf-8')

def verify_password(password: str, password_hash: str) -> bool:
    """Verify password against bcrypt hash (backward-compatible dengan SHA256)."""
    try:
        if not isinstance(password_hash, str) or not password_hash:
            return False
        # New bcrypt format
        if password_hash.startswith("bcrypt$"):
            hash_value = password_hash[7:].encode('utf-8')
            return bcrypt.checkpw(password.encode(), hash_value)
        # Legacy SHA256 format
        if re.fullmatch(r"[0-9a-f]{64}", password_hash):
            return hashlib.sha256(password.encode()).hexdigest() == password_hash
        # Legacy plaintext format (auto-upgrade after successful login)
        return password == password_hash
    except Exception as e:
        print(f"Password verification error: {e}")
        return False


WARNET_ADMIN_CODE_SECRET = "RR_WARNET_CFG_LOCK_V1"


def generate_warnet_admin_code(client_ip: str, day: datetime = None) -> str:
    """Generate kode akses pengaturan client berdasarkan IP + tanggal."""
    ip = (client_ip or "").strip()
    if not ip:
        return ""
    day = day or datetime.now()
    date_key = day.strftime("%Y%m%d")
    digest = hashlib.sha256(f"{ip}|{date_key}|{WARNET_ADMIN_CODE_SECRET}".encode("utf-8")).hexdigest().upper()
    return f"{digest[0:4]}-{digest[8:12]}-{digest[20:24]}"


# ─── WARNET SOCKET SERVER FOR CLIENT APPS ──────────────────────────────────────

class TokenManager:
    """JWT token manager untuk warnet client authentication."""
    SECRET_KEY = "rr_billing_warnet_secret_2024"
    ALGORITHM = "HS256"
    TOKEN_EXPIRY_DAYS = 180  # 6 months
    
    @classmethod
    def generate_token(cls, client_id: str) -> str:
        """Generate JWT token dengan 6-month expiry."""
        now = datetime.now()
        exp_time = now + timedelta(days=cls.TOKEN_EXPIRY_DAYS)
        
        payload = {
            "client_id": client_id,
            "iat": int(now.timestamp()),
            "exp": int(exp_time.timestamp())
        }
        return jwt.encode(payload, cls.SECRET_KEY, algorithm=cls.ALGORITHM)
    
    @classmethod
    def verify_token(cls, token: str) -> dict:
        """Verify token dan return payload, raise exception jika invalid."""
        try:
            payload = jwt.decode(token, cls.SECRET_KEY, algorithms=[cls.ALGORITHM])
            return payload
        except jwt.ExpiredSignatureError as e:
            raise Exception(f"Token expired: {e}")
        except jwt.InvalidSignatureError as e:
            raise Exception(f"Invalid signature: {e}")
        except jwt.DecodeError as e:
            raise Exception(f"Decode error: {e}")
        except jwt.InvalidTokenError as e:
            raise Exception(f"Invalid token: {e}")
        except ValueError as e:
            # jwt.decode can raise ValueError for malformed tokens
            raise Exception(f"Malformed token: {e}")
        except Exception as e:
            raise Exception(f"Token verification error ({type(e).__name__}): {e}")
    
    @classmethod
    def is_token_valid(cls, token: str) -> bool:
        """Quick check if token is valid."""
        try:
            cls.verify_token(token)
            return True
        except:
            return False


class WarnetSocketServer:
    """Socket server untuk warnet client app connections.
    Supports both TCP and WebSocket connections.
    """
    
    def __init__(self, config_manager=None, listen_port=5000, ws_port=5001, app=None):
        self.config_manager = config_manager or ConfigManager
        self.listen_port = listen_port
        self.ws_port = ws_port
        self.listen_address = "0.0.0.0"
        self.server_socket = None
        self.ws_server = None
        self.running = False
        self.sessions = {}  # {session_token: {client_id, last_heartbeat, address}}
        self.sessions_lock = threading.Lock()
        self.token_manager = TokenManager()
        self.app = app  # Reference to main app untuk query kursi/PC data
        self.pending_commands = {}  # {pc_id: [cmd_dict, ...]}
        self.pending_commands_lock = threading.Lock()
    
    def start(self):
        """Start socket server di background thread."""
        if self.running:
            return
        
        self.running = True
        # TCP server
        server_thread = threading.Thread(target=self._accept_connections, daemon=True)
        server_thread.start()
        print(f"[WARNET SERVER] TCP started on {self.listen_address}:{self.listen_port}")
        # WebSocket server
        self._start_ws()
    
    # ── Pending Commands Queue (untuk C# client) ──────────────
    def queue_pending_command(self, pc_id, cmd, **params):
        """Queue a command (LOCK/UNLOCK etc) for a specific PC.
        The command will be delivered on the next GET_STATUS/heartbeat from that PC."""
        with self.pending_commands_lock:
            if pc_id not in self.pending_commands:
                self.pending_commands[pc_id] = []
            cmd_dict = {"cmd": cmd, "timestamp": int(time.time())}
            cmd_dict.update(params)
            self.pending_commands[pc_id].append(cmd_dict)
            print(f"[WARNET SERVER] Queued command {cmd} for PC {pc_id}")

    def pop_pending_commands(self, pc_id):
        """Pop and return all pending commands for a PC (JSON-safe list)."""
        with self.pending_commands_lock:
            cmds = self.pending_commands.pop(pc_id, [])
        return cmds

    def stop(self):
        """Stop socket server."""
        self.running = False
        if self.server_socket:
            try:
                self.server_socket.close()
            except:
                pass
        self._stop_ws()
    
    def _start_ws(self):
        """Start WebSocket server in the asyncio event loop."""
        ws_loop = _WSLoopThread.get_loop()
        asyncio.run_coroutine_threadsafe(self._ws_serve(), ws_loop)
    
    def _stop_ws(self):
        if self.ws_server:
            try:
                self.ws_server.close()
            except:
                pass
            self.ws_server = None
    
    async def _ws_serve(self):
        """Async WebSocket server coroutine."""
        try:
            self.ws_server = await websockets.serve(
                self._handle_ws_client,
                self.listen_address,
                self.ws_port,
                ping_interval=20,
                ping_timeout=10
            )
            print(f"[WARNET SERVER] WebSocket started on ws://{self.listen_address}:{self.ws_port}")
            await asyncio.Future()  # Run forever
        except Exception as e:
            print(f"[WARNET SERVER] WebSocket server error: {e}")
    
    async def _handle_ws_client(self, websocket):
        """Handle a single WebSocket client connection."""
        session_token = None
        address = websocket.remote_address
        try:
            async for message in websocket:
                try:
                    data = json.loads(message)
                except json.JSONDecodeError:
                    await websocket.send(json.dumps({"type": "ERROR", "message": "Invalid JSON"}))
                    continue
                
                response = self._process_message(data)
                
                # Track / update session
                if data.get("type") == "AUTH" and response.get("status") == "OK":
                    session_token = response.get("session_token")
                    with self.sessions_lock:
                        self.sessions[session_token] = {
                            "client_id": data.get("client_id"),
                            "last_heartbeat": time.time(),
                            "address": address,
                            "transport": "websocket",
                        }
                
                if session_token and data.get("type") in ("COMMAND", "PING", "GET_STATUS"):
                    with self.sessions_lock:
                        if session_token in self.sessions:
                            self.sessions[session_token]["last_heartbeat"] = time.time()
                
                await websocket.send(json.dumps(response))
        except websockets.ConnectionClosed:
            pass
        except Exception as e:
            print(f"[WARNET SERVER WS] Client error ({address}): {e}")
        finally:
            if session_token:
                with self.sessions_lock:
                    self.sessions.pop(session_token, None)
    
    def _accept_connections(self):
        """Accept incoming client connections."""
        try:
            self.server_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            self.server_socket.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            self.server_socket.bind((self.listen_address, self.listen_port))
            self.server_socket.listen(20)
            
            while self.running:
                try:
                    self.server_socket.settimeout(1)
                    client_socket, address = self.server_socket.accept()
                    client_thread = threading.Thread(
                        target=self._handle_client,
                        args=(client_socket, address),
                        daemon=True
                    )
                    client_thread.start()
                except socket.timeout:
                    continue
                except Exception as e:
                    if self.running:
                        print(f"[WARNET SERVER] Accept error: {e}")
        
        except Exception as e:
            print(f"[WARNET SERVER] Server error: {e}")
        finally:
            self.running = False
    
    def _handle_client(self, client_socket, address):
        """Handle single client connection with line-based message framing."""
        session_token = None
        recv_buffer = b""
        try:
            while self.running:
                client_socket.settimeout(5)
                try:
                    chunk = client_socket.recv(4096)
                    if not chunk:
                        break
                    recv_buffer += chunk
                except socket.timeout:
                    continue
                except ConnectionError:
                    break
                
                # Process complete lines from buffer
                while b'\n' in recv_buffer:
                    line, recv_buffer = recv_buffer.split(b'\n', 1)
                    line = line.strip()
                    if not line:
                        continue
                    
                    try:
                        message = json.loads(line.decode('utf-8'))
                    except json.JSONDecodeError:
                        print(f"[WARNET SERVER] Invalid JSON from {address}: {line[:200]}")
                        continue
                    
                    response = self._process_message(message)
                    
                    # Track session if AUTH successful
                    if message.get("type") == "AUTH" and response.get("status") == "OK":
                        session_token = response.get("session_token")
                        with self.sessions_lock:
                            self.sessions[session_token] = {
                                "client_id": message.get("client_id"),
                                "last_heartbeat": time.time(),
                                "address": address
                            }
                    
                    # Update heartbeat
                    if session_token and message.get("type") in ["COMMAND", "PING", "GET_STATUS"]:
                        with self.sessions_lock:
                            if session_token in self.sessions:
                                self.sessions[session_token]["last_heartbeat"] = time.time()
                    
                    client_socket.sendall(json.dumps(response).encode('utf-8') + b'\n')
        
        except Exception as e:
            print(f"[WARNET SERVER] Client error ({address}): {e}")
        finally:
            try:
                client_socket.close()
            except:
                pass
            
            # Remove session on disconnect
            if session_token:
                with self.sessions_lock:
                    self.sessions.pop(session_token, None)
    
    def _process_message(self, message: dict) -> dict:
        """Process incoming message and return response."""
        msg_type = message.get("type")
        
        if msg_type == "AUTH":
            return self._handle_auth(message)
        elif msg_type == "COMMAND":
            return self._handle_command(message)
        elif msg_type == "PING":
            return self._handle_ping(message)
        elif msg_type == "GET_STATUS":
            return self._handle_get_status(message)
        elif msg_type == "REQUEST":
            return self._handle_request(message)
        else:
            return {"type": "ERROR", "message": f"Unknown message type: {msg_type}"}
    
    def _handle_auth(self, message: dict) -> dict:
        """Handle AUTH: verify client_id and password, return token."""
        client_id = message.get("client_id")
        password = message.get("password")
        
        cfg = self.config_manager.load()
        warnet_clients = cfg.get("warnet_clients", [])
        
        # Find client
        client_data = None
        for c in warnet_clients:
            if c.get("client_id") == client_id:
                client_data = c
                break
        
        if not client_data:
            return {
                "type": "AUTH_RESPONSE",
                "status": "FAIL",
                "message": f"Client {client_id} not found"
            }
        
        # Verify password
        password_hash = client_data.get("password_hash", "")
        if not verify_password(password, password_hash):
            return {
                "type": "AUTH_RESPONSE",
                "status": "FAIL",
                "message": "Invalid password"
            }
        
        # Generate token
        session_token = self.token_manager.generate_token(client_id)
        
        # Return token + PC list
        return {
            "type": "AUTH_RESPONSE",
            "status": "OK",
            "client_id": client_id,
            "session_token": session_token,
            "pcs": client_data.get("pcs", []),
            "timestamp": int(time.time()),
            "admin_code_secret": WARNET_ADMIN_CODE_SECRET,
        }
    
    def _handle_command(self, message: dict) -> dict:
        """Handle COMMAND: execute action on PC."""
        session_token = message.get("session_token")
        pc_id = message.get("pc_id")
        action = message.get("action")
        
        # Verify token is present
        if not session_token:
            return {
                "type": "COMMAND_RESPONSE",
                "status": "FAIL",
                "message": "Missing session_token"
            }
        
        # Debug: log token details
        # print(f"[WARNET SERVER] Token received: {session_token[:50]}...")
        
        # Verify token
        try:
            payload = self.token_manager.verify_token(session_token)
            client_id = payload.get("client_id")
        except Exception as e:
            # Debug error
            error_str = f"{type(e).__name__}: {str(e)}"
            print(f"[WARNET SERVER] Token verification failed: {error_str}")
            return {
                "type": "COMMAND_RESPONSE",
                "status": "FAIL",
                "message": f"Token error: {error_str}"
            }
        
        # Get client config
        cfg = self.config_manager.load()
        warnet_clients = cfg.get("warnet_clients", [])
        
        client_data = None
        for c in warnet_clients:
            if c.get("client_id") == client_id:
                client_data = c
                break
        
        if not client_data:
            return {
                "type": "COMMAND_RESPONSE",
                "status": "FAIL",
                "message": f"Client {client_id} not found"
            }
        
        # Verify PC ownership
        pc_data = None
        for pc in client_data.get("pcs", []):
            if pc.get("pc_id") == pc_id:
                pc_data = pc
                break
        
        if not pc_data:
            return {
                "type": "COMMAND_RESPONSE",
                "status": "FAIL",
                "message": f"PC {pc_id} not found for client {client_id}"
            }
        
        # Verify action is allowed
        allowed_actions = client_data.get("allowed_actions", ["ON", "OFF", "VOL+", "VOL-"])
        if action not in allowed_actions:
            return {
                "type": "COMMAND_RESPONSE",
                "status": "FAIL",
                "message": f"Action {action} not allowed"
            }
        
        # Execute command
        success = self._execute_pc_command(pc_data, action)
        
        # Log activity
        self._log_warnet_activity({
            "timestamp": datetime.now().isoformat(),
            "client_id": client_id,
            "pc_id": pc_id,
            "action": action,
            "success": success
        })
        
        return {
            "type": "COMMAND_RESPONSE",
            "status": "OK" if success else "FAIL",
            "message": f"{action} executed on {pc_id}",
            "timestamp": int(time.time())
        }
    
    def _handle_ping(self, message: dict) -> dict:
        """Handle PING: heartbeat."""
        return {
            "type": "PONG",
            "server_timestamp": int(time.time())
        }
    
    def _handle_get_status(self, message: dict) -> dict:
        """Handle GET_STATUS: return real billing status for a PC.
        Request: {"type": "GET_STATUS", "session_token": "...", "pc_id": "PC_1"}
        Response: {"type": "STATUS_RESPONSE", "status": "OK", "billing": {...}}
        """
        session_token = message.get("session_token")
        pc_id = message.get("pc_id")
        
        # Verify token
        try:
            payload = self.token_manager.verify_token(session_token)
            client_id = payload.get("client_id")
        except Exception as e:
            return {
                "type": "STATUS_RESPONSE",
                "status": "FAIL",
                "message": f"Token error: {str(e)}"
            }
        
        # Query real billing data from app state
        billing_status = {
            "pc_id": pc_id,
            "time_left": 0,
            "paket_aktif": "-",
            "total_biaya": 0,
            "is_playing": False,
            "timestamp": int(time.time())
        }
        
        # Determine if this request comes from a warnet client
        cfg = self.config_manager.load()
        is_warnet_client = any(
            c.get("client_id") == client_id
            for c in cfg.get("warnet_clients", [])
        )
        
        # TV loop — only for non-warnet clients
        if not is_warnet_client and self.app and hasattr(self.app, '_semua_kartu_tv'):
            for kursi in self.app._semua_kartu_tv:
                if hasattr(kursi, 'label_tv'):
                    if kursi.paket_aktif:
                        billing_status["paket_aktif"] = kursi.paket_aktif
                    if kursi.sisa_waktu > 0:
                        billing_status["time_left"] = kursi.sisa_waktu
                    if kursi.paket_harga_tetap > 0 or kursi.biaya_pesanan > 0:
                        billing_status["total_biaya"] = kursi.paket_harga_tetap + kursi.biaya_pesanan
                    if kursi.paket_aktif and kursi.sisa_waktu > 0:
                        billing_status["is_playing"] = True
                    break
        
        # Warnet cards — always run for warnet clients
        if is_warnet_client and self.app and hasattr(self.app, '_semua_kartu_warnet'):
            for kursi in self.app._semua_kartu_warnet:
                if getattr(kursi, '_pc_id', None) == pc_id:
                    if kursi.paket_aktif:
                        billing_status["paket_aktif"] = kursi.paket_aktif
                    elif getattr(kursi, '_last_transaction_item', None) or getattr(kursi, '_last_riwayat_idx', None) is not None:
                        # Session just ended — send "SELESAI" to trigger lock screen on client
                        billing_status["paket_aktif"] = "SELESAI"
                    if kursi.sisa_waktu > 0:
                        billing_status["time_left"] = kursi.sisa_waktu
                    if getattr(kursi, 'is_bebas', False) and getattr(kursi, 'waktu_mulai', None):
                        # Main Bebas: total_biaya = tarif/menit × menit terpakai + pesanan,
                        # sama dengan estimator "Total berjalan" di kartu (agar angka
                        # di PC client cocok dengan dashboard server).
                        tarif_menit = hitung_tarif_per_menit(kursi.get_paket_data())
                        total_detik = kursi.menit_dipakai_awal * 60 + int(
                            (datetime.now() - kursi.waktu_mulai).total_seconds())
                        biaya_waktu = tarif_menit * (max(total_detik, 0) / 60)
                        billing_status["total_biaya"] = int(
                            kursi._total_setelah_diskon(biaya_waktu + kursi.biaya_pesanan))
                    elif kursi.paket_harga_tetap > 0 or kursi.biaya_pesanan > 0:
                        billing_status["total_biaya"] = kursi.paket_harga_tetap + kursi.biaya_pesanan
                    if kursi.paket_aktif is not None:
                        billing_status["is_playing"] = True
                    if kursi.is_bebas:
                        billing_status["is_playing"] = True
                    if hasattr(kursi, 'is_on') and not kursi.is_on:
                        billing_status["is_playing"] = False
                        billing_status["time_left"] = 0
                    # Sumber kebenaran kedua utk lock state: kartu menandai pc_locked
                    # (LOCK via waktu_habis/selesai_manual/manual_off). Client memakai
                    # field ini sbg fallback bila pending_commands hilang (restart dll).
                    if getattr(kursi, 'pc_locked', False):
                        billing_status["is_locked"] = True
                    print(f"[GET_STATUS] pc={pc_id} client={client_id} "
                          f"kartu={kursi.label_kursi} pc_id_kartu={getattr(kursi, '_pc_id', None)} "
                          f"paket={billing_status.get('paket_aktif')} "
                          f"time_left={billing_status.get('time_left')} "
                          f"is_playing={billing_status.get('is_playing')} "
                          f"is_locked={billing_status.get('is_locked', False)}", flush=True)
                    break
        
        # Kartu yang ditandai terkunci (pc_locked) & tidak ada sesi aktif:
        # pastikan LOCK terkirim ulang — recovery saat client reboot /
        # server restart (pending_commands di server hilang). Debounce 10 dtk
        # (client idempotent: LockWorkstation skip bila sudah _isLocked).
        if is_warnet_client and self.app and hasattr(self.app, '_semua_kartu_warnet'):
            for kursi in self.app._semua_kartu_warnet:
                if getattr(kursi, '_pc_id', None) != pc_id:
                    continue
                if getattr(kursi, 'pc_locked', False):
                    sesi_aktif = bool(getattr(kursi, 'paket_aktif', None)) and getattr(kursi, 'sisa_waktu', 0) > 0
                    if not sesi_aktif:
                        now = time.time()
                        if now - getattr(kursi, '_last_lock_requeue', 0) >= 10:
                            kursi._last_lock_requeue = now
                            reason = getattr(kursi, '_pc_lock_reason', '') or "waktu_habis"
                            message = getattr(kursi, '_pc_lock_message', '') or f"PC {kursi.label_kursi} terkunci."
                            self.queue_pending_command(pc_id, "LOCK", reason=reason, message=message)
                break
        
        # Check for pending commands for this PC
        pending = self.pop_pending_commands(pc_id)
        if pending:
            billing_status["pending_commands"] = pending

        return {
            "type": "STATUS_RESPONSE",
            "status": "OK",
            "billing": billing_status
        }
    
    def _handle_request(self, message: dict) -> dict:
        """Handle REQUEST from client: add_time, order, chat."""
        session_token = message.get("session_token")
        pc_id = message.get("pc_id")
        request_type = message.get("request_type", "")
        data = message.get("data", {})
        
        try:
            payload = self.token_manager.verify_token(session_token)
            client_id = payload.get("client_id")
        except Exception as e:
            return {
                "type": "REQUEST_RESPONSE",
                "status": "FAIL",
                "message": f"Token error: {str(e)}"
            }
        
        log_msg = f"[WARNET] Request dari {client_id}/{pc_id}: type={request_type} data={data}"
        print(log_msg, flush=True)
        self._log_warnet_activity({
            "type": "request",
            "client_id": client_id,
            "pc_id": pc_id,
            "request_type": request_type,
            "data": data,
            "timestamp": int(time.time())
        })
        
        if request_type == "add_time":
            # Client minta tambah waktu: cari kartu warnet by pc_id, lalu tambahkan
            # paket langsung (jalan di main thread via after) dan catat transaksi.
            paket = data.get("package", "")
            kursi = None
            app = self.app
            if app is not None and hasattr(app, '_semua_kartu_warnet'):
                for k in app._semua_kartu_warnet:
                    if getattr(k, '_pc_id', None) == pc_id:
                        kursi = k
                        break
            if kursi is None:
                return {
                    "type": "REQUEST_RESPONSE",
                    "status": "FAIL",
                    "message": f"Kursi untuk PC {pc_id} belum ada di dashboard warnet."
                }
            if not paket:
                return {
                    "type": "REQUEST_RESPONSE",
                    "status": "FAIL",
                    "message": "Nama paket kosong."
                }
            try:
                paket_map = app.get_paket_data("Warnet", for_warnet=True)
            except Exception:
                paket_map = {}
            info = paket_map.get(paket) if isinstance(paket_map, dict) else None
            if not isinstance(info, dict):
                return {
                    "type": "REQUEST_RESPONSE",
                    "status": "FAIL",
                    "message": f"Paket {paket!r} tidak dikenal di grup tarif Warnet."
                }
            harga = int(info.get("harga", 0) or 0)
            menit = int(info.get("menit", 0) or 0)
            if menit <= 0:
                return {
                    "type": "REQUEST_RESPONSE",
                    "status": "FAIL",
                    "message": f"Paket {paket!r} tanpa durasi tidak bisa ditambahkan via add_time."
                }

            def _apply():
                try:
                    kursi._on_paket_confirm(paket, harga, menit, {}, 0)
                except Exception as e:
                    print(f"[WARNET] add_time apply error: {e}", flush=True)

            try:
                app.after(0, _apply)
            except Exception as e:
                return {
                    "type": "REQUEST_RESPONSE",
                    "status": "FAIL",
                    "message": f"Gagal memproses add_time: {e}"
                }
            return {
                "type": "REQUEST_RESPONSE",
                "status": "OK",
                "message": f"Tambah waktu {paket} ({menit} menit) diproses."
            }
        elif request_type == "order":
            return {
                "type": "REQUEST_RESPONSE",
                "status": "OK",
                "message": "Pesanan diterima."
            }
        elif request_type == "chat":
            return {
                "type": "REQUEST_RESPONSE",
                "status": "OK",
                "message": "Pesan terkirim ke admin."
            }
        else:
            return {
                "type": "REQUEST_RESPONSE",
                "status": "FAIL",
                "message": f"Unknown request type: {request_type}"
            }
    
    def _execute_pc_command(self, pc_data: dict, action: str) -> bool:
        """Execute command on TV/PC via ADB."""
        pc_ip = pc_data.get("ip")
        adb_port = pc_data.get("adb_port", 5555)
        if not pc_ip:
            return False
        try:
            action_map = {
                "ON":   lambda: ADBHelper.power_toggle(pc_ip, port=adb_port),
                "OFF":  lambda: ADBHelper.power_toggle(pc_ip, port=adb_port),
                "VOL+": lambda: ADBHelper.volume(pc_ip, naik=True, port=adb_port),
                "VOL-": lambda: ADBHelper.volume(pc_ip, naik=False, port=adb_port),
                "HOME": lambda: ADBHelper.home(pc_ip, port=adb_port),
            }
            if action not in action_map:
                return False
            ok, _, _ = action_map[action]()
            return ok
        except Exception as e:
            print(f"[WARNET SERVER] ADB error {action} on {pc_ip}:{adb_port}: {e}")
            return False
    
    def _log_warnet_activity(self, activity: dict):
        """Log warnet activity to config file."""
        try:
            cfg = self.config_manager.load()
            if "warnet_activity_log" not in cfg:
                cfg["warnet_activity_log"] = []
            
            cfg["warnet_activity_log"].append(activity)
            
            # Keep only last 1000 activities
            if len(cfg["warnet_activity_log"]) > 1000:
                cfg["warnet_activity_log"] = cfg["warnet_activity_log"][-1000:]
            
            self.config_manager.save(cfg)
        except Exception as e:
            print(f"[WARNET SERVER] Logging error: {e}")


def is_valid_username(username: str) -> bool:
    return bool(re.fullmatch(r"[a-z0-9_.]{4,20}", username))


def is_valid_email(email: str) -> bool:
    return bool(re.fullmatch(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", email))


def is_valid_phone(phone: str) -> bool:
    return bool(re.fullmatch(r"\+?[0-9]{7,15}", phone))


def is_valid_password(password: str) -> bool:
    return len(password) >= 8 and bool(re.search(r"[A-Za-z]", password)) and bool(re.search(r"[0-9]", password))


def is_valid_kasir_password(password: str) -> bool:
    # Aturan APTV2: min 6 karakter, wajib huruf besar & angka
    return (len(password) >= 6 and any(c.isupper() for c in password) and any(c.isdigit() for c in password))

EMAIL_VERIFICATION_EXPIRY_MINUTES = 30
EMAIL_VERIFICATION_CODE_LENGTH = 6
EMAIL_VERIFICATION_CHARS = "0123456789"


def generate_verification_code(length: int = EMAIL_VERIFICATION_CODE_LENGTH) -> str:
    return "".join(random.choice(EMAIL_VERIFICATION_CHARS) for _ in range(length))


def sanitize_text(value: str) -> str:
    return value.strip()


def subprocess_no_window_kwargs() -> dict:
    if os.name != "nt":
        return {}
    kwargs = {}
    creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0)
    if creationflags:
        kwargs["creationflags"] = creationflags
    startup_cls = getattr(subprocess, "STARTUPINFO", None)
    if startup_cls is not None:
        startupinfo = startup_cls()
        startupinfo.dwFlags |= getattr(subprocess, "STARTF_USESHOWWINDOW", 0)
        startupinfo.wShowWindow = 0
        kwargs["startupinfo"] = startupinfo
    return kwargs

# ─── TEMA ───────────────────────────────────────────────────────────────────
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

# ─── TEMA PRESETS ────────────────────────────────────────────────────────────
THEMES = {
    "Dark (Default)": {
        "bg": "#0A0A18", "panel": "#151530", "card": "#1E1E3E",
        "accent": "#00FFCC", "accent2": "#6C63FF",
        "red": "#FF5252", "green": "#00E676", "yellow": "#FFD740",
        "text": "#E8E8FF", "muted": "#8888BB", "btn": "#181838",
        "border": "#3A3A6A", "orange": "#FF8C00",
    },
    "Dark Blue": {
        "bg": "#0B0F2A", "panel": "#131844", "card": "#1C2360",
        "accent": "#4FC3F7", "accent2": "#7C4DFF",
        "red": "#EF5350", "green": "#66BB6A", "yellow": "#FFEE58",
        "text": "#E3E8FF", "muted": "#7A80B0", "btn": "#151B48",
        "border": "#2E3570", "orange": "#FFA726",
    },
    "Light": {
        "bg": "#F0F2FF", "panel": "#E0E4F8", "card": "#FFFFFF",
        "accent": "#00897B", "accent2": "#5C6BC0",
        "red": "#E53935", "green": "#43A047", "yellow": "#FDD835",
        "text": "#1A1A2E", "muted": "#7878A0", "btn": "#D8DCF0",
        "border": "#B0B4D0", "orange": "#EF6C00",
    },
    "Midnight": {
        "bg": "#050510", "panel": "#0A0A20", "card": "#101030",
        "accent": "#00E5FF", "accent2": "#AA00FF",
        "red": "#FF1744", "green": "#00E676", "yellow": "#FFEA00",
        "text": "#C0C0E0", "muted": "#606090", "btn": "#0D0D28",
        "border": "#202050", "orange": "#FF9100",
    },
}

# ─── SET DEFAULT THEME ───────────────────────────────────────────────────────
_default_theme = THEMES["Dark (Default)"]
C_BG     = _default_theme["bg"]
C_PANEL  = _default_theme["panel"]
C_CARD   = _default_theme["card"]
C_ACCENT  = _default_theme["accent"]
C_ACCENT2 = _default_theme["accent2"]
C_RED    = _default_theme["red"]
C_GREEN  = _default_theme["green"]
C_YELLOW = _default_theme["yellow"]
C_TEXT   = _default_theme["text"]
C_MUTED  = _default_theme["muted"]
C_BTN    = _default_theme["btn"]
C_BORDER = _default_theme["border"]
C_ORANGE = _default_theme["orange"]

FONT_TITLE  = ("Russo One",  20, "bold")
FONT_SUB    = ("Russo One",  14, "bold")
FONT_BODY   = ("Courier New", 14)
FONT_SMALL  = ("Courier New", 12)
FONT_LABEL  = ("Consolas",  13)

# ─── FILE KONFIGURASI ─────────────────────────────────────────────────────────
def get_app_base_dir() -> str:
    if getattr(sys, "frozen", False):
        return os.path.dirname(os.path.abspath(sys.executable))
    return os.path.dirname(os.path.abspath(__file__))


APP_BASE_DIR = get_app_base_dir()


def app_path(filename: str) -> str:
    return os.path.join(APP_BASE_DIR, filename)


def _qris_file() -> str:
    """Path file qris.png (QR pembayaran aktivasi): folder aplikasi → bundle
    PyInstaller (_MEIPASS). Return "" bila tidak ada di keduanya."""
    for base in (APP_BASE_DIR, getattr(sys, "_MEIPASS", "")):
        if base:
            p = os.path.join(base, "qris.png")
            if os.path.isfile(p):
                return p
    return ""


CONFIG_FILE   = app_path("rr_billing_config.json")
LICENSE_FILE  = app_path("rr_billing_license.json")
RIWAYAT_FILE  = app_path("rr_billing_riwayat.json")

# ─── SKALA UI (font & widget) ─────────────────────────────────────────────
# ui_scale: perbesar font/widget untuk layar PC beresolusi tinggi (font
# 10px -> 13px pada 1.30). Diatur lewat rr_billing_config.json; fallback 1.30.
def _load_ui_scale() -> float:
    try:
        with open(CONFIG_FILE, "r", encoding="utf-8") as _f:
            v = json.load(_f).get("ui_scale")
        return max(1.0, min(2.0, float(v)))
    except Exception:
        return 1.30


UI_SCALE = _load_ui_scale()
ctk.set_widget_scaling(UI_SCALE)
ctk.set_window_scaling(UI_SCALE)

# ─── DATA HARGA (default) ─────────────────────────────────────────────────────
# Harga dikelompokkan per "Grup Tarif" (mis. PS3, PS4, Room VIP),
# karena tiap jenis device bisa punya harga sewa berbeda.
# Tiap grup punya paket sendiri {nama: {"harga":int, "menit":int}},
# dan "Main Bebas" di dalam grup itu otomatis dihitung dari paket acuan grup itu.
NAMA_GRUP_DEFAULT = "Reguler"

_PAKET_STANDAR = {
    "30 Menit":   {"harga": 5_000,  "menit": 30},
    "1 Jam":      {"harga": 10_000, "menit": 60},
    "2 Jam":      {"harga": 18_000, "menit": 120},
    "3 Jam":      {"harga": 25_000, "menit": 180},
    "5 Jam":      {"harga": 35_000, "menit": 300},
    "Overnight":  {"harga": 50_000, "menit": 540},
    "Main Bebas": {"harga": 0,      "menit": 0},
}

# Grup tarif bawaan saat aplikasi pertama kali dijalankan (belum ada config tersimpan).
# User bisa menambah/mengubah/menghapus grup ini secara bebas lewat tab Kontrol Harga.
DEFAULT_GRUP_TARIF = {
    "Reguler":   {k: dict(v) for k, v in _PAKET_STANDAR.items()},
    "PS3":       {k: dict(v) for k, v in _PAKET_STANDAR.items()},
    "PS4":       {k: dict(v) for k, v in _PAKET_STANDAR.items()},
    "Room VIP":  {k: dict(v) for k, v in _PAKET_STANDAR.items()},
}

# Grup yang tidak boleh dihapus user (minimal harus ada 1 grup default fallback).
GRUP_TERKUNCI = {"Reguler"}

# Nama paket acuan tarif per-menit untuk Main Bebas di tiap grup.
# Tarif/menit Main Bebas = harga paket ini ÷ menit paket ini (di dalam grup yang sama).
PAKET_ACUAN_BEBAS = "1 Jam"

DEFAULT_MENU_MAKANAN = {
    "Indomie Goreng":    8_000,
    "Indomie Kuah":      8_000,
    "Kentang Goreng":   12_000,
    "Burger":           20_000,
    "Roti Bakar":       10_000,
    "Nasi Goreng":      15_000,
    "Sosis Bakar":       7_000,
}

DEFAULT_MENU_MINUMAN = {
    "Air Mineral":        3_000,
    "Es Teh Manis":       5_000,
    "Es Jeruk":           6_000,
    "Kopi Hitam":         6_000,
    "Susu Coklat":        8_000,
    "Jus Mangga":        10_000,
    "Soda Gembira":      10_000,
    "Pocari Sweat":       8_000,
}

def fmt_rp(n):
    return f"Rp {n:,.0f}".replace(",", ".")


def fmt_durasi(menit):
    """Format menit jadi teks 'X jam Y menit' yang enak dibaca."""
    if menit <= 0:
        return "Bebas"
    jam, sisa = divmod(menit, 60)
    if jam and sisa:
        return f"{jam} jam {sisa} menit"
    if jam:
        return f"{jam} jam"
    return f"{sisa} menit"

LOGO_RESIZE_MAX = 768  # 3x dari ukuran lama (256px) — logo di header web booking


def logo_gambar_b64(path: str, label_widget=None, tampil_error: bool = False) -> str:
    """Ubah file gambar jadi dataURL logo rental: <=768px (PNG, tetap transparan).
    Bila melampaui 700KB base64 -> otomatis JPEG q85 (transparansi hilang)
    supaya dokumen call_meta tetap di bawah limit 1MB Firestore."""
    try:
        import base64 as _b64
        from PIL import Image as _PIL
        from io import BytesIO as _Bio
    except Exception as e:
        if tampil_error and label_widget is not None:
            try:
                label_widget.configure(text=f"✖ Modul gambar: {e}")
            except Exception:
                pass
        return ""
    try:
        img = _PIL.open(path)
        img.load()
        img = img.convert("RGBA")
        img.thumbnail((LOGO_RESIZE_MAX, LOGO_RESIZE_MAX))
        out = _Bio()
        img.save(out, format="PNG", optimize=True)
        b64 = _b64.b64encode(out.getvalue()).decode("ascii")
        if len(b64) > 700000:
            rgb = img.convert("RGB")
            out2 = _Bio()
            rgb.save(out2, format="JPEG", quality=85)
            b64 = _b64.b64encode(out2.getvalue()).decode("ascii")
            if label_widget is not None:
                try:
                    label_widget.configure(text=f"✔ Logo terpasang ({len(b64) // 1024} KB, JPEG)")
                except Exception:
                    pass
            return "data:image/jpeg;base64," + b64
        if label_widget is not None:
            try:
                label_widget.configure(text=f"✔ Logo terpasang ({len(b64) // 1024} KB)")
            except Exception:
                pass
        return "data:image/png;base64," + b64
    except Exception as e:
        if label_widget is not None:
            try:
                label_widget.configure(text=f"✖ Gagal: {e}")
            except Exception:
                pass
        if tampil_error:
            try:
                messagebox.showerror("Gagal", f"Gagal memproses gambar:\n{e}")
            except Exception:
                pass
        return ""

DEFAULT_PORT = 5555
APP_VERSION = "2.4.13"
# Video promosi bawaan — disembunyikan (hidden attribute) supaya tidak bisa
# dihapus/diganti; satu-satunya video yang diputar user NON-LIFETIME.
PROMO_VIDEO_DEFAULT = "rr_promo_1785840135101.mp4"


# ═══════════════════════════════════════════════════════════════════════════════
#  HELPER LOGO
# ═══════════════════════════════════════════════════════════════════════════════
def _get_logo_path():
    """Cari logo.png di folder aplikasi; fallback ke folder bundle PyInstaller."""
    app_logo = app_path("logo.png")
    if os.path.exists(app_logo):
        return app_logo
    meipass = getattr(sys, "_MEIPASS", "")
    if meipass:
        bundled_logo = os.path.join(meipass, "logo.png")
        if os.path.exists(bundled_logo):
            return bundled_logo
    return app_logo


def load_ctk_image(size=(54, 54)):
    """
    Muat logo.png sebagai CTkImage untuk dipakai di CTkLabel.
    Kembalikan CTkImage jika berhasil, None jika gagal.
    """
    path = _get_logo_path()
    if not os.path.exists(path):
        return None
    try:
        img = Image.open(path).convert("RGBA").resize(size, Image.LANCZOS)
        return ctk.CTkImage(light_image=img, dark_image=img, size=size)
    except Exception as e:
        print(f"[logo] Gagal muat CTkImage: {e}")
        return None


def set_window_icon(window):
    """
    Set logo.png sebagai ikon window (titlebar & taskbar).
    Bekerja di Windows dan Linux/macOS.
    """
    path = _get_logo_path()
    if not os.path.exists(path):
        return
    try:
        img = Image.open(path)
        # Buat beberapa ukuran supaya taskbar & titlebar sama-sama bagus
        sizes = [(16,16),(32,32),(48,48),(64,64),(128,128)]
        icons = []
        for s in sizes:
            resized = img.resize(s, Image.LANCZOS)
            icons.append(ImageTk.PhotoImage(resized))
        # Simpan referensi agar tidak di-garbage-collect
        window._icon_images = icons
        window.iconphoto(True, *icons[::-1])  # urutan besar → kecil
    except Exception as e:
        print(f"[logo] Gagal set window icon: {e}")


def center_window(win, master=None, width=None, height=None):
    """Center a toplevel window relative to master or screen."""
    win.update_idletasks()
    geom = win.geometry().split("+")[0]
    if width is None or height is None:
        parts = geom.split("x")
        if len(parts) == 2:
            width, height = int(parts[0]), int(parts[1])
    if master is not None:
        try:
            master.update_idletasks()
            mx = master.winfo_x()
            my = master.winfo_y()
            mw = master.winfo_width()
            mh = master.winfo_height()
            x = mx + max((mw - width) // 2, 0)
            y = my + max((mh - height) // 2, 0)
        except Exception:
            x = max((win.winfo_screenwidth() - width) // 2, 0)
            y = max((win.winfo_screenheight() - height) // 2, 0)
    else:
        x = max((win.winfo_screenwidth() - width) // 2, 0)
        y = max((win.winfo_screenheight() - height) // 2, 0)
    win.geometry(f"{width}x{height}+{x}+{y}")


# ═══════════════════════════════════════════════════════════════════════════════
#  KONFIGURASI MANAGER
# ═══════════════════════════════════════════════════════════════════════════════
_CONFIG_LOCK = threading.RLock()


class ConfigManager:
    """Simpan & load data harga, user, dan lisensi dengan file locking untuk thread-safety."""
    @staticmethod
    def _decrypt_all(obj):
        if isinstance(obj, dict):
            decrypted = {}
            for k, v in obj.items():
                if isinstance(k, str) and k.endswith("_enc") and isinstance(v, str):
                    decrypted[k] = CryptoConfig.decrypt(v)
                elif k == "password" and isinstance(v, str) and not v.startswith("bcrypt$") and len(v) != 64:
                    decrypted[k] = v
                else:
                    decrypted[k] = ConfigManager._decrypt_all(v)
            return decrypted
        elif isinstance(obj, list):
            return [ConfigManager._decrypt_all(i) for i in obj]
        return obj

    @staticmethod
    def _encrypt_all(obj):
        if isinstance(obj, dict):
            encrypted = {}
            for k, v in obj.items():
                if isinstance(k, str) and k.endswith("_enc") and isinstance(v, str):
                    encrypted[k] = CryptoConfig.encrypt(v)
                else:
                    encrypted[k] = ConfigManager._encrypt_all(v)
            return encrypted
        elif isinstance(obj, list):
            return [ConfigManager._encrypt_all(i) for i in obj]
        return obj

    @staticmethod
    def load():
        with _CONFIG_LOCK:
            if os.path.exists(CONFIG_FILE):
                # Retry selama ~5 detik: jika file sedang ditulis proses lain,
                # baca ulang daripada mengembalikan {} (yang bisa menimpa config
                # saat di-save dan membuat fitur tampak "user tidak ditemukan").
                for _attempt in range(100):
                    try:
                        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                            data = json.load(f)
                        # Decrypt _enc fields
                        data = ConfigManager._decrypt_all(data)
                        # Auto-migrate: if users have plain 'password' (no password_enc), migrate
                        users = data.get("users", {})
                        migrated = False
                        for uname, u in users.items():
                            if isinstance(u, dict) and "password" in u and "password_enc" not in u:
                                pw = u.pop("password")
                                if pw:
                                    u["password_enc"] = CryptoConfig.encrypt(pw)
                                else:
                                    u["password_enc"] = ""
                                migrated = True
                            # Also check if role is missing
                            if isinstance(u, dict) and "role" not in u:
                                u["role"] = "kasir"
                        if migrated:
                            data["users"] = users
                            # Don't auto-save here; let the caller handle it
                        return data
                    except Exception:
                        time.sleep(0.05)
                # JSON korup permanen (semua retry gagal): backup file dulu
                # supaya isi lama tidak hilang saat save() menimpa dengan {}.
                try:
                    _bak = CONFIG_FILE + ".bak_corrupt"
                    if not os.path.exists(_bak):
                        import shutil
                        shutil.copy2(CONFIG_FILE, _bak)
                        print(f"[CONFIG] rr_billing_config.json korup — backup ke "
                              f"{os.path.basename(_bak)}; data dipakai: kosong", flush=True)
                except Exception:
                    pass
            return {}

    @staticmethod
    def save(data):
        """Simpan config dengan file locking + tulis atomik (tmp → replace)."""
        with _CONFIG_LOCK:
            # Encrypt _enc fields before saving
            data = ConfigManager._encrypt_all(data)
            lock_file = CONFIG_FILE + ".lock"
            max_retry = 10
            retry_delay = 0.05
            tmp_file = CONFIG_FILE + ".tmp"

            for attempt in range(max_retry):
                try:
                    # Buat lock file jika belum ada
                    with open(lock_file, 'a') as lock:
                        if os.name == 'nt':  # Windows
                            # Windows: gunakan win32 file locking
                            import msvcrt
                            msvcrt.locking(lock.fileno(), msvcrt.LK_NBLCK, 1)
                            try:
                                with open(tmp_file, "w", encoding="utf-8") as f:
                                    json.dump(data, f, indent=2, ensure_ascii=False)
                                os.replace(tmp_file, CONFIG_FILE)
                            finally:
                                msvcrt.locking(lock.fileno(), msvcrt.LK_UNLCK, 1)
                        else:  # Unix/Linux/Mac
                            if fcntl is not None:
                                fcntl.flock(lock.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
                                try:
                                    with open(tmp_file, "w", encoding="utf-8") as f:
                                        json.dump(data, f, indent=2, ensure_ascii=False)
                                    os.replace(tmp_file, CONFIG_FILE)
                                finally:
                                    fcntl.flock(lock.fileno(), fcntl.LOCK_UN)
                            else:
                                # fcntl not available, write without lock
                                with open(tmp_file, "w", encoding="utf-8") as f:
                                    json.dump(data, f, indent=2, ensure_ascii=False)
                                os.replace(tmp_file, CONFIG_FILE)
                    return  # Success
                except (IOError, OSError) as e:
                    if attempt < max_retry - 1:
                        time.sleep(retry_delay)
                        continue
                    else:
                        print(f"Config save error after {max_retry} retries: {e}")
                        raise

    @staticmethod
    def update(mutator):
        """Load → mutate → save secara atomik dalam satu lock.
        Gunakan ini untuk penulis config yang bisa dipanggil dari thread lain
        (mis. TimerService._sync_timer_state) agar tidak ada dua penulis yang
        saling menimpa key (bug kehilangan menu/tarif saat R3).
        Jika config tidak dapat dibaca, TIDAK menyimpan apa pun (mencegah wipe)."""
        with _CONFIG_LOCK:
            cfg = ConfigManager.load()
            if not cfg and os.path.exists(CONFIG_FILE):
                raise IOError("Config tidak dapat dibaca; simpan dibatalkan.")
            result = mutator(cfg)
            if result is not None:
                cfg = result
            ConfigManager.save(cfg)
            return cfg

    @staticmethod
    def get(key, default=None):
        d = ConfigManager.load()
        return d.get(key, default)

    @staticmethod
    def set(key, value):
        ConfigManager.update(lambda d: d.__setitem__(key, value) or d)


def _load_theme():
    name = ConfigManager.get("app_theme", "Dark (Default)")
    theme = THEMES.get(name, THEMES["Dark (Default)"])
    _apply_theme_globals(theme)
    return name

def _apply_theme_globals(theme):
    global C_BG, C_PANEL, C_CARD, C_ACCENT, C_ACCENT2
    global C_RED, C_GREEN, C_YELLOW, C_TEXT, C_MUTED, C_BTN, C_BORDER, C_ORANGE
    C_BG     = theme["bg"]
    C_PANEL  = theme["panel"]
    C_CARD   = theme["card"]
    C_ACCENT  = theme["accent"]
    C_ACCENT2 = theme["accent2"]
    C_RED    = theme["red"]
    C_GREEN  = theme["green"]
    C_YELLOW = theme["yellow"]
    C_TEXT   = theme["text"]
    C_MUTED  = theme["muted"]
    C_BTN    = theme["btn"]
    C_BORDER = theme["border"]
    C_ORANGE = theme["orange"]

# Muat tema dari config saat startup
_load_theme()


DEFAULT_EMAIL_SETTINGS = {
    "smtp_server": "",
    "smtp_port": 587,
    "smtp_username": "",
    "smtp_password_enc": "",
    "from_address": "",
    "use_tls": True,
}


def _normalize_email_settings(settings: dict | None) -> dict:
    merged = dict(DEFAULT_EMAIL_SETTINGS)
    if isinstance(settings, dict):
        merged.update(settings)
    try:
        merged["smtp_port"] = int(merged.get("smtp_port", 587))
    except Exception:
        merged["smtp_port"] = 587
    use_tls = merged.get("use_tls", True)
    if isinstance(use_tls, str):
        merged["use_tls"] = use_tls.strip().lower() in ("1", "true", "yes", "on")
    else:
        merged["use_tls"] = bool(use_tls)
    for key in ("smtp_server", "smtp_username", "smtp_password_enc", "from_address"):
        if merged.get(key) is None:
            merged[key] = ""
    return merged


def _set_email_settings(settings: dict) -> None:
    cfg = ConfigManager.load()
    cfg["email_settings"] = _normalize_email_settings(settings)
    ConfigManager.save(cfg)


def _get_email_settings():
    cfg = ConfigManager.load()
    settings = cfg.get("email_settings")
    if not isinstance(settings, dict):
        settings = cfg.get("smtp_settings", {})
    return _normalize_email_settings(settings)


def _email_configured() -> bool:
    settings = _get_email_settings()
    required = ["smtp_server", "smtp_port", "smtp_username", "from_address"]
    # Password bisa berupa smtp_password_enc (terenkripsi) ATAU smtp_password (plaintext lama)
    has_pass = str(settings.get("smtp_password_enc") or settings.get("smtp_password", "")).strip()
    return all(str(settings.get(k, "")).strip() for k in required) and bool(has_pass)


def _send_verification_email(to_email: str, username: str, code: str) -> tuple:
    settings = _get_email_settings()
    if not _email_configured():
        return False, "Email belum dikonfigurasi."

    msg = EmailMessage()
    msg["Subject"] = "RR Billing Pro - Email Verification Code"
    msg["From"] = settings.get("from_address")
    msg["To"] = to_email
    msg.set_content(
        f"Halo {username},\n\n"
        f"Kode verifikasi email Anda: {code}\n\n"
        f"Masukkan kode ini di aplikasi untuk menyelesaikan pendaftaran.\n"
        f"Kode berlaku selama {EMAIL_VERIFICATION_EXPIRY_MINUTES} menit.\n\n"
        "Terima kasih."
    )

    try:
        server = settings.get("smtp_server")
        port = int(settings.get("smtp_port", 587))
        username_smtp = settings.get("smtp_username")
        password_smtp = settings.get("smtp_password_enc") or settings.get("smtp_password", "")
        use_tls = bool(settings.get("use_tls", True))

        if use_tls:
            context = ssl.create_default_context()
            with smtplib.SMTP(server, port, timeout=10) as smtp:
                smtp.starttls(context=context)
                smtp.login(username_smtp, password_smtp)
                smtp.send_message(msg)
        else:
            with smtplib.SMTP(server, port, timeout=10) as smtp:
                smtp.login(username_smtp, password_smtp)
                smtp.send_message(msg)
        return True, "Email verifikasi telah dikirim."
    except Exception as e:
        return False, str(e)

# ─── AUDIT LOGGER ───────────────────────────────────────────────────────────
AUDIT_FILE = app_path("rr_billing_audit.jsonl")

# ─── RIWAYAT PESANAN WEB (QR CALL) ──────────────────────────────────────────
QR_PESAN_LOG = os.path.join(APP_BASE_DIR, "qr_pesanan_log.json")
QR_PESAN_LOG_LOCK = threading.Lock()


class AuditLogger:
    @staticmethod
    def _append_line(line: str):
        lock_file = AUDIT_FILE + ".lock"
        max_retry = 10
        retry_delay = 0.05

        for attempt in range(max_retry):
            try:
                with open(lock_file, 'a') as lock:
                    if os.name == 'nt':
                        import msvcrt
                        msvcrt.locking(lock.fileno(), msvcrt.LK_NBLCK, 1)
                        try:
                            with open(AUDIT_FILE, 'a', encoding='utf-8') as f:
                                f.write(line + '\n')
                        finally:
                            msvcrt.locking(lock.fileno(), msvcrt.LK_UNLCK, 1)
                    else:
                        if fcntl is not None:
                            fcntl.flock(lock.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
                            try:
                                with open(AUDIT_FILE, 'a', encoding='utf-8') as f:
                                    f.write(line + '\n')
                            finally:
                                fcntl.flock(lock.fileno(), fcntl.LOCK_UN)
                        else:
                            with open(AUDIT_FILE, 'a', encoding='utf-8') as f:
                                f.write(line + '\n')
                return
            except (IOError, OSError):
                if attempt < max_retry - 1:
                    time.sleep(retry_delay)
                    continue
                raise

    @staticmethod
    def log(action: str, username: str = "", status: str = "", details: dict = None):
        payload = {
            "timestamp": datetime.now().isoformat(),
            "action": action,
            "username": username,
            "status": status,
            "details": details or {},
        }
        try:
            AuditLogger._append_line(json.dumps(payload, ensure_ascii=False))
        except Exception as e:
            print(f"Audit log error: {e}")


# ─── BACKGROUND TIMER SERVICE ──────────────────────────────────────────────
class TimerService:
    """Background timer service that ticks all active sessions every second.
    Replaces the per-card after() timer with a centralized tick loop,
    so timers keep running even when cards are scrolled off-screen."""

    TICK_INTERVAL = 1.0
    SYNC_INTERVAL = 30

    def __init__(self, app=None):
        self.app = app
        self._thread = None
        self._running = False
        self._tick_count = 0
        self._lock = threading.Lock()

    def start(self, app):
        self.app = app
        if self._running:
            return
        self._running = True
        self._thread = threading.Thread(target=self._run, daemon=True, name="TimerService")
        self._thread.start()
        print("[TIMER] Background timer service started")

    def stop(self):
        self._running = False
        if self._thread:
            self._thread.join(timeout=2)
            self._thread = None
        print("[TIMER] Background timer service stopped")

    def _run(self):
        while self._running:
            try:
                self._tick()
            except Exception as e:
                print(f"[TIMER] Tick error: {e}")
            time.sleep(self.TICK_INTERVAL)

    def _tick(self):
        app = self.app
        if not app:
            return
        self._tick_count += 1

        now = time.time()
        for kartu in getattr(app, '_semua_kartu_tv', []):
            try:
                if (kartu.paket_aktif and kartu.sisa_waktu > 0
                        and not kartu._timer_paused
                        and not getattr(kartu, '_billing_paused', False)):
                    kartu.sisa_waktu = max(0, kartu.sisa_waktu - 1)
                    if kartu.sisa_waktu <= 0:
                        app.after_idle(lambda k=kartu: k._timer_habis())
                    else:
                        app.after_idle(lambda k=kartu: k._update_timer_display())
            except Exception:
                pass

        for kartu in getattr(app, '_semua_kartu_warnet', []):
            try:
                if kartu.paket_aktif and kartu.sisa_waktu > 0 and not kartu._timer_paused:
                    kartu.sisa_waktu = max(0, kartu.sisa_waktu - 1)
                    if kartu.sisa_waktu <= 0:
                        app.after_idle(lambda k=kartu: k._timer_habis())
                    else:
                        app.after_idle(lambda k=kartu: k._update_timer_display())
            except Exception:
                pass

        if self._tick_count % self.SYNC_INTERVAL == 0:
            self._sync_timer_state()

    def _sync_timer_state(self):
        try:
            state = {"tv": {}, "warnet": {}}
            for kartu in getattr(self.app, '_semua_kartu_tv', []):
                if kartu.paket_aktif and kartu.sisa_waktu > 0:
                    state["tv"][str(kartu.nomor)] = {
                        "label": kartu.label_tv,
                        "paket": kartu.paket_aktif,
                        "sisa_waktu": kartu.sisa_waktu,
                        "is_bebas": kartu.is_bebas,
                        "waktu_mulai": kartu.waktu_mulai.isoformat() if kartu.waktu_mulai else None,
                        "biaya_pesanan": kartu.biaya_pesanan,
                        "paket_harga_tetap": kartu.paket_harga_tetap,
                        "pesanan_aktif": dict(kartu.pesanan_aktif),
                        "daftar_paket_sesi": list(kartu.daftar_paket_sesi),
                        "lunas_paket": list(kartu.lunas_paket),
                        "harga_paket_sesi": list(kartu.harga_paket_sesi),
                        "lunas_pesanan": dict(kartu.lunas_pesanan),
                        "menit_dipakai_awal": kartu.menit_dipakai_awal,
                        "diskoni": kartu.diskoni,
                        "diskoni_mode": kartu.diskoni_mode,
                        "paid": kartu.paid,
                    }
            for kartu in getattr(self.app, '_semua_kartu_warnet', []):
                if kartu.paket_aktif and kartu.sisa_waktu > 0:
                    state["warnet"][str(getattr(kartu, '_pc_id', kartu.nomor))] = {
                        "label": getattr(kartu, 'label_kursi', str(kartu.nomor)),
                        "paket": kartu.paket_aktif,
                        "sisa_waktu": kartu.sisa_waktu,
                        "is_bebas": kartu.is_bebas,
                        "waktu_mulai": kartu.waktu_mulai.isoformat() if kartu.waktu_mulai else None,
                        "biaya_pesanan": kartu.biaya_pesanan,
                        "paket_harga_tetap": kartu.paket_harga_tetap,
                        "pesanan_aktif": dict(kartu.pesanan_aktif),
                        "daftar_paket_sesi": list(kartu.daftar_paket_sesi),
                        "lunas_paket": list(kartu.lunas_paket),
                        "harga_paket_sesi": list(kartu.harga_paket_sesi),
                        "lunas_pesanan": dict(kartu.lunas_pesanan),
                        "menit_dipakai_awal": kartu.menit_dipakai_awal,
                        "diskoni": kartu.diskoni,
                        "diskoni_mode": kartu.diskoni_mode,
                        "paid": kartu.paid,
                    }
            try:
                ConfigManager.update(lambda cfg: cfg.__setitem__("timer_state", state) or
                                     cfg.__setitem__("timer_state_updated", datetime.now().isoformat()) or
                                     cfg)
            except Exception as e:
                print(f"[TIMER] Sync error: {e}")
        except Exception as e:
            print(f"[TIMER] Sync error: {e}")

    @staticmethod
    def restore_timer_state(app):
        try:
            cfg = ConfigManager.load()
            state = cfg.get("timer_state", {})
            if not state:
                return
            tv_state = state.get("tv", {})
            warnet_state = state.get("warnet", {})
            restored_count = 0
            for kartu in getattr(app, '_semua_kartu_tv', []):
                key = str(kartu.nomor)
                if key in tv_state:
                    s = tv_state[key]
                    if s.get("paket") and s.get("sisa_waktu", 0) > 0:
                        kartu.paket_aktif = s["paket"]
                        kartu.sisa_waktu = s["sisa_waktu"]
                        kartu.is_bebas = s.get("is_bebas", False)
                        kartu.biaya_pesanan = s.get("biaya_pesanan", 0)
                        kartu.paket_harga_tetap = s.get("paket_harga_tetap", 0)
                        kartu.pesanan_aktif = dict(s.get("pesanan_aktif", {}))
                        kartu.daftar_paket_sesi = list(s.get("daftar_paket_sesi", []))
                        kartu.lunas_paket = list(s.get("lunas_paket", []))
                        kartu.harga_paket_sesi = list(s.get("harga_paket_sesi", []))
                        kartu.lunas_pesanan = dict(s.get("lunas_pesanan", {}))
                        kartu.menit_dipakai_awal = s.get("menit_dipakai_awal", 0)
                        kartu.diskoni = s.get("diskoni", 0)
                        kartu.diskoni_mode = s.get("diskoni_mode", "nominal")
                        if s.get("waktu_mulai"):
                            try:
                                kartu.waktu_mulai = datetime.fromisoformat(s["waktu_mulai"])
                            except Exception:
                                kartu.waktu_mulai = datetime.now()
                        kartu.is_on = True
                        kartu.paid = s.get("paid", True)
                        restored_count += 1
            for kartu in getattr(app, '_semua_kartu_warnet', []):
                key = str(getattr(kartu, '_pc_id', None) or kartu.nomor)
                if key in warnet_state:
                    s = warnet_state[key]
                    if s.get("paket") and s.get("sisa_waktu", 0) > 0:
                        kartu.paket_aktif = s["paket"]
                        kartu.sisa_waktu = s["sisa_waktu"]
                        kartu.is_bebas = s.get("is_bebas", False)
                        kartu.biaya_pesanan = s.get("biaya_pesanan", 0)
                        kartu.paket_harga_tetap = s.get("paket_harga_tetap", 0)
                        kartu.pesanan_aktif = dict(s.get("pesanan_aktif", {}))
                        kartu.daftar_paket_sesi = list(s.get("daftar_paket_sesi", []))
                        kartu.lunas_paket = list(s.get("lunas_paket", []))
                        kartu.harga_paket_sesi = list(s.get("harga_paket_sesi", []))
                        kartu.lunas_pesanan = dict(s.get("lunas_pesanan", {}))
                        kartu.menit_dipakai_awal = s.get("menit_dipakai_awal", 0)
                        kartu.diskoni = s.get("diskoni", 0)
                        kartu.diskoni_mode = s.get("diskoni_mode", "nominal")
                        if s.get("waktu_mulai"):
                            try:
                                kartu.waktu_mulai = datetime.fromisoformat(s["waktu_mulai"])
                            except Exception:
                                kartu.waktu_mulai = datetime.now()
                        kartu.is_on = True
                        kartu.paid = s.get("paid", True)
                        restored_count += 1
                        print(f"[TIMER] Restored warnet {kartu.label_kursi}: {s.get('paket')} sisa {s.get('sisa_waktu')}s", flush=True)
            # Relink kartu yang dipulihkan ke baris riwayat yang sudah tercatat
            # (source+label) agar timer habis tidak membuat baris duplikat.
            try:
                for kartu in getattr(app, '_semua_kartu_tv', []):
                    if not getattr(kartu, 'paket_aktif', None):
                        continue
                    for i, meta in enumerate(app.riwayat_meta):
                        if not isinstance(meta, dict) or meta.get("source") != "tv":
                            continue
                        if meta.get("_linked"):
                            continue
                        row = app.riwayat_transaksi[i] if i < len(app.riwayat_transaksi) else ()
                        if len(row) > 2 and row[2] == getattr(kartu, 'label_tv', None):
                            kartu._last_riwayat_idx = i
                            kartu._last_cloud_id = meta.get("cloud_id")
                            kartu._last_transaction_item = None
                            meta["_linked"] = True
                            break
                for kartu in getattr(app, '_semua_kartu_warnet', []):
                    if not getattr(kartu, 'paket_aktif', None):
                        continue
                    for i, meta in enumerate(app.riwayat_meta):
                        if not isinstance(meta, dict) or meta.get("source") != "warnet":
                            continue
                        if meta.get("_linked"):
                            continue
                        row = app.riwayat_transaksi[i] if i < len(app.riwayat_transaksi) else ()
                        if len(row) > 2 and row[2] == getattr(kartu, 'label_kursi', None):
                            kartu._last_riwayat_idx = i
                            kartu._last_cloud_id = meta.get("cloud_id")
                            kartu._last_transaction_item = None
                            meta["_linked"] = True
                            break
            except Exception:
                pass
            print(f"[TIMER] Restored {restored_count} active sessions")
        except Exception as e:
            print(f"[TIMER] Restore error: {e}")


# Modul lisensi terpisah (rr_license.py). Kalau ada rr_keygen.py (binding per-
# username) ia akan dipakai otomatis oleh LicenseManager.aktivasi(); kalau
# tidak ada, LicenseManager sudah punya fallback aman di dalamnya sendiri.
from rr_license import LicenseManager, LicenseGenerator, get_machine_id, get_edition_limits


# ─── AES-256 ENCRYPTION FOR SENSITIVE CONFIG DATA ─────────────────────────
class CryptoConfig:
    """AES-256 encryption for sensitive config values using cryptography.fernet."""
    SALT_FILE = app_path(".rr_salt")
    KEY_FILE = app_path(".rr_key")
    _cached_key = None

    @staticmethod
    def _derive_key(password: str, salt: bytes) -> bytes:
        if not CRYPTO_AVAILABLE:
            return b""
        kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=salt, iterations=100000)
        return base64.urlsafe_b64encode(kdf.derive(password.encode()))

    @staticmethod
    def _get_or_create_salt() -> bytes:
        if os.path.exists(CryptoConfig.SALT_FILE):
            with open(CryptoConfig.SALT_FILE, "rb") as f:
                return f.read()
        salt = os.urandom(16)
        with open(CryptoConfig.SALT_FILE, "wb") as f:
            f.write(salt)
        return salt

    @staticmethod
    def _get_machine_key() -> bytes:
        if CryptoConfig._cached_key is not None:
            return CryptoConfig._cached_key
        machine_id = get_machine_id()[:32]
        salt = CryptoConfig._get_or_create_salt()
        key = CryptoConfig._derive_key(machine_id, salt)
        CryptoConfig._cached_key = key
        return key

    @staticmethod
    def encrypt(plaintext: str) -> str:
        if not plaintext or not CRYPTO_AVAILABLE:
            return plaintext
        try:
            key = CryptoConfig._get_machine_key()
            f = Fernet(key)
            return f.encrypt(plaintext.encode()).decode()
        except Exception as e:
            print(f"[CRYPTO] Encrypt error: {e}")
            return plaintext

    @staticmethod
    def decrypt(ciphertext: str) -> str:
        if not ciphertext or not CRYPTO_AVAILABLE:
            return ciphertext
        try:
            key = CryptoConfig._get_machine_key()
            f = Fernet(key)
            return f.decrypt(ciphertext.encode()).decode()
        except Exception as e:
            print(f"[CRYPTO] Decrypt error: {e}")
            return ciphertext

    @staticmethod
    def encrypt_dict(d: dict) -> dict:
        out = {}
        for k, v in d.items():
            if k.endswith("_enc") and isinstance(v, str):
                out[k] = CryptoConfig.encrypt(v)
            else:
                out[k] = v
        return out

    @staticmethod
    def decrypt_dict(d: dict) -> dict:
        out = {}
        for k, v in d.items():
            if k.endswith("_enc") and isinstance(v, str):
                out[k] = CryptoConfig.decrypt(v)
            else:
                out[k] = v
        return out


# ═══════════════════════════════════════════════════════════════════════════════
#  HELPER ANDROID TV REMOTE (menggantikan ADB)
# ═══════════════════════════════════════════════════════════════════════════════
class ADBHelper:
    """Wrapper untuk Android TV Remote v2 — menggantikan ADB sepenuhnya."""

    _instances: dict[str, tv_mesin.AndroidTVRemote] = {}
    _connection_methods: dict[str, str] = {}
    _cert_paths: dict[str, str] = {}
    _lock = threading.Lock()
    _reconnect_locks: dict[str, threading.Lock] = {}
    _reconnect_locks_guard = threading.Lock()
    _adb_exe = None

    @classmethod
    def _adb_binary(cls):
        """Path adb.exe: bundle lokal (bin/adb/adb.exe) dulu, fallback PATH.
        Konsisten di semua PC user — versi adb lama sering gagal menangani
        TV Android 11+ (ADB TLS / streaming install) yang menimbulkan
        error samar seperti avc: denied di sisi TV."""
        if cls._adb_exe:
            return cls._adb_exe
        if getattr(sys, "frozen", False):
            base = os.path.dirname(os.path.abspath(sys.executable))
        else:
            base = os.path.dirname(os.path.abspath(__file__))
        bundled = os.path.join(base, "bin", "adb", "adb.exe")
        cls._adb_exe = bundled if os.path.isfile(bundled) else "adb"
        return cls._adb_exe

    @classmethod
    def _reconnect_lock(cls, ip: str) -> threading.Lock:
        """Lock per-IP agar reconnect/power-check tidak saling menumpuk
        (dua thread connect_blocking bersamaan untuk TV yang sama)."""
        with cls._reconnect_locks_guard:
            lock = cls._reconnect_locks.get(ip)
            if lock is None:
                lock = threading.Lock()
                cls._reconnect_locks[ip] = lock
            return lock

    @classmethod
    def _get_remote(cls, ip: str) -> tv_mesin.AndroidTVRemote:
        if ip not in cls._instances:
            with cls._lock:
                if ip not in cls._instances:
                    cls._instances[ip] = tv_mesin.AndroidTVRemote()
        return cls._instances[ip]

    @classmethod
    def adb_tersedia(cls):
        try:
            import androidtvremote2
            return True
        except ImportError:
            return False

    @classmethod
    def set_connection_method(cls, ip, method="atpv2"):
        with cls._lock:
            cls._connection_methods[ip] = method

    @classmethod
    def get_connection_method(cls, ip):
        return cls._connection_methods.get(ip, "atpv2")

    @classmethod
    def pair_dan_connect(cls, ip, pair_port, kode_pairing=None):
        try:
            result = tv_mesin.pair_tv_sync(ip, api_port=pair_port, pair_port=pair_port + 1)
            if result.get("status") != "pairing_started":
                return False, None, result.get("message", "Pairing gagal dimulai")
            remote_obj = result.get("remote")
            pin = kode_pairing or "000000"
            fin = tv_mesin.finish_pair_sync(remote_obj, pin)
            if fin.get("status") != "paired":
                return False, None, fin.get("message", "PIN salah atau pairing gagal")
            cls.set_connection_method(ip, "atpv2")
            rem = cls._get_remote(ip)
            conn = rem.connect_blocking(ip)
            if conn.get("status") == "connected":
                return True, 0, f"Paired & connected ke {ip}"
            return False, None, conn.get("message", "Gagal connect setelah pair")
        except Exception as e:
            return False, None, str(e)

    @classmethod
    def connect(cls, ip, port=0, method="atpv2"):
        cls.set_connection_method(ip, method)
        rem = cls._get_remote(ip)
        res = rem.connect_blocking(ip) if method == "atpv2" else ADBHelper._adb_connect(ip, port)
        if res.get("status") == "connected":
            return True, res.get("message", f"Terhubung ke {ip}")
        return False, res.get("message", "Gagal terhubung")

    @classmethod
    def _adb_connect(cls, ip, port=5555):
        import subprocess
        try:
            r = subprocess.run([cls._adb_binary(), "connect", f"{ip}:{port}"],
                               capture_output=True, text=True, timeout=10,
                               **subprocess_no_window_kwargs())
            if "connected" in r.stdout.lower() or "already connected" in r.stdout.lower():
                return {"status": "connected", "message": r.stdout.strip()}
            return {"status": "error", "message": r.stdout.strip() or r.stderr.strip()}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    @classmethod
    def disconnect(cls, ip, port=0):
        rem = cls._instances.pop(ip, None)
        if rem:
            rem.disconnect()
        cls._connection_methods.pop(ip, None)
        return True, "Disconnected"

    @classmethod
    def list_devices(cls):
        devices = {}
        for ip, rem in list(cls._instances.items()):
            if rem.is_connected():
                devices[ip] = "device"
        return devices, ""

    @classmethod
    def status_untuk_ip(cls, ip, port=0):
        rem = cls._instances.get(ip)
        if rem and rem.is_connected():
            method = cls._connection_methods.get(ip, "atpv2")
            return "device", method
        return "not_found", ""

    @classmethod
    def check_connection(cls, ip):
        rem = cls._instances.get(ip)
        if not rem:
            return False, "Tidak ada instance"
        return rem.is_connected(), ""

    @classmethod
    def check_connection_detail(cls, ip):
        rem = cls._instances.get(ip)
        if not rem:
            return {"status": "error", "message": "Tidak ada instance"}
        return rem.check_connection_blocking()

    @classmethod
    def power_toggle(cls, ip, port=0):
        """Kirim KEYCODE_POWER via atpv2. Kalau remote belum terhubung,
        coba reconnect dulu (sertifikat pairing tersimpan per IP) sebelum
        mengirim — supaya perintah tidak gagal senyap saat koneksi putus."""
        rem = cls._get_remote(ip)
        if not rem.is_connected():
            try:
                ok_c, _, msg_c = cls.cek_dan_reconnect(ip, port)
                if not ok_c:
                    return False, f"Tidak terhubung (reconnect gagal: {str(msg_c)[:100]})", str(msg_c)
            except Exception as e:
                return False, f"Tidak terhubung (reconnect error: {e})", str(e)
        res = rem.turn_off_blocking()
        ok = res.get("status") == "ok"
        return ok, res.get("message", ""), res.get("message", "")

    @classmethod
    def volume(cls, ip, naik=True, port=0):
        rem = cls._get_remote(ip)
        res = rem.volume_up_blocking() if naik else rem.volume_down_blocking()
        ok = res.get("status") == "ok"
        return ok, res.get("message", ""), res.get("message", "")

    @classmethod
    def home(cls, ip, port=0):
        rem = cls._get_remote(ip)
        res = rem.send_home_blocking()
        ok = res.get("status") == "ok"
        return ok, res.get("message", ""), res.get("message", "")

    @classmethod
    def send_key(cls, ip, key_name, port=0):
        rem = cls._get_remote(ip)
        res = rem.send_key_blocking(key_name)
        ok = res.get("status") == "ok"
        return ok, res.get("message", ""), res.get("message", "")

    @classmethod
    def cek_dan_reconnect(cls, ip, port=0):
        if not cls.adb_tersedia():
            return False, "no_adb", "androidtvremote2 tidak terinstal."
        with cls._reconnect_lock(ip):
            status_awal, _ = cls.status_untuk_ip(ip)
            if status_awal == "device":
                return True, status_awal, f"Sudah terhubung ({ip})."
            ok, msg = cls.connect(ip)
            if ok:
                return True, status_awal, "Reconnect berhasil."
            return False, status_awal, msg or "Reconnect gagal."

    # ── TV Client Features ─────────────────────────────────────────────────
    @classmethod
    def adb_shell(cls, ip, command, timeout=15, port=5555):
        """Jalankan perintah ADB shell pada TV target."""
        try:
            r = subprocess.run([cls._adb_binary(), "-s", f"{ip}:{port}", "shell", command],
                               capture_output=True, text=True, timeout=timeout,
                               **subprocess_no_window_kwargs())
            if r.returncode == 0:
                return True, r.stdout.strip()
            return False, r.stderr.strip() or r.stdout.strip()
        except subprocess.TimeoutExpired:
            return False, "Timeout"
        except FileNotFoundError:
            return False, "adb.exe tidak ditemukan"
        except Exception as e:
            return False, str(e)

    @classmethod
    def tv_power_state(cls, ip, port=5555):
        """Status nyala TV sesungguhnya: True=HIDUP, False=MATI, None=tidak terdeteksi."""
        try:
            rem = cls._instances.get(ip)
            if rem and rem.is_connected() and rem.is_on is not None:
                return bool(rem.is_on)
        except Exception:
            pass
        ok, out = cls.adb_shell(ip, "dumpsys power", timeout=8, port=port)
        if ok and out:
            for line in out.splitlines():
                if line.strip().startswith("mWakefulness="):
                    return line.strip().split("=", 1)[1].strip() == "Awake"
        return None

    @classmethod
    def send_intent(cls, ip, action, data_uri=None, extra_args=None):
        """Kirim intent Android via ADB shell."""
        cmd = f"am start -a {action}"
        if data_uri:
            cmd += f" -d \"{data_uri}\""
        if extra_args:
            cmd += " " + extra_args
        return cls.adb_shell(ip, cmd)

    @classmethod
    def send_video_url(cls, ip, url):
        """Buka URL video di TV menggunakan intent VIEW."""
        return cls.send_intent(ip, "android.intent.action.VIEW", data_uri=url)

    @classmethod
    def send_warning_broadcast(cls, ip, message, duration=30):
        """Kirim broadcast warning ke TV untuk menampilkan banner."""
        extras = f"--es message \"{message}\" --ei duration {duration}"
        return cls.send_intent(ip, "rr_billing.SHOW_WARNING", extra_args=extras)

    @classmethod
    def adb_push(cls, ip, local_path, remote_path="/sdcard/Download/"):
        """Push file ke TV via ADB."""
        try:
            r = subprocess.run([cls._adb_binary(), "-s", f"{ip}:5555", "push", local_path, remote_path],
                               capture_output=True, text=True, timeout=120,
                               **subprocess_no_window_kwargs())
            if r.returncode == 0:
                return True, r.stdout.strip()
            return False, r.stderr.strip() or r.stdout.strip()
        except subprocess.TimeoutExpired:
            return False, "Timeout — file terlalu besar"
        except FileNotFoundError:
            return False, "adb.exe tidak ditemukan"
        except Exception as e:
            return False, str(e)

    @classmethod
    def _pm_install(cls, ip, remote_path, timeout=240):
        """Jalankan pm install remote_path di TV via adb shell.
        Mengembalikan (ok, pesan)."""
        return cls.adb_shell(ip, "pm install -r " + remote_path, timeout=timeout)

    @classmethod
    def _ambil_diagnosa_avc(cls, ip, batas=25):
        """Ambil baris SELinux avc: denied dari TV (dmesg + logcat) untuk
        ditampilkan ke user — diagnosa cepat mengapa install ditolak ROM."""
        bagian = []
        try:
            ok, out = cls.adb_shell(
                ip, "logcat -d -b all -t 500 2>/dev/null | grep -aiE 'avc|installd|PackageInstaller' | tail -%d" % batas,
                timeout=25)
            if ok and out:
                bagian.append("— logcat —\n" + "\n".join(out.splitlines()[-batas:]))
        except Exception:
            pass
        try:
            ok, out = cls.adb_shell(ip, "dmesg | grep -aiE 'avc|denied' | tail -%d" % batas,
                                    timeout=20)
            if ok and out:
                bagian.append("— dmesg —\n" + "\n".join(out.splitlines()[-batas:]))
        except Exception:
            pass
        return "\n\n".join(bagian) if bagian else "(tidak ada log avc yang bisa diambil)"

    @classmethod
    def adb_install(cls, ip, apk_path):
        """Install APK ke TV dengan urutan jalur anti-SELinux:

        1. push ke /data/local/tmp (context shell_data_file — diizinkan
           hampir semua ROM) lalu `pm install -r`.
        2. kalau ROM menolak jalur tsb → push ke /sdcard/Download lalu
           `pm install -r`.
        3. kalau masih gagal dengan avc: denied → streaming adb install.
        4. tetap gagal → lampirkan diagnosa dmesg/logcat avc di pesan error.
        """
        if not os.path.isfile(apk_path):
            return False, "File APK tidak ditemukan"
        remote_tmp = "/data/local/tmp/rr_tv_client.apk"
        remote_sd = "/sdcard/Download/rr_tv_client.apk"

        # 1) /data/local/tmp
        ok, pesan = cls.adb_push(ip, apk_path, remote_tmp)
        if ok:
            ok2, pesan2 = cls._pm_install(ip, remote_tmp)
            if ok2 and "Success" in (pesan2 or ""):
                return True, pesan2.strip()
            pesan = pesan + " | pm: " + (pesan2 or "-")
        # 2) /sdcard/Download
        ok, pesan = cls.adb_push(ip, apk_path, remote_sd)
        if ok:
            ok2, pesan2 = cls._pm_install(ip, remote_sd)
            if ok2 and "Success" in (pesan2 or ""):
                return True, pesan2.strip()
            pesan = pesan + " | pm: " + (pesan2 or "-")
        # 3) streaming adb install
        try:
            r = subprocess.run([cls._adb_binary(), "-s", f"{ip}:5555", "install", "-r", apk_path],
                               capture_output=True, text=True, timeout=180,
                               **subprocess_no_window_kwargs())
            if r.returncode == 0 and "Success" in (r.stdout or ""):
                return True, r.stdout.strip()
            pesan = pesan + " | adb install: " + (r.stderr.strip() or r.stdout.strip())
        except subprocess.TimeoutExpired:
            pesan = pesan + " | adb install: Timeout"
        except FileNotFoundError:
            pesan = pesan + " | adb.exe tidak ditemukan"
        except Exception as e:
            pesan = pesan + f" | adb install: {e}"
        # 4) diagnosa avc
        if "avc" in pesan.lower() or "denied" in pesan.lower():
            diagnosa = cls._ambil_diagnosa_avc(ip)
            return False, ("Install ditolak oleh pengaman ROM (SELinux avc: denied).\n\n"
                           f"{pesan}\n\n{diagnosa}")
        return False, pesan or "Instalasi ditolak TV (hasil tidak jelas)"

    @classmethod
    def adb_install_with_progress(cls, ip, apk_path, progress_cb=None):
        """Install APK ke TV via ADB dengan status persentase.

        Fase 1: push APK (0–80%) — persen di-parse dari output streaming adb.
        Fase 2: pm install (85–100%) — tanpa persen, ditandai indeterminate.
        progress_cb(persen: int, pesan: str) dipanggil dari thread ini.

        Jalur target: /data/local/tmp dulu (anti avc: denied), fallback
        /sdcard/Download, lalu streaming adb install.
        """
        def _set(pct, pesan):
            if progress_cb:
                try:
                    progress_cb(pct, pesan)
                except Exception:
                    pass

        if not os.path.isfile(apk_path):
            return False, "File APK tidak ditemukan"

        jalur_gagal = []
        for remote in ("/data/local/tmp/rr_tv_client.apk",
                       "/sdcard/Download/rr_tv_client.apk"):
            _set(2, f"Mengirim APK ke TV ({remote})…")
            try:
                proc = subprocess.Popen(
                    [cls._adb_binary(), "-s", f"{ip}:5555", "push", apk_path, remote],
                    stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                    text=True, encoding="utf-8", errors="replace",
                    **subprocess_no_window_kwargs())
            except FileNotFoundError:
                return False, "adb.exe tidak ditemukan"
            except Exception as e:
                return False, str(e)
            try:
                for line in iter(proc.stdout.readline, ""):
                    m = re.search(r"\[\s*(\d+)\s*%\]", line or "")
                    if m:
                        try:
                            pct = max(0, min(100, int(m.group(1))))
                            _set(pct * 8 // 10, f"Mengirim APK ke TV… {pct}%")
                        except Exception:
                            pass
                proc.wait(timeout=120)
            except Exception:
                try:
                    proc.kill()
                except Exception:
                    pass
                return False, "Timeout — file APK terlalu besar"
            if proc.returncode != 0:
                jalur_gagal.append(f"{remote}: push ditolak (SELinux?)")
                continue
            _set(85, "Memasang APK… mohon tunggu (bisa 1–3 menit).")
            ok, out = cls._pm_install(ip, remote)
            if ok and "Success" in (out or ""):
                _set(100, "✅ Selesai")
                return True, out.strip()
            jalur_gagal.append(f"{remote}: {out or 'pm install ditolak'}")

        # Fallback streaming adb install
        _set(86, "Mencoba jalur streaming…")
        try:
            r = subprocess.run([cls._adb_binary(), "-s", f"{ip}:5555",
                                "install", "-r", apk_path],
                               capture_output=True, text=True, timeout=180,
                               **subprocess_no_window_kwargs())
            if r.returncode == 0 and "Success" in (r.stdout or ""):
                _set(100, "✅ Selesai")
                return True, r.stdout.strip()
            jalur_gagal.append("streaming: " + (r.stderr.strip() or r.stdout.strip()))
        except FileNotFoundError:
            return False, "adb.exe tidak ditemukan"
        except Exception as e:
            jalur_gagal.append(f"streaming: {e}")

        pesan = "\n".join(jalur_gagal)
        if "avc" in pesan.lower() or "denied" in pesan.lower():
            diagnosa = cls._ambil_diagnosa_avc(ip)
            return False, ("Install ditolak oleh pengaman ROM (SELinux avc: denied).\n\n"
                           f"{pesan}\n\n{diagnosa}")
        return False, pesan or "Instalasi ditolak TV (hasil tidak jelas)"

    @classmethod
    def ping_host(cls, host, port=5555, timeout=3):
        """Cek apakah host TV reachable via TCP socket."""
        try:
            with socket.create_connection((host, port), timeout=timeout):
                return True
        except Exception:
            return False

    @classmethod
    def get_connected_tvs(cls):
        """Kembalikan daftar IP TV yang sedang terhubung untuk dropdown."""
        return list(cls._instances.keys())


class DialogWarnetAdminCode(ctk.CTkToplevel):
    """Dialog kecil untuk generate kode admin pengaturan client dari IP."""

    def __init__(self, master):
        super().__init__(master)
        self.title("Generator Kode Pengaturan Client")
        self.geometry("470x260")
        self.configure(fg_color=C_BG)
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()

        ctk.CTkLabel(
            self,
            text="🔐 Generator Kode Pengaturan Client",
            font=FONT_TITLE,
            text_color=C_ACCENT,
        ).pack(pady=(14, 8))

        ctk.CTkLabel(
            self,
            text="Masukkan IP PC client, lalu generate kode untuk buka tombol ⚙ di app client.",
            font=FONT_SMALL,
            text_color=C_MUTED,
        ).pack(pady=(0, 10))

        form = ctk.CTkFrame(self, fg_color=C_CARD, corner_radius=10)
        form.pack(fill="x", padx=16, pady=6)
        form.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(form, text="IP Client", font=FONT_LABEL, text_color=C_MUTED).grid(row=0, column=0, padx=10, pady=10, sticky="w")
        self.ip_var = tk.StringVar()
        self.entry_ip = ctk.CTkEntry(form, textvariable=self.ip_var)
        self.entry_ip.grid(row=0, column=1, padx=10, pady=10, sticky="ew")

        ctk.CTkLabel(form, text="Kode Hari Ini", font=FONT_LABEL, text_color=C_MUTED).grid(row=1, column=0, padx=10, pady=(0, 10), sticky="w")
        self.code_var = tk.StringVar(value="-")
        self.entry_code = ctk.CTkEntry(form, textvariable=self.code_var, state="readonly")
        self.entry_code.grid(row=1, column=1, padx=10, pady=(0, 10), sticky="ew")

        btns = ctk.CTkFrame(self, fg_color="transparent")
        btns.pack(fill="x", padx=16, pady=(8, 6))
        ctk.CTkButton(btns, text="Generate", fg_color=C_ACCENT2, hover_color="#5A0FCC", command=self._generate).pack(side="left")
        ctk.CTkButton(btns, text="Copy Kode", fg_color=C_GREEN, hover_color="#2F7A2F", command=self._copy_code).pack(side="left", padx=8)
        ctk.CTkButton(btns, text="Tutup", fg_color=C_RED, hover_color="#7A1A1A", command=self.destroy).pack(side="right")

        self.info = ctk.CTkLabel(self, text="", font=FONT_SMALL, text_color=C_MUTED)
        self.info.pack(fill="x", padx=16, pady=(0, 8))
        self.entry_ip.focus_set()

    def _generate(self):
        ip = self.ip_var.get().strip()
        if not ip:
            messagebox.showwarning("IP Belum Diisi", "Masukkan IP client terlebih dahulu.", parent=self)
            return
        code = generate_warnet_admin_code(ip)
        if not code:
            messagebox.showwarning("Gagal Generate", "Kode gagal dibuat. Periksa IP.", parent=self)
            return
        self.code_var.set(code)
        self.info.configure(text=f"IP: {ip} | Tanggal: {datetime.now().strftime('%d-%m-%Y')}")

    def _copy_code(self):
        code = self.code_var.get().strip()
        if not code or code == "-":
            messagebox.showwarning("Belum Ada Kode", "Generate kode terlebih dahulu.", parent=self)
            return
        self.clipboard_clear()
        self.clipboard_append(code)
        self.update()
        messagebox.showinfo("Berhasil", "Kode berhasil disalin ke clipboard.", parent=self)


# ═══════════════════════════════════════════════════════════════════════════════
#  DIALOG OVERLAY SETTING (TV)
# ═══════════════════════════════════════════════════════════════════════════════

class DialogOverlaySetting(ctk.CTkToplevel):
    """Pengaturan tampilan overlay countdown di TV (mode + menit terakhir)."""

    MODE_OPTIONS = [
        ("always", "Selalu tampil"),
        ("last_minutes", "Hanya N menit terakhir"),
        ("hidden", "Sembunyi (tidak tampil)"),
    ]

    def __init__(self, master, mode="always", last_minutes=5, on_save=None):
        super().__init__(master)
        self.on_save = on_save
        self.title("Pengaturan Overlay TV")
        self.geometry("480x300")
        self.configure(fg_color=C_BG)
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()

        ctk.CTkLabel(self, text="⚙️  OVERLAY COUNTDOWN TV",
                     font=FONT_TITLE, text_color=C_ACCENT).pack(pady=(16, 4))
        ctk.CTkLabel(self, text="Kapan widget waktu di pojok kanan atas TV ditampilkan?",
                     font=FONT_SMALL, text_color=C_MUTED).pack(pady=(0, 10))

        self.mode_var = tk.StringVar(value=mode)
        for value, label in self.MODE_OPTIONS:
            ctk.CTkRadioButton(self, text=label, variable=self.mode_var, value=value,
                               font=FONT_BODY, fg_color=C_ACCENT2,
                               command=self._on_mode_change).pack(anchor="w", padx=30, pady=4)

        row = ctk.CTkFrame(self, fg_color="transparent")
        row.pack(fill="x", padx=30, pady=(8, 4))
        ctk.CTkLabel(row, text="Tampil sejak sisa waktu (menit):", font=FONT_BODY,
                     text_color=C_TEXT).pack(side="left")
        self.minutes_var = tk.StringVar(value=str(last_minutes))
        self.entry_minutes = ctk.CTkEntry(row, textvariable=self.minutes_var,
                                          width=70, height=32,
                                          font=("Consolas", 13, "bold"))
        self.entry_minutes.pack(side="left", padx=10)
        self._on_mode_change()

        btns = ctk.CTkFrame(self, fg_color="transparent")
        btns.pack(fill="x", padx=30, pady=(10, 14))
        ctk.CTkButton(btns, text="✖ Batal", width=110, height=36,
                      fg_color=C_RED, hover_color="#7A1A1A",
                      font=FONT_SUB, command=self.destroy).pack(side="right")
        ctk.CTkButton(btns, text="✅ Simpan", width=130, height=36,
                      fg_color=C_GREEN, hover_color="#2F7A2F",
                      font=FONT_SUB, command=self._save).pack(side="right", padx=(0, 8))

    def _on_mode_change(self):
        state = "normal" if self.mode_var.get() == "last_minutes" else "disabled"
        self.entry_minutes.configure(state=state)

    def _save(self):
        mode = self.mode_var.get()
        try:
            minutes = max(1, min(60, int(self.minutes_var.get().strip() or 5)))
        except ValueError:
            messagebox.showwarning("Input Salah", "Jumlah menit harus angka 1–60.", parent=self)
            return
        if callable(self.on_save):
            self.on_save(mode, minutes)
        self.destroy()


# ═══════════════════════════════════════════════════════════════════════════════
#  DIALOG INSTALL APK
# ═══════════════════════════════════════════════════════════════════════════════

class DialogInstallAPK(ctk.CTkToplevel):
    """Dialog untuk install APK billing client ke Android TV via ADB."""

    def __init__(self, master, config_manager):
        super().__init__(master)
        self.config_manager = config_manager
        self.install_thread = None
        self._stop_flag = False

        self.title("📱 Install APK ke Android TV")
        self.geometry("560x520")
        self.configure(fg_color=C_BG)
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()

        self._build_ui()

    def _build_ui(self):
        title_frame = ctk.CTkFrame(self, fg_color=C_PANEL, corner_radius=0)
        title_frame.pack(fill="x")
        ctk.CTkLabel(title_frame, text="📱  INSTALL APK CLIENT VIA ADB",
                     font=FONT_TITLE, text_color=C_ACCENT).pack(side="left", padx=18, pady=12)

        # ── APK file selection ──
        file_frame = ctk.CTkFrame(self, fg_color=C_CARD, corner_radius=10)
        file_frame.pack(fill="x", padx=14, pady=(10, 4))
        file_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(file_frame, text="File APK", font=FONT_LABEL,
                     text_color=C_MUTED).grid(row=0, column=0, padx=10, pady=8, sticky="w")
        self.apk_var = tk.StringVar()
        ctk.CTkEntry(file_frame, textvariable=self.apk_var).grid(
            row=0, column=1, padx=6, pady=8, sticky="ew")
        ctk.CTkButton(file_frame, text="Browse", width=80, height=28,
                      fg_color=C_BTN, command=self._browse_apk
                      ).grid(row=0, column=2, padx=(0, 10), pady=8)

        # ── Target TV selection ──
        tv_frame = ctk.CTkFrame(self, fg_color=C_CARD, corner_radius=10)
        tv_frame.pack(fill="x", padx=14, pady=4)
        tv_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(tv_frame, text="Target TV", font=FONT_LABEL,
                     text_color=C_MUTED).grid(row=0, column=0, padx=10, pady=8, sticky="w")
        connected = ADBHelper.get_connected_tvs()
        if not connected:
            connected = ["Tidak ada TV terhubung"]
        self.tv_var = tk.StringVar(value=connected[0])
        self.tv_menu = ctk.CTkOptionMenu(tv_frame, values=connected,
                                          variable=self.tv_var,
                                          fg_color=C_BTN, button_color=C_ACCENT2,
                                          button_hover_color="#5A0FCC",
                                          text_color=C_TEXT,
                                          font=("Consolas", 11),
                                          dropdown_font=("Consolas", 10))
        self.tv_menu.grid(row=0, column=1, padx=6, pady=8, sticky="ew")
        ctk.CTkButton(tv_frame, text="🔄 Refresh", width=90, height=28,
                      fg_color=C_BTN, command=self._refresh_tv_list
                      ).grid(row=0, column=2, padx=(0, 10), pady=8)

        # ── Action buttons ──
        action_frame = ctk.CTkFrame(self, fg_color="transparent")
        action_frame.pack(fill="x", padx=14, pady=(6, 2))
        self.btn_install = ctk.CTkButton(action_frame, text="📱 Install APK",
                                          fg_color=C_ACCENT2, hover_color="#5A0FCC",
                                          height=36, font=("Russo One", 10, "bold"),
                                          command=self._start_install)
        self.btn_install.pack(side="left")
        ctk.CTkButton(action_frame, text="✖ Tutup", width=100, height=36,
                      fg_color=C_RED, font=("Russo One", 10, "bold"),
                      text_color="white", command=self.destroy
                      ).pack(side="right")

        # ── Log output ──
        log_frame = ctk.CTkFrame(self, fg_color=C_CARD, corner_radius=10)
        log_frame.pack(fill="both", expand=True, padx=14, pady=4)
        log_frame.grid_columnconfigure(0, weight=1)
        log_frame.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(log_frame, text="Log Install", font=FONT_SUB,
                     text_color=C_TEXT).grid(row=0, column=0, sticky="w", padx=10, pady=(6, 0))
        self.txt_log = ctk.CTkTextbox(log_frame, height=120, fg_color=C_BG,
                                      font=("Consolas", 10))
        self.txt_log.grid(row=1, column=0, sticky="nsew", padx=8, pady=(2, 6))

        self.progress = ctk.CTkProgressBar(self, height=6, fg_color=C_BTN,
                                           progress_color=C_ACCENT)
        self.progress.pack(fill="x", padx=14, pady=(2, 10))
        self.progress.set(0)

    def _browse_apk(self):
        path = filedialog.askopenfilename(
            parent=self, title="Pilih file APK",
            filetypes=[("APK", "*.apk"), ("All files", "*.*")]
        )
        if path:
            self.apk_var.set(path)

    def _refresh_tv_list(self):
        connected = ADBHelper.get_connected_tvs()
        if not connected:
            connected = ["Tidak ada TV terhubung"]
        self.tv_menu.configure(values=connected)
        self.tv_var.set(connected[0])
        self._append_log("🔄 Daftar TV diperbarui", C_ACCENT)

    def _append_log(self, text, color=C_TEXT):
        self.after(0, lambda: self.txt_log.insert("end", text + "\n"))
        self.after(0, lambda: self.txt_log.see("end"))

    def _start_install(self):
        if self.install_thread and self.install_thread.is_alive():
            messagebox.showwarning("Sedang Berjalan", "Proses install sedang berlangsung.", parent=self)
            return

        apk_path = self.apk_var.get().strip()
        if not apk_path or not os.path.exists(apk_path):
            messagebox.showerror("File Tidak Ditemukan",
                                 f"APK tidak ditemukan:\n{apk_path}", parent=self)
            return

        target_ip = self.tv_var.get().strip()
        if not target_ip or target_ip == "Tidak ada TV terhubung":
            messagebox.showwarning("TV Tidak Valid", "Pilih TV target terlebih dahulu.", parent=self)
            return

        self.btn_install.configure(state="disabled", text="Menginstall...")
        self.txt_log.delete("1.0", "end")
        self.progress.set(0)
        self._stop_flag = False

        self.install_thread = threading.Thread(
            target=self._install_worker,
            args=(target_ip, apk_path),
            daemon=True
        )
        self.install_thread.start()

    def _install_worker(self, ip, apk_path):
        fname = os.path.basename(apk_path)
        self._append_log(f"[{datetime.now().strftime('%H:%M:%S')}] 🚀 Install {fname} ke {ip}...", C_TEXT)

        # ADBHelper.adb_install sudah multi-jalur: /data/local/tmp →
        # /sdcard/Download → streaming — dengan diagnosa avc bila ditolak.
        self._append_log(f"[{datetime.now().strftime('%H:%M:%S')}] 📲 Install APK (mohon tunggu)...", C_YELLOW)
        ok, msg = ADBHelper.adb_install(ip, apk_path)
        if ok:
            self._append_log(f"[{datetime.now().strftime('%H:%M:%S')}] ✅ Install berhasil!", C_GREEN)
            self.after(0, lambda: self.progress.set(1))
        else:
            self._append_log(f"[{datetime.now().strftime('%H:%M:%S')}] ❌ Install gagal: {msg}", C_RED)
            self.after(0, lambda: self.progress.set(0))

        self.after(0, lambda: self.btn_install.configure(state="normal", text="📱 Install APK"))


# ═══════════════════════════════════════════════════════════════════════════════
#  DIALOG DEPLOY CLIENT
# ═══════════════════════════════════════════════════════════════════════════════

class DialogDeployClient(ctk.CTkToplevel):
    """Dialog untuk deploy/update client app ke PC warnet via SSH.
    Paket client = 3 exe + INSTALL_CLIENT.bat (+ rr_billing_config.json opsional)."""

    PKG_HINTS = [
        os.path.join(APP_BASE_DIR, "RRBillingPro_Client_Package"),
        os.path.join(APP_BASE_DIR, "BillingClientCSharp", "dist"),
    ]

    def __init__(self, master, warnet_server, config_manager):
        super().__init__(master)
        self.warnet_server = warnet_server
        self.config_manager = config_manager
        self.pc_list = []
        self.selected_indices = set()
        self.deploy_thread = None
        self._stop_flag = False

        self.title("🚀 Deploy Client ke PC Warnet")
        self.geometry("640x620")
        self.configure(fg_color=C_BG)
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()

        self._build_ui()
        self._load_pc_list()
        self._auto_detect_pkg()

    def _build_ui(self):
        # Title
        title_frame = ctk.CTkFrame(self, fg_color=C_PANEL, corner_radius=0)
        title_frame.pack(fill="x")
        ctk.CTkLabel(title_frame, text="🚀  DEPLOY CLIENT APLIKASI",
                     font=FONT_TITLE, text_color=C_ACCENT).pack(side="left", padx=18, pady=12)

        # ── Paket selection ─────────────────────────────────────────────
        file_frame = ctk.CTkFrame(self, fg_color=C_CARD, corner_radius=10)
        file_frame.pack(fill="x", padx=14, pady=(10, 4))
        file_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(file_frame, text="Folder Paket", font=FONT_LABEL,
                     text_color=C_MUTED).grid(row=0, column=0, padx=10, pady=8, sticky="w")
        self.exe_var = tk.StringVar()
        self.entry_exe = ctk.CTkEntry(file_frame, textvariable=self.exe_var)
        self.entry_exe.grid(row=0, column=1, padx=6, pady=8, sticky="ew")
        ctk.CTkButton(file_frame, text="Browse", width=80, height=28,
                      fg_color=C_BTN, command=self._browse_pkg
                      ).grid(row=0, column=2, padx=(0, 10), pady=8)

        self.var_keep_config = tk.BooleanVar(value=True)
        ctk.CTkCheckBox(file_frame, text="Pertahankan rr_billing_config.json di client "
                                          "(jangan timpa saat update)",
                        variable=self.var_keep_config, font=FONT_SMALL,
                        text_color=C_MUTED).grid(row=1, column=1, columnspan=2,
                                                 padx=6, pady=(0, 8), sticky="w")

        # ── SSH credentials ───────────────────────────────────────────────
        ssh_frame = ctk.CTkFrame(self, fg_color=C_CARD, corner_radius=10)
        ssh_frame.pack(fill="x", padx=14, pady=4)
        ssh_frame.grid_columnconfigure(1, weight=1)
        ssh_frame.grid_columnconfigure(3, weight=1)

        ctk.CTkLabel(ssh_frame, text="SSH User", font=FONT_LABEL,
                     text_color=C_MUTED).grid(row=0, column=0, padx=10, pady=8, sticky="w")
        self.ssh_user_var = tk.StringVar(value="Administrator")
        ctk.CTkEntry(ssh_frame, textvariable=self.ssh_user_var).grid(
            row=0, column=1, padx=6, pady=8, sticky="ew")

        ctk.CTkLabel(ssh_frame, text="Password", font=FONT_LABEL,
                     text_color=C_MUTED).grid(row=0, column=2, padx=10, pady=8, sticky="w")
        self.ssh_pass_var = tk.StringVar()
        ctk.CTkEntry(ssh_frame, textvariable=self.ssh_pass_var, show="*").grid(
            row=0, column=3, padx=6, pady=8, sticky="ew")

        # ── PC target list ────────────────────────────────────────────────
        list_frame = ctk.CTkFrame(self, fg_color=C_CARD, corner_radius=10)
        list_frame.pack(fill="both", expand=True, padx=14, pady=4)
        list_frame.grid_columnconfigure(0, weight=1)
        list_frame.grid_rowconfigure(1, weight=1)

        top_row = ctk.CTkFrame(list_frame, fg_color="transparent")
        top_row.grid(row=0, column=0, sticky="ew", padx=8, pady=(6, 2))
        ctk.CTkLabel(top_row, text="Daftar PC Target", font=FONT_SUB,
                     text_color=C_TEXT).pack(side="left")
        self.lbl_selected = ctk.CTkLabel(top_row, text="0 terpilih",
                                         font=FONT_SMALL, text_color=C_MUTED)
        self.lbl_selected.pack(side="right")

        # Scrollable PC list
        self.pc_scroll = ctk.CTkScrollableFrame(list_frame, fg_color=C_BG,
                                                 corner_radius=8)
        self.pc_scroll.grid(row=1, column=0, sticky="nsew", padx=8, pady=2)

        # Action buttons
        action_frame = ctk.CTkFrame(list_frame, fg_color="transparent")
        action_frame.grid(row=2, column=0, sticky="ew", padx=8, pady=(4, 6))
        ctk.CTkButton(action_frame, text="Select All", width=90, height=28,
                      fg_color=C_BTN, command=self._select_all
                      ).pack(side="left", padx=(0, 6))
        ctk.CTkButton(action_frame, text="Deselect All", width=100, height=28,
                      fg_color=C_BTN, command=self._deselect_all
                      ).pack(side="left", padx=(0, 6))
        self.btn_deploy = ctk.CTkButton(action_frame, text="Deploy ke (0) PC",
                                        fg_color=C_ACCENT2, hover_color="#5A0FCC",
                                        height=32, command=self._start_deploy)
        self.btn_deploy.pack(side="right")
        self.btn_config = ctk.CTkButton(action_frame, text="⚙️ Config Client",
                                        fg_color=C_BTN, hover_color="#3A3A2A",
                                        border_width=1, border_color=C_YELLOW,
                                        text_color=C_YELLOW,
                                        height=32, command=self._open_config_client)
        self.btn_config.pack(side="right", padx=(0, 6))

        # ── Log output ────────────────────────────────────────────────────
        log_frame = ctk.CTkFrame(self, fg_color=C_CARD, corner_radius=10)
        log_frame.pack(fill="x", padx=14, pady=4)
        log_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(log_frame, text="Log", font=FONT_SUB,
                     text_color=C_TEXT).grid(row=0, column=0, sticky="w", padx=10, pady=(6, 0))
        self.txt_log = ctk.CTkTextbox(log_frame, height=80, fg_color=C_BG,
                                      font=("Consolas", 10))
        self.txt_log.grid(row=1, column=0, sticky="ew", padx=8, pady=(2, 6))

        # ── Progress bar ──────────────────────────────────────────────────
        self.progress = ctk.CTkProgressBar(self, height=6, fg_color=C_BTN,
                                           progress_color=C_ACCENT)
        self.progress.pack(fill="x", padx=14, pady=(2, 10))
        self.progress.set(0)

    def _auto_detect_pkg(self):
        for p in self.PKG_HINTS:
            if os.path.isdir(p):
                self.exe_var.set(p)
                return

    def _browse_pkg(self):
        path = filedialog.askdirectory(
            parent=self,
            title="Pilih folder paket client (berisi 3 exe + INSTALL_CLIENT.bat)",
            initialdir=os.path.dirname(self.exe_var.get() or os.path.abspath(__file__))
        )
        if path:
            self.exe_var.set(path)

    def _load_pc_list(self):
        """Load daftar PC dari warnet_server sessions + config."""
        seen = set()
        self.pc_list = []

        server = self.warnet_server
        active_sessions = {}
        if server and getattr(server, 'running', False):
            with server.sessions_lock:
                active_sessions = {
                    sid: s.get("client_id", "")
                    for sid, s in server.sessions.items()
                }

        cfg = self.config_manager.load()
        warnet_clients = cfg.get("warnet_clients", [])

        for client in warnet_clients:
            cid = client.get("client_id", "")
            is_active = cid in active_sessions.values()
            for pc in client.get("pcs", []):
                ip = pc.get("ip", "")
                if ip and ip not in seen:
                    seen.add(ip)
                    self.pc_list.append({
                        "ip": ip,
                        "name": pc.get("name", ""),
                        "client_id": cid,
                        "active": is_active,
                    })

        self._render_pc_list()

    def _render_pc_list(self):
        for w in self.pc_scroll.winfo_children():
            w.destroy()
        self.selected_indices.clear()
        self._pc_checkboxes = []

        if not self.pc_list:
            ctk.CTkLabel(self.pc_scroll, text="Tidak ada PC client terdaftar.\n"
                         "Tambah PC di config atau tunggu client connect.",
                         font=FONT_BODY, text_color=C_MUTED).pack(pady=20)
            return

        for i, pc in enumerate(self.pc_list):
            ip = pc["ip"]
            name = pc["name"]
            cid = pc["client_id"]
            active = pc["active"]

            frame = ctk.CTkFrame(self.pc_scroll, fg_color=C_BTN if not active else C_CARD,
                                 corner_radius=6)
            frame.pack(fill="x", padx=4, pady=2)

            var = tk.BooleanVar(value=active)
            cb = ctk.CTkCheckBox(frame, text="", variable=var,
                                 command=lambda idx=i: self._toggle_select(idx))
            cb.pack(side="left", padx=(8, 4), pady=6)
            self._pc_checkboxes.append((var, cb))

            info = f"{ip}  —  {name}  ({cid})"
            lbl = ctk.CTkLabel(frame, text=info, font=("Consolas", 11),
                               anchor="w", text_color=C_TEXT)
            lbl.pack(side="left", fill="x", expand=True, padx=4, pady=6)

            status_text = "● Connected" if active else "○ Disconnected"
            status_color = C_GREEN if active else C_MUTED
            ctk.CTkLabel(frame, text=status_text, font=FONT_SMALL,
                         text_color=status_color).pack(side="right", padx=(4, 10), pady=6)

            if active:
                self.selected_indices.add(i)
                var.set(True)

        self._update_selected_count()

    def _toggle_select(self, idx):
        var, _ = self._pc_checkboxes[idx]
        if var.get():
            self.selected_indices.add(idx)
        else:
            self.selected_indices.discard(idx)
        self._update_selected_count()

    def _select_all(self):
        for i, (var, cb) in enumerate(self._pc_checkboxes):
            var.set(True)
            self.selected_indices.add(i)
        self._update_selected_count()

    def _deselect_all(self):
        for i, (var, cb) in enumerate(self._pc_checkboxes):
            var.set(False)
            self.selected_indices.discard(i)
        self._update_selected_count()

    def _update_selected_count(self):
        n = len(self.selected_indices)
        self.lbl_selected.configure(text=f"{n} terpilih")
        self.btn_deploy.configure(text=f"Deploy ke ({n}) PC")

    def _append_log(self, text, color=C_TEXT):
        tag = self.txt_log.index("end-1c")
        self.txt_log.insert("end", text + "\n")
        self.txt_log.see("end")
        self.update()

    def _start_deploy(self):
        if self.deploy_thread and self.deploy_thread.is_alive():
            messagebox.showwarning("Sedang Berjalan", "Proses deploy sedang berlangsung.", parent=self)
            return

        pkg_path = self.exe_var.get().strip()
        if not pkg_path or not os.path.isdir(pkg_path):
            messagebox.showerror("Folder Tidak Ditemukan",
                                 f"Folder paket tidak ditemukan:\n{pkg_path}", parent=self)
            return

        if not self.selected_indices:
            messagebox.showwarning("Tidak Ada PC Dipilih",
                                   "Pilih minimal 1 PC target.", parent=self)
            return

        ssh_user = self.ssh_user_var.get().strip()
        ssh_pass = self.ssh_pass_var.get()
        if not ssh_user or not ssh_pass:
            messagebox.showerror("SSH Login Diperlukan",
                                 "Isi username dan password SSH.", parent=self)
            return

        targets = [self.pc_list[i]["ip"] for i in sorted(self.selected_indices)]
        self.btn_deploy.configure(state="disabled", text="Deploying...")
        self.txt_log.delete("1.0", "end")
        self.progress.set(0)
        self._stop_flag = False

        keep_config = self.var_keep_config.get()
        self.deploy_thread = threading.Thread(
            target=self._deploy_worker,
            args=(targets, pkg_path, ssh_user, ssh_pass, keep_config),
            daemon=True
        )
        self.deploy_thread.start()

    def _open_config_client(self):
        DialogConfigClient(self)

    def _deploy_worker(self, targets, pkg_path, ssh_user, ssh_pass, keep_config):
        # Import here to avoid circular import at module level
        from scripts.deploy_manager import DeployManager

        dm = DeployManager(ssh_username=ssh_user, ssh_password=ssh_pass, timeout=15)
        total = len(targets)
        completed = 0

        for ip in targets:
            if self._stop_flag:
                self.after(0, lambda: self._append_log(f"⛔ Dibatalakan oleh user.", C_YELLOW))
                break

            self.after(0, lambda i=ip: self._append_log(f"[{datetime.now().strftime('%H:%M:%S')}] "
                         f"🚀 Deploy ke {i}...", C_TEXT))
            self.after(0, lambda: self.update())

            # Test SSH first
            ok, msg = dm.test_connection(ip)
            if not ok:
                self.after(0, lambda i=ip, m=msg: self._append_log(
                    f"[{datetime.now().strftime('%H:%M:%S')}] ❌ {i}: SSH gagal — {m}", C_RED))
                completed += 1
                self.after(0, lambda c=completed, t=total: self.progress.set(c / t))
                continue

            # Deploy paket lengkap (stop service -> kill -> copy -> install -> start)
            result = dm.deploy_package(ip, pkg_path, keep_config=keep_config)
            if result.get("success"):
                self.after(0, lambda i=ip: self._append_log(
                    f"[{datetime.now().strftime('%H:%M:%S')}] ✅ {i}: {result.get('message', 'Deploy selesai')}", C_GREEN))
                for step_name, ok_step, step_msg in result.get("steps", []):
                    self.after(0, lambda s=step_name, m=step_msg: self._append_log(
                        f"       · {s}: {m}", C_MUTED))
            else:
                self.after(0, lambda i=ip, m=result.get("message", ""): self._append_log(
                    f"[{datetime.now().strftime('%H:%M:%S')}] ❌ {i}: {m}", C_RED))

            completed += 1
            self.after(0, lambda c=completed, t=total: self.progress.set(c / t))

        self.after(0, lambda: self._append_log(
            f"[{datetime.now().strftime('%H:%M:%S')}] ✅ Selesai — {completed}/{total} PC", C_ACCENT))
        self.after(0, lambda: self.btn_deploy.configure(
            state="normal", text=f"Deploy ke ({len(targets)}) PC"))


# ═══════════════════════════════════════════════════════════════════════════════
#  CONFIG CLIENT (WARNET) — dialog kelola client & PC + generate paket client
# ═══════════════════════════════════════════════════════════════════════════════
class DialogConfigClient(ctk.CTkToplevel):
    """Dialog untuk mengelola warnet_clients di config dan menyiapkan
    paket client (3 exe + rr_billing_config.json) siap-copy ke PC warnet."""

    DEPLOY_EXE_DIR = os.path.join(APP_BASE_DIR, "BillingClientCSharp", "dist")
    DEPLOY_OUT_DIR = os.path.join(APP_BASE_DIR, "deploy_warnet")
    SERVER_PORT = 5000

    def __init__(self, master):
        super().__init__(master)
        self.config_manager = ConfigManager
        self.title("⚙️ Config Client")
        self.geometry("660x680")
        self.configure(fg_color=C_BG)
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()

        self._clients = []
        self._gen_pcs = []
        self._load_from_config()

        self._build_ui()
        self._refresh_client_combo()
        self._refresh_pc_combo()
        self._refresh_gen_combo()

    # ── Load / save config ──────────────────────────────────────────────
    def _load_from_config(self):
        cfg = self.config_manager.load()
        self._clients = cfg.get("warnet_clients", []) or []
        for c in self._clients:
            c.setdefault("pcs", [])
        self._server_ip = cfg.get("warnet_server_host", "") or _detect_local_ip()

    def _simpan_config(self):
        cfg = self.config_manager.load()
        server_ip = self.entry_server_ip.get().strip()
        if server_ip:
            cfg["warnet_server_host"] = server_ip
        cfg["warnet_clients"] = self._clients
        self.config_manager.save(cfg)
        self.lbl_status.configure(
            text="💾 Tersimpan ke rr_billing_config.json", text_color=C_YELLOW)

    def _find_client(self, client_id):
        for c in self._clients:
            if c.get("client_id") == client_id:
                return c
        return None

    # ── UI ──────────────────────────────────────────────────────────────
    def _build_ui(self):
        title_frame = ctk.CTkFrame(self, fg_color=C_PANEL, corner_radius=0)
        title_frame.pack(fill="x")
        ctk.CTkLabel(title_frame, text="⚙️  CONFIG CLIENT",
                     font=FONT_TITLE, text_color=C_YELLOW).pack(side="left", padx=18, pady=12)

        # A. INFO SERVER
        info_frame = ctk.CTkFrame(self, fg_color=C_CARD, corner_radius=10)
        info_frame.pack(fill="x", padx=14, pady=(10, 4))
        info_frame.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(info_frame, text="🌐 INFO SERVER", font=FONT_SUB,
                     text_color=C_YELLOW).grid(row=0, column=0, columnspan=3,
                                               sticky="w", padx=10, pady=(8, 2))
        ctk.CTkLabel(info_frame, text="IP Server", font=FONT_LABEL,
                     text_color=C_YELLOW).grid(row=1, column=0, padx=10, pady=6, sticky="w")
        self.entry_server_ip = ctk.CTkEntry(info_frame)
        self.entry_server_ip.insert(0, self._server_ip)
        self.entry_server_ip.grid(row=1, column=1, columnspan=2, padx=6, pady=6, sticky="ew")
        ctk.CTkLabel(info_frame, text=f"Port: {self.SERVER_PORT}", font=FONT_LABEL,
                     text_color=C_YELLOW).grid(row=2, column=0, padx=10, pady=6, sticky="w")
        ctk.CTkLabel(info_frame, text="Nilai IP + Port di atas yang harus diisi "
                                      "di config client tiap PC warnet.",
                     font=FONT_SMALL, text_color=C_YELLOW).grid(
            row=3, column=0, columnspan=3, sticky="w", padx=10, pady=(2, 8))

        # B. KELOLA CLIENT & PC
        kelola_frame = ctk.CTkFrame(self, fg_color=C_CARD, corner_radius=10)
        kelola_frame.pack(fill="both", expand=True, padx=14, pady=4)
        kelola_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(kelola_frame, text="📋 KELOLA CLIENT & PC", font=FONT_SUB,
                     text_color=C_YELLOW).grid(row=0, column=0, sticky="w",
                                               padx=10, pady=(8, 2))
        self.cb_client = ctk.CTkComboBox(kelola_frame, values=[""], width=300,
                                         state="readonly", command=self._on_pick_client)
        self.cb_client.grid(row=1, column=0, sticky="w", padx=10, pady=4)
        self.cb_pc = ctk.CTkComboBox(kelola_frame, values=[""], width=300,
                                     state="readonly")
        self.cb_pc.grid(row=2, column=0, sticky="w", padx=10, pady=4)

        btn_row = ctk.CTkFrame(kelola_frame, fg_color="transparent")
        btn_row.grid(row=3, column=0, sticky="ew", padx=10, pady=4)
        ctk.CTkButton(btn_row, text="➕ Tambah Client", width=120, height=30,
                      fg_color=C_BTN, hover_color="#3A3A2A", border_width=1,
                      border_color=C_YELLOW, text_color=C_YELLOW,
                      command=self._tambah_client).pack(side="left", padx=(0, 6))
        ctk.CTkButton(btn_row, text="🔑 Ganti Password", width=130, height=30,
                      fg_color=C_BTN, hover_color="#3A3A2A", border_width=1,
                      border_color=C_YELLOW, text_color=C_YELLOW,
                      command=self._ganti_password).pack(side="left", padx=(0, 6))
        ctk.CTkButton(btn_row, text="🗑 Hapus Client", width=115, height=30,
                      fg_color=C_BTN, hover_color="#3A3A2A", border_width=1,
                      border_color=C_YELLOW, text_color=C_YELLOW,
                      command=self._hapus_client).pack(side="left", padx=(0, 6))
        ctk.CTkButton(btn_row, text="➕ Tambah PC", width=105, height=30,
                      fg_color=C_BTN, hover_color="#3A3A2A", border_width=1,
                      border_color=C_YELLOW, text_color=C_YELLOW,
                      command=self._tambah_pc).pack(side="left", padx=(0, 6))
        ctk.CTkButton(btn_row, text="🗑 Hapus PC", width=95, height=30,
                      fg_color=C_BTN, hover_color="#3A3A2A", border_width=1,
                      border_color=C_YELLOW, text_color=C_YELLOW,
                      command=self._hapus_pc).pack(side="left", padx=(0, 6))

        # C. GENERATE PAKET CLIENT
        gen_frame = ctk.CTkFrame(self, fg_color=C_CARD, corner_radius=10)
        gen_frame.pack(fill="x", padx=14, pady=4)
        gen_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(gen_frame, text="📦 GENERATE PAKET CLIENT", font=FONT_SUB,
                     text_color=C_YELLOW).grid(row=0, column=0, columnspan=3,
                                               sticky="w", padx=10, pady=(8, 2))
        ctk.CTkLabel(gen_frame, text="Pilih PC", font=FONT_LABEL,
                     text_color=C_YELLOW).grid(row=1, column=0, padx=10, pady=4, sticky="w")
        self.cb_gen = ctk.CTkComboBox(gen_frame, values=[""], width=320,
                                      state="readonly")
        self.cb_gen.grid(row=1, column=1, padx=6, pady=4, sticky="ew")
        ctk.CTkButton(gen_frame, text="📦 Siapkan Paket", width=130, height=30,
                      fg_color=C_ACCENT2, hover_color="#5A0FCC",
                      command=self._siapkan_paket).grid(row=1, column=2, padx=6, pady=4)
        ctk.CTkButton(gen_frame, text="📁 Buka Folder Deploy", width=160, height=30,
                      fg_color=C_BTN, hover_color="#3A3A2A", border_width=1,
                      border_color=C_YELLOW, text_color=C_YELLOW,
                      command=self._buka_folder).grid(row=2, column=1, columnspan=2,
                                                      sticky="e", padx=6, pady=(2, 8))

        # Footer
        foot = ctk.CTkFrame(self, fg_color="transparent")
        foot.pack(fill="x", padx=14, pady=(4, 10))
        self.lbl_status = ctk.CTkLabel(foot, text="", font=FONT_SMALL, text_color=C_YELLOW)
        self.lbl_status.pack(side="left")
        ctk.CTkButton(foot, text="💾 Simpan", width=110, height=32,
                      fg_color=C_GREEN, hover_color="#2E7D32",
                      command=self._simpan_config).pack(side="right")
        ctk.CTkButton(foot, text="Tutup", width=90, height=32,
                      fg_color=C_BTN, command=self.destroy).pack(side="right", padx=6)

    # ── Combo refresh ───────────────────────────────────────────────────
    def _client_labels(self):
        return [f"{c.get('client_id','?')}  ({c.get('location','-')}) — {len(c.get('pcs',[]))} PC"
                for c in self._clients]

    def _on_pick_client(self, _=None):
        self._refresh_pc_combo()

    def _refresh_client_combo(self):
        self.cb_client.configure(values=self._client_labels() or [""])
        if self._clients:
            self.cb_client.set(self._client_labels()[0])
        else:
            self.cb_client.set("")

    def _selected_client(self):
        labels = self._client_labels()
        cur = self.cb_client.get()
        if cur in labels:
            return self._clients[labels.index(cur)]
        return None

    def _refresh_pc_combo(self):
        c = self._selected_client()
        vals = [f"{p.get('pc_id','?')} · {p.get('ip','-')} · {p.get('name','-')}"
                for p in (c.get("pcs", []) if c else [])]
        self.cb_pc.configure(values=vals or [""])
        self.cb_pc.set(vals[0] if vals else "")

    def _selected_pc(self):
        c = self._selected_client()
        if not c:
            return None, None
        vals = [f"{p.get('pc_id','?')} · {p.get('ip','-')} · {p.get('name','-')}"
                for p in c.get("pcs", [])]
        cur = self.cb_pc.get()
        if cur in vals:
            return c, c["pcs"][vals.index(cur)]
        return c, None

    def _refresh_gen_combo(self):
        self._gen_pcs = []
        labels = []
        for c in self._clients:
            for p in c.get("pcs", []):
                self._gen_pcs.append({"client_id": c.get("client_id", ""),
                                      "pc_id": p.get("pc_id", ""),
                                      "ip": p.get("ip", ""),
                                      "name": p.get("name", "")})
                labels.append(f"{c.get('client_id','?')} / {p.get('pc_id','?')} "
                              f"({p.get('name','?')})")
        self.cb_gen.configure(values=labels or [""])
        self.cb_gen.set(labels[0] if labels else "")

    # ── Input helper ────────────────────────────────────────────────────
    def _ask_fields(self, title, fields):
        dlg = ctk.CTkToplevel(self)
        dlg.title(title)
        dlg.geometry(f"420x{130 + len(fields) * 46}")
        dlg.configure(fg_color=C_BG)
        dlg.transient(self)
        dlg.grab_set()
        result = {}
        entries = []
        for i, (key, label, default) in enumerate(fields):
            ctk.CTkLabel(dlg, text=label, font=FONT_LABEL,
                         text_color=C_YELLOW).grid(row=i, column=0, padx=10, pady=6, sticky="w")
            e = ctk.CTkEntry(dlg, width=230, show="*" if key == "password" else "")
            e.insert(0, default or "")
            e.grid(row=i, column=1, padx=10, pady=6)
            entries.append((key, e))
        state = {"ok": False}

        def _ok():
            for k, e in entries:
                result[k] = e.get().strip()
            state["ok"] = True
            dlg.destroy()

        ctk.CTkButton(dlg, text="OK", width=90, height=30, fg_color=C_ACCENT2,
                      command=_ok).grid(row=len(fields), column=0, padx=10, pady=10)
        ctk.CTkButton(dlg, text="Batal", width=90, height=30, fg_color=C_BTN,
                      command=dlg.destroy).grid(row=len(fields), column=1, padx=10, pady=10)
        dlg.bind("<Return>", lambda _e: _ok())
        dlg.wait_window()
        return result if state["ok"] else None

    # ── CRUD client / PC ────────────────────────────────────────────────
    def _next_id(self, prefix, used):
        i = 1
        while f"{prefix}{i}" in used:
            i += 1
        return f"{prefix}{i}"

    def _tambah_client(self):
        used = {c.get("client_id", "") for c in self._clients}
        default_id = self._next_id("WARNET_", used)
        vals = self._ask_fields("➕ Tambah Client", [
            ("client_id", "Client ID", default_id),
            ("password", "Password", "admin123"),
            ("location", "Lokasi", "Warnet Lokasi 1"),
        ])
        if not vals or not vals.get("client_id") or not vals.get("password"):
            return
        if vals["client_id"] in used:
            messagebox.showwarning("⚠ Sudah Ada", f"Client ID '{vals['client_id']}' sudah dipakai.",
                                   parent=self)
            return
        self._clients.append({
            "client_id": vals["client_id"],
            "password_hash": hash_password(vals["password"]),
            "password_enc": vals["password"],
            "location": vals.get("location", ""),
            "pcs": [],
            "allowed_actions": [],
            "created_at": datetime.now().isoformat(timespec="seconds"),
        })
        self._refresh_client_combo()
        self.cb_client.set(self._client_labels()[-1])
        self._refresh_pc_combo()
        self._refresh_gen_combo()
        self._simpan_config()
        self.lbl_status.configure(text=f"✅ Client '{vals['client_id']}' ditambahkan.",
                                  text_color=C_YELLOW)

    def _ganti_password(self):
        c = self._selected_client()
        if not c:
            messagebox.showinfo("Pilih Client", "Pilih client dulu dari daftar.", parent=self)
            return
        vals = self._ask_fields("🔑 Ganti Password", [("password", "Password Baru", "")])
        if not vals or not vals.get("password"):
            return
        c["password_hash"] = hash_password(vals["password"])
        c["password_enc"] = vals["password"]
        self._simpan_config()
        self.lbl_status.configure(text=f"✅ Password '{c['client_id']}' diganti.",
                                  text_color=C_YELLOW)

    def _hapus_client(self):
        c = self._selected_client()
        if not c:
            messagebox.showinfo("Pilih Client", "Pilih client dulu dari daftar.", parent=self)
            return
        if not messagebox.askyesno("🗑 Hapus Client",
                                   f"Hapus client '{c.get('client_id')}' beserta semua PC-nya?",
                                   parent=self):
            return
        self._clients.remove(c)
        self._refresh_client_combo()
        self._refresh_pc_combo()
        self._refresh_gen_combo()
        self._simpan_config()
        self.lbl_status.configure(text=f"🗑 Client '{c.get('client_id')}' dihapus.",
                                  text_color=C_YELLOW)

    def _tambah_pc(self):
        c = self._selected_client()
        if not c:
            messagebox.showinfo("Pilih Client", "Pilih client dulu dari daftar.", parent=self)
            return
        used = {p.get("pc_id", "") for p in c.get("pcs", [])}
        default_pc = self._next_id("PC_", used)
        vals = self._ask_fields("➕ Tambah PC", [
            ("pc_id", "PC ID", default_pc),
            ("ip", "IP PC Warnet", ""),
            ("name", "Nama Kursi", ""),
        ])
        if not vals or not vals.get("pc_id"):
            return
        if vals["pc_id"] in used:
            messagebox.showwarning("⚠ Sudah Ada", f"PC ID '{vals['pc_id']}' sudah dipakai.",
                                   parent=self)
            return
        c["pcs"].append({"pc_id": vals["pc_id"], "ip": vals.get("ip", ""),
                         "name": vals.get("name", "")})
        self._refresh_pc_combo()
        self._refresh_gen_combo()
        self._simpan_config()
        self.lbl_status.configure(
            text=f"✅ PC '{vals['pc_id']}' ditambahkan ke '{c.get('client_id')}'.",
            text_color=C_YELLOW)

    def _hapus_pc(self):
        c, p = self._selected_pc()
        if not c or not p:
            messagebox.showinfo("Pilih PC", "Pilih PC dulu dari daftar.", parent=self)
            return
        if not messagebox.askyesno("🗑 Hapus PC",
                                   f"Hapus PC '{p.get('pc_id')}' dari '{c.get('client_id')}'?",
                                   parent=self):
            return
        c["pcs"].remove(p)
        self._refresh_pc_combo()
        self._refresh_gen_combo()
        self._simpan_config()
        self.lbl_status.configure(text=f"🗑 PC '{p.get('pc_id')}' dihapus.", text_color=C_YELLOW)

    # ── Generate paket client ───────────────────────────────────────────
    def _siapkan_paket(self):
        idx = self.cb_gen.current()
        if idx < 0 or idx >= len(self._gen_pcs):
            messagebox.showinfo("Pilih PC", "Pilih PC dulu di bagian Generate Paket Client.",
                                parent=self)
            return
        info = self._gen_pcs[idx]
        c = self._find_client(info["client_id"])
        if not c:
            return
        plain = c.get("password_enc", "")
        if not plain:
            messagebox.showerror("Password Tidak Ada",
                                 f"Client '{info['client_id']}' belum punya password "
                                 "(set via Ganti Password).", parent=self)
            return
        server_ip = self.entry_server_ip.get().strip() or _detect_local_ip()
        out_dir = os.path.join(self.DEPLOY_OUT_DIR, f"{info['client_id']}_{info['pc_id']}")
        try:
            os.makedirs(out_dir, exist_ok=True)
            cfg = {
                "server_host": server_ip,
                "server_port": self.SERVER_PORT,
                "client_id": info["client_id"],
                "password": plain,
                "pc_id": info["pc_id"],
            }
            with open(os.path.join(out_dir, "rr_billing_config.json"), "w",
                      encoding="utf-8") as f:
                json.dump(cfg, f, indent=2, ensure_ascii=False)

            copied = []
            if os.path.isdir(self.DEPLOY_EXE_DIR):
                for exe in ("BillingClientService.exe", "BillingLockScreenUI.exe",
                            "BillingClientApp.exe"):
                    src = os.path.join(self.DEPLOY_EXE_DIR, exe)
                    if os.path.exists(src):
                        shutil.copy2(src, os.path.join(out_dir, exe))
                        copied.append(exe)

            self._tulis_install_bat(out_dir)
            self._tulis_tray_bat(out_dir)
            self._tulis_baca_dulu(out_dir, server_ip)

            msg = f"✅ Paket siap di: {out_dir}"
            if copied:
                msg += f"  ({len(copied)} exe)"
            self.lbl_status.configure(text=msg, text_color=C_YELLOW)
            messagebox.showinfo("📦 Paket Siap", msg + "\n\nCopy folder ini ke PC warnet "
                                "lalu jalankan install_service.bat (Run as administrator).",
                                parent=self)
        except Exception as e:
            self.lbl_status.configure(text=f"❌ Gagal: {e}", text_color=C_RED)
            messagebox.showerror("Gagal", str(e), parent=self)

    def _tulis_install_bat(self, out_dir):
        content = (
            "@echo off\r\n"
            "net session >nul 2>&1\r\n"
            "if %errorlevel% neq 0 (\r\n"
            "  echo Jalankan sebagai Administrator! Klik kanan file ini lalu pilih \"Run as administrator\".\r\n"
            "  pause\r\n"
            "  exit /b 1\r\n"
            ")\r\n"
            "echo Menginstall RR Billing Client Service...\r\n"
            "BillingClientService.exe -i\r\n"
            "if %errorlevel% neq 0 (\r\n"
            "  echo Gagal menginstall service.\r\n"
            "  pause\r\n"
            "  exit /b 1\r\n"
            ")\r\n"
            "echo Memulai service...\r\n"
            "net start RRBillingClientService\r\n"
            "echo.\r\n"
            "echo Selesai! Client akan tersambung ke server billing.\r\n"
            "echo Jalankan install_tray.bat untuk tray info billing.\r\n"
            "pause\r\n"
        )
        with open(os.path.join(out_dir, "install_service.bat"), "w",
                  encoding="ascii") as f:
            f.write(content)

    def _tulis_tray_bat(self, out_dir):
        content = (
            "@echo off\r\n"
            "set STARTUP=%APPDATA%\\Microsoft\\Windows\\Start Menu\\Programs\\Startup\r\n"
            "set SCRIPT_TMP=%TEMP%\\rr_mk_tray_link.vbs\r\n"
            "echo Set WshShell = WScript.CreateObject(\"WScript.Shell\") > \"%SCRIPT_TMP%\"\r\n"
            "echo Set sc = WshShell.CreateShortcut(\"%STARTUP%\\RR Billing Client.lnk\") >> \"%SCRIPT_TMP%\"\r\n"
            "echo sc.TargetPath = \"%~dp0BillingClientApp.exe\" >> \"%SCRIPT_TMP%\"\r\n"
            "echo sc.WorkingDirectory = \"%~dp0\" >> \"%SCRIPT_TMP%\"\r\n"
            "echo sc.Save >> \"%SCRIPT_TMP%\"\r\n"
            "cscript //nologo \"%SCRIPT_TMP%\"\r\n"
            "del \"%SCRIPT_TMP%\"\r\n"
            "echo Shortcut tray dibuat. Tray akan otomatis jalan saat login Windows.\r\n"
            "pause\r\n"
        )
        with open(os.path.join(out_dir, "install_tray.bat"), "w", encoding="ascii") as f:
            f.write(content)

    def _tulis_baca_dulu(self, out_dir, server_ip):
        lines = [
            "=== RR BILLING PRO — PAKET CLIENT WARNET ===",
            "",
            f"Server billing: {server_ip}:5000",
            "",
            "LANGKAH DI PC WARNET (Windows 10):",
            "1. Copy SELURUH isi folder ini ke satu folder di PC warnet,",
            "   misalnya C:\\RRBillingClient\\",
            "2. Jalankan install_service.bat — KLIK KANAN lalu 'Run as administrator'.",
            "   (menginstall & memulai BillingClientService, konek ke server)",
            "3. Jalankan install_tray.bat sekali (shortcut tray ke Startup).",
            "",
            "Cek hasil:",
            "- Di PC admin: Dashboard Warnet -> 'Client: Tersambung'",
            "- File rr_billing_config.json berisi IP server; ubah bila IP server berubah.",
            "",
            "CATATAN:",
            "- IP server HARUS statis di jaringan warnet.",
            "- Jika PC di-lock oleh server saat waktu habis, lock screen akan muncul.",
        ]
        with open(os.path.join(out_dir, "BACA_DULU.txt"), "w",
                  encoding="utf-8") as f:
            f.write("\n".join(lines))

    def _buka_folder(self):
        try:
            os.makedirs(self.DEPLOY_OUT_DIR, exist_ok=True)
            os.startfile(self.DEPLOY_OUT_DIR)
        except Exception as e:
            self.lbl_status.configure(text=f"❌ Gagal buka folder: {e}", text_color=C_RED)


def _detect_local_ip() -> str:
    """Deteksi IPv4 lokal PC admin (bukan 127.0.0.1)."""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        try:
            s.connect(("8.8.8.8", 80))
            return s.getsockname()[0]
        finally:
            s.close()
    except Exception:
        pass
    try:
        for info in socket.getaddrinfo(socket.gethostname(), None):
            ip = info[4][0]
            if ip and not ip.startswith("127.") and "." in ip:
                return ip
    except Exception:
        pass
    return "192.168.1.100"


# ═══════════════════════════════════════════════════════════════════════════════
#  LOGIN PAGE
# ═══════════════════════════════════════════════════════════════════════════════
class LoginPage(ctk.CTkFrame):
    # Default passwords (hashed dengan bcrypt untuk keamanan)
    # Jika tidak ada users di config.json, gunakan ini sebagai fallback
    DEFAULT_USERS = {
        # Tidak ada default user - user harus register atau admin setup
    }

    LOGIN_MAX_ATTEMPTS = 5
    LOGIN_LOCK_DURATION = timedelta(minutes=1)

    def __init__(self, master, on_login_success):
        super().__init__(master, fg_color=C_BG, corner_radius=0)
        self.on_login_success = on_login_success
        self._attempt         = 0
        self._locked_until    = None
        self._lp_username_verified = None  # username yg terverifikasi di lupa password

        # Satu container — isinya diganti saat pindah view
        # Dengan width minimum agar tidak shrink terlalu kecil
        self._view_container = ctk.CTkFrame(self, fg_color="transparent")
        self._view_container.pack(expand=True, fill="both", padx=20, pady=(14, 10), anchor="n")

        self._show_login_view()

    # ══════════════════════════════════════════════════════════════════════════
    #  VIEW 1 — LOGIN
    # ══════════════════════════════════════════════════════════════════════════
    def _show_login_view(self):
        for w in self._view_container.winfo_children():
            w.destroy()

        scroll = ctk.CTkScrollableFrame(self._view_container, fg_color="transparent")
        scroll.pack(fill="both", expand=True)

        outer = ctk.CTkFrame(scroll, fg_color=C_PANEL,
                              corner_radius=20, border_width=2,
                              border_color=C_ACCENT2)
        outer.pack(anchor="n", pady=(10, 10), padx=4, fill="x")

        # Logo
        logo_ico = ctk.CTkFrame(outer, fg_color=C_CARD, corner_radius=12,
                                  width=152, height=62)
        logo_ico.pack(pady=(28, 8))
        logo_ico.pack_propagate(False)
        ctk_img = load_ctk_image(size=(143, 55))
        if ctk_img:
            ctk.CTkLabel(logo_ico, text="", image=ctk_img).place(
                relx=0.5, rely=0.5, anchor="center")
        else:
            ctk.CTkLabel(logo_ico, text="🎮", font=("Arial", 32)).place(
                relx=0.5, rely=0.5, anchor="center")

        ctk.CTkLabel(outer, text="RR BILLING PRO",
                     font=("Russo One", 22, "bold"),
                     text_color=C_ACCENT).pack(pady=(4, 0))
        ctk.CTkLabel(outer, text="Sistem Billing Rental TV & PS",
                     font=("Courier New", 11), text_color=C_MUTED).pack(pady=(0, 16))

        self.lbl_lic = ctk.CTkLabel(outer, text="Silakan login untuk melanjutkan.",
                                     font=("Courier New", 10), text_color=C_MUTED)
        self.lbl_lic.pack(pady=(0, 10))

        ctk.CTkFrame(outer, height=1, fg_color=C_BORDER).pack(
            fill="x", padx=30, pady=(0, 18))

        # Input username
        ctk.CTkLabel(outer, text="Username", font=("Consolas", 11),
                     text_color=C_MUTED).pack(padx=40)
        self.entry_user = ctk.CTkEntry(
            outer, placeholder_text="Masukkan username",
            fg_color=C_BTN, text_color=C_ACCENT,
            border_color=C_BORDER, font=("Consolas", 14),
            height=40, width=320, justify="center")
        self.entry_user.pack(pady=(2, 10), padx=40)

        # Input password
        ctk.CTkLabel(outer, text="Password", font=("Consolas", 11),
                     text_color=C_MUTED).pack(padx=40)
        self.entry_pass = ctk.CTkEntry(
            outer, placeholder_text="Masukkan password",
            fg_color=C_BTN, text_color=C_ACCENT,
            border_color=C_BORDER, font=("Consolas", 14),
            height=40, width=320, show="●", justify="center")
        self.entry_pass.pack(pady=(2, 6), padx=40)
        self.entry_pass.bind("<Return>", lambda e: self._login())

        # Status error
        self.lbl_status = ctk.CTkLabel(outer, text="",
                                        font=("Consolas", 11), text_color=C_RED)
        self.lbl_status.pack(pady=(0, 8))

        # Tombol Masuk
        self.btn_login = ctk.CTkButton(
            outer, text="🔓  MASUK", width=280, height=44,
            fg_color=C_ACCENT2, hover_color=C_ACCENT,
            font=("Russo One", 13, "bold"), text_color="white",
            command=self._login)
        self.btn_login.pack(pady=(0, 4), padx=30)

        # Google Login
        self.btn_google = ctk.CTkButton(
            outer, text="🌐  MASUK DENGAN GOOGLE", width=280, height=38,
            fg_color="#DB4437", hover_color="#C53829",
            font=("Russo One", 11, "bold"), text_color="white",
            command=self._login_google)
        self.btn_google.pack(pady=(4, 2), padx=30)

        # Login Kasir (sub-akun admin) — APTV2 style
        self.btn_kasir_login = ctk.CTkButton(
            outer, text="👤  LOGIN KASIR", width=280, height=38,
            fg_color="#0A2A2A", hover_color="#0A3A3A",
            border_width=1, border_color=C_GREEN,
            font=("Russo One", 11, "bold"), text_color=C_GREEN,
            command=self._show_kasir_login_view)
        self.btn_kasir_login.pack(pady=(6, 2), padx=30)

        ctk.CTkLabel(outer,
                     text="Hubungi administrator untuk akses.",
                     font=("Courier New", 10), text_color=C_MUTED).pack(pady=(0, 4))

        # Tombol Daftar
        ctk.CTkLabel(outer, text="Rental baru? Belum punya akun?",
                     font=("Courier New", 10), text_color=C_MUTED).pack(pady=(8, 4))
        ctk.CTkButton(
            outer, text="📝  DAFTAR RENTAL BARU", width=280, height=36,
            fg_color="transparent", hover_color=C_BTN,
            border_width=1, border_color=C_ACCENT,
            font=("Russo One", 11, "bold"), text_color=C_ACCENT,
            command=self._show_daftar_view).pack(pady=(0, 6), padx=30)

        # Tombol Lupa Password
        ctk.CTkButton(
            outer, text="🔓  Lupa Password?", width=280, height=34,
            fg_color="transparent", hover_color=C_BTN,
            border_width=1, border_color=C_RED,
            font=("Russo One", 10, "bold"), text_color=C_RED,
            command=self._show_lupa_password_view).pack(pady=(0, 10), padx=30)

        # Version di login page
        ctk.CTkLabel(outer, text=f"v{APP_VERSION}",
                     font=("Courier New", 10), text_color=C_MUTED).pack(pady=(0, 6))

    # ══════════════════════════════════════════════════════════════════════════
    #  VIEW 2 — DAFTAR RENTAL BARU
    # ══════════════════════════════════════════════════════════════════════════
    def _show_daftar_view(self):
        for w in self._view_container.winfo_children():
            w.destroy()

        outer = ctk.CTkFrame(self._view_container, fg_color=C_PANEL,
                              corner_radius=20, border_width=2,
                              border_color=C_ACCENT2)
        outer.pack()

        # Logo section (consistent with login view)
        logo_ico = ctk.CTkFrame(outer, fg_color=C_CARD, corner_radius=12,
                                 width=152, height=62)
        logo_ico.pack(pady=(28, 8))
        logo_ico.pack_propagate(False)
        ctk_img = load_ctk_image(size=(143, 55))
        if ctk_img:
            ctk.CTkLabel(logo_ico, text="", image=ctk_img).place(
                relx=0.5, rely=0.5, anchor="center")
        else:
            ctk.CTkLabel(logo_ico, text="🎮", font=("Arial", 32)).place(
                relx=0.5, rely=0.5, anchor="center")

        # Title section
        ctk.CTkLabel(outer, text="DAFTAR AKUN BARU",
                     font=("Russo One", 22, "bold"),
                     text_color=C_ACCENT).pack(pady=(4, 0))
        ctk.CTkLabel(outer, text="Sistem Billing Rental TV & PS",
                     font=("Courier New", 11), text_color=C_MUTED).pack(pady=(0, 16))

        ctk.CTkLabel(outer,
                     text="Buat akun admin untuk rental Anda — trial otomatis aktif",
                     font=FONT_SMALL, text_color=C_MUTED).pack(pady=(10, 4))
        ctk.CTkLabel(outer,
                     text="Email wajib valid karena kami akan mengirim kode verifikasi.",
                     font=FONT_SMALL, text_color=C_MUTED).pack(pady=(0, 8))

        scroll = ctk.CTkScrollableFrame(outer, fg_color="transparent",
                                         width=500, height=550)
        scroll.pack(padx=20, pady=4)

        # Akun login
        akun_f = ctk.CTkFrame(scroll, fg_color=C_CARD, corner_radius=10)
        akun_f.pack(fill="x", pady=6)
        ctk.CTkLabel(akun_f, text="🔐  AKUN LOGIN", font=FONT_SUB,
                     text_color=C_ACCENT2).pack(anchor="w", padx=14, pady=(10, 6))
        self.d_username  = self._input_field(akun_f, "Username", "mis. rentalku01")
        self.d_password  = self._input_field(akun_f, "Password", "minimal 6 karakter", show="●")
        self.d_password2 = self._input_field(akun_f, "Konfirmasi Password", "ulangi password", show="●")

        # Profil rental
        profil_f = ctk.CTkFrame(scroll, fg_color=C_CARD, corner_radius=10)
        profil_f.pack(fill="x", pady=6)
        ctk.CTkLabel(profil_f, text="🏢  PROFIL RENTAL", font=FONT_SUB,
                     text_color=C_ACCENT2).pack(anchor="w", padx=14, pady=(10, 6))
        self.d_nama_pemilik = self._input_field(profil_f, "Nama Pemilik",  "mis. Budi Santoso")
        self.d_nama_rental  = self._input_field(profil_f, "Nama Rental PS","mis. RR Game Center")
        self.d_email        = self._input_field(profil_f, "Email / Gmail", "mis. nama@gmail.com")
        self.d_hp           = self._input_field(profil_f, "No HP / WhatsApp","mis. 0812xxxxxxx")

        ctk.CTkLabel(profil_f, text="Alamat", font=FONT_LABEL,
                     text_color=C_MUTED, anchor="w").pack(anchor="w", padx=14, pady=(4, 2))
        self.d_alamat = ctk.CTkTextbox(profil_f, height=56,
                                        fg_color=C_BTN, text_color=C_TEXT,
                                        border_color=C_BORDER, border_width=1,
                                        font=FONT_BODY)
        self.d_alamat.pack(fill="x", padx=14, pady=(0, 10))

        self.d_logo_b64 = ""
        logo_row = ctk.CTkFrame(profil_f, fg_color="transparent")
        logo_row.pack(fill="x", padx=14, pady=(0, 10))
        ctk.CTkLabel(logo_row, text="🖼 Logo Rental (opsional, tampil di halaman booking)",
                     font=FONT_LABEL, text_color=C_MUTED, anchor="w").pack(anchor="w", pady=(0, 4))
        self.lbl_logo_daftar = ctk.CTkLabel(logo_row, text="Belum ada logo",
                                            font=FONT_SMALL, text_color=C_MUTED, anchor="w")
        self.lbl_logo_daftar.pack(side="left", padx=(0, 10))
        ctk.CTkButton(logo_row, text="📁 Pilih Gambar", width=120, height=30,
                      fg_color=C_ACCENT2, font=("Russo One", 9, "bold"),
                      command=lambda: self._daftar_pilih_logo()).pack(side="left")

        # Tombol dalam scroll
        btn_frame = ctk.CTkFrame(scroll, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(12, 0))
        
        ctk.CTkButton(btn_frame, text="✅  DAFTAR SEKARANG", width=180, height=40,
                      fg_color=C_ACCENT2, hover_color="#5A0FCC",
                      font=("Russo One", 11, "bold"), text_color="white",
                      command=self._submit_daftar).pack(side="left", padx=6)
        ctk.CTkButton(btn_frame, text="❌  BATAL", width=140, height=40,
                      fg_color=C_RED, hover_color="#8B0000",
                      font=("Russo One", 11, "bold"), text_color="white",
                      command=self._show_login_view).pack(side="left", padx=6)

        # Status message
        self.lbl_daftar_status = ctk.CTkLabel(outer, text="",
                                               font=FONT_LABEL, text_color=C_RED,
                                               wraplength=420, justify="center")
        self.lbl_daftar_status.pack(pady=(6, 2))

    def _input_field(self, parent, label, placeholder, show=None):
        ctk.CTkLabel(parent, text=label, font=FONT_LABEL,
                     text_color=C_MUTED, anchor="w").pack(anchor="w", padx=14, pady=(4, 2))
        kw = {"show": show} if show else {}
        e = ctk.CTkEntry(parent, placeholder_text=placeholder,
                          fg_color=C_BTN, text_color=C_TEXT,
                          border_color=C_BORDER, font=FONT_BODY, height=34, **kw)
        e.pack(fill="x", padx=14, pady=(0, 6))
        return e

    def _ambil_logo_b64(self, label_widget=None, tampil_error: bool = False) -> str:
        """Pilih file gambar logo rental -> resize <=768px -> dataURL.
        Return "" bila dibatalkan/gagal."""
        try:
            self.lift()
            self.focus_force()
        except Exception:
            pass
        try:
            from tkinter import filedialog
            path = filedialog.askopenfilename(
                title="Pilih Logo Rental",
                filetypes=[("Gambar", "*.png *.jpg *.jpeg *.webp *.bmp *.gif"),
                           ("Semua file", "*.*")],
                parent=self)
        except Exception as e:
            if tampil_error:
                messagebox.showerror("Gagal Buka Dialog", str(e), parent=self)
            return ""
        if not path:
            return ""
        return logo_gambar_b64(path, label_widget=label_widget, tampil_error=tampil_error)

    def _daftar_pilih_logo(self):
        self.d_logo_b64 = self._ambil_logo_b64(self.lbl_logo_daftar)

    def _submit_daftar(self):
        username     = sanitize_text(self.d_username.get()).lower()
        password     = self.d_password.get().strip()
        password2    = self.d_password2.get().strip()
        nama_pemilik = sanitize_text(self.d_nama_pemilik.get())
        nama_rental  = sanitize_text(self.d_nama_rental.get())
        email        = sanitize_text(self.d_email.get()).lower()
        no_hp        = sanitize_text(self.d_hp.get())
        alamat       = sanitize_text(self.d_alamat.get("1.0", "end"))

        if not all([username, password, password2, nama_pemilik,
                    nama_rental, email, no_hp, alamat]):
            self.lbl_daftar_status.configure(
                text="⚠  Semua field wajib diisi.", text_color=C_YELLOW)
            return
        if not is_valid_username(username):
            self.lbl_daftar_status.configure(
                text="⚠  Username harus 4-20 karakter, huruf, angka, titik, atau garis bawah.",
                text_color=C_YELLOW)
            return
        if not is_valid_password(password):
            self.lbl_daftar_status.configure(
                text="⚠  Password harus minimal 8 karakter, dan berisi huruf serta angka.",
                text_color=C_YELLOW)
            return
        if password != password2:
            self.lbl_daftar_status.configure(
                text="⚠  Konfirmasi password tidak cocok.", text_color=C_YELLOW)
            return
        if not is_valid_email(email):
            self.lbl_daftar_status.configure(
                text="⚠  Format email tidak valid.", text_color=C_YELLOW)
            return
        if not is_valid_phone(no_hp):
            self.lbl_daftar_status.configure(
                text="⚠  Nomor HP tidak valid. Gunakan 7-15 digit, boleh diawali +.",
                text_color=C_YELLOW)
            return

        cfg   = ConfigManager.load()
        users = cfg.get("users", dict(LoginPage.DEFAULT_USERS))
        pending = cfg.get("pending_email_verifications", {})

        if username in users:
            self.lbl_daftar_status.configure(
                text=f"✖  Username '{username}' sudah dipakai.", text_color=C_RED)
            return
        if username in pending:
            self.lbl_daftar_status.configure(
                text=f"⚠  Username '{username}' sudah memiliki verifikasi tertunda.",
                text_color=C_YELLOW)
            return

        if not _email_configured():
            self.lbl_daftar_status.configure(
                text="⚠  Verifikasi email belum tersedia karena SMTP belum dikonfigurasi.",
                text_color=C_YELLOW)
            return

        verification_code = generate_verification_code()
        verification_expires = datetime.now() + timedelta(minutes=EMAIL_VERIFICATION_EXPIRY_MINUTES)

        pending[username] = {
            "code": verification_code,
            "expires_at": verification_expires.isoformat(),
            "email": email,
            "data": {
                "password_enc": hash_password(password),
                "role": "admin",
                "profil": {
                    "nama_pemilik":   nama_pemilik,
                    "nama_rental":    nama_rental,
                    "email":          email,
                    "no_hp":          no_hp,
                    "alamat":         alamat,
                    "logo":           getattr(self, "d_logo_b64", "") or "",
                    "tanggal_daftar": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                },
            },
        }
        cfg["pending_email_verifications"] = pending
        ConfigManager.save(cfg)

        sukses, pesan = _send_verification_email(email, username, verification_code)
        if not sukses:
            self.lbl_daftar_status.configure(
                text=f"✖  Gagal kirim kode verifikasi: {pesan}",
                text_color=C_RED)
            return

        AuditLogger.log(
            action="registration_pending",
            username=username,
            status="pending",
            details={"email": email}
        )

        self._show_verification_dialog(username)

    def _save_pending_verification(self, username: str, payload: dict):
        cfg = ConfigManager.load()
        pending = cfg.get("pending_email_verifications", {})
        pending[username] = payload
        cfg["pending_email_verifications"] = pending
        ConfigManager.save(cfg)

    def _load_pending_verification(self, username: str) -> dict:
        cfg = ConfigManager.load()
        return cfg.get("pending_email_verifications", {}).get(username, {})

    def _clear_pending_verification(self, username: str):
        cfg = ConfigManager.load()
        pending = cfg.get("pending_email_verifications", {})
        if username in pending:
            pending.pop(username)
            cfg["pending_email_verifications"] = pending
            ConfigManager.save(cfg)

    def _show_verification_dialog(self, username: str):
        pending = self._load_pending_verification(username)
        if not pending:
            messagebox.showerror("✖ Error", "Data verifikasi tidak ditemukan. Coba daftar ulang.")
            return

        dialog = ctk.CTkToplevel(self)
        dialog.title("📧 Verifikasi Email")
        dialog.geometry("420x260")
        dialog.grab_set()

        ctk.CTkLabel(dialog, text="Verifikasi Email", font=("Russo One", 16, "bold"),
                     text_color=C_ACCENT).pack(pady=(20, 10))
        ctk.CTkLabel(dialog,
                     text=f"Kode dikirim ke {pending.get('email')}. Masukkan kode di bawah.",
                     font=FONT_BODY, text_color=C_TEXT, wraplength=380).pack(pady=(0, 10))

        entry_code = ctk.CTkEntry(dialog, placeholder_text="Kode verifikasi 6 digit",
                                  fg_color=C_BTN, text_color=C_ACCENT,
                                  border_color=C_BORDER, font=FONT_BODY, height=40)
        entry_code.pack(fill="x", padx=30, pady=(0, 10))
        try:
            entry_code.focus()
        except Exception:
            pass

        status_label = ctk.CTkLabel(dialog, text="", font=FONT_LABEL, text_color=C_RED,
                                    wraplength=380, justify="center")
        status_label.pack(pady=(0, 10))

        def submit_code():
            code = sanitize_text(entry_code.get())
            if not code:
                status_label.configure(text="⚠  Masukkan kode verifikasi.", text_color=C_YELLOW)
                return
            success, msg = self._submit_verification_code(username, code)
            if success:
                dialog.destroy()
            else:
                status_label.configure(text=msg, text_color=C_RED)

        def resend_code():
            code = generate_verification_code()
            pending_data = self._load_pending_verification(username)
            if not pending_data:
                status_label.configure(text="✖  Data verifikasi hilang.", text_color=C_RED)
                return
            pending_data["code"] = code
            pending_data["expires_at"] = (datetime.now() + timedelta(minutes=EMAIL_VERIFICATION_EXPIRY_MINUTES)).isoformat()
            self._save_pending_verification(username, pending_data)
            sukses, pesan = _send_verification_email(pending_data.get("email", ""), username, code)
            if sukses:
                status_label.configure(text="✅  Kode baru telah dikirim.", text_color=C_GREEN)
                AuditLogger.log(
                    action="email_verification_resent",
                    username=username,
                    status="pending",
                    details={"email": pending_data.get("email", "")}
                )
            else:
                status_label.configure(text=f"✖  Gagal kirim ulang: {pesan}", text_color=C_RED)

        ctk.CTkButton(dialog, text="✅  Verifikasi", width=160, height=38,
                      fg_color=C_ACCENT2, hover_color="#5A0FCC",
                      font=("Russo One", 10, "bold"), text_color="white",
                      command=submit_code).pack(pady=(0, 6))
        ctk.CTkButton(dialog, text="🔁  Kirim Ulang Kode", width=160, height=34,
                      fg_color="transparent", hover_color=C_BTN,
                      border_width=1, border_color=C_ACCENT,
                      font=("Russo One", 10, "bold"), text_color=C_ACCENT,
                      command=resend_code).pack(pady=(0, 8))

    def _submit_verification_code(self, username: str, code: str) -> tuple:
        pending = self._load_pending_verification(username)
        if not pending:
            return False, "Data verifikasi tidak ditemukan. Coba daftar ulang."

        expires_at = pending.get("expires_at")
        if not expires_at:
            return False, "Data kadaluarsa tidak valid."

        try:
            expires = datetime.fromisoformat(expires_at)
        except Exception:
            return False, "Data kadaluarsa tidak valid."

        if datetime.now() > expires:
            self._clear_pending_verification(username)
            return False, "Kode verifikasi sudah kadaluarsa. Daftar ulang untuk mengirim kode baru."

        if code != pending.get("code"):
            AuditLogger.log(
                action="email_verification_failed",
                username=username,
                status="invalid_code",
                details={"attempt_code": code}
            )
            return False, "Kode verifikasi tidak cocok."

        data = pending.get("data", {})
        cfg = ConfigManager.load()
        users = cfg.get("users", dict(LoginPage.DEFAULT_USERS))
        if username in users:
            self._clear_pending_verification(username)
            return False, "Username sudah terdaftar."

        users[username] = {
            "password_enc": data.get("password_enc") or data.get("password"),
            "role": data.get("role", "admin"),
        }
        cfg["users"] = users
        profil_semua = cfg.get("profil_rental", {})
        profil_semua[username] = data.get("profil", {})
        cfg["profil_rental"] = profil_semua
        self._clear_pending_verification(username)
        ConfigManager.save(cfg)

        AuditLogger.log(
            action="registration",
            username=username,
            status="verified",
            details={"email": pending.get("email")}
        )

        messagebox.showinfo("✅ Terverifikasi", "Email berhasil diverifikasi. Silakan login.")
        self._show_login_view()
        self.entry_user.delete(0, "end")
        self.entry_user.insert(0, username)
        try:
            self.entry_pass.focus()
        except Exception:
            pass
        return True, ""

    # ══════════════════════════════════════════════════════════════════════════
    #  VIEW 3 — LUPA PASSWORD
    # ══════════════════════════════════════════════════════════════════════════
    def _show_lupa_password_view(self):
        for w in self._view_container.winfo_children():
            w.destroy()
        self._lp_username_verified = None

        self.lp_outer = ctk.CTkFrame(self._view_container, fg_color=C_PANEL,
                                      corner_radius=20, border_width=2,
                                      border_color=C_RED)
        self.lp_outer.pack(fill="both", expand=True)

        # Header merah
        hdr = ctk.CTkFrame(self.lp_outer, fg_color=C_RED, corner_radius=0)
        hdr.pack(fill="x")
        ctk.CTkLabel(hdr, text="🔓  LUPA PASSWORD",
                     font=("Russo One", 14, "bold"),
                     text_color="white").pack(pady=14)

        ctk.CTkLabel(self.lp_outer,
                     text="Masukkan username dan email yang saat daftar.\n"
                          "Jika cocok, kamu bisa set password baru.",
                     font=FONT_SMALL, text_color=C_MUTED,
                     justify="center").pack(pady=(14, 4))

        # ── STEP 1: Verifikasi ────────────────────────────────────────────────
        step1 = ctk.CTkFrame(self.lp_outer, fg_color=C_CARD, corner_radius=12)
        step1.pack(fill="x", padx=28, pady=(6, 4))

        # Indikator step
        step_row = ctk.CTkFrame(step1, fg_color="transparent")
        step_row.pack(fill="x", padx=14, pady=(12, 8))
        for no, label, aktif in [("1", "Verifikasi Identitas", True),
                                   ("2", "Set Password Baru",   False)]:
            dot = ctk.CTkFrame(step_row, fg_color=C_ACCENT2 if aktif else C_BTN,
                                corner_radius=12, width=24, height=24)
            dot.pack(side="left", padx=(0, 4))
            dot.pack_propagate(False)
            ctk.CTkLabel(dot, text=no, font=("Russo One", 9, "bold"),
                          text_color="white").place(relx=0.5, rely=0.5, anchor="center")
            ctk.CTkLabel(step_row, text=label, font=FONT_LABEL,
                          text_color=C_ACCENT2 if aktif else C_MUTED).pack(side="left", padx=(0, 16))

        ctk.CTkFrame(step1, height=1, fg_color=C_BORDER).pack(fill="x", padx=14, pady=(0, 10))

        ctk.CTkLabel(step1, text="Username:", font=FONT_LABEL,
                     text_color=C_MUTED, anchor="w").pack(anchor="w", padx=14, pady=(0, 2))
        self.lp_entry_username = ctk.CTkEntry(
            step1, placeholder_text="Username akun Anda",
            fg_color=C_BTN, text_color=C_ACCENT,
            border_color=C_BORDER,
            font=("Consolas", 12, "bold"), height=36)
        self.lp_entry_username.pack(fill="x", padx=14, pady=(0, 10))

        ctk.CTkLabel(step1, text="Email yang didaftarkan:", font=FONT_LABEL,
                     text_color=C_MUTED, anchor="w").pack(anchor="w", padx=14, pady=(0, 2))
        self.lp_entry_email = ctk.CTkEntry(
            step1, placeholder_text="mis. nama@gmail.com",
            fg_color=C_BTN, text_color=C_ACCENT,
            border_color=C_BORDER,
            font=("Consolas", 12, "bold"), height=36)
        self.lp_entry_email.pack(fill="x", padx=14, pady=(0, 10))
        self.lp_entry_email.bind("<Return>", lambda e: self._verifikasi_identitas())

        self.btn_verif = ctk.CTkButton(
            step1, text="✅  Verifikasi Identitas", height=38,
            fg_color=C_ACCENT2, hover_color="#4A20C8",
            font=("Russo One", 10, "bold"), text_color="white",
            command=self._verifikasi_identitas)
        self.btn_verif.pack(fill="x", padx=14, pady=(0, 14))

        # ── STEP 2: Set password baru (hidden, muncul setelah verifikasi) ─────
        self.lp_step2 = ctk.CTkFrame(self.lp_outer, fg_color=C_CARD, corner_radius=12)
        # Belum di-pack — ditampilkan setelah verifikasi sukses

        step2_hdr = ctk.CTkFrame(self.lp_step2, fg_color=C_GREEN, corner_radius=8)
        step2_hdr.pack(fill="x", padx=14, pady=(12, 8))
        ctk.CTkLabel(step2_hdr,
                     text="✅  Identitas Terverifikasi — Set Password Baru",
                     font=FONT_SUB, text_color=C_BG).pack(pady=8)

        ctk.CTkLabel(self.lp_step2, text="Password Baru:", font=FONT_LABEL,
                     text_color=C_MUTED, anchor="w").pack(anchor="w", padx=14, pady=(4, 2))
        self.lp_entry_pass1 = ctk.CTkEntry(
            self.lp_step2, placeholder_text="Minimal 6 karakter",
            fg_color=C_BTN, text_color=C_ACCENT,
            border_color=C_BORDER,
            font=("Consolas", 12, "bold"), height=36, show="●")
        self.lp_entry_pass1.pack(fill="x", padx=14, pady=(0, 10))

        ctk.CTkLabel(self.lp_step2, text="Konfirmasi Password:", font=FONT_LABEL,
                     text_color=C_MUTED, anchor="w").pack(anchor="w", padx=14, pady=(0, 2))
        self.lp_entry_pass2 = ctk.CTkEntry(
            self.lp_step2, placeholder_text="Ulangi password baru",
            fg_color=C_BTN, text_color=C_ACCENT,
            border_color=C_BORDER,
            font=("Consolas", 12, "bold"), height=36, show="●")
        self.lp_entry_pass2.pack(fill="x", padx=14, pady=(0, 10))
        self.lp_entry_pass2.bind("<Return>", lambda e: self._submit_reset_password())

        ctk.CTkButton(
            self.lp_step2, text="🔒  Simpan Password Baru", height=40,
            fg_color=C_GREEN, hover_color="#20CC00",
            font=("Russo One", 11, "bold"), text_color=C_BG,
            command=self._submit_reset_password
        ).pack(fill="x", padx=14, pady=(0, 14))

        # Status
        self.lp_lbl_status = ctk.CTkLabel(self.lp_outer, text="",
                                            font=FONT_LABEL,
                                            text_color=C_RED,
                                            wraplength=380,
                                            justify="center")
        self.lp_lbl_status.pack(pady=(4, 6))

        # Kembali login
        ctk.CTkButton(self.lp_outer, text="← KEMBALI LOGIN", height=36,
                      fg_color="transparent", hover_color=C_BTN,
                      border_width=1, border_color=C_MUTED,
                      font=("Russo One", 9, "bold"), text_color=C_MUTED,
                      command=self._show_login_view
                      ).pack(padx=28, pady=(0, 22), fill="x")

    def _verifikasi_identitas(self):
        username = sanitize_text(self.lp_entry_username.get()).lower()
        email    = sanitize_text(self.lp_entry_email.get()).lower()

        if not username or not email:
            self.lp_lbl_status.configure(
                text="⚠  Username dan email wajib diisi.",
                text_color=C_YELLOW)
            return

        if not is_valid_username(username):
            self.lp_lbl_status.configure(
                text="⚠  Username tidak valid.",
                text_color=C_YELLOW)
            return

        if not is_valid_email(email):
            self.lp_lbl_status.configure(
                text="⚠  Format email tidak valid.",
                text_color=C_YELLOW)
            return

        cfg          = ConfigManager.load()
        users        = cfg.get("users", {})
        profil_semua = cfg.get("profil_rental", {})

        # Cek username terdaftar
        if username not in users:
            self.lp_lbl_status.configure(
                text="✖  Username tidak ditemukan.",
                text_color=C_RED)
            return

        # Cek email cocok
        profil = profil_semua.get(username, {})
        email_terdaftar = profil.get("email", "").strip().lower()

        if not email_terdaftar:
            self.lp_lbl_status.configure(
                text="✖  Akun ini tidak punya email terdaftar.\n"
                     "Hubungi admin/developer untuk reset password.",
                text_color=C_RED)
            return

        if email != email_terdaftar:
            self.lp_lbl_status.configure(
                text="✖  Email tidak cocok dengan data yang terdaftar.",
                text_color=C_RED)
            return

        # Generate verification code
        code = generate_verification_code()
        exp_time = (datetime.now() + timedelta(minutes=EMAIL_VERIFICATION_EXPIRY_MINUTES)).isoformat()
        
        # Store in pending verification
        pending_verif = cfg.get("pending_forgot_password_verifications", {})
        pending_verif[username] = {
            "code": code,
            "expires_at": exp_time,
            "email": email
        }
        cfg["pending_forgot_password_verifications"] = pending_verif
        ConfigManager.save(cfg)
        
        # Send email verification code
        self.lp_lbl_status.configure(
            text="📧  Mengirim kode verifikasi ke email...",
            text_color=C_MUTED)
        self.lp_lbl_status.update()
        
        threading.Thread(target=self._send_forgot_password_code, args=(username, email, code), daemon=True).start()

    def _send_forgot_password_code(self, username: str, email: str, code: str):
        """Send verification code via email."""
        smtp_settings = _get_email_settings()
        if not _email_configured():
            self.after(0, lambda: self.lp_lbl_status.configure(
                text="⚠  SMTP belum dikonfigurasi.",
                text_color=C_YELLOW))
            return

        try:
            msg = EmailMessage()
            msg['Subject'] = "🔐 Kode Verifikasi Lupa Password - RR Billing PRO"
            msg['From'] = smtp_settings.get('from_address', smtp_settings.get('smtp_username'))
            msg['To'] = email

            html_body = f"""
<html>
<body style="font-family: Arial, sans-serif; background-color: #f4f4f4; padding: 20px;">
    <div style="max-width: 600px; margin: 0 auto; background-color: white; border-radius: 10px; padding: 30px; box-shadow: 0 0 10px rgba(0,0,0,0.1);">
        <h2 style="color: #1e1e4a; text-align: center;">RR BILLING PRO</h2>
        <h3 style="color: #ff3366; text-align: center;">Permintaan Ubah Password</h3>
        
        <p style="color: #333; font-size: 14px;">Hai <strong>{username}</strong>,</p>
        <p style="color: #666; font-size: 14px;">Kami menerima permintaan untuk mengubah password akun Anda. Gunakan kode verifikasi di bawah ini:</p>
        
        <div style="background-color: #0d0d1a; border: 2px solid #7b2fff; border-radius: 8px; padding: 20px; text-align: center; margin: 20px 0;">
            <p style="font-size: 28px; font-weight: bold; color: #00ffcc; margin: 0; letter-spacing: 4px;">{code}</p>
            <p style="color: #6060a0; margin: 10px 0 0 0; font-size: 12px;">Kode berlaku selama 30 menit</p>
        </div>
        
        <p style="color: #666; font-size: 14px;">Langkah-langkah:</p>
        <ol style="color: #666; font-size: 14px;">
            <li>Masukkan kode verifikasi di aplikasi</li>
            <li>Set password baru Anda</li>
            <li>Login dengan password baru</li>
        </ol>
        
        <p style="color: #999; font-size: 12px; margin-top: 30px; border-top: 1px solid #eee; padding-top: 20px;">
            Jika Anda tidak membuat permintaan ini, abaikan email ini atau hubungi admin kami segera.
        </p>
    </div>
</body>
</html>
"""
            msg.set_content(f"Kode verifikasi: {code}")
            msg.add_alternative(html_body, subtype='html')

            context = ssl.create_default_context()
            with smtplib.SMTP(smtp_settings['smtp_server'], smtp_settings['smtp_port'], timeout=10) as smtp:
                if smtp_settings.get("use_tls", True):
                    smtp.starttls(context=context)
                smtp.login(smtp_settings['smtp_username'],
                           smtp_settings.get('smtp_password_enc') or smtp_settings.get('smtp_password', ""))
                smtp.send_message(msg)

            self.after(0, lambda: self._show_forgot_password_verify_code_view())
        except Exception as e:
            self.after(0, lambda e=e: self.lp_lbl_status.configure(
                text=f"❌  Gagal kirim email: {str(e)[:50]}",
                text_color=C_RED))

    def _show_forgot_password_verify_code_view(self):
        """Show verification code input form."""
        username = sanitize_text(self.lp_entry_username.get()).lower()
        
        # Clear and redesign the view - destroy only step 1 form, keep status and step2
        if hasattr(self, 'lp_step1_verif') and self.lp_step1_verif.winfo_exists():
            self.lp_step1_verif.destroy()

        # Re-create step1 with verification code input
        self.lp_step1_verif = ctk.CTkFrame(self.lp_outer, fg_color=C_CARD, corner_radius=12)
        self.lp_step1_verif.pack(fill="x", padx=28, pady=(6, 4), before=self.lp_lbl_status)

        ctk.CTkLabel(self.lp_step1_verif, text="📧  Cek Email Anda",
                     font=("Russo One", 12, "bold"),
                     text_color=C_GREEN).pack(pady=(12, 4))

        ctk.CTkLabel(self.lp_step1_verif, text="Kami sudah mengirim kode verifikasi ke email Anda.\nMasukkan kode 6 digit di bawah.",
                     font=FONT_SMALL, text_color=C_MUTED,
                     justify="center").pack(pady=(0, 10))

        ctk.CTkLabel(self.lp_step1_verif, text="Kode Verifikasi (6 digit):", font=FONT_LABEL,
                     text_color=C_MUTED, anchor="w").pack(anchor="w", padx=14, pady=(0, 2))
        
        self.lp_entry_verif_code = ctk.CTkEntry(
            self.lp_step1_verif, placeholder_text="mis. 123456",
            fg_color=C_BTN, text_color=C_ACCENT,
            border_color=C_BORDER,
            font=("Consolas", 16, "bold"), height=40)
        self.lp_entry_verif_code.pack(fill="x", padx=14, pady=(0, 10))
        self.lp_entry_verif_code.bind("<Return>", lambda e: self._verify_forgot_password_code())
        try:
            self.lp_entry_verif_code.focus()
        except Exception:
            pass

        ctk.CTkButton(
            self.lp_step1_verif, text="✅  Verifikasi Kode", height=38,
            fg_color=C_ACCENT2, hover_color="#4A20C8",
            font=("Russo One", 10, "bold"), text_color="white",
            command=self._verify_forgot_password_code).pack(fill="x", padx=14, pady=(0, 10))

        ctk.CTkButton(
            self.lp_step1_verif, text="📧  Kirim Ulang Kode", height=36,
            fg_color=C_BTN, hover_color=C_BORDER,
            font=("Russo One", 9, "bold"), text_color=C_MUTED,
            command=lambda: self._verifikasi_identitas()).pack(fill="x", padx=14, pady=(0, 14))

    def _verify_forgot_password_code(self):
        """Verify the code sent via email."""
        username = sanitize_text(self.lp_entry_username.get()).lower()
        code_input = sanitize_text(self.lp_entry_verif_code.get()).strip()

        if not code_input:
            self.lp_lbl_status.configure(
                text="⚠  Kode verifikasi wajib diisi.",
                text_color=C_YELLOW)
            return

        cfg = ConfigManager.load()
        pending = cfg.get("pending_forgot_password_verifications", {})
        
        if username not in pending:
            self.lp_lbl_status.configure(
                text="⚠  Data verifikasi tidak ditemukan. Mulai ulang dari awal.",
                text_color=C_YELLOW)
            return

        pending_data = pending[username]
        exp_dt = datetime.fromisoformat(pending_data["expires_at"])
        
        if datetime.now() > exp_dt:
            del pending[username]
            cfg["pending_forgot_password_verifications"] = pending
            ConfigManager.save(cfg)
            self.lp_lbl_status.configure(
                text="⚠  Kode verifikasi sudah kadaluarsa (berlaku 30 menit).\nSilakan minta kode baru.",
                text_color=C_YELLOW)
            return

        if code_input != pending_data["code"]:
            self.lp_lbl_status.configure(
                text="✖  Kode verifikasi tidak cocok.",
                text_color=C_RED)
            return

        # Code verified!
        self._lp_username_verified = username
        del pending[username]
        cfg["pending_forgot_password_verifications"] = pending
        ConfigManager.save(cfg)

        self.lp_lbl_status.configure(
            text=f"✅  Kode terverifikasi! Set password baru di bawah.",
            text_color=C_GREEN)

        # Hide verification code input step
        if hasattr(self, 'lp_step1_verif'):
            self.lp_step1_verif.pack_forget()
        
        # Show password reset form
        self.lp_step2.pack(fill="x", padx=28, pady=(6, 4))
        try:
            self.lp_entry_pass1.focus()
        except Exception:
            pass

    def _submit_reset_password(self):
        if not self._lp_username_verified:
            self.lp_lbl_status.configure(
                text="⚠  Lakukan verifikasi identitas dulu (Langkah 1).",
                text_color=C_YELLOW)
            return

        pw1 = self.lp_entry_pass1.get().strip()
        pw2 = self.lp_entry_pass2.get().strip()

        if not is_valid_password(pw1):
            self.lp_lbl_status.configure(
                text="⚠  Password baru harus minimal 8 karakter dan berisi huruf serta angka.",
                text_color=C_YELLOW)
            return
        if pw1 != pw2:
            self.lp_lbl_status.configure(
                text="⚠  Konfirmasi password tidak cocok.",
                text_color=C_YELLOW)
            self.lp_entry_pass2.delete(0, "end")
            try:
                self.lp_entry_pass2.focus()
            except Exception:
                pass
            return

        cfg   = ConfigManager.load()
        users = cfg.get("users", {})
        if self._lp_username_verified not in users:
            self.lp_lbl_status.configure(
                text="✖  User tidak ditemukan. Coba lagi.",
                text_color=C_RED)
            return

        # Simpan password baru (dengan bcrypt) — akan dienkripsi otomatis oleh ConfigManager.save()
        users[self._lp_username_verified]["password_enc"] = hash_password(pw1)
        cfg["users"] = users
        ConfigManager.save(cfg)
        AuditLogger.log(
            action="password_reset",
            username=self._lp_username_verified,
            status="success",
            details={"method": "forgot_password"}
        )

        uname = self._lp_username_verified
        self._lp_username_verified = None

        # Kembali ke login dengan pesan sukses
        self._show_login_view()
        self.entry_user.delete(0, "end")
        self.entry_user.insert(0, uname)
        try:
            self.entry_pass.focus()
        except Exception:
            pass
        self.lbl_status.configure(
            text=f"✅  Password '{uname}' berhasil direset — silakan login.",
            text_color=C_GREEN)

    # ══════════════════════════════════════════════════════════════════════════
    #  HELPER BERSAMA
    # ══════════════════════════════════════════════════════════════════════════
    def _cek_status_lisensi(self):
        current_user = getattr(self, 'current_user', None) or ""
        if not current_user:
            self.lbl_lic.configure(text="Silakan login untuk melanjutkan.", text_color=C_MUTED)
            return
        status = LicenseManager.get_status(current_user=current_user)
        if status["status"] == "expired":
            self.lbl_lic.configure(
                text=f"⚠ {status['pesan']} — Login masih bisa, tapi fitur terbatas.",
                text_color=C_RED)
        elif status["status"] == "trial":
            self.lbl_lic.configure(
                text=f"🕐 {status['pesan']}", text_color=C_YELLOW)
        else:
            self.lbl_lic.configure(text="✅ Lisensi Aktif", text_color=C_GREEN)

    def _login(self):
        if self._locked_until and datetime.now() < self._locked_until:
            sisa = int((self._locked_until - datetime.now()).total_seconds())
            self.lbl_status.configure(text=f"⛔ Terkunci — coba lagi dalam {sisa}s")
            return

        username  = sanitize_text(self.entry_user.get().lower())
        password  = self.entry_pass.get().strip()

        if not username or not password:
            self.lbl_status.configure(
                text="⚠  Username dan password wajib diisi.", text_color=C_YELLOW)
            return

        if not is_valid_username(username):
            self.lbl_status.configure(
                text="⚠  Username harus 4-20 karakter alfanumerik, titik, atau garis bawah.",
                text_color=C_YELLOW)
            return

        users = ConfigManager.get("users", self.DEFAULT_USERS)
        if not isinstance(users, dict):
            users = {}
        user_data = users.get(username) if isinstance(users, dict) else None
        # Read from password_enc (auto-decrypted by ConfigManager.load)
        password_hash = user_data.get("password_enc") if isinstance(user_data, dict) else ""
        if not password_hash:
            password_hash = user_data.get("password", "") if isinstance(user_data, dict) else ""

        # Akun kasir tidak bisa login lewat form admin — pakai tombol LOGIN KASIR
        if user_data and isinstance(user_data, dict) and user_data.get("role", "kasir") == "kasir":
            self.lbl_status.configure(
                text="⚠ Akun kasir — gunakan tombol '👤 LOGIN KASIR' di bawah.",
                text_color=C_YELLOW)
            return

        # Gunakan verify_password() yang support bcrypt dan legacy SHA256
        if user_data and verify_password(password, password_hash):
            self._attempt = 0
            AuditLogger.log(
                action="login_success",
                username=username,
                status="success",
                details={"role": user_data.get("role", "kasir")}
            )
            self.on_login_success(username, user_data.get("role", "kasir"))
            return

        self._attempt += 1
        if self._attempt >= self.LOGIN_MAX_ATTEMPTS:
            self._locked_until = datetime.now() + self.LOGIN_LOCK_DURATION
            self.lbl_status.configure(
                text=f"⛔ {self.LOGIN_MAX_ATTEMPTS}x salah — terkunci {int(self.LOGIN_LOCK_DURATION.total_seconds()/60)} menit",
                text_color=C_RED)
            AuditLogger.log(
                action="login_failed",
                username=username,
                status="locked",
                details={"attempts": self._attempt}
            )
        else:
            self.lbl_status.configure(
                text=f"✖ Username/Password salah ({self._attempt}/{self.LOGIN_MAX_ATTEMPTS})",
                text_color=C_RED)
            AuditLogger.log(
                action="login_failed",
                username=username,
                status="failed",
                details={"attempts": self._attempt}
            )

    def _kasir_accounts(self):
        users = ConfigManager.get("users", self.DEFAULT_USERS)
        if not isinstance(users, dict):
            return []
        return [uname for uname, u in users.items()
                if isinstance(u, dict) and u.get("role", "kasir") == "kasir"]

    def _show_kasir_login_view(self):
        for w in self._view_container.winfo_children():
            w.destroy()

        outer = ctk.CTkFrame(self._view_container, fg_color=C_PANEL,
                             corner_radius=20, border_width=2, border_color=C_GREEN)
        outer.pack(anchor="n", pady=(10, 10), padx=4, fill="x")

        ctk.CTkLabel(outer, text="RR BILLING PRO",
                     font=("Russo One", 22, "bold"), text_color=C_GREEN).pack(pady=(26, 0))
        ctk.CTkLabel(outer, text="LOGIN KASIR",
                     font=("Russo One", 12, "bold"), text_color=C_YELLOW).pack(pady=(2, 4))
        ctk.CTkLabel(outer, text="Pilih akun kasir Anda, lalu masukkan password.",
                     font=("Courier New", 10), text_color=C_MUTED).pack(pady=(0, 12))

        kasir_users = self._kasir_accounts()
        if not kasir_users:
            ctk.CTkLabel(outer,
                         text="Belum ada akun kasir.\nMinta admin mendaftarkan Anda terlebih dahulu.",
                         font=("Courier New", 11), text_color=C_YELLOW,
                         justify="center").pack(pady=10)
            ctk.CTkButton(outer, text="← Kembali", fg_color="transparent", hover_color=C_BTN,
                          border_width=1, border_color=C_BORDER, font=("Russo One", 11),
                          text_color=C_MUTED, command=self._show_login_view).pack(
                              pady=(4, 22), padx=30, fill="x")
            return

        ctk.CTkLabel(outer, text="Akun Kasir", font=("Consolas", 11),
                     text_color=C_MUTED).pack(padx=40)
        self.opt_kasir = ctk.CTkOptionMenu(outer, values=kasir_users,
                                           variable=ctk.StringVar(value=kasir_users[0]),
                                           fg_color=C_BTN, button_color=C_ACCENT2,
                                           button_hover_color="#5A0FCC",
                                           text_color=C_TEXT,
                                           dropdown_fg_color=C_CARD,
                                           dropdown_text_color=C_TEXT,
                                           font=("Consolas", 12), width=280)
        self.opt_kasir.pack(pady=(2, 10), padx=40)

        ctk.CTkLabel(outer, text="Password", font=("Consolas", 11),
                     text_color=C_MUTED).pack(padx=40)
        self.entry_kasir_pass = ctk.CTkEntry(
            outer, placeholder_text="Masukkan password",
            fg_color=C_BTN, text_color=C_ACCENT,
            border_color=C_BORDER, font=("Consolas", 14),
            height=40, width=320, show="●", justify="center")
        self.entry_kasir_pass.pack(pady=(2, 6), padx=40)
        self.entry_kasir_pass.bind("<Return>", lambda e: self._login_kasir())

        show_kasir_pw = ctk.CTkCheckBox(outer, text="👁 Lihat Password",
                                        fg_color=C_ACCENT2, hover_color=C_ACCENT,
                                        font=("Consolas", 10), text_color=C_MUTED,
                                        command=lambda: self.entry_kasir_pass.configure(
                                            show="" if show_kasir_pw.get() else "●"))
        show_kasir_pw.pack(pady=(0, 6))

        self.lbl_kasir_status = ctk.CTkLabel(outer, text="", font=("Consolas", 11),
                                             text_color=C_RED)
        self.lbl_kasir_status.pack(pady=(0, 8))

        ctk.CTkButton(outer, text="🔓  MASUK SEBAGAI KASIR", width=280, height=44,
                      fg_color=C_GREEN, hover_color="#0A6A3A",
                      font=("Russo One", 13, "bold"), text_color="white",
                      command=self._login_kasir).pack(pady=(0, 4), padx=30)

        ctk.CTkButton(outer, text="←  Kembali ke Login Admin", width=280, height=34,
                      fg_color="transparent", hover_color=C_BTN,
                      border_width=1, border_color=C_BORDER,
                      font=("Russo One", 10, "bold"), text_color=C_MUTED,
                      command=self._show_login_view).pack(pady=(6, 22), padx=30)

    def _login_kasir(self):
        username = self.opt_kasir.get() if hasattr(self, "opt_kasir") else ""
        username = sanitize_text(username).lower()
        password = self.entry_kasir_pass.get() if hasattr(self, "entry_kasir_pass") else ""
        password = password.strip()
        if not username or not password:
            if hasattr(self, "lbl_kasir_status"):
                self.lbl_kasir_status.configure(
                    text="⚠ Pilih akun kasir dan isi password.", text_color=C_YELLOW)
            return
        users = ConfigManager.get("users", self.DEFAULT_USERS)
        user_data = users.get(username) if isinstance(users, dict) else None
        if not user_data or not isinstance(user_data, dict) or user_data.get("role", "kasir") != "kasir":
            if hasattr(self, "lbl_kasir_status"):
                self.lbl_kasir_status.configure(text="✖ Akun bukan kasir.", text_color=C_RED)
            return
        password_hash = user_data.get("password_enc") or user_data.get("password", "")
        if verify_password(password, password_hash):
            self._attempt = 0
            AuditLogger.log(action="login_success", username=username, status="success",
                            details={"role": "kasir", "via": "login_kasir"})
            self.on_login_success(username, "kasir")
            return
        if hasattr(self, "lbl_kasir_status"):
            self.lbl_kasir_status.configure(text="✖ Password salah.", text_color=C_RED)

    def _login_google(self):
        # Aturan: login Google maksimal 10x per hari PER AKUN (per email).
        # Counter dihitung setelah auth sukses di bawah.
        self.lbl_status.configure(text="⏳ Membuka browser untuk login Google...")
        self.update()
        auth = get_firebase_auth()
        ok, msg = auth.login_with_google()
        if ok:
            email = auth.get_email() or ""
            # Batasan 10x/hari per akun Google (anti-spam auth ke Google/Firebase)
            try:
                limits = dict(ConfigManager.get("google_login_limits", {}) or {})
                today = date.today().isoformat()
                ent = limits.get(email) or {}
                if not isinstance(ent, dict) or ent.get("date") != today:
                    ent = {"date": today, "count": 0}
                if int(ent.get("count", 0) or 0) >= 10:
                    self.lbl_status.configure(
                        text="✖ Batas login Google hari ini (10x) untuk akun ini tercapai.",
                        text_color=C_RED)
                    return
                ent["count"] = int(ent.get("count", 0) or 0) + 1
                limits[email] = ent
                cfg = ConfigManager.load()
                cfg["google_login_limits"] = limits
                ConfigManager.save(cfg)
            except Exception:
                pass
            nama = auth.get_display_name() or ""
            self.lbl_status.configure(text="✅ Login Google berhasil", text_color=C_GREEN)
            uname = self._find_username_by_email(email)
            # Cari username Firestore yang punya licenseStatus aktif untuk email ini
            try:
                fc = FirestoreClient()
                fb_results = fc.query_where_equal("billingps_users", "email", email)
                for r in fb_results:
                    fb_uname = r.get("_id", "")
                    # Bersihkan prefix _user_
                    if fb_uname.startswith("_user_"):
                        bare = fb_uname[6:]
                        bare_doc = fc.get_user_doc(bare)
                        if bare_doc and bare_doc.get("licenseStatus", {}).get("status") == "active":
                            _LOGGER.info("Using Firestore username '%s' (active licenseStatus)", bare)
                            uname = bare
                            break
                    elif r.get("licenseStatus", {}).get("status") == "active":
                        _LOGGER.info("Using Firestore username '%s' (active licenseStatus)", fb_uname)
                        uname = fb_uname
                        break
            except Exception:
                pass
            if not uname:
                try:
                    fc = FirestoreClient()
                    uname = fc.find_username_by_email(email)
                except Exception:
                    pass
            # Sinkron email ke doc cloud agar login Gmail berikutnya langsung cocok
            if uname and email:
                try:
                    fc = FirestoreClient()
                    fc.set_user_doc(uname, {"email": email}, merge=True)
                except Exception:
                    pass
            if uname:
                # Pastikan akun terdaftar di config lokal (username dari Firestore
                # belum tentu ada di 'users'), agar ganti password & fitur lain jalan.
                try:
                    def _provisi(cfg):
                        users = cfg.get("users")
                        if not isinstance(users, dict):
                            users = {}
                        if uname not in users:
                            users[uname] = {"password_enc": hash_password("google_" + uname),
                                            "role": "admin", "email": email}
                            cfg["users"] = users
                        return cfg
                    ConfigManager.update(_provisi)
                except Exception:
                    pass
                self.on_login_success(uname, "admin")
            else:
                self._show_register_google_dialog(email, nama)
        else:
            self.lbl_status.configure(text=f"✖ {msg}", text_color=C_RED)

    def _show_register_google_dialog(self, email, nama_google):
        username_auto = email.split("@")[0].replace(".", "_").replace("-", "_").lower()
        dlg = ctk.CTkToplevel(self.winfo_toplevel())
        dlg.title("📝  Lengkapi Data Rental")
        dlg.geometry("420x440")
        dlg.configure(fg_color=C_BG)
        dlg.transient(self.winfo_toplevel())
        dlg.resizable(False, False)
        dlg.grab_set()

        scroll = ctk.CTkScrollableFrame(dlg, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=16, pady=16)

        ctk.CTkLabel(scroll, text="🎉  Akun Baru Ditemukan!",
                     font=("Russo One", 16, "bold"), text_color=C_ACCENT).pack(pady=(8, 4))
        ctk.CTkLabel(scroll, text="Lengkapi data rental Anda untuk melanjutkan.",
                     font=FONT_BODY, text_color=C_MUTED).pack(pady=(0, 16))

        card = ctk.CTkFrame(scroll, fg_color=C_CARD, corner_radius=14)
        card.pack(fill="x")

        fields = [
            ("email",     "📧  Email",          email, False),
            ("nama",      "👤  Nama Pemilik",   nama_google, False),
            ("username",  "🔑  Username",       username_auto, False),
        ]
        entries = {}
        for key, label, val, _ in fields:
            ctk.CTkLabel(card, text=label, font=FONT_LABEL, text_color=C_MUTED).pack(anchor="w", padx=18, pady=(10, 2))
            e = ctk.CTkEntry(card, fg_color=C_BTN, text_color=C_TEXT,
                             border_color=C_BORDER, font=FONT_BODY, height=34)
            e.insert(0, val)
            e.configure(state="readonly")
            e.pack(fill="x", padx=18, pady=(0, 2))
            entries[key] = e

        # Editable fields
        ctk.CTkLabel(card, text="🏪  Nama Rental *", font=FONT_LABEL, text_color=C_ACCENT2).pack(anchor="w", padx=18, pady=(10, 2))
        entry_rental = ctk.CTkEntry(card, placeholder_text="contoh: RR Gaming PS", fg_color=C_BTN, text_color=C_TEXT,
                                     border_color=C_BORDER, font=FONT_BODY, height=34)
        entry_rental.pack(fill="x", padx=18, pady=(0, 2))

        ctk.CTkLabel(card, text="📍  Alamat Rental *", font=FONT_LABEL, text_color=C_ACCENT2).pack(anchor="w", padx=18, pady=(10, 2))
        entry_alamat = ctk.CTkTextbox(card, height=60, fg_color=C_BTN, text_color=C_TEXT,
                                       border_color=C_BORDER, font=FONT_BODY)
        entry_alamat.pack(fill="x", padx=18, pady=(0, 2))

        ctk.CTkLabel(card, text="📱  No. WhatsApp *", font=FONT_LABEL, text_color=C_ACCENT2).pack(anchor="w", padx=18, pady=(10, 2))
        entry_wa = ctk.CTkEntry(card, placeholder_text="08xxxxxxxxxx", fg_color=C_BTN, text_color=C_TEXT,
                                 border_color=C_BORDER, font=FONT_BODY, height=34)
        entry_wa.pack(fill="x", padx=18, pady=(0, 16))

        lbl_error = ctk.CTkLabel(scroll, text="", font=FONT_LABEL, text_color=C_RED)
        lbl_error.pack(pady=(4, 2))

        def _save():
            try:
                rental = entry_rental.get().strip()
                alamat = entry_alamat.get("1.0", "end").strip()
                wa = entry_wa.get().strip()
                if not rental or not alamat or not wa:
                    lbl_error.configure(text="⚠  Nama Rental, Alamat, dan No WA wajib diisi.", text_color=C_YELLOW)
                    return
                uname = username_auto
                cfg = ConfigManager.load()
                users = cfg.get("users", {})
                if uname in users:
                    uname = uname + "_" + str(len(users) + 1)
                users[uname] = {"password_enc": hash_password("google_" + uname), "role": "admin", "email": email}
                cfg["users"] = users
                profil = cfg.get("profil_rental", {})
                profil[uname] = {
                    "nama_pemilik":   nama_google or rental,
                    "nama_rental":    rental,
                    "email":          email,
                    "hp":             wa,
                    "alamat":         alamat,
                    "tanggal_daftar": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "sumber":         "google",
                }
                cfg["profil_rental"] = profil
                ConfigManager.save(cfg)
                # Sinkron email ke doc cloud agar login Gmail berikutnya langsung cocok
                try:
                    from firestore_sync import FirestoreClient as _FC
                    _FC().set_user_doc(uname, {"email": email}, merge=True)
                except Exception:
                    pass
                AuditLogger.log(action="register_google", username=uname, status="success",
                                details={"email": email, "sumber": "google"})
                LicenseManager._set_trial_status_in_config(uname, date.today())
                dlg.destroy()
                self.on_login_success(uname, "admin")
            except Exception as e:
                import traceback
                traceback.print_exc()
                lbl_error.configure(text=f"✖  Error: {e}", text_color=C_RED)

        btn_frame = ctk.CTkFrame(scroll, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(8, 4))
        ctk.CTkButton(btn_frame, text="✅  Daftar & Masuk", height=40,
                      fg_color=C_ACCENT2, font=("Russo One", 12, "bold"), text_color="white",
                      command=_save).pack(fill="x")
        ctk.CTkButton(btn_frame, text="❌  Batal", height=36,
                      fg_color=C_RED, font=("Russo One", 10, "bold"), text_color="white",
                      command=dlg.destroy).pack(fill="x", pady=(6, 0))

    def _find_username_by_email(self, email: str):
        try:
            profil = ConfigManager.get("profil_rental", {})
            for uname, p in profil.items():
                if isinstance(p, dict) and p.get("email", "").lower() == email.lower():
                    return uname
            users = ConfigManager.get("users", {})
            for uname, u in users.items():
                if isinstance(u, dict) and u.get("email", "").lower() == email.lower():
                    return uname
        except Exception:
            pass
        return None


# ═══════════════════════════════════════════════════════════════════════════════
#  DIALOG TES KONEKSI (Android Remote v2 — tanpa port)
# ═══════════════════════════════════════════════════════════════════════════════
class DialogGantiPort(ctk.CTkToplevel):
    """Tes koneksi ulang ke TV via Android Remote v2 (tidak perlu port)."""
    def __init__(self, master, label_tv, ip, port_lama, on_confirm):
        super().__init__(master)
        self.title(f"Tes Koneksi — {label_tv}")
        self.geometry("360x240")
        self.configure(fg_color=C_BG)
        self.transient(master)
        self.resizable(False, False)
        self.label_tv = label_tv
        self.ip = ip
        self.on_confirm = on_confirm
        self._connected = False
        self._build()
        center_window(self, master, width=360, height=240)
        self.after(50, self.grab_set)

    def _build(self):
        ctk.CTkLabel(self, text=f"📺  {self.label_tv}",
                     font=FONT_TITLE, text_color=C_ACCENT).pack(pady=(22, 4))
        ctk.CTkLabel(self, text="Tes koneksi ke TV via Android Remote v2",
                     font=FONT_BODY, text_color=C_MUTED).pack(pady=(0, 14))
        info = ctk.CTkFrame(self, fg_color=C_PANEL, corner_radius=10)
        info.pack(fill="x", padx=28, pady=4)
        ctk.CTkLabel(info, text=f"IP: {self.ip}",
                     font=("Consolas", 12, "bold"), text_color=C_ACCENT).pack(pady=8)
        self.lbl_status = ctk.CTkLabel(self, text="⬤  Belum diuji",
                                        font=FONT_BODY, text_color=C_MUTED)
        self.lbl_status.pack(pady=8)
        btn_f = ctk.CTkFrame(self, fg_color="transparent")
        btn_f.pack(pady=6)
        self.btn_tes = ctk.CTkButton(btn_f, text="🔗  Tes Koneksi", width=140, height=36,
                                      fg_color=C_BTN, border_width=1, border_color=C_GREEN,
                                      font=FONT_SUB, text_color=C_GREEN,
                                      command=self._tes_koneksi)
        self.btn_tes.pack(side="left", padx=8)
        self.btn_simpan = ctk.CTkButton(btn_f, text="✅  Simpan", width=100, height=36,
                                         fg_color=C_ACCENT2, font=FONT_SUB, text_color="white",
                                         state="disabled", command=self._konfirmasi)
        self.btn_simpan.pack(side="left", padx=8)
        ctk.CTkButton(btn_f, text="✖  Tutup", width=100, height=36,
                      fg_color=C_RED, font=FONT_SUB, text_color="white",
                      command=self.destroy).pack(side="left", padx=8)

    def _tes_koneksi(self):
        self.btn_tes.configure(text="⏳...", state="disabled")
        threading.Thread(target=self._connect_thread, daemon=True).start()

    def _connect_thread(self):
        try:
            sukses, status, pesan = ADBHelper.cek_dan_reconnect(self.ip)
            self.after(0, self._update_status, sukses, pesan)
        except Exception as e:
            self.after(0, self._update_status, False, str(e))

    def _update_status(self, sukses, pesan=""):
        self.btn_tes.configure(state="normal", text="🔗  Tes Koneksi")
        if sukses:
            self._connected = True
            self.lbl_status.configure(text=f"✅  TERHUBUNG — {self.ip}", text_color=C_GREEN)
            self.btn_simpan.configure(state="normal")
        else:
            self.lbl_status.configure(text=f"✖  GAGAL  {pesan}", text_color=C_RED)
            self.btn_simpan.configure(state="disabled")

    def _konfirmasi(self):
        if self._connected:
            self.on_confirm(None)
            self.destroy()


class DialogGantiIP(ctk.CTkToplevel):
    def __init__(self, master, label_tv, ip_lama, port_lama, on_confirm):
        super().__init__(master)
        self.title(f"Ganti IP — {label_tv}")
        self.geometry("360x300")
        self.configure(fg_color=C_BG)
        self.transient(master)
        self.resizable(False, False)
        self.label_tv = label_tv
        self.ip_lama = ip_lama
        self.on_confirm = on_confirm
        self._connected = False
        self._build()
        center_window(self, master, width=360, height=300)
        self.after(50, self.grab_set)

    def _build(self):
        ctk.CTkLabel(self, text=f"📺  {self.label_tv}", font=FONT_TITLE, text_color=C_ACCENT).pack(pady=(16, 4))
        ctk.CTkLabel(self, text="Masukkan IP baru lalu tes koneksi",
                     font=FONT_BODY, text_color=C_MUTED).pack(pady=(0, 10))

        row = ctk.CTkFrame(self, fg_color=C_PANEL, corner_radius=10)
        row.pack(fill="x", padx=18, pady=4)

        ip_row = ctk.CTkFrame(row, fg_color="transparent")
        ip_row.pack(fill="x", padx=10, pady=(10, 4))
        ctk.CTkLabel(ip_row, text="IP:", font=FONT_LABEL, text_color=C_MUTED, width=40).pack(side="left")
        self.entry_ip = ctk.CTkEntry(ip_row, fg_color=C_BTN, text_color=C_ACCENT,
                                     border_color=C_BORDER, font=("Consolas", 12, "bold"), height=34)
        self.entry_ip.pack(side="left", fill="x", expand=True, padx=(0, 8))
        self.entry_ip.insert(0, self.ip_lama)

        self.btn_tes = ctk.CTkButton(self, text="🔗 Tes Koneksi", width=140, height=34,
                                      fg_color=C_BTN, border_width=1, border_color=C_GREEN,
                                      font=FONT_SUB, text_color=C_GREEN,
                                      command=self._tes_koneksi)
        self.btn_tes.pack(pady=(6, 4))

        self.lbl_status = ctk.CTkLabel(self, text="⬤  Belum diuji",
                                        font=FONT_BODY, text_color=C_MUTED)
        self.lbl_status.pack(pady=(0, 8))

        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.pack(fill="x", padx=18, pady=(8, 18))
        btn_frame.columnconfigure(0, weight=1)
        btn_frame.columnconfigure(1, weight=1)
        ctk.CTkButton(btn_frame, text="✖  Batal", fg_color=C_RED, font=FONT_SUB, height=36, command=self.destroy).grid(row=0, column=0, sticky="we", padx=(0, 6))
        ctk.CTkButton(btn_frame, text="✅  Simpan", fg_color=C_ACCENT2, font=FONT_SUB, height=36, command=self._konfirmasi).grid(row=0, column=1, sticky="we", padx=(6, 0))

        self.entry_ip.focus_set()

    def _tes_koneksi(self):
        ip = self.entry_ip.get().strip()
        if not ip:
            self.lbl_status.configure(text="⚠  IP tidak valid", text_color=C_YELLOW)
            return
        self.btn_tes.configure(text="⏳...", state="disabled")
        threading.Thread(target=self._connect_thread, args=(ip,), daemon=True).start()

    def _connect_thread(self, ip, method="atpv2"):
        try:
            ADBHelper.set_connection_method(ip, method)
            sukses, status_awal, pesan = ADBHelper.cek_dan_reconnect(ip)
        except Exception as e:
            sukses, status_awal, pesan = False, 'error', str(e)
        self.after(0, self._update_status, sukses, ip, pesan)

    def _update_status(self, sukses, ip, pesan=""):
        self.btn_tes.configure(state="normal", text="🔗 Tes Koneksi")
        if sukses:
            self._connected = True
            self.lbl_status.configure(text=f"✅ Terhubung — {ip}", text_color=C_GREEN)
        else:
            self._connected = False
            self.lbl_status.configure(text=f"✖ Gagal: {pesan}", text_color=C_RED)
            if "sertifikat" in pesan.lower() or "pairing" in pesan.lower():
                if messagebox.askyesno("Pairing Diperlukan",
                                       "TV ini belum pernah di-pair. Lakukan pairing sekarang?",
                                       parent=self):
                    threading.Thread(target=self._pairing_flow, args=(ip,), daemon=True).start()

    def _pairing_flow(self, ip):
        try:
            result = tv_mesin.pair_tv_sync(ip)
            if result.get("status") != "pairing_started":
                self.after(0, lambda: self.lbl_status.configure(
                    text=f"✖ Gagal pairing: {result.get('message', '?')}", text_color=C_RED))
                return
            remote_obj = result.get("remote")
            pin_holder = {"value": None}
            def on_pin(pin):
                pin_holder["value"] = pin
            def on_cancel():
                pin_holder["value"] = "CANCEL"
            self.after(0, lambda: DialogPinInput(self, on_pin, on_cancel))
            while pin_holder["value"] is None:
                import time as _t
                _t.sleep(0.1)
            if pin_holder["value"] == "CANCEL":
                self.after(0, lambda: self.lbl_status.configure(text="⛔ Pairing dibatalkan", text_color=C_YELLOW))
                return
            fin = tv_mesin.finish_pair_sync(remote_obj, pin_holder["value"])
            if fin.get("status") != "paired":
                self.after(0, lambda: self.lbl_status.configure(
                    text=f"✖ PIN salah: {fin.get('message', '?')}", text_color=C_RED))
                return
            ADBHelper.set_connection_method(ip, "atpv2")
            sukses2, _, pesan2 = ADBHelper.cek_dan_reconnect(ip)
            if sukses2:
                self.after(0, lambda: self.lbl_status.configure(
                    text=f"✅ Pairing & terhubung — {ip}", text_color=C_GREEN))
                self.after(0, lambda: setattr(self, '_connected', True))
            else:
                self.after(0, lambda: self.lbl_status.configure(
                    text=f"✖ Gagal connect setelah pair: {pesan2}", text_color=C_RED))
        except Exception as e:
            self.after(0, lambda e=e: self.lbl_status.configure(
                text=f"✖ Error: {e}", text_color=C_RED))

    def _konfirmasi(self):
        new_ip = self.entry_ip.get().strip()
        if not new_ip:
            messagebox.showwarning("Input Salah", "IP tidak valid.")
            return
        if not self._connected:
            self.lbl_status.configure(text="⚠  Silakan tes koneksi terlebih dahulu", text_color=C_YELLOW)
            return
        self.on_confirm(new_ip, 0)
        self.destroy()


# ═══════════════════════════════════════════════════════════════════════════════
#  DIALOG TAMBAH TV
# ═══════════════════════════════════════════════════════════════════════════════
class DialogPinInput(ctk.CTkToplevel):
    """Popup kecil untuk input kode pairing alfanumerik dari TV."""
    def __init__(self, master, on_submit, on_cancel):
        super().__init__(master)
        self.on_submit = on_submit
        self.on_cancel = on_cancel
        self.title("Kode Pairing dari TV")
        self.geometry("360x200")
        self.configure(fg_color=C_BG)
        self.transient(master)
        self.resizable(False, False)
        self.grab_set()
        ctk.CTkLabel(self, text="🔐  Masukkan kode pairing yang muncul di TV",
                     font=FONT_BODY, text_color=C_YELLOW).pack(pady=(18, 4))
        ctk.CTkLabel(self, text="(alfanumerik, 4-6 karakter)",
                     font=FONT_SMALL, text_color=C_MUTED).pack(pady=(0, 6))
        self.entry_pin = ctk.CTkEntry(self, placeholder_text="contoh: A1B2C3",
                                       font=("Consolas", 20, "bold"), width=200, height=44,
                                       fg_color=C_BTN, text_color=C_GREEN, border_color=C_GREEN,
                                       justify="center")
        self.entry_pin.pack(pady=4)
        self.entry_pin.focus_set()
        btn_f = ctk.CTkFrame(self, fg_color="transparent")
        btn_f.pack(pady=8)
        ctk.CTkButton(btn_f, text="✅  Konfirmasi", width=120,
                      fg_color="#1A3A1A", border_width=1, border_color=C_GREEN,
                      text_color=C_GREEN, command=self._submit).pack(side="left", padx=6)
        ctk.CTkButton(btn_f, text="✖  Batal", width=100,
                      fg_color=C_RED, command=self._cancel).pack(side="left", padx=6)

    def _submit(self):
        raw = self.entry_pin.get().strip().upper()
        if not raw:
            messagebox.showwarning("Kode Kosong", "Masukkan kode pairing dari TV", parent=self)
            return
        if len(raw) < 4 or len(raw) > 6:
            messagebox.showwarning("Kode Salah", "Kode pairing terdiri dari 4-6 karakter", parent=self)
            return
        self.on_submit(raw)
        self.destroy()

    def _cancel(self):
        self.on_cancel()
        self.destroy()


class DialogKonfirmasiBayar(ctk.CTkToplevel):
    """Konfirmasi status pembayaran (BAYAR=LUNAS / BELUM= TAGIHAN) yang muncul
    setelah user menekan MULAI SESI / TAMBAHKAN PESANAN."""
    def __init__(self, master, on_submit, judul="Status Pembayaran",
                 rincian="", default_lunas=True):
        super().__init__(master)
        self.on_submit = on_submit
        self.title("Konfirmasi Pembayaran")
        center_window(self, master, width=400, height=280)
        self.configure(fg_color=C_BG)
        self.transient(master)
        self.resizable(False, False)
        ctk.CTkLabel(self, text=f"💳  {judul}",
                     font=("Russo One", 14, "bold"), text_color=C_ACCENT).pack(pady=(16, 2))
        if rincian:
            ctk.CTkLabel(self, text=rincian, font=FONT_SMALL,
                         text_color=C_MUTED, wraplength=340).pack(pady=(0, 10))
        ctk.CTkLabel(self, text="Pilih status pembayaran:",
                     font=FONT_BODY, text_color=C_TEXT).pack(pady=(0, 4))
        btn_f = ctk.CTkFrame(self, fg_color="transparent")
        btn_f.pack(pady=6)
        ctk.CTkButton(btn_f, text="✅  BAYAR  (LUNAS)", width=170, height=46,
                      fg_color=C_GREEN, hover_color="#2E7D32", text_color="white",
                      font=("Russo One", 12, "bold"),
                      command=lambda: self._finish(True)).pack(side="left", padx=8)
        ctk.CTkButton(btn_f, text="⏳  TAGIHAN\n(BELUM BAYAR)", width=170, height=46,
                      fg_color="#FFCC00", hover_color="#E6B800", text_color="black",
                      font=("Russo One", 11, "bold"),
                      command=lambda: self._finish(False)).pack(side="left", padx=8)
        self.after(50, self.grab_set)
        self.lift()

    def _finish(self, paid):
        try:
            self.on_submit(paid)
        finally:
            self.destroy()


class DialogPelangganAkhir(ctk.CTkToplevel):
    """Input data pelanggan saat sesi selesai dan memilih TAGIHAN (belum bayar)."""
    def __init__(self, master, on_submit):
        super().__init__(master)
        self.on_submit = on_submit
        self.title("Data Pelanggan Tagihan")
        center_window(self, master, width=380, height=260)
        self.configure(fg_color=C_BG)
        self.transient(master)
        self.resizable(False, False)
        ctk.CTkLabel(self, text="👤  Data Pelanggan (TAGIHAN)",
                     font=("Russo One", 13, "bold"), text_color="#FFCC00").pack(pady=(14, 2))
        ctk.CTkLabel(self, text="Transaksi dicatat sebagai TAGIHAN (belum bayar).\n"
                                "Isi data pelanggan untuk pencatatan:",
                     font=FONT_SMALL, text_color=C_MUTED).pack(pady=(0, 6))
        f1 = ctk.CTkFrame(self, fg_color="transparent")
        f1.pack(fill="x", padx=24, pady=2)
        ctk.CTkLabel(f1, text="Nama:", width=60, anchor="w", font=FONT_LABEL,
                     text_color=C_MUTED).pack(side="left")
        self.entry_nama = ctk.CTkEntry(f1, width=220, font=FONT_BODY,
                                       fg_color=C_BTN, text_color=C_TEXT)
        self.entry_nama.pack(side="left")
        f2 = ctk.CTkFrame(self, fg_color="transparent")
        f2.pack(fill="x", padx=24, pady=2)
        ctk.CTkLabel(f2, text="No. HP:", width=90, anchor="w", font=FONT_LABEL,
                     text_color=C_MUTED).pack(side="left")
        self.entry_hp = ctk.CTkEntry(f2, width=220, font=FONT_BODY,
                                     fg_color=C_BTN, text_color=C_TEXT)
        self.entry_hp.pack(side="left")
        btn_f = ctk.CTkFrame(self, fg_color="transparent")
        btn_f.pack(pady=(10, 12))
        ctk.CTkButton(btn_f, text="💾  SIMPAN TAGIHAN", width=150,
                      fg_color="#FFCC00", hover_color="#E6B800", text_color="black",
                      font=("Russo One", 11, "bold"), command=self._submit).pack(side="left", padx=6)
        ctk.CTkButton(btn_f, text="✖ Batal", width=100, fg_color=C_RED,
                      command=self.destroy).pack(side="left", padx=6)
        self.after(50, self.grab_set)
        self.entry_nama.focus_set()

    def _submit(self):
        nama = self.entry_nama.get().strip()
        if not nama:
            messagebox.showwarning("Nama Kosong", "Nama pelanggan wajib diisi.", parent=self)
            return
        self.on_submit(nama, self.entry_hp.get().strip())
        self.destroy()


class DialogTambahTV(ctk.CTkToplevel):
    """Form tambah TV: Grup Tarif, Nama, IP → langsung pairing + input PIN."""
    def __init__(self, master, nomor_tv, on_confirm, on_close_cb, daftar_grup=None):
        super().__init__(master)
        self.title(f"Tambah TV #{nomor_tv}")
        self.geometry("380x420")
        self.configure(fg_color=C_BG)
        self.transient(master)
        self.resizable(False, False)
        self.nomor_tv = nomor_tv
        self.on_confirm = on_confirm
        self.on_close_cb = on_close_cb
        self._confirmed = False
        self._pairing_remote = None
        self._pairing_ip = ""
        self.daftar_grup = daftar_grup or [NAMA_GRUP_DEFAULT]
        self.grup_var = ctk.StringVar(value=self.daftar_grup[0])
        self.protocol("WM_DELETE_WINDOW", self._on_close)
        self._build()
        self._build_status_bar()
        center_window(self, master, width=380, height=420)
        self.after(50, self.grab_set)

    def _build(self):
        ctk.CTkLabel(self, text=f"📺  Tambah TV #{self.nomor_tv}",
                     font=FONT_TITLE, text_color=C_ACCENT).pack(pady=(18, 6))
        # Grup
        grup_f = ctk.CTkFrame(self, fg_color=C_PANEL, corner_radius=10)
        grup_f.pack(fill="x", padx=28)
        ctk.CTkLabel(grup_f, text="🏷  Grup Tarif",
                     font=FONT_LABEL, text_color=C_MUTED).pack(anchor="w", padx=14, pady=(10, 2))
        self.opt_grup = ctk.CTkOptionMenu(grup_f, values=self.daftar_grup, variable=self.grup_var,
                                          fg_color=C_BTN, button_color=C_ACCENT2,
                                          text_color=C_TEXT, font=FONT_BODY,
                                          dropdown_fg_color=C_CARD, dropdown_text_color=C_TEXT)
        self.opt_grup.pack(fill="x", padx=14, pady=(0, 10))
        # Nama
        nama_f = ctk.CTkFrame(self, fg_color=C_PANEL, corner_radius=10)
        nama_f.pack(fill="x", padx=28, pady=(8, 0))
        ctk.CTkLabel(nama_f, text="✏️  Nama TV",
                     font=FONT_LABEL, text_color=C_MUTED).pack(anchor="w", padx=14, pady=(8, 2))
        self.entry_nama = ctk.CTkEntry(nama_f, placeholder_text=f"TV {self.nomor_tv}",
                                       fg_color=C_BTN, text_color=C_TEXT,
                                       border_color=C_BORDER, font=FONT_BODY, height=34)
        self.entry_nama.pack(fill="x", padx=14, pady=(0, 10))
        # IP
        ip_f = ctk.CTkFrame(self, fg_color=C_PANEL, corner_radius=10)
        ip_f.pack(fill="x", padx=28, pady=(8, 0))
        ctk.CTkLabel(ip_f, text="🌐  IP Address TV",
                     font=FONT_LABEL, text_color=C_MUTED).pack(anchor="w", padx=14, pady=(8, 2))
        self.entry_ip = ctk.CTkEntry(ip_f, placeholder_text="192.168.1.xxx",
                                      fg_color=C_BTN, text_color=C_ACCENT,
                                      border_color=C_BORDER, font=("Consolas", 13, "bold"), height=34)
        self.entry_ip.pack(fill="x", padx=14, pady=(0, 10))
    def _build_status_bar(self):
        self.lbl_status = ctk.CTkLabel(self, text="", font=FONT_SMALL, text_color=C_MUTED)
        self.lbl_status.pack(pady=(8, 2))
        btn_f = ctk.CTkFrame(self, fg_color="transparent")
        btn_f.pack(pady=(4, 14))
        self.btn_tambah = ctk.CTkButton(btn_f, text="✅  Tambah TV", width=140, height=36,
                                         fg_color=C_ACCENT2, font=FONT_SUB, text_color="white",
                                         command=self._tambah_tv)
        self.btn_tambah.pack(side="left", padx=6)
        ctk.CTkButton(btn_f, text="✖  Batal", width=100, height=36,
                      fg_color=C_RED, font=FONT_SUB, text_color="white",
                      command=self._on_close).pack(side="left", padx=6)

    def _tambah_tv(self):
        ip = self.entry_ip.get().strip()
        nama = self.entry_nama.get().strip()
        if not ip:
            self.lbl_status.configure(text="⚠  Isi IP Address TV", text_color=C_YELLOW)
            return
        if not nama:
            self.lbl_status.configure(text="⚠  Isi Nama TV", text_color=C_YELLOW)
            return
        self._pairing_ip = ip
        self.btn_tambah.configure(state="disabled", text="⏳  Pairing...")
        self.lbl_status.configure(text="⏳  Memulai pairing...", text_color=C_YELLOW)
        threading.Thread(target=self._pair_start_thread, args=(ip,), daemon=True).start()

    def _pair_start_thread(self, ip):
        try:
            result = tv_mesin.pair_tv_sync(ip)
            self.after(0, self._on_pair_started, result)
        except Exception as e:
            self.after(0, self._on_pair_error, str(e))

    def _on_pair_started(self, result):
        if result.get("status") == "pairing_started":
            self._pairing_remote = result.get("remote")
            self.lbl_status.configure(
                text=f"✅  Pairing dimulai — masukkan PIN dari TV ({result.get('device_name', 'TV')})",
                text_color=C_GREEN)
            DialogPinInput(self,
                on_submit=self._on_pin_submit,
                on_cancel=self._on_pin_cancel)
        else:
            self._pair_error(result.get("message", "Gagal memulai pairing"))

    def _on_pair_error(self, msg):
        self._pair_error(msg)

    def _pair_error(self, msg):
        self.lbl_status.configure(text=f"✖  {msg}", text_color=C_RED)
        self.btn_tambah.configure(state="normal", text="✅  Tambah TV")

    def _on_pin_submit(self, pin):
        self.lbl_status.configure(text="⏳  Menyelesaikan pairing...", text_color=C_YELLOW)
        threading.Thread(target=self._finish_thread, args=(pin,), daemon=True).start()

    def _on_pin_cancel(self):
        self.lbl_status.configure(text="✖  Pairing dibatalkan", text_color=C_RED)
        self.btn_tambah.configure(state="normal", text="✅  Tambah TV")
        self._pairing_remote = None

    def _finish_thread(self, pin):
        remote = self._pairing_remote
        if not remote:
            self.after(0, self._pair_error, "Tidak ada sesi pairing")
            return
        try:
            fin = tv_mesin.finish_pair_sync(remote, pin)
            if fin.get("status") == "paired":
                self.after(0, self._connect_after_pair)
            else:
                self.after(0, self._pair_error, fin.get("message", "PIN salah"))
        except Exception as e:
            self.after(0, self._pair_error, str(e))

    def _connect_after_pair(self):
        ip = self._pairing_ip
        self.lbl_status.configure(text="⏳  Pair OK — menghubungkan...", text_color=C_GREEN)
        try:
            ADBHelper.set_connection_method(ip, "atpv2")
            rem = tv_mesin.AndroidTVRemote()
            conn = rem.connect_blocking(ip)
            if conn.get("status") == "connected":
                ADBHelper._instances[ip] = rem
                self._confirmed = True
                nama = self.entry_nama.get().strip() or f"TV {self.nomor_tv}"
                grup = self.grup_var.get() or NAMA_GRUP_DEFAULT
                self.on_confirm(ip, nama, 0, grup)
                self.destroy()
            else:
                self._pair_error(conn.get("message", "Gagal connect"))
        except Exception as e:
            self._pair_error(str(e))

    def _on_close(self):
        if not self._confirmed:
            self.on_close_cb()
        self.destroy()


# ═══════════════════════════════════════════════════════════════════════════════
#  DIALOG TAMBAH PESANAN (MAKANAN/MINUMAN SAAT SESI BERJALAN)
# ═══════════════════════════════════════════════════════════════════════════════
class DialogTambahPesanan(ctk.CTkToplevel):
    """Dialog untuk tambah pesanan makanan/minuman saat sesi TV sedang berjalan."""
    def __init__(self, master, tv_label, on_confirm, makanan_data, minuman_data, pesanan_aktif=None,
                 paket_harga=0, paket_label=""):
        super().__init__(master)
        self.title(f"Tambah Pesanan — {tv_label}")
        self.geometry("380x460")
        self.configure(fg_color=C_BG)
        self.transient(master)
         
        self.tv_label = tv_label
        self.on_confirm = on_confirm
        self.makanan_data = makanan_data or {}
        self.minuman_data = minuman_data or {}
        self.pesanan_aktif = pesanan_aktif or {}
        self.paket_harga = paket_harga or 0
        self.paket_label = paket_label or ""
        self.order_qty = {}
         
        self._build()
        center_window(self, master, width=380, height=460)
        self.after(50, self.grab_set)
     
    def _build(self):
        # Header
        hdr = ctk.CTkFrame(self, fg_color="transparent")
        hdr.pack(fill="x", padx=12, pady=(12, 8))
        ctk.CTkLabel(hdr, text=f"🛒 Pesanan Tambahan — {self.tv_label}",
                     font=("Russo One", 13, "bold"), text_color=C_ACCENT).pack(anchor="w")
        ctk.CTkLabel(hdr, text="Pilih item untuk ditambahkan atau perbarui jumlah",
                     font=FONT_BODY, text_color=C_MUTED).pack(anchor="w")
        
        # Total display
        self.lbl_total = ctk.CTkLabel(self, text="Total Pesanan: Rp 0",
                                       font=("Russo One", 12, "bold"), text_color=C_YELLOW)
        self.lbl_total.pack(pady=(0, 2))
        self.lbl_paket_info = ctk.CTkLabel(self, text="",
                                            font=FONT_SMALL, text_color=C_MUTED)
        self.lbl_paket_info.pack(pady=(0, 8))
        
        # Scrollable content
        scroll = ctk.CTkScrollableFrame(self, fg_color=C_BG)
        scroll.pack(fill="both", expand=True, padx=12, pady=0)
        
        # Makanan section
        if self.makanan_data:
            self._build_menu_section(scroll, "🍔  MAKANAN", self.makanan_data)
        
        # Minuman section
        if self.minuman_data:
            self._build_menu_section(scroll, "🥤  MINUMAN", self.minuman_data)
        
        if not self.makanan_data and not self.minuman_data:
            ctk.CTkLabel(scroll, text="Tidak ada item tersedia",
                        font=FONT_BODY, text_color=C_MUTED).pack(pady=20)
        
        # Buttons
        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.pack(fill="x", padx=12, pady=(8, 12))
        
        ctk.CTkButton(btn_frame, text="✅  TAMBAHKAN PESANAN",
                     fg_color=C_ACCENT2, hover_color=C_ACCENT,
                     font=("Russo One", 11, "bold"), height=44,
                     command=self._confirm).pack(fill="x", pady=(0, 6))
        ctk.CTkButton(btn_frame, text="✖  BATAL",
                     fg_color=C_RED, hover_color="#CC0033",
                     font=("Russo One", 10, "bold"), height=38,
                     command=self.destroy).pack(fill="x")

        self._update_total()
    
    def _build_menu_section(self, parent, title, menu_dict):
        """Build collapsible menu section."""
        section = ctk.CTkFrame(parent, fg_color="transparent")
        section.pack(fill="x", pady=6)
        
        # Header
        header = ctk.CTkFrame(section, fg_color=C_PANEL, corner_radius=8)
        header.pack(fill="x")
        ctk.CTkLabel(header, text=title,
                    font=("Russo One", 10, "bold"), text_color=C_ACCENT2).pack(anchor="w", padx=10, pady=8)
        
        # Content
        content = ctk.CTkFrame(section, fg_color=C_CARD, corner_radius=6)
        content.pack(fill="x", padx=4, pady=(0, 4))
        
        for nama, harga in menu_dict.items():
            row = ctk.CTkFrame(content, fg_color="transparent")
            row.pack(fill="x", padx=10, pady=4)
            
            # Item name & price
            ctk.CTkLabel(row, text=f"{nama}  •  {fmt_rp(harga)}",
                        font=FONT_LABEL, text_color=C_TEXT, anchor="w").pack(side="left", fill="x", expand=True)
            
            # Qty control
            var = ctk.IntVar(value=self.pesanan_aktif.get(nama, 0))
            self.order_qty[nama] = var
            
            ctk.CTkButton(row, text="−", width=24, height=24, fg_color=C_BTN, hover_color=C_RED,
                         font=("Consolas", 11, "bold"),
                         command=lambda v=var: (v.set(max(0, v.get()-1)), self._update_total())
                         ).pack(side="left", padx=2)
            ctk.CTkLabel(row, textvariable=var, width=24,
                        font=FONT_LABEL, text_color=C_ACCENT).pack(side="left")
            ctk.CTkButton(row, text="+", width=24, height=24, fg_color=C_BTN, hover_color=C_GREEN,
                         font=("Consolas", 11, "bold"),
                         command=lambda v=var: (v.set(v.get()+1), self._update_total())
                         ).pack(side="left", padx=2)
    
    def _update_total(self):
        """Update total pesanan (termasuk qty yang sudah ada)."""
        all_menu = {**self.makanan_data, **self.minuman_data}
        total = sum(all_menu.get(nm, 0) * v.get() for nm, v in self.order_qty.items())
        if self.paket_harga > 0 or self.paket_label:
            total_sesi = self.paket_harga + total
            self.lbl_total.configure(text=f"Total Sesi: {fmt_rp(total_sesi)}")
            self.lbl_paket_info.configure(
                text=f"Paket: {self.paket_label} ({fmt_rp(self.paket_harga)}) + Pesanan {fmt_rp(total)}")
        else:
            self.lbl_total.configure(text=f"Total Pesanan: {fmt_rp(total)}")
            self.lbl_paket_info.configure(text="")
    
    def _confirm(self):
        """Confirm and return new order data + status pembayaran popup."""
        pesanan_baru = {nm: v.get() for nm, v in self.order_qty.items() if v.get() > 0}
        if not pesanan_baru:
            messagebox.showwarning("Pesanan Kosong", "Pilih minimal satu item pesanan.", parent=self)
            return
        total = sum((self.makanan_data | self.minuman_data).get(nm, 0) * q
                    for nm, q in pesanan_baru.items())
        DialogKonfirmasiBayar(
            self,
            lambda paid, p=pesanan_baru: self._finish(p, paid),
            judul="Tambah Pesanan — Status",
            rincian=f"{', '.join(f'{q}x {n}' for n, q in pesanan_baru.items())}\n{fmt_rp(total)}",
        ).lift()

    def _finish(self, pesanan_baru, paid):
        try:
            self.on_confirm(pesanan_baru, paid)
        finally:
            self.destroy()


# ═══════════════════════════════════════════════════════════════════════════════
#  DIALOG PAKET + PESANAN
# ═══════════════════════════════════════════════════════════════════════════════
class DialogPaket(ctk.CTkToplevel):
    def __init__(self, master, tv_label, on_confirm, paket_data, makanan_data, minuman_data, nama_grup="Reguler", for_warnet=False):
        self.for_warnet = for_warnet
        super().__init__(master)
        self.app = master
        self.title(f"Paket & Pesanan — {tv_label}")
        self.geometry("460x620")  # Lebih kecil & compact
        self.configure(fg_color=C_BG)
        self.transient(master)
         
        self.on_confirm   = on_confirm
        self.tv_label     = tv_label
        self.paket_data   = paket_data or {"Paket Default": {"harga": 50000, "menit": 60}}
        self.makanan_data = makanan_data or {}
        self.minuman_data = minuman_data or {}
        self.nama_grup    = nama_grup
        self.paket_var    = ctk.StringVar(value=list(self.paket_data.keys())[0])
        self.pesanan_qty  = {}
        self.diskon_var   = ctk.StringVar(value="0")
        self.diskon_mode_var = ctk.StringVar(value="nominal")
        
        # State untuk collapse/expand
        self.expanded_groups = {"paket": True, "makanan": False, "minuman": False}
        
        self._build()
        center_window(self, master, width=460, height=620)
        self.after(50, self.grab_set)
    
    def _build(self):
        # Header
        hdr = ctk.CTkFrame(self, fg_color="transparent")
        hdr.pack(fill="x", padx=12, pady=(12, 2))
        ctk.CTkLabel(hdr, text=f"📺  {self.tv_label}",
                     font=("Russo One", 14, "bold"), text_color=C_ACCENT).pack(anchor="w")
        ctk.CTkLabel(hdr, text=f"🏷 {self.nama_grup}  •  Total: ",
                     font=("Courier New", 9), text_color=C_MUTED).pack(anchor="w")
        
        # Total display inline
        self.lbl_total = ctk.CTkLabel(self, text="Rp 0",
                                       font=("Russo One", 13, "bold"), text_color=C_YELLOW)
        self.lbl_total.pack(pady=(0, 8))
        
        # Scrollable content - HANYA PAKET (makanan/minuman di tab TV)
        scroll = ctk.CTkScrollableFrame(self, fg_color=C_BG)
        scroll.pack(fill="both", expand=True, padx=12, pady=0)
        
        # Grup 1: PAKET SAJA
        paket_title = "⏱  PAKET WARNET" if getattr(self, 'for_warnet', False) else "⏱  PAKET RENTAL PS"
        self._build_collapsible_group(
            scroll, "paket", paket_title,
            self._build_paket_content
        )

        # Grup booking online (hanya kartu TV — warnet tidak memakai)
        if not getattr(self, 'for_warnet', False):
            self._build_booking_section(scroll)

        # DISCOUNT section
        diskon_frame = ctk.CTkFrame(self, fg_color="transparent")
        diskon_frame.pack(fill="x", padx=12, pady=(4, 2))
        diskon_row = ctk.CTkFrame(diskon_frame, fg_color="transparent")
        diskon_row.pack(fill="x")
        ctk.CTkLabel(diskon_row, text="Diskon:", font=("Russo One", 10, "bold"), text_color=C_ACCENT2).pack(side="left", padx=(0, 6))
        self.diskon_entry = ctk.CTkEntry(diskon_row, textvariable=self.diskon_var, width=80,
                                          font=("Courier New", 12, "bold"), fg_color=C_BTN, text_color=C_TEXT)
        self.diskon_entry.pack(side="left", padx=(0, 8))
        ctk.CTkRadioButton(diskon_row, text="Nominal (Rp)", variable=self.diskon_mode_var, value="nominal",
                           font=("Courier New", 9), fg_color=C_ACCENT, hover_color=C_ACCENT2,
                           command=self._update_total).pack(side="left", padx=(0, 4))
        ctk.CTkRadioButton(diskon_row, text="Persen (%)", variable=self.diskon_mode_var, value="persen",
                           font=("Courier New", 9), fg_color=C_ACCENT, hover_color=C_ACCENT2,
                           command=self._update_total).pack(side="left")

        # STATUS PEMBAYARAN — dipilih lewat DialogKonfirmasiBayar setelah
        # user menekan MULAI SESI (bukan radio di sini).

        # BUTTONS — paling bawah
        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.pack(fill="x", padx=12, pady=(8, 12))
        
        btn_mulai = ctk.CTkButton(btn_frame, text="✅  MULAI SESI",
                      fg_color=C_ACCENT2, hover_color=C_ACCENT,
                      font=("Russo One", 11, "bold"), height=44,
                      command=self._handle_mulai_sesi)
        btn_mulai.pack(fill="x", pady=(0, 6))
        
        ctk.CTkButton(btn_frame, text="✖  BATAL",
                      fg_color=C_RED, hover_color="#CC0033",
                      font=("Russo One", 10, "bold"), height=38,
                      command=self.destroy).pack(fill="x")
        
        self._update_total()
    
    def _build_collapsible_group(self, parent, group_id, title, content_builder):
        """Buat collapsible group dengan header dan content."""
        group_container = ctk.CTkFrame(parent, fg_color="transparent")
        group_container.pack(fill="x", pady=4)
        
        # Header dengan toggle
        header = ctk.CTkFrame(group_container, fg_color=C_PANEL, corner_radius=8)
        header.pack(fill="x")
        header.bind("<Button-1>", lambda e: self._toggle_group(group_id, content_frame, header_btn))
        
        header_btn = ctk.CTkButton(
            header, text=f"{'▼' if self.expanded_groups[group_id] else '▶'}  {title}",
            fg_color="transparent", hover_color=C_CARD,
            font=("Russo One", 10, "bold"), text_color=C_ACCENT2,
            anchor="w", width=450
        )
        header_btn.pack(fill="x", padx=10, pady=8)
        header_btn.bind("<Button-1>", lambda e: self._toggle_group(group_id, content_frame, header_btn))
        
        # Content frame
        content_frame = ctk.CTkFrame(group_container, fg_color="transparent")
        if self.expanded_groups[group_id]:
            content_frame.pack(fill="x", padx=2, pady=(0, 4))
            content_builder(content_frame)
        
        # Store reference untuk toggle
        group_container.content_frame = content_frame
        group_container.header_btn = header_btn
    
    def _toggle_group(self, group_id, content_frame, header_btn):
        """Toggle expand/collapse untuk group."""
        self.expanded_groups[group_id] = not self.expanded_groups[group_id]
        
        if self.expanded_groups[group_id]:
            # Expand
            content_frame.pack(fill="x", padx=2, pady=(0, 4))
            header_btn.configure(text=header_btn.cget("text").replace("▶", "▼"))
        else:
            # Collapse
            content_frame.pack_forget()
            btn_text = header_btn.cget("text").replace("▼", "▶")
            header_btn.configure(text=btn_text)
    
    def _build_paket_content(self, parent):
        """Build paket options inside collapsible group."""
        cf = ctk.CTkFrame(parent, fg_color=C_CARD, corner_radius=6)
        cf.pack(fill="x", padx=8, pady=4)
        
        for nama, info in self.paket_data.items():
            harga = info.get("harga", 0)
            menit = info.get("menit", 0)
            
            row = ctk.CTkFrame(cf, fg_color="transparent")
            row.pack(fill="x", padx=10, pady=3)
            
            color = C_GREEN if nama == "Main Bebas" else C_ORANGE if nama == "Reguler" else C_TEXT
            rb = ctk.CTkRadioButton(row, text=nama, variable=self.paket_var, value=nama,
                                     font=FONT_LABEL, text_color=color,
                                     fg_color=C_ACCENT, hover_color=C_ACCENT2,
                                     command=self._update_total)
            rb.pack(side="left")
            
            if nama == "Main Bebas":
                tarif_menit = hitung_tarif_per_menit(self.paket_data)
                harga_txt = f"≈ {fmt_rp(tarif_menit)}/menit"
                durasi_txt = "Bebas"
            else:
                harga_txt = "Sesuai Durasi" if harga == 0 and menit == 0 else fmt_rp(harga)
                durasi_txt = fmt_durasi(menit) if menit > 0 else "—"
            
                ctk.CTkLabel(row, text=f"{harga_txt} • {durasi_txt}", font=FONT_LABEL,
                        text_color=C_MUTED).pack(side="right")

    # ── Booking online (kartu TV saja) ─────────────────────────────────────
    def _build_booking_section(self, parent):
        """Tombol daftar PAKET BOOKING + area hasil (collapse toggle)."""
        self._booking_frame_shown = False
        self._booking_btn = ctk.CTkButton(
            parent, text="📅  PAKET BOOKING  (klik untuk buka)",
            fg_color=C_PANEL, hover_color=C_CARD, height=36,
            font=("Russo One", 10, "bold"), text_color=C_ORANGE,
            command=self._toggle_booking)
        self._booking_btn.pack(fill="x", padx=8, pady=(2, 0))
        self._booking_frame = ctk.CTkFrame(parent, fg_color="transparent")
        self._booking_rows = []

    def _toggle_booking(self):
        if self._booking_frame_shown and self._booking_frame.winfo_manager():
            self._booking_frame.pack_forget()
            self._booking_frame_shown = False
            self._booking_btn.configure(text="📅  PAKET BOOKING  (klik untuk buka)")
            self._booking_rows = []
            return
        try:
            kartu = self.app._qr_cari_kartu(self.tv_label)
        except Exception:
            kartu = None
        if kartu is not None and not kartu.sesi_kosong():
            messagebox.showwarning(
                "Kartu Dipakai",
                f"Kartu '{self.tv_label}' sedang dipakai sesi lain.\n"
                "Selesaikan sesi dulu sebelum memakai booking online.",
                parent=self)
            return
        self._booking_btn.configure(text="⏳  Memuat booking...")
        try:
            self.app._booking_fetch_valid(self.tv_label, self._render_booking)
        except Exception:
            self._booking_btn.configure(text="📅  PAKET BOOKING  (klik untuk buka)")

    def _render_booking(self, rows):
        for w in self._booking_frame.winfo_children():
            w.destroy()
        if not rows:
            ctk.CTkLabel(self._booking_frame,
                         text=f"Tidak ada booking aktif untuk {self.tv_label}.",
                         font=FONT_SMALL, text_color=C_MUTED).pack(fill="x", padx=14, pady=8)
        else:
            for b in rows:
                try:
                    self._build_booking_row(b)
                except Exception:
                    continue
        self._booking_frame.pack(fill="x", padx=8, pady=(0, 4))
        self._booking_frame_shown = True
        self._booking_btn.configure(text="📅  TUTUP DAFTAR BOOKING")

    def _build_booking_row(self, b):
        did = str(b.get("_id", ""))
        status_bayar = str(b.get("statusBayar", "") or "")
        sb_txt, sb_col = self.app._booking_status_bayar(b)
        row = ctk.CTkFrame(self._booking_frame, fg_color=C_CARD, corner_radius=6)
        row.pack(fill="x", padx=6, pady=2)
        info = f"{did[:8].upper()} • {b.get('namaPelanggan', '-')} • {b.get('jam', '')[:5]}"
        ctk.CTkLabel(row, text=info, font=("Consolas", 9, "bold"),
                     text_color=C_TEXT, anchor="w").pack(fill="x", padx=8, pady=(4, 0))
        sub = (f"{b.get('paket', '')}  •  {fmt_rp(int(b.get('totalHarga', 0) or 0))}"
               + (f"  ({sb_txt})" if sb_txt else ""))
        ctk.CTkLabel(row, text=sub, font=FONT_SMALL,
                     text_color=sb_col, anchor="w").pack(fill="x", padx=8, pady=(0, 3))
        row.bind("<Button-1>", lambda e, bb=dict(b): self._mulai_booking(bb))
        for w in row.winfo_children():
            w.bind("<Button-1>", lambda e, bb=dict(b): self._mulai_booking(bb))

    def _mulai_booking(self, b):
        try:
            kartu = self.app._qr_cari_kartu(self.tv_label)
        except Exception:
            kartu = None
        if kartu is not None and not kartu.sesi_kosong():
            messagebox.showwarning(
                "Kartu Dipakai",
                f"Kartu '{self.tv_label}' sudah dipakai sesi lain.",
                parent=self)
            return
        grupk = str(b.get("grup", "") or "")
        paket = str(b.get("paket", "") or "")
        pmap = self.app.get_paket_data(grupk) if grupk else {}
        if not isinstance(pmap, dict) or paket not in pmap:
            messagebox.showerror("Paket Tidak Dikenal",
                                 f"Paket '{paket}' tidak ada di grup '{grupk or '-'}'.",
                                 parent=self)
            return
        info = pmap[paket]
        harga = int((info.get("harga", 0) if isinstance(info, dict) else info) or 0)
        menit = int((info.get("menit", 60) if isinstance(info, dict) else 60) or 60)
        pesanan = {str(k): int(v or 0) for k, v in (b.get("pesanan") or {}).items()
                   if int(v or 0) > 0}
        all_menu = {**dict(self.app.menu_makanan), **dict(self.app.menu_minuman)}
        total_pesanan = sum(all_menu.get(nm, 0) * qty for nm, qty in pesanan.items())
        metode = str(b.get("metode", "") or "")
        sb = str(b.get("statusBayar", "") or "")
        did = str(b.get("_id", ""))
        try:
            sb_info = self.app._booking_status_bayar(b)[0]
        except Exception:
            sb_info = ""
        if not messagebox.askyesno(
                "Mulai dari Booking",
                f"Mulai sesi booking {did[:8].upper()}?\n"
                f"{b.get('namaPelanggan', '-')} • {paket} • {fmt_rp(int(b.get('totalHarga', 0) or 0))}\n"
                f"Metode: {metode or 'tempat'}"
                + (f"\nStatus: {sb_info}" if sb_info else ""),
                parent=self):
            return
        if metode == "lunas" or sb == "lunas_transfer":
            self.on_confirm(paket, harga, menit, pesanan, total_pesanan,
                            0, "nominal", True, booking=b)
            self.destroy()
        elif metode == "dp":
            self.on_confirm(paket, harga, menit, pesanan, total_pesanan,
                            0, "nominal", False, booking=b)
            self.destroy()
        else:
            # Belum bayar (bayar di tempat) — langsung mulai dengan status
            # TAGIHAN, TANPA popup konfirmasi pembayaran. Konfirmasi bayar
            # hanya muncul saat sesi selesai (waktu habis / tombol Selesai).
            self.on_confirm(paket, harga, menit, pesanan, total_pesanan,
                            0, "nominal", False, booking=b)
            self.destroy()

    def _build_menu_content(self, parent, menu_dict):
        """Build menu items (makanan/minuman) inside collapsible group."""
        cf = ctk.CTkFrame(parent, fg_color=C_CARD, corner_radius=6)
        cf.pack(fill="x", padx=8, pady=4)
        
        if not menu_dict:
            ctk.CTkLabel(cf, text="(Tidak ada item)", font=FONT_SMALL, text_color=C_MUTED).pack(pady=8)
            return
        
        for nama, harga in menu_dict.items():
            row = ctk.CTkFrame(cf, fg_color="transparent")
            row.pack(fill="x", padx=10, pady=3)
            
            ctk.CTkLabel(row, text=f"{nama}  •  {fmt_rp(harga)}",
                        font=FONT_LABEL, text_color=C_TEXT, anchor="w").pack(side="left", fill="x", expand=True)
            
            var = ctk.IntVar(value=0)
            self.pesanan_qty[nama] = var
            
            ctk.CTkButton(row, text="−", width=24, height=24, fg_color=C_BTN, hover_color=C_RED,
                         font=("Consolas", 11, "bold"),
                         command=lambda v=var: (v.set(max(0, v.get()-1)), self._update_total())
                         ).pack(side="left", padx=2)
            ctk.CTkLabel(row, textvariable=var, width=24,
                        font=FONT_LABEL, text_color=C_ACCENT).pack(side="left")
            ctk.CTkButton(row, text="+", width=24, height=24, fg_color=C_BTN, hover_color=C_GREEN,
                         font=("Consolas", 11, "bold"),
                         command=lambda v=var: (v.set(v.get()+1), self._update_total())
                         ).pack(side="left", padx=2)

    def _get_diskon_amount(self, subtotal):
        try:
            val = int(self.diskon_var.get())
        except ValueError:
            return 0
        if val <= 0:
            return 0
        if self.diskon_mode_var.get() == "persen":
            return subtotal * val // 100
        return val

    def _update_total(self):
        """Update total calculation with discount."""
        info = self.paket_data.get(self.paket_var.get(), {})
        harga_paket = info.get("harga", 0)
        menit_paket = info.get("menit", 0)
        all_menu = {**self.makanan_data, **self.minuman_data}
        total_pesanan = sum(all_menu.get(nm, 0) * v.get() for nm, v in self.pesanan_qty.items())
        
        nm = self.paket_var.get()
        if nm == "Main Bebas":
            tarif_menit = hitung_tarif_per_menit(self.paket_data)
            total_txt = f"≈ {fmt_rp(tarif_menit)}/menit"
            if total_pesanan > 0:
                total_txt += f" + {fmt_rp(total_pesanan)}"
        else:
            subtotal = harga_paket + total_pesanan
            diskon_val = self._get_diskon_amount(subtotal)
            total = subtotal - diskon_val
            if diskon_val > 0:
                total_txt = f"{fmt_rp(subtotal)} - {fmt_rp(diskon_val)} = {fmt_rp(total)}"
            else:
                total_txt = fmt_rp(total)
        
        self.lbl_total.configure(text=total_txt)

    def _on_mulai_sesi(self):
        """Handle MULAI SESI button click - simplified."""
        print(f"[DEBUG] _on_mulai_sesi() called")
        try:
            paket_nm = self.paket_var.get()
            print(f"[DEBUG] Paket selected: {paket_nm}")
            info = self.paket_data.get(paket_nm, {})
            paket_harga = info.get("harga", 0)
            paket_menit = info.get("menit", 0)
            pesanan = {}  # Kosong - makanan/minuman di tab TV

            print(f"[DEBUG] Asking payment status for: {paket_nm}")
            DialogKonfirmasiBayar(
                self, 
                lambda paid: self._lanjutkan_konfirmasi(paket_nm, paket_harga, paket_menit, pesanan, 0, 0, "nominal", paid),
                judul="Mulai Sesi — Pilih Status",
                rincian=f"{paket_nm} • {fmt_rp(paket_harga)}",
            ).lift()
        except Exception as e:
            print(f"[ERROR] Exception in _on_mulai_sesi: {e}")
            import traceback
            traceback.print_exc()

    def _handle_mulai_sesi(self):
        """Actual handler - called via button command."""
        if messagebox.askyesno("Konfirmasi", "Mulai sesi sekarang?", parent=self):
            self._confirm()

    def _confirm(self):
        """Confirm with payment status popup."""
        try:
            paket_nm    = self.paket_var.get()
            info        = self.paket_data.get(paket_nm, {})
            paket_harga = info.get("harga", 0)
            paket_menit = info.get("menit", 0)
            all_menu    = {**self.makanan_data, **self.minuman_data}
            pesanan     = {nm: v.get() for nm, v in self.pesanan_qty.items() if v.get() > 0}

            # Hitung total pesanan (makanan/minuman) — paket_harga dihitung terpisah
            total_pesanan = sum(all_menu.get(nm, 0) * qty for nm, qty in pesanan.items())

            try:
                diskoni = int(self.diskon_var.get())
            except ValueError:
                diskoni = 0
            diskoni_mode = self.diskon_mode_var.get()

            sub_total = paket_harga + total_pesanan
            DialogKonfirmasiBayar(
                self,
                lambda paid, pp=paket_nm, ph=paket_harga, pm=paket_menit,
                       p=pesanan, tp=total_pesanan, d=diskoni, dm=diskoni_mode:
                    self._lanjutkan_konfirmasi(pp, ph, pm, p, tp, d, dm, paid),
                judul="Konfirmasi Pembayaran",
                rincian=f"{paket_nm} • {fmt_rp(paket_harga)}" +
                        (f" + Pesanan {fmt_rp(total_pesanan)}" if total_pesanan else ""),
            ).lift()
        except Exception as e:
            messagebox.showerror("Error", f"Terjadi kesalahan: {str(e)}")

    def _lanjutkan_konfirmasi(self, paket_nm, paket_harga, paket_menit, pesanan,
                              total_pesanan, diskoni, diskoni_mode, paid, booking=None):
        """Setelah user memilih BAYAR / TAGIHAN, teruskan ke on_confirm."""
        self.on_confirm(paket_nm, paket_harga, paket_menit, pesanan,
                        total_pesanan, diskoni, diskoni_mode, paid, booking)
        print(f"[DEBUG] on_confirm dipanggil dengan paid={paid}")
        self.destroy()


def hitung_tarif_per_menit(paket_data):
    """
    Tarif per menit untuk Main Bebas, diturunkan dari harga & durasi
    paket acuan (default: '1 Jam'). Kalau paket acuan tidak ada / durasinya 0,
    fallback ke paket non-Main Bebas / non-Reguler pertama yang punya durasi > 0.
    """
    acuan = paket_data.get(PAKET_ACUAN_BEBAS)
    if acuan and acuan.get("menit", 0) > 0:
        return acuan["harga"] / acuan["menit"]
    for nama, info in paket_data.items():
        if nama in ("Main Bebas",) :
            continue
        if info.get("menit", 0) > 0:
            return info["harga"] / info["menit"]
    return 0


# ═══════════════════════════════════════════════════════════════════════════════
#  DIALOG TAMBAH WARNET
# ═══════════════════════════════════════════════════════════════════════════════
class DialogTambahWarnet(ctk.CTkToplevel):
    def __init__(self, master, on_confirm, on_close_cb, daftar_grup=None, lock_group=False, pc_options=None):
        super().__init__(master)
        self.title("Tambah Kursi Warnet")
        self.geometry("460x340")
        self.configure(fg_color=C_BG)
        self.transient(master)
        self.resizable(False, False)
        self.on_confirm = on_confirm
        self.on_close_cb = on_close_cb
        self.daftar_grup = daftar_grup or [NAMA_GRUP_DEFAULT]
        self.lock_group = bool(lock_group)
        self.grup_var = ctk.StringVar(value=self.daftar_grup[0])
        self.pc_options = pc_options or []
        self.selected_pc = None
        self._confirmed = False
        self._build()
        center_window(self, master, width=460, height=340)
        self.protocol("WM_DELETE_WINDOW", self._on_close)
        self.after(50, self.grab_set)

    def _build(self):
        ctk.CTkLabel(self, text="➕ Tambah Kursi Warnet", font=FONT_TITLE, text_color=C_ACCENT).pack(pady=(14, 6))

        # PC Client selection
        pc_frame = ctk.CTkFrame(self, fg_color=C_PANEL, corner_radius=10)
        pc_frame.pack(fill="x", padx=18, pady=(0, 6))
        ctk.CTkLabel(pc_frame, text="PC Client:", width=90, anchor="w", font=FONT_LABEL, text_color=C_MUTED).pack(side="left", padx=(10, 0))
        pc_labels = [p["label"] for p in self.pc_options] if self.pc_options else ["Tidak ada PC terhubung"]
        self.pc_var = ctk.StringVar(value=pc_labels[0])
        self.opt_pc = ctk.CTkOptionMenu(
            pc_frame, values=pc_labels, variable=self.pc_var,
            fg_color=C_BTN, button_color=C_ACCENT2, button_hover_color="#5A0FCC",
            text_color=C_TEXT, font=FONT_BODY, dropdown_font=FONT_BODY,
            dropdown_fg_color=C_CARD, dropdown_text_color=C_TEXT,
        )
        self.opt_pc.pack(side="left", fill="x", expand=True, padx=(6, 10))

        # Nama Kursi
        row = ctk.CTkFrame(self, fg_color=C_PANEL, corner_radius=10)
        row.pack(fill="x", padx=18, pady=(0, 6))
        ctk.CTkLabel(row, text="Nama:", width=90, anchor="w", font=FONT_LABEL, text_color=C_MUTED).pack(side="left", padx=(10, 0))
        self.entry_nama = ctk.CTkEntry(row, fg_color=C_BTN, text_color=C_ACCENT,
                                      border_color=C_BORDER, font=("Consolas", 12, "bold"), height=34)
        self.entry_nama.pack(side="left", fill="x", expand=True, padx=(6, 10))

        # Auto-fill nama from selected PC
        if self.pc_options:
            first = self.pc_options[0]
            if first.get("name"):
                self.entry_nama.insert(0, first["name"])

        self.opt_pc.configure(command=self._on_pc_selected)

        # Grup Tarif
        grp_row = ctk.CTkFrame(self, fg_color=C_PANEL, corner_radius=10)
        grp_row.pack(fill="x", padx=18, pady=(0, 6))
        ctk.CTkLabel(grp_row, text="Grup:", width=90, anchor="w", font=FONT_LABEL, text_color=C_MUTED).pack(side="left", padx=(10, 0))
        self.opt_grup = ctk.CTkOptionMenu(
            grp_row, values=self.daftar_grup, variable=self.grup_var,
            fg_color=C_BTN, button_color=C_ACCENT2, button_hover_color="#5A0FCC",
            text_color=C_TEXT, font=FONT_BODY, dropdown_font=FONT_BODY,
            dropdown_fg_color=C_CARD, dropdown_text_color=C_TEXT,
        )
        self.opt_grup.pack(side="left", fill="x", expand=True, padx=(6, 10))
        if getattr(self, 'lock_group', False):
            try:
                self.opt_grup.configure(state="disabled")
            except Exception:
                try:
                    self.opt_grup.pack_forget()
                    lbl = ctk.CTkLabel(grp_row, textvariable=self.grup_var, font=FONT_BODY, text_color=C_TEXT)
                    lbl.pack(side="left", fill="x", expand=True, padx=(6,10))
                except Exception:
                    pass

        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.pack(fill="x", padx=18, pady=(8, 14))
        ctk.CTkButton(btn_frame, text="✅ Tambah Kursi", fg_color=C_ACCENT2,
                      hover_color=C_ACCENT, font=FONT_SUB, text_color="white",
                      command=self._on_confirm).pack(side="left", fill="x", expand=True, padx=(0, 6), pady=0, ipady=6)
        ctk.CTkButton(btn_frame, text="✖ Batal", fg_color=C_RED,
                      hover_color="#FF5C5C", font=FONT_SUB, text_color="white",
                      command=self._on_close).pack(side="left", fill="x", expand=True, padx=(6, 0), pady=0, ipady=6)

    def _on_pc_selected(self, choice):
        for p in self.pc_options:
            if p["label"] == choice:
                current = self.entry_nama.get().strip()
                # Only auto-fill if field is empty or matches previous name
                if not current or current == self.pc_options[0].get("name", ""):
                    self.entry_nama.delete(0, "end")
                    self.entry_nama.insert(0, p.get("name", ""))
                break

    def _on_confirm(self):
        if not self.pc_options:
            messagebox.showwarning("Tidak Ada PC", "Tidak ada PC client yang terhubung.", parent=self)
            return
        nama = self.entry_nama.get().strip()
        grup = self.grup_var.get().strip() if self.grup_var.get() else ""
        if not nama:
            messagebox.showwarning("Input Salah", "Nama kursi wajib diisi.", parent=self)
            return
        if not grup:
            messagebox.showwarning("Input Salah", "Pilih grup tarif untuk kursi warnet.", parent=self)
            return
        # Find selected PC
        choice = self.pc_var.get()
        selected = None
        for p in self.pc_options:
            if p["label"] == choice:
                selected = p
                break
        if not selected:
            messagebox.showwarning("Pilih PC", "Pilih PC client yang akan ditambahkan.", parent=self)
            return
        self._confirmed = True
        self.on_confirm(nama, grup, selected)
        self.destroy()

    def _on_close(self):
        if not self._confirmed and self.on_close_cb:
            self.on_close_cb()
        self.destroy()


# ═══════════════════════════════════════════════════════════════════════════════
#  VIRTUAL REMOTE DIALOG
# ═══════════════════════════════════════════════════════════════════════════════

class VirtualRemoteDialog(ctk.CTkToplevel):
    def __init__(self, master, label_tv, ip, port):
        super().__init__(master, fg_color=C_BG)
        self.title(f"🕹️ Remote — {label_tv}")
        self.geometry("380x680")
        self.resizable(False, False)
        self.grab_set()
        self.ip = ip
        self.port = port
        self.label_tv = label_tv

        self._build()

    def _build(self):
        title = ctk.CTkLabel(self, text=f"🕹️  {self.label_tv}", font=("Russo One", 14, "bold"), text_color=C_ACCENT)
        title.pack(pady=(16, 6))

        sep = ctk.CTkFrame(self, height=2, fg_color=C_ACCENT2)
        sep.pack(fill="x", padx=30, pady=(0, 12))

        main = ctk.CTkFrame(self, fg_color="transparent")
        main.pack(expand=True, fill="both", padx=20)

        # ── Row 1: Power & System ──
        r1 = ctk.CTkFrame(main, fg_color="transparent")
        r1.pack(pady=(0, 10))
        for txt, key in [("⚡ POWER", "POWER"), ("📺 TV PWR", "TV_POWER"), ("⚙ SETTINGS", "SETTINGS")]:
            ctk.CTkButton(r1, text=txt, width=100, height=38,
                          fg_color=C_RED if "POWER" in txt else C_BTN,
                          hover_color="#FF6666" if "POWER" in txt else C_ACCENT2,
                          font=("Russo One", 9, "bold"), text_color="white",
                          border_width=1, border_color=C_RED if "POWER" in txt else C_ACCENT2,
                          command=lambda k=key: self._send(k)).pack(side="left", padx=4)

        # ── Row 2: D-Pad ──
        dpad_frame = ctk.CTkFrame(main, fg_color="transparent")
        dpad_frame.pack(pady=(0, 10))

        dpad_grid = ctk.CTkFrame(dpad_frame, fg_color=C_CARD, corner_radius=8)
        dpad_grid.pack()

        # UP
        up_btn = ctk.CTkButton(dpad_grid, text="▲", width=60, height=36,
                                fg_color=C_BTN, hover_color=C_ACCENT2,
                                font=("Russo One", 14, "bold"), text_color="white",
                                command=lambda: self._send("DPAD_UP"))
        up_btn.grid(row=0, column=1, padx=2, pady=2)

        # LEFT, OK, RIGHT
        for col, (txt, key) in enumerate([("◀", "DPAD_LEFT"), ("● OK", "DPAD_CENTER"), ("▶", "DPAD_RIGHT")], start=0):
            ctk.CTkButton(dpad_grid, text=txt, width=60, height=36,
                          fg_color=C_ACCENT2 if "OK" in txt else C_BTN,
                          hover_color=C_ACCENT if "OK" in txt else C_ACCENT2,
                          font=("Russo One", 10, "bold") if "OK" in txt else ("Russo One", 14, "bold"),
                          text_color="white",
                          command=lambda k=key: self._send(k)).grid(row=1, column=col, padx=2, pady=2)

        # DOWN
        down_btn = ctk.CTkButton(dpad_grid, text="▼", width=60, height=36,
                                  fg_color=C_BTN, hover_color=C_ACCENT2,
                                  font=("Russo One", 14, "bold"), text_color="white",
                                  command=lambda: self._send("DPAD_DOWN"))
        down_btn.grid(row=2, column=1, padx=2, pady=2)

        # ── Row 3: Volume & Media ──
        r3 = ctk.CTkFrame(main, fg_color="transparent")
        r3.pack(pady=(0, 10))
        for txt, key in [("🔊 VOL+", "VOLUME_UP"), ("🔉 VOL−", "VOLUME_DOWN"), ("🔇 MUTE", "MUTE")]:
            ctk.CTkButton(r3, text=txt, width=100, height=38,
                          fg_color=C_GREEN if "VOL+" in txt else (C_YELLOW if "VOL−" in txt else C_BTN),
                          hover_color=C_ACCENT2,
                          font=("Russo One", 9, "bold"), text_color="black" if "VOL+" in txt else "white",
                          border_width=1,
                          border_color=C_GREEN if "VOL+" in txt else (C_YELLOW if "VOL−" in txt else C_ACCENT2),
                          command=lambda k=key: self._send(k)).pack(side="left", padx=4)

        # ── Row 4: Navigation + Media ──
        r4 = ctk.CTkFrame(main, fg_color="transparent")
        r4.pack(pady=(0, 10))
        for txt, key in [("🏠 HOME", "HOME"), ("◀ BACK", "BACK"), ("▶⏸ PLAY", "MEDIA_PLAY_PAUSE")]:
            ctk.CTkButton(r4, text=txt, width=100, height=38,
                          fg_color=C_ACCENT, hover_color=C_ACCENT2,
                          font=("Russo One", 9, "bold"), text_color="black" if "HOME" in txt else "white",
                          border_width=1, border_color=C_ACCENT,
                          command=lambda k=key: self._send(k)).pack(side="left", padx=4)

        # ── Row 5: Sleep/Wake ──
        r5 = ctk.CTkFrame(main, fg_color="transparent")
        r5.pack(pady=(0, 10))
        for txt, key in [("💤 SLEEP", "SLEEP"), ("☀ WAKEUP", "WAKEUP")]:
            ctk.CTkButton(r5, text=txt, width=140, height=38,
                          fg_color=C_BTN, hover_color=C_ACCENT2,
                          font=("Russo One", 9, "bold"), text_color="white",
                          border_width=1, border_color=C_ACCENT2,
                          command=lambda k=key: self._send(k)).pack(side="left", padx=4)

        # ── Row 6: TV Client Features ──
        r6 = ctk.CTkFrame(main, fg_color="transparent")
        r6.pack(pady=(0, 10))
        for txt, cmd in [
            ("📺 Kirim Video", self._send_video),
            ("💾 Cache Media", self._cache_media),
            ("⚠ Kirim Peringatan", self._send_warning),
            ("🖼 Ganti Logo Lock", self._ganti_logo_lock),
        ]:
            ctk.CTkButton(r6, text=txt, width=100, height=38,
                          fg_color=C_BTN, hover_color=C_ACCENT2,
                          font=("Russo One", 9, "bold"), text_color="white",
                          border_width=1, border_color=C_ACCENT2,
                          command=cmd).pack(side="left", padx=4)

        # ── Close button ──
        sep2 = ctk.CTkFrame(self, height=1, fg_color=C_BORDER)
        sep2.pack(fill="x", padx=30, pady=(0, 10))

        ctk.CTkButton(self, text="✕ TUTUP", height=40,
                      fg_color=C_RED, hover_color="#FF5C5C",
                      font=("Russo One", 11, "bold"), text_color="white",
                      command=self.destroy).pack(padx=30, pady=(0, 16), fill="x")

    def _send(self, key_name):
        threading.Thread(target=self._send_thread, args=(key_name,), daemon=True).start()

    def _send_thread(self, key_name):
        try:
            ok, out, err = ADBHelper.send_key(self.ip, key_name)
        except Exception:
            ok = False
        try:
            self.after(0, lambda: self._flash_feedback(ok))
        except Exception:
            pass

    def _flash_feedback(self, ok):
        try:
            if not self.winfo_exists():
                return
            if ok:
                self.configure(fg_color="#0A2A0A")
                self.after(150, lambda: self.configure(fg_color=C_BG))
        except Exception:
            pass

    # ── TV Client Feature Methods ──────────────────────────────────────────
    def _send_video(self):
        dlg = ctk.CTkToplevel(self)
        dlg.title("📺 Kirim Video ke TV")
        dlg.geometry("400x200")
        dlg.configure(fg_color=C_BG)
        dlg.resizable(False, False)
        dlg.transient(self)
        dlg.grab_set()
        ctk.CTkLabel(dlg, text=f"Masukkan URL Video untuk {self.label_tv}",
                     font=FONT_SUB, text_color=C_ACCENT).pack(pady=(16, 8))
        url_var = tk.StringVar()
        ctk.CTkEntry(dlg, textvariable=url_var, font=("Consolas", 11),
                     fg_color=C_BTN, text_color=C_TEXT).pack(fill="x", padx=20, pady=6)
        lbl_status = ctk.CTkLabel(dlg, text="", font=FONT_SMALL, text_color=C_MUTED)
        lbl_status.pack(pady=4)

        def _do_send():
            url = url_var.get().strip()
            if not url:
                lbl_status.configure(text="URL tidak boleh kosong", text_color=C_RED)
                return
            lbl_status.configure(text="Mengirim video...", text_color=C_YELLOW)
            ok, msg = ADBHelper.send_video_url(self.ip, url)
            lbl_status.configure(text=msg if ok else f"Gagal: {msg}",
                                 text_color=C_GREEN if ok else C_RED)

        ctk.CTkButton(dlg, text="🚀 Kirim", height=36,
                      fg_color=C_ACCENT2, font=FONT_SUB, text_color="white",
                      command=lambda: threading.Thread(target=_do_send, daemon=True).start()
                      ).pack(pady=10, padx=20, fill="x")
        ctk.CTkButton(dlg, text="✖ Batal", height=30,
                      fg_color=C_RED, font=FONT_SMALL, text_color="white",
                      command=dlg.destroy).pack(pady=(0, 10), padx=20, fill="x")

    def _cache_media(self):
        path = filedialog.askopenfilename(
            parent=self, title="Pilih file media",
            filetypes=[("Media", "*.mp4 *.jpg *.jpeg *.png *.gif"), ("All files", "*.*")]
        )
        if not path:
            return
        dlg = ctk.CTkToplevel(self)
        dlg.title("💾 Cache Media ke TV")
        dlg.geometry("400x160")
        dlg.configure(fg_color=C_BG)
        dlg.resizable(False, False)
        dlg.transient(self)
        dlg.grab_set()
        fname = os.path.basename(path)
        ctk.CTkLabel(dlg, text=f"Mengirim {fname} ke {self.label_tv}",
                     font=FONT_SUB, text_color=C_ACCENT).pack(pady=(16, 8))
        lbl = ctk.CTkLabel(dlg, text="Memulai transfer...", font=FONT_BODY, text_color=C_YELLOW)
        lbl.pack(pady=8)
        bar = ctk.CTkProgressBar(dlg, height=6, fg_color=C_BTN, progress_color=C_ACCENT)
        bar.pack(fill="x", padx=20, pady=4)
        bar.set(0)

        def _do_push():
            ok, msg = ADBHelper.adb_push(self.ip, path)
            self.after(0, lambda: bar.set(1))
            self.after(0, lambda: lbl.configure(
                text=f"✅ {fname} dikirim" if ok else f"❌ Gagal: {msg}",
                text_color=C_GREEN if ok else C_RED))

        threading.Thread(target=_do_push, daemon=True).start()
        ctk.CTkButton(dlg, text="✖ Tutup", height=32,
                      fg_color=C_BTN, font=FONT_SMALL, text_color="white",
                      command=dlg.destroy).pack(pady=10, padx=20, fill="x")

    def _send_warning(self):
        dlg = ctk.CTkToplevel(self)
        dlg.title("⚠ Kirim Peringatan ke TV")
        dlg.geometry("420x260")
        dlg.configure(fg_color=C_BG)
        dlg.resizable(False, False)
        dlg.transient(self)
        dlg.grab_set()
        ctk.CTkLabel(dlg, text=f"Banner Peringatan untuk {self.label_tv}",
                     font=FONT_SUB, text_color=C_ACCENT).pack(pady=(14, 6))
        ctk.CTkLabel(dlg, text="Teks peringatan:", font=FONT_LABEL,
                     text_color=C_MUTED).pack(anchor="w", padx=20)
        msg_var = tk.StringVar(value="Waktu akan habis!")
        ctk.CTkEntry(dlg, textvariable=msg_var, font=("Consolas", 11),
                     fg_color=C_BTN, text_color=C_TEXT).pack(fill="x", padx=20, pady=4)
        dur_frame = ctk.CTkFrame(dlg, fg_color="transparent")
        dur_frame.pack(fill="x", padx=20, pady=6)
        ctk.CTkLabel(dur_frame, text="Durasi (detik):", font=FONT_LABEL,
                     text_color=C_MUTED).pack(side="left", padx=(0, 8))
        dur_var = tk.StringVar(value="30")
        ctk.CTkEntry(dur_frame, textvariable=dur_var, width=70,
                      font=("Consolas", 11), fg_color=C_BTN, text_color=C_TEXT).pack(side="left")
        lbl_status = ctk.CTkLabel(dlg, text="", font=FONT_SMALL, text_color=C_MUTED)
        lbl_status.pack(pady=6)

        def _do_warning():
            msg = msg_var.get().strip()
            if not msg:
                lbl_status.configure(text="Teks tidak boleh kosong", text_color=C_RED)
                return
            try:
                dur = int(dur_var.get().strip())
            except ValueError:
                dur = 30
            lbl_status.configure(text="Mengirim peringatan...", text_color=C_YELLOW)
            ok, out = ADBHelper.send_warning_broadcast(self.ip, msg, dur)
            lbl_status.configure(text="⚠ Peringatan terkirim" if ok else f"❌ Gagal: {out}",
                                 text_color=C_GREEN if ok else C_RED)

        ctk.CTkButton(dlg, text="⚠ Kirim Peringatan", height=36,
                      fg_color=C_RED, font=FONT_SUB, text_color="white",
                      command=lambda: threading.Thread(target=_do_warning, daemon=True).start()
                      ).pack(pady=6, padx=20, fill="x")
        ctk.CTkButton(dlg, text="✖ Batal", height=30,
                      fg_color=C_BTN, font=FONT_SMALL, text_color="white",
                      command=dlg.destroy).pack(pady=(0, 10), padx=20, fill="x")

    def _ganti_logo_lock(self):
        """Ganti logo lock screen (global untuk semua kartu TV).

        File disimpan sebagai logo_lock.png di folder media_promo; client TV
        mengambilnya via TvMediaServer saat LOCK_SCREEN (waktu sewa habis).
        Hanya tersedia untuk lisensi LIFETIME."""
        app = self.winfo_toplevel()
        if not app._lisensi_lifetime():
            messagebox.showwarning(
                "Fitur Lisensi LIFETIME",
                "Fitur logo lock TV hanya tersedia untuk lisensi LIFETIME.\n\n"
                "Logo tetap memakai RR BILLING PRO.\n"
                "Lakukan aktivasi lisensi LIFETIME untuk mengganti logo.",
                parent=self)
            return
        ms = getattr(app, 'tv_media_server', None)
        if not ms or not ms.running:
            messagebox.showwarning(
                "Logo Lock",
                "Server media (port 8082) tidak berjalan.\n"
                "Mulai ulang aplikasi billing lalu coba lagi.",
                parent=self)
            return
        path = filedialog.askopenfilename(
            parent=self, title="Pilih logo lock screen (PNG/JPG)",
            filetypes=[("Gambar", "*.png *.jpg *.jpeg"), ("All files", "*.*")]
        )
        if not path:
            return
        try:
            app._simpan_logo_lock(path)
        except Exception as e:
            messagebox.showerror("Logo Lock", f"Gagal menyalin logo:\n{e}", parent=self)
            return
        # Beri tahu semua TV yang terhubung: refresh logo lockscreen sekarang.
        try:
            logo_url = app._tv_logo_url()
            hub = getattr(app, 'tv_ws_hub', None)
            if hub and logo_url:
                n = hub.broadcast_update_logo(logo_url)
                print(f"[TV LOGO] broadcast UPDATE_LOGO ke {n} TV")
        except Exception as e:
            print(f"[TV LOGO] gagal broadcast: {e}")
        messagebox.showinfo(
            "Logo Lock",
            "✅ Logo lock screen berhasil diganti.\n"
            "Berlaku untuk semua TV saat waktu sewa habis.",
            parent=self)


class DialogQrKartu(ctk.CTkToplevel):
    """Preview QR panggil kasir untuk satu TV + simpan PNG / salin link."""

    def __init__(self, master, label_tv, url, path_png):
        super().__init__(master)
        self.title(f"QR Panggil Kasir — {label_tv}")
        self.geometry("420x560")
        self.configure(fg_color=C_BG)
        self.resizable(False, False)
        self.transient(master)
        self._url = url

        ctk.CTkLabel(self, text="📱  QR PANGGIL KASIR",
                     font=("Russo One", 14, "bold"),
                     text_color=C_ACCENT).pack(pady=(16, 2))
        ctk.CTkLabel(self, text=f"{label_tv} — kode unik (QR TV lain tidak berfungsi di TV ini)",
                     font=FONT_SMALL, text_color=C_MUTED).pack(pady=(0, 8))

        qr_img = None
        try:
            from PIL import Image
            if path_png and os.path.isfile(path_png):
                img = Image.open(path_png)
                img.thumbnail((280, 280), Image.LANCZOS)
                qr_img = ctk.CTkImage(light_image=img, dark_image=img, size=img.size)
        except Exception:
            qr_img = None

        if qr_img:
            box = ctk.CTkFrame(self, fg_color="white", corner_radius=12,
                               width=300, height=300)
            box.pack(pady=6)
            box.pack_propagate(False)
            ctk.CTkLabel(box, text="", image=qr_img).place(relx=0.5, rely=0.5,
                                                           anchor="center")
        else:
            ctk.CTkLabel(self, text="(Gagal membuat gambar QR)",
                         font=FONT_BODY, text_color=C_RED).pack(pady=20)

        ctk.CTkLabel(self, text="Pelanggan scan → pilih layanan/keluhan → kasir dapat notif",
                     font=FONT_SMALL, text_color=C_MUTED).pack(pady=(8, 8))

        btn_row = ctk.CTkFrame(self, fg_color="transparent")
        btn_row.pack(fill="x", padx=14, pady=(2, 4))
        ctk.CTkButton(btn_row, text="💾 Simpan PNG", width=120, height=36,
                      fg_color=C_ACCENT2, font=("Russo One", 9, "bold"),
                      text_color="white",
                      command=self._simpan_png).pack(side="left", padx=4)
        ctk.CTkButton(btn_row, text="📋 Salin Link", width=110, height=36,
                      fg_color=C_BTN, font=("Russo One", 9, "bold"),
                      text_color="white",
                      command=self._salin_link).pack(side="left", padx=4)
        ctk.CTkButton(btn_row, text="🌐 Buka", width=80, height=36,
                      fg_color=C_BTN, font=("Russo One", 9, "bold"),
                      text_color="white",
                      command=self._buka_web).pack(side="left", padx=4)
        ctk.CTkButton(self, text="✖ Tutup", height=30,
                      fg_color=C_BTN, font=FONT_SMALL, text_color="white",
                      command=self.destroy).pack(pady=(0, 10), padx=14, fill="x")

        # Perbaiki jenis font Tk yang salah jika perlu (konsistensi)
        self._qr_path = path_png

    def _qrcode_src(self):
        src = self._qr_path
        if src and os.path.isfile(src):
            return src
        import qrcode
        folder = os.path.join(APP_BASE_DIR, "qr_panggilan")
        os.makedirs(folder, exist_ok=True)
        src = os.path.join(folder, "qr_temp.png")
        qrcode.make(self._url).save(src)
        return src

    def _simpan_png(self):
        try:
            src = self._qrcode_src()
            dest = filedialog.asksaveasfilename(
                parent=self, title="Simpan QR sebagai PNG",
                defaultextension=".png", filetypes=[("PNG", "*.png")],
                initialfile="qr_panggil.png")
            if not dest:
                return
            import shutil
            shutil.copyfile(src, dest)
            messagebox.showinfo("✅ QR Tersimpan",
                                f"QR tersimpan di:\n{dest}\n\nCetak lalu "
                                "tempel di dekat TV.", parent=self)
        except Exception as e:
            messagebox.showerror("Simpan QR", f"Gagal: {e}", parent=self)

    def _salin_link(self):
        try:
            self.clipboard_clear()
            self.clipboard_append(self._url)
            self.update()
            messagebox.showinfo("📋 Link Tersalin", self._url, parent=self)
        except Exception as e:
            messagebox.showerror("Salin Link", str(e), parent=self)

    def _buka_web(self):
        import webbrowser
        webbrowser.open(self._url)


# ═══════════════════════════════════════════════════════════════════════════════
#  KARTU TV
# ═══════════════════════════════════════════════════════════════════════════════
class DialogStatusClient(ctk.CTkToplevel):
    """Dialog status APK client TV (terpasang/belum) + install."""
    PACKAGE = "com.rrbillingpro.tvclient"

    def __init__(self, master, label_tv, ip, on_install, role="admin"):
        super().__init__(master)
        self.title(f"Status Client TV — {label_tv}")
        self.geometry("480x360")
        self.configure(fg_color=C_BG)
        self.resizable(False, False)
        self.transient(master)
        self._ip = ip
        self._on_install = on_install
        self._label_tv = label_tv
        self._is_admin = (role or "admin") == "admin"
        # Hub WebSocket untuk fallback: APK bisa aktif walau ADB port 5555 mati.
        self._hub = getattr(master.winfo_toplevel(), 'tv_ws_hub', None)
        self._busy = False

        ctk.CTkLabel(self, text=f"📺  {label_tv}   ({ip})",
                     font=FONT_TITLE, text_color=C_ACCENT).pack(pady=(18, 6))
        self.lbl_adb = ctk.CTkLabel(self, text="ADB: memeriksa…",
                                    font=("Courier New", 12, "bold"),
                                    text_color=C_YELLOW)
        self.lbl_adb.pack(pady=2)
        self.lbl_apk = ctk.CTkLabel(self, text="APK client: memeriksa…",
                                    font=("Courier New", 12, "bold"),
                                    text_color=C_YELLOW)
        self.lbl_apk.pack(pady=2)
        self.lbl_info = ctk.CTkLabel(self, text="",
                                     font=FONT_SMALL, text_color=C_MUTED)
        self.lbl_info.pack(pady=2)
        self.progress = ctk.CTkProgressBar(self, height=8, fg_color=C_BTN,
                                           progress_color=C_ACCENT)
        self.progress.pack(fill="x", padx=18, pady=(4, 2))
        self.progress.set(0)
        self.lbl_progress = ctk.CTkLabel(self, text="",
                                         font=("Courier New", 9),
                                         text_color=C_MUTED)
        self.lbl_progress.pack(pady=(0, 2))

        self.btn_install = ctk.CTkButton(self, text="🔧 INSTALL APK", height=38,
                                         fg_color=C_ACCENT2,
                                         font=("Russo One", 11, "bold"),
                                         command=self._pilih_apk)
        self.btn_install.pack(pady=(18, 4))
        self.btn_install.configure(state="disabled")
        self.btn_upgrade = None
        if self._is_admin:
            self.btn_upgrade = ctk.CTkButton(self, text="⬆ UPGRADE APK (Unduh & Pasang Versi Baru)",
                                             height=38, fg_color=C_ACCENT,
                                             hover_color="#0F766E",
                                             font=("Russo One", 11, "bold"),
                                             command=self._upgrade_apk)
            self.btn_upgrade.pack(pady=(6, 4))
            self.btn_upgrade.configure(state="disabled")
        ctk.CTkButton(self, text="Tutup", height=32, fg_color=C_BTN,
                      text_color=C_MUTED, font=("Russo One", 10, "bold"),
                      command=self.destroy).pack(pady=4)
        self._cek()

    def _cek(self):
        self._busy = True
        threading.Thread(target=self._cek_thread, daemon=True).start()

    def _cek_thread(self):
        hasil = self._cek_apk()
        self.after(0, self._tampilkan, hasil)

    def _cek_apk(self):
        konek = ADBHelper._adb_connect(self._ip)
        if konek.get("status") != "connected":
            # Fallback: APK client mungkin tetap aktif via WebSocket walau port
            # ADB 5555 ditutup (biasa di Android TV yang tidak punya ADB debug).
            if self._hub and self._hub.is_meja_connected(self._label_tv):
                return {"adb": False, "ws": True,
                        "pesan": konek.get("message", "Gagal konek ADB")}
            return {"adb": False, "pesan": konek.get("message", "Gagal konek ADB")}
        ok, out = ADBHelper.adb_shell(self._ip, "pm list packages " + self.PACKAGE)
        if not ok:
            return {"adb": True, "pesan": out or "Gagal menjalankan perintah ADB."}
        terpasang = self.PACKAGE in (out or "")
        versi = ""
        if terpasang:
            ok2, out2 = ADBHelper.adb_shell(self._ip,
                                            "dumpsys package " + self.PACKAGE)
            m = re.search(r"versionName=([0-9.]+)", out2 or "")
            if m:
                versi = m.group(1)
        return {"adb": True, "terpasang": terpasang, "versi": versi}

    def _tampilkan(self, hasil):
        self._busy = False
        self.progress.set(0)
        self.lbl_progress.configure(text="")
        if self.btn_upgrade:
            self.btn_upgrade.configure(state="disabled")
        if not hasil.get("adb"):
            self.lbl_adb.configure(text="ADB: ❌ GAGAL", text_color=C_RED)
            if hasil.get("ws"):
                self.lbl_apk.configure(text="APK client: ✅ AKTIF (WebSocket)",
                                       text_color=C_GREEN)
                self.lbl_info.configure(
                    text="Aplikasi TV terhubung via WebSocket.\n"
                         "ADB (port 5555) tidak tersedia — install perlu "
                         "mengaktifkan ADB debugging di TV.")
            else:
                self.lbl_apk.configure(text=f"Pesan: {str(hasil.get('pesan', ''))[:60]}",
                                       text_color=C_RED)
                self.lbl_info.configure(text="Pastikan ADB debugging di TV aktif.")
            return
        self.lbl_adb.configure(text="ADB: ✅ Terhubung", text_color=C_GREEN)
        if hasil.get("terpasang"):
            v = hasil.get("versi") or "?"
            self.lbl_apk.configure(text=f"APK client: ✅ TERPASANG (v{v})",
                                   text_color=C_GREEN)
            self.btn_install.configure(state="disabled",
                                       text="🔧 INSTALL APK (sudah terpasang)")
        else:
            self.lbl_apk.configure(text="APK client: ❌ BELUM TERPASANG",
                                   text_color=C_RED)
            self.btn_install.configure(state="normal")
        if self.btn_upgrade:
            self.btn_upgrade.configure(state="normal")

    def _pilih_apk(self):
        if self._busy:
            return
        base = os.path.dirname(os.path.abspath(sys.argv[0]))
        default = os.path.join(base, "android_tv_client", "app", "build",
                               "outputs", "apk", "debug", "app-debug.apk")
        if not os.path.isfile(default):
            default = os.path.expanduser("~")
        path = filedialog.askopenfilename(
            parent=self, title="Pilih file APK client",
            initialdir=os.path.dirname(default),
            filetypes=[("APK file", "*.apk")])
        if not path:
            return
        self._busy = True
        self.btn_install.configure(state="disabled", text="⏳ Menginstall…")
        if self.btn_upgrade:
            self.btn_upgrade.configure(state="disabled")
        self.progress.set(0)
        self.lbl_progress.configure(text="")
        self.lbl_info.configure(text="Mengirim APK via ADB… mohon tunggu.")
        threading.Thread(target=self._install_thread, args=(path,), daemon=True).start()

    def _set_progress(self, pct, pesan=""):
        try:
            self.after(0, lambda: (self.progress.set(max(0.0, min(1.0, int(pct) / 100.0))),
                                   self.lbl_progress.configure(text=pesan)))
        except Exception:
            pass

    def _install_thread(self, path):
        ok, pesan = ADBHelper.adb_install_with_progress(
            self._ip, path, self._set_progress)
        self.after(0, self._install_selesai, ok, pesan)

    def _install_selesai(self, ok, pesan):
        self._busy = False
        self.btn_install.configure(text="🔧 INSTALL APK")
        self.lbl_info.configure(text="")
        self.progress.set(1 if ok else 0)
        self.lbl_progress.configure(text="✅ 100% — Install selesai" if ok else "✖ Gagal")
        if self.btn_upgrade:
            self.btn_upgrade.configure(state="normal")
        if ok:
            self.lbl_apk.configure(text="APK client: ✅ TERPASANG", text_color=C_GREEN)
            messagebox.showinfo("✅ Install Berhasil",
                                f"APK terpasang di {self._ip}\n{pesan}", parent=self)
        else:
            self.lbl_apk.configure(text="APK client: ❌ GAGAL DIPASANG", text_color=C_RED)
            messagebox.showerror("❌ Install Gagal", str(pesan), parent=self)

    # ── Upgrade APK versi terbaru (unduh dari URL → pasang via ADB) ─────────
    def _upgrade_apk(self):
        """Unduh APK client terbaru lalu pasang via ADB.

        Sumber URL (prioritas):
        1. Config 'apk_tv_url' (URL APK langsung atau manifest.json).
        2. Auto-detect: asset .apk pada release GitHub terbaru (dari
           'update_manifest_url', pola github.com/<owner>/<repo>).
        3. URL resmi bawaan (fallback terakhir tanpa dialog manual).
        """
        if self._busy:
            return
        url = ""
        source = ""
        # 1) URL yang sudah diatur di config (diisi otomatis saat rilis)
        url = str(ConfigManager.get("apk_tv_url") or "").strip()
        if url:
            source = "config 'apk_tv_url'"
        # 2) Auto-detect dari GitHub release terbaru
        if not url:
            manifest_url = str(ConfigManager.get("update_manifest_url") or "").strip()
            if manifest_url:
                try:
                    from scripts import check_update
                    found = check_update.find_latest_apk_url(manifest_url)
                    if found:
                        url = found
                        source = "GitHub release terbaru"
                except Exception:
                    pass
        # 3) URL resmi bawaan (tanpa dialog manual)
        if not url:
            url = ("https://github.com/dedekemoking-commits/rr_billing_pro_windows/"
                   "releases/latest/download/RRBillingPro-TV.apk")
            source = "URL resmi otomatis"
        if url:
            try:
                ConfigManager.set("apk_tv_url", url)
            except Exception:
                pass
        self._busy = True
        self.btn_upgrade.configure(state="disabled", text="⏳ Mengunduh…")
        self.btn_install.configure(state="disabled")
        self.progress.set(0)
        self.lbl_progress.configure(text="")
        self.lbl_info.configure(text=f"Mengunduh APK… ({source}) mohon tunggu.")
        threading.Thread(target=self._upgrade_thread,
                         args=(str(url).strip(),), daemon=True).start()

    def _upgrade_thread(self, url):
        def _ui(fn):
            try:
                self.after(0, fn)
            except Exception:
                pass

        def _dl_progress(d, t):
            pct = int(d * 80 // max(t, 1)) if t else 0
            self._set_progress(pct, f"Mengunduh APK… {pct}%")

        tmpdir = tempfile.mkdtemp(prefix="rr_tv_apk_")
        try:
            from scripts import check_update
            apk_path = None
            if url.lower().endswith((".json",)):
                # Manifest JSON (opsional): asset_url + sha256 untuk verifikasi
                try:
                    import urllib.request as _ur
                    with _ur.urlopen(url, timeout=30) as r:
                        mf = json.loads(r.read().decode("utf-8"))
                except Exception as e:
                    raise ValueError(f"Gagal baca manifest: {e}")
                asset = str(mf.get("asset_url") or "").strip()
                if not asset:
                    raise ValueError("Manifest tidak punya field asset_url")
                apk_path = os.path.join(
                    tmpdir, os.path.basename(asset.split("?")[0]) or "tv_client.apk")
                check_update.download_asset(
                    asset, apk_path, mf.get("sha256") or None,
                    progress_cb=_dl_progress)
            else:
                apk_path = os.path.join(tmpdir, "tv_client.apk")
                check_update.download_asset(
                    url, apk_path, None, progress_cb=_dl_progress)

            # Cegah downgrade: bandingkan versi APK baru vs yang terpasang di TV
            try:
                new_code, new_name = check_update.read_apk_version(apk_path)
            except Exception:
                new_code, new_name = None, None
            if new_code is not None:
                ok3, out3 = ADBHelper.adb_shell(
                    self._ip, "dumpsys package " + self.PACKAGE)
                m = re.search(r"versionCode=(\d+)", out3 or "")
                cur_code = int(m.group(1)) if m else None
                if cur_code is not None and new_code <= cur_code:
                    _ui(lambda: self._upgrade_selesai(
                        True,
                        f"APK di sumber (v{new_name or new_code}) tidak lebih baru "
                        f"dari yang terpasang di TV (v{cur_code}).\n\n"
                        "Tidak perlu upgrade. Pastikan APK versi terbaru sudah "
                        "di-upload ke GitHub release."))
                    return

            self._set_progress(82, "Memasang APK via ADB…")
            ok, pesan = ADBHelper.adb_install_with_progress(
                self._ip, apk_path, self._set_progress)
            if not ok:
                if "INSTALL_FAILED_VERSION_DOWNGRADE" in (pesan or ""):
                    pesan = ("Versi APK lebih lama dari yang terpasang di TV — "
                             "upgrade ditolak (downgrade).\n\n"
                             "Upload APK versi terbaru ke GitHub release, lalu "
                             "coba lagi.")
                _ui(lambda: self._upgrade_selesai(False, pesan))
                return
            versi = ""
            ok2, out2 = ADBHelper.adb_shell(self._ip,
                                            "dumpsys package " + self.PACKAGE)
            m = re.search(r"versionName=([0-9.]+)", out2 or "")
            if m:
                versi = m.group(1)
            pesan_akhir = f"APK terpasang di {self._ip}."
            if versi:
                pesan_akhir += f"\nVersi baru: v{versi}"
            _ui(lambda: self._upgrade_selesai(True, pesan_akhir))
        except Exception as e:
            _ui(lambda msg=str(e): self._upgrade_selesai(False, msg))
        finally:
            try:
                shutil.rmtree(tmpdir, ignore_errors=True)
            except Exception:
                pass

    def _upgrade_selesai(self, ok, pesan):
        self._busy = False
        self.lbl_info.configure(text="")
        self.progress.set(1 if ok else 0)
        self.lbl_progress.configure(text="✅ 100% — Upgrade selesai" if ok else "✖ Gagal")
        if self.btn_upgrade:
            self.btn_upgrade.configure(
                state="normal", text="⬆ UPGRADE APK (Unduh & Pasang Versi Baru)")
        if ok:
            self.lbl_apk.configure(text="APK client: ✅ TERPASANG", text_color=C_GREEN)
            messagebox.showinfo("✅ Upgrade APK Berhasil",
                                f"{self._label_tv} ({self._ip})\n{pesan}",
                                parent=self)
        else:
            self.lbl_apk.configure(text="APK client: ❌ UPGRADE GAGAL", text_color=C_RED)
            messagebox.showerror("❌ Upgrade APK Gagal", str(pesan), parent=self)


class KartuTV(tk.Canvas):
    def __init__(self, master, nomor, ip, port, label_tv, on_transaksi,
                 get_paket_data, get_makanan_data, get_minuman_data,
                 get_semua_kartu, nama_grup="Reguler", is_first=False,
                 get_daftar_grup=None, on_ganti_grup=None, on_hapus=None,
                 role="admin", **kwargs):
        super().__init__(master, highlightthickness=0, bd=0,
                         bg="white", **kwargs)
        self.role         = role or "admin"
        self.nomor        = nomor
        self.ip           = ip
        self.port         = port
        self.label_tv     = label_tv
        self.on_transaksi = on_transaksi
        self.nama_grup        = nama_grup
        self.get_paket_data   = get_paket_data
        self.get_makanan_data = get_makanan_data
        self.get_minuman_data = get_minuman_data
        self.get_semua_kartu  = get_semua_kartu
        self.get_daftar_grup  = get_daftar_grup or (lambda: [nama_grup])
        self.on_ganti_grup    = on_ganti_grup
        self.on_hapus         = on_hapus
        self.is_first     = is_first
        self.is_on        = False
        self.connected    = True

        self.paket_aktif   = None
        self.sisa_waktu    = 0
        self.is_bebas      = False
        self.menit_dipakai_awal = 0
        self.waktu_mulai   = None
        self.pesanan_aktif = {}
        self.biaya_pesanan = 0
        self.paket_harga_tetap = 0
        self.daftar_paket_sesi = []
        # Status pembayaran per item: lunas_paket[i] sejajar daftar_paket_sesi[i],
        # harga_paket_sesi[i] = harga paket ke-i, lunas_pesanan[nm] = status item pesanan.
        self.lunas_paket = []
        self.harga_paket_sesi = []
        self.lunas_pesanan = {}
        self.diskoni       = 0
        self.diskoni_mode  = "nominal"
        self._timer_job    = None
        self._last_transaction_item = None
        self._last_riwayat_idx = None
        self._last_cloud_id = None
        self.paid          = True   # status pembayaran sesi (sinkron ke riwayat)
        self._warning_blink_on = False
        self._timer_paused = False
        self._timer_was_running = False

        self._ids = {}
        self._btn_states = {}
        self._card_w = 260
        self._BLACK_BTNS = frozenset({
            "vol_up", "vol_dn", "home", "remote", "apk",
            "shop", "paket", "pause", "ip", "pindah",
            "video", "gambar", "logo", "qr", "ganti_nama", "hapus",
        })

        self._build()
        self.connected = False
        self._online_check_running = False
        self._power_check_running = False
        self._tv_reachable = True
        # Auto-off TV tanpa paket (anti-kasir nakal): idle_on_seconds = akumulasi
        # detik layar nyala tanpa sesi; idle_escalated = pernah di-auto-mati
        # (ambang turun dari 10 menit ke 5 menit sampai paket sah dibuka).
        self.idle_on_seconds = 0
        self.idle_escalated = False
        self._power_state_known = False
        self.after(3000, self._online_checker)
        self.after(2000, self._refresh_client_badge)
        self.after(2000, self._tv_power_checker)
 
        self.bind("<Configure>", self._on_card_resize)

    def _build(self):
        if self._ids:
            self.delete("all")
            self._ids = {}
        self._card_bg_color = C_CARD
        self._build_inner()

    def _on_card_resize(self, event):
        if event.width > 50 and abs(event.width - self._card_w) > 5:
            self._card_w = event.width
            job = getattr(self, "_resize_job", None)
            if job:
                try:
                    self.after_cancel(job)
                except Exception:
                    pass
            self._resize_job = self.after(200, lambda: self.winfo_exists() and self._build())

    def _build_inner(self):
        W = self._card_w
        y = 2

        # ── Header (C_ACCENT2 background) ───────────────────────────────────
        hdr_h, hdr_y = 48, y
        self._ids['hdr'] = self.create_rectangle(2, hdr_y, W-2, hdr_y+hdr_h,
            fill=C_ACCENT2, outline="", width=0, tags="hdr")
        y += 4

        display_label = self.label_tv
        if display_label.upper().startswith("TV "):
            display_label = display_label[3:]

        # TV name
        self._ids['tv_name'] = self.create_text(10, y+10,
            text=display_label, font=("Russo One", 13, "bold"),
            fill="white", anchor="w", tags="tv_name")

        # Badge status pembayaran (sebelah nama)
        self._ids['paid_badge'] = self.create_text(10, y+10,
            text="", font=("Courier New", 10, "bold"),
            fill=C_GREEN, anchor="w", tags="paid_badge")

        # "Nama" button
        self._draw_canvas_btn("ganti_nama", W-126, y-2, 76, 24, "Nama",
            "black", "white", ("Russo One", 10, "bold"), self._buka_ganti_nama)

        # "✖" button
        self._draw_canvas_btn("hapus", W-46, y-2, 36, 24, "✖",
            "black", "white", ("Russo One", 10, "bold"), self._confirm_hapus)

        y += 26
        # IP + badge
        self._ids['ip_footer'] = self.create_text(10, y+6,
            text=f"IP: {self.ip}", font=("Courier New", 10, "bold"),
            fill="white", anchor="w", tags="ip_footer")
        _method = ADBHelper.get_connection_method(self.ip)
        _badge_text = "ONLINE" if _method == "atpv2" else "ADB"
        _badge_color = "#00D68F" if _method == "atpv2" else "#FFAA00"
        self._ids['metode_badge'] = self.create_text(W-10, y+6,
            text=_badge_text, font=("Courier New", 10, "bold"),
            fill=_badge_color, anchor="e", tags="metode_badge")
        self._ids['online_dot'] = self.create_text(W-60, y+6,
            text="●", font=("Courier New", 10, "bold"),
            fill=C_GREEN, anchor="e", tags="online_dot")
        self._ids['cli_badge'] = self.create_text(W-82, y+6,
            text="CLI✗", font=("Courier New", 10, "bold"),
            fill=C_MUTED, anchor="e", tags="cli_badge")
        y = hdr_y + hdr_h + 4

        # ── Status row ──────────────────────────────────────────────────────
        srow_y = y
        self._ids['lbl_power'] = self.create_text(8, y+10,
            text="● HIDEN", font=("Courier New", 11, "bold"),
            fill=C_GREEN, anchor="w", tags="lbl_power")
        self._ids['lbl_grup'] = self.create_text(80, y+10,
            text=f"\u21bb Reguler", font=("Courier New", 11, "bold"),
            fill=C_ACCENT2, anchor="w", tags="lbl_grup")
        self.tag_bind("lbl_grup", "<Button-1>", lambda e: self._buka_ganti_grup())
        self._ids['lbl_paket'] = self.create_text(W-8, y+10,
            text="\u2014", font=("Courier New", 10),
            fill=C_MUTED, anchor="e", tags="lbl_paket")
        y += 24

        # ── Timer ────────────────────────────────────────────────────────────
        self._ids['lbl_timer'] = self.create_text(W//2, y+14,
            text="00:00:00", font=("Russo One", 20, "bold"),
            fill=C_ACCENT2, anchor="center", tags="lbl_timer")
        y += 32

        # ── Estimasi ─────────────────────────────────────────────────────────
        self._ids['lbl_estimasi'] = self.create_text(W//2, y+6,
            text="", font=("Courier New", 10),
            fill=C_YELLOW, anchor="center", tags="lbl_estimasi")
        y += 16

        # ── Button Row 1 ────────────────────────────────────────────────────
        btn_h, gap_b = 30, 4
        n_btn, btn_cols = 6, 6
        avail = W - 8 - (n_btn - 1) * gap_b
        bw = avail // btn_cols
        bx = 4

        r1y = y
        btn_defs1 = [
            ("power", "\u26a1 PWR", C_RED, C_RED, self._toggle_power),
            ("vol_up", "VOL+", "black", "white", lambda: self._adb_action(lambda: ADBHelper.volume(self.ip, naik=True, port=self.port))),
            ("vol_dn", "VOL\u2212", "black", "white", lambda: self._adb_action(lambda: ADBHelper.volume(self.ip, naik=False, port=self.port))),
            ("home", "HOME", "black", "white", lambda: self._adb_action(lambda: ADBHelper.home(self.ip, port=self.port))),
            ("remote", "RMT", "black", "white", self._buka_remote),
            ("apk", "APK", "black", "white", self._buka_status_client),
        ]
        for i, (key, txt, bg, fg, cmd) in enumerate(btn_defs1):
            self._draw_canvas_btn(key, bx + i*(bw+gap_b), r1y, bw, btn_h, txt, bg, fg, ("Courier New", 10, "bold"), cmd)
        y = r1y + btn_h + 4

        # ── Button Row 2 ────────────────────────────────────────────────────
        r2y = y
        n_btn2, btn_cols2 = 6, 6
        avail2 = W - 8 - (n_btn2 - 1) * gap_b
        bw2 = avail2 // btn_cols2
        btn_defs2 = [
            ("selesai", "SELESAI", C_BTN, C_RED, self._klik_selesai, True),
            ("shop", "SHOP", "black", "white", self._buka_tambah_pesanan, True),
            ("paket", "PAKET", "black", "white", self._pilih_paket, False),
            ("pause", "PAUSE", "black", "white", self._toggle_billing_pause, True),
            ("ip", "IP", "black", "white", self._buka_ganti_ip, False),
            ("pindah", "Pindah", "black", "white", self._klik_pindah, False),
        ]
        for i, (key, txt, bg, fg, cmd, disabled) in enumerate(btn_defs2):
            self._draw_canvas_btn(key, bx + i*(bw2+gap_b), r2y, bw2, btn_h, txt, bg, fg, ("Russo One", 9, "bold"), cmd)
            if disabled:
                self._disable_btn(key)
        y = r2y + btn_h + 4

        # ── Button Row 3 (Media promosi + QR) ─────────────────────────────
        r3y = y
        n_btn3, btn_cols3 = 4, 4
        avail3 = W - 8 - (n_btn3 - 1) * gap_b
        bw3 = avail3 // btn_cols3
        btn_defs3 = [
            ("video", "🎬 VIDEO", "black", "white", self._buka_media_video),
            ("gambar", "🖼 GAMBAR", "black", "white", self._buka_media_gambar),
            ("logo", "🖼 LOGO", "black", "white", self._ganti_logo_kartu),
            ("qr", "📱 QR", "black", "white", self._buka_qr_kartu),
        ]
        for i, (key, txt, bg, fg, cmd) in enumerate(btn_defs3):
            self._draw_canvas_btn(key, bx + i*(bw3+gap_b), r3y, bw3, btn_h, txt, bg, fg, ("Russo One", 9, "bold"), cmd)
        y = r3y + btn_h + 6
        # Kasir tidak boleh mengubah IP TV / media promosi / logo
        if self.role != "admin":
            for k in ("ip", "video", "gambar", "logo"):
                self._disable_btn(k)
        self._update_paid_badge()

        # Update card background height
        total_h = y
        self._ids['_card_h'] = total_h
        self.configure(height=total_h, width=W)
        self._ids['bg'] = self.create_rectangle(0, 0, W, total_h, fill=C_CARD, outline=C_BORDER, width=1, tags="bg")
        self.tag_lower(self._ids['bg'])
        self._redraw_bg()

    def _redraw_bg(self):
        W = self._card_w
        h = self._ids.get('_card_h', 200)
        self.coords(self._ids['bg'], 0, 0, W, h)
        self.tag_lower(self._ids['bg'])

    def _draw_canvas_btn(self, key, x, y, w, h, text, bg, fg, font, cmd):
        rect = self.create_rectangle(x, y, x+w, y+h, fill=bg, outline=fg, width=1, tags=f"btn_{key}")
        txt_id = self.create_text(x+w//2, y+h//2, text=text, font=font,
                                  fill=self._btn_text_color(bg, fg), tags=(f"btn_{key}", f"btn_{key}_txt"))
        self._ids[f'btn_{key}'] = rect
        self._ids[f'btn_{key}_txt'] = txt_id
        self._btn_states[key] = "normal"
        self.tag_bind(f"btn_{key}", "<Button-1>", lambda e, k=key, c=cmd: c() if self._btn_states.get(k) != "disabled" else None)
        self.tag_bind(f"btn_{key}", "<Enter>", lambda e, k=key, bg=bg: self._btn_hover(k, bg))
        self.tag_bind(f"btn_{key}", "<Leave>", lambda e, k=key, bg=bg: self._btn_leave(k, bg))
        return rect, txt_id

    def _btn_text_color(self, bg, fg):
        if bg == "black":
            return "white"
        if bg == C_ACCENT:
            return "black"
        if bg == C_BTN:
            return fg
        return "white"

    def _btn_hover(self, key, bg):
        if self._btn_states.get(key) == "disabled":
            return
        if bg == "black":
            self.itemconfig(self._ids[f'btn_{key}'], fill="#333333")
        elif bg == C_RED:
            self.itemconfig(self._ids[f'btn_{key}'], fill="#FF6666")
        elif bg == C_GREEN:
            self.itemconfig(self._ids[f'btn_{key}'], fill="#66BB6A")
        elif bg == C_ACCENT:
            self.itemconfig(self._ids[f'btn_{key}'], fill="#66FFE0")
        elif bg == C_BTN:
            self.itemconfig(self._ids[f'btn_{key}'], fill=C_ACCENT2)

    def _btn_leave(self, key, bg):
        if self._btn_states.get(key) == "disabled":
            self._disable_btn(key)
        else:
            self.itemconfig(self._ids[f'btn_{key}'], fill=bg)

    def _enable_btn(self, key, bg=C_BTN, fg=C_ACCENT2):
        self._btn_states[key] = "normal"
        self.itemconfig(self._ids[f'btn_{key}'], fill=bg, outline=fg)
        self.itemconfig(self._ids[f'btn_{key}_txt'], fill=fg)

    def _disable_btn(self, key):
        self._btn_states[key] = "disabled"
        if key in self._BLACK_BTNS:
            self.itemconfig(self._ids[f'btn_{key}'], fill="black", outline=C_BORDER)
        else:
            self.itemconfig(self._ids[f'btn_{key}'], fill=C_BTN, outline=C_BORDER)
        self.itemconfig(self._ids[f'btn_{key}_txt'], fill=C_MUTED)

    # ── Util status sesi ─────────────────────────────────────────────────────
    def sesi_kosong(self):
        return self.paket_aktif is None

    # ── Status pembayaran (sinkron ke riwayat) ───────────────────────────────
    def _update_paid_badge(self):
        """Update badge 'LUNAS' / 'TAGIHAN' (atau gabungan) di header, tepat di sebelah nama."""
        if 'paid_badge' not in self._ids:
            return
        display_label = self.label_tv
        if display_label.upper().startswith("TV "):
            display_label = display_label[3:]
        # Estimasi lebar nama (Russo One 11 bold ≈ 7.5px/char) + jarak 6px
        name_w = len(display_label) * 7 + 6
        if self.sesi_kosong():
            text, color = "", C_GREEN
        elif self.is_bebas:
            text, color = "⏳ TAGIHAN", "#FFCC00"
        else:
            lunas, tagihan = self._split_payment()
            if lunas > 0 and tagihan > 0:
                text, color = "LUNAS + TAGIHAN", "#FF9933"
            elif lunas > 0:
                text, color = "● LUNAS", C_GREEN
            else:
                text, color = "⏳ TAGIHAN", "#FFCC00"
        self.itemconfig(self._ids['paid_badge'], text=text, fill=color)
        try:
            # Lebarkan sedikit saat teks gabungan agar tidak tumpang tindih
            cur_x, cur_y = self.coords(self._ids['tv_name'])
        except Exception:
            cur_x, cur_y = 10, 5
        self.coords(self._ids['paid_badge'], cur_x + len(display_label) * 7 + 6, cur_y)

    def _update_bayar_buttons(self):
        """Warna tombol SUDAH/BELUM BAYAR sesuai state aktif (lunas = hijau)."""
        if 'btn_bayar_lunas' not in self._ids:
            return
        lunas, tagihan = self._split_payment()
        if tagihan <= 0:
            self.itemconfig(self._ids['btn_bayar_lunas'], fill=C_GREEN, outline=C_GREEN)
            self.itemconfig(self._ids['btn_bayar_lunas_txt'], fill="white")
            self.itemconfig(self._ids['btn_bayar_belum'], fill="black", outline=C_BORDER)
            self.itemconfig(self._ids['btn_bayar_belum_txt'], fill=C_MUTED)
        else:
            self.itemconfig(self._ids['btn_bayar_belum'], fill="#FFCC00", outline="#FFCC00")
            self.itemconfig(self._ids['btn_bayar_belum_txt'], fill="black")
            self.itemconfig(self._ids['btn_bayar_lunas'], fill="black", outline=C_BORDER)
            self.itemconfig(self._ids['btn_bayar_lunas_txt'], fill=C_MUTED)

    def _bind_last_transaction(self):
        """Simpan referensi index/cloud_id baris riwayat milik sesi ini."""
        try:
            app = self.winfo_toplevel()
            item_id = getattr(self, '_last_transaction_item', None)
            idx = app._tree_item_to_index.get(item_id, -1)
            if idx < 0:
                idx = getattr(app, "_last_catat_idx", -1)
            self._last_riwayat_idx = idx
            if 0 <= idx < len(app.riwayat_meta):
                self._last_cloud_id = app.riwayat_meta[idx].get("cloud_id")
        except Exception:
            pass

    def _set_paid(self, paid):
        """Tandai sesi lunas/belum lunas, sinkronkan ke baris riwayat terkait."""
        if self.sesi_kosong() or self.is_bebas:
            return
        if self._last_transaction_item is None and self._last_riwayat_idx is None:
            return
        app = self.winfo_toplevel()
        idx = app._resolve_session_idx(self)
        row_needs_sync = False
        if 0 <= idx < len(app.riwayat_meta):
            row_needs_sync = bool(app.riwayat_meta[idx].get('paid', True)) != bool(paid)
        if self.paid == paid and not row_needs_sync:
            return
        self.paid = paid
        # Semua item ikut memilih tombol
        self.lunas_paket = [paid] * (len(self.daftar_paket_sesi or []) or 1)
        if self.pesanan_aktif:
            self.lunas_pesanan = {nm: paid for nm in self.pesanan_aktif}
        self._update_paid_badge()
        self._update_bayar_buttons()
        if idx >= 0 and hasattr(app, '_set_transaksi_paid_idx'):
            app._set_transaksi_paid_idx(idx, paid)
        elif hasattr(app, '_set_transaksi_paid'):
            app._set_transaksi_paid(self._last_transaction_item, paid)

    # ── TV WebSocket Hub helpers (Overlay & Lockscreen Android TV) ──────────
    def _split_payment(self):
        """Split tagihan sesi jadi (lunas_total, tagihan_total) PER ITEM.

        Status dihitung granular per paket (sejajar daftar_paket_sesi) dan
        per item pesanan (lunas_pesanan). Diskon dialokasikan proporsional
        terhadap subtotal lunas/tagihan.
        """
        try:
            total = self._total_setelah_diskon()
        except Exception:
            total = getattr(self, "paket_harga_tetap", 0) + getattr(self, "biaya_pesanan", 0)
        if self.sesi_kosong():
            return 0, 0
        try:
            subtotal = self.paket_harga_tetap + self.biaya_pesanan
        except Exception:
            subtotal = 0
        if subtotal <= 0:
            if getattr(self, "paid", True):
                return total, 0
            return 0, total

        # ── subtotal LUNAS per item ──
        lunas_sub = 0
        # paket per segmen sesi
        harga_paket = getattr(self, "harga_paket_sesi", None) or []
        lunas_paket = getattr(self, "lunas_paket", None) or []
        if harga_paket:
            for i, h in enumerate(harga_paket):
                paid_i = lunas_paket[i] if i < len(lunas_paket) else getattr(self, "paid", True)
                if paid_i:
                    lunas_sub += h
        else:
            # fallback sesi lama (tanpa data per item): pakai paid keseluruhan
            if getattr(self, "paid", True):
                lunas_sub += self.paket_harga_tetap

        # pesanan per item
        lunas_pesanan = getattr(self, "lunas_pesanan", None) or {}
        all_menu = {}
        try:
            all_menu = {**self.get_makanan_data(), **self.get_minuman_data()}
        except Exception:
            pass
        for nm, qty in (self.pesanan_aktif or {}).items():
            paid_i = lunas_pesanan.get(nm, getattr(self, "paid", True))
            if paid_i:
                lunas_sub += all_menu.get(nm, 0) * qty

        lunas_sub = min(lunas_sub, subtotal)
        if lunas_sub >= subtotal:
            lunas_res = total
        elif lunas_sub <= 0:
            lunas_res = 0
        else:
            # diskon proporsional terhadap subtotal lunas
            lunas_res = round(total * lunas_sub / subtotal)
        lunas_res = min(lunas_res, total)
        tagihan_res = max(0, total - lunas_res)
        # DP booking: nominal yang sudah dibayar pelanggan via transfer tetap
        # dihitung LUNAS walau item belum ditandai — tampil "bayar X / tagihan Y"
        try:
            dp_pre = int(getattr(self, "dp_bayar_awal", 0) or 0)
        except Exception:
            dp_pre = 0
        if dp_pre > 0 and tagihan_res > 0:
            lunas_res = min(total, max(lunas_res, dp_pre))
            tagihan_res = max(0, total - lunas_res)
        return lunas_res, tagihan_res

    def _hub(self):
        app = self.winfo_toplevel()
        return getattr(app, 'tv_ws_hub', None)

    def _ws_send_start(self, sisa_detik):
        hub = self._hub()
        if not hub:
            return
        try:
            lunas, tagihan = self._split_payment()
            hub.send_start_timer(self.label_tv, sisa_detik, self._total_setelah_diskon(),
                                 lunas_total=lunas, tagihan_total=tagihan)
        except Exception as e:
            print(f"[TV WS HUB] send_start error {self.label_tv}: {e}")
        # Paket sah dibuka → batalkan hitungan auto-off & kembali ke 10 menit.
        self.idle_on_seconds = 0
        self.idle_escalated = False

    def _ws_send_total(self, total):
        hub = self._hub()
        if not hub:
            return
        try:
            lunas, tagihan = self._split_payment()
            hub.send_update_total(self.label_tv, total,
                                  lunas_total=lunas, tagihan_total=tagihan)
        except Exception as e:
            print(f"[TV WS HUB] send_update_total error {self.label_tv}: {e}")

    def _ws_send_sync(self):
        hub = self._hub()
        if not hub:
            return
        try:
            lunas, tagihan = self._split_payment()
            hub.send_sync_timer(self.label_tv, self.sisa_waktu, self._total_setelah_diskon(),
                                lunas_total=lunas, tagihan_total=tagihan)
        except Exception as e:
            print(f"[TV WS HUB] send_sync_timer error {self.label_tv}: {e}")

    def _ws_send_stop(self):
        hub = self._hub()
        if not hub:
            return
        try:
            hub.send_stop_timer(self.label_tv)
            hub.send_unlock_screen(self.label_tv)
        except Exception as e:
            print(f"[TV WS HUB] send_stop error {self.label_tv}: {e}")

    def _ws_send_lock(self, pesan, detail):
        hub = self._hub()
        if not hub:
            return
        try:
            hub.send_lock_screen(self.label_tv, pesan, detail)
        except Exception as e:
            print(f"[TV WS HUB] send_lock error {self.label_tv}: {e}")

    # ── Auto-off TV tanpa paket aktif (anti-kasir nakal) ────────────────────
    # Dipanggil poller aplikasi tiap 30 detik. Logika:
    #   - layar nyala + tanpa sesi  → akumulasi detik
    #   - ambang 10 menit (pertama) → 5 menit setelah pernah di-auto-mati
    #   - paket sah dibuka          → reset counter & kembali ke 10 menit
    #   - layar mati                → reset counter
    #   - status layar tak diketahui (APK tidak terhubung) → skip
    def _tv_idle_check(self):
        hub = self._hub()
        if not hub:
            return
        state = hub.get_screen_state(self.label_tv)
        if state is None:
            return
        if not state:
            self.idle_on_seconds = 0
            return
        if not self.sesi_kosong():
            self.idle_on_seconds = 0
            self.idle_escalated = False
            return
        self.idle_on_seconds += 30
        app = self.winfo_toplevel()
        threshold = (getattr(app, 'tv_auto_off_sec', 300)
                     if self.idle_escalated
                     else getattr(app, 'tv_auto_off_first_sec', 600))
        if self.idle_on_seconds >= threshold:
            self.idle_on_seconds = 0
            self.idle_escalated = True
            self._tv_sleep_now(threshold)

    def _tv_sleep_now(self, grace_detik, alasan=None):
        """Matikan TV (sleep) dengan upaya berlapis:
        1) power_toggle atpv2 (auto-reconnect + retry max 3x, selang 2 dtk),
        2) KEYCODE_SLEEP via atpv2 (beberapa TV hanya merespons SLEEP),
        3) fallback ADB: adb shell input keyevent KEYCODE_POWER / KEYCODE_SLEEP,
        4) verifikasi status layar ±9 dtk -> audit jujur (ok hanya jika TV
           benar-benar mati; gagal dicatat dengan alasan, tidak menipu log).
        """
        app = self.winfo_toplevel()
        user = getattr(app, "current_user", "") or ""
        if alasan is None:
            alasan = f"Layar nyala tanpa paket aktif selama {grace_detik // 60} menit"
        print(f"[TV SLEEP] {self.label_tv} ({self.ip}): {alasan}")

        def runner():
            hasil, pesan = False, ""
            # 1) atpv2 POWER dengan retry
            for _ in range(3):
                try:
                    ok, out, _ = ADBHelper.power_toggle(self.ip, port=self.port)
                    if ok:
                        hasil, pesan = True, "atpv2 POWER terkirim"
                        break
                    pesan = str(out)[:120]
                except Exception as e:
                    pesan = str(e)
                time.sleep(2)
            # 2) KEYCODE_SLEEP via atpv2
            if not hasil:
                try:
                    rem = ADBHelper._get_remote(self.ip)
                    res = rem.sleep_blocking()
                    if res.get("status") == "ok":
                        hasil, pesan = True, "atpv2 SLEEP terkirim"
                except Exception as e:
                    pesan = str(e)[:120]
            # 3) Fallback ADB (port 5555 / port kartu bila terbuka)
            if not hasil:
                for kunci in ("KEYCODE_POWER", "KEYCODE_SLEEP", "223"):
                    try:
                        ok_adb, out_adb = ADBHelper.adb_shell(self.ip, f"input keyevent {kunci}",
                                                              timeout=8, port=self.port)
                        if ok_adb:
                            hasil, pesan = True, f"fallback ADB keyevent {kunci}: {out_adb[:80]}"
                            break
                    except Exception as e:
                        pesan = str(e)[:120]
            # 4) Verifikasi status layar sungguhan (polls 3x, selang 3 dtk)
            if hasil:
                state = None
                for _ in range(3):
                    time.sleep(3)
                    try:
                        state = ADBHelper.tv_power_state(self.ip, self.port)
                    except Exception:
                        state = None
                    if state is False:
                        break
                    if state is None:
                        break  # tidak bisa diverifikasi; pertahankan hasil kirim
                if state is False:
                    pesan = pesan + " — terverifikasi MATI"
                elif state is True:
                    hasil, pesan = False, "perintah terkirim tapi layar masih nyala (verifikasi gagal)"
                else:
                    pesan = pesan + " — tanpa verifikasi (tak terdeteksi)"
            status = "ok" if hasil else "gagal"
            AuditLogger.log("TV_AUTO_OFF", user, status, {
                "label_tv": self.label_tv,
                "ip": self.ip,
                "alasan": alasan,
                "hasil": pesan,
                "jam": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            })
            print(f"[TV SLEEP] {self.label_tv}: {status.upper()} — {pesan}")

        threading.Thread(target=runner, daemon=True).start()

    def _tv_lock_media_urls(self, app):
        """URL logo lock (global) + video promosi aktif untuk detail LOCK_SCREEN.

        User NON-LIFETIME: logo default client (drawable bawaan) + video promosi
        BAWAAN (PROMO_VIDEO_DEFAULT) — media custom tidak pernah terpakai."""
        return app._tv_media_urls_now()

    def _ws_send_pause(self, paused):
        hub = self._hub()
        if not hub:
            return
        try:
            if paused:
                hub.send_pause_timer(self.label_tv)
            else:
                hub.send_resume_timer(self.label_tv, self.sisa_waktu)
        except Exception as e:
            print(f"[TV WS HUB] send_pause error {self.label_tv}: {e}")

    def _toggle_billing_pause(self):
        self._billing_paused = not getattr(self, '_billing_paused', False)
        if self._billing_paused:
            if self._timer_job:
                self.after_cancel(self._timer_job)
                self._timer_job = None
            self._ws_send_pause(True)
            self._enable_btn("pause", "black", "white")
            self.itemconfig(self._ids['btn_pause_txt'], text="PAUSED")
            self.itemconfig(self._ids['lbl_timer'], text="II PAUSED", fill=C_YELLOW)
            self.itemconfig(self._ids['lbl_estimasi'], text="", fill=C_YELLOW)
        else:
            self._ws_send_pause(False)
            self._enable_btn("pause", "black", "white")
            self.itemconfig(self._ids['btn_pause_txt'], text="PAUSE")
            if self.is_bebas:
                self._tick_bebas()
            elif self.sisa_waktu > 0:
                self._tick_waktu()

    def _buka_ganti_port(self):
        DialogGantiPort(self.winfo_toplevel(), self.label_tv, self.ip, self.port,
                        on_confirm=self._terapkan_port_baru)

    def _buka_ganti_ip(self):
        def _on_confirm(new_ip, new_port):
            try:
                self._terapkan_ip_baru(new_ip, new_port)
            except Exception:
                pass
        DialogGantiIP(self.winfo_toplevel(), self.label_tv, self.ip, self.port,
                      on_confirm=_on_confirm)
 
    def _buka_ganti_nama(self):
        dlg = ctk.CTkInputDialog(text="Nama TV baru:", title=f"✏️ Ganti Nama TV — {self.label_tv}")
        nama_baru = dlg.get_input()
        if not nama_baru:
            return
        nama_baru = nama_baru.strip()
        if not nama_baru:
            return
        lama = self.label_tv
        self.label_tv = nama_baru
        display_label = nama_baru
        if display_label.upper().startswith("TV "):
            display_label = display_label[3:]
        self.itemconfig(self._ids['tv_name'], text=display_label)
        self.itemconfig(self._ids['ip_footer'], text=f"IP: {self.ip}")
        self._update_paid_badge()
        self._simpan_daftar_tv()
        AuditLogger.log(action="rename_tv", username="", status="success", details={"old": lama, "new": nama_baru})
 
    def _confirm_hapus(self):
        if messagebox.askyesno("Hapus TV", f"Hapus {self.label_tv} dari dashboard?"):
            if callable(self.on_hapus):
                self.on_hapus(self)
 
    def _terapkan_ip_baru(self, new_ip, new_port):
        self.ip = new_ip
        self.port = new_port
        self.itemconfig(self._ids['ip_footer'], text=f"IP: {self.ip}")
        _method = ADBHelper.get_connection_method(self.ip)
        badge_text = "ONLINE" if _method == "atpv2" else "ADB"
        badge_color = "#00D68F" if _method == "atpv2" else "#FFAA00"
        self.itemconfig(self._ids['metode_badge'], text=badge_text, fill=badge_color)
        dot_id = self._ids.get('online_dot')
        if dot_id:
            self.itemconfig(dot_id, fill=C_YELLOW)
        self._simpan_daftar_tv()
        self.after(1000, self._online_checker)

    def _buka_ganti_grup(self):
        if not self.sesi_kosong():
            messagebox.showwarning("⚠ Sesi Sedang Berjalan",
                                    "Selesaikan sesi yang sedang berjalan dulu sebelum "
                                    "mengganti grup tarif TV ini.")
            return
        daftar = self.get_daftar_grup()
        if not daftar:
            return
        dlg = ctk.CTkToplevel(self.winfo_toplevel())
        dlg.title(f"Ganti Grup Tarif — {self.label_tv}")
        dlg.geometry("420x160")
        dlg.configure(fg_color=C_BG)
        dlg.grab_set()
        dlg.resizable(False, False)
        ctk.CTkLabel(dlg, text=f"🏷  Grup Tarif untuk {self.label_tv}",
                     font=("Russo One", 13, "bold"), text_color=C_ACCENT).pack(pady=(20, 12))
        var_grup = ctk.StringVar(value=self.nama_grup if self.nama_grup in daftar else daftar[0])
        opt = ctk.CTkOptionMenu(dlg, values=daftar, variable=var_grup,
                                 fg_color=C_BTN, button_color=C_ACCENT2,
                                 button_hover_color="#5A0FCC", text_color=C_TEXT,
                                 font=("Courier New", 12), dropdown_font=("Courier New", 11),
                                 dropdown_fg_color=C_CARD, dropdown_text_color=C_TEXT)
        opt.pack(padx=25, pady=8, fill="x")

        def terapkan():
            grup_baru = var_grup.get()
            self.nama_grup = grup_baru
            self.itemconfig(self._ids['lbl_grup'], text=f"\U0001f3f7 {grup_baru}")
            if self.on_ganti_grup:
                self.on_ganti_grup(self, grup_baru)
            dlg.destroy()

        ctk.CTkButton(dlg, text="✅  TERAPKAN GRUP", height=40,
                      fg_color=C_ACCENT2, hover_color=C_ACCENT,
                      font=("Russo One", 12, "bold"), text_color="white",
                      command=terapkan).pack(pady=16, padx=25, fill="x")

    def _terapkan_port_baru(self, port_baru):
        self.port = port_baru
        self._simpan_daftar_tv()

    def _simpan_daftar_tv(self):
        """Simpan daftar kartu TV (ip/nama/port/grup) ke config agar otomatis
        dimuat ulang saat login berikutnya (tanpa tambah manual)."""
        app = self.winfo_toplevel()
        if hasattr(app, "_simpan_daftar_tv"):
            app._simpan_daftar_tv()

    def _cek_koneksi_adb(self):
        if getattr(self, "_cek_busy", False): return
        self._cek_busy = True
        threading.Thread(target=self._cek_koneksi_thread, daemon=True).start()

    def _cek_koneksi_thread(self):
        sukses, status_awal, pesan = ADBHelper.cek_dan_reconnect(self.ip, self.port)
        self.after(0, self._cek_koneksi_selesai, sukses, status_awal, pesan)

    def _refresh_metode_badge(self):
        _method = ADBHelper.get_connection_method(self.ip)
        badge_text = "ONLINE" if _method == "atpv2" else "ADB"
        badge_color = "#00D68F" if _method == "atpv2" else "#FFAA00"
        self.itemconfig(self._ids['metode_badge'], text=badge_text, fill=badge_color)

    def _online_checker(self):
        if self._online_check_running:
            return
        self._online_check_running = True
        threading.Thread(target=self._online_check_thread, daemon=True).start()

    def _online_check_thread(self):
        reachable = ADBHelper.ping_host(self.ip, timeout=2)
        self.after(0, self._update_online_status, reachable)
        self._online_check_running = False
        try:
            self.after(30000, self._online_checker)
        except Exception:
            pass

    def _update_online_status(self, reachable):
        if not self.winfo_exists():
            return
        # TV dianggap online jika APK-nya terhubung WebSocket, walau port ADB
        # 5555 ditutup (sumber akurat — status layar juga datang dari sini).
        app = self.winfo_toplevel()
        hub = getattr(app, 'tv_ws_hub', None)
        ws_connected = bool(hub and hub.is_meja_connected(self.label_tv))
        online = bool(reachable or ws_connected)
        self.connected = online
        self._tv_reachable = online
        dot_id = self._ids.get('online_dot')
        if dot_id:
            self.itemconfig(dot_id, fill=C_GREEN if online else C_RED)
        badge_id = self._ids.get('metode_badge')
        if badge_id:
            badge_text = "ONLINE" if online else "OFFLINE"
            badge_color = C_GREEN if online else C_RED
            self.itemconfig(badge_id, text=badge_text, fill=badge_color)
        if ws_connected:
            # Status HIDUP/MATI diambil dari WebSocket APK (akurat) — tidak
            # perlu spawn adb dumpsys tiap menit.
            self._auto_reconnect_remote()
            return
        if reachable:
            self._auto_reconnect_remote()
            self._tv_power_checker()
        else:
            self._apply_tv_power_state(False)

    def _auto_reconnect_remote(self):
        """Reconnect otomatis remote atpv2 (androidtvremote2) saat TV online
        tapi remote belum terhubung — tanpa perlu re-pairing (sertifikat
        tersimpan per IP). Hanya jalan bila file sertifikat pairing ada."""
        if getattr(self, "_reconnect_running", False):
            return
        if not ADBHelper.adb_tersedia():
            return
        rem = ADBHelper._get_remote(self.ip)
        if rem and rem.is_connected():
            return
        certfile, _ = tv_mesin._cert_paths_for_ip(self.ip)
        if not os.path.isfile(certfile):
            return
        self._reconnect_running = True

        def runner():
            try:
                ok, _, pesan = ADBHelper.cek_dan_reconnect(self.ip, self.port)
                print(f"[TV RECONNECT] {self.label_tv} ({self.ip}): "
                      f"{'OK' if ok else str(pesan)[:120]}")
            except Exception as e:
                print(f"[TV RECONNECT] {self.label_tv}: {e}")
            finally:
                self._reconnect_running = False

        threading.Thread(target=runner, daemon=True).start()

    # ── Cek status nyala TV sesungguhnya (HIDUP/MATI) ───────────────────────
    def _tv_power_checker(self):
        if getattr(self, "_power_check_running", False):
            return
        app = self.winfo_toplevel()
        hub = getattr(app, 'tv_ws_hub', None)
        if hub and hub.is_meja_connected(self.label_tv):
            # Status layar sudah diketahui dari APK via WebSocket.
            return
        # TV offline → lewati (hindari spawn adb.exe tiap menit); cek lagi nanti
        if not getattr(self, "_tv_reachable", True):
            try:
                self.after(60000, self._tv_power_checker)
            except Exception:
                pass
            return
        self._power_check_running = True
        threading.Thread(target=self._tv_power_check_thread, daemon=True).start()

    def _tv_power_check_thread(self):
        try:
            state = ADBHelper.tv_power_state(self.ip, self.port)
        except Exception:
            state = None
        self._power_check_running = False
        try:
            self.after(0, self._apply_tv_power_state, state)
        except Exception:
            return
        try:
            self.after(60000, self._tv_power_checker)
        except Exception:
            pass

    def _apply_tv_power_state(self, state):
        if not self.winfo_exists():
            return
        if state is True:
            self.is_on = True
            self._power_state_known = True
            self.itemconfig(self._ids['lbl_power'], text="\U0001f4fa HIDUP", fill=C_GREEN)
            self.itemconfig(self._ids['btn_power'], fill="#3A0000")
        elif state is False:
            self.is_on = False
            self._power_state_known = True
            self.itemconfig(self._ids['lbl_power'], text="\U0001f4f5 MATI", fill=C_MUTED)
            self.itemconfig(self._ids['btn_power'], fill=C_BTN)

    def _cek_koneksi_selesai(self, sukses, status_awal, pesan):
        self._cek_busy = False
        self._refresh_metode_badge()
        if sukses:
            self.connected = True
            messagebox.showinfo("✅ Koneksi ADB", f"TV: {self.label_tv}\n{pesan}")
        else:
            self.connected = False
            if messagebox.askyesno("⚠ Koneksi Gagal",
                                    f"TV: {self.label_tv}\n{pesan}\n\nBuka dialog Ganti Port?"):
                self._buka_ganti_port()

    def _toggle_power(self):
        self.is_on = not self.is_on
        if self.is_on:
            self.itemconfig(self._ids['lbl_power'], text="\U0001f4fa HIDUP", fill=C_GREEN)
            self.itemconfig(self._ids['btn_power'], fill="#3A0000")
        else:
            self.itemconfig(self._ids['lbl_power'], text="\U0001f4f5 MATI", fill=C_MUTED)
            self.itemconfig(self._ids['btn_power'], fill=C_BTN)
        self._adb_action(lambda: ADBHelper.power_toggle(self.ip, port=self.port))

    def _adb_action(self, fn):
        def runner():
            ok, out, err = fn()
        threading.Thread(target=runner, daemon=True).start()

    def _buka_remote(self):
        VirtualRemoteDialog(self.winfo_toplevel(), self.label_tv, self.ip, self.port)

    # ── Status APK client TV (badge CLI + cek/install) ──────────────────────
    def _buka_status_client(self):
        DialogStatusClient(self.winfo_toplevel(), self.label_tv, self.ip,
                           self._install_apk_client, role=self.role)

    # ── Media promosi (video/gambar) ke client TV ───────────────────────────
    def _get_media_server(self):
        app = self.winfo_toplevel()
        return getattr(app, 'tv_media_server', None)

    def _pilih_media(self, kategori, filetypes, judul):
        app = self.winfo_toplevel()
        if not app._lisensi_lifetime():
            messagebox.showwarning(
                "Fitur Lisensi LIFETIME",
                f"Fitur {('video' if kategori == 'video' else 'gambar')} promosi "
                "hanya tersedia untuk lisensi LIFETIME.\n\n"
                "TV tetap memutar video promosi bawaan.\n"
                "Lakukan aktivasi lisensi LIFETIME untuk mengganti media promosi.",
                parent=app)
            return
        ms = self._get_media_server()
        if not ms or not ms.running:
            messagebox.showwarning("Media Server Mati",
                                   "Server media (port 8082) tidak berjalan.\n"
                                   "Cek konfigurasi tv_ws_enabled di rr_billing_config.json.",
                                   parent=self.winfo_toplevel())
            return
        path = filedialog.askopenfilename(
            parent=self.winfo_toplevel(), title=judul,
            initialdir=os.path.join(APP_BASE_DIR, "media_promo"),
            filetypes=filetypes)
        if not path:
            return
        if os.path.basename(path).lower() == PROMO_VIDEO_DEFAULT.lower():
            messagebox.showwarning(
                "Video Promosi Bawaan",
                "Video promosi bawaan tidak dapat dipilih/diganti.",
                parent=self.winfo_toplevel())
            return
        app = self.winfo_toplevel()
        hub = getattr(app, 'tv_ws_hub', None)
        if kategori == "video":
            self._kirim_media_video(app, hub, ms, path)
        else:
            self._kirim_media_image(app, hub, ms, path)

    def _kirim_media_image(self, app, hub, ms, path):
        try:
            filename = ms.simpan_file(path)
            ms.set_current("image", filename)
            # quote(): nama file ber-spasi/karakter non-ASCII aman untuk
            # semua versi Android.
            url = f"http://{app._get_lan_ip()}:{ms.port}/media/{quote(filename)}"
            if hub:
                hub.send_show_media(self.label_tv, "image", url)
            messagebox.showinfo("✅ Media Terkirim",
                                f"TV: {self.label_tv}\n{os.path.basename(path)}\n"
                                f"(gambar)",
                                parent=self.winfo_toplevel())
        except Exception as e:
            messagebox.showerror("Gagal Kirim Media", str(e),
                                 parent=self.winfo_toplevel())

    def _kirim_media_video(self, app, hub, ms, path):
        """Video diproses di background: normalisasi format (ffmpeg) supaya
        jalan di SEMUA Android TV, lalu dikirim ke TV.

        Video yang sudah H.264+faststart dikirim langsung; lainnya di-remux/
        transcode ke MP4 H.264 Main L4.0 + AAC + faststart (format yang didukung
        semua merk/versi Android, termasuk box lama Android 11 ke bawah)."""
        parent_win = self.winfo_toplevel()
        if not ffmpeg_path():
            tanya = messagebox.askyesno(
                "ffmpeg Belum Terpasang",
                "Untuk menjamin video promo bisa diputar di SEMUA TV Android, "
                "video perlu dinormalisasi dengan ffmpeg (±87 MB, diunduh sekali).\n\n"
                "Unduh sekarang? (tanpa ffmpeg, video tetap dikirim apa adanya — "
                "bisa gagal diputar di TV Android lama)",
                parent=parent_win)
            if not tanya:
                self._kirim_media_image_raw(app, hub, ms, path, "video")
                return
            try:
                ok = self._unduh_ffmpeg_dialog(app)
            except Exception as e:
                ok = False
                self._kirim_media_image_raw(app, hub, ms, path, "video",
                                            f"ffmpeg gagal diunduh: {e}")
            if not ok:
                return
        # ── Window progress ───────────────────────────────────────────────────
        win = tk.Toplevel(parent_win)
        win.title("Siapkan Video Promosi")
        win.geometry("440x170")
        win.transient(app)
        win.resizable(False, False)
        tk.Label(win, text="Menganalisis & menyiapkan video...").pack(pady=(14, 6))
        bar = ttk.Progressbar(win, mode="indeterminate", length=380)
        bar.pack(padx=20, pady=4)
        lbl_status = tk.Label(win, text="", font=("Segoe UI", 11))
        lbl_status.pack(pady=2)
        state = {"batal": False, "mulai": time.time()}

        def on_batal():
            state["batal"] = True
        btn_batal = tk.Button(win, text="Batal", command=on_batal)
        btn_batal.pack(pady=8)
        bar.start(12)

        def update_status(progress: float) -> None:
            # progress = detik video yang sudah diproses
            try:
                m = int(progress // 60)
                s = int(progress % 60)
                lbl_status.config(text=f"Mengonversi video... {m}m {s:02d}s")
            except Exception:
                pass

        def selesai(ok, pesan):
            try:
                win.destroy()
            except Exception:
                pass
            if ok:
                messagebox.showinfo("✅ Video Terkirim", pesan,
                                    parent=self.winfo_toplevel())
            else:
                messagebox.showerror("Video Gagal", pesan,
                                     parent=self.winfo_toplevel())

        def worker():
            tmp_out = os.path.join(
                tempfile.gettempdir(),
                f"rr_promo_{int(time.time() * 1000)}.mp4")
            try:
                action = prepare_video(
                    path, tmp_out,
                    progress_cb=update_status,
                    cancel=lambda: state["batal"])
                filename = ms.simpan_file(tmp_out)
                ms.set_current("video", filename)
                url = f"http://{app._get_lan_ip()}:{ms.port}/media/{quote(filename)}"
                if hub:
                    hub.send_show_media(self.label_tv, "video", url)
                ket = {"copy": "langsung (format sudah kompatibel)",
                       "remux": "dirapikan (faststart)",
                       "transcode": "dikonversi ke H.264 + faststart"}.get(action, action)
                app.after(0, lambda: selesai(
                    True, f"TV: {self.label_tv}\n{os.path.basename(path)}\n"
                          f"→ {os.path.basename(filename)}\n({ket})"))
            except Exception as e:
                app.after(0, lambda msg=str(e): selesai(False, msg))
            finally:
                try:
                    if os.path.isfile(tmp_out):
                        os.remove(tmp_out)
                except Exception:
                    pass

        threading.Thread(target=worker, daemon=True).start()

    def _kirim_media_image_raw(self, app, hub, ms, path, media_type, info=""):
        """Fallback: kirim file apa adanya (tanpa normalisasi)."""
        try:
            filename = ms.simpan_file(path)
            ms.set_current(media_type, filename)
            url = f"http://{app._get_lan_ip()}:{ms.port}/media/{quote(filename)}"
            if hub:
                hub.send_show_media(self.label_tv, media_type, url)
            pesan = (f"TV: {self.label_tv}\n{os.path.basename(path)}\n"
                     f"({media_type} — dikirim apa adanya)")
            if info:
                pesan += f"\n\n{info}"
            messagebox.showinfo("⚠ Media Terkirim (tanpa normalisasi)", pesan,
                                parent=self.winfo_toplevel())
        except Exception as e:
            messagebox.showerror("Gagal Kirim Media", str(e),
                                 parent=self.winfo_toplevel())

    def _unduh_ffmpeg_dialog(self, app) -> bool:
        """Unduh ffmpeg.exe sekali (dari GitHub release) dengan progress bar."""
        import urllib.request

        dest = ffmpeg_target_path()
        os.makedirs(os.path.dirname(dest), exist_ok=True)
        win = tk.Toplevel(app)
        win.title("Unduh ffmpeg (±87 MB)")
        win.geometry("460x150")
        win.transient(app)
        win.resizable(False, False)
        tk.Label(win, text="Mengunduh ffmpeg untuk normalisasi video...").pack(pady=(14, 6))
        bar = ttk.Progressbar(win, mode="determinate", length=400, maximum=100)
        bar.pack(padx=20, pady=4)
        lbl = tk.Label(win, text="0%", font=("Segoe UI", 11))
        lbl.pack()
        win.update_idletasks()

        hasil = {}

        def reporthook(count, block_size, total_size):
            if total_size and total_size > 0:
                pct = min(100.0, count * block_size * 100.0 / total_size)
                try:
                    bar["value"] = pct
                    lbl.config(text=f"{int(pct)}%  ({count * block_size / 1048576:.0f} MB)")
                except Exception:
                    pass

        def worker():
            part = dest + ".part"
            try:
                urllib.request.urlretrieve(FFMPEG_URL, part, reporthook=reporthook)
                os.replace(part, dest)
                # verifikasi bisa dijalankan
                rc = subprocess.run([dest, "-version"],
                                    stdout=subprocess.DEVNULL,
                                    stderr=subprocess.DEVNULL).returncode
                hasil["ok"] = (rc == 0)
                hasil["pesan"] = "" if rc == 0 else "ffmpeg tidak dapat dijalankan"
            except Exception as e:
                hasil["ok"] = False
                hasil["pesan"] = str(e)
            finally:
                try:
                    if os.path.isfile(part):
                        os.remove(part)
                except Exception:
                    pass
                app.after(0, finish)

        def finish():
            try:
                win.destroy()
            except Exception:
                pass
            if hasil.get("ok"):
                messagebox.showinfo("✅ ffmpeg Siap",
                                    "ffmpeg berhasil diunduh.\nVideo promosi "
                                    "sekarang akan dinormalisasi otomatis.",
                                    parent=app)
            else:
                messagebox.showerror("Unduh ffmpeg Gagal",
                                     hasil.get("pesan", "Coba lagi nanti."),
                                     parent=app)

        threading.Thread(target=worker, daemon=True).start()
        try:
            win.wait_window()  # tunggu unduhan selesai (thread menutup window)
        except Exception:
            pass
        return bool(hasil.get("ok"))

    def _buka_media_video(self):
        self._pilih_media(
            "video",
            [("Video", "*.mp4 *.webm *.3gp *.ts *.mkv *.mov *.avi"),
             ("Semua", "*.*")],
            "Pilih video promosi")

    def _buka_media_gambar(self):
        self._pilih_media("image", [("Gambar", "*.jpg *.jpeg *.png *.gif *.webp"),
                                    ("Semua", "*.*")],
                          "Pilih gambar promosi")

    def _ganti_logo_kartu(self):
        """Ganti logo lock screen dari kartu TV (global untuk semua TV).

        Koneksi sama seperti tombol VIDEO/GAMBAR — tanpa ADB/port 5555: file
        disalin ke media_promo/logo_lock.png via TvMediaServer (HTTP 8082);
        client TV memakainya saat LOCK_SCREEN (waktu sewa habis).
        Hanya tersedia untuk lisensi LIFETIME."""
        app = self.winfo_toplevel()
        if not app._lisensi_lifetime():
            messagebox.showwarning(
                "Fitur Lisensi LIFETIME",
                "Fitur logo kartu TV hanya tersedia untuk lisensi LIFETIME.\n\n"
                "Logo tetap memakai RR BILLING PRO.\n"
                "Lakukan aktivasi lisensi LIFETIME untuk mengganti logo.",
                parent=app)
            return
        ms = self._get_media_server()
        if not ms or not ms.running:
            messagebox.showwarning("Media Server Mati",
                                   "Server media (port 8082) tidak berjalan.\n"
                                   "Cek konfigurasi tv_ws_enabled di rr_billing_config.json.",
                                   parent=app)
            return
        path = filedialog.askopenfilename(
            parent=app, title="Pilih logo lock screen (PNG/JPG)",
            initialdir=os.path.join(APP_BASE_DIR, "media_promo"),
            filetypes=[("Gambar", "*.png *.jpg *.jpeg"), ("All files", "*.*")])
        if not path:
            return
        try:
            app._simpan_logo_lock(path)
        except Exception as e:
            messagebox.showerror("Logo Lock", f"Gagal menyalin logo:\n{e}", parent=app)
            return
        # Beri tahu semua TV yang terhubung: refresh logo lockscreen sekarang.
        try:
            logo_url = app._tv_logo_url()
            hub = getattr(app, 'tv_ws_hub', None)
            if hub and logo_url:
                n = hub.broadcast_update_logo(logo_url)
                print(f"[TV LOGO] broadcast UPDATE_LOGO ke {n} TV")
        except Exception as e:
            print(f"[TV LOGO] gagal broadcast: {e}")
        messagebox.showinfo("✅ Logo Lock",
                            "Logo lock berhasil diganti.\n"
                            "Berlaku untuk semua TV saat waktu sewa habis.",
                            parent=app)

    def _buka_qr_kartu(self):
        """Buka dialog QR panggil kasir untuk TV ini (QR unik per TV)."""
        app = self.winfo_toplevel()
        kode = app._qr_generate_untuk(self.label_tv)
        if not kode:
            messagebox.showwarning(
                "QR Gagal Dibuat",
                "Gagal membuat kode QR untuk TV ini.\nPeriksa file rr_billing_config.json.",
                parent=app)
            return
        url = app._qr_url(self.label_tv, kode, self.nama_grup)
        path = app._qr_simpan_png(self.label_tv, url)
        DialogQrKartu(app, self.label_tv, url, path)

    def _refresh_client_badge(self):
        try:
            if not self.winfo_exists():
                return
            app = self.winfo_toplevel()
            hub = getattr(app, 'tv_ws_hub', None)
            online = bool(hub and hub.is_meja_connected(self.label_tv))
            badge_id = self._ids.get('cli_badge')
            if badge_id:
                self.itemconfig(badge_id,
                                text="CLI\u2713" if online else "CLI\u2717",
                                fill=C_GREEN if online else C_MUTED)
            # Status HIDUP/MATI: prioritas status layar dari APK (WebSocket) —
            # paling akurat. Fallback ADB; kalau belum ada sumber → "?" abu-abu
            # (bukan label basi).
            ws_state = hub.get_screen_state(self.label_tv) if hub else None
            if ws_state is True:
                self._apply_tv_power_state(True)
            elif ws_state is False:
                self._apply_tv_power_state(False)
            elif not self._power_state_known:
                self.itemconfig(self._ids['lbl_power'],
                                text="\U0001f4fa ?", fill=C_MUTED)
                self.itemconfig(self._ids['btn_power'], fill=C_BTN)
            self.after(10000, self._refresh_client_badge)
        except Exception:
            pass

    def _install_apk_client(self, apk_path):
        if not apk_path or not os.path.isfile(apk_path):
            messagebox.showerror("APK Tidak Ditemukan",
                                 "File APK tidak ditemukan.",
                                 parent=self.winfo_toplevel())
            return
        threading.Thread(target=self._install_apk_thread,
                         args=(apk_path,), daemon=True).start()

    def _install_apk_thread(self, apk_path):
        ok, pesan = ADBHelper.adb_install(self.ip, apk_path)
        self.after(0, lambda: messagebox.showinfo(
            "✅ Install APK" if ok else "❌ Install Gagal",
            f"TV: {self.label_tv}\n{pesan}",
            parent=self.winfo_toplevel()))

    def _pilih_paket(self):
        if self.is_bebas:
            return
        DialogPaket(self.winfo_toplevel(), self.label_tv, self._on_paket_confirm,
                    self.get_paket_data(), self.get_makanan_data(), self.get_minuman_data(),
                    nama_grup=self.nama_grup)

    def _buka_tambah_pesanan(self):
        """Buka dialog untuk tambah pesanan makanan/minuman saat sesi berjalan."""
        if not self.paket_aktif:
            messagebox.showwarning("Tidak Ada Sesi", "Mulai sesi terlebih dahulu untuk memesan.")
            return
        
        # Buka DialogTambahPesanan untuk add order
        DialogTambahPesanan(self.winfo_toplevel(), self.label_tv, 
                           self._on_tambah_pesanan_confirm,
                           self.get_makanan_data(), self.get_minuman_data(),
                           pesanan_aktif=self.pesanan_aktif.copy(),
                           paket_harga=self.paket_harga_tetap, paket_label=self.paket_aktif)

    def _on_tambah_pesanan_confirm(self, pesanan_baru, paid=True, _stok_delta=None):
        """Callback saat user confirm tambah pesanan."""
        app = self.winfo_toplevel()
        # Validasi stok sebelum pesanan diterima (blokir kalau tidak cukup)
        if hasattr(app, '_stok_validate_orders'):
            stok_check = pesanan_baru if _stok_delta is None else {
                k: v for k, v in _stok_delta.items() if int(v or 0) > 0}
            ok, pesan, nm, sisa = app._stok_validate_orders(stok_check)
            if not ok:
                messagebox.showwarning("⚠ Stok Tidak Mencukupi", pesan, parent=app)
                return
        old_pesanan = dict(getattr(self, 'pesanan_aktif', {}) or {})
        # Merge pesanan baru ke pesanan_aktif
        all_menu = {**self.get_makanan_data(), **self.get_minuman_data()}
        total_baru = 0
        
        for nama, qty in pesanan_baru.items():
            self.pesanan_aktif[nama] = qty
            self.lunas_pesanan[nama] = paid  # status per item sesuai popup
        
        # Hitung ulang total berdasarkan semua pesanan aktif
        total_baru = sum(all_menu.get(nama, 0) * qty for nama, qty in self.pesanan_aktif.items())
        subtotal_paket = self.paket_harga_tetap
        total_semua = self._total_setelah_diskon(subtotal_paket + total_baru) if not self.is_bebas else total_baru
        self.biaya_pesanan = total_baru
        
        # Update display
        if self.is_bebas:
            self.itemconfig(self._ids['lbl_paket'], text=f"Main Bebas \U0001f579 +Pesanan {fmt_rp(total_baru)}")
        else:
            self.itemconfig(self._ids['lbl_paket'], text=f"{self.paket_aktif} | {fmt_rp(total_semua)}")
        self._ws_send_total(total_semua if not self.is_bebas else total_baru)
        self._update_paid_badge()
        self._update_bayar_buttons()

        # Update recorded transaction row if this fixed package session already has one.
        if not self.is_bebas and (getattr(self, '_last_transaction_item', None) or getattr(self, '_last_riwayat_idx', None) is not None):
            app = self.winfo_toplevel()
            idx = app._resolve_session_idx(self)
            if idx >= 0 and hasattr(app, 'tree') and hasattr(app, '_tree_item_to_index'):
                item_id = app._tree_item_for_idx(idx)
                if item_id is None:
                    item_id = self._last_transaction_item
                if idx < len(app.riwayat_transaksi):
                    waktu = app.riwayat_transaksi[idx][0] if len(app.riwayat_transaksi[idx]) > 0 else datetime.now().strftime("%Y-%m-%d %H:%M")
                    _bkid = app.riwayat_meta[idx].get("booking_id") if idx < len(app.riwayat_meta) else None
                    updated_row = app._format_riwayat_row(waktu, self.label_tv, self.paket_aktif, self.pesanan_aktif, total_semua, paid=self.paid, booking_id=_bkid)
                    app.riwayat_transaksi[idx] = updated_row
                    if item_id is not None and app.tree.exists(item_id):
                        app.tree.item(item_id, values=updated_row)
                    try:
                        all_menu_app = {**app.menu_makanan, **app.menu_minuman}
                        pesanan_total_baru = sum(all_menu_app.get(nm, 0) * qty for nm, qty in self.pesanan_aktif.items())
                        paket_harga_baru = total_semua - pesanan_total_baru
                        if paket_harga_baru < 0:
                            paket_harga_baru = 0
                        if idx < len(app.riwayat_meta):
                            app.riwayat_meta[idx]['paket_harga'] = paket_harga_baru
                            app.riwayat_meta[idx]['pesanan_total'] = pesanan_total_baru
                            app.riwayat_meta[idx]['total'] = total_semua
                            app.riwayat_meta[idx]['pesanan'] = {
                                str(k): int(v) for k, v in self.pesanan_aktif.items()}
                            threading.Thread(
                                target=app._upsert_tx_cloud_from_index, args=(idx,), daemon=True).start()
                    except Exception:
                        pass
                    if hasattr(app, '_refresh_riwayat_summary'):
                        app._refresh_riwayat_summary()
                    if hasattr(app, '_save_riwayat'):
                        app._save_riwayat()

        # Terapkan perubahan stok (selisih qty lama → baru; QR pakai delta eksplisit)
        try:
            if hasattr(app, '_stok_terapkan'):
                if _stok_delta is not None:
                    app._stok_terapkan(_stok_delta)
                else:
                    delta = {}
                    for nama, qty in pesanan_baru.items():
                        d = int(qty or 0) - int(old_pesanan.get(nama, 0) or 0)
                        if d:
                            delta[nama] = d
                    app._stok_terapkan(delta)
        except Exception:
            pass

    def _on_paket_confirm(self, paket_nm, paket_harga, paket_menit, pesanan, total_pesanan, diskoni=0, diskoni_mode="nominal", paid=True, booking=None):
        app = self.winfo_toplevel()
        # Validasi stok makanan/minuman sebelum paket+pesanan diterima
        if pesanan and hasattr(app, '_stok_validate_orders'):
            ok, pesan, nm, sisa = app._stok_validate_orders(pesanan)
            if not ok:
                messagebox.showwarning("⚠ Stok Tidak Mencukupi", pesan, parent=app)
                return
        previous_session = not self.sesi_kosong()
        old_pesanan = dict(getattr(self, 'pesanan_aktif', {}) or {})
        self.menit_dipakai_awal = 0
        self.diskoni = diskoni
        self.diskoni_mode = diskoni_mode

        # Status pembayaran sesi DIAMBIL DARI DOKUMEN BOOKING (bukan input
        # manual): lunas/lunas_transfer → LUNAS penuh; dp → bayar sejumlah DP
        # + tagihan sisa; biasa → mengikuti pilihan kasir (DialogKonfirmasiBayar).
        if booking is not None:
            try:
                _b_metode = str(booking.get("metode", "") or "")
                _b_sb = str(booking.get("statusBayar", "") or "")
                if (_b_metode == "lunas") or (_b_sb == "lunas_transfer"):
                    paid = True
                else:
                    paid = False
            except Exception:
                pass

        if not previous_session:
            # Sesi baru: reset pesanan dan biaya
            self.daftar_paket_sesi = [paket_nm]
            self.harga_paket_sesi = [paket_harga]
            self.lunas_paket = [paid]
            self.pesanan_aktif = pesanan
            all_menu = {**self.get_makanan_data(), **self.get_minuman_data()}
            self.lunas_pesanan = {nm: paid for nm in pesanan}
            self.biaya_pesanan = sum(all_menu.get(nm, 0) * qty for nm, qty in pesanan.items())
            # DP booking: nominal yang sudah dibayar pelanggan via transfer
            self.dp_bayar_awal = 0
            if booking is not None:
                try:
                    self.dp_bayar_awal = int(booking.get("nominalDp", 0) or 0)
                except Exception:
                    self.dp_bayar_awal = 0
        else:
            # Tambah paket ke sesi aktif: pertahankan pesanan lama, merge pesanan baru (jika ada)
            self.daftar_paket_sesi.append(paket_nm)
            self.harga_paket_sesi.append(paket_harga)
            self.lunas_paket.append(paid)
            for nm, qty in pesanan.items():
                self.pesanan_aktif[nm] = qty
                self.lunas_pesanan[nm] = paid
            if pesanan:
                all_menu = {**self.get_makanan_data(), **self.get_minuman_data()}
                self.biaya_pesanan += sum(all_menu.get(nm, 0) * qty for nm, qty in pesanan.items())
            # biaya_pesanan yang sudah ada dari SHOP tetap terjaga

        if paket_nm == "Main Bebas":
            self.is_bebas    = True
            self.sisa_waktu  = 0
            self.waktu_mulai = datetime.now()
            self.paket_harga_tetap = 0
            self.paket_aktif = paket_nm
            self.itemconfig(self._ids['lbl_paket'], text="Main Bebas \U0001f579 (berjalan)", fill=C_GREEN)
            self.itemconfig(self._ids['lbl_timer'], fill=C_YELLOW)
            self.itemconfig(self._ids['lbl_estimasi'], text=f"Total berjalan: {fmt_rp(self.biaya_pesanan)}", fill=C_YELLOW)
        else:
            self.is_bebas    = False
            self.paket_aktif = paket_nm
            if previous_session and self.paket_harga_tetap:
                self.sisa_waktu += paket_menit * 60
                self.paket_harga_tetap += paket_harga
            else:
                self.sisa_waktu  = paket_menit * 60
                self.paket_harga_tetap = paket_harga
            self.waktu_mulai = datetime.now()
            self.itemconfig(self._ids['lbl_paket'], text=f"{paket_nm} | {fmt_rp(self._total_setelah_diskon())}", fill="black")
            self.itemconfig(self._ids['lbl_timer'], fill=C_ACCENT)

        if self._timer_job:
            self.after_cancel(self._timer_job)

        self._billing_paused = False
        self._enable_btn("selesai", C_BTN, C_RED)
        self._enable_btn("shop", "black", "white")
        self._enable_btn("pause", "black", "white")
        self.itemconfig(self._ids['btn_pause_txt'], text="PAUSE")
        self._card_bg_color = "white"
        self.itemconfig(self._ids['bg'], fill="white")
        self._warning_blink_on = False

        if not self.is_bebas and self.sisa_waktu > 0:
            h, rem = divmod(self.sisa_waktu, 3600)
            m, s = divmod(rem, 60)
            self.itemconfig(self._ids['lbl_timer'], text=f"{h:02d}:{m:02d}:{s:02d}")

        if self.is_bebas:
            self._tick_bebas()
            self._disable_btn("paket")
        elif self.sisa_waktu > 0:
            self._tick_waktu()
            self._enable_btn("paket", "black", "white")
        else:
            self.itemconfig(self._ids['lbl_timer'], text="\u221e BEBAS", fill=C_GREEN)
        self._ws_send_total(self._total_setelah_diskon())

        if not self.is_bebas:
            if (self._last_transaction_item or self._last_riwayat_idx is not None) and previous_session:
                app = self.winfo_toplevel()
                idx = app._resolve_session_idx(self)
                total_int = self._total_setelah_diskon()
                if idx >= 0 and hasattr(app, 'tree') and hasattr(app, '_tree_item_to_index'):
                    item_id = app._tree_item_for_idx(idx)
                    if item_id is None:
                        item_id = self._last_transaction_item
                    if idx < len(app.riwayat_transaksi):
                        waktu = app.riwayat_transaksi[idx][0] if len(app.riwayat_transaksi[idx]) > 0 else datetime.now().strftime("%Y-%m-%d %H:%M")
                        _bkid2 = app.riwayat_meta[idx].get("booking_id") if idx < len(app.riwayat_meta) else None
                        updated_row = app._format_riwayat_row(waktu, self.label_tv, self.paket_aktif, self.pesanan_aktif, total_int, paid=self.paid, booking_id=_bkid2)
                        app.riwayat_transaksi[idx] = updated_row
                        if item_id is not None and app.tree.exists(item_id):
                            app.tree.item(item_id, values=updated_row)
                        try:
                            all_menu = {**app.menu_makanan, **app.menu_minuman}
                            pesanan_total_baru = sum(all_menu.get(nm, 0) * qty for nm, qty in self.pesanan_aktif.items())
                            paket_harga_baru = total_int - pesanan_total_baru
                            if paket_harga_baru < 0:
                                paket_harga_baru = 0
                            if idx < len(app.riwayat_meta):
                                app.riwayat_meta[idx]['paket_harga'] = paket_harga_baru
                                app.riwayat_meta[idx]['pesanan_total'] = pesanan_total_baru
                                app.riwayat_meta[idx]['total'] = total_int
                                app.riwayat_meta[idx]['diskoni'] = self.diskoni
                                app.riwayat_meta[idx]['diskoni_mode'] = self.diskoni_mode
                                app.riwayat_meta[idx]['pesanan'] = {
                                    str(k): int(v) for k, v in self.pesanan_aktif.items()}
                                threading.Thread(
                                    target=app._upsert_tx_cloud_from_index, args=(idx,), daemon=True).start()
                        except Exception:
                            pass
                    if hasattr(app, '_refresh_riwayat_summary'):
                        app._refresh_riwayat_summary()
                    if hasattr(app, '_save_riwayat'):
                        app._save_riwayat()
            else:
                total_int_baru = self._total_setelah_diskon(self.paket_harga_tetap + total_pesanan)
                if booking is not None:
                    # Mulai dari booking: link ke baris riwayat yang sudah dicatat
                    # saat konfirmasi (tanpa membuat baris baru / dobel bayar).
                    idx_b = -1
                    try:
                        idx_b = app._riwayat_idx_by_booking(str(booking.get("_id", "") or ""))
                    except Exception:
                        idx_b = -1
                    if idx_b >= 0:
                        self._last_riwayat_idx = idx_b
                        self._last_transaction_item = app._tree_item_for_idx(idx_b)
                        try:
                            self._last_cloud_id = app.riwayat_meta[idx_b].get("cloud_id")
                        except Exception:
                            self._last_cloud_id = None
                        if idx_b < len(app.riwayat_transaksi):
                            waktu_b = (app.riwayat_transaksi[idx_b][0]
                                       if len(app.riwayat_transaksi[idx_b]) > 0
                                       else datetime.now().strftime("%Y-%m-%d %H:%M"))
                            updated_row = app._format_riwayat_row(
                                waktu_b, self.label_tv, self.paket_aktif,
                                self.pesanan_aktif, total_int_baru, paid=paid,
                                booking_id=str(booking.get("_id", "") or ""))
                            app.riwayat_transaksi[idx_b] = updated_row
                            item_b = app._tree_item_for_idx(idx_b)
                            if item_b is not None and app.tree.exists(item_b):
                                app.tree.item(item_b, values=updated_row)
                            try:
                                bmenu = {**app.menu_makanan, **app.menu_minuman}
                                pesanan_total_b = sum(
                                    bmenu.get(nm, 0) * qty for nm, qty in self.pesanan_aktif.items())
                                paket_harga_b = max(0, total_int_baru - pesanan_total_b)
                                if idx_b < len(app.riwayat_meta):
                                    app.riwayat_meta[idx_b]['paket_harga'] = paket_harga_b
                                    app.riwayat_meta[idx_b]['pesanan_total'] = pesanan_total_b
                                    app.riwayat_meta[idx_b]['total'] = total_int_baru
                                    app.riwayat_meta[idx_b]['diskoni'] = self.diskoni
                                    app.riwayat_meta[idx_b]['diskoni_mode'] = self.diskoni_mode
                                    app.riwayat_meta[idx_b]['pesanan'] = {
                                        str(k): int(v) for k, v in self.pesanan_aktif.items()}
                                    threading.Thread(
                                        target=app._upsert_tx_cloud_from_index,
                                        args=(idx_b,), daemon=True).start()
                            except Exception:
                                pass
                            if hasattr(app, '_refresh_riwayat_summary'):
                                app._refresh_riwayat_summary()
                            if hasattr(app, '_save_riwayat'):
                                app._save_riwayat()
                    else:
                        # Baris konfirmasi tidak ada (mis. riwayat dibersihkan) — catat baru
                        self._last_transaction_item = self.on_transaksi(
                            self.label_tv, paket_nm, pesanan, total_int_baru, paid=paid,
                            booking_meta={
                                "booking_id": str(booking.get("_id", "") or ""),
                                "kode": str(booking.get("_id", "") or "")[:8].upper(),
                                "metode": str(booking.get("metode", "") or ""),
                                "status_bayar": str(booking.get("statusBayar", "") or ""),
                                "pelanggan": str(booking.get("namaPelanggan", "") or ""),
                                "no_hp": str(booking.get("noHp", "") or ""),
                            })
                        self._bind_last_transaction()
                else:
                    self._last_transaction_item = self.on_transaksi(
                        self.label_tv, paket_nm, pesanan, total_int_baru, paid=paid)
                    self._bind_last_transaction()
        else:
            self._last_transaction_item = None
            self._last_riwayat_idx = None
            self._last_cloud_id = None

        # Status pembayaran: hanya untuk paket berwaktu (bukan Main Bebas)
        if self.is_bebas:
            self.paid = True
            self.lunas_paket = [True] * (len(self.daftar_paket_sesi or []) or 1)
            self.lunas_pesanan = {nm: True for nm in self.pesanan_aktif}
        else:
            if not previous_session:
                self.paid = paid   # status bayar sesuai pilihan kasir di popup
            self._update_bayar_buttons()
        self._update_paid_badge()

        if not self.is_bebas and not previous_session:
            self._ws_send_total(self._total_setelah_diskon())

        app = self.winfo_toplevel()
        if hasattr(app, '_refresh_dashboard_total_pesanan'):
            app._refresh_dashboard_total_pesanan()

        # Kirim sinyal ke overlay Android TV (START_TIMER / RESUME update)
        if self.is_bebas:
            self._ws_send_start(-1)
        else:
            self._ws_send_start(self.sisa_waktu)

        # Terapkan perubahan stok (selisih qty lama → baru)
        try:
            if pesanan and hasattr(app, '_stok_terapkan'):
                delta = {}
                for nm, qty in pesanan.items():
                    d = int(qty or 0) - int(old_pesanan.get(nm, 0) or 0)
                    if d:
                        delta[nm] = d
                app._stok_terapkan(delta)
        except Exception:
            pass

        # Tandai booking online selesai dipakai (tidak muncul lagi di daftar)
        if booking is not None:
            try:
                bid_m = str(booking.get("_id", "") or "")
                if bid_m:
                    def _tandai_dimulai():
                        try:
                            from firestore_sync import get_firestore_client
                            get_firestore_client().set_document(
                                f"bookings/{bid_m}",
                                {"sesiDimulai": True,
                                 "sesiDimulaiAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
                                merge=True)
                        except Exception:
                            pass
                    threading.Thread(target=_tandai_dimulai, daemon=True).start()
            except Exception:
                pass

        # Popup cetak struk setelah sesi baru dimulai (konfirmasi pembayaran)
        if not previous_session:
            try:
                self._buka_print_struk(self._trans_struk())
            except Exception:
                pass

    # ── Timer paket berwaktu (mundur) ───────────────────────────────────────
    # Decrement sisa_waktu kini dikelola TimerService (thread background);
    # _tick_waktu hanya refresh tampilan. Waktu habis ditangani _timer_habis.
    def _tick_waktu(self):
        if self._timer_paused:
            return
        self._update_timer_display()

    # ── Timer Main Bebas (maju, hitung estimasi biaya berjalan) ─────────────
    def _tick_bebas(self):
        if self._timer_paused or not self.is_bebas or self.waktu_mulai is None:
            return
        total_detik = self.menit_dipakai_awal * 60 + int((datetime.now() - self.waktu_mulai).total_seconds())
        h, rem = divmod(total_detik, 3600)
        m, s   = divmod(rem, 60)
        self.itemconfig(self._ids['lbl_timer'], text=f"{h:02d}:{m:02d}:{s:02d}", fill="black")

        tarif_menit = hitung_tarif_per_menit(self.get_paket_data())
        menit_berjalan = total_detik / 60
        biaya_waktu_berjalan = tarif_menit * menit_berjalan
        total_berjalan = self._total_setelah_diskon(biaya_waktu_berjalan + self.biaya_pesanan)
        self.itemconfig(self._ids['lbl_estimasi'],
            text=f"Total berjalan: {fmt_rp(total_berjalan)}",
            fill="black"
        )
        self._ws_send_total(total_berjalan)

        if not self._timer_paused:
            self._timer_job = self.after(1000, self._tick_bebas)

    # ── Update timer display (called from TimerService) ─────────────────────
    def _update_timer_display(self):
        if self._timer_paused:
            return
        if self.is_bebas:
            total_detik = self.menit_dipakai_awal * 60 + int((datetime.now() - self.waktu_mulai).total_seconds()) if self.waktu_mulai else 0
            h, rem = divmod(total_detik, 3600)
            m, s   = divmod(rem, 60)
            self.itemconfig(self._ids['lbl_timer'], text=f"{h:02d}:{m:02d}:{s:02d}", fill="black")
            tarif_menit = hitung_tarif_per_menit(self.get_paket_data())
            menit_berjalan = total_detik / 60
            biaya_waktu_berjalan = tarif_menit * menit_berjalan
            total_berjalan = self._total_setelah_diskon(biaya_waktu_berjalan + self.biaya_pesanan)
            self.itemconfig(self._ids['lbl_estimasi'],
                text=f"Total berjalan: {fmt_rp(total_berjalan)}",
                fill="black"
            )
            self._ws_send_total(total_berjalan)
        else:
            h, rem = divmod(self.sisa_waktu, 3600)
            m, s   = divmod(rem, 60)
            self.itemconfig(self._ids['lbl_timer'], text=f"{h:02d}:{m:02d}:{s:02d}")
            if self.sisa_waktu <= 120:
                self._warning_blink_on = not self._warning_blink_on
                if self._warning_blink_on:
                    self.itemconfig(self._ids['bg'], fill=C_RED)
                    self.itemconfig(self._ids['lbl_timer'], fill="white")
                else:
                    self.itemconfig(self._ids['bg'], fill="white")
                    self.itemconfig(self._ids['lbl_timer'], fill=C_ACCENT2)
            else:
                self.itemconfig(self._ids['bg'], fill="white")
                self.itemconfig(self._ids['lbl_timer'], fill=C_ACCENT2)
            self._ws_send_sync()

    # ── Timer habis handler (called from TimerService) ──────────────────────
    def _timer_habis(self):
        if self.sesi_kosong():
            return
        self.itemconfig(self._ids['lbl_timer'], text="WAKTU HABIS \u23f9", fill=C_RED)
        total_akhir = self._total_setelah_diskon()
        pesanan_txt = ", ".join(f"{nm}\u00d7{qty}" for nm, qty in self.pesanan_aktif.items()) or "Tidak ada pesanan"
        paket_txt = f"{self.paket_aktif or '-'} ({fmt_rp(self.paket_harga_tetap)})"
        app = self.winfo_toplevel()
        idx = app._resolve_session_idx(self)
        if idx >= 0 and hasattr(app, 'tree') and hasattr(app, '_tree_item_to_index'):
            item_id = app._tree_item_for_idx(idx)
            if item_id is None:
                item_id = self._last_transaction_item
            if idx < len(app.riwayat_transaksi):
                waktu = app.riwayat_transaksi[idx][0] if len(app.riwayat_transaksi[idx]) > 0 else datetime.now().strftime("%Y-%m-%d %H:%M")
                _bkid3 = app.riwayat_meta[idx].get("booking_id") if idx < len(app.riwayat_meta) else None
                updated_row = app._format_riwayat_row(waktu, self.label_tv, self.paket_aktif, self.pesanan_aktif, total_akhir, paid=self.paid, booking_id=_bkid3)
                app.riwayat_transaksi[idx] = updated_row
                if item_id is not None and app.tree.exists(item_id):
                    app.tree.item(item_id, values=updated_row)
                try:
                    all_menu = {**app.menu_makanan, **app.menu_minuman}
                    pesanan_total = sum(all_menu.get(nm, 0) * qty for nm, qty in self.pesanan_aktif.items())
                    paket_harga = total_akhir - pesanan_total
                    if paket_harga < 0:
                        paket_harga = 0
                    app.riwayat_meta[idx]['paket_harga'] = paket_harga
                    app.riwayat_meta[idx]['pesanan_total'] = pesanan_total
                    app.riwayat_meta[idx]['total'] = total_akhir
                    app.riwayat_meta[idx]['diskoni'] = self.diskoni
                    app.riwayat_meta[idx]['diskoni_mode'] = self.diskoni_mode
                    app.riwayat_meta[idx]['pesanan'] = {
                        str(k): int(v) for k, v in self.pesanan_aktif.items()}
                    threading.Thread(
                        target=app._upsert_tx_cloud_from_index, args=(idx,), daemon=True).start()
                except Exception:
                    pass
                if hasattr(app, '_refresh_riwayat_summary'):
                    app._refresh_riwayat_summary()
                if hasattr(app, '_save_riwayat'):
                    app._save_riwayat()
        else:
            self.on_transaksi(self.label_tv, self.paket_aktif, self.pesanan_aktif, total_akhir,
                              diskoni=self.diskoni, diskoni_mode=self.diskoni_mode)
        # Kunci layar TV client (Android): tampilkan Lockscreen fullscreen
        # dengan rincian pesanan lengkap (Sewa / Makanan / Minuman / Total).
        # Dikirim SEGERA saat waktu habis (sebelum dialog blokir) supaya TV
        # langsung menampilkan semua detail pesanan.
        daftar_sewa = " + ".join(self.daftar_paket_sesi) or (self.paket_aktif or "-")
        app_menu = self.winfo_toplevel()
        menu_makanan = getattr(app_menu, 'menu_makanan', {}) or {}
        menu_minuman = getattr(app_menu, 'menu_minuman', {}) or {}
        logo_url, promo_url = self._tv_lock_media_urls(app_menu)
        lunas_now, tagihan_now = self._split_payment()
        sewa_lunas = all(self.lunas_paket) if getattr(self, "lunas_paket", None) else (
            True if getattr(self, "paid", True) else False)
        print(f"[TV TIMER] {self.label_tv}: WAKTU HABIS -> kirim LOCK_SCREEN "
              f"logo={logo_url!r} promo={promo_url!r}")
        self._ws_send_lock("WAKTU SEWA HABIS", {
            "meja": self.label_tv,
            "sewa": daftar_sewa,
            "sewa_harga": fmt_rp(self.paket_harga_tetap),
            "sewa_lunas": sewa_lunas,
            "lunas_total": fmt_rp(lunas_now),
            "tagihan_total": fmt_rp(tagihan_now),
            "makanan": [
                {"item": f"{qty}x {nm}", "harga": fmt_rp(menu_makanan.get(nm, 0) * qty),
                 "lunas": (getattr(self, "lunas_pesanan", {}).get(nm, getattr(self, "paid", True)))}
                for nm, qty in self.pesanan_aktif.items() if nm in menu_makanan
            ],
            "minuman": [
                {"item": f"{qty}x {nm}", "harga": fmt_rp(menu_minuman.get(nm, 0) * qty),
                 "lunas": (getattr(self, "lunas_pesanan", {}).get(nm, getattr(self, "paid", True)))}
                for nm, qty in self.pesanan_aktif.items() if nm in menu_minuman
            ],
            "fnb": fmt_rp(self.biaya_pesanan),
            "total": fmt_rp(total_akhir),
            "logo_url": logo_url,
            "promo_url": promo_url,
        })

        messagebox.showwarning(
            "\u23f0 Waktu TV Habis",
            f"TV: {self.label_tv}\n"
            f"Paket: {paket_txt}\n"
            f"Pesanan: {pesanan_txt} ({fmt_rp(self.biaya_pesanan)})\n"
            f"TOTAL: {fmt_rp(total_akhir)}",
            parent=self.winfo_toplevel(),
        )

        # Jika masih ada sisa TAGIHAN, kasir diminta konfirmasi:
        #  - BAYAR   -> sesi ditandai lunas (riwayat jadi LUNAS).
        #  - TAGIHAN -> input nama + no HP pelanggan untuk catatan.
        if tagihan_now > 0:
            DialogKonfirmasiBayar(
                self.winfo_toplevel(),
                lambda paid: self._akhir_sesi_setelah_konfirmasi(paid),
                judul="Selesaikan Tagihan",
                rincian=f"Tagihan belum terbayar: {fmt_rp(tagihan_now)}\n"
                        f"Sudah terbayar (LUNAS): {fmt_rp(lunas_now)}",
            ).lift()
        else:
            self._akhir_sesi_tanpa_tagih()

    def _akhir_sesi_setelah_konfirmasi(self, paid):
        """Lanjutan sesi habis setelah konfirmasi BAYAR / TAGIHAN:
        BAYAR -> semua item ditandai lunas (riwayat jadi LUNAS);
        TAGIHAN -> isi nama & no HP pelanggan untuk catatan tagihan.
        """
        if paid:
            self._set_paid(True)
            self._akhir_sesi_tanpa_tagih()
        else:
            DialogPelangganAkhir(
                self.winfo_toplevel(),
                lambda nama, hp: self._simpan_pelanggan_tagihan(nama, hp),
            ).lift()

    def _simpan_pelanggan_tagihan(self, nama, hp):
        """Catat nama & no HP pelanggan (kasus TAGIHAN) ke baris riwayat."""
        try:
            app = self.winfo_toplevel()
            idx = app._resolve_session_idx(self)
            if idx >= 0 and hasattr(app, '_tree_item_to_index'):
                if idx < len(app.riwayat_meta):
                    app.riwayat_meta[idx]['nama_pelanggan'] = nama
                    app.riwayat_meta[idx]['no_hp'] = hp
                    threading.Thread(
                        target=app._upsert_tx_cloud_from_index, args=(idx,), daemon=True).start()
                if hasattr(app, '_save_riwayat'):
                    app._save_riwayat()
                if hasattr(app, '_refresh_riwayat_summary'):
                    app._refresh_riwayat_summary()
                # Status riwayat tetap BELUM LUNAS (tagihan)
                if hasattr(app, '_set_transaksi_paid_idx'):
                    app._set_transaksi_paid_idx(idx, False)
        except Exception as e:
            print(f"[TAGIHAN] Gagal simpan pelanggan {self.label_tv}: {e}")
        self._akhir_sesi_tanpa_tagih()

    def _akhir_sesi_tanpa_tagih(self):
        """Tutup sesi: unlock TV, sleep 2 dtk, reset kartu."""
        print(f"[TV TIMER] {self.label_tv}: dialog kasir OK -> UNLOCK + sleep 2dtk")
        trans = self._trans_struk()
        self._ws_send_stop()
        self._tv_sleep_now(2, alasan="Waktu habis - kasir klik OK (sleep 2 dtk)")
        self._reset_sesi()
        self._buka_print_struk(trans)

    def _trans_struk(self):
        """Data transaksi untuk popup cetak struk (panggil SEBELUM sesi direset)."""
        try:
            menit = self._total_menit_terpakai() or 0
        except Exception:
            menit = self.menit_dipakai_awal
        return {
            "pc": getattr(self, "label_tv", None) or "-",
            "paket": self.paket_aktif or "Main Bebas",
            "menit": int(menit or 0),
            "harga": int(self._total_setelah_diskon() or 0),
        }

    def _buka_print_struk(self, trans):
        """Buka popup cetak struk di window utama (app._print_receipt)."""
        try:
            app = self.winfo_toplevel()
            if hasattr(app, "_print_receipt"):
                app._print_receipt(trans)
        except Exception as e:
            print(f"[CETAK] gagal buka popup struk {self.label_tv}: {e}")

    # ── Pause/Resume timer untuk virtual scroll ────────────────────────────
    def _pause_timer(self):
        self._timer_paused = True
        self._timer_was_running = self._timer_job is not None
        if self._timer_job:
            self.after_cancel(self._timer_job)
            self._timer_job = None

    def _resume_timer(self):
        self._timer_paused = False
        if self._timer_was_running:
            self._timer_was_running = False
            if self.is_bebas:
                self._tick_bebas()
            elif self.sisa_waktu > 0:
                self._tick_waktu()

    # ── Hitung total menit yang sudah dipakai sesi saat ini ─────────────────
    def _total_menit_terpakai(self):
        if self.is_bebas:
            if self.waktu_mulai is None:
                return self.menit_dipakai_awal
            detik_berjalan = (datetime.now() - self.waktu_mulai).total_seconds()
            return self.menit_dipakai_awal + detik_berjalan / 60
        return None  # tidak relevan untuk paket berwaktu

    def _total_setelah_diskon(self, subtotal=None):
        if subtotal is None:
            subtotal = self.paket_harga_tetap + self.biaya_pesanan
        if self.diskoni <= 0:
            return subtotal
        if self.diskoni_mode == "persen":
            diskon = subtotal * self.diskoni // 100
        else:
            diskon = self.diskoni
        return max(0, subtotal - diskon)

    def _klik_selesai(self):
        if self.sesi_kosong():
            return
        if self.is_bebas:
            menit_total = self._total_menit_terpakai()
            tarif_menit = hitung_tarif_per_menit(self.get_paket_data())
            biaya_waktu = tarif_menit * menit_total
            total_akhir = self._total_setelah_diskon(biaya_waktu + self.biaya_pesanan)
            self._last_transaction_item = self.on_transaksi(
                self.label_tv, "Main Bebas", self.pesanan_aktif, total_akhir,
                diskoni=self.diskoni, diskoni_mode=self.diskoni_mode, paid=False)
            self._bind_last_transaction()
            DialogPelangganAkhir(
                self.winfo_toplevel(),
                lambda nama, hp: self._simpan_pelanggan_tagihan(nama, hp),
            ).lift()
            return
        else:
            total_akhir = self._total_setelah_diskon()
            pesanan_txt = ", ".join(f"{nm}×{qty}" for nm, qty in self.pesanan_aktif.items()) or "Tidak ada pesanan"
            paket_txt = f"{self.paket_aktif} ({fmt_rp(self.paket_harga_tetap)})"
            if not messagebox.askyesno(
                    "⏹ Selesai",
                    f"TV: {self.label_tv}\n"
                    f"Paket: {paket_txt}\n"
                    f"Pesanan: {pesanan_txt} ({fmt_rp(self.biaya_pesanan)})\n"
                    f"TOTAL: {fmt_rp(total_akhir)}\n\n"
                    f"Akhiri sesi ini?"):
                return
            # Transaksi paket berwaktu sudah dicatat saat konfirmasi awal,
            # jadi di sini tidak dicatat ulang — cukup tutup sesi.
            lunas_now, tagihan_now = self._split_payment()
            if tagihan_now > 0:
                DialogKonfirmasiBayar(
                    self.winfo_toplevel(),
                    lambda paid: self._klik_selesai_tutup_sesi(paid),
                    judul="Selesaikan Tagihan",
                    rincian=f"Tagihan belum terbayar: {fmt_rp(tagihan_now)}\n"
                            f"Sudah terbayar (LUNAS): {fmt_rp(lunas_now)}",
                ).lift()
                return

        self.itemconfig(self._ids['lbl_timer'], text="SELESAI \u23f9", fill=C_MUTED)
        self.itemconfig(self._ids['lbl_estimasi'], text="")
        
        # Sembunyikan overlay & tutup lock screen di TV client (unlock dulu,
        # biar lockscreen tertutup sebelum TV masuk mode sleep).
        self._ws_send_stop()

        # TV masuk mode sleep SETELAH kasir klik OK/SELESAI
        self._tv_sleep_now(2, alasan="Kasir klik SELESAI (sleep 2 dtk)")

        self._reset_sesi()

    def _klik_selesai_tutup_sesi(self, paid):
        """Konfirmasi tagihan di tombol SELESAI: BAYAR -> riwayat lunas;
        TAGIHAN -> dialog data pelanggan, lalu tetap tutup sesi."""
        if paid:
            self._set_paid(True)
        else:
            DialogPelangganAkhir(
                self.winfo_toplevel(),
                lambda nama, hp: self._simpan_pelanggan_tagihan(nama, hp),
            ).lift()
        # bukan -> sudah disimpan di _simpan_pelanggan_tagihan
        if paid:
            self.itemconfig(self._ids['lbl_timer'], text="SELESAI \u23f9", fill=C_MUTED)
            self.itemconfig(self._ids['lbl_estimasi'], text="")
            trans = self._trans_struk()
            self._ws_send_stop()
            self._tv_sleep_now(2, alasan="Kasir klik SELESAI (sleep 2 dtk)")
            self._reset_sesi()
            self._buka_print_struk(trans)

    def _reset_sesi(self):
        if self._timer_job:
            self.after_cancel(self._timer_job)
            self._timer_job = None
        self.paket_aktif   = None
        self.sisa_waktu    = 0
        self.is_bebas      = False
        self.waktu_mulai   = None
        self.menit_dipakai_awal = 0
        self.pesanan_aktif = {}
        self.biaya_pesanan = 0
        self.paket_harga_tetap = 0
        self.daftar_paket_sesi = []
        self.lunas_paket = []
        self.harga_paket_sesi = []
        self.lunas_pesanan = {}
        self.diskoni       = 0
        self.diskoni_mode  = "nominal"
        self._billing_paused = False
        self.paid          = True
        self.itemconfig(self._ids['lbl_paket'], text="\u2014", fill=C_MUTED)
        self.itemconfig(self._ids['lbl_timer'], text="00:00:00", fill=C_MUTED)
        self.itemconfig(self._ids['bg'], fill=C_CARD)
        self._warning_blink_on = False
        self.itemconfig(self._ids['lbl_estimasi'], text="")
        self._disable_btn("selesai")
        self._disable_btn("shop")
        self._enable_btn("paket", "black", "white")
        if "pause" in self._ids:
            self._disable_btn("pause")
            self.itemconfig(self._ids['btn_pause_txt'], text="PAUSE")
        self._update_paid_badge()
        app = self.winfo_toplevel()
        if hasattr(app, '_refresh_dashboard_total_pesanan'):
            app._refresh_dashboard_total_pesanan()

    # ── Tombol PINDAH TV ────────────────────────────────────────────────────
    def _klik_pindah(self):
        if self.sesi_kosong():
            return
        semua = self.get_semua_kartu()
        kandidat = [k for k in semua if k is not self and k.sesi_kosong()]
        if not kandidat:
            messagebox.showwarning("⚠ Tidak Ada TV Kosong",
                                    "Semua TV lain sedang dipakai.\n"
                                    "Selesaikan dulu salah satu sesi, atau tambah TV baru.")
            return
        self._buka_dialog_pilih_tujuan(kandidat)

    def _buka_dialog_pilih_tujuan(self, kandidat):
        dlg = ctk.CTkToplevel(self.winfo_toplevel())
        dlg.title("↔ Pindah TV")
        dlg.geometry("360x420")
        dlg.configure(fg_color=C_BG)
        dlg.grab_set()
        dlg.resizable(False, False)

        ctk.CTkLabel(dlg, text=f"↔  Pindah dari {self.label_tv}",
                     font=FONT_TITLE, text_color=C_ACCENT).pack(pady=(18, 4))
        ctk.CTkLabel(dlg, text="Pilih TV tujuan — sisa waktu & sesi akan dipindah",
                     font=FONT_BODY, text_color=C_MUTED, wraplength=320).pack(pady=(0, 12))

        scroll = ctk.CTkScrollableFrame(dlg, fg_color=C_BG, height=280)
        scroll.pack(fill="both", expand=True, padx=16)

        for k in kandidat:
            btn = ctk.CTkButton(
                scroll, text=f"📺  {k.label_tv}   ({k.ip}:{k.port})",
                height=40, fg_color=C_CARD, hover_color=C_ACCENT2,
                border_width=1, border_color=C_BORDER,
                font=FONT_BODY, text_color=C_TEXT, anchor="w",
                command=lambda target=k: self._konfirmasi_pindah(target, dlg))
            btn.pack(fill="x", pady=4, padx=4)

        ctk.CTkButton(dlg, text="✖  Batal", height=34, width=120,
                      fg_color=C_RED, font=FONT_SUB, text_color="white",
                      command=dlg.destroy).pack(pady=12)

    def _konfirmasi_pindah(self, target, dlg):
        if not messagebox.askyesno(
                "↔ Konfirmasi Pindah",
                f"Pindahkan sesi dari {self.label_tv} ke {target.label_tv}?\n\n"
                f"Sisa waktu / lama main akan dilanjutkan di {target.label_tv},\n"
                f"dan {self.label_tv} akan menjadi kosong."):
            return

        # ── Salin status sesi ke TV tujuan ───────────────────────────────────
        target.paket_aktif   = self.paket_aktif
        target.is_bebas      = self.is_bebas
        target.pesanan_aktif = dict(self.pesanan_aktif)
        target.biaya_pesanan = self.biaya_pesanan
        target.paket_harga_tetap = self.paket_harga_tetap
        target.diskoni       = self.diskoni
        target.diskoni_mode  = self.diskoni_mode
        target.paid          = self.paid

        if self.is_bebas:
            target.menit_dipakai_awal = self._total_menit_terpakai()
            target.sisa_waktu  = 0
            target.waktu_mulai = datetime.now()
            target.itemconfig(target._ids['lbl_paket'], text="Main Bebas \U0001f579 (berjalan)", fill=C_GREEN)
            target.itemconfig(target._ids['lbl_timer'], fill=C_YELLOW)
            target.itemconfig(target._ids['lbl_estimasi'], text=f"Total berjalan: {fmt_rp(target.biaya_pesanan)}", fill=C_YELLOW)
        else:
            target.sisa_waktu  = self.sisa_waktu
            target.waktu_mulai = datetime.now()
            harga_tampil = self._total_setelah_diskon()
            target.itemconfig(target._ids['lbl_paket'], text=f"{self.paket_aktif} | {fmt_rp(harga_tampil)}", fill="black")
            target.itemconfig(target._ids['lbl_timer'], fill=C_ACCENT)

        if target._timer_job:
            target.after_cancel(target._timer_job)
        if target.is_bebas:
            target._tick_bebas()
            target._disable_btn("paket")
        elif target.sisa_waktu > 0:
            target._tick_waktu()
            target._enable_btn("paket", "black", "white")
        else:
            target.itemconfig(target._ids['lbl_timer'], text="\u221e BEBAS", fill=C_GREEN)
        target._enable_btn("selesai", C_BTN, C_RED)
        target._enable_btn("shop", "black", "white")
        target._update_paid_badge()

        # ── Kosongkan TV asal ────────────────────────────────────────────────
        self._reset_sesi()

        dlg.destroy()
        messagebox.showinfo("✅ Berhasil Pindah",
                            f"Sesi telah dipindah ke {target.label_tv}.")


# ═══════════════════════════════════════════════════════════════════════════════
#  KARTU WARNET
# ═══════════════════════════════════════════════════════════════════════════════
class KartuWarnet(tk.Canvas):
    def __init__(self, master, nomor, label_kursi, on_transaksi,
                 get_paket_data, get_makanan_data, get_minuman_data,
                get_semua_kartu=None, get_daftar_grup=None, on_ganti_grup=None,
                on_hapus=None, nama_grup=None, **kwargs):
        super().__init__(master, highlightthickness=0, bd=0,
                         bg="white", **kwargs)
        self.nomor             = nomor
        self.label_kursi       = label_kursi
        self.on_transaksi      = on_transaksi
        self.nama_grup         = nama_grup or NAMA_GRUP_DEFAULT
        self.get_paket_data    = get_paket_data or (lambda: {})
        self.get_makanan_data  = get_makanan_data
        self.get_minuman_data  = get_minuman_data
        self.get_semua_kartu   = get_semua_kartu or (lambda: [])
        self.get_daftar_grup   = get_daftar_grup or (lambda: [self.nama_grup])
        self.on_ganti_grup     = on_ganti_grup
        self.on_hapus          = on_hapus

        self.paket_aktif       = None
        self.sisa_waktu        = 0
        self.is_bebas          = False
        self.menit_dipakai_awal= 0
        self.waktu_mulai       = None
        self.pesanan_aktif     = {}
        self.biaya_pesanan     = 0
        self.paket_harga_tetap = 0
        self.daftar_paket_sesi = []
        self.lunas_paket       = []
        self.harga_paket_sesi  = []
        self.lunas_pesanan     = {}
        self.diskoni           = 0
        self.diskoni_mode      = "nominal"
        self._timer_job        = None
        self._timer_paused = False
        self._timer_was_running = False
        self._last_transaction_item = None
        self._last_riwayat_idx = None
        self._last_cloud_id = None
        self.paid             = True   # status pembayaran sesi (sinkron ke riwayat)
        self.is_on            = False
        self.pc_locked        = False   # LOCK terakhir yang dikirim ke client (persisten lintas reconnect)
        self._pc_lock_reason  = ""
        self._pc_lock_message = ""
        self._last_lock_requeue = 0
        self._warning_blink_on = False
        self._client_id       = None
        self._pc_id           = None
        self._pc_ip           = None
        self._ids = {}
        self._btn_states = {}
        self._card_w = 260
        self._card_bg_color = C_CARD

        self._build()
        self.bind("<Configure>", self._on_card_resize)
        self.after(2000, self._periodic_pc_status)

    def _build(self):
        if self._ids:
            self.delete("all")
            self._ids = {}
        self._card_bg_color = C_CARD
        self._build_inner()

    def _on_card_resize(self, event):
        if event.width > 50 and abs(event.width - self._card_w) > 5:
            self._card_w = event.width
            job = getattr(self, "_resize_job", None)
            if job:
                try:
                    self.after_cancel(job)
                except Exception:
                    pass
            self._resize_job = self.after(200, lambda: self.winfo_exists() and self._build())

    def _build_inner(self):
        W = self._card_w
        y = 2
        hdr_h, hdr_y = 38, y
        self._ids['hdr'] = self.create_rectangle(2, hdr_y, W-2, hdr_y+hdr_h,
            fill=C_ACCENT2, outline="", width=0, tags="hdr")

        # Kursi name
        self._ids['lbl_kursi'] = self.create_text(10, hdr_y+12,
            text=self.label_kursi, font=("Russo One", 11, "bold"),
            fill="white", anchor="w", tags="lbl_kursi")

        # Badge status pembayaran (sebelah nama)
        self._ids['paid_badge'] = self.create_text(10, hdr_y+12,
            text="", font=("Courier New", 10, "bold"),
            fill=C_GREEN, anchor="w", tags="paid_badge")

        # PC status text
        pc_connected = False
        app = self.winfo_toplevel() if self.winfo_exists() else None
        if hasattr(self, '_client_id') and self._client_id and app and hasattr(app, 'warnet_server'):
            server = app.warnet_server
            if getattr(server, 'running', False):
                with server.sessions_lock:
                    sessions = list(server.sessions.values())
                pc_connected = any(s.get("client_id") == self._client_id for s in sessions)
        pc_text = "● Connected" if pc_connected else "● Disconnected"
        pc_color = C_GREEN if pc_connected else C_RED
        self._ids['lbl_pc_status'] = self.create_text(W-52, hdr_y+12,
            text=pc_text, font=("Courier New", 11, "bold"),
            fill=pc_color, anchor="e", tags="lbl_pc_status")

        # Delete button
        self._draw_canvas_btn("hapus", W-44, hdr_y+4, 36, 24, "✖",
            C_RED, C_RED, ("Russo One", 10, "bold"), self._confirm_hapus)

        y = hdr_y + hdr_h + 4
        # Status row
        self._ids['lbl_grup'] = self.create_text(70, y+10,
            text=f"\u21bb {self.nama_grup}", font=("Courier New", 11, "bold"),
            fill=C_ACCENT2, anchor="w", tags="lbl_grup")
        self._ids['lbl_paket'] = self.create_text(W-8, y+10,
            text="\u2014", font=("Courier New", 10),
            fill=C_MUTED, anchor="e", tags="lbl_paket")
        y += 24

        # Timer
        self._ids['lbl_timer'] = self.create_text(W//2, y+14,
            text="00:00:00", font=("Russo One", 20, "bold"),
            fill=C_ACCENT2, anchor="center", tags="lbl_timer")
        y += 32

        # Estimasi
        self._ids['lbl_estimasi'] = self.create_text(W//2, y+6,
            text="", font=("Courier New", 10),
            fill=C_YELLOW, anchor="center", tags="lbl_estimasi")
        y += 16

        # Button Row 1
        btn_h, gap_b = 30, 4
        bw = (W - 8 - gap_b) // 2
        bx = 4
        r1y = y
        btn_defs1 = [
            ("buka", "\U0001F513 BUKA", C_GREEN, self._buka_unlock),
            ("status", "ON", C_BTN, self._toggle_power),
        ]
        for i, (key, txt, col, cmd) in enumerate(btn_defs1):
            self._draw_canvas_btn(key, bx + i*(bw+gap_b), r1y, bw, btn_h, txt, col, col, ("Russo One", 10, "bold"), cmd)
        y = r1y + btn_h + 4

        # Button Row 2
        r2y = y
        bw2 = (W - 8 - 4 * gap_b) // 5
        btn_defs2 = [
            ("selesai", "SELESAI", C_BTN, C_RED, self._klik_selesai, True),
            ("shop", "SHOP", C_BTN, C_ACCENT, self._buka_tambah_pesanan, True),
            ("paket", "PAKET", C_BTN, C_ACCENT2, self._pilih_paket, False),
            ("ip", "IP", C_BTN, C_ACCENT2, self._buka_ganti_ip, False),
            ("pindah", "Pindah PC", C_BTN, C_ACCENT2, self._klik_pindah, False),
        ]
        for i, (key, txt, bg, fg, cmd, disabled) in enumerate(btn_defs2):
            self._draw_canvas_btn(key, bx + i*(bw2+gap_b), r2y, bw2, btn_h, txt, bg, fg, ("Russo One", 10, "bold"), cmd)
            if disabled:
                self._disable_btn(key)
        y = r2y + btn_h + 4

        # Button Row 3 (Status pembayaran)
        r3y = y
        n_btn3, btn_cols3 = 2, 2
        avail3 = W - 8 - (n_btn3 - 1) * gap_b
        bw3 = avail3 // btn_cols3
        btn_defs3 = [
            ("bayar_lunas", "✅ SUDAH BAYAR", "black", "white", lambda: self._set_paid(True)),
            ("bayar_belum", "⏳ BELUM BAYAR", "black", "white", lambda: self._set_paid(False)),
        ]
        for i, (key, txt, bg, fg, cmd) in enumerate(btn_defs3):
            self._draw_canvas_btn(key, bx + i*(bw3+gap_b), r3y, bw3, btn_h, txt, bg, fg, ("Russo One", 9, "bold"), cmd)
            self._disable_btn(key)
            self.tag_bind(f"btn_{key}", "<Enter>", lambda e, k=key: self._update_bayar_buttons())
            self.tag_bind(f"btn_{key}", "<Leave>", lambda e, k=key: self._update_bayar_buttons())
        y = r3y + btn_h + 6
        self._update_paid_badge()

        total_h = y
        self._ids['_card_h'] = total_h
        self.configure(height=total_h, width=W)
        self._ids['bg'] = self.create_rectangle(0, 0, W, total_h, fill=C_CARD, outline=C_BORDER, width=1, tags="bg")
        self.tag_lower(self._ids['bg'])

    def _redraw_bg(self):
        W = self._card_w
        h = self._ids.get('_card_h', 200)
        self.coords(self._ids['bg'], 0, 0, W, h)
        self.tag_lower(self._ids['bg'])

    def _draw_canvas_btn(self, key, x, y, w, h, text, bg, fg, font, cmd):
        rect = self.create_rectangle(x, y, x+w, y+h, fill=bg, outline=fg, width=1, tags=f"btn_{key}")
        txt_id = self.create_text(x+w//2, y+h//2, text=text, font=font,
                                  fill=self._btn_text_color(bg, fg), tags=(f"btn_{key}", f"btn_{key}_txt"))
        self._ids[f'btn_{key}'] = rect
        self._ids[f'btn_{key}_txt'] = txt_id
        self._btn_states[key] = "normal"
        self.tag_bind(f"btn_{key}", "<Button-1>", lambda e, c=cmd: c())
        self.tag_bind(f"btn_{key}", "<Enter>", lambda e, k=key, bg=bg: self._btn_hover(k, bg))
        self.tag_bind(f"btn_{key}", "<Leave>", lambda e, k=key, bg=bg: self._btn_leave(k, bg))
        return rect, txt_id

    def _btn_text_color(self, bg, fg):
        if bg == C_ACCENT:
            return "black"
        if bg == C_BTN:
            return fg
        return "white"

    def _btn_hover(self, key, bg):
        if self._btn_states.get(key) == "disabled":
            return
        if bg == C_RED:
            self.itemconfig(self._ids[f'btn_{key}'], fill="#FF6666")
        elif bg == C_GREEN or bg == C_YELLOW:
            self.itemconfig(self._ids[f'btn_{key}'], fill="#66BB6A")
        elif bg == C_ACCENT:
            self.itemconfig(self._ids[f'btn_{key}'], fill="#66FFE0")
        elif bg == C_BTN:
            self.itemconfig(self._ids[f'btn_{key}'], fill=C_ACCENT2)

    def _btn_leave(self, key, bg):
        if self._btn_states.get(key) == "disabled":
            self._disable_btn(key)
        else:
            self.itemconfig(self._ids[f'btn_{key}'], fill=bg)

    def _enable_btn(self, key, bg=C_BTN, fg=C_ACCENT2):
        self._btn_states[key] = "normal"
        self.itemconfig(self._ids[f'btn_{key}'], fill=bg, outline=fg)
        self.itemconfig(self._ids[f'btn_{key}_txt'], fill=fg)

    def _disable_btn(self, key):
        self._btn_states[key] = "disabled"
        self.itemconfig(self._ids[f'btn_{key}'], fill=C_BTN, outline=C_BORDER)
        self.itemconfig(self._ids[f'btn_{key}_txt'], fill=C_MUTED)

    def sesi_kosong(self):
        return self.paket_aktif is None

    # ── Status pembayaran (sinkron ke riwayat) ───────────────────────────────
    def _update_paid_badge(self):
        """Update badge 'LUNAS' / 'TAGIHAN' (atau gabungan) di header."""
        if 'paid_badge' not in self._ids:
            return
        name_w = len(self.label_kursi) * 7 + 6
        if self.sesi_kosong():
            text, color = "", C_GREEN
        elif self.is_bebas:
            text, color = "⏳ TAGIHAN", "#FFCC00"
        else:
            lunas, tagihan = self._split_payment()
            if lunas > 0 and tagihan > 0:
                text, color = "LUNAS + TAGIHAN", "#FF9933"
            elif lunas > 0:
                text, color = "● LUNAS", C_GREEN
            else:
                text, color = "⏳ TAGIHAN", "#FFCC00"
        self.itemconfig(self._ids['paid_badge'], text=text, fill=color)
        try:
            cur_x, cur_y = self.coords(self._ids['lbl_kursi'])
        except Exception:
            cur_x, cur_y = 10, 5
        self.coords(self._ids['paid_badge'], cur_x + len(self.label_kursi) * 7 + 6, cur_y)

    def _split_payment(self):
        """Split total sesi jadi (lunas_total, tagihan_total) per item status."""
        try:
            subtotal = self.paket_harga_tetap + self.biaya_pesanan
            total = self._total_setelah_diskon()
        except Exception:
            total = getattr(self, "paket_harga_tetap", 0) + getattr(self, "biaya_pesanan", 0)
            subtotal = total
        if self.sesi_kosong():
            return 0, 0
        if subtotal <= 0:
            if getattr(self, "paid", True):
                return total, 0
            return 0, total
        lunas_sub = 0
        harga_paket = getattr(self, "harga_paket_sesi", None) or []
        lunas_paket = getattr(self, "lunas_paket", None) or []
        if harga_paket:
            for i, h in enumerate(harga_paket):
                paid_i = lunas_paket[i] if i < len(lunas_paket) else getattr(self, "paid", True)
                if paid_i:
                    lunas_sub += h
        else:
            if getattr(self, "paid", True):
                lunas_sub += self.paket_harga_tetap
        lunas_pesanan = getattr(self, "lunas_pesanan", None) or {}
        all_menu = {}
        try:
            all_menu = {**self.get_makanan_data(), **self.get_minuman_data()}
        except Exception:
            pass
        for nm, qty in (self.pesanan_aktif or {}).items():
            paid_i = lunas_pesanan.get(nm, getattr(self, "paid", True))
            if paid_i:
                lunas_sub += all_menu.get(nm, 0) * qty
        lunas_sub = min(lunas_sub, subtotal)
        if lunas_sub >= subtotal:
            return total, 0
        if lunas_sub <= 0:
            return 0, total
        lunas = round(total * lunas_sub / subtotal)
        return lunas, max(0, total - lunas)

    def _update_bayar_buttons(self):
        """Warna tombol SUDAH/BELUM BAYAR sesuai state aktif (lunas = hijau)."""
        if 'btn_bayar_lunas' not in self._ids:
            return
        lunas, tagihan = self._split_payment()
        if tagihan <= 0:
            self.itemconfig(self._ids['btn_bayar_lunas'], fill=C_GREEN, outline=C_GREEN)
            self.itemconfig(self._ids['btn_bayar_lunas_txt'], fill="white")
            self.itemconfig(self._ids['btn_bayar_belum'], fill="black", outline=C_BORDER)
            self.itemconfig(self._ids['btn_bayar_belum_txt'], fill=C_MUTED)
        else:
            self.itemconfig(self._ids['btn_bayar_belum'], fill="#FFCC00", outline="#FFCC00")
            self.itemconfig(self._ids['btn_bayar_belum_txt'], fill="black")
            self.itemconfig(self._ids['btn_bayar_lunas'], fill="black", outline=C_BORDER)
            self.itemconfig(self._ids['btn_bayar_lunas_txt'], fill=C_MUTED)

    def _bind_last_transaction(self):
        """Simpan referensi index/cloud_id baris riwayat milik sesi ini."""
        try:
            app = self.winfo_toplevel()
            item_id = getattr(self, '_last_transaction_item', None)
            idx = app._tree_item_to_index.get(item_id, -1)
            if idx < 0:
                idx = getattr(app, "_last_catat_idx", -1)
            self._last_riwayat_idx = idx
            if 0 <= idx < len(app.riwayat_meta):
                self._last_cloud_id = app.riwayat_meta[idx].get("cloud_id")
        except Exception:
            pass

    def _set_paid(self, paid):
        """Tandai sesi lunas/belum lunas, sinkronkan ke baris riwayat terkait."""
        if self.sesi_kosong() or self.is_bebas:
            return
        if self._last_transaction_item is None and self._last_riwayat_idx is None:
            return
        app = self.winfo_toplevel()
        idx = app._resolve_session_idx(self)
        row_needs_sync = False
        if 0 <= idx < len(app.riwayat_meta):
            row_needs_sync = bool(app.riwayat_meta[idx].get('paid', True)) != bool(paid)
        if self.paid == paid and not row_needs_sync:
            return
        self.paid = paid
        # Semua item ikut memilih tombol
        self.lunas_paket = [paid] * (len(self.daftar_paket_sesi or []) or 1)
        if self.pesanan_aktif:
            self.lunas_pesanan = {nm: paid for nm in self.pesanan_aktif}
        self._update_paid_badge()
        self._update_bayar_buttons()
        if idx >= 0 and hasattr(app, '_set_transaksi_paid_idx'):
            app._set_transaksi_paid_idx(idx, paid)
        elif hasattr(app, '_set_transaksi_paid'):
            app._set_transaksi_paid(self._last_transaction_item, paid)

    def _update_pc_status(self):
        app = self.winfo_toplevel()
        connected = False
        if self._client_id and self._pc_id and hasattr(app, 'warnet_server'):
            server = app.warnet_server
            if getattr(server, 'running', False):
                with server.sessions_lock:
                    sessions = list(server.sessions.values())
                connected = any(s.get("client_id") == self._client_id for s in sessions)
        text = "● Connected" if connected else "● Disconnected"
        color = C_GREEN if connected else C_RED
        if 'lbl_pc_status' in self._ids:
            self.itemconfig(self._ids['lbl_pc_status'], text=text, fill=color)

    def _periodic_pc_status(self):
        if not self.winfo_exists():
            return
        self._update_pc_status()
        self.after(10000, self._periodic_pc_status)

    def _buka_ganti_grup(self):
        daftar = self.get_daftar_grup() or []
        if not daftar:
            messagebox.showwarning("Grup Tarif", "Tidak ada grup tarif yang tersedia untuk dipilih.", parent=self)
            return

        dlg = ctk.CTkToplevel(self.winfo_toplevel())
        dlg.title("Ganti Grup Tarif Warnet")
        dlg.geometry("360x180")
        dlg.configure(fg_color=C_BG)
        dlg.resizable(False, False)
        dlg.grab_set()

        ctk.CTkLabel(dlg, text=f"Kursi: {self.label_kursi}", font=FONT_SUB,
                     text_color=C_ACCENT).pack(pady=(16, 6))
        ctk.CTkLabel(dlg, text="Pilih grup tarif untuk kursi warnet ini:",
                     font=FONT_BODY, text_color=C_MUTED).pack(padx=16)

        var_grup = tk.StringVar(value=self.nama_grup)
        ctk.CTkOptionMenu(dlg, values=daftar, variable=var_grup,
                          fg_color=C_BTN, button_color=C_ACCENT2,
                          button_hover_color="#5A0FCC",
                          text_color=C_TEXT, font=FONT_BODY,
                          dropdown_font=FONT_BODY,
                          dropdown_fg_color=C_CARD,
                          dropdown_text_color=C_TEXT,
                          width=280).pack(pady=14)

        def terapkan():
            grup_baru = var_grup.get()
            if not grup_baru or grup_baru == self.nama_grup:
                dlg.destroy()
                return
            self.nama_grup = grup_baru
            self.itemconfig(self._ids['lbl_grup'], text=f"\u21bb {grup_baru}")
            if callable(self.on_ganti_grup):
                self.on_ganti_grup(self, grup_baru)
            dlg.destroy()

        ctk.CTkButton(dlg, text="Terapkan", width=120, height=36,
                      fg_color=C_ACCENT2, hover_color=C_ACCENT,
                      font=FONT_SMALL, text_color="white",
                      command=terapkan).pack(pady=(0, 10))

    def _buka_ganti_ip(self):
        messagebox.showinfo("IP Client", f"IP: {self._pc_ip or '\u2014'}\nPC ID: {self._pc_id or '\u2014'}\nClient: {self._client_id or '\u2014'}", parent=self)

    def _confirm_hapus(self):
        if messagebox.askyesno("Hapus Kursi", f"Hapus {self.label_kursi} dari dashboard warnet?"):
            if callable(self.on_hapus):
                self.on_hapus(self)

    def _toggle_power(self):
        self.is_on = not self.is_on
        app = self.winfo_toplevel()
        if self.is_on:
            self.itemconfig(self._ids['btn_status'], fill=C_GREEN, outline=C_GREEN)
            self.itemconfig(self._ids['btn_status_txt'], text="ON", fill=C_GREEN)
            self._send_lock_command(app, "UNLOCK", "manual_on",
                                    f"PC {self.label_kursi} dinyalakan admin.")
        else:
            self.itemconfig(self._ids['btn_status'], fill=C_BTN, outline=C_BORDER)
            self.itemconfig(self._ids['btn_status_txt'], text="OFF", fill=C_MUTED)
            self._send_lock_command(app, "LOCK", "manual_off",
                                    f"PC {self.label_kursi} dimatikan admin.")

    # Tombol "BUKA" khusus: SELALU kirim UNLOCK, tidak mengubah is_on.
    # Menghindari jebakan tombol PWR yang (jika is_on masih True) justru
    # mematikan PC / mengirim LOCK.
    def _buka_unlock(self):
        app = self.winfo_toplevel()
        self.pc_locked = False
        self._pc_lock_reason = ""
        self._pc_lock_message = ""
        self._send_lock_command(app, "UNLOCK", "manual_on",
                                f"PC {self.label_kursi} dibuka kunci oleh admin.")

    def _send_lock_command(self, app, cmd, reason, message):
        if not (hasattr(app, 'warnet_server') and getattr(self, '_pc_id', None)):
            if not getattr(self, '_pc_id', None):
                messagebox.showinfo(
                    "Client Belum Terhubung",
                    f"{self.label_kursi} belum terhubung ke client PC.\n"
                    "Hubungkan client lalu ulangi.", parent=self.winfo_toplevel())
            return
        # Tandai state lock kartu (dipakai untuk re-queue LOCK saat client reconnect)
        self.pc_locked = (cmd == "LOCK")
        if self.pc_locked:
            self._pc_lock_reason = reason
            self._pc_lock_message = message
        server = app.warnet_server
        ok = server.queue_pending_command(self._pc_id, cmd, reason=reason, message=message)
        with server.sessions_lock:
            active = any(s.get("client_id") == getattr(self, '_client_id', None)
                         for s in list(server.sessions.values()))
        if not active:
            messagebox.showwarning(
                "Client Offline",
                f"{self.label_kursi} tidak terhubung ke server.\n"
                "Perintah akan dikirim saat client kembali terhubung.", parent=self.winfo_toplevel())

    def _dummy_action(self, action):
        messagebox.showinfo("Info", f"{action} tidak tersedia untuk dashboard warnet.", parent=self)

    def _pilih_paket(self):
        try:
            paket_data = self.get_paket_data() if callable(self.get_paket_data) else {}
        except Exception as e:
            paket_data = {}

        DialogPaket(self.winfo_toplevel(), self.label_kursi, self._on_paket_confirm,
                    paket_data, self.get_makanan_data(), self.get_minuman_data(),
                    nama_grup=self.nama_grup, for_warnet=True)

    def _buka_tambah_pesanan(self):
        if not self.paket_aktif:
            messagebox.showwarning("Tidak Ada Sesi", "Mulai sesi terlebih dahulu untuk memesan.", parent=self)
            return
        DialogTambahPesanan(self.winfo_toplevel(), self.label_kursi,
                            self._on_tambah_pesanan_confirm,
                            self.get_makanan_data(), self.get_minuman_data(),
                            pesanan_aktif=self.pesanan_aktif.copy(),
                            paket_harga=self.paket_harga_tetap, paket_label=self.paket_aktif)

    def _on_tambah_pesanan_confirm(self, pesanan_baru, paid=True, _stok_delta=None):
        app = self.winfo_toplevel()
        # Validasi stok sebelum pesanan diterima (blokir kalau tidak cukup)
        if hasattr(app, '_stok_validate_orders'):
            stok_check = pesanan_baru if _stok_delta is None else {
                k: v for k, v in _stok_delta.items() if int(v or 0) > 0}
            ok, pesan, nm, sisa = app._stok_validate_orders(stok_check)
            if not ok:
                messagebox.showwarning("⚠ Stok Tidak Mencukupi", pesan, parent=app)
                return
        old_pesanan = dict(getattr(self, 'pesanan_aktif', {}) or {})
        all_menu = {**self.get_makanan_data(), **self.get_minuman_data()}
        for nama, qty in pesanan_baru.items():
            self.pesanan_aktif[nama] = qty
            self.lunas_pesanan[nama] = paid
        total_baru = sum(all_menu.get(nama, 0) * qty for nama, qty in self.pesanan_aktif.items())
        self.biaya_pesanan = total_baru

        if self.is_bebas:
            self.itemconfig(self._ids['lbl_paket'], text=f"Main Bebas +Pesanan {fmt_rp(total_baru)}")
        else:
            total_semua = self._total_setelah_diskon(self.paket_harga_tetap + total_baru)
            self.itemconfig(self._ids['lbl_paket'], text=f"{self.paket_aktif} | {fmt_rp(total_semua)}")
        self._update_paid_badge()
        self._update_bayar_buttons()

        if not self.is_bebas and (getattr(self, '_last_transaction_item', None) or getattr(self, '_last_riwayat_idx', None) is not None):
            app = self.winfo_toplevel()
            idx = app._resolve_session_idx(self)
            total_int = self._total_setelah_diskon()
            if idx >= 0 and hasattr(app, 'tree') and hasattr(app, '_tree_item_to_index'):
                item_id = app._tree_item_for_idx(idx)
                if item_id is None:
                    item_id = self._last_transaction_item
                if idx < len(app.riwayat_transaksi):
                    waktu = app.riwayat_transaksi[idx][0] if len(app.riwayat_transaksi[idx]) > 0 else datetime.now().strftime("%Y-%m-%d %H:%M")
                    updated_row = app._format_riwayat_row(waktu, self.label_kursi, self.paket_aktif, self.pesanan_aktif, total_int, paid=self.paid)
                    app.riwayat_transaksi[idx] = updated_row
                    if item_id is not None and app.tree.exists(item_id):
                        app.tree.item(item_id, values=updated_row)
                    try:
                        all_menu = {**app.menu_makanan, **app.menu_minuman}
                        pesanan_total = sum(all_menu.get(nm, 0) * qty for nm, qty in self.pesanan_aktif.items())
                        paket_harga = total_int - pesanan_total
                        if paket_harga < 0:
                            paket_harga = 0
                        if idx < len(app.riwayat_meta):
                            app.riwayat_meta[idx]['paket_harga'] = paket_harga
                            app.riwayat_meta[idx]['pesanan_total'] = pesanan_total
                            app.riwayat_meta[idx]['total'] = total_int
                            app.riwayat_meta[idx]['diskoni'] = self.diskoni
                            app.riwayat_meta[idx]['diskoni_mode'] = self.diskoni_mode
                            app.riwayat_meta[idx]['pesanan'] = {
                                str(k): int(v) for k, v in self.pesanan_aktif.items()}
                            threading.Thread(
                                target=app._upsert_tx_cloud_from_index, args=(idx,), daemon=True).start()
                    except Exception:
                        pass
                    if hasattr(app, '_refresh_riwayat_summary'):
                        app._refresh_riwayat_summary()
                    if hasattr(app, '_save_riwayat'):
                        app._save_riwayat()
        app = self.winfo_toplevel()
        if hasattr(app, '_refresh_warnet_footer'):
            app._refresh_warnet_footer()

        # Terapkan perubahan stok (selisih qty lama → baru; QR pakai delta eksplisit)
        try:
            if hasattr(app, '_stok_terapkan'):
                if _stok_delta is not None:
                    app._stok_terapkan(_stok_delta)
                else:
                    delta = {}
                    for nama, qty in pesanan_baru.items():
                        d = int(qty or 0) - int(old_pesanan.get(nama, 0) or 0)
                        if d:
                            delta[nama] = d
                    app._stok_terapkan(delta)
        except Exception:
            pass

    def _on_paket_confirm(self, paket_nm, paket_harga, paket_menit, pesanan, total_pesanan, diskoni=0, diskoni_mode="nominal", paid=True, booking=None):
        app = self.winfo_toplevel()
        # Validasi stok makanan/minuman sebelum paket+pesanan diterima
        if pesanan and hasattr(app, '_stok_validate_orders'):
            ok, pesan, nm, sisa = app._stok_validate_orders(pesanan)
            if not ok:
                messagebox.showwarning("⚠ Stok Tidak Mencukupi", pesan, parent=app)
                return
        previous_session = not self.sesi_kosong()
        old_pesanan = dict(getattr(self, 'pesanan_aktif', {}) or {})
        self.menit_dipakai_awal = 0
        self.diskoni = diskoni
        self.diskoni_mode = diskoni_mode

        if not previous_session:
            self.pesanan_aktif = pesanan
            all_menu = {**self.get_makanan_data(), **self.get_minuman_data()}
            self.biaya_pesanan = sum(all_menu.get(nm, 0) * qty for nm, qty in pesanan.items())
            self.daftar_paket_sesi = [paket_nm]
            self.harga_paket_sesi = [paket_harga]
            self.lunas_paket = [paid]
            self.lunas_pesanan = {nm: paid for nm in pesanan}
        else:
            for nm, qty in pesanan.items():
                self.pesanan_aktif[nm] = qty
                self.lunas_pesanan[nm] = paid
            if pesanan:
                all_menu = {**self.get_makanan_data(), **self.get_minuman_data()}
                self.biaya_pesanan += sum(all_menu.get(nm, 0) * qty for nm, qty in pesanan.items())
            self.daftar_paket_sesi.append(paket_nm)
            self.harga_paket_sesi.append(paket_harga)
            self.lunas_paket.append(paid)

        if paket_nm == "Main Bebas":
            self.is_bebas = True
            self.sisa_waktu = 0
            self.waktu_mulai = datetime.now()
            self.paket_harga_tetap = 0
            self.paket_aktif = paket_nm
            self.itemconfig(self._ids['lbl_paket'], text="Main Bebas (berjalan)", fill=C_GREEN)
            self.itemconfig(self._ids['lbl_timer'], fill=C_YELLOW)
            self.itemconfig(self._ids['lbl_estimasi'], text=f"Total berjalan: {fmt_rp(self.biaya_pesanan)}", fill=C_YELLOW)
        else:
            self.is_bebas = False
            self.paket_aktif = paket_nm
            if previous_session and self.paket_harga_tetap:
                self.sisa_waktu += paket_menit * 60
                self.paket_harga_tetap += paket_harga
            else:
                self.sisa_waktu = paket_menit * 60
                self.paket_harga_tetap = paket_harga
            self.waktu_mulai = datetime.now()
            self.itemconfig(self._ids['lbl_paket'], text=f"{paket_nm} | {fmt_rp(self._total_setelah_diskon())}",
                             fill=C_YELLOW)
            self.itemconfig(self._ids['lbl_timer'], fill=C_ACCENT)

        if self._timer_job:
            self.after_cancel(self._timer_job)

        self.is_on = True
        self.itemconfig(self._ids['btn_status'], fill=C_GREEN, outline=C_GREEN)
        self.itemconfig(self._ids['btn_status_txt'], text="ON", fill=C_GREEN)
        self._enable_btn("selesai", C_BTN, C_RED)
        self._enable_btn("shop", C_BTN, C_ACCENT2)
        self.itemconfig(self._ids['btn_shop_txt'], text="SHOP")
        self.itemconfig(self._ids['bg'], fill="white")
        self._warning_blink_on = False

        if not self.is_bebas and self.sisa_waktu > 0:
            h, rem = divmod(self.sisa_waktu, 3600)
            m, s = divmod(rem, 60)
            self.itemconfig(self._ids['lbl_timer'], text=f"{h:02d}:{m:02d}:{s:02d}")

        if self.is_bebas:
            self._tick_bebas()
        elif self.sisa_waktu > 0:
            self._tick_waktu()
        else:
            self.itemconfig(self._ids['lbl_timer'], text="\u221e BEBAS", fill=C_GREEN)

        if not self.is_bebas:
            if (self._last_transaction_item or self._last_riwayat_idx is not None) and previous_session:
                app = self.winfo_toplevel()
                idx = app._resolve_session_idx(self)
                total_int = self._total_setelah_diskon()
                if idx >= 0 and hasattr(app, 'tree') and hasattr(app, '_tree_item_to_index'):
                    item_id = app._tree_item_for_idx(idx)
                    if item_id is None:
                        item_id = self._last_transaction_item
                    if idx < len(app.riwayat_transaksi):
                        waktu = app.riwayat_transaksi[idx][0] if len(app.riwayat_transaksi[idx]) > 0 else datetime.now().strftime("%Y-%m-%d %H:%M")
                        updated_row = app._format_riwayat_row(waktu, self.label_kursi, self.paket_aktif, self.pesanan_aktif, total_int, paid=self.paid)
                        app.riwayat_transaksi[idx] = updated_row
                        if item_id is not None and app.tree.exists(item_id):
                            app.tree.item(item_id, values=updated_row)
                        try:
                            all_menu = {**app.menu_makanan, **app.menu_minuman}
                            pesanan_total_baru = sum(all_menu.get(nm, 0) * qty for nm, qty in self.pesanan_aktif.items())
                            paket_harga_baru = total_int - pesanan_total_baru
                            if paket_harga_baru < 0:
                                paket_harga_baru = 0
                            if idx < len(app.riwayat_meta):
                                app.riwayat_meta[idx]['paket_harga'] = paket_harga_baru
                                app.riwayat_meta[idx]['pesanan_total'] = pesanan_total_baru
                                app.riwayat_meta[idx]['total'] = total_int
                                app.riwayat_meta[idx]['diskoni'] = self.diskoni
                                app.riwayat_meta[idx]['diskoni_mode'] = self.diskoni_mode
                                app.riwayat_meta[idx]['pesanan'] = {
                                    str(k): int(v) for k, v in self.pesanan_aktif.items()}
                                threading.Thread(
                                    target=app._upsert_tx_cloud_from_index, args=(idx,), daemon=True).start()
                        except Exception:
                            pass
                        if hasattr(app, '_refresh_riwayat_summary'):
                            app._refresh_riwayat_summary()
                        if hasattr(app, '_save_riwayat'):
                            app._save_riwayat()
            else:
                total_int_baru = self._total_setelah_diskon(self.paket_harga_tetap + total_pesanan)
                self._last_transaction_item = self.on_transaksi(
                    self.label_kursi, paket_nm, pesanan, total_int_baru, source='warnet',
                    diskoni=self.diskoni, diskoni_mode=self.diskoni_mode, paid=paid)
                self._bind_last_transaction()
        else:
            self._last_transaction_item = None
            self._last_riwayat_idx = None
            self._last_cloud_id = None

        # Status pembayaran: hanya untuk paket berwaktu (bukan Main Bebas)
        if self.is_bebas:
            self.paid = True
            self.lunas_paket = [True] * (len(self.daftar_paket_sesi or []) or 1)
            self.lunas_pesanan = {nm: True for nm in self.pesanan_aktif}
            self._disable_btn("bayar_lunas")
            self._disable_btn("bayar_belum")
        else:
            if not previous_session:
                self.paid = paid   # status bayar sesuai pilihan kasir di popup
            self._enable_btn("bayar_lunas", "black", "white")
            self._enable_btn("bayar_belum", "black", "white")
            self._update_bayar_buttons()
        self._update_paid_badge()

        app = self.winfo_toplevel()
        if hasattr(app, 'warnet_server') and getattr(self, '_pc_id', None):
            self.pc_locked = False
            self._pc_lock_reason = ""
            self._pc_lock_message = ""
            app.warnet_server.queue_pending_command(
                self._pc_id, "UNLOCK",
                reason="sesi_baru",
                message=f"Sesi baru dimulai untuk {self.label_kursi}"
            )
        if hasattr(app, '_refresh_warnet_footer'):
            app._refresh_warnet_footer()

        # Terapkan perubahan stok (selisih qty lama → baru)
        try:
            if pesanan and hasattr(app, '_stok_terapkan'):
                delta = {}
                for nm, qty in pesanan.items():
                    d = int(qty or 0) - int(old_pesanan.get(nm, 0) or 0)
                    if d:
                        delta[nm] = d
                app._stok_terapkan(delta)
        except Exception:
            pass

        # Popup cetak struk setelah sesi baru dimulai (konfirmasi pembayaran)
        if not previous_session:
            try:
                self._buka_print_struk(self._trans_struk())
            except Exception:
                pass

    def _tick_waktu(self):
        if self._timer_paused:
            return
        self._update_timer_display()

    def _tick_bebas(self):
        if self._timer_paused or not self.is_bebas or self.waktu_mulai is None:
            return
        total_detik = self.menit_dipakai_awal * 60 + int((datetime.now() - self.waktu_mulai).total_seconds())
        h, rem = divmod(total_detik, 3600)
        m, s = divmod(rem, 60)
        self.itemconfig(self._ids['lbl_timer'], text=f"{h:02d}:{m:02d}:{s:02d}", fill=C_YELLOW)
        tarif_menit = hitung_tarif_per_menit(self.get_paket_data())
        menit_berjalan = total_detik / 60
        biaya_waktu_berjalan = tarif_menit * menit_berjalan
        total_berjalan = self._total_setelah_diskon(biaya_waktu_berjalan + self.biaya_pesanan)
        self.itemconfig(self._ids['lbl_estimasi'],
            text=f"Total berjalan: {fmt_rp(total_berjalan)}",
            fill=C_YELLOW)
        if not self._timer_paused:
            self._timer_job = self.after(1000, self._tick_bebas)

    # ── Update timer display (called from TimerService) ─────────────────────
    def _update_timer_display(self):
        if self._timer_paused:
            return
        if self.is_bebas:
            total_detik = self.menit_dipakai_awal * 60 + int((datetime.now() - self.waktu_mulai).total_seconds()) if self.waktu_mulai else 0
            h, rem = divmod(total_detik, 3600)
            m, s = divmod(rem, 60)
            self.itemconfig(self._ids['lbl_timer'], text=f"{h:02d}:{m:02d}:{s:02d}", fill=C_YELLOW)
            tarif_menit = hitung_tarif_per_menit(self.get_paket_data())
            menit_berjalan = total_detik / 60
            biaya_waktu_berjalan = tarif_menit * menit_berjalan
            total_berjalan = self._total_setelah_diskon(biaya_waktu_berjalan + self.biaya_pesanan)
            self.itemconfig(self._ids['lbl_estimasi'],
                text=f"Total berjalan: {fmt_rp(total_berjalan)}",
                fill=C_YELLOW)
        else:
            h, rem = divmod(self.sisa_waktu, 3600)
            m, s = divmod(rem, 60)
            self.itemconfig(self._ids['lbl_timer'], text=f"{h:02d}:{m:02d}:{s:02d}")
            if self.sisa_waktu <= 120:
                self._warning_blink_on = not self._warning_blink_on
                if self._warning_blink_on:
                    self.itemconfig(self._ids['bg'], fill=C_RED)
                    self.itemconfig(self._ids['lbl_timer'], fill="white")
                else:
                    self.itemconfig(self._ids['bg'], fill="white")
                    self.itemconfig(self._ids['lbl_timer'], fill=C_ACCENT2)
            else:
                self.itemconfig(self._ids['bg'], fill="white")
                self.itemconfig(self._ids['lbl_timer'], fill=C_ACCENT2)

    # ── Timer habis handler (called from TimerService) ──────────────────────
    def _timer_habis(self):
        if self.sesi_kosong():
            return
        self.itemconfig(self._ids['lbl_timer'], text="WAKTU HABIS", fill=C_RED)
        total_akhir = self._total_setelah_diskon()
        pesanan_txt = ", ".join(f"{nm}\u00d7{qty}" for nm, qty in self.pesanan_aktif.items()) or "Tidak ada pesanan"
        paket_txt = f"{self.paket_aktif or '-'} ({fmt_rp(self.paket_harga_tetap)})"
        app = self.winfo_toplevel()
        idx = app._resolve_session_idx(self)
        if idx >= 0 and hasattr(app, 'tree') and hasattr(app, '_tree_item_to_index'):
            item_id = app._tree_item_for_idx(idx)
            if item_id is None:
                item_id = self._last_transaction_item
            if idx < len(app.riwayat_transaksi):
                waktu = app.riwayat_transaksi[idx][0] if len(app.riwayat_transaksi[idx]) > 0 else datetime.now().strftime("%Y-%m-%d %H:%M")
                updated_row = app._format_riwayat_row(waktu, self.label_kursi, self.paket_aktif, self.pesanan_aktif, total_akhir, paid=self.paid)
                app.riwayat_transaksi[idx] = updated_row
                if item_id is not None and app.tree.exists(item_id):
                    app.tree.item(item_id, values=updated_row)
                try:
                    all_menu = {**app.menu_makanan, **app.menu_minuman}
                    pesanan_total = sum(all_menu.get(nm, 0) * qty for nm, qty in self.pesanan_aktif.items())
                    paket_harga = total_akhir - pesanan_total
                    if paket_harga < 0:
                        paket_harga = 0
                    if idx < len(app.riwayat_meta):
                        app.riwayat_meta[idx]['paket_harga'] = paket_harga
                        app.riwayat_meta[idx]['pesanan_total'] = pesanan_total
                        app.riwayat_meta[idx]['total'] = total_akhir
                        app.riwayat_meta[idx]['diskoni'] = self.diskoni
                        app.riwayat_meta[idx]['diskoni_mode'] = self.diskoni_mode
                        app.riwayat_meta[idx]['pesanan'] = {
                            str(k): int(v) for k, v in self.pesanan_aktif.items()}
                        threading.Thread(
                            target=app._upsert_tx_cloud_from_index, args=(idx,), daemon=True).start()
                except Exception:
                    pass
                if hasattr(app, '_refresh_riwayat_summary'):
                    app._refresh_riwayat_summary()
                if hasattr(app, '_save_riwayat'):
                    app._save_riwayat()
        else:
            self.on_transaksi(self.label_kursi, self.paket_aktif or '-', self.pesanan_aktif, total_akhir, source='warnet',
                              diskoni=self.diskoni, diskoni_mode=self.diskoni_mode)
        if hasattr(app, 'warnet_server') and getattr(self, '_pc_id', None):
            self.pc_locked = True
            self._pc_lock_reason = "waktu_habis"
            self._pc_lock_message = f"Waktu PC {self.label_kursi} telah habis."
            print(f"[WARNET] {self.label_kursi} waktu habis -> queue LOCK ke "
                  f"pc_id={self._pc_id} (is_on={getattr(self, 'is_on', None)})", flush=True)
            app.warnet_server.queue_pending_command(
                self._pc_id, "LOCK",
                reason="waktu_habis",
                time_left=0,
                message=f"Waktu PC {self.label_kursi} telah habis."
            )
        elif hasattr(app, 'warnet_server'):
            print(f"[WARNET] {self.label_kursi} waktu habis TAPI _pc_id kosong — "
                  f"LOCK TIDAK dikirim! pc_id=None, client_id={getattr(self, '_client_id', None)}",
                  flush=True)
        messagebox.showwarning(
            "Waktu PC Habis",
            f"PC: {self.label_kursi}\n"
            f"Paket: {paket_txt}\n"
            f"Pesanan: {pesanan_txt} ({fmt_rp(self.biaya_pesanan)})\n"
            f"TOTAL: {fmt_rp(total_akhir)}",
            parent=self.winfo_toplevel(),
        )
        # Jika masih ada sisa TAGIHAN, kasir diminta konfirmasi (BAYAR/TAGIHAN).
        lunas_now, tagihan_now = self._split_payment()
        if tagihan_now > 0:
            DialogKonfirmasiBayar(
                self.winfo_toplevel(),
                lambda paid: self._warnet_akhir_sesi_setelah_konfirmasi(paid),
                judul="Selesaikan Tagihan",
                rincian=f"Tagihan belum terbayar: {fmt_rp(tagihan_now)}\n"
                        f"Sudah terbayar (LUNAS): {fmt_rp(lunas_now)}",
            ).lift()
        else:
            trans = self._trans_struk()
            self._reset_sesi()
            self._buka_print_struk(trans)

    def _warnet_akhir_sesi_setelah_konfirmasi(self, paid):
        """Konfirmasi tagihan sesi habis warnet: BAYAR -> riwayat lunas;
        TAGIHAN -> isi nama & no HP pelanggan, riwayat tetap belum lunas."""
        if paid:
            self._set_paid(True)
            trans = self._trans_struk()
            self._reset_sesi()
            self._buka_print_struk(trans)
        else:
            DialogPelangganAkhir(
                self.winfo_toplevel(),
                lambda nama, hp: self._simpan_pelanggan_tagihan(nama, hp),
            ).lift()

    def _trans_struk(self):
        """Data transaksi untuk popup cetak struk (panggil SEBELUM sesi direset)."""
        return {
            "pc": getattr(self, "label_kursi", None) or "-",
            "paket": self.paket_aktif or "Main Bebas",
            "menit": int(getattr(self, "menit_dipakai_awal", 0) or 0),
            "harga": int(self._total_setelah_diskon() or 0),
        }

    def _buka_print_struk(self, trans):
        """Buka popup cetak struk di window utama (app._print_receipt)."""
        try:
            app = self.winfo_toplevel()
            if hasattr(app, "_print_receipt"):
                app._print_receipt(trans)
        except Exception as e:
            print(f"[CETAK] gagal buka popup struk {self.label_kursi}: {e}")

    def _simpan_pelanggan_tagihan(self, nama, hp):
        """Catat nama & no HP pelanggan (kasus TAGIHAN) ke baris riwayat."""
        try:
            app = self.winfo_toplevel()
            idx = app._resolve_session_idx(self)
            if idx >= 0 and hasattr(app, '_tree_item_to_index'):
                if idx < len(app.riwayat_meta):
                    app.riwayat_meta[idx]['nama_pelanggan'] = nama
                    app.riwayat_meta[idx]['no_hp'] = hp
                    threading.Thread(
                        target=app._upsert_tx_cloud_from_index, args=(idx,), daemon=True).start()
                if hasattr(app, '_save_riwayat'):
                    app._save_riwayat()
                if hasattr(app, '_refresh_riwayat_summary'):
                    app._refresh_riwayat_summary()
                # Status riwayat tetap BELUM LUNAS (tagihan)
                if hasattr(app, '_set_transaksi_paid_idx'):
                    app._set_transaksi_paid_idx(idx, False)
        except Exception as e:
            print(f"[TAGIHAN] Gagal simpan pelanggan {self.label_kursi}: {e}")
        trans = self._trans_struk()
        self._reset_sesi()
        self._buka_print_struk(trans)

    def _pause_timer(self):
        self._timer_paused = True
        self._timer_was_running = self._timer_job is not None
        if self._timer_job:
            self.after_cancel(self._timer_job)
            self._timer_job = None

    def _resume_timer(self):
        self._timer_paused = False
        if self._timer_was_running:
            self._timer_was_running = False
            if self.is_bebas:
                self._tick_bebas()
            elif self.sisa_waktu > 0:
                self._tick_waktu()

    def _total_menit_terpakai(self):
        if self.is_bebas:
            if self.waktu_mulai is None:
                return self.menit_dipakai_awal
            detik_berjalan = (datetime.now() - self.waktu_mulai).total_seconds()
            return self.menit_dipakai_awal + detik_berjalan / 60
        return None

    def _total_setelah_diskon(self, subtotal=None):
        if subtotal is None:
            subtotal = self.paket_harga_tetap + self.biaya_pesanan
        if self.diskoni <= 0:
            return subtotal
        if self.diskoni_mode == "persen":
            diskon = subtotal * self.diskoni // 100
        else:
            diskon = self.diskoni
        return max(0, subtotal - diskon)

    def _klik_selesai(self):
        if self.sesi_kosong():
            return
        if self.is_bebas:
            menit_total = self._total_menit_terpakai()
            tarif_menit = hitung_tarif_per_menit(self.get_paket_data())
            biaya_waktu = tarif_menit * menit_total
            total_akhir = self._total_setelah_diskon(biaya_waktu + self.biaya_pesanan)
            self._last_transaction_item = self.on_transaksi(
                self.label_kursi, "Main Bebas", self.pesanan_aktif, total_akhir, source='warnet',
                diskoni=self.diskoni, diskoni_mode=self.diskoni_mode, paid=False)
            self._bind_last_transaction()
            DialogPelangganAkhir(
                self.winfo_toplevel(),
                lambda nama, hp: self._simpan_pelanggan_tagihan(nama, hp),
            ).lift()
            return
        else:
            total_akhir = self._total_setelah_diskon()
            pesanan_txt = ", ".join(f"{nm}\u00d7{qty}" for nm, qty in self.pesanan_aktif.items()) or "Tidak ada pesanan"
            paket_txt = f"{self.paket_aktif} ({fmt_rp(self.paket_harga_tetap)})"
            if not messagebox.askyesno(
                    "Selesai",
                    f"Kursi: {self.label_kursi}\n"
                    f"Paket: {paket_txt}\n"
                    f"Pesanan: {pesanan_txt} ({fmt_rp(self.biaya_pesanan)})\n"
                    f"TOTAL: {fmt_rp(total_akhir)}\n\n"
                    f"Akhiri sesi ini?"):
                return
            lunas_now, tagihan_now = self._split_payment()
            if tagihan_now > 0:
                DialogKonfirmasiBayar(
                    self.winfo_toplevel(),
                    lambda paid: self._warnet_selesai_tutup_sesi(paid),
                    judul="Selesaikan Tagihan",
                    rincian=f"Tagihan belum terbayar: {fmt_rp(tagihan_now)}\n"
                            f"Sudah terbayar (LUNAS): {fmt_rp(lunas_now)}",
                ).lift()
                return
        self._warnet_tutup_sesi_final()

    def _warnet_selesai_tutup_sesi(self, paid):
        """Konfirmasi tagihan tombol SELESAI warnet: BAYAR -> riwayat lunas;
        TAGIHAN -> dialog data pelanggan; lalu tutup sesi."""
        if paid:
            self._set_paid(True)
            trans = self._trans_struk()
            self._warnet_tutup_sesi_final()
            self._buka_print_struk(trans)
        else:
            DialogPelangganAkhir(
                self.winfo_toplevel(),
                lambda nama, hp: self._simpan_pelanggan_tagihan(nama, hp),
            ).lift()

    def _warnet_tutup_sesi_final(self):
        self.itemconfig(self._ids['lbl_timer'], text="SELESAI", fill=C_MUTED)
        self.itemconfig(self._ids['lbl_estimasi'], text="")
        app = self.winfo_toplevel()
        if hasattr(app, 'warnet_server') and getattr(self, '_pc_id', None):
            self.pc_locked = True
            self._pc_lock_reason = "selesai_manual"
            self._pc_lock_message = f"Sesi {self.label_kursi} dihentikan admin."
            app.warnet_server.queue_pending_command(
                self._pc_id, "LOCK",
                reason="selesai_manual",
                message=f"Sesi {self.label_kursi} dihentikan admin."
            )
        self._reset_sesi()

    def _reset_sesi(self):
        if self._timer_job:
            self.after_cancel(self._timer_job)
            self._timer_job = None
        self.paket_aktif = None
        self.sisa_waktu = 0
        self.is_bebas = False
        self.waktu_mulai = None
        self.menit_dipakai_awal = 0
        self.pesanan_aktif = {}
        self.biaya_pesanan = 0
        self.paket_harga_tetap = 0
        self.daftar_paket_sesi = []
        self.lunas_paket = []
        self.harga_paket_sesi = []
        self.lunas_pesanan = {}
        self.diskoni           = 0
        self.diskoni_mode      = "nominal"
        self.paid              = True
        self.itemconfig(self._ids['lbl_paket'], text="\u2014", fill=C_MUTED)
        self.itemconfig(self._ids['lbl_timer'], text="00:00:00", fill=C_MUTED)
        self.itemconfig(self._ids['bg'], fill=C_CARD)
        self._warning_blink_on = False
        self.itemconfig(self._ids['lbl_estimasi'], text="")
        self._disable_btn("selesai")
        self._disable_btn("shop")
        self._disable_btn("bayar_lunas")
        self._disable_btn("bayar_belum")
        self.itemconfig(self._ids['btn_shop_txt'], text="SHOP")
        self._update_paid_badge()
        app = self.winfo_toplevel()
        if hasattr(app, '_refresh_warnet_footer'):
            app._refresh_warnet_footer()

    def _klik_pindah(self):
        if self.sesi_kosong():
            return
        semua = self.get_semua_kartu()
        kandidat = [k for k in semua if k is not self and k.sesi_kosong()]
        if not kandidat:
            messagebox.showwarning("Tidak Ada Kursi Kosong",
                                    "Semua kursi lain sedang dipakai.\n"
                                    "Selesaikan dulu salah satu sesi, atau tambah kursi baru.")
            return
        self._buka_dialog_pilih_tujuan(kandidat)

    def _buka_dialog_pilih_tujuan(self, kandidat):
        dlg = ctk.CTkToplevel(self.winfo_toplevel())
        dlg.title("Pindah Kursi")
        dlg.geometry("360x420")
        dlg.configure(fg_color=C_BG)
        dlg.grab_set()
        dlg.resizable(False, False)

        ctk.CTkLabel(dlg, text=f"Pindah dari {self.label_kursi}",
                     font=FONT_TITLE, text_color=C_ACCENT).pack(pady=(18, 4))
        ctk.CTkLabel(dlg, text="Pilih kursi tujuan \u2014 sisa waktu & sesi akan dipindah",
                     font=FONT_BODY, text_color=C_MUTED, wraplength=320).pack(pady=(0, 12))

        scroll = ctk.CTkScrollableFrame(dlg, fg_color=C_BG, height=280)
        scroll.pack(fill="both", expand=True, padx=16)

        for k in kandidat:
            btn = ctk.CTkButton(
                scroll, text=f"  {k.label_kursi}",
                height=40, fg_color=C_CARD, hover_color=C_ACCENT2,
                border_width=1, border_color=C_BORDER,
                font=FONT_BODY, text_color=C_TEXT, anchor="w",
                command=lambda target=k: self._konfirmasi_pindah(target, dlg))
            btn.pack(fill="x", pady=4, padx=4)

        ctk.CTkButton(dlg, text="Batal", height=34, width=120,
                      fg_color=C_RED, font=FONT_SUB, text_color="white",
                      command=dlg.destroy).pack(pady=12)

    def _konfirmasi_pindah(self, target, dlg):
        if not messagebox.askyesno(
                "Konfirmasi Pindah",
                f"Pindahkan sesi dari {self.label_kursi} ke {target.label_kursi}?\n\n"
                f"Sisa waktu / lama main akan dilanjutkan di {target.label_kursi},\n"
                f"dan {self.label_kursi} akan menjadi kosong."):
            return

        target.paket_aktif = self.paket_aktif
        target.is_bebas = self.is_bebas
        target.pesanan_aktif = dict(self.pesanan_aktif)
        target.biaya_pesanan = self.biaya_pesanan
        target.paket_harga_tetap = self.paket_harga_tetap
        target.diskoni = self.diskoni
        target.diskoni_mode = self.diskoni_mode
        target._last_transaction_item = self._last_transaction_item
        target._last_riwayat_idx = self._last_riwayat_idx
        target._last_cloud_id = self._last_cloud_id
        target.paid = self.paid

        if self.is_bebas:
            target.menit_dipakai_awal = self._total_menit_terpakai()
            target.sisa_waktu = 0
            target.waktu_mulai = datetime.now()
            target.itemconfig(target._ids['lbl_paket'], text="Main Bebas (berjalan)", fill=C_GREEN)
            target.itemconfig(target._ids['lbl_timer'], fill=C_YELLOW)
            target.itemconfig(target._ids['lbl_estimasi'], text=f"Total berjalan: {fmt_rp(target.biaya_pesanan)}", fill=C_YELLOW)
        else:
            target.sisa_waktu = self.sisa_waktu
            target.waktu_mulai = datetime.now()
            target.itemconfig(target._ids['lbl_paket'], text=f"{self.paket_aktif} | {fmt_rp(self._total_setelah_diskon())}",
                               fill=C_YELLOW)
            target.itemconfig(target._ids['lbl_timer'], fill=C_ACCENT)

        target._enable_btn("selesai", C_BTN, C_RED)
        target._enable_btn("shop", C_BTN, C_ACCENT2)
        target.itemconfig(target._ids['btn_shop_txt'], text="SHOP")
        if target.is_bebas:
            target._disable_btn("bayar_lunas")
            target._disable_btn("bayar_belum")
        else:
            target._enable_btn("bayar_lunas", "black", "white")
            target._enable_btn("bayar_belum", "black", "white")
            target._update_bayar_buttons()
        target._update_paid_badge()

        if target._timer_job:
            target.after_cancel(target._timer_job)
            target._timer_job = None
        if target.is_bebas:
            target._tick_bebas()
        elif target.sisa_waktu > 0:
            target._tick_waktu()

        self._reset_sesi()
        dlg.destroy()
        messagebox.showinfo("Berhasil Pindah",
                            f"Sesi telah dipindah ke {target.label_kursi}.")


# ═══════════════════════════════════════════════════════════════════════════════
#  APLIKASI UTAMA
# ═══════════════════════════════════════════════════════════════════════════════
class AutoRentApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("RR BILLING PRO — Billing TV System")
        self.geometry("620x900")
        self.resizable(False, False)
        self.configure(fg_color=C_BG)

        # Font default widget tkinter (dialog tk.Toplevel dkk) ikut diperbesar.
        try:
            import tkinter.font as _tkfont
            for _name in ("TkDefaultFont", "TkTextFont", "TkMenuFont",
                          "TkHeadingFont", "TkTooltipFont", "TkCaptionFont",
                          "TkSmallCaptionFont", "TkIconFont"):
                try:
                    _tkfont.nametofont(_name).configure(
                        size=_tkfont.nametofont(_name).cget("size") + 3)
                except Exception:
                    pass
        except Exception:
            pass

        # ── Set ikon window dari logo.png ─────────────────────────────────────
        set_window_icon(self)

        self.current_user  = None
        self.current_role  = None
        self.jumlah_tv     = 0
        self.jumlah_warnet = 0
        self.riwayat_transaksi = []
        self.riwayat_meta = []  # parallel metadata: dicts with keys: source('tv'|'warnet'), pesanan_total(int), total(int)
        self._tree_item_to_index = {}
        self._riwayat_filter_tanggal = None  # "YYYY-MM-DD" atau None = semua tanggal
        self._riwayat_filter_kasir = "SEMUA"  # username kasir atau "SEMUA"
        self._pending_tx_uploads = []  # antrian transaksi menunggu upload ke Firestore
        self._tambah_btn_enabled = True
        self._tambah_warnet_btn_enabled = True
        self._semua_kartu_tv  = []   # daftar semua KartuTV yang sedang ada di Dashboard
        self._semua_kartu_warnet = []  # daftar kartu warnet
        self._timer_service = None

        cfg = ConfigManager.load()
        self.grup_tarif   = self._migrasi_grup_tarif(cfg.get("grup_tarif"), cfg.get("paket_main"))
        self.menu_makanan = cfg.get("menu_makanan",  dict(DEFAULT_MENU_MAKANAN))
        self.menu_minuman = cfg.get("menu_minuman",  dict(DEFAULT_MENU_MINUMAN))
        # Stok makanan/minuman: {"makanan": {nama: qty}, "minuman": {nama: qty}}.
        # Item TANPA entri stok = tidak dilacak (backward compatible, tanpa blokir).
        self.stok     = cfg.get("stok",     {})
        self.stok_min = cfg.get("stok_min", {})
        if not isinstance(self.stok, dict):
            self.stok = {}
        if not isinstance(self.stok_min, dict):
            self.stok_min = {}
        self._stok_tracked_lbl = None  # label badge stok menipis di dashboard
        self.current_tab  = None
        self._bg_image_path = cfg.get("app_bg_image", "")
        self._bersihkan_grup_warnet_bocor()

        # QR pembayaran aktivasi: pastikan qris.png selalu ada di folder app
        # (disalin dari bundle EXE bila folder baru/deploy tanpa file).
        self._ensure_bundled_qris()

        # ── Start Warnet Socket Server ──────────────────────────────────────────
        ws_port = cfg.get("warnet_ws_port", 5001)
        self.warnet_server = WarnetSocketServer(app=self, ws_port=ws_port)
        self.warnet_server.start()

        # ── Start TV WebSocket Hub (Overlay & Lockscreen Android TV) ────────────
        self.tv_overlay_mode = cfg.get("tv_overlay_mode", "always") or "always"
        try:
            self.tv_overlay_last_minutes = int(cfg.get("tv_overlay_last_minutes", 5) or 5)
        except Exception:
            self.tv_overlay_last_minutes = 5
        self.tv_ws_hub = None
        self.tv_test_api = None
        self.tv_media_server = None
        if cfg.get("tv_ws_enabled", True):
            try:
                self.tv_ws_hub = TvWsHub(
                    app=self,
                    port=cfg.get("warnet_tv_ws_port", 8080),
                    get_nama_rental=self._get_nama_rental_dinamis,
                    state_extra=self._media_state_extra,
                )
                self.tv_ws_hub._r4_log_dir = app_log_dir()
                self.tv_ws_hub.start()
                self.tv_test_api = TvTestApi(self.tv_ws_hub, port=cfg.get("warnet_tv_api_port", 8081))
                self.tv_test_api.start()
                self.tv_media_server = TvMediaServer(
                    media_dir=os.path.join(APP_BASE_DIR, "media_promo"),
                    port=cfg.get("warnet_media_port", 8082),
                    qr_page_dir=os.path.join(APP_BASE_DIR, "qr_page"),
                )
                self.tv_media_server.start()
                self._ensure_default_promo()
                self._set_media_default_promo()
            except Exception as e:
                print(f"[TV WS HUB] Gagal start: {e}")
                self.tv_ws_hub = None
                self.tv_test_api = None
                self.tv_media_server = None

        self._timer_service = TimerService()
        self._timer_service.start(self)

        # ── Auto-off TV tanpa paket aktif (anti-kasir nakal) ─────────────────
        # Ambang: 10 menit pertama, 5 menit setelah pernah di-auto-mati (reset
        # ke 10 menit saat paket sah dibuka). 0/nonaktif lewat tv_auto_off_enabled.
        self.tv_auto_off_enabled = bool(cfg.get("tv_auto_off_enabled", True))
        try:
            self.tv_auto_off_first_sec = int(cfg.get("tv_auto_off_first_minutes", 10) or 10) * 60
        except Exception:
            self.tv_auto_off_first_sec = 600
        try:
            self.tv_auto_off_sec = int(cfg.get("tv_auto_off_minutes", 5) or 5) * 60
        except Exception:
            self.tv_auto_off_sec = 300
        self.after(30000, self._tv_idle_guard)
        self.after(20000, self._stok_poll_badge)
        self.after(10000, self._tv_ip_watcher_tick)

        self.protocol("WM_DELETE_WINDOW", self._on_app_close)

        self._show_login()

        threading.Thread(
            target=lambda: get_firebase_auth().ensure_anonymous(),
            daemon=True
        ).start()

    def _on_app_close(self):
        self._stop_session_poller()
        self._stop_idle_watcher()
        self._clear_session_cloud_quick()
        if self._timer_service:
            self._timer_service.stop()
        if getattr(self, 'tv_test_api', None):
            try:
                self.tv_test_api.stop()
            except Exception:
                pass
        if getattr(self, 'tv_ws_hub', None):
            try:
                self.tv_ws_hub.stop()
            except Exception:
                pass
        if getattr(self, 'tv_media_server', None):
            try:
                self.tv_media_server.stop()
            except Exception:
                pass
        self.warnet_server.stop()
        try:
            loop = _WSLoopThread.get_loop()
            for task in asyncio.all_tasks(loop):
                task.cancel()
        except Exception:
            pass
        self.destroy()

    def _get_nama_rental_dinamis(self) -> str:
        """Nama rental dari profil user aktif untuk dikirim ke overlay TV.

        User NON-LIFETIME: selalu RR Billing Pro (nama rental baru tampil
        setelah aktivasi lisensi LIFETIME)."""
        try:
            if not self._lisensi_lifetime():
                return "RR Billing Pro"
            cfg = ConfigManager.load()
            uname = getattr(self, 'current_user', None) or ""
            profil = cfg.get("profil_rental", {})
            if uname and isinstance(profil.get(uname), dict):
                nama = profil[uname].get("nama_rental")
                if nama:
                    return str(nama)
        except Exception:
            pass
        return "RR Billing Pro"

    # ── Watcher IP LAN untuk URL media TV ────────────────────────────────────
    def _tv_media_urls_now(self):
        """URL logo lock (global) + video promosi aktif untuk detail LOCK_SCREEN.

        User NON-LIFETIME: logo default client (drawable bawaan) + video promosi
        BAWAAN (PROMO_VIDEO_DEFAULT) — media custom tidak pernah terpakai."""
        ms = getattr(self, 'tv_media_server', None)
        if not ms or not ms.running:
            return "", ""
        base = f"http://{self._get_lan_ip()}:{ms.port}/media/"
        if not self._lisensi_lifetime():
            logo_url = ""
            promo_url = ""
            default_path = os.path.join(ms.media_dir, PROMO_VIDEO_DEFAULT)
            if os.path.isfile(default_path):
                promo_url = base + quote(PROMO_VIDEO_DEFAULT)
            return logo_url, promo_url
        logo_url = self._tv_logo_url()
        promo_url = ""
        cur = getattr(ms, 'current_media', None) or {}
        if cur.get("type") == "video" and cur.get("filename"):
            promo_url = base + quote(cur["filename"])
        return logo_url, promo_url

    def _promo_default_media_url(self) -> str:
        """URL video promosi BAWAAN (NON-LIFETIME) — dipakai hub TV untuk
        mengirim SHOW_MEDIA sekali saat TV baru terhubung (idle). Kosong bila
        server media tidak jalan / video default tidak ada."""
        ms = getattr(self, 'tv_media_server', None)
        if not ms or not ms.running:
            return ""
        path = os.path.join(ms.media_dir, PROMO_VIDEO_DEFAULT)
        if not os.path.isfile(path):
            return ""
        base = f"http://{self._get_lan_ip()}:{ms.port}/media/"
        try:
            versi = int(os.path.getmtime(path))
        except Exception:
            versi = 0
        return f"{base}{quote(PROMO_VIDEO_DEFAULT)}?v={versi}"

    def _tv_ip_watcher_tick(self):
        """Pantau perubahan IP LAN (dipanggil tiap 30 detik via self.after).

        IP mesin kasir bisa berubah (DHCP/NIC). URL logo/promo yang tersimpan
        di Prefs client TV jadi basi — promo gagal diputar saat TV bangun.
        Bila IP berubah, kirim ulang LOCK_SCREEN dengan URL segar ke semua TV
        yang terkunci (client menyimpan promo_url dari LOCK_SCREEN tanpa
        memutar); TV tidak terkunci memakai URL baru saat lock berikutnya."""
        try:
            self._tv_ip_watcher_check()
        except Exception as e:
            print(f"[TV IP] watcher error: {e}", flush=True)
        try:
            self.after(30000, self._tv_ip_watcher_tick)
        except Exception:
            pass

    def _tv_ip_watcher_check(self):
        try:
            ip = self._get_lan_ip()
            prev = getattr(self, "_tv_ip_watcher_last_ip", None)
            if prev is not None and ip and ip != prev:
                print(f"[TV IP] LAN IP berubah {prev} -> {ip}; refresh URL media TV terkunci",
                      flush=True)
                self._refresh_locked_tv_media_urls()
            self._tv_ip_watcher_last_ip = ip
        except Exception as e:
            print(f"[TV IP] watcher check error: {e}", flush=True)

    def _refresh_locked_tv_media_urls(self):
        hub = self.tv_ws_hub
        if not hub or not hub.running:
            return
        try:
            logo_url, promo_url = self._tv_media_urls_now()
            if not promo_url and not logo_url:
                return
            for mid in hub.get_connected_ids():
                if not hub.is_locked(mid):
                    continue
                detail = dict(hub.locked_summary().get(mid, {}) or {})
                detail["logo_url"] = logo_url
                detail["promo_url"] = promo_url
                hub.send_lock_screen(mid, "WAKTU SEWA HABIS", detail)
        except Exception as e:
            print(f"[TV IP] refresh media error: {e}", flush=True)

    def _get_lan_ip(self) -> str:
        """IP LAN mesin kasir untuk URL media client TV.

        Prioritas:
          1. Konfigurasi manual: tv_media_ip di rr_billing_config.json
          2. UDP connect ke IP TV pertama yang dikenal (daftar_tv + cert
             android_tv_certs) — OS memilih sendiri interface yang menuju TV.
             TV umumnya di LAN 192.168.1.x, sedangkan NIC default-route bisa
             di subnet lain (modem/tethering 172.16.x.x dst.) — URL media
             WAJIB pakai IP yang satu jaringan dengan TV, kalau tidak client
             TV tak pernah bisa mengunduh video/gambar.
          3. NIC lokal yang subnet-nya cocok dengan IP TV.
          4. NIC default-route (perilaku lama).
        """
        # 1) Override manual dari config
        try:
            manual = ConfigManager.load().get("tv_media_ip", "").strip()
            if manual:
                return manual
        except Exception:
            pass
        # Kumpulkan IP TV yang dikenal (daftar_tv + cert pairing)
        tv_ips = set()
        try:
            cfg = ConfigManager.load()
            for item in cfg.get("daftar_tv", []) or []:
                ip = str(item.get("ip", "")).strip()
                if ip:
                    tv_ips.add(ip)
        except Exception:
            pass
        for d in (APP_BASE_DIR, os.path.join(APP_BASE_DIR, "android_tv_certs")):
            try:
                if not os.path.isdir(d):
                    continue
                for fn in os.listdir(d):
                    if fn.startswith("cert_") and fn.endswith(".pem"):
                        ip = fn[5:-4].replace("_", ".")
                        octets = ip.split(".")
                        if (len(octets) == 4
                                and all(o.isdigit() and 0 <= int(o) <= 255 for o in octets)):
                            tv_ips.add(ip)
            except OSError:
                continue
        # 2) NIC lokal yang subnet-nya menuju TV — pakai trik UDP connect:
        #    socket.connect() ke IP TV (tanpa handshake) membuat OS memilih
        #    sendiri interface + IP lokal yang tepat untuk mencapai TV itu.
        #    Paling andal (tidak bergantung enumerasi NIC), bekerja walau TV
        #    sedang mati (UDP connect tidak mengirim paket nyata).
        for tip in sorted(tv_ips, key=lambda x: [int(o) for o in x.split(".")]):
            ip = self._local_ip_for_target(tip)
            if ip:
                return ip
        # 3) NIC lokal yang paling cocok subnet dengan TV
        if tv_ips:
            best, best_score = "", -1
            for local in self._local_ipv4s():
                score = sum(1 for tip in tv_ips if self._same_lan(local, tip))
                if score > best_score:
                    best, best_score = local, score
            if best:
                return best
        # 4) Fallback: NIC default-route
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            try:
                s.connect(("8.8.8.8", 80))
                ip = s.getsockname()[0]
            finally:
                s.close()
            if ip and ip != "0.0.0.0":
                return ip
        except Exception:
            pass
        return "127.0.0.1"

    @staticmethod
    def _local_ip_for_target(target_ip: str) -> str:
        """IP lokal yang akan dipakai OS untuk menjangkau target_ip."""
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            try:
                s.connect((target_ip, 9))
                ip = s.getsockname()[0]
            finally:
                s.close()
            if ip and ip != "0.0.0.0":
                return ip
        except Exception:
            pass
        return ""

    def _local_ipv4s(self):
        """Semua alamat IPv4 lokal mesin (Windows: gethostbyname_ex/getaddrinfo)."""
        ips = set()
        try:
            for ip in socket.gethostbyname_ex(socket.gethostname())[2]:
                ips.add(ip)
        except Exception:
            pass
        try:
            for res in socket.getaddrinfo(socket.gethostname(), None, socket.AF_INET):
                ips.add(res[4][0])
        except Exception:
            pass
        return [ip for ip in ips
                if ip != "127.0.0.1" and not ip.startswith("169.254.")]

    @staticmethod
    def _same_lan(a: str, b: str) -> bool:
        """Cocokkan subnet /24 (skala warnet: 1 router = 1 LAN)."""
        try:
            return a.split(".")[:3] == b.split(".")[:3]
        except Exception:
            return False

    def _tv_logo_url(self) -> str:
        """URL logo lock dengan cache-buster (mtime file) — URL berubah setiap
        file diganti supaya client selalu mengunduh versi terbaru.

        User NON-LIFETIME: logo_url kosong -> client TV memakai drawable
        bawaan (RR BILLING PRO)."""
        if not self._lisensi_lifetime():
            return ""
        ms = getattr(self, 'tv_media_server', None)
        if not ms or not ms.running:
            return ""
        path = os.path.join(ms.media_dir, "logo_lock.png")
        if not os.path.isfile(path):
            return ""
        base = f"http://{self._get_lan_ip()}:{ms.port}/media/"
        try:
            versi = int(os.path.getmtime(path))
        except Exception:
            versi = 0
        return f"{base}logo_lock.png?v={versi}"

    def _simpan_logo_lock(self, path) -> str:
        """Simpan logo lock sebagai PNG ASLI (di-normalisasi) di media_promo.

        Upload boleh PNG/JPG; tanpa normalisasi file JPG yang disalin sebagai
        logo_lock.png bisa gagal tampil di client TV (konten JPEG dengan
        Content-Type PNG). Return path dest; raise Exception bila gagal."""
        ms = getattr(self, 'tv_media_server', None)
        if not ms or not ms.running:
            raise RuntimeError("Server media (port 8082) tidak berjalan.")
        dest = os.path.join(ms.media_dir, "logo_lock.png")
        try:
            from PIL import Image
            img = Image.open(path)
            img.load()
            if img.mode not in ("RGB", "RGBA"):
                img = img.convert("RGBA")
            # Batasi ukuran agar client TV (BitmapFactory) bisa decode tanpa
            # OOM — maks 1920x1080 (Full HD), aspek dipertahankan.
            max_w, max_h = 1920, 1080
            if img.width > max_w or img.height > max_h:
                img.thumbnail((max_w, max_h), Image.LANCZOS)
            img.save(dest, "PNG")
        except Exception:
            # PIL gagal (file korup/dll) — fallback salin mentah.
            import shutil
            shutil.copyfile(path, dest)
        return dest

    def _sembunyikan_file(self, path: str) -> None:
        """Set atribut hidden Windows (tidak terlihat di Explorer) — dipakai
        untuk melindungi video promosi bawaan dari penghapusan."""
        if os.name != "nt":
            return
        try:
            import ctypes
            FILE_ATTRIBUTE_HIDDEN = 0x2
            ctypes.windll.kernel32.SetFileAttributesW(path, FILE_ATTRIBUTE_HIDDEN)
        except Exception:
            pass

    def _ensure_bundled_qris(self) -> None:
        """Pastikan qris.png (QR pembayaran aktivasi) SELALU ada di folder app.

        Saat dijalankan sebagai EXE onefile, gambar tersimpan di bundle
        _MEIPASS (folder temp) — disalin ke folder aplikasi supaya dialog QR
        pembayaran selalu bekerja tanpa menyalin file manual ke tiap deploy."""
        try:
            if os.path.isfile(app_path("qris.png")):
                return
            meipass = getattr(sys, "_MEIPASS", "")
            src = os.path.join(meipass, "qris.png") if meipass else ""
            if src and os.path.isfile(src):
                import shutil
                shutil.copyfile(src, app_path("qris.png"))
                print(f"[QRIS] qris.png disalin dari bundle ke {APP_BASE_DIR}")
        except Exception as e:
            print(f"[QRIS] ensure qris gagal: {e}")

    def _ensure_default_promo(self) -> None:
        """Pastikan video promosi bawaan (pilihan user NON-LIFETIME) SELALU ada
        di media_promo. Jika tidak ada, salin dari bundle PyInstaller
        (folder mode / _MEIPASS) supaya user tanpa lisensi LIFETIME tetap
        mendapat video default."""
        try:
            media_dir = os.path.join(APP_BASE_DIR, "media_promo")
            dest = os.path.join(media_dir, PROMO_VIDEO_DEFAULT)
            if os.path.isfile(dest):
                self._sembunyikan_file(dest)
                return
            sumber = []
            meipass = getattr(sys, "_MEIPASS", "")
            if meipass:
                sumber.append(os.path.join(meipass, "media_promo", PROMO_VIDEO_DEFAULT))
            sumber.append(os.path.join(media_dir, PROMO_VIDEO_DEFAULT))
            for path in sumber:
                if os.path.isfile(path):
                    os.makedirs(media_dir, exist_ok=True)
                    with open(path, "rb") as src_f:
                        data = src_f.read()
                    with open(dest, "wb") as dst_f:
                        dst_f.write(data)
                    self._sembunyikan_file(dest)
                    print(f"[MEDIA] Default promo disalin dari bundle: {PROMO_VIDEO_DEFAULT} ({len(data)} byte)")
                    return
            print(f"[MEDIA] PERINGATAN: default promo {PROMO_VIDEO_DEFAULT} tidak ada di bundle")
        except Exception as e:
            print(f"[MEDIA] ensure default promo gagal: {e}")

    def _set_media_default_promo(self) -> None:
        """Sembunyikan video promosi bawaan; user NON-LIFETIME dikunci memakai
        video default itu (media custom tidak dipakai)."""
        try:
            default_path = os.path.join(APP_BASE_DIR, "media_promo", PROMO_VIDEO_DEFAULT)
            if os.path.isfile(default_path):
                self._sembunyikan_file(default_path)
            if (not self._lisensi_lifetime()
                    and getattr(self, 'tv_media_server', None) is not None):
                self.tv_media_server.set_current("video", PROMO_VIDEO_DEFAULT)
        except Exception as e:
            print(f"[MEDIA] set media default gagal: {e}")

    def _ambil_logo_b64(self, label_widget=None, tampil_error: bool = False) -> str:
        """Pilih file gambar logo rental -> resize <=768px -> dataURL.
        Return "" bila dibatalkan/gagal (versi AutoRentApp)."""
        try:
            self.lift()
            self.focus_force()
        except Exception:
            pass
        try:
            from tkinter import filedialog
            path = filedialog.askopenfilename(
                title="Pilih Logo Rental",
                filetypes=[("Gambar", "*.png *.jpg *.jpeg *.webp *.bmp *.gif"),
                           ("Semua file", "*.*")],
                parent=self)
        except Exception as e:
            if tampil_error:
                messagebox.showerror("Gagal Buka Dialog", str(e), parent=self)
            return ""
        if not path:
            return ""
        return logo_gambar_b64(path, label_widget=label_widget, tampil_error=tampil_error)

    def _profil_pilih_logo(self, logo_box: dict, label_widget):
        b64 = self._ambil_logo_b64(label_widget, tampil_error=True)
        if b64:
            logo_box["b64"] = b64
            try:
                label_widget.configure(text="✔ Logo terpasang", text_color=C_GREEN)
            except Exception:
                pass

    def _profil_hapus_logo(self, logo_box: dict, label_widget):
        logo_box["b64"] = ""
        try:
            label_widget.configure(text="Belum ada logo", text_color=C_MUTED)
        except Exception:
            pass

    def _profil_pilih_qr(self, qr_box: dict, label_widget):
        b64 = self._ambil_logo_b64(label_widget, tampil_error=True)
        if b64:
            qr_box["b64"] = b64
            try:
                label_widget.configure(text="✔ QR pembayaran terpasang", text_color=C_GREEN)
            except Exception:
                pass

    def _profil_hapus_qr(self, qr_box: dict, label_widget):
        qr_box["b64"] = ""
        try:
            label_widget.configure(text="Belum ada QR pembayaran", text_color=C_MUTED)
        except Exception:
            pass

    def _simpan_profil_rental(self, data: dict) -> None:
        """Simpan profil rental user aktif; LIFETIME -> broadcast UPDATE_RENTAL
        ke semua TV (popup kanan atas). Non-LIFETIME -> tersimpan tapi TV
        tetap 'RR Billing Pro'."""
        try:
            cfg = ConfigManager.load()
            profil = cfg.get("profil_rental", {}) or {}
            if not isinstance(profil, dict):
                profil = {}
            lama = profil.get(self.current_user, {}) or {}
            if not isinstance(lama, dict):
                lama = {}
            lama.update(data)
            profil[self.current_user] = lama
            cfg["profil_rental"] = profil
            ConfigManager.save(cfg)
        except Exception as e:
            messagebox.showerror("Gagal Simpan Profil", str(e), parent=self)
            return
        if self._lisensi_lifetime():
            hub = getattr(self, 'tv_ws_hub', None)
            if hub:
                try:
                    n = hub.broadcast_update_rental(data.get("nama_rental", ""))
                    print(f"[TV RENTAL] broadcast UPDATE_RENTAL ke {n} TV")
                except Exception as e:
                    print(f"[TV RENTAL] gagal broadcast: {e}")
            pesan = ("Profil rental tersimpan.\n"
                     f"Popup TV kini menampilkan: {data.get('nama_rental')}")
        else:
            pesan = ("Profil rental tersimpan.\n"
                     "Popup kanan atas TV tetap menampilkan 'RR BILLING PRO'.\n"
                     "Aktifkan lisensi LIFETIME untuk mengganti nama popup TV.")
        messagebox.showinfo("✅ Profil Tersimpan", pesan, parent=self)

    # ══════════════════════════════════════════════════════════════════════════
    #  QR PANGGIL KASIR (halaman web pelanggan -> Firestore -> popup kasir)
    # ══════════════════════════════════════════════════════════════════════════
    # Halaman QR di-host di Firebase Hosting (rrbillingpro.web.app) —
    # bisa diakses pelanggan dari jaringan MANA PUN (internet), bukan cuma
    # Wi-Fi kasir. Override via config qr_page_url bila mau host lain.
    # Local server (TvMediaServer /qr/) tetap melayani sebagai cadangan.
    QR_PAGE_BASE = "https://rrbillingpro.web.app/call.html"
    QR_RATE_LIMIT = 90  # detik antar panggilan per TV
    QR_PIN_TTL = 240  # detik PIN berlaku (sejak ditulis kasir)

    def _qr_token_baru(self) -> str:
        import secrets
        return secrets.token_urlsafe(9).replace("-", "").replace("_", "")[:8].upper()

    def _qr_host_web(self) -> str:
        """Base URL halaman QR (default Firebase Hosting — internet)."""
        try:
            cfg = ConfigManager.load()
            return str(cfg.get("qr_page_url", "") or "").strip() or self.QR_PAGE_BASE
        except Exception:
            return self.QR_PAGE_BASE

    def _qr_url(self, nama_tv: str, kode: str, nama_grup: str = "") -> str:
        from urllib.parse import quote
        owner = self._resolve_license_user() or ""
        url = (f"{self._qr_host_web()}?tv={quote(nama_tv)}&k={kode}"
               f"&o={quote(owner)}")
        if nama_grup:
            url += f"&g={quote(nama_grup)}"
        return url

    def _qr_grup_tv(self, nama_tv: str) -> str:
        """Grup tarif kartu TV (untuk QR web: daftar paket sesuai grup)."""
        try:
            for kartu in list(getattr(self, "_semua_kartu_tv", []) or []):
                if getattr(kartu, "label_tv", "") == nama_tv:
                    return getattr(kartu, "nama_grup", "") or ""
        except Exception:
            pass
        return ""

    def _qr_simpan_png(self, nama_tv: str, url: str) -> str:
        """Simpan QR PNG ke folder qr_panggilan/<TV>.png. Return path ("" gagal)."""
        try:
            import qrcode
            folder = os.path.join(APP_BASE_DIR, "qr_panggilan")
            os.makedirs(folder, exist_ok=True)
            aman = "".join(c if c.isalnum() or c in " -_" else "_" for c in nama_tv).strip()
            path = os.path.join(folder, f"{aman or 'TV'}.png")
            qrcode.make(url).save(path)
            return path
        except Exception as e:
            self._qr_log(f"simpan png gagal: {e}")
            return ""

    def _qr_ip_tv(self, nama_tv: str) -> str:
        try:
            for item in (ConfigManager.load().get("daftar_tv", []) or []):
                if str(item.get("nama", "")) == nama_tv:
                    return str(item.get("ip", "")).strip()
        except Exception:
            pass
        return ""

    def _qr_generate_untuk(self, nama_tv: str) -> str:
        """Ambil kode unik TV; jika belum ada ATAU IP TV berubah -> kode BARU
        (QR lama tidak berlaku lagi). Return kode; "" bila gagal."""
        try:
            cfg = ConfigManager.load()
            peta = cfg.get("qr_call", {}) or {}
            if not isinstance(peta, dict):
                peta = {}
            lama = peta.get(nama_tv) or {}
            if isinstance(lama, dict):
                kode_lama = str(lama.get("kode", ""))
                ip_lama = str(lama.get("ip", ""))
            else:
                kode_lama, ip_lama = "", ""
            ip_kini = self._qr_ip_tv(nama_tv)
            if kode_lama and (not ip_kini or ip_kini == ip_lama):
                # QR lama tetap berlaku — namun PNG di-refresh agar URL terkini
                # (termasuk param grup g= & host server) ikut tercetak ulang.
                self._qr_simpan_png(nama_tv,
                                    self._qr_url(nama_tv, kode_lama,
                                                 self._qr_grup_tv(nama_tv)))
                return kode_lama
            kode = self._qr_token_baru()
            peta[nama_tv] = {"kode": kode, "ip": ip_kini}
            cfg["qr_call"] = peta
            ConfigManager.save(cfg)
            self._qr_simpan_png(nama_tv,
                                self._qr_url(nama_tv, kode, self._qr_grup_tv(nama_tv)))
            self._qr_log(f"{nama_tv}: kode baru ({ip_kini})")
            return kode
        except Exception as e:
            self._qr_log(f"generate gagal: {e}")
            return ""

    def _qr_selaras_semua(self):
        """Auto-generate QR untuk SEMUA kartu TV (saat login/pairing) —
        termasuk deteksi IP berubah -> QR baru. Halaman QR di-host Firebase
        (internet) jadi host stabil — refresh PNG tetap jalan tiap login."""
        for kartu in list(getattr(self, "_semua_kartu_tv", []) or []):
            try:
                self._qr_generate_untuk(kartu.label_tv)
            except Exception:
                pass

    def _qr_log(self, msg: str):
        """Log QR/PIN ke console (jika ada) DAN ke file qr_debug.log —
        exe di-build tanpa console, jadi file ini satu-satunya jejak
        di PC kasir untuk diagnosa dari jarak jauh."""
        try:
            line = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}"
            print(line, flush=True)
            with open(os.path.join(APP_BASE_DIR, "qr_debug.log"), "a", encoding="utf-8") as f:
                f.write(line + "\n")
        except Exception:
            try:
                print(msg, flush=True)
            except Exception:
                pass

    def _qr_pastikan_halaman(self) -> str:
        """Jamin qr_page/call.html ada di sebelah app. Untuk EXE (frozen):
        file di-bundle via spec datas (masuk _MEIPASS) — diekstrak ke
        APP_BASE_DIR/qr_page/ sekali, supaya bisa diedit oleh kasir."""
        try:
            folder = os.path.join(APP_BASE_DIR, "qr_page")
            dst = os.path.join(folder, "call.html")
            if os.path.isfile(dst):
                return dst
            src = ""
            if getattr(sys, "frozen", False) and getattr(sys, "_MEIPASS", ""):
                cand = os.path.join(sys._MEIPASS, "qr_page", "call.html")
                if os.path.isfile(cand):
                    src = cand
            if not src:
                cand = os.path.join(os.path.dirname(os.path.abspath(__file__)), "qr_page", "call.html")
                if os.path.isfile(cand):
                    src = cand
            if src:
                import shutil
                os.makedirs(folder, exist_ok=True)
                shutil.copyfile(src, dst)
                self._qr_log(f"qr_page/call.html diekstrak ke {dst}")
            return dst if os.path.isfile(dst) else ""
        except Exception as e:
            self._qr_log(f"qr_page ekstrak gagal: {e}")
            return ""

    def _start_call_poller(self):
        try:
            self._qr_pastikan_halaman()
        except Exception:
            pass
        # Bersihkan dokumen basi (lebih dari 24 jam) dari calls & qr_sessions —
        # menekan biaya read poller (Firestore tagih 1 read per dokumen yang
        # dikembalikan). Best-effort: di-throttle otomatis saat 429.
        try:
            import time as _t
            from firestore_sync import get_firestore_client
            fs = get_firestore_client()
            potong = _t.time() - 86400
            for koleksi, field_ts in (("calls", "ts"), ("qr_sessions", "created")):
                try:
                    for d in fs.query_all(koleksi, limit=100):
                        ts = d.get(field_ts)
                        try:
                            ts = float(ts) / 1000.0 if ts and float(ts) > 1e12 else (float(ts) if ts else 0)
                        except Exception:
                            ts = 0
                        if ts and ts < potong:
                            fs.delete_document(f"{koleksi}/{d.get('_id', '')}")
                except Exception:
                    pass
        except Exception:
            pass
        try:
            from firestore_sync import CallPoller
            self._call_poller = CallPoller(interval=6.0, limit=5, order_field="ts")
            self._call_poller.start(self._qr_panggilan_masuk)
            self._qr_log("CallPoller dimulai (calls, order ts DESC)")
        except Exception as e:
            self._qr_log(f"CallPoller gagal start: {e}")
        # PIN sesi QR (qr_sessions) — verifikasi kehadiran pelanggan di depan TV
        try:
            from firestore_sync import CallPoller
            self._pin_actif = {}          # tv -> {'sid','pin','t0','owner'}
            self._pin_hide_last = {}      # tv -> ts HIDE_PIN terakhir (kooldown)
            self._pin_loop_stop = threading.Event()
            self._pin_poller = CallPoller(collection="qr_sessions", interval=4.0, limit=10,
                                          order_field="created")
            self._pin_poller.start(self._qr_pin_proses)
            t = threading.Thread(target=self._qr_pin_loop, daemon=True)
            t.start()
            self._pin_loop_thread = t
            self._qr_log("PinPoller dimulai (qr_sessions, order created DESC)")
        except Exception as e:
            self._qr_log(f"PinPoller gagal start: {e}")
        # Booking Online (website /b/<user>) — booking baru milik user ini ->
        # popup kasir + bukti DP; status diubah lewat tombol Konfirmasi/Tolak.
        try:
            from firestore_sync import CallPoller
            self._booking_seen = set()
            self._booking_poller = CallPoller(collection="bookings", interval=8.0,
                                              limit=15, order_field="createdAt")
            self._booking_poller.start(self._qr_booking_masuk)
            self._qr_log("BookingPoller dimulai (bookings, order createdAt DESC)")
        except Exception as e:
            self._qr_log(f"BookingPoller gagal start: {e}")

    def _stop_call_poller(self):
        p = getattr(self, "_call_poller", None)
        if p:
            try:
                p.stop()
            except Exception:
                pass
        self._call_poller = None
        p = getattr(self, "_pin_poller", None)
        if p:
            try:
                p.stop()
            except Exception:
                pass
        self._pin_poller = None
        ev = getattr(self, "_pin_loop_stop", None)
        if ev:
            ev.set()
        self._pin_loop_stop = None
        try:
            self.after(0, self._qr_pin_hapus_popup)
        except Exception:
            pass
        self._pin_actif = {}
        self._pin_hide_last = {}
        p = getattr(self, "_booking_poller", None)
        if p:
            try:
                p.stop()
            except Exception:
                pass
        self._booking_poller = None

    # ── PIN Sesi QR ──────────────────────────────────────────────────────────
    _PIN_HURUF = "ABCDEFGHJKLMNPQRSTUVWXYZ"
    _PIN_ANGKA = "23456789"

    def _qr_pin_baru(self) -> str:
        """PIN 5 karakter alfanumerik: 2 huruf + 3 angka (tanpa 0/O/1/I).
        Selalu BERUBAH setiap kali QR discan (satu PIN per sesi)."""
        import random
        h = [random.choice(self._PIN_HURUF) for _ in range(2)]
        a = [random.choice(self._PIN_ANGKA) for _ in range(3)]
        pin = list(h + a)
        random.shuffle(pin)
        return "".join(pin)

    def _qr_cari_tv_oleh_kode(self, kode: str):
        """Cari nama TV di config qr_call yang kodenya cocok — untuk QR lama
        yang nama TV-nya sudah berubah (mis. 'TV 1' -> '1') tapi kodenya
        masih tercatat. Return (tv, pi) atau (None, None)."""
        try:
            peta = (ConfigManager.load().get("qr_call", {}) or {})
            for tv, pi in peta.items():
                if isinstance(pi, dict) and str(pi.get("kode", "") or "") == str(kode):
                    return tv, pi
        except Exception:
            pass
        return None, None

    def _qr_pin_proses(self, doc: dict):
        """Thread PinPoller: state machine sesi PIN (dipanggil di thread).

        Catatan: TIDAK memakai guard _pin_seen — dokumen yang sama wajib
        diproses berulang setiap poll, karena alur verifikasi PIN (pin_user
        dari web) hanya bisa dilihat pada poll berikutnya setelah web
        menulis pin_user. Guard lama justru memblokir verifikasi selamanya.
        """
        try:
            did = str(doc.get("_id", ""))
            if not did:
                return
            tv = str(doc.get("tv", ""))
            kode = str(doc.get("kode", ""))
            if not tv or not kode:
                return
            peta = (ConfigManager.load().get("qr_call", {}) or {})
            pi = peta.get(tv) if isinstance(peta, dict) else None
            if not isinstance(pi, dict) or kode != str(pi.get("kode", "")):
                # QR lama / nama TV berubah: coba cocokkan lewat kode dulu.
                tv2, pi2 = self._qr_cari_tv_oleh_kode(kode)
                if tv2 is not None and isinstance(pi2, dict):
                    tv, pi = tv2, pi2
                    self._qr_log(f"PIN: sesi {did} tv={doc.get('tv')!r} dipetakan ke {tv!r} via kode")
                else:
                    self._qr_log(f"PIN tolak: tv={tv!r} kode={kode!r} tidak dikenal di qr_call")
                    return  # QR TV lain / kedaluwarsa — tolak
            import time as _t
            now = _t.time()
            status = str(doc.get("status", "awaiting"))
            owner = str(doc.get("owner", "")).strip()
            aktif = self._pin_actif.get(tv)
            sid_sekarang = (aktif or {}).get("sid")

            if status == "awaiting":
                pin = str(doc.get("pin", "") or "")
                if not pin:
                    # PIN sudah di memori untuk sesi yang sama (write Firestore
                    # tertunda) -> pakai ulang, jangan generate PIN baru.
                    if aktif and sid_sekarang == did and (aktif or {}).get("pin"):
                        pin = (aktif or {}).get("pin")
                        self._set_pin_doc(did, {"pin": pin, "pin_set_at": int(now * 1000)})
                        self._qr_pin_set_tv(owner, tv, pin)
                        return
                    # Satu sesi aktif per TV: sesi lama hangus diganti yang baru
                    if aktif and sid_sekarang and sid_sekarang != did:
                        self._qr_pin_selesai(tv, sid_sekarang, reason="diganti")
                    pin = self._qr_pin_baru()
                    self._set_pin_doc(did, {"pin": pin, "pin_set_at": int(now * 1000)})
                    self._pin_actif[tv] = {"sid": did, "pin": pin, "t0": now, "owner": owner}
                    self._qr_pin_set_tv(owner, tv, pin)
                    return
                # PIN sudah terpasang (sinkronisasi lintas-restart)
                self._pin_actif[tv] = {"sid": did, "pin": pin, "t0": now, "owner": owner}
                pin_set_at = float(doc.get("pin_set_at", 0) or 0) / 1000.0
                if pin_set_at and now - pin_set_at > self.QR_PIN_TTL:
                    self._set_pin_doc(did, {"status": "expired", "reason": "ttl"})
                    self._qr_pin_selesai(tv, did, reason="expired")
                    return
                pin_user = str(doc.get("pin_user", "") or "").strip().upper()
                if pin_user:
                    if pin_user == str(pin).strip().upper():
                        self._set_pin_doc(did, {"status": "verified", "pin_user": ""})
                        self._qr_pin_selesai(tv, did, reason="ok", hapus_doc=False)
                    else:
                        tries = int(doc.get("tries", 0) or 0) + 1
                        if tries >= 3:
                            self._set_pin_doc(did, {"status": "blocked", "pin_user": "", "tries": tries})
                            self._qr_pin_selesai(tv, did, reason="blocked")
                        else:
                            self._set_pin_doc(did, {"tries": tries, "pin_user": ""})
            elif status in ("blocked", "expired"):
                self._qr_pin_selesai(tv, did, reason=status)
            elif status == "verified":
                self._qr_pin_selesai(tv, did, reason="ok", hapus_doc=False)
                created_ms = float(doc.get("created", 0) or 0)
                if created_ms and now - created_ms / 1000.0 > 300:
                    try:
                        from firestore_sync import get_firestore_client
                        get_firestore_client().delete_document(f"qr_sessions/{did}")
                    except Exception:
                        pass
        except Exception as e:
            self._qr_log(f"pin proses error: {e}")

    def _set_pin_doc(self, did: str, data: dict):
        try:
            from firestore_sync import get_firestore_client
            get_firestore_client().set_document(f"qr_sessions/{did}", data, merge=True)
        except Exception as e:
            self._qr_log(f"update sesi {did} gagal: {e}")

    def _qr_pin_set_tv(self, owner: str, tv: str, pin: str):
        """Tampilkan PIN di layar TV (kiri atas) via WebSocket hub (app RR Billing TV)."""
        try:
            hub = getattr(self, "tv_ws_hub", None)
            if hub is not None:
                hub.send_show_pin(tv, pin)
            else:
                self._qr_log(f"tampil PIN TV {tv} gagal: hub WS tidak aktif")
        except Exception as e:
            self._qr_log(f"tampil PIN TV {tv} gagal: {e}")

    def _qr_pin_clear_tv(self, owner: str, tv: str):
        """Sembunyikan PIN di TV — idempotent dengan kooldown 10 dtk supaya
        HIDE_PIN yang gagal (WS putus sesaat) tertolong kiriman berikutnya
        (dokumen verified dipoll sampai ~900 dtk)."""
        if not tv:
            return
        try:
            import time as _t
            last = getattr(self, "_pin_hide_last", {}).get(tv, 0)
            if _t.time() - last < 10:
                return
            self._pin_hide_last[tv] = _t.time()
            hub = getattr(self, "tv_ws_hub", None)
            if hub is not None:
                hub.send_hide_pin(tv)
            else:
                self._qr_log(f"sembunyi PIN TV {tv} gagal (hub WS tidak aktif)")
        except Exception as e:
            self._qr_log(f"sembunyi PIN TV {tv} gagal: {e}")

    def _qr_pin_selesai(self, tv: str, sid: str, reason: str = "", hapus_doc: bool = True):
        """Akhiri sesi PIN: sembunyikan overlay TV, hapus dokumen sesi
        (kecuali status verified yang butuh dibaca web). HIDE_PIN dikirim
        SELALU untuk alasan terminal — bukan hanya saat entry masih ada —
        karena sesi bisa berakhir setelah entry-nya dipop."""
        akt = self._pin_actif.get(tv)
        if akt and (not sid or akt.get("sid") == sid):
            self._pin_actif.pop(tv, None)
        if reason:
            self._qr_log(f"sesi PIN {tv}: {reason}")
            self._qr_pin_clear_tv("", tv)
        if hapus_doc and sid:
            try:
                from firestore_sync import get_firestore_client
                get_firestore_client().delete_document(f"qr_sessions/{sid}")
            except Exception as e:
                self._qr_log(f"hapus sesi {sid} gagal: {e}")

    def _qr_pin_loop(self):
        """Re-trigger overlay PIN di TV tiap ~6 detik selama sesi aktif,
        karena overlay app TV auto-hilang 8 detik. Sesi lewat TTL dihanguskan."""
        import time as _t
        while True:
            ev = getattr(self, "_pin_loop_stop", None)
            if ev is None or ev.is_set():
                break
            now = _t.time()
            for tv, akt in list((self._pin_actif or {}).items()):
                try:
                    sid = akt.get("sid")
                    # Kalau dokumen sesi sudah hilang (dihapus lewat jalur lain /
                    # sesi web dibatalkan), entry memori ini YATIM: loop bakal
                    # terus menampilkan PIN selamanya. Cek keberadaan dokumen
                    # tiap tayang — 2 kali berturut-turut hilang = hangus.
                    if sid and (akt.get("doc_none") or 0) < 2:
                        try:
                            from firestore_sync import get_firestore_client
                            d = get_firestore_client().get_document(f"qr_sessions/{sid}")
                            if d is None:
                                akt["doc_none"] = (akt.get("doc_none") or 0) + 1
                                if akt["doc_none"] >= 2:
                                    self._qr_pin_selesai(tv, sid, reason="hilang", hapus_doc=False)
                                    continue
                            else:
                                akt["doc_none"] = 0
                        except Exception:
                            pass
                    if now - (akt.get("t0") or now) > self.QR_PIN_TTL:
                        self._set_pin_doc(sid, {"status": "expired", "reason": "ttl"})
                        self._qr_pin_selesai(tv, sid, reason="expired")
                        continue
                    self._qr_pin_set_tv(akt.get("owner", ""), tv, akt.get("pin", ""))
                    # Sesi bisa berakhir TEPAT saat SHOW ini dikirim (verifikasi
                    # PIN lewat poller) — SHOW terakhir bisa mendarat di TV
                    # SETELAH HIDE dari _qr_pin_selesai. Cek ulang & balas HIDE
                    # kalau entry sesi sudah tidak aktif lagi.
                    if (self._pin_actif.get(tv) or {}).get("sid") != akt.get("sid"):
                        self._qr_pin_clear_tv("", tv)
                except Exception as e:
                    self._qr_log(f"pin loop {tv}: {e}")
            _t.sleep(6)

    def _qr_parse_item_text(self, s):
        """Parse string item lama (dari call.html versi lama / keluhan) jadi
        struktur item: {tipe, nama, qty, harga, menit, status}."""
        s = str(s or "").strip()
        tipe, nama, harga, qty, menit = "makanan", s, 0, 1, 0
        m = re.search(r"\(\s*Rp\s*([\d.,]+)\s*\)", s, re.I)
        if m:
            try:
                harga = int(re.sub(r"[^\d]", "", m.group(1)) or 0)
            except Exception:
                harga = 0
            nama = s[:m.start()].strip().lstrip("•·—").strip()
        nama = nama.replace("🍔", "").replace("🥤", "").replace("➕", "").strip()
        low = nama.lower()
        if low.startswith("paket"):
            tipe = "paket"
            if ":" in nama:
                nama = nama.split(":", 1)[1].strip()
        elif "🥤" in s:
            tipe = "minuman"
        elif "🍔" in s:
            tipe = "makanan"
        if not nama:
            nama = str(s or "?").strip() or "?"
        return {"tipe": tipe, "nama": nama, "qty": qty, "harga": harga,
                "menit": menit, "status": "baru"}

    def _qr_parse_items(self, doc: dict) -> list:
        """Kumpulkan item terstruktur dari dokumen calls. Format BARU (`items`)
        dikirim web (t: paket|makanan|minuman, n, q, h, men, g); format LAMA
        (`item` string) dijadikan fallback lewat _qr_parse_item_text."""
        raw = doc.get("items")
        if isinstance(raw, list) and raw:
            out = []
            for x in raw:
                try:
                    if isinstance(x, dict):
                        tipe = str(x.get("t") or x.get("tipe") or "makanan")
                        if tipe not in ("paket", "makanan", "minuman"):
                            tipe = "makanan"
                        out.append({
                            "tipe": tipe,
                            "nama": str(x.get("n") or x.get("nama") or "?"),
                            "qty": int(x.get("q") or x.get("qty") or 1),
                            "harga": int(x.get("h") or x.get("harga") or 0),
                            "menit": int(x.get("men") or x.get("menit") or 0),
                            "status": "baru",
                        })
                    else:
                        out.append(self._qr_parse_item_text(x))
                except Exception:
                    continue
            return out
        item_str = doc.get("item")
        if isinstance(item_str, list) and item_str:
            return [self._qr_parse_item_text(x) for x in item_str]
        return []

    def _qr_pesan_log_load(self) -> list:
        try:
            with QR_PESAN_LOG_LOCK:
                if os.path.exists(QR_PESAN_LOG):
                    with open(QR_PESAN_LOG, "r", encoding="utf-8") as f:
                        data = json.load(f)
                    return data if isinstance(data, list) else []
        except Exception as e:
            print(f"[QR] gagal baca riwayat pesanan: {e}")
        return []

    def _qr_pesan_log_save(self, rows: list):
        try:
            with QR_PESAN_LOG_LOCK:
                with open(QR_PESAN_LOG, "w", encoding="utf-8") as f:
                    json.dump(rows, f, ensure_ascii=False, indent=1)
        except Exception as e:
            print(f"[QR] gagal simpan riwayat pesanan: {e}")

    def _qr_pesan_log_append(self, order: dict):
        rows = [r for r in self._qr_pesan_log_load() if r.get("id") != order.get("id")]
        rows.append(order)
        self._qr_pesan_log_save(rows)

    def _qr_pesan_log_update(self, oid: str, items, status=None):
        """Perbarui item-status (dan opsional status order) di riwayat lokal."""
        rows = self._qr_pesan_log_load()
        for r in rows:
            if r.get("id") != oid:
                continue
            if items is not None:
                r["items"] = items
            if status:
                r["status"] = status
            break
        self._qr_pesan_log_save(rows)

    def _qr_cari_kartu(self, tv):
        """Cari kartu TV di dashboard berdasar nama (label) dari QR call."""
        tv_s = str(tv or "").strip()
        if not tv_s:
            return None
        bagi = tv_s.upper()
        for k in getattr(self, "_semua_kartu_tv", []):
            lt = str(getattr(k, "label_tv", "") or "").strip()
            lu = lt.upper()
            if lu == bagi:
                return k
            if tv_s.isdigit():
                if lu == f"TV {tv_s}" or lu == f"TV{tv_s}":
                    return k
        return None

    def _logo_rental_b64(self) -> str:
        """Logo rental user aktif (dataURL PNG) untuk call_meta — kosong bila
        belum dipasang. Dipakai halaman booking online (header & status)."""
        try:
            cfg = ConfigManager.load()
            uname = getattr(self, "current_user", None) or ""
            profil = cfg.get("profil_rental", {}) or {}
            if isinstance(profil, dict) and isinstance(profil.get(uname), dict):
                logo = profil[uname].get("logo") or ""
                if isinstance(logo, str) and logo.startswith("data:image/"):
                    return logo
        except Exception:
            pass
        return ""

    def _get_daftar_tv_nama(self) -> list:
        """Nama kartu TV dari config — daftar perangkat di halaman booking web."""
        daftar = []
        try:
            cfg = ConfigManager.load()
            tvs = cfg.get("daftar_tv", []) or []
            if isinstance(tvs, list):
                for t in tvs:
                    if isinstance(t, dict):
                        nama = str(t.get("nama", "")).strip()
                        if nama:
                            daftar.append(nama)
        except Exception:
            pass
        return daftar

    def _booking_push_tv_status(self):
        """Push status TV (aktif/tidak) + daftar perangkat (TV & PC warnet,
        lengkap dengan grup tarif) ke call_meta tiap 8 dtk — halaman booking
        web menampilkan kartu perangkat dengan status realtime."""
        def worker():
            try:
                from firestore_sync import FirestoreClient
                owner = self._resolve_license_user()
                if not owner:
                    return
                tv_status = {}
                for kartu in list(getattr(self, "_semua_kartu_tv", []) or []):
                    try:
                        aktif = (bool(getattr(kartu, "paket_aktif", None))
                                 and getattr(kartu, "sisa_waktu", 0) > 0
                                 and not getattr(kartu, "_timer_paused", False)
                                 and not getattr(kartu, "_billing_paused", False))
                        tv_status[str(getattr(kartu, "label_tv", ""))] = {
                            "aktif": bool(aktif),
                            "paket": str(getattr(kartu, "paket_aktif", "") or "-") if aktif else "",
                        }
                    except Exception:
                        continue
                devices = {"TV": [], "PC": []}
                try:
                    cfg = ConfigManager.load()
                    tvs = cfg.get("daftar_tv", []) or []
                    if isinstance(tvs, list):
                        for t in tvs:
                            if not isinstance(t, dict):
                                continue
                            nama = str(t.get("nama", "")).strip()
                            if not nama:
                                continue
                            st = tv_status.get(nama, {}) or {}
                            devices["TV"].append({
                                "nama": nama,
                                "grup": str(t.get("nama_grup", "") or "Reguler"),
                                "aktif": bool(st.get("aktif", False)),
                                "paket": str(st.get("paket", "") or ""),
                            })
                except Exception:
                    pass
                try:
                    for kursi in list(getattr(self, "_semua_kartu_warnet", []) or []):
                        try:
                            nama_k = str(getattr(kursi, "label_kursi", "") or "")
                            if not nama_k:
                                continue
                            aktif = (bool(getattr(kursi, "paket_aktif", None))
                                     and getattr(kursi, "sisa_waktu", 0) > 0
                                     and not getattr(kursi, "_timer_paused", False))
                            devices["PC"].append({
                                "nama": nama_k,
                                "grup": str(getattr(kursi, "nama_grup", "") or "Reguler"),
                                "aktif": bool(aktif),
                                "paket": str(getattr(kursi, "paket_aktif", "") or "") if aktif else "",
                            })
                        except Exception:
                            continue
                except Exception:
                    pass
                pus = (ConfigManager.load().get("profil_rental", {}) or {}).get(owner, {}) or {}
                if not isinstance(pus, dict):
                    pus = {}
                FirestoreClient().set_document(
                    f"call_meta/{owner}",
                    {"tv_status": tv_status,
                     "devices": devices,
                     "daftar_tv": self._get_daftar_tv_nama(),
                     "no_hp": str(pus.get("no_hp") or pus.get("hp") or "").strip(),
                     "updatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
                    merge=True)
            except Exception as e:
                print(f"[BOOKING] push tv_status gagal: {e}")
        try:
            threading.Thread(target=worker, daemon=True).start()
        except Exception:
            pass
        try:
            self.after(8000, self._booking_push_tv_status)
        except Exception:
            pass

    def _qr_panggilan_masuk(self, doc: dict):
        """Dipanggil thread CallPoller — validasi, simpan ke riwayat lokal
        (SEBELUM dokumen dihapus), lalu tampilkan di main thread.

        Catatan: panggilan yang masuk dalam jendela rate-limit TIDAK dibuang —
        tetap dicatat di riwayat (agar tidak hilang) dan dokumennya dihapus."""
        try:
            did = str(doc.get("_id", ""))
            if not did or did in self._qr_seen:
                return
            kode = str(doc.get("kode", ""))
            tv = str(doc.get("tv", ""))
            if not tv or not kode:
                return
            peta = (ConfigManager.load().get("qr_call", {}) or {})
            pi = peta.get(tv) if isinstance(peta, dict) else None
            if not isinstance(pi, dict) or kode != str(pi.get("kode", "")):
                tv2, pi2 = self._qr_cari_tv_oleh_kode(kode)
                if tv2 is not None and isinstance(pi2, dict):
                    tv, pi = tv2, pi2
                    self._qr_log(f"panggilan {did}: tv={doc.get('tv')!r} dipetakan ke {tv!r} via kode")
                else:
                    self._qr_log(f"panggilan {did} diabaikan: tv={tv!r} kode={kode!r} tak dikenal")
                    return  # milik kasir/QR lain — biarkan, pemiliknya yang menghapus
            self._qr_seen.add(did)
            import time as _t
            now = _t.time()
            rate_ok = now - self._qr_last_call.get(tv, 0) >= self.QR_RATE_LIMIT
            if rate_ok:
                self._qr_last_call[tv] = now
            jenis = str(doc.get("jenis", "keluhan"))
            item_text = doc.get("item") or []
            items = self._qr_parse_items(doc) or []
            catatan = str(doc.get("catatan", ""))
            # ══ RIWAYAT LOKAL — disimpan SEBELUM dokumen calls dihapus ══
            order = {
                "id": did,
                "waktu": datetime.now().strftime("%Y-%m-%d %H:%M"),
                "tv": tv,
                "jenis": jenis,
                "item": [str(x) for x in item_text],
                "items": items,
                "catatan": catatan,
                "status": "baru",
            }
            self._qr_pesan_log_append(order)
            self.after(0, self._qr_riwayat_refresh)
            try:
                from firestore_sync import get_firestore_client
                get_firestore_client().delete_document(f"calls/{did}")
            except Exception as e:
                self._qr_log(f"gagal hapus panggilan {did}: {e}")
            if not rate_ok:
                self._qr_log(f"panggilan {did} masuk jendela rate-limit — dicatat di riwayat")
                return
            self._qr_log(f"panggilan masuk: tv={tv} did={did} jenis={jenis} item={len(items)}")
            payload = {"tv": tv, "jenis": jenis, "item": item_text,
                       "items": items, "catatan": catatan, "id": did}
            self.after(0, lambda p=payload: self._qr_tampil_panggilan(p))
        except Exception as e:
            self._qr_log(f"panggilan masuk error: {e}")

    def _qr_booking_masuk(self, doc: dict):
        """Dipanggil thread BookingPoller — validasi milik user ini & status
        'baru', lalu tampilkan popup kasir di main thread."""
        try:
            did = str(doc.get("_id", ""))
            if not did or did in getattr(self, "_booking_seen", set()):
                return
            if str(doc.get("status", "")) != "baru":
                return
            owner_b = str(doc.get("owner", "")).strip().lower()
            uname = (getattr(self, "current_user", None) or "").strip().lower()
            if not uname or owner_b != uname:
                return  # booking milik rental lain
            self._booking_seen.add(did)
            self._qr_log(f"booking masuk: {did} {str(doc.get('namaPelanggan', ''))} "
                         f"{str(doc.get('perangkat', ''))} "
                         f"metode={str(doc.get('metode', ''))} "
                         f"bukti={'ada' if doc.get('bukti') else '-'}")
            self.after(0, lambda d=dict(doc): self._qr_tampil_booking(d))
        except Exception as e:
            self._qr_log(f"booking masuk error: {e}")

    def _qr_tampil_booking(self, d: dict, bunyi: bool = True):
        """Popup detail booking (baru masuk / klik dari tab Booking): data per
        kode + Lihat Bukti + Konfirmasi / Tolak. Hasil ditulis balik ke Firestore."""
        did = str(d.get("_id", ""))
        from tkinter import simpledialog
        win = ctk.CTkToplevel(self)
        win.title("📅 Booking Online")
        win.geometry("440x600")
        win.attributes("-topmost", True)
        win.transient(self)
        try:
            win.after(100, lambda: win.attributes("-topmost", False))
        except Exception:
            pass

        ctk.CTkLabel(win, text="📅 Booking Baru", font=("Russo One", 15, "bold"),
                     text_color=C_ACCENT2).pack(pady=(14, 4))
        sub = f"{str(d.get('namaPelanggan', '-'))} • {str(d.get('perangkat', '-'))}"
        ctk.CTkLabel(win, text=sub, font=FONT_SUB, text_color=C_TEXT).pack(pady=(0, 8))

        # Body scrollable: detail panjang (pesanan banyak / catatan) TIDAK
        # boleh menutupi tombol Konfirmasi / Tolak / Lihat Bukti.
        body = ctk.CTkScrollableFrame(win, fg_color=C_PANEL, corner_radius=10)
        body.pack(fill="both", expand=True, padx=18, pady=(0, 6))
        rows = [
            ("Kode", did[:8].upper()),
            ("Nama", str(d.get("namaPelanggan", "-"))),
            ("WhatsApp", str(d.get("noHp", "-"))),
            ("Perangkat", str(d.get("perangkat", "-"))),
            ("Paket", f"{d.get('grup', '')} → {d.get('paket', '')}"),
            ("Mulai", f"{d.get('tanggal', '')} {d.get('jam', '')}"),
            ("Total", f"Rp {int(d.get('totalHarga', 0) or 0):,}"),
            ("Metode", {"lunas": "Lunas (transfer penuh)", "dp": "DP (transfer sebagian)"}
             .get(str(d.get("metode", "")), "Bayar di tempat")),
        ]
        sb = str(d.get("statusBayar", "") or "")
        if sb == "lunas_transfer":
            rows.append(("Status Bayar", "✅ LUNAS VIA TRANSFER"))
            nom_t = int(d.get("nominalTransfer", 0) or 0)
            if nom_t:
                rows.append(("Ditransfer", f"Rp {nom_t:,}"))
        elif sb == "dp":
            dpn = int(d.get("nominalDp", 0) or 0)
            sisa = int(d.get("sisaBayar", 0) or 0)
            rows.append(("Status Bayar", f"💸 TAGIHAN SISA — Rp {sisa:,}"))
            if dpn:
                rows.append(("DP Dibayar", f"Rp {dpn:,}  (sisa Rp {sisa:,})"))
        else:
            rows.append(("Status Bayar", "⏳ BELUM BAYAR (bayar di tempat)"))
        if d.get("metode") in ("dp", "lunas"):
            try:
                pus = (ConfigManager.load().get("profil_rental", {}) or {}).get(self.current_user, {}) or {}
                if isinstance(pus, dict):
                    nama_dana = str(pus.get("nama_dana", "") or "").strip()
                    no_dana = str(pus.get("no_dana", "") or "").strip()
                    if no_dana or nama_dana:
                        rows.append(("DANA", f"{nama_dana} — {no_dana}" if nama_dana else no_dana))
            except Exception:
                pass
        pesanan = d.get("pesanan") or {}
        if isinstance(pesanan, dict) and pesanan:
            try:
                mkd = {**dict(getattr(self, "menu_makanan", {}) or {}),
                       **dict(getattr(self, "menu_minuman", {}) or {})}
                txt = ", ".join(f"{nm} x{q}" for nm, q in pesanan.items() if int(q or 0) > 0)
                harga_p = sum(int(mkd.get(nm, 0) or 0) * int(q or 0) for nm, q in pesanan.items() if int(q or 0) > 0)
                if txt:
                    rows.append(("Pesanan", txt))
                    rows.append(("+Pesanan", f"Rp {harga_p:,}"))
            except Exception:
                pass
        for label, value in rows:
            r = ctk.CTkFrame(body, fg_color="transparent")
            r.pack(fill="x", padx=14, pady=3)
            ctk.CTkLabel(r, text=label, width=90, anchor="w",
                         font=FONT_LABEL, text_color=C_MUTED).pack(side="left")
            ctk.CTkLabel(r, text=str(value), anchor="w",
                         font=FONT_BODY, text_color=C_TEXT).pack(side="left", expand=True, fill="x")
        catatan = str(d.get("catatan", "") or "")
        if catatan:
            ctk.CTkLabel(body, text="Catatan: " + catatan, wraplength=350,
                         font=FONT_SMALL, text_color=C_MUTED,
                         anchor="w", justify="left").pack(fill="x", padx=14, pady=(2, 10))

        btns = ctk.CTkFrame(win, fg_color="transparent")
        btns.pack(fill="x", padx=18, pady=(0, 14))
        label_state = ctk.CTkLabel(win, text="", font=FONT_SMALL, text_color=C_MUTED)
        label_state.pack(pady=(0, 8))

        st_sekarang = str(d.get("status", "") or "")
        btns_all = []

        def buka_bukti():
            bukti = str(d.get("bukti", "") or "")
            if not bukti.startswith("data:image/"):
                messagebox.showinfo("Tidak Ada", "Tidak ada bukti transfer pada booking ini.", parent=win)
                return
            try:
                import base64 as _b64
                folder = app_path("invoices_bukti")
                os.makedirs(folder, exist_ok=True)
                raw = bukti.split(",", 1)[1]
                ext = "png" if "image/png" in bukti else "jpg"
                path = os.path.join(folder, f"booking_bukti_{did}.{ext}")
                with open(path, "wb") as f:
                    f.write(_b64.b64decode(raw))
                os.startfile(path)
            except Exception as e:
                messagebox.showerror("Gagal", f"Gagal membuka bukti:\n{e}", parent=win)

        def _simpan(patch: dict, ok_msg: str, err_msg: str, after_ok=None):
            for b in btns_all:
                try:
                    b.configure(state="disabled")
                except Exception:
                    pass
            try:
                label_state.configure(text="⏳ Menyimpan...", text_color=C_MUTED)
            except Exception:
                pass

            def worker():
                ok = False
                try:
                    from firestore_sync import get_firestore_client
                    ok = get_firestore_client().set_document(f"bookings/{did}", patch, merge=True)
                except Exception as e:
                    self._qr_log(f"booking {did} update gagal: {e}")

                def done():
                    if ok:
                        try:
                            label_state.configure(text=ok_msg, text_color=C_GREEN)
                        except Exception:
                            pass
                        self._qr_log(f"booking {did} -> {patch.get('status')}")
                        if after_ok is not None:
                            try:
                                after_ok()
                            except Exception:
                                pass
                        try:
                            win.after(700, win.destroy)
                        except Exception:
                            pass
                    else:
                        try:
                            label_state.configure(text=err_msg, text_color=C_RED)
                            for b in btns_all:
                                b.configure(state="normal")
                        except Exception:
                            pass

                try:
                    self.after(0, done)
                except Exception:
                    pass

            try:
                threading.Thread(target=worker, daemon=True).start()
            except Exception:
                pass

        def konfirmasi():
            _simpan({"status": "dikonfirmasi", "kasir": str(getattr(self, "current_user", "")),
                     "alasan": "", "updatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
                    "✅ Booking dikonfirmasi", "✖ Gagal menyimpan — cek internet",
                    after_ok=konfirmasi_after_ok)

        def konfirmasi_after_ok():
            """Setelah status jadi dikonfirmasi — catat ke riwayat kas langsung
            (anti-curang: kasir tidak bisa menerima booking tanpa mencatat)."""
            if self._riwayat_idx_by_booking(did) >= 0:
                return
            try:
                perangkat = str(d.get("perangkat", "") or "").strip()
                if not perangkat:
                    return
                pesanan = {str(k): int(v or 0) for k, v in (d.get("pesanan") or {}).items()
                           if int(v or 0) > 0}
                total = int(d.get("totalHarga", 0) or 0)
                metode = str(d.get("metode", "") or "")
                sb = str(d.get("statusBayar", "") or "")
                paid_awal = (metode == "lunas") or (sb == "lunas_transfer")
                self._catat_transaksi(
                    perangkat,
                    str(d.get("paket", "") or ""),
                    pesanan,
                    total,
                    source="tv",
                    paid=paid_awal,
                    booking_meta={
                        "booking_id": did,
                        "kode": did[:8].upper(),
                        "metode": metode,
                        "status_bayar": sb,
                        "pelanggan": str(d.get("namaPelanggan", "") or ""),
                        "no_hp": str(d.get("noHp", "") or ""),
                    })
                try:
                    self._booking_riwayat_refresh()
                except Exception:
                    pass
            except Exception as e:
                self._qr_log(f"catat transaksi booking {did} gagal: {e}")

        def lunas_sisa():
            sisa = int(d.get("sisaBayar", 0) or 0)
            dpn = int(d.get("nominalDp", 0) or 0)
            total = int(d.get("totalHarga", 0) or 0)
            if not messagebox.askyesno(
                    "Lunas Sisa",
                    f"Tandai sisa Rp {sisa:,} dari total Rp {total:,} sebagai LUNAS?\n"
                    f"(DP Rp {dpn:,} + sisa Rp {sisa:,})",
                    parent=win):
                return
            pat = {"statusBayar": "lunas_transfer",
                   "nominalTransfer": total,
                   "pelunasanSisa": sisa,
                   "lunasAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                   "updatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                   "kasir": str(getattr(self, "current_user", ""))}
            def _after_lunas():
                idx = self._riwayat_idx_by_booking(did)
                if idx >= 0:
                    try:
                        self._set_transaksi_paid_idx(idx, True)
                    except Exception:
                        pass
                perangkat = str(d.get("perangkat", "") or "").strip()
                kursi = self._qr_cari_kartu(perangkat) if perangkat else None
                if kursi is not None:
                    try:
                        if hasattr(kursi, "_set_paid"):
                            kursi._set_paid(True)
                    except Exception:
                        pass
                try:
                    messagebox.showinfo("Lunas",
                                        f"✅ Sisa Rp {sisa:,} booking {did[:8].upper()} "
                                        f"ditandai LUNAS via transfer.", parent=win)
                except Exception:
                    pass
                self._booking_riwayat_refresh()
            _simpan(pat, "💰 Sisa ditandai LUNAS", "✖ Gagal menyimpan — cek internet",
                    after_ok=_after_lunas)

        def tolak():
            alasan = simpledialog.askstring("Tolak Booking", "Alasan penolakan:", parent=win)
            if alasan is None:
                return
            _simpan({"status": "ditolak", "kasir": str(getattr(self, "current_user", "")),
                     "alasan": alasan.strip(), "updatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
                    "✖ Booking ditolak", "✖ Gagal menyimpan — cek internet")

        btn_konf = ctk.CTkButton(btns, text="✅ Konfirmasi", height=38, fg_color=C_GREEN,
                                 hover_color="#15803D", font=("Russo One", 11, "bold"),
                                 text_color="#06210F", command=konfirmasi)
        btn_konf.pack(side="left", expand=True, fill="x", padx=(0, 5))
        btns_all.append(btn_konf)
        # Booking yang sudah dikonfirmasi tidak bisa ditolak lagi — tombol
        # Tolak tidak dibuat sama sekali agar tampilan bersih.
        btn_tolak = None
        if st_sekarang != "dikonfirmasi":
            btn_tolak = ctk.CTkButton(btns, text="✖ Tolak", height=38, fg_color=C_RED,
                                      hover_color="#8B0000", font=("Russo One", 11, "bold"),
                                      text_color="white", command=tolak)
            btn_tolak.pack(side="left", expand=True, fill="x", padx=(0, 5))
            btns_all.append(btn_tolak)
        btn_bukti = ctk.CTkButton(btns, text="🖼 Lihat Bukti", height=38, fg_color=C_ACCENT2,
                                  hover_color="#5A0FCC", font=("Russo One", 11, "bold"),
                                  text_color="white", command=buka_bukti,
                                  state="normal" if d.get("bukti") else "disabled")
        btn_bukti.pack(side="left", expand=True, fill="x")
        btn_lunas_sisa = ctk.CTkButton(btns, text="💰 Lunas Sisa", height=38, fg_color=C_ORANGE,
                                       hover_color="#B37500", font=("Russo One", 11, "bold"),
                                       text_color="#1A1200", command=lunas_sisa,
                                       state="normal" if str(d.get("statusBayar", "")) == "dp" else "disabled")
        btn_lunas_sisa.pack(side="left", expand=True, fill="x", padx=(5, 0))
        btns_all.append(btn_lunas_sisa)

        try:
            if bunyi:
                threading.Thread(target=self._qr_bel, daemon=True).start()
        except Exception:
            pass

    def _qr_bel(self):
        try:
            import winsound
            for _ in range(3):
                winsound.Beep(1400, 180)
                import time as _t
                _t.sleep(0.08)
        except Exception:
            pass

    def _qr_tampil_panggilan(self, p: dict):
        """Popup panggilan (per TV). Tidak modal & bisa di-minimize — operator
        tetap bisa membuka sesi/aktivitas lain sambil pesanan menunggu."""
        try:
            threading.Thread(target=self._qr_bel, daemon=True).start()
        except Exception:
            pass
        tv = str(p.get("tv", "?"))
        tv_label = f"TV {tv}" if tv.isdigit() else tv
        jenis = str(p.get("jenis", "keluhan"))
        catatan = str(p.get("catatan", ""))
        oid = str(p.get("id") or "")
        items = p.get("items") or []

        popups = getattr(self, "_qr_popup_windows", {})
        win = popups.get(tv)
        if not (win and win.winfo_exists()):
            win = ctk.CTkToplevel(self)
            win.title(f"🔔 Panggilan — {tv_label}")
            win.geometry("470x540")
            win.minsize(400, 320)
            win.configure(fg_color=C_BG)
            hdr = ctk.CTkFrame(win, fg_color=C_PANEL, corner_radius=0, height=46)
            hdr.pack(fill="x")
            hdr.pack_propagate(False)
            ctk.CTkLabel(hdr, text=f"🔔  PANGGILAN DARI {tv_label}",
                         font=("Russo One", 13, "bold"),
                         text_color=C_ACCENT).pack(side="left", padx=12)
            ctk.CTkButton(hdr, text="📋 Riwayat", width=92, height=26,
                          fg_color=C_BTN, hover_color=C_ACCENT2,
                          border_width=1, border_color=C_ACCENT2,
                          font=("Russo One", 9, "bold"), text_color=C_ACCENT2,
                          command=self._buka_riwayat_pesanan).pack(side="right", padx=4, pady=9)
            ctk.CTkButton(hdr, text="✖ Tutup", width=76, height=26,
                          fg_color=C_RED, hover_color="#B71C1C",
                          font=("Russo One", 9, "bold"), text_color="white",
                          command=lambda: self._qr_tutup_panggilan(tv)).pack(side="right", padx=(4, 8), pady=9)
            body = ctk.CTkScrollableFrame(win, fg_color=C_PANEL, corner_radius=8)
            body.pack(fill="both", expand=True, padx=10, pady=10)
            win.body = body
            win._q_orders = {}          # oid -> {'widgets': {idx: (btn, lbl)}}
            win._q_items = {}           # oid -> daftar item terstruktur
            popups[tv] = win
        self._qr_tambah_section(win, tv, jenis, oid, items, p.get("catatan", ""))
        win.lift()
        win.attributes("-topmost", True)
        win.attributes("-topmost", False)

    def _qr_tambah_section(self, win, tv, jenis, oid, items, catatan):
        """Tambah satu blok pesanan (per panggilan) ke popup; tiap item punya
        tombol ✓ SUDAH untuk langsung menambahkan ke tagihan kartu TV."""
        label_jenis = {"pesanan": "🛒  PESANAN", "paket": "➕  TAMBAH PAKET",
                       "makanan": "🍔  MAKANAN", "minuman": "🥤  MINUMAN",
                       "keluhan": "📢  KELUHAN"}.get(jenis, jenis.upper())
        sec = ctk.CTkFrame(win.body, fg_color=C_CARD, corner_radius=8)
        sections = getattr(win, "_q_sections", None)
        if sections is None:
            sections = []
            win._q_sections = sections
        if sections:
            # Pack sebelum section pesanan pertama — JANGAN pakai
            # winfo_children()[0] (frame internal CTkScrollableFrame tidak
            # pack-managed → TclError "isn't packed").
            sec.pack(fill="x", padx=2, pady=5, before=sections[0])
        else:
            sec.pack(fill="x", padx=2, pady=5)
        sections.insert(0, sec)
        if items:
            jam = datetime.now().strftime("%H:%M")
            ctk.CTkLabel(sec, text=f"{label_jenis}  •  {jam}",
                         font=("Russo One", 11, "bold"),
                         text_color=C_ACCENT, anchor="w").pack(fill="x", padx=10, pady=(6, 0))
        widgets = {}
        if items:
            for idx, it in enumerate(items):
                row_f = ctk.CTkFrame(sec, fg_color="transparent")
                row_f.pack(fill="x", padx=8, pady=1)
                nama = str(it.get("nama", "?"))
                qty = int(it.get("qty", 1) or 1)
                harga = int(it.get("harga", 0) or 0)
                label_txt = f"•  {nama}" + (f"  x{qty}" if qty > 1 else "")
                if harga:
                    label_txt += f"    {fmt_rp(int(harga) * qty)}"
                st = str(it.get("status", "baru"))
                is_paket = str(it.get("tipe")) == "paket"
                txt_color = C_GREEN if st == "sudah" else C_TEXT
                icon = "➕" if is_paket else ("🍔" if it.get("tipe") == "makanan" else "🥤")
                lbl = ctk.CTkLabel(row_f, text=f"{icon} {label_txt}",
                                   font=FONT_BODY, anchor="w", text_color=txt_color)
                lbl.pack(side="left", fill="x", expand=True)
                btn_pay = ctk.CTkButton(row_f, text="✅  SUDAH BAYAR", width=118, height=26,
                                        fg_color=C_GREEN, hover_color="#2E7D32",
                                        text_color="white",
                                        font=("Russo One", 9, "bold"),
                                        command=lambda oid=oid, i=idx: self._qr_mark_sudah(
                                            oid, i, win, True))
                btn_unpaid = ctk.CTkButton(row_f, text="⏳  BELUM BAYAR", width=118, height=26,
                                           fg_color="#FFB300", hover_color="#E6A400",
                                           text_color="black",
                                           font=("Russo One", 9, "bold"),
                                           command=lambda oid=oid, i=idx: self._qr_mark_sudah(
                                               oid, i, win, False))
                if st != "baru":
                    btn_pay.configure(state="disabled", fg_color="#2E7D32",
                                      hover_color="#2E7D32")
                    btn_unpaid.configure(state="disabled", fg_color="#B26A00",
                                         hover_color="#B26A00")
                btn_pay.pack(side="right", padx=2, pady=3)
                btn_unpaid.pack(side="right", padx=2, pady=3)
                widgets[idx] = (btn_pay, btn_unpaid, lbl)
        if catatan:
            ctk.CTkLabel(sec, text=f"✏️  {catatan}",
                         font=FONT_SMALL, text_color=C_MUTED,
                         anchor="w", wraplength=380).pack(fill="x", padx=10, pady=(0, 4))
        elif not items:
            ctk.CTkLabel(sec, text="(panggilan tanpa pesanan — hanya panggil kasir)",
                         font=FONT_SMALL, text_color=C_MUTED).pack(pady=6)
        win._q_orders[oid] = {"widgets": widgets}
        win._q_items[oid] = items or []

    def _qr_tutup_panggilan(self, tv: str):
        win = getattr(self, "_qr_popup_windows", {}).pop(tv, None)
        if win and win.winfo_exists():
            try:
                win.destroy()
            except Exception:
                pass

    def _qr_pak_info(self, kartu, nama: str):
        """Cari harga & menit paket di grup kartu TV. Return (harga, menit)."""
        try:
            d = kartu.get_paket_data() or {}
        except Exception:
            d = {}
        if not d:
            try:
                d = self.get_paket_data(getattr(kartu, "nama_grup", None)) or {}
            except Exception:
                d = {}
        info = d.get(nama) or {}
        if isinstance(info, dict):
            try:
                return int(info.get("harga", 0) or 0), int(info.get("menit", 0) or 0)
            except Exception:
                return 0, 0
        try:
            return int(info or 0), 0
        except Exception:
            return 0, 0

    def _qr_mark_sudah(self, oid, idx, win=None, paid=True):
        """Operator menekan tombol aksi pada item pesanan web:
        - ✅ SUDAH BAYAR  → item langsung masuk kartu TV, status LUNAS.
        - ⏳ BELUM BAYAR  → item langsung masuk kartu TV, status TAGIHAN.
        - paket          → langsung masuk ke kartu TV (tambah waktu + biaya);
                           kalau kartu kosong, sesi otomatis dimulai.
        - makanan/minuman → langsung masuk ke tagihan kartu TV.
        Status item diperbarui di riwayat lokal & tampilan popup/riwayat."""
        import traceback
        try:
            rows = self._qr_pesan_log_load()
            row = next((r for r in rows if r.get("id") == str(oid)), None)
            if row is None:
                return
            items = row.get("items", [])
            if idx < 0 or idx >= len(items):
                return
            litem = items[idx]
            if str(litem.get("status", "baru")) != "baru":
                messagebox.showinfo("Sudah Diproses", "Item ini sudah ditandai.",
                                    parent=self)
                return
            tv = str(row.get("tv", ""))
            kartu = self._qr_cari_kartu(tv)
            if kartu is None:
                messagebox.showerror(
                    "Kartu TV Tidak Ditemukan",
                    f"Kartu TV '{tv}' tidak ada di dashboard.\n"
                    "Tambahkan/nyalakan kartu TV itu dulu, lalu klik tombol lagi.",
                    parent=self)
                return

            tipe = str(litem.get("tipe", ""))
            nama = str(litem.get("nama", "?"))
            qty = int(litem.get("qty", 1) or 1)
            harga_item = int(litem.get("harga", 0) or 0)
            lbl_status = "LUNAS" if paid else "TAGIHAN"
            info = ""
            try:
                if tipe == "paket":
                    harga, menit = self._qr_pak_info(kartu, nama)
                    harga = harga or harga_item
                    if harga <= 0:
                        messagebox.showwarning(
                            "Paket Tidak Dikenali",
                            f"Paket '{nama}' tidak ada di data tarif "
                            f"grup {getattr(kartu, 'nama_grup', '?')}.\n"
                            "Cek nama grup/paket di menu Tarif, lalu coba lagi.",
                            parent=self)
                        return
                    sesi_kosong = kartu.sesi_kosong()
                    kartu._on_paket_confirm(nama, harga, menit, {}, 0, 0, "nominal",
                                            paid=paid)
                    if sesi_kosong:
                        info = (f"TV {tv}: sesi baru '{nama}' ({fmt_rp(harga)}) "
                                f"otomatis berjalan ({menit:g} menit) — {lbl_status}.")
                    else:
                        info = (f"TV {tv}: paket '{nama}' +{menit:g} mnt → "
                                f"total {fmt_rp(kartu._total_setelah_diskon())} ({lbl_status}).")
                else:
                    pesanan = {nama: qty}
                    kartu._on_tambah_pesanan_confirm(pesanan, paid=paid,
                                                     _stok_delta=dict(pesanan))
                    if kartu.sesi_kosong():
                        info = (f"TV {tv}: {nama} masuk ke tagihan ({lbl_status}). Ses ini "
                                "belum aktif — mulai paket dari kartu TV untuk timer.")
                    else:
                        sub = f"({fmt_rp(harga_item * qty)}) " if harga_item else ""
                        info = f"TV {tv}: {nama} masuk ke tagihan {sub}({lbl_status})."
            except Exception as e:
                traceback.print_exc()
                messagebox.showerror("Gagal Tambah Pesanan",
                                     f"{tipe} '{nama}':\n{e}", parent=self)
                return

            litem["status"] = "sudah"
            litem["paid"] = bool(paid)
            row["ditangani"] = True
            if all(str(x.get("status", "baru")) == "sudah" for x in items):
                row["status"] = "selesai"
            self._qr_pesan_log_save(rows)
            self._qr_panggilan_ui_sudah(win, oid, idx, info, paid)
            self._qr_riwayat_refresh()
        except Exception as e:
            traceback.print_exc()
            messagebox.showerror("Error", f"Gagal menandai: {e}", parent=self)

    def _qr_panggilan_ui_sudah(self, win, oid, idx, info, paid=True):
        """Update tombol/label item di popup + notice singkat setelah aksi."""
        try:
            if win is not None and win.winfo_exists():
                o = getattr(win, "_q_orders", {}).get(str(oid))
                trio = (o or {}).get("widgets", {}).get(idx)
                if trio and len(trio) == 3:
                    btn_pay, btn_unpaid, lbl = trio
                    try:
                        btn_pay.configure(text="✅  LUNAS", state="disabled",
                                          fg_color="#2E7D32", hover_color="#2E7D32")
                        btn_unpaid.configure(text="⏳  TAGIHAN", state="disabled",
                                             fg_color="#B26A00", hover_color="#B26A00")
                        lbl.configure(text_color=C_GREEN if paid else "#FF9800")
                    except Exception:
                        pass
            if info:
                try:
                    self.after(0, lambda t=info: self._qr_notice_popup(t))
                except Exception:
                    pass
        except Exception:
            pass

    def _qr_notice_popup(self, text):
        try:
            ct = ctk.CTkToplevel(self)
            ct.title("✓  Sudah Diproses")
            ct.geometry("380x120")
            ct.configure(fg_color=C_BG)
            ctk.CTkLabel(ct, text=text, font=FONT_SMALL, text_color=C_TEXT,
                         wraplength=330).pack(pady=18)
            ct.after(2600, ct.destroy)
            ct.lift()
        except Exception:
            pass

    def _buka_riwayat_pesanan(self):
        """Window riwayat pesanan web — non-modal, bisa di-minimize; operator
        tetap bisa bekerja di TV lain sambil popup ini terbuka."""
        win = getattr(self, "_qr_his_win", None)
        if win and win.winfo_exists():
            win.lift()
            self._qr_riwayat_refresh()
            return
        win = ctk.CTkToplevel(self)
        win.title("📋 Riwayat Pesanan Web")
        win.geometry("980x560")
        win.minsize(760, 380)
        win.configure(fg_color=C_BG)
        self._qr_his_win = win

        hdr = ctk.CTkFrame(win, fg_color=C_PANEL, corner_radius=0, height=48)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        ctk.CTkLabel(hdr, text="📋  RIWAYAT PESANAN WEB",
                     font=("Russo One", 14, "bold"), text_color=C_ACCENT).pack(side="left", padx=14)
        ctk.CTkButton(hdr, text="✖ Tutup", width=76, height=28,
                      fg_color=C_RED, font=("Russo One", 10, "bold"), text_color="white",
                      command=win.destroy).pack(side="right", padx=(4, 10), pady=9)
        ctk.CTkButton(hdr, text="⏳ Tandai BELUM BAYAR", width=170, height=28,
                      fg_color="#FFB300", hover_color="#E6A400", text_color="black",
                      font=("Russo One", 10, "bold"),
                      command=lambda: self._qr_his_mark_selected(False)).pack(side="right", padx=4, pady=9)
        ctk.CTkButton(hdr, text="✅ Tandai SUDAH BAYAR", width=170, height=28,
                      fg_color=C_GREEN, hover_color="#2E7D32", text_color="white",
                      font=("Russo One", 10, "bold"),
                      command=lambda: self._qr_his_mark_selected(True)).pack(side="right", padx=4, pady=9)
        ctk.CTkButton(hdr, text="↻ Muat Ulang", width=120, height=28,
                      fg_color=C_BTN, border_width=1, border_color=C_ACCENT2,
                      font=("Russo One", 10, "bold"), text_color=C_ACCENT2,
                      command=self._qr_riwayat_refresh).pack(side="right", padx=4, pady=9)

        wrap = ctk.CTkFrame(win, fg_color=C_PANEL, corner_radius=8)
        wrap.pack(fill="both", expand=True, padx=10, pady=10)
        # Style lokal supaya font tabel riwayat QR ikut besar walau tab Riwayat
        # belum pernah dibuka (style "Treeview" global baru dibuat di sana).
        _st = ttk.Style()
        _st.configure("QRDiag.Treeview",
                      background=C_CARD, fieldbackground=C_CARD,
                      foreground=C_TEXT, rowheight=32,
                      font=("Consolas", 12))
        _st.configure("QRDiag.Treeview.Heading",
                      background=C_PANEL, foreground=C_ACCENT,
                      font=("Russo One", 11, "bold"), relief="flat")
        cols = ("waktu", "tv", "jenis", "item", "qty", "harga", "status", "catatan")
        tree = ttk.Treeview(wrap, columns=cols, show="headings",
                            selectmode="extended", style="QRDiag.Treeview")
        tree.heading("waktu", text="Waktu")
        tree.heading("tv", text="TV")
        tree.heading("jenis", text="Jenis")
        tree.heading("item", text="Item")
        tree.heading("qty", text="Qty")
        tree.heading("harga", text="Harga")
        tree.heading("status", text="Keterangan")
        tree.heading("catatan", text="Catatan")
        tree.column("waktu", width=130, anchor="w")
        tree.column("tv", width=48, anchor="center")
        tree.column("jenis", width=90, anchor="w")
        tree.column("item", width=240, anchor="w")
        tree.column("qty", width=46, anchor="center")
        tree.column("harga", width=90, anchor="e")
        tree.column("status", width=130, anchor="center")
        tree.column("catatan", width=200, anchor="w")
        tree.tag_configure("s", foreground=C_GREEN)
        tree.tag_configure("b", foreground="#FF9800")
        tree.tag_configure("t", foreground="#E53935")
        sb = ttk.Scrollbar(wrap, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=sb.set)
        tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        # Status HANYA diubah kasir via klik kanan (menu) — klik kiri tidak
        # boleh langsung menandai ditangani.
        tree.bind("<Button-3>", self._qr_riwayat_right_click)
        win._qr_tree = tree
        win._qr_wrap = wrap

        self._qr_riwayat_refresh()
        win.lift()

    def _qr_his_mark_selected(self, paid=True):
        win = getattr(self, "_qr_his_win", None)
        if not (win and win.winfo_exists()):
            return
        tree = getattr(win, "_qr_tree", None)
        if tree is None:
            return
        sel = tree.selection()
        if not sel:
            messagebox.showinfo("Pilih Item", "Pilih baris yang mau ditandai.",
                                parent=self)
            return
        for iid in list(sel):
            oid, idx = (iid.split("|", 1) + ["0"])[:2]
            try:
                idx = int(idx)
            except Exception:
                idx = -1
            rows = self._qr_pesan_log_load()
            row = next((r for r in rows if r.get("id") == oid), None)
            if row and 0 <= idx < len(row.get("items", [])):
                self._qr_mark_sudah(str(oid), idx, None, paid)
            else:
                tree.selection_remove(iid)

    def _qr_log_set_ditangani(self, oid, ditangani):
        """Tandai order (riwayat web) sebagai Sudah/Belum Ditangani."""
        rows = self._qr_pesan_log_load()
        for r in rows:
            if r.get("id") == str(oid):
                r["ditangani"] = bool(ditangani)
                break
        self._qr_pesan_log_save(rows)
        self._qr_riwayat_refresh()

    def _qr_riwayat_right_click(self, event):
        tree = event.widget
        iid = tree.identify_row(event.y)
        if not iid:
            return
        tree.selection_set(iid)
        oid, idx = (iid.split("|", 1) + ["0"])[:2]
        try:
            idx = int(idx)
        except Exception:
            idx = -1
        menu = tk.Menu(tree, tearoff=0)
        menu.add_command(label="✅  Tandai Sudah Ditangani",
                         command=lambda o=oid: self._qr_log_set_ditangani(o, True))
        menu.add_command(label="⏳  Tandai Belum Ditangani",
                         command=lambda o=oid: self._qr_log_set_ditangani(o, False))
        if idx >= 0:
            menu.add_separator()
            menu.add_command(label="✅  Tandai SUDAH BAYAR",
                             command=lambda o=oid, i=idx: self._qr_mark_sudah(o, i, None, True))
            menu.add_command(label="⏳  Tandai BELUM BAYAR",
                             command=lambda o=oid, i=idx: self._qr_mark_sudah(o, i, None, False))
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    def _qr_riwayat_refresh(self):
        """Refresh isi tree riwayat + status bar (terbaru di paling atas)."""
        win = getattr(self, "_qr_his_win", None)
        if not (win and win.winfo_exists()):
            return
        tree = getattr(win, "_qr_tree", None)
        if tree is None:
            return
        rows = self._qr_pesan_log_load()
        try:
            rows.sort(key=lambda r: datetime.strptime(str(r.get("waktu", "")),
                                                      "%Y-%m-%d %H:%M"))
        except Exception:
            pass
        tree.delete(*tree.get_children())
        belum = 0
        for r in reversed(rows):
            tv = str(r.get("tv", "?"))
            tv_label = f"TV {tv}" if tv.isdigit() else tv
            oid = str(r.get("id", ""))
            catatan = str(r.get("catatan", ""))
            dit = bool(r.get("ditangani", False))
            items = r.get("items", []) or []
            if not items:
                if not dit:
                    belum += 1
                tree.insert("", "end",
                            iid=f"{oid}|x",
                            values=(str(r.get("waktu", "")), tv_label,
                                    str(r.get("jenis", "")).upper(),
                                    "(panggilan tanpa item)", "1", "",
                                    "Sudah Ditangani" if dit else "Belum Ditangani",
                                    catatan),
                            tags=("s" if dit else "b",))
                continue
            if not dit and any(str(x.get("status", "baru")) == "baru" for x in items):
                belum += 1
            for idx, it in enumerate(items):
                st = str(it.get("status", "baru"))
                tipe = str(it.get("tipe", ""))
                harga = int(it.get("harga", 0) or 0)
                qty = int(it.get("qty", 1) or 1)
                butir = f"{it.get('nama', '?')}  ({tipe.upper()})"
                if harga:
                    butir += f"  • {fmt_rp(harga)}"
                sudah = st != "baru"
                st_show = "Sudah Ditangani" if (sudah or dit) else "Belum Ditangani"
                tree.insert("", "end", iid=f"{oid}|{idx}",
                            values=(str(r.get("waktu", "")), tv_label,
                                    str(r.get("jenis", "")).upper(),
                                    butir, qty,
                                    fmt_rp(harga * qty) if harga else "—",
                                    st_show,
                                    catatan if idx == len(items) - 1 else ""),
                            tags=("s" if (sudah or dit) else "b",))
        footer = getattr(win, "_qr_footer", None)
        if footer is None:
            footer = ctk.CTkLabel(win, text="", font=FONT_SMALL, text_color=C_MUTED)
            footer.pack(pady=4)
            win._qr_footer = footer
        footer.configure(text=f"Total pesanan: {len(rows)}  •  Belum Ditangani: {belum}")
        if win and win.winfo_exists():
            win.after(3000, self._qr_riwayat_refresh)

    def _qr_push_menu_bg(self):
        """Push menu/nama rental ke Firestore call_meta/<owner> — dipakai halaman
        web pelanggan. Dijalankan berkala (menu jarang berubah)."""
        def worker():
            try:
                from firestore_sync import FirestoreClient
                owner = self._resolve_license_user()
                if not owner:
                    return
                pus = (ConfigManager.load().get("profil_rental", {}) or {}).get(owner, {}) or {}
                if not isinstance(pus, dict):
                    pus = {}
                nama_rental = str(pus.get("nama_rental", "") or "").strip() or "RR Billing Pro"
                paket_grup = {}
                for g in (getattr(self, "grup_tarif", {}) or {}):
                    d = self.get_paket_data(g)
                    if d:
                        paket_grup[g] = {
                            n: {"harga": int(v.get("harga", 0) if isinstance(v, dict) else v),
                                "menit": int(v.get("menit", 60) if isinstance(v, dict) else 60)}
                            for n, v in d.items()
                        }
                data = {
                    "nama_rental": nama_rental,
                    "logo": self._logo_rental_b64(),
                    "daftar_tv": self._get_daftar_tv_nama(),
                    "paket_grup": paket_grup,
                    "makanan": dict(getattr(self, "menu_makanan", {}) or {}),
                    "minuman": dict(getattr(self, "menu_minuman", {}) or {}),
                    "stok": getattr(self, "stok", {}) or {},
                    "stok_min": getattr(self, "stok_min", {}) or {},
                    "updatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                }
                for k in ("nama_dana", "no_dana", "alamat", "no_hp"):
                    if k == "no_hp":
                        v = str(pus.get("no_hp") or pus.get("hp") or "").strip()
                    else:
                        v = str(pus.get(k, "") or "").strip()
                    if v:
                        data[k] = v
                if str(pus.get("qr_pembayaran", "") or "").strip():
                    data["qr_pembayaran"] = str(pus.get("qr_pembayaran", "")).strip()
                FirestoreClient().set_document(f"call_meta/{owner}", data, merge=True)
            except Exception as e:
                print(f"[QR] push menu gagal: {e}")
        try:
            threading.Thread(target=worker, daemon=True).start()
        except Exception:
            pass
        try:
            self._stok_push_queued = False
        except Exception:
            pass
        try:
            self.after(300000, self._qr_push_menu_bg)  # refresh berkala (5 menit)
        except Exception:
            pass

    def _media_state_extra(self, meja_id: str) -> list:
        """TIDAK lagi mengirim ulang SHOW_MEDIA saat client reconnect.

        Dulu media diputar ulang pada SETIAP reconnect WS (blip jaringan,
        restart app TV, restart kasir) — membuat video promo "selalu play"
        di waktu yang salah. Sejak R3, media saat TV bangun dari tidur
        diputar oleh APK sendiri (onScreenWake: 1x per ACTION_SCREEN_ON,
        URL terakhir yang diterima). Di sini selalu kosong.
        """
        return []

    def _lakukan_aktivasi(self):
        kode = self.entry_kode.get().strip()
        if not kode:
            self.lbl_akt_status.configure(
                text="⚠  Masukkan kode aktivasi dulu.", text_color=C_YELLOW)
            return

        uname = self.current_user or ""

        # LicenseManager.aktivasi sudah otomatis mencoba binding username
        # (kalau rr_keygen.py tersedia) dengan fallback ke universal/machine.
        sukses, pesan = LicenseManager.aktivasi(
            kode,
            username=uname,
            binding_mode="username" if uname else "machine",
            promo_add_tv=self._get_promo_add_tv_from_current(),
        )

        AuditLogger.log(
            action="activation_attempt",
            username=uname,
            status="success" if sukses else "failed",
            details={
                "binding_mode": "username" if uname else "machine",
                "message": pesan,
            }
        )
        if sukses:
            self.lbl_akt_status.configure(text=f"✅  {pesan}", text_color=C_GREEN)
            messagebox.showinfo("🎉 Aktivasi Berhasil", pesan)
            self._save_promo_override_to_license()
            self._rebuild_sidebar_lic()

            # Sync ke Firestore
            threading.Thread(target=self._sync_aktivasi_ke_cloud, args=(kode,), daemon=True).start()
        else:
            self.lbl_akt_status.configure(text=f"✖  {pesan}", text_color=C_RED)

    def _revoke_license(self):
        """Revoke license: deactivate locally, sync to Firestore, clear file, log audit."""
        uname = self.current_user or ""
        if not messagebox.askyesno("🛑 Revoke License",
                                   "Yakin ingin mencabut lisensi?\n\n"
                                   "Lisensi akan dinonaktifkan dan tidak bisa digunakan lagi.\n\n"
                                   "Alasan pencabutan (opsional):\n"
                                   "(akan dikirim ke server)"):
            return
        dlg = ctk.CTkToplevel(self)
        dlg.title("Alasan Revoke")
        dlg.configure(fg_color=C_BG)
        dlg.geometry("400x200")
        dlg.transient(self)
        dlg.grab_set()
        ctk.CTkLabel(dlg, text="📝  Alasan pencabutan lisensi:",
                     font=FONT_SUB, text_color=C_ACCENT2).pack(pady=(16, 8))
        entry_alasan = ctk.CTkEntry(dlg, placeholder_text="Contoh: Ganti PC / Lisensi bermasalah",
                                     fg_color=C_BTN, text_color=C_TEXT,
                                     border_color=C_BORDER, font=FONT_BODY,
                                     height=34, width=340)
        entry_alasan.pack(padx=20, pady=(0, 12))
        def _do_revoke():
            alasan = entry_alasan.get().strip()
            threading.Thread(target=self._revoke_license_thread, args=(alasan, dlg), daemon=True).start()
        ctk.CTkButton(dlg, text="🛑  Ya, Revoke", width=140, height=34,
                      fg_color=C_RED, hover_color="#7A1A1A",
                      font=("Russo One", 10, "bold"), text_color="white",
                      command=_do_revoke).pack(pady=(0, 16))

    def _revoke_license_thread(self, alasan: str, dlg=None):
        try:
            # 1. Deactivate locally
            sukses, pesan = LicenseManager.deactivate()
            if not sukses:
                self.after(0, lambda: messagebox.showerror("Error", pesan))
                return
            # 2. Send revocation to Firestore
            try:
                fc = FirestoreClient()
                lic = LicenseManager.load()
                kode = lic.get("kode_aktivasi", "")
                if kode:
                    record = fc.find_license_by_code(kode)
                    if record:
                        doc_id = record.get("_id", "")
                        if doc_id:
                            fc.revoke_license(doc_id, reason=alasan)
                # Also write licenseStatus to user doc
                fc.write_license_status(self.current_user if self.current_user else "", {
                    "status": "revoked",
                    "pesan": "Lisensi dicabut oleh pengguna.",
                    "expiresAt": "",
                })
            except Exception as e:
                _LOGGER.warning("Cloud revoke error: %s", e)
            # 3. Clear local license file (set revoked=True)
            lic_data = LicenseManager.load()
            lic_data["aktif"] = False
            lic_data["revoked"] = True
            lic_data["revoked_at"] = datetime.now().isoformat()
            LicenseManager.save(lic_data)
            # 4. Log to audit
            AuditLogger.log(
                action="license_revoked",
                username=self.current_user or "",
                status="success",
                details={"reason": alasan}
            )
            if dlg:
                self.after(0, dlg.destroy)
            self.after(0, lambda: messagebox.showinfo("✅ License Revoked",
                        "Lisensi telah dicabut.\nAktivasi ulang diperlukan untuk menggunakan aplikasi."))
            self.after(200, self._rebuild_sidebar_lic)
        except Exception as e:
            _LOGGER.exception("Revoke error: %s", e)
            if dlg:
                self.after(0, dlg.destroy)
            self.after(0, lambda e=e: messagebox.showerror("Error", f"Gagal revoke: {e}"))

    def _sync_aktivasi_ke_cloud(self, kode: str):
        """Sync aktivasi ke Firestore: update license doc + write licenseStatus ke user doc."""
        uname = self.current_user or ""
        if not uname:
            return
        try:
            import datetime as _dt
            fc = FirestoreClient()
            # 1. Cari license record by kode (opsional — kode keygen tidak
            # selalu masuk koleksi licenses/; jangan blokir penulisan status)
            record = fc.find_license_by_code(kode)
            doc_id = (record or {}).get("_id", "")
            if record and record.get("revoked"):
                _LOGGER.warning("Sync aktivasi dilewati — lisensi %s sudah direvoke", kode)
                return
            if not doc_id:
                _LOGGER.info("License code not in Firestore (keygen code?) — sync tetap lanjut: %s", kode)

            # 2. Dapatkan expiry + maxTv dari license record dan lisensi lokal
            status = LicenseManager.get_status(current_user=uname)
            expiry_str = status.get("expiry", "")

            # 3. Activate license (update activatedDevices) bila record ada
            if record and doc_id:
                fc.activate_license(doc_id, expiry=expiry_str, device_type="desktop")

            # 4. Hitung maxTv dari license record (prioritas), fallback ke promo, lalu ke edition
            lic_local = LicenseManager.load()
            max_tv = (record or {}).get("maxTv", 0) or 0
            if max_tv <= 0:
                max_tv = lic_local.get("promo_add_tv", 0) or 0
            if max_tv <= 0:
                ed = lic_local.get("edition", "")
                max_tv = {"BULANAN": 5, "3BULAN": 10, "TAHUNAN": 15, "LIFETIME": 999999}.get(ed, 0)
            if max_tv <= 0:
                pkg_name = ((record or {}).get("package", "BULANAN") or "BULANAN").upper()
                pkg_max_tv = {"BULANAN": 5, "3BULAN": 10, "TAHUNAN": 15, "LIFETIME": 999999}
                max_tv = pkg_max_tv.get(pkg_name, 5)
            promo_add = lic_local.get("promo_add_tv", 0)

            # 5. Write licenseStatus ke user doc
            ls = {
                "status": "active",
                "pesan": f"Lisensi aktif hingga {expiry_str}",
                "expiresAt": expiry_str,
                "maxTv": max_tv,
                "maxPc": max_tv,
                "promoAddTv": promo_add,
                "cloud_restored": True,
            }
            fc.write_license_status(uname, ls)
            _LOGGER.info("Cloud activation sync success for %s (maxTv=%d, promo=%d)", uname, max_tv, promo_add)
        except Exception as e:
            _LOGGER.warning("Cloud activation sync error: %s", e)

    def _start_license_poller(self):
        """Start background poller untuk memantau perubahan lisensi dari cloud."""
        uname = self.current_user or ""
        if not uname:
            return
        try:
            from firestore_sync import LicensePoller
            self._license_poller = LicensePoller(uname, interval=180.0)
            self._license_poller.start(self._on_cloud_license_update)
            _LOGGER.info("License poller started for %s", uname)
        except Exception as e:
            _LOGGER.warning("License poller start error: %s", e)

    def _stop_license_poller(self):
        """Stop license poller."""
        poller = getattr(self, "_license_poller", None)
        if poller:
            try:
                poller.stop()
            except Exception:
                pass
            self._license_poller = None

    # ── SINGLE SESSION (1 AKUN = 1 PC) ─────────────────────────────────
    @staticmethod
    def _get_device_id():
        """Device ID persisten per PC (UUID di config; dibuat otomatis sekali)."""
        cfg = ConfigManager.load()
        dev = str(cfg.get("device_id") or "").strip()
        if not dev:
            dev = uuid.uuid4().hex[:12].upper()
            cfg["device_id"] = dev
            ConfigManager.save(cfg)
        return dev

    @staticmethod
    def _get_pc_name():
        try:
            return socket.gethostname() or "PC"
        except Exception:
            return "PC"

    def _session_payload(self, login_at=None):
        return {
            "deviceId": self._get_device_id(),
            "pcName": self._get_pc_name(),
            "lastSeen": int(time.time() * 1000),
            "loginAt": int(login_at or getattr(self, "_session_login_at", 0) or time.time() * 1000),
        }

    def _register_session(self):
        uname = getattr(self, "current_user", None) or ""
        if not uname:
            return
        try:
            if not getattr(self, "_session_login_at", None):
                self._session_login_at = int(time.time() * 1000)
            fc = FirestoreClient()
            fc.set_user_doc(uname, {"activeSession": self._session_payload()}, merge=True)
        except Exception as e:
            _LOGGER.warning("Session register error: %s", e)

    def _clear_session_cloud(self):
        """Hapus field activeSession (best-effort) supaya akun bisa dipakai PC lain."""
        uname = getattr(self, "current_user", None) or ""
        if not uname:
            return
        try:
            fc = FirestoreClient()
            fc.set_user_doc(uname, {"activeSession": None}, merge=True)
        except Exception:
            pass

    def _clear_session_cloud_quick(self):
        """Clear sesi di background dengan batas waktu 2 dtk (jangan blokir UI)."""
        t = threading.Thread(target=self._clear_session_cloud, daemon=True)
        t.start()
        t.join(timeout=2.0)

    def _conflicting_session_pc(self):
        """Kembalikan dict sesi PC lain yang masih aktif (grace 90 dtk), atau None."""
        uname = getattr(self, "current_user", None) or ""
        if not uname:
            return None
        try:
            fc = FirestoreClient()
            doc = fc.get_user_doc(uname) or {}
            s = doc.get("activeSession") or {}
            dev = str(s.get("deviceId") or "")
            last = 0
            try:
                last = int(s.get("lastSeen") or 0)
            except Exception:
                pass
            if dev and dev != self._get_device_id() and (time.time() * 1000 - last) < 90000:
                return {
                    "pcName": str(s.get("pcName") or "PC lain"),
                    "deviceId": dev,
                    "loginAt": int(s.get("loginAt") or 0),
                }
        except Exception:
            pass
        return None

    def _start_session_poller(self):
        """Heartbeat 15 dtk di thread + deteksi sesi lain (tanpa beban di thread UI)."""
        self._stop_session_poller()
        self._session_poller_stop = False
        self._session_poller_id = getattr(self, "_session_poller_id", 0) + 1
        self._session_conflict_answered = False
        self._session_conflict_confirmed = False
        poller_id = self._session_poller_id

        def _run():
            while True:
                if getattr(self, "_session_poller_stop", True):
                    return
                if getattr(self, "_session_poller_id", 0) != poller_id:
                    return
                try:
                    konflik = self._conflicting_session_pc()
                    aksi = self._session_conflict_action(konflik)
                    if aksi == "kick":
                        # Saya PC lama, sesi diambil PC lain → kick otomatis
                        self.after(0, lambda k=konflik: self._force_logout_remote(
                            f"akun ini dipakai di PC lain ({k.get('pcName', 'PC lain')})"))
                        return
                    elif aksi == "ask":
                        # Saya PC yang baru login → minta konfirmasi sekali
                        self.after(0, lambda k=konflik: self._on_session_login_check(k))
                    elif aksi == "register":
                        self._register_session()
                except Exception as e:
                    _LOGGER.warning("Session tick error: %s", e)
                time.sleep(15)

        self._session_poller_thread = threading.Thread(
            target=_run, daemon=True, name="SessionPoller")
        self._session_poller_thread.start()

    def _stop_session_poller(self):
        self._session_poller_stop = True
        self._session_poller_thread = None

    def _session_conflict_action(self, konflik):
        """Aksi saat ada konflik sesi: 'kick' (saya PC lama → logout otomatis),
        'ask' (saya PC baru → minta konfirmasi), 'register' (tidak konflik / sudah disetujui),
        'wait' (konflik ada tapi belum diputuskan user)."""
        if not konflik:
            return "register"
        mine = int(getattr(self, "_session_login_at", 0) or 0)
        theirs = int(konflik.get("loginAt") or 0)
        if theirs > mine:
            return "kick"
        if not getattr(self, "_session_conflict_answered", False):
            return "ask"
        return "register" if getattr(self, "_session_conflict_confirmed", False) else "wait"

    def _on_session_login_check(self, konflik):
        """UI thread: konfirmasi saat login di PC2 dan ada sesi aktif di PC lain."""
        if getattr(self, "_session_conflict_answered", False):
            return
        self._session_conflict_answered = True
        pc_lain = (konflik or {}).get("pcName") or "PC lain"
        try:
            lanjut = messagebox.askyesno(
                "🔒 AKUN SUDAH LOGIN DI PC LAIN",
                f"Akun '{self.current_user}' sedang aktif di PC: {pc_lain}.\n\n"
                "Jika kamu lanjutkan, PC tersebut akan otomatis logout.\n\n"
                "Lanjutkan login di PC ini?")
        except Exception:
            lanjut = True
        if not lanjut:
            self._show_login()
            return
        self._session_conflict_confirmed = True

    def _load_promo_bg(self):
        """Fetch promo settings di thread (jangan blokir thread UI saat login)."""
        try:
            data = FirestoreClient().fetch_promo_settings()
            if data is not None:
                self._promo_data = data
        except Exception:
            pass

    def _force_logout_remote(self, reason):
        """Logout otomatis karena sesi diambil PC lain (harus dipanggil di thread UI)."""
        self._stop_session_poller()
        self._stop_idle_watcher()
        try:
            threading.Thread(target=self._clear_session_cloud, daemon=True).start()
        except Exception:
            pass
        try:
            messagebox.showwarning("🔒 DIPINDAHKAN",
                                   f"Kamu telah logout otomatis karena {reason}.\n"
                                   "Sesi lama otomatis berakhir.")
        except Exception:
            pass
        self._show_login()

    # ── AUTO-LOCK SAAT IDLE ────────────────────────────────────────────
    def _start_idle_watcher(self):
        """Kunci otomatis ke halaman login setelah X menit tanpa aktivitas
        (idle_lock_minutes di config; 0/nonaktif = default).
        Handler hanya mencatat waktu (tanpa after-churn); satu poller 5 dtk yang cek selisihnya."""
        self._stop_idle_watcher()
        try:
            cfg = ConfigManager.load()
            minutes = int(cfg.get("idle_lock_minutes", 0) or 0)
        except Exception:
            minutes = 0
        if minutes <= 0:
            return  # fitur dinonaktifkan (default)
        self._idle_timeout_ms = minutes * 60000
        self._idle_last_evt = time.time()
        self._idle_stop = False

        def _mark(_evt=None):
            self._idle_last_evt = time.time()

        def _tick():
            if getattr(self, "_idle_stop", True):
                return
            if time.time() - self._idle_last_evt > self._idle_timeout_ms / 1000:
                try:
                    self.after(0, self._auto_lock)
                except Exception:
                    pass
                return
            try:
                self._idle_after = self.after(5000, _tick)
            except Exception:
                pass

        self._idle_mark = _mark
        self.bind_all("<Key>", _mark, add="+")
        self.bind_all("<Button-1>", _mark, add="+")
        self._idle_after = self.after(5000, _tick)

    def _stop_idle_watcher(self):
        self._idle_stop = True
        if getattr(self, "_idle_after", None):
            try:
                self.after_cancel(self._idle_after)
            except Exception:
                pass
            self._idle_after = None
        try:
            self.unbind_all("<Key>")
            self.unbind_all("<Button-1>")
        except Exception:
            pass

    def _auto_lock(self):
        if (getattr(self, "current_user", None) or "") == "":
            return
        self._stop_idle_watcher()
        self._stop_session_poller()
        try:
            threading.Thread(target=self._clear_session_cloud, daemon=True).start()
        except Exception:
            pass
        try:
            menit = max(1, self._idle_timeout_ms // 60000)
            messagebox.showinfo("🔒 Terkunci Otomatis",
                                f"Aplikasi terkunci karena tidak ada aktivitas selama {menit} menit.\n"
                                "Silakan login kembali.")
        except Exception:
            pass
        self._show_login()

    def _on_cloud_license_update(self, ls: Optional[dict]):
        """Callback ketika cloud license status berubah."""
        if not ls:
            return
        try:
            import datetime as _dt
            import pathlib as _pl
            lic_path = _pl.Path("rr_billing_license.json")
            existing = {}
            if lic_path.exists():
                existing = json.loads(lic_path.read_text())
            # Lisensi file = cache cloud; binding mengikuti admin yang login
            # (kasir ikut admin_utama). Lisensi milik user lain tidak ditimpa,
            # kecuali admin yang login adalah pemilik data mesin ini.
            existing_user = existing.get("username", "")
            resolved_user = (self._resolve_license_user() or (self.current_user or "")).strip()
            if existing.get("aktif") and existing_user and existing_user != resolved_user:
                admin_utama = None
                try:
                    for _u in (ConfigManager.get("users", {}) or {}).values():
                        if isinstance(_u, dict) and _u.get("admin_utama"):
                            admin_utama = _u.get("admin_utama")
                except Exception:
                    admin_utama = None
                if not (admin_utama and resolved_user == admin_utama):
                    return
            status = ls.get("status", "")
            expires_at = ls.get("expiresAt", "")

            # ── Lisensi dicabut / kedaluwarsa dari cloud → nonaktifkan lokal ──
            if status in ("revoked", "expired", "inactive") and existing.get("aktif"):
                existing["aktif"] = False
                existing["revoked"] = True
                existing["revoked_from_cloud"] = True
                existing["revoked_at"] = _dt.datetime.now().isoformat()
                if status == "revoked":
                    existing["revoke_reason"] = ls.get("pesan", "Lisensi dicabut di cloud.")
                LicenseManager.save(existing)
                _LOGGER.warning("Lisensi lokal dinonaktifkan dari cloud (status=%s)", status)
                try:
                    self.after(200, self._rebuild_sidebar_lic)
                    self.after(400, self._cek_lisensi_saat_start)
                except Exception:
                    pass
                return

            if status == "active" and expires_at:
                expires = _dt.datetime.fromisoformat(expires_at.replace("Z", "+00:00"))
                if expires.tzinfo is None:
                    expires = expires.replace(tzinfo=_dt.timezone.utc)
                if expires > _dt.datetime.now(_dt.timezone.utc):
                    existing["aktif"] = True
                    # Proteksi: jangan memotong masa lisensi lokal yang lebih lama
                    # (cloud kadang hanya memegang 1 dari 2 kode yang di-stack).
                    try:
                        _old_exp_raw = existing.get("expiry", "")
                        _old_exp = _dt.datetime.fromisoformat(
                            (_old_exp_raw + "T00:00:00" if "T" not in _old_exp_raw else _old_exp_raw).replace("Z", "+00:00"))
                        _new_exp = _dt.datetime.fromisoformat(
                            (expires_at + "T00:00:00" if "T" not in expires_at else expires_at).replace("Z", "+00:00"))
                        if not _old_exp_raw or _new_exp > _old_exp:
                            existing["expiry"] = expires_at
                    except Exception:
                        existing["expiry"] = expires_at
                    existing["firebase_sync"] = True
                    existing["binding_mode"] = "username"
                    existing["username"] = resolved_user or self.current_user
                    mtv = ls.get("maxTv") or 0
                    if mtv <= 0:
                        _LOGGER.warning(
                            "Cloud licenseStatus tanpa maxTv (user=%s) — edition dipertahankan",
                            resolved_user or self.current_user)
                    if mtv >= 999999:
                        _new_ed = "LIFETIME"
                    elif mtv >= 15:
                        _new_ed = "TAHUNAN"
                    elif mtv >= 10:
                        _new_ed = "3BULAN"
                    elif mtv > 0:
                        _new_ed = "BULANAN"
                    else:
                        _new_ed = ""
                    if _new_ed:
                        _ed_rank = {"BULANAN": 0, "3BULAN": 1, "TAHUNAN": 2, "LIFETIME": 3}
                        _old_ed = str(existing.get("edition", "")).strip().upper() if "edition" in existing else ""
                        if not _old_ed or _ed_rank.get(_new_ed, -1) > _ed_rank.get(_old_ed, -1):
                            existing["edition"] = _new_ed
                    existing.setdefault("promo_add_tv", 0)
                    existing.setdefault("promo_add_warnet", 0)
                    LicenseManager.save(existing)
                    try:
                        self.after(200, self._rebuild_sidebar_lic)
                        self.after(500, self._save_promo_override_to_license)
                    except Exception:
                        pass
        except Exception as e:
            _LOGGER.warning("Cloud license update error: %s", e)

    @staticmethod
    def _migrasi_grup_tarif(data_grup, data_paket_lama):
        """
        Pastikan struktur self.grup_tarif selalu:
            {nama_grup: {nama_paket: {"harga":int, "menit":int}, ...}, ...}

        Tiga kemungkinan sumber data saat startup:
        1. Config baru sudah punya "grup_tarif"  -> dipakai langsung (dengan sanitasi).
        2. Config lama (sebelum fitur grup) cuma punya "paket_main" flat
           -> dijadikan grup tunggal "Reguler", grup PS3/PS4/Room VIP lain
              TIDAK otomatis dibuat (supaya tidak mengejutkan user lama yang
              sudah custom harga "Reguler"-nya).
        3. Belum ada config sama sekali -> pakai DEFAULT_GRUP_TARIF penuh
           (Reguler, PS3, PS4, Room VIP).
        """
        menit_lama = {
            "30 Menit": 30, "1 Jam": 60, "2 Jam": 120,
            "3 Jam": 180, "5 Jam": 300, "Overnight": 540,
            "Main Bebas": 0, "Reguler": 60,
        }

        def _bersihkan_satu_grup(paket_dict):
            hasil = {}
            for nama, val in paket_dict.items():
                if isinstance(val, dict):
                    hasil[nama] = {"harga": int(val.get("harga", 0)), "menit": int(val.get("menit", 0))}
                else:
                    hasil[nama] = {"harga": int(val), "menit": menit_lama.get(nama, 60)}
            if "Main Bebas" not in hasil:
                hasil["Main Bebas"] = {"harga": 0, "menit": 0}
            return hasil

        if data_grup:
            return {nama: _bersihkan_satu_grup(paket) for nama, paket in data_grup.items()}

        if data_paket_lama:
            return {NAMA_GRUP_DEFAULT: _bersihkan_satu_grup(data_paket_lama)}

        return {nama: {k: dict(v) for k, v in paket.items()} for nama, paket in DEFAULT_GRUP_TARIF.items()}

    def _warnet_group_names(self):
        """Return set of nama grup warnet-only (grup yang TIDAK boleh dipakai kartu TV).

        Gabungan dua lapis: config 'warnet_only_groups' + semua kunci 'grup_tarif_warnet'.
        Grup dalam set ini hanya boleh diikat ke kartu PC/warnet.
        """
        cfg = ConfigManager.load()
        names = set(cfg.get('warnet_only_groups', []) or [])
        for g in (cfg.get('grup_tarif_warnet', {}) or {}).keys():
            names.add(g)
        return names

    def _is_warnet_group(self, nama_grup):
        """True jika nama_grup adalah grup khusus warnet (tidak boleh dipakai TV)."""
        if not nama_grup:
            return False
        return nama_grup in self._warnet_group_names()

    def daftar_nama_grup(self, for_warnet=False):
        """Return list of group names.
        If for_warnet is False (default), exclude groups marked as warnet-only in config 'warnet_only_groups'
        dan semua grup di 'grup_tarif_warnet' — grup warnet tidak pernah muncul untuk kartu TV.
        If for_warnet is True, include warnet-only groups and any groups defined specifically for warnet in
        config key 'grup_tarif_warnet'. "Warnet" group (if exists) is always first.
        """
        cfg = ConfigManager.load()
        warnet_only = self._warnet_group_names()
        if for_warnet:
            names = list(self.grup_tarif.keys()) if getattr(self, 'grup_tarif', None) else []
            # Include any warnet-specific groups defined separately in config
            warnet_map = cfg.get('grup_tarif_warnet', {}) or {}
            for g in warnet_map.keys():
                if g not in names:
                    names.append(g)
            # Ensure "Warnet" group is first if it exists
            if 'Warnet' in names:
                names.remove('Warnet')
                names.insert(0, 'Warnet')
            return names or [NAMA_GRUP_DEFAULT]
        # Default (for PS/TV): exclude warnet-only groups
        names = [g for g in (self.grup_tarif.keys() if getattr(self, 'grup_tarif', None) else []) if g not in warnet_only]
        return names or [NAMA_GRUP_DEFAULT]

    def daftar_semua_grup(self):
        """Return list of ALL group names untuk Kontrol Harga (shared + warnet).
        User bisa edit harga untuk semua grup.
        """
        cfg = ConfigManager.load()
        names = list(self.grup_tarif.keys()) if getattr(self, 'grup_tarif', None) else []
        # Include warnet-specific groups
        warnet_map = cfg.get('grup_tarif_warnet', {}) or {}
        for g in warnet_map.keys():
            if g not in names:
                names.append(g)
        return names or [NAMA_GRUP_DEFAULT]

    # ── Login ──────────────────────────────────────────────────────────────────
    def _show_login(self):
        # Login tampil dengan skala asli (sebelum pembesaran ui_scale) — ukuran
        # teks & jendela dikembalikan ke semula; main layout pakai UI_SCALE lagi
        # saat _on_login.
        ctk.set_widget_scaling(1.0)
        ctk.set_window_scaling(1.0)
        self._stop_license_poller()
        self._stop_call_poller()
        self._stop_session_poller()
        self._stop_idle_watcher()
        try:
            threading.Thread(target=self._clear_session_cloud, daemon=True).start()
        except Exception:
            pass
        self._promo_data = None
        self.geometry("620x900")
        self.resizable(False, False)
        for w in self.winfo_children():
            w.destroy()
        login = LoginPage(self, on_login_success=self._on_login)
        login.pack(fill="both", expand=True)

    def _on_login(self, username, role):
        # Kembalikan skala UI (font besar) untuk layout utama — login sudah
        # selesai dibuat dengan skala 1.0 (lihat _show_login).
        ctk.set_widget_scaling(UI_SCALE)
        ctk.set_window_scaling(UI_SCALE)
        self.current_user = username
        self.current_role = role
        self.current_user_email = ""
        try:
            users = ConfigManager.get("users", {})
            rec = users.get(username) if isinstance(users, dict) else None
            if isinstance(rec, dict):
                self.current_user_email = str(rec.get("email", "") or "")
        except Exception:
            pass
        self.geometry("1280x800")
        self.resizable(True, True)
        self.state("zoomed")
        for w in self.winfo_children():
            w.destroy()
        self._build_layout()
        self._muat_daftar_tv()
        self._muat_daftar_warnet()
        # QR panggil kasir: state per sesi
        self._qr_seen = set()
        self._qr_last_call = {}
        self._qr_popup_windows = {}
        # Auto-generate QR setiap TV (IP berubah -> QR baru)
        try:
            self._qr_selaras_semua()
        except Exception:
            pass
        # Sinkronkan tvList ke Firestore setelah kartu dimuat (APK TV membaca ini)
        try:
            self._jadwal_sync_tv_list()
        except Exception:
            pass
        TimerService.restore_timer_state(self)
        self._load_riwayat()
        self._cek_adb_global_saat_start()

        # Firebase auth + auto-restore kartu TV di thread agar UI tidak
        # terkunci saat jaringan lambat (fix freeze "5 detik not responding").
        def _startup_async():
            try:
                get_firebase_auth().ensure_anonymous()
            except Exception:
                pass
            try:
                self._auto_restore_daftar_tv()
            except Exception as e:
                print(f"[TV] Auto-restore error: {e}", flush=True)

        threading.Thread(target=_startup_async, daemon=True).start()

        # Restore license from Firestore in background, then re-check license status
        # This syncs license from Android (which writes to billingps_users/{user}/licenseStatus)
        def _restore_and_recheck():
            try:
                self._try_restore_license_from_cloud()
            except Exception as e:
                _LOGGER.warning("Cloud license restore error: %s", e)
            try:
                self.after(100, self._cek_lisensi_saat_start)
            except Exception:
                pass

        threading.Thread(target=_restore_and_recheck, daemon=True).start()

        # Initial check (may show expired briefly until cloud restore completes)
        self._cek_lisensi_saat_start()

        # Start background update checker (non-blocking)
        try:
            threading.Thread(target=self._start_update_checker, daemon=True).start()
        except Exception:
            pass

        # Start periodic background checker with sidebar notification
        try:
            threading.Thread(target=self._check_for_updates_background, daemon=True).start()
        except Exception:
            pass

        # Fetch promo settings once after login (di thread, jangan blokir UI)
        try:
            threading.Thread(target=self._load_promo_bg, daemon=True).start()
        except Exception:
            pass

        # Start license poller untuk pantau perubahan dari cloud
        try:
            self._start_license_poller()
        except Exception:
            pass

        # QR panggil kasir: poller + menu cloud (halaman web pelanggan)
        try:
            self._start_call_poller()
        except Exception:
            pass
        try:
            self._qr_push_menu_bg()
        except Exception:
            pass
        try:
            self._booking_push_tv_status()
        except Exception:
            pass

        # Sinkronkan seluruh riwayat lokal ke cloud (idempoten, dedupe by id),
        # lalu retry tiap 30 detik jika ada antrian yang gagal.
        try:
            self._pending_tx_uploads = [
                t for t in (
                    self._build_tx_cloud(r, m)
                    for r, m in zip(self.riwayat_transaksi, self.riwayat_meta)
                ) if t
            ]
            threading.Thread(target=self._flush_cloud_uploads, daemon=True).start()
            self._schedule_cloud_retry()
        except Exception:
            pass

        # Single-session: daftarkan sesi di cloud + pantau jika dipakai PC lain
        try:
            self._session_login_at = int(time.time() * 1000)
            self._start_session_poller()
        except Exception as e:
            _LOGGER.warning("Session start error: %s", e)

        # Auto-import riwayat cloud saat login (admin; dedupe by id, tanpa popup)
        if (self.current_role or "") == "admin":
            try:
                threading.Thread(
                    target=self._import_riwayat_from_cloud,
                    kwargs={"silent": True},
                    daemon=True).start()
            except Exception:
                pass

        # Auto-lock saat idle (kunci ke login setelah 10 menit tidak aktif)
        try:
            self._start_idle_watcher()
        except Exception:
            pass

        # Watchdog: deteksi hang UI (heartbeat >5 dtk) → tulis diagnostic_hang.log
        try:
            self._start_watchdog()
        except Exception:
            pass

    def _start_watchdog(self):
        """Heartbeat 1 dtk dari main thread; monitor thread menulis stack dump
        semua thread ke diagnostic_hang.log bila UI tidak bernapas >5 dtk."""
        if getattr(self, "_wd_started", False):
            return
        self._wd_started = True
        self._wd_last_beat = time.time()

        def _beat():
            self._wd_last_beat = time.time()
            try:
                self.after(1000, _beat)
            except Exception:
                pass

        def _monitor():
            while True:
                time.sleep(2)
                if time.time() - self._wd_last_beat <= 5:
                    continue
                try:
                    lines = [
                        f"--- RRBILLINGPRO HANG DIAGNOSTIC {datetime.now().isoformat()} "
                        f"(UI tidak responsif >5 dtk) ---",
                        f"PID {os.getpid()} | app v{APP_VERSION} | user: {getattr(self, 'current_user', '')}",
                    ]
                    frames = sys._current_frames()
                    for th in threading.enumerate():
                        frm = frames.get(th.ident)
                        if frm is not None:
                            try:
                                import traceback
                                lines.append(f"\n=== Thread '{th.name}' (daemon={th.daemon}, ident={th.ident}) ===")
                                lines.append("".join(traceback.format_stack(frm)))
                            except Exception:
                                pass
                    try:
                        with open(app_path("diagnostic_hang.log"), "a", encoding="utf-8") as f:
                            f.write("\n".join(lines) + "\n")
                    except Exception:
                        pass
                except Exception:
                    pass
                self._wd_last_beat = time.time()

        threading.Thread(target=_monitor, daemon=True).start()
        try:
            self.after(1000, _beat)
        except Exception:
            pass

    def _try_restore_license_from_cloud(self):
        """Coba restore lisensi dari Firestore setelah login dengan 4 sumber."""
        if not self.current_user:
            return

        # Lisensi file = cache cloud; binding mengikuti admin mesin ini
        # (kasir ikut admin_utama). Hapus lisensi lokal CUKUP bila login yang
        # datang adalah pemilik mesin — jangan biarkan admin lain merebutnya.
        import pathlib as _pl_try
        _lic_p = _pl_try.Path("rr_billing_license.json")
        resolved_user = (self._resolve_license_user() or (self.current_user or "")).strip()
        if _lic_p.exists():
            try:
                _old = json.loads(_lic_p.read_text())
                _old_user = _old.get("username") or ""
                if _old_user and _old_user != resolved_user:
                    admin_utama = None
                    try:
                        for _u in (ConfigManager.get("users", {}) or {}).values():
                            if isinstance(_u, dict) and _u.get("admin_utama"):
                                admin_utama = _u.get("admin_utama")
                    except Exception:
                        admin_utama = None
                    if admin_utama and resolved_user != admin_utama:
                        _LOGGER.info("Lisensi milik '%s' dipertahankan (pemilik mesin '%s').",
                                     _old_user, admin_utama)
                        return
                    _lic_p.unlink()
                    _LOGGER.info("Removed old license for user '%s' to restore cloud license for '%s'",
                                 _old_user, resolved_user)
            except Exception:
                pass

        _EDITION_RANK = {"BULANAN": 0, "3BULAN": 1, "TAHUNAN": 2, "LIFETIME": 3}

        def _save_edition_from_max_tv(lic: dict, max_tv: int):
            # maxTv tidak valid/hilang (<=0) → jangan ubah edition:
            # LIFETIME (expiry 2099+) tetap LIFETIME, jangan downgrade ke BULANAN.
            if max_tv >= 999999:
                _new = "LIFETIME"
            elif max_tv >= 15:
                _new = "TAHUNAN"
            elif max_tv >= 10:
                _new = "3BULAN"
            elif max_tv > 0:
                _new = "BULANAN"
            else:
                return
            _old = str(lic.get("edition", "")).strip().upper() if "edition" in lic else ""
            if _old and _EDITION_RANK.get(_old, -1) >= _EDITION_RANK.get(_new, -1):
                return  # jangan pernah downgrade edition dari lisensi lokal
            lic["edition"] = _new

        def _write_cloud_license(expires_at: str, kode_aktivasi: str = "", max_tv: int = 0, promo_add_tv: int = 0):
            import pathlib as _pl
            import datetime as _dt
            lic_path = _pl.Path("rr_billing_license.json")
            lic = {}
            if lic_path.exists():
                lic = json.loads(lic_path.read_text())
            # Lisensi file = cache cloud. Binding mengikuti ADMIN yang login
            # (kasir ikut admin_utama). Lisensi milik user lain TIDAK ditimpa,
            # kecuali admin yang login sekarang adalah pemilik data mesin ini
            # (terdaftar sebagai admin_utama pada akun di config).
            existing_user = lic.get("username", "")
            resolved_user = (self._resolve_license_user() or (self.current_user or "")).strip()
            if lic.get("aktif") and existing_user and existing_user != resolved_user:
                admin_utama = None
                try:
                    for _u in (ConfigManager.get("users", {}) or {}).values():
                        if isinstance(_u, dict) and _u.get("admin_utama"):
                            admin_utama = _u.get("admin_utama")
                except Exception:
                    admin_utama = None
                if not (admin_utama and resolved_user == admin_utama):
                    return
            lic["aktif"] = True
            if not max_tv or max_tv <= 0:
                # maxTv hilang/0 di cloud: jangan tebak 5 — edition dipertahankan
                # (rank guard di _save_edition_from_max_tv mencegah downgrade).
                _LOGGER.warning("Cloud license tanpa maxTv valid (%s); edition lokal dipertahankan",
                                (lic.get("username") or "") or (self.current_user or ""))
            # Proteksi: jangan memotong masa lisensi lokal yang lebih lama
            try:
                _old_exp_raw = lic.get("expiry", "")
                _old_exp = _dt.datetime.fromisoformat(
                    (_old_exp_raw + "T00:00:00" if "T" not in _old_exp_raw else _old_exp_raw).replace("Z", "+00:00"))
                _new_exp = _dt.datetime.fromisoformat(
                    (expires_at + "T00:00:00" if "T" not in expires_at else expires_at).replace("Z", "+00:00"))
                if not _old_exp_raw or _new_exp > _old_exp:
                    lic["expiry"] = expires_at
            except Exception:
                lic["expiry"] = expires_at
            lic["firebase_sync"] = True
            lic["binding_mode"] = "username"
            lic["username"] = resolved_user or self.current_user
            _save_edition_from_max_tv(lic, max_tv)
            if kode_aktivasi:
                lic["kode_aktivasi"] = kode_aktivasi
            lic["promo_add_tv"] = promo_add_tv or lic.get("promo_add_tv", 0)
            lic["promo_add_warnet"] = lic["promo_add_tv"]
            LicenseManager.save(lic)
            try:
                self.after(200, self._rebuild_sidebar_lic)
                self.after(500, self._save_promo_override_to_license)
            except Exception:
                pass

        try:
            fc = FirestoreClient()
            ls = fc.fetch_license_status_by_username(self.current_user)
            if ls and ls.get("status") == "active" and ls.get("expiresAt"):
                import datetime as _dt
                expires = _dt.datetime.fromisoformat(ls["expiresAt"].replace("Z", "+00:00"))
                if expires.tzinfo is None:
                    expires = expires.replace(tzinfo=_dt.timezone.utc)
                if expires > _dt.datetime.now(_dt.timezone.utc):
                    _LOGGER.info("Cloud license restore: found via licenseStatus, maxTv=%s", ls.get("maxTv"))
                    _write_cloud_license(ls["expiresAt"], max_tv=ls.get("maxTv") or 0, promo_add_tv=ls.get("promoAddTv") or 0)
                    # Re-check lisensi setelah restore agar UI update
                    self.after(1000, self._cek_lisensi_saat_start)
                return
            # Fallback: cari langsung di koleksi licenses/
            try:
                import datetime as _dt
                ld = fc.get_document(f"licenses/{self.current_user}")
                if ld:
                    if ld.get("expiry") and not ld.get("revoked"):
                        expires = _dt.datetime.fromisoformat(ld["expiry"].replace("Z", "+00:00"))
                        if expires.tzinfo is None:
                            expires = expires.replace(tzinfo=_dt.timezone.utc)
                        if expires > _dt.datetime.now(_dt.timezone.utc):
                            _LOGGER.info("Cloud license restore: found via licenses/")
                            _write_cloud_license(ld["expiry"], ld.get("kode", ""), max_tv=ld.get("maxTv") or 0)
                            self.after(1000, self._cek_lisensi_saat_start)
                            return
            except Exception:
                pass
            # Fallback: cari di invoices/
            try:
                import datetime as _dt
                invs = fc.query_where_equal("invoices", "username", self.current_user)
                for iv in invs:
                    if iv.get("revoked"):
                        continue
                    if iv.get("status", "").upper() == "CONFIRMED" and iv.get("kodeLisensi"):
                        kode = iv["kodeLisensi"]
                        import rr_keygen
                        pkg_name = iv.get("package", "BULANAN").upper()
                        pkg = rr_keygen.PAKET_INFO.get(pkg_name)
                        if pkg:
                            expires_at = _dt.datetime.fromtimestamp(iv.get("expiresAt", 0)).isoformat() if iv.get("expiresAt") else pkg.get("expiresAt", "")
                            _LOGGER.info("Cloud license restore: found via invoices/")
                            _write_cloud_license(expires_at, kode, max_tv=pkg.get("maxTv") or 0)
                            self.after(1000, self._cek_lisensi_saat_start)
                            return
            except Exception:
                pass
        except Exception as e:
            _LOGGER.warning("Cloud license restore error: %s", e)

    def _cek_adb_global_saat_start(self):
        if self.current_role == "admin" and not ADBHelper.adb_tersedia():
            self.after(500, lambda: messagebox.showwarning(
                "⚠ androidtvremote2 Tidak Terinstal",
                "Package 'androidtvremote2' tidak ditemukan.\n"
                "Jalankan: pip install androidtvremote2\n"
                "Koneksi ke TV tidak akan berfungsi tanpa package ini."))

    def _cek_lisensi_saat_start(self):
        status = LicenseManager.get_status(current_user=self._resolve_license_user())
        if status["status"] == "active":
            # Re-enable semua tab
            for tab_key, btn in self.nav_btns.items():
                btn.configure(state="normal", text_color=C_TEXT)
            self._rebuild_sidebar_lic()
            return
        if status["status"] == "expired":
            # Disable semua tab kecuali Aktivasi
            for tab_key, btn in self.nav_btns.items():
                if tab_key != "aktivasi":
                    btn.configure(state="disabled", text_color=C_MUTED)
            
            # Tampilkan warning
            self.after(600, lambda: messagebox.showwarning(
                "⚠ LISENSI HABIS — AKSES TERBATAS",
                f"{status['pesan']}\n\n"
                f"Semua fitur telah dikunci.\n"
                f"Silakan aktifkan lisensi untuk melanjutkan menggunakan aplikasi."))
            
            # Paksa show tab Aktivasi
            self.after(700, lambda: self._show_tab("aktivasi"))

    # ── Layout ─────────────────────────────────────────────────────────────────
    def _build_layout(self):
        self.sidebar = ctk.CTkScrollableFrame(self, width=185, fg_color=C_PANEL,
                                              corner_radius=0, label_text="")
        self.sidebar.pack(side="left", fill="y")

        self.content = ctk.CTkFrame(self, fg_color=C_BG, corner_radius=0)
        self.content.pack(side="left", fill="both", expand=True)

        self._build_sidebar()

        self.frames = {}
        for name in ["dashboard", "warnet", "harga", "stok", "riwayat", "booking", "wifi", "aktivasi", "profil", "log_aplikasi", "users"]:
            f = ctk.CTkFrame(self.content, fg_color=C_BG, corner_radius=0)
            self.frames[name] = f

        self._setup_dashboard()
        self._setup_warnet()
        self._setup_harga()
        self._setup_stok()
        self._setup_riwayat()
        self._setup_booking()
        self._setup_wifi()
        self._setup_aktivasi()
        self._setup_profil()
        self._setup_log_aplikasi()
        # Admin-only kasir management tab (APTV2-style)
        if self.current_role == "admin":
            self._setup_users()
        self._show_tab("dashboard")

    def _build_sidebar(self):
        self._sidebar_text_widgets = []

        # Toggle button
        logo_f = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        logo_f.pack(pady=(22, 6))

        # ── LOGO SIDEBAR: coba logo.png, fallback emoji ───────────────────────
        ico_bg = ctk.CTkFrame(logo_f, fg_color=C_PANEL, corner_radius=10,
                               width=128, height=54)
        ico_bg.pack()
        ico_bg.pack_propagate(False)

        ctk_img_sidebar = load_ctk_image(size=(120, 46))
        if ctk_img_sidebar:
            lbl_sb_logo = ctk.CTkLabel(ico_bg, text="", image=ctk_img_sidebar)
        else:
            lbl_sb_logo = ctk.CTkLabel(ico_bg, text="🎮", font=("Arial", 24))
        lbl_sb_logo.place(relx=0.5, rely=0.5, anchor="center")

        ctk.CTkLabel(logo_f, text="RR BILLING",
                     font=("Russo One", 13, "bold"),
                     text_color=C_ACCENT, justify="center").pack(pady=(6, 0))
        ctk.CTkLabel(logo_f, text="PRO",
                     font=("Russo One", 11, "bold"),
                     text_color=C_ACCENT2).pack()

        status = LicenseManager.get_status(current_user=self._resolve_license_user())
        lic_color = C_GREEN if status["status"] == "active" else C_YELLOW if status["status"] == "trial" else C_RED
        self.lbl_sidebar_license_status = ctk.CTkLabel(self.sidebar, text=status["pesan"],
                                                       font=("Courier New", 10), text_color=lic_color,
                                                       wraplength=165)
        self.lbl_sidebar_license_status.pack(pady=(2, 10))

        ctk.CTkLabel(self.sidebar, text=f"👤 {self.current_user} [{self.current_role}]",
                     font=("Courier New", 10), text_color=C_MUTED).pack(pady=(0, 2))

        sep = ctk.CTkFrame(self.sidebar, height=1, fg_color=C_BORDER)
        sep.pack(fill="x", padx=10, pady=4)

        # Kasir (sub-akun admin): hanya tab yang aman — Dashboard TV, Warnet,
        # Histori Aktivasi (read-only) & Riwayat. Admin melihat semua menu.
        is_admin = (self.current_role or "kasir") == "admin"
        if is_admin:
            nav_items = [
                ("📺", "Dashboard TV",    "dashboard"),
                ("🖥️", "Dashboard Warnet", "warnet"),
                ("📜", "Riwayat",         "riwayat"),
                ("📅", "Booking",         "booking"),
                ("💰", "Kontrol Harga",   "harga"),
                ("📦", "Stok",            "stok"),
                ("🔗", "Panduan Koneksi",  "wifi"),
                ("🔓", "Aktivasi",        "aktivasi"),
                ("👤", "Profil",          "profil"),
                ("📋", "Log Aplikasi",    "log_aplikasi"),
                ("👥", "Manajemen Kasir", "users"),
            ]
        else:
            nav_items = [
                ("📺", "Dashboard TV",    "dashboard"),
                ("🖥️", "Dashboard Warnet", "warnet"),
                ("📋", "Histori Aktivasi", "aktivasi"),
                ("📜", "Riwayat",         "riwayat"),
                ("📅", "Booking",         "booking"),
            ]
        self.nav_btns = {}
        for ico, label, key in nav_items:
            row = ctk.CTkFrame(self.sidebar, fg_color="transparent")
            row.pack(fill="x", padx=10, pady=2)
            ctk.CTkLabel(row, text=ico, font=("Segoe UI Emoji", 14),
                         width=28, anchor="center").pack(side="left", padx=(2, 0))
            btn = ctk.CTkButton(
                row, text=f"  {label}", anchor="w", height=40,
                font=("Russo One", 12, "bold"),
                fg_color="transparent", hover_color="#1E1E4A",
                text_color=C_TEXT, corner_radius=8,
                command=lambda k=key: self._show_tab(k))
            btn.pack(side="left", fill="x", expand=True, padx=(0, 2))
            self.nav_btns[key] = btn

        sep2 = ctk.CTkFrame(self.sidebar, height=1, fg_color=C_BORDER)
        sep2.pack(fill="x", padx=10, pady=(10, 4))

        row_keluar = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        row_keluar.pack(fill="x", padx=10, pady=2)
        ctk.CTkLabel(row_keluar, text="🚪", font=("Segoe UI Emoji", 14),
                     width=28, anchor="center").pack(side="left", padx=(2, 0))
        ctk.CTkButton(row_keluar, text="  Keluar", anchor="w", height=36,
                      font=("Russo One", 11, "bold"),
                      fg_color="transparent", hover_color="#3A0000",
                      text_color=C_RED, corner_radius=8,
                      command=self._logout).pack(side="left", fill="x", expand=True, padx=(0, 2))

        self._row_update = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        self._row_update.pack(fill="x", padx=10, pady=2)
        ctk.CTkLabel(self._row_update, text="🔄", font=("Segoe UI Emoji", 14),
                     width=28, anchor="center").pack(side="left", padx=(2, 0))
        ctk.CTkButton(self._row_update, text="  Download & Install", anchor="w", height=36,
                      font=("Russo One", 10, "bold"),
                      fg_color="transparent", hover_color="#1A3A5A",
                      text_color="#3A8AFF", corner_radius=8,
                      command=self._download_and_install_update).pack(side="left", fill="x", expand=True, padx=(0, 2))

        ctk.CTkLabel(self.sidebar, text=f"v{APP_VERSION} — 2026",
                     font=("Courier New", 10), text_color=C_MUTED).pack(side="bottom", pady=12)

    def _logout(self):
        if messagebox.askyesno("Keluar", "Yakin ingin keluar / ganti akun?"):
            AuditLogger.log(
                action="logout",
                username=self.current_user or "",
                status="success",
                details={"role": self.current_role or ""}
            )
            self._show_login()

    def _on_check_update(self):
        """Triggered by UI button. Reads manifest URL from config and runs check in background."""
        manifest = ConfigManager.get('update_manifest_url') or ""
        if not manifest.strip():
            messagebox.showinfo("Cek Pembaruan", 
                "Fitur pembaruan belum dikonfigurasi.\n\n"
                "Untuk mengaktifkan:\n"
                "1. Upload manifest.json ke GitHub release\n"
                "2. Setel 'update_manifest_url' di config.json\n\n"
                "Atau klik 'Download & Install' di sidebar untuk memasukkan URL manifest.")
            return
        # Run in thread to avoid blocking UI
        threading.Thread(target=self._check_update_thread, args=(manifest.strip(),), daemon=True).start()

    def _detect_update(self):
        """Ambil manifest + verifikasi tanda tangan + bandingkan versi (TANPA mengunduh).

        Return dict manifest jika ada pembaruan, None jika sudah terbaru,
        raise ValueError jika manifest tidak valid / signature salah.
        """
        from scripts import check_update
        manifest = ConfigManager.get('update_manifest_url') or ""
        if not manifest.strip():
            return None
        mf = check_update.fetch_manifest(manifest.strip())
        if not check_update.verify_manifest(mf, ConfigManager.get('update_pubkey_path') or None):
            raise ValueError('Signature manifest tidak valid')
        if str(mf.get('version', '')).strip() == str(APP_VERSION).strip():
            return None
        return mf

    def _check_update_thread(self, manifest_url: str):
        try:
            mf = self._detect_update()
            if mf is None:
                res = f"Versi terbaru terpasang ({APP_VERSION})."
            else:
                res = (f"Pembaruan tersedia: v{mf.get('version')} "
                       f"(versi Anda: v{APP_VERSION})\n\n"
                       "Klik 'Download & Install' di sidebar untuk mengunduh dan memasangnya.")
            self.after(0, lambda: messagebox.showinfo("Cek Pembaruan", res))
        except Exception as e:
            self.after(0, lambda e=e: messagebox.showerror("Cek Pembaruan - Error", str(e)))
    
    def _show_tab(self, key):
        role = self.current_role or "kasir"
        kasir_allowed = {"dashboard", "warnet", "aktivasi", "riwayat", "booking"}
        if role != "admin" and key not in kasir_allowed:
            messagebox.showwarning("⚠ AKSES TERBATAS", "Hanya admin yang dapat mengakses fitur ini.")
            return
        if key == "riwayat" and role == "admin":
            self.after_idle(self._refresh_kasir_filter_options)
        status = LicenseManager.get_status(current_user=self._resolve_license_user())
        if status["status"] == "expired" and key != "aktivasi":
            if role != "admin":
                messagebox.showwarning(
                    "⚠ AKSES TERBATAS",
                    "Lisensi telah habis. Silakan hubungi admin untuk memperpanjang lisensi.")
                return
            messagebox.showwarning(
                "⚠ AKSES TERBATAS",
                "Lisensi Anda telah habis. Silakan aktifkan lisensi untuk mengakses fitur lain.")
            self._show_tab("aktivasi")
            return
        
        self.current_tab = key
        for k, f in self.frames.items():
            f.pack_forget()
        self.frames[key].pack(fill="both", expand=True)
        # Tab Stok dibangun ulang setiap dibuka supaya item baru/hapus/rename
        # dari Kontrol Harga langsung terlihat.
        if key == "stok" and hasattr(self, "scroll_stok"):
            try:
                self._build_stok_section()
            except Exception:
                pass
        for k, btn in self.nav_btns.items():
            btn.configure(fg_color=C_ACCENT2 if k == key else "transparent",
                          text_color="white" if k == key else C_TEXT)
        if key in ("dashboard", "warnet") and ConfigManager.get("app_bg_image", ""):
            self.after(50, lambda: self._apply_bg_image(key))

    def get_paket_data(self, nama_grup=None, for_warnet=False):
        """Return paket dict for a group.
        If for_warnet=True, prefer warnet-specific groups stored under config key
        'grup_tarif_warnet'. Falls back to shared grup_tarif or standard paket.
        """
        cfg = ConfigManager.load()
        if for_warnet:
            warnet_map = cfg.get('grup_tarif_warnet', {}) or {}
            # Cek grup warnet dulu (exact, lalu case-insensitive)
            if nama_grup:
                k_found = None
                if nama_grup in warnet_map:
                    k_found = nama_grup
                else:
                    lower_map = {k.lower(): k for k in warnet_map.keys()}
                    k_found = lower_map.get(nama_grup.lower()) if isinstance(nama_grup, str) else None
                if k_found is not None:
                    grp = warnet_map[k_found]
                    if isinstance(grp, dict):
                        return {k: {"harga": int(v.get("harga", 0)), "menit": int(v.get("menit", 0))} for k, v in grp.items()}
                    return {k: {"harga": int(v), "menit": 60} for k, v in grp.items()}
                # Kartu warnet bisa juga mengikat ke grup shared (Reguler, PS3, dll.)
                # -> pakai harga shared yang sama seperti kartu TV
                if nama_grup in self.grup_tarif and nama_grup not in self._warnet_group_names():
                    return {k: dict(v) for k, v in self.grup_tarif[nama_grup].items()}
            # Fallback: grup 'Warnet', lalu grup warnet pertama
            if 'Warnet' in warnet_map:
                grp = warnet_map['Warnet']
                if isinstance(grp, dict):
                    return {k: {"harga": int(v.get("harga", 0)), "menit": int(v.get("menit", 0))} for k, v in grp.items()}
                return {k: {"harga": int(v), "menit": 60} for k, v in grp.items()}
            if warnet_map:
                first = next(iter(warnet_map.values()))
                if isinstance(first, dict):
                    return {k: {"harga": int(v.get("harga", 0)), "menit": int(v.get("menit", 0))} for k, v in first.items()}
                return {k: {"harga": int(v), "menit": 60} for k, v in first.items()}
            # Jika tidak ada konfigurasi warnet sama sekali, jangan ambil dari grup PS — kembalikan kosong
            return {}
        
        # Default behavior (kartu TV/PS): pakai shared grup_tarif saja.
        # Grup warnet TIDAK PERNAH diberikan ke kartu TV — fallback ke Reguler.
        warnet_only = self._warnet_group_names()
        if nama_grup:
            # Check shared grup_tarif first
            if nama_grup in self.grup_tarif and nama_grup not in warnet_only:
                return {k: dict(v) for k, v in self.grup_tarif[nama_grup].items()}
        
        if NAMA_GRUP_DEFAULT in self.grup_tarif:
            return {k: dict(v) for k, v in self.grup_tarif[NAMA_GRUP_DEFAULT].items()}
        if self.grup_tarif:
            first = next(iter(self.grup_tarif.values()))
            return {k: dict(v) for k, v in first.items()}
        return {k: dict(v) for k, v in _PAKET_STANDAR.items()}

    def get_makanan_data(self): return dict(self.menu_makanan)
    def get_minuman_data(self): return dict(self.menu_minuman)
    def get_semua_kartu(self):  return list(self._semua_kartu_tv)

    # ── Stok makanan & minuman ─────────────────────────────────────────────────
    def _stok_map(self, nama):
        """Cari kategori item: 'makanan' | 'minuman' | None (item tidak dikenal)."""
        if nama in self.menu_makanan:
            return "makanan"
        if nama in self.menu_minuman:
            return "minuman"
        return None

    def _stok_get(self, nama):
        """Stok saat ini utk item; None = item tidak dilacak stoknya."""
        kat = self._stok_map(nama)
        if kat is None:
            return None
        m = (self.stok.get(kat) or {}) if isinstance(self.stok, dict) else {}
        val = m.get(nama)
        if val is None:
            return None
        try:
            return max(0, int(val))
        except Exception:
            return None

    def _stok_min_get(self, nama):
        kat = self._stok_map(nama)
        if kat is None:
            return None
        m = (self.stok_min.get(kat) or {}) if isinstance(self.stok_min, dict) else {}
        val = m.get(nama)
        if val is None:
            return None
        try:
            return max(0, int(val))
        except Exception:
            return None

    def _stok_set(self, nama, qty, kat=None):
        """Set stok item (qty >= 0). Item jadi 'dilacak' sejak pertama di-set."""
        kat = kat or self._stok_map(nama)
        if kat not in ("makanan", "minuman"):
            return False
        if not isinstance(self.stok, dict):
            self.stok = {}
        if not isinstance(self.stok.get(kat), dict):
            self.stok[kat] = {}
        try:
            self.stok[kat][nama] = max(0, int(qty))
        except Exception:
            return False
        return True

    def _stok_min_set(self, nama, qty, kat=None):
        kat = kat or self._stok_map(nama)
        if kat not in ("makanan", "minuman"):
            return False
        if not isinstance(self.stok_min, dict):
            self.stok_min = {}
        if not isinstance(self.stok_min.get(kat), dict):
            self.stok_min[kat] = {}
        try:
            self.stok_min[kat][nama] = max(0, int(qty))
        except Exception:
            return False
        return True

    def _stok_save(self):
        """Simpan stok & stok_min ke config (thread-safe)."""
        try:
            cfg = ConfigManager.load()
            cfg["stok"] = self.stok
            cfg["stok_min"] = self.stok_min
            ConfigManager.save(cfg)
        except Exception:
            pass

    def _stok_validate_orders(self, pesanan):
        """Cek apakah semua item bisa dipenuhi stok. Return (ok, pesan, nama,
        stok_sisa). Pesan None = semua aman."""
        if not pesanan:
            return True, None, None, None
        for nama, qty in pesanan.items():
            sisa = self._stok_get(nama)
            if sisa is None:
                continue  # tidak dilacak
            if int(qty or 0) > sisa:
                return (False,
                        f"Stok '{nama}' tidak mencukupi.\n"
                        f"Tersedia: {sisa}, diminta: {int(qty or 0)}.\n"
                        "Kurangi jumlah atau restok dulu.",
                        nama, sisa)
        return True, None, None, None

    def _stok_terapkan(self, delta):
        """Terapkan selisih stok dari pesanan: delta > 0 = kurangi, delta < 0 =
        kembalikan (mis. kasir mengurangi qty item yang sudah pernah dipesan)."""
        if not delta:
            return
        berubah = False
        for nama, d in delta.items():
            d = int(d or 0)
            if d == 0:
                continue
            sisa = self._stok_get(nama)
            if sisa is None:
                continue  # tidak dilacak
            if self._stok_set(nama, max(0, sisa - d), self._stok_map(nama)):
                berubah = True
        if berubah:
            self._stok_save()
            self._stok_push_bg()
            self._stok_update_badge()

    def _stok_push_bg(self):
        """Push stok terbaru ke call_meta Firestore (halaman web pelanggan)."""
        try:
            if not getattr(self, "_stok_push_queued", False):
                self._stok_push_queued = True
                self.after(3000, self._qr_push_menu_bg)
        except Exception:
            pass

    def _stok_tracked_items(self):
        """Daftar (kategori, nama, stok, stok_min) utk semua item menu."""
        out = []
        for kat, menu in (("makanan", self.menu_makanan), ("minuman", self.menu_minuman)):
            m = (self.stok.get(kat) or {}) if isinstance(self.stok, dict) else {}
            mn = (self.stok_min.get(kat) or {}) if isinstance(self.stok_min, dict) else {}
            for nama in menu:
                try:
                    stok = max(0, int(m.get(nama, 0) or 0))
                except Exception:
                    stok = 0
                try:
                    mini = max(0, int(mn.get(nama, 0) or 0))
                except Exception:
                    mini = 0
                out.append((kat, nama, stok, mini))
        return out

    def _stok_menipis(self):
        """Item dilacak dengan stok <= stok_min (atau stok 0)."""
        items = []
        for kat, nama, stok, mini in self._stok_tracked_items():
            mini_eff = mini
            if mini_eff <= 0 and stok <= 0:
                mini_eff = 0
            if stok <= mini_eff:
                items.append((kat, nama, stok, mini))
        return items

    def _stok_update_badge(self):
        """Perbarui badge 'stok menipis' di Dashboard TV & tab Stok."""
        try:
            items = self._stok_menipis()
            by_kat = {}
            for kat, nama, stok, mini in items:
                by_kat.setdefault(kat, []).append((nama, stok, mini))
            if not by_kat:
                for lbl in (getattr(self, "_stok_tracked_lbl", None),
                            getattr(self, "_stok_dash_lbl", None)):
                    if lbl is not None:
                        try:
                            lbl.configure(text="", fg_color="transparent")
                        except Exception:
                            pass
                return
            txt = "📦 Stok menipis: "
            txt += " | ".join(
                f"{'Makanan' if k == 'makanan' else 'Minuman'} {len(v)} item"
                for k, v in by_kat.items())
            for lbl in (getattr(self, "_stok_tracked_lbl", None),
                        getattr(self, "_stok_dash_lbl", None)):
                if lbl is not None:
                    try:
                        lbl.configure(text=txt, fg_color="#3A3A2A")
                    except Exception:
                        pass
        except Exception:
            pass

    def _stok_poll_badge(self):
        try:
            self._stok_update_badge()
        except Exception:
            pass
        try:
            self.after(60000, self._stok_poll_badge)
        except Exception:
            pass

    def _remap_stok_rename(self, kat, nama_asli, nama_baru):
        """Ikuti rename item: pindahkan entri stok lama ke nama baru."""
        if not isinstance(self.stok, dict) or not isinstance(self.stok.get(kat), dict):
            return
        m = self.stok[kat]
        if nama_asli in m and nama_asli != nama_baru:
            m[nama_baru] = m.pop(nama_asli)
        mn = self.stok_min.get(kat)
        if isinstance(mn, dict) and nama_asli in mn and nama_asli != nama_baru:
            mn[nama_baru] = mn.pop(nama_asli)

    def _hapus_stok_item(self, kat, nama):
        if isinstance(self.stok, dict) and isinstance(self.stok.get(kat), dict):
            self.stok[kat].pop(nama, None)
        if isinstance(self.stok_min, dict) and isinstance(self.stok_min.get(kat), dict):
            self.stok_min[kat].pop(nama, None)

    def _stok_default_stok_min(self):
        """Default stok_min utk item yang baru pertama kali dilacak."""
        return 3


    def _tv_idle_guard(self):
        """Poller auto-mati TV yang menyala tanpa paket aktif (anti-kasir nakal).

        Bekerja hanya jika tv_auto_off_enabled = true; status layar diambil
        dari WebSocket APK (sumber akurat), bukan ADB."""
        try:
            if getattr(self, "tv_auto_off_enabled", False):
                for kartu in list(getattr(self, "_semua_kartu_tv", [])):
                    try:
                        kartu._tv_idle_check()
                    except Exception:
                        pass
        except Exception:
            pass
        try:
            self.after(30000, self._tv_idle_guard)
        except Exception:
            pass

    def _refresh_dashboard_total_pesanan(self):
        total = sum(getattr(kartu, 'biaya_pesanan', 0) for kartu in self._semua_kartu_tv)
        if hasattr(self, 'lbl_dashboard_total_pesanan'):
            self.lbl_dashboard_total_pesanan.configure(text=f"Total Pesanan: {fmt_rp(total)}")

    # ── Filter Riwayat (tanggal & kasir) ──────────────────────────────────────
    def _riwayat_kasir_choices(self):
        if (self.current_role or "") == "admin":
            choices = ["SEMUA"]
            users = ConfigManager.get("users", {}) or {}
            if isinstance(users, dict):
                for uname, u in users.items():
                    if isinstance(u, dict) and u.get("role", "kasir") == "kasir":
                        choices.append(uname)
            return choices or ["SEMUA"]
        return [self.current_user or "SEMUA"]

    def _row_matches_filter(self, row):
        if not row:
            return False
        if getattr(self, "_riwayat_filter_range", None):
            waktu = str(row[0] if len(row) > 0 else "")
            hari = waktu[:10]
            if not hari or not (self._riwayat_filter_range[0] <= hari <= self._riwayat_filter_range[1]):
                return False
        elif self._riwayat_filter_tanggal:
            waktu = str(row[0] if len(row) > 0 else "")
            if not waktu.startswith(self._riwayat_filter_tanggal):
                return False
        kasir_filter = self._riwayat_filter_kasir or "SEMUA"
        if kasir_filter != "SEMUA":
            kasir = row[1] if len(row) > 1 else ""
            if kasir != kasir_filter:
                return False
        cari = (getattr(self, "_riwayat_filter_cari", "") or "").strip().lower()
        if cari:
            hay = " ".join(str(c or "") for c in row[:6]).lower()
            if cari not in hay:
                return False
        return True

    def _riwayat_filter_indices(self):
        out = []
        status_f = getattr(self, "_riwayat_filter_status", "SEMUA") or "SEMUA"
        for i, r in enumerate(self.riwayat_transaksi):
            if not self._row_matches_filter(r):
                continue
            if status_f != "SEMUA":
                paid = self.riwayat_meta[i].get('paid', True) if i < len(self.riwayat_meta) else True
                if (status_f == "LUNAS") != paid:
                    continue
            out.append(i)
        return out

    def _render_riwayat_tree(self):
        if not hasattr(self, "tree"):
            return
        self._tree_item_to_index = {}
        for item in self.tree.get_children():
            self.tree.delete(item)
        for idx in reversed(self._riwayat_filter_indices()):
            row = self.riwayat_transaksi[idx]
            paid = self.riwayat_meta[idx].get('paid', True) if idx < len(self.riwayat_meta) else True
            tag = "paid" if paid else "unpaid"
            item_id = self.tree.insert("", 0, values=row, tags=(tag,))
            self._tree_item_to_index[item_id] = idx
        self._refresh_riwayat_summary()
        self._remap_kartu_transaction_items()

    def _set_filter_tanggal(self, mode):
        self._riwayat_filter_range = None
        if mode is None:
            self._riwayat_filter_tanggal = None
            self.lbl_filter_tanggal.configure(text="Semua Tanggal")
        elif mode == "7_hari":
            self._riwayat_filter_tanggal = None
            awal = datetime.now() - timedelta(days=6)
            akhir = datetime.now()
            self._riwayat_filter_range = (awal.strftime("%Y-%m-%d"), akhir.strftime("%Y-%m-%d"))
            self.lbl_filter_tanggal.configure(
                text=f"7 Hari Terakhir ({self._riwayat_filter_range[0]} → {self._riwayat_filter_range[1]})")
        elif mode == "bulan_ini":
            self._riwayat_filter_tanggal = None
            now = datetime.now()
            awal = now.replace(day=1)
            self._riwayat_filter_range = (awal.strftime("%Y-%m-%d"), now.strftime("%Y-%m-%d"))
            self.lbl_filter_tanggal.configure(text=f"Bulan Ini ({awal.strftime('%b %Y')})")
        elif mode == "hari_ini":
            self._riwayat_filter_tanggal = datetime.now().strftime("%Y-%m-%d")
            self.lbl_filter_tanggal.configure(
                text=f"Hari Ini ({self._riwayat_filter_tanggal})")
        elif mode == "kemarin":
            kemarin = datetime.now() - timedelta(days=1)
            self._riwayat_filter_tanggal = kemarin.strftime("%Y-%m-%d")
            self.lbl_filter_tanggal.configure(
                text=f"Kemarin ({self._riwayat_filter_tanggal})")
        else:
            self._riwayat_filter_tanggal = mode
            try:
                tgl = datetime.strptime(mode, "%Y-%m-%d")
                self.lbl_filter_tanggal.configure(text=tgl.strftime("%d %b %Y"))
            except Exception:
                self.lbl_filter_tanggal.configure(text=mode)
        self._render_riwayat_tree()

    def _on_cari_ketik(self, _evt=None):
        """Debounce pencarian riwayat: rebuild tree hanya setelah berhenti mengetik 300 ms."""
        job = getattr(self, "_cari_debounce_job", None)
        if job:
            try:
                self.after_cancel(job)
            except Exception:
                pass
        self._cari_debounce_job = self.after(300, self._apply_riwayat_filter)

    def _apply_riwayat_filter(self):
        if hasattr(self, "opt_filter_kasir"):
            self._riwayat_filter_kasir = self.opt_filter_kasir.get()
        if hasattr(self, "opt_filter_status"):
            self._riwayat_filter_status = self.opt_filter_status.get()
        if hasattr(self, "ent_filter_cari"):
            self._riwayat_filter_cari = self.ent_filter_cari.get()
        self._render_riwayat_tree()

    def _refresh_kasir_filter_options(self):
        if (self.current_role or "") != "admin":
            return
        if not hasattr(self, "opt_filter_kasir"):
            return
        choices = self._riwayat_kasir_choices()
        self.opt_filter_kasir.configure(values=choices)
        if (self._riwayat_filter_kasir or "SEMUA") not in choices:
            self._riwayat_filter_kasir = "SEMUA"
            self.opt_filter_kasir.set("SEMUA")
            self._render_riwayat_tree()

    def _pilih_tanggal_kalender(self, dlg, tgl_str):
        self._riwayat_filter_tanggal = tgl_str
        try:
            tgl = datetime.strptime(tgl_str, "%Y-%m-%d")
            self.lbl_filter_tanggal.configure(text=tgl.strftime("%d %b %Y"))
        except Exception:
            self.lbl_filter_tanggal.configure(text=tgl_str)
        dlg.destroy()
        self._render_riwayat_tree()

    def _buka_kalender(self):
        """Kalender sederhana untuk cek riwayat per hari (tanpa dependensi baru)."""
        import calendar as _cal
        dlg = ctk.CTkToplevel(self.winfo_toplevel())
        dlg.title("📅 Pilih Tanggal")
        dlg.configure(fg_color=C_BG)
        dlg.transient(self.winfo_toplevel())
        dlg.resizable(False, False)
        dlg.grab_set()

        now = datetime.now()
        bulan = now.month
        tahun = now.year

        hdr_f = ctk.CTkFrame(dlg, fg_color="transparent")
        hdr_f.pack(pady=(10, 4))
        lbl_periode = ctk.CTkLabel(hdr_f, text="", font=("Russo One", 13, "bold"),
                                   text_color=C_ACCENT)
        grid_f = ctk.CTkFrame(dlg, fg_color=C_PANEL, corner_radius=10)
        grid_f.pack(padx=12, pady=(0, 8))

        def _render():
            lbl_periode.configure(text=f"{_cal.month_name[bulan]} {tahun}")
            for w in grid_f.winfo_children():
                w.destroy()
            for i, nama in enumerate(["Sn", "Sl", "Rb", "Km", "Jm", "Sb", "Mg"]):
                ctk.CTkLabel(grid_f, text=nama, font=FONT_SMALL, text_color=C_MUTED,
                             width=38).grid(row=0, column=i, padx=1, pady=1)
            for r, week in enumerate(_cal.monthcalendar(tahun, bulan), 1):
                for c, day in enumerate(week):
                    if day == 0:
                        continue
                    tgl_str = f"{tahun}-{bulan:02d}-{day:02d}"
                    is_today = (day == now.day and bulan == now.month and tahun == now.year)
                    ctk.CTkButton(grid_f, text=str(day), width=38, height=32,
                                  fg_color=C_ACCENT2 if is_today else C_BTN,
                                  hover_color=C_ACCENT,
                                  font=("Consolas", 11, "bold") if is_today else ("Consolas", 11),
                                  text_color="white" if is_today else C_TEXT,
                                  command=lambda t=tgl_str: self._pilih_tanggal_kalender(dlg, t)
                                  ).grid(row=r, column=c, padx=1, pady=1)

        def _geser(delta):
            nonlocal bulan, tahun
            bulan += delta
            if bulan < 1:
                bulan, tahun = 12, tahun - 1
            elif bulan > 12:
                bulan, tahun = 1, tahun + 1
            _render()

        ctk.CTkButton(hdr_f, text="‹", width=34, height=30, fg_color=C_BTN,
                      hover_color=C_ACCENT2, font=("Russo One", 14, "bold"),
                      text_color=C_TEXT, command=lambda: _geser(-1)).pack(side="left", padx=(0, 6))
        lbl_periode.pack(side="left", padx=8)
        ctk.CTkButton(hdr_f, text="›", width=34, height=30, fg_color=C_BTN,
                      hover_color=C_ACCENT2, font=("Russo One", 14, "bold"),
                      text_color=C_TEXT, command=lambda: _geser(1)).pack(side="left", padx=(6, 0))
        _render()

    def _refresh_riwayat_summary(self):
        # Ringkasan mengikuti filter aktif (tanggal + kasir), bukan hanya user sendiri
        user_indices = self._riwayat_filter_indices()
        user_metas = [self.riwayat_meta[i] for i in user_indices if i < len(self.riwayat_meta)]

        def paket_only(m):
            return m.get('paket_harga', m.get('total', 0) - m.get('pesanan_total', 0))
        total_tv_paket = sum(paket_only(m) for m in user_metas if m.get('source') == 'tv')
        total_warnet_paket = sum(paket_only(m) for m in user_metas if m.get('source') == 'warnet')
        total_pesanan = sum(m.get('pesanan_total', 0) for m in user_metas)
        total_all = sum(m.get('total', 0) for m in user_metas)
        paid_count = sum(1 for m in user_metas if m.get('paid', True))
        unpaid_count = sum(1 for m in user_metas if not m.get('paid', True))
        def _no_rp(n):
            return f"{n:,.0f}".replace(",", ".")
        summary_text = (
            f"TV Rp {_no_rp(total_tv_paket)}  |  PC Rp {_no_rp(total_warnet_paket)}  |  "
            f"F&B Rp {_no_rp(total_pesanan)}  |  Total Rp {_no_rp(total_all)}"
        )
        # Keep a short summary in the left label as before
        short_text = f"Total Transaksi: {len(user_indices)}  |  Total Pendapatan: {fmt_rp(total_all)}"
        self.lbl_rekap.configure(text=short_text)
        if hasattr(self, 'lbl_rekap_footer'):
            # show full breakdown in footer
            status_text = f"  |  ✅ Lunas: {paid_count}  |  ⏳ Belum: {unpaid_count}"
            self.lbl_rekap_footer.configure(text=summary_text + status_text)

    # ── Lisensi sub-akun kasir ────────────────────────────────────────────────
    def _resolve_license_user(self):
        """Kasir adalah sub-akun admin utama: lisensi mengikuti admin (APTV2).
        Admin menggunakan usernamenya sendiri."""
        if (self.current_role or "") != "kasir":
            return self.current_user or ""
        try:
            users = ConfigManager.get("users", {}) or {}
            u = users.get(self.current_user) or {}
            if isinstance(u, dict):
                return u.get("admin_utama") or self.current_user or ""
            return self.current_user or ""
        except Exception:
            return self.current_user or ""

    def _lisensi_lifetime(self) -> bool:
        """True bila lisensi efektif = LIFETIME (label LIFETIME atau expiry >= 2099).

        Dipakai untuk mengunci fitur branding TV (logo, nama rental, media promosi)
        yang hanya tersedia bagi user berlisensi LIFETIME."""
        try:
            lic = LicenseManager.load()
            return LicenseManager._effective_edition(lic) == "LIFETIME"
        except Exception:
            return False

    def _cloud_upload_target(self):
        """Target doc Firestore untuk upload transaksi: admin utama untuk kasir,
        username sendiri untuk admin."""
        return self._resolve_license_user()

    @staticmethod
    def _now_iso_utc():
        # Format sama dengan Android: yyyy-MM-dd'T'HH:mm:ss'Z' (waktu lokal, huruf Z literal)
        return datetime.now().strftime("%Y-%m-%dT%H:%M:%SZ")

    def _build_tx_cloud(self, row, meta):
        """Konversi baris riwayat lokal -> dict transaksi Firestore (kompatibel APTV2:
        billingps_users/{user}.transaksiList). Kembalikan None jika meta belum punya cloud_id."""
        try:
            if not isinstance(meta, dict):
                return None
            cloud_id = meta.get("cloud_id") or ""
            if not cloud_id:
                return None
            waktu_str = str(row[0] if len(row) > 0 else "") or ""
            if len(waktu_str) >= 16:
                waktu_iso = waktu_str[:16].replace(" ", "T") + ":00Z"
            else:
                waktu_iso = self._now_iso_utc()
            total = int(meta.get("total", 0))
            paket_harga = int(meta.get("paket_harga", 0) or 0)
            if paket_harga <= 0:
                paket_harga = max(0, total - int(meta.get("pesanan_total", 0) or 0))
            pesanan_raw = meta.get("pesanan") or {}
            pesanan = {}
            if isinstance(pesanan_raw, dict):
                for k, v in pesanan_raw.items():
                    try:
                        pesanan[str(k)] = int(v)
                    except Exception:
                        pesanan[str(k)] = 0
            all_menu = {**self.menu_makanan, **self.menu_minuman}
            pesananHarga = {}
            for nm, qty in pesanan.items():
                try:
                    pesananHarga[nm] = int(all_menu.get(nm, 0)) * qty
                except Exception:
                    pesananHarga[nm] = 0
            return {
                "id": cloud_id,
                "waktu": waktu_iso,
                "kasir": str(row[1] if len(row) > 1 else "") or (self.current_user or ""),
                "kota": str(row[2] if len(row) > 2 else "") or "",
                "paket": str(meta.get("paket_raw") or ""),
                "total": total,
                "pesanan": pesanan,
                "paketHarga": paket_harga,
                "pesananHarga": pesananHarga,
                "tvJenisPs": "TV" if meta.get("source") == "tv" else "PC",
                "paid": bool(meta.get("paid", True)),
            }
        except Exception:
            return None

    def _backfill_cloud_ids(self):
        """Berikan cloud_id stabil ke baris riwayat lama (tanpa cloud_id) agar bisa
        diupload ke Firestore tanpa duplikat. Id dihasilkan deterministik dari isi baris."""
        import hashlib
        changed = False
        for i, meta in enumerate(self.riwayat_meta):
            if not isinstance(meta, dict):
                continue
            if meta.get("cloud_id"):
                continue
            try:
                row = self.riwayat_transaksi[i]
                stable = hashlib.md5(f"{tuple(row)}".encode("utf-8")).hexdigest()[:12]
                meta["cloud_id"] = f"tx_{stable}"
                meta.setdefault("paket_raw", "")
                meta.setdefault("pesanan", {})
                changed = True
            except Exception:
                continue
        if changed:
            self._save_riwayat()

    def _flush_cloud_uploads(self):
        """Upload antrian transaksi ke billingps_users/{target}.transaksiList.
        Dipanggil dari thread background; idempoten via id transaksi."""
        try:
            if not self._pending_tx_uploads:
                return
            target = self._cloud_upload_target()
            if not target:
                self._pending_tx_uploads.clear()
                return
            pending = list(self._pending_tx_uploads)
            fc = FirestoreClient()
            ok = fc.push_transactions(target, pending)
            if ok:
                sent_ids = {t.get("id") for t in pending}
                self._pending_tx_uploads = [
                    t for t in self._pending_tx_uploads if t.get("id") not in sent_ids]
        except Exception as e:
            _LOGGER.warning("Gagal upload transaksi ke cloud: %s", e)

    def _schedule_cloud_retry(self):
        try:
            self._cloud_retry_job = self.after(30000, self._cloud_retry_tick)
        except Exception:
            pass

    def _upsert_tx_cloud_from_index(self, idx):
        """Upload ulang (replace by id) transaksi ke cloud setelah detailnya
        berubah di riwayat lokal (mis. pesanan makanan/minuman ditambahkan)."""
        try:
            if idx is None or not (0 <= idx < len(self.riwayat_transaksi)):
                return
            if idx >= len(self.riwayat_meta):
                return
            row = self.riwayat_transaksi[idx]
            meta = self.riwayat_meta[idx]
            tx = self._build_tx_cloud(row, meta)
            if not tx:
                return
            target = self._cloud_upload_target()
            if not target:
                return
            fc = FirestoreClient()
            ok = fc.upsert_transactions(target, [tx])
            if ok:
                _LOGGER.info("Transaksi %s di-update di cloud", tx.get("id"))
        except Exception as e:
            _LOGGER.warning("Gagal update transaksi cloud: %s", e)

    def _cloud_retry_tick(self):
        try:
            if self._pending_tx_uploads:
                threading.Thread(target=self._flush_cloud_uploads, daemon=True).start()
        except Exception:
            pass
        self._schedule_cloud_retry()

    # ══════════════════════════════════════════════════════════════════════════
    #  TAB 1: Dashboard
    # ══════════════════════════════════════════════════════════════════════════
    def _setup_warnet(self):
        f = self.frames["warnet"]
        hdr = ctk.CTkFrame(f, fg_color=C_PANEL, height=96, corner_radius=0)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        # Baris 1: judul + statistik (sebelah kiri)
        row1 = ctk.CTkFrame(hdr, fg_color="transparent")
        row1.pack(fill="x", pady=(10, 2))
        ctk.CTkLabel(row1, text="💻  DASHBOARD WARNET",
                     font=FONT_TITLE, text_color=C_ACCENT).pack(side="left", padx=18)
        self.lbl_total_warnet = ctk.CTkLabel(row1, text="Total Kursi: 0",
                                             font=FONT_BODY, text_color=C_MUTED)
        self.lbl_total_warnet.pack(side="left", padx=20)
        self.lbl_socket_warnet = ctk.CTkLabel(row1, text="Client: Belum tersambung",
                                               font=FONT_BODY, text_color=C_MUTED)
        self.lbl_socket_warnet.pack(side="left", padx=12)

        # Baris 2: tombol aksi — baris sendiri supaya semua terlihat (tidak
        # terpotong di kiri meski window sempit / font membesar).
        row2 = ctk.CTkFrame(hdr, fg_color="transparent")
        row2.pack(fill="x", pady=(4, 8))
        self.btn_tambah_warnet = ctk.CTkButton(row2, text="➕  Tambah Kursi", width=150, height=32,
                                             fg_color=C_ACCENT2, hover_color="#5A0FCC",
                                             font=("Russo One", 10, "bold"),
                                             command=self._buka_dialog_tambah_warnet)
        self.btn_tambah_warnet.pack(side="right", padx=8)
        if (self.current_role or "kasir") != "admin":
            self.btn_tambah_warnet.pack_forget()
        self.btn_deploy_client = ctk.CTkButton(row2, text="🚀 Deploy Client", width=145, height=32,
                                                fg_color=C_GREEN, hover_color="#2F7A2F",
                                                font=("Russo One", 10, "bold"),
                                                text_color="#000000",
                                                command=self._open_deploy_client_dialog)
        self.btn_deploy_client.pack(side="right", padx=8)
        self.btn_upload_logo = ctk.CTkButton(row2, text="🖼 Logo Lock", width=120, height=32,
                                                fg_color=C_ACCENT, hover_color="#5A0FCC",
                                                font=("Russo One", 10, "bold"),
                                                command=self._buka_upload_logo)
        self.btn_upload_logo.pack(side="right", padx=8)
        if (self.current_role or "kasir") != "admin":
            self.btn_upload_logo.pack_forget()
        self.btn_warnet_admin_code = ctk.CTkButton(
            row2,
            text="🔐 Kode Client",
            width=140,
            height=32,
            fg_color=C_BTN,
            hover_color=C_ACCENT2,
            border_width=1,
            border_color=C_ACCENT2,
            font=("Russo One", 10, "bold"),
            text_color=C_ACCENT2,
            command=self._open_warnet_admin_code_generator,
        )
        self.btn_warnet_admin_code.pack(side="right", padx=10)
        self._schedule_warnet_status_refresh()
 
        self.scroll_warnet = ctk.CTkScrollableFrame(f, fg_color=C_BG)
        self.scroll_warnet.pack(fill="both", expand=True, padx=6, pady=6)
        # column frames sebagai window items langsung di _parent_canvas
        try:
            warnet_canvas = self.scroll_warnet._parent_canvas
        except AttributeError:
            warnet_canvas = None
        self._scroll_canvas_warnet = warnet_canvas
        self._warnet_card_windows = []  # (window_id, kartu)
        self._warnet_layout_debounce_job = None
        self._warnet_last_view_pos = (0, 0)
        if warnet_canvas:
            warnet_canvas.bind("<Configure>",
                               lambda e: self._debounced_layout_warnet())
            warnet_canvas.bind("<MouseWheel>", lambda e: self.after_idle(self._update_visible_warnet_cards))
            self.after_idle(self._debounced_layout_warnet)
            self.after(500, self._schedule_warnet_visibility_check)

    def _debounced_layout_warnet(self):
        if self._warnet_layout_debounce_job:
            self.after_cancel(self._warnet_layout_debounce_job)
        self._warnet_layout_debounce_job = self.after(30, self._layout_warnet_columns)

    def _layout_warnet_columns(self):
        canvas = self._scroll_canvas_warnet
        if not canvas:
            return
        cards = self._warnet_card_windows
        n = len(cards)
        if n == 0:
            canvas.configure(scrollregion=(0,0,0,0))
            return
        total_w = max(10, canvas.winfo_width())
        num_cols = 3
        gap = 10
        col_w = max(1, (total_w - (num_cols - 1) * gap) // num_cols)
        # Pass 1: set lebar kartu
        for wid, kartu in cards:
            try: canvas.itemconfig(wid, width=col_w)
            except Exception: pass
        self.update_idletasks()
        # Pass 2: hitung scrollregion dan posisikan kartu yg terlihat
        max_x = 0
        max_y = 0
        for idx, (wid, kartu) in enumerate(cards):
            col = idx % num_cols
            row = idx // num_cols
            card_x = col * (col_w + gap)
            card_y = row * (kartu.winfo_reqheight() + gap)
            canvas.coords(wid, card_x, card_y)
            rx = card_x + col_w
            ry = card_y + kartu.winfo_reqheight()
            if rx > max_x: max_x = rx
            if ry > max_y: max_y = ry
        canvas.configure(scrollregion=(0, 0, max_x + gap, max_y + gap))
        self.after_idle(self._update_visible_warnet_cards)

    def _update_visible_warnet_cards(self):
        canvas = self._scroll_canvas_warnet
        if not canvas or not self._warnet_card_windows:
            return
        cards = self._warnet_card_windows
        n = len(cards)
        if n == 0:
            return
        total_w = max(10, canvas.winfo_width())
        num_cols = 3
        gap = 10
        col_w = max(1, (total_w - (num_cols - 1) * gap) // num_cols)
        view_top = canvas.canvasy(0)
        view_bot = view_top + canvas.winfo_height()
        margin = 100
        for idx, (wid, kartu) in enumerate(cards):
            col = idx % num_cols
            row = idx // num_cols
            card_x = col * (col_w + gap)
            card_y = row * (kartu.winfo_reqheight() + gap)
            in_view = card_y < view_bot + margin and card_y + kartu.winfo_reqheight() > view_top - margin
            if in_view:
                try: canvas.coords(wid, card_x, card_y)
                except Exception: pass
                try: canvas.itemconfig(wid, state="normal")
                except Exception: pass
            else:
                try: canvas.coords(wid, -9999, -9999)
                except Exception: pass
                try: canvas.itemconfig(wid, state="hidden")
                except Exception: pass

    def _schedule_warnet_visibility_check(self):
        if not self._scroll_canvas_warnet:
            return
        canvas = self._scroll_canvas_warnet
        cur_pos = (canvas.canvasx(0), canvas.canvasy(0))
        last_pos = getattr(self, '_warnet_last_view_pos', (0, 0))
        if abs(cur_pos[0] - last_pos[0]) > 5 or abs(cur_pos[1] - last_pos[1]) > 5:
            self._warnet_last_view_pos = cur_pos
            self._update_visible_warnet_cards()
        if self._scroll_canvas_warnet:
            self.after(300, self._schedule_warnet_visibility_check)

    def _setup_dashboard(self):
        f = self.frames["dashboard"]
        hdr = ctk.CTkFrame(f, fg_color=C_PANEL, height=54, corner_radius=0)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        ctk.CTkLabel(hdr, text="📺  DASHBOARD TV",
                     font=FONT_TITLE, text_color=C_ACCENT).pack(side="left", padx=18, pady=14)
        self.lbl_total_tv = ctk.CTkLabel(hdr, text="Total TV: 0",
                                          font=FONT_BODY, text_color=C_MUTED)
        self.lbl_total_tv.pack(side="left", padx=20)
        self.btn_install_client = ctk.CTkButton(hdr, text="📱 Install Client", width=140, height=34,
                                                fg_color=C_GREEN, hover_color="#2E7D32",
                                                font=("Russo One", 10, "bold"),
                                                text_color="#000000",
                                                command=self._open_install_apk_dialog)
        self.btn_install_client.pack(side="right", padx=5, pady=10)
        self.btn_overlay_setting = ctk.CTkButton(hdr, text="⚙ Overlay", width=110, height=34,
                                                 fg_color=C_BTN, hover_color=C_ACCENT2,
                                                 border_width=1, border_color=C_ACCENT2,
                                                 font=("Russo One", 10, "bold"),
                                                 text_color=C_ACCENT2,
                                                 command=self._open_overlay_setting)
        self.btn_overlay_setting.pack(side="right", padx=5, pady=10)
        if MATPLOTLIB_AVAILABLE:
            self.btn_revenue = ctk.CTkButton(hdr, text="📊 Revenue", width=120, height=34,
                                             fg_color=C_BTN, hover_color=C_ACCENT2,
                                             border_width=1, border_color=C_ORANGE,
                                             font=("Russo One", 10, "bold"),
                                             text_color=C_ORANGE,
                                             command=self._show_revenue_charts_dialog)
            self.btn_revenue.pack(side="right", padx=5, pady=10)
        self.btn_tambah = ctk.CTkButton(hdr, text="➕  Tambah TV", width=150, height=34,
                                         fg_color=C_ACCENT2, hover_color="#5A0FCC",
                                         font=("Russo One", 10, "bold"),
                                         command=self._buka_dialog_tambah)
        self.btn_tambah.pack(side="right", padx=8, pady=10)
        self.btn_riwayat_qr = ctk.CTkButton(hdr, text="📋  Riwayat Pesanan", width=150, height=34,
                                            fg_color=C_BTN, hover_color=C_ACCENT2,
                                            border_width=1, border_color=C_ORANGE,
                                            font=("Russo One", 10, "bold"),
                                            text_color=C_ORANGE,
                                            command=self._buka_riwayat_pesanan)
        self.btn_riwayat_qr.pack(side="right", padx=2, pady=10)
        self._stok_dash_lbl = ctk.CTkLabel(hdr, text="", font=("Russo One", 10, "bold"),
                                           text_color="#FFB300", fg_color="transparent",
                                           corner_radius=6, cursor="hand2")
        self._stok_dash_lbl.bind("<Button-1>", lambda e: self._show_tab("stok"))
        if (self.current_role or "kasir") == "admin":
            self._stok_dash_lbl.pack(side="right", padx=4, pady=10)
        if (self.current_role or "kasir") != "admin":
            self.btn_tambah.pack_forget()
            self.btn_overlay_setting.pack_forget()
        self.scroll_dash = ctk.CTkScrollableFrame(f, fg_color=C_BG)
        self.scroll_dash.pack(fill="both", expand=True, padx=6, pady=6)
        # column frames sebagai window items langsung di _parent_canvas
        try:
            dash_canvas = self.scroll_dash._parent_canvas
        except AttributeError:
            dash_canvas = None
        self._scroll_canvas_dash = dash_canvas
        self._dash_card_windows = []  # (window_id, kartu)
        # Dashboard dibuat ulang setiap login — kartu dimuat ulang dari config
        # (daftar_tv) oleh _muat_daftar_tv, jadi reset daftar & counter di sini.
        self._semua_kartu_tv = []
        self.jumlah_tv = 0
        self._dash_layout_debounce_job = None
        self._dash_last_view_pos = (0, 0)
        if dash_canvas:
            dash_canvas.bind("<Configure>", lambda e: self._debounced_layout_dash())
            dash_canvas.bind("<MouseWheel>", lambda e: self.after_idle(self._update_visible_dash_cards))
            self.after_idle(self._debounced_layout_dash)
            self.after(500, self._schedule_dash_visibility_check)

    def _debounced_layout_dash(self):
        if self._dash_layout_debounce_job:
            self.after_cancel(self._dash_layout_debounce_job)
        self._dash_layout_debounce_job = self.after(30, self._layout_dash_columns)

    def _layout_dash_columns(self):
        canvas = self._scroll_canvas_dash
        if not canvas:
            return
        cards = self._dash_card_windows
        n = len(cards)
        if n == 0:
            canvas.configure(scrollregion=(0,0,0,0))
            return
        total_w = max(10, canvas.winfo_width())
        num_cols = 3
        gap = 10
        col_w = max(1, (total_w - (num_cols - 1) * gap) // num_cols)
        # Pass 1: set lebar kartu
        for wid, kartu in cards:
            try: canvas.itemconfig(wid, width=col_w)
            except Exception: pass
        self.update_idletasks()
        # Pass 2: hitung scrollregion dan posisikan kartu yg terlihat
        max_x = 0
        max_y = 0
        for idx, (wid, kartu) in enumerate(cards):
            col = idx % num_cols
            row = idx // num_cols
            card_x = col * (col_w + gap)
            card_y = row * (kartu.winfo_reqheight() + gap)
            canvas.coords(wid, card_x, card_y)
            rx = card_x + col_w
            ry = card_y + kartu.winfo_reqheight()
            if rx > max_x: max_x = rx
            if ry > max_y: max_y = ry
        canvas.configure(scrollregion=(0, 0, max_x + gap, max_y + gap))
        self.after_idle(self._update_visible_dash_cards)

    def _update_visible_dash_cards(self):
        canvas = self._scroll_canvas_dash
        if not canvas or not self._dash_card_windows:
            return
        cards = self._dash_card_windows
        n = len(cards)
        if n == 0:
            return
        total_w = max(10, canvas.winfo_width())
        num_cols = 3
        gap = 10
        col_w = max(1, (total_w - (num_cols - 1) * gap) // num_cols)
        view_top = canvas.canvasy(0)
        view_bot = view_top + canvas.winfo_height()
        margin = 100
        for idx, (wid, kartu) in enumerate(cards):
            col = idx % num_cols
            row = idx // num_cols
            card_x = col * (col_w + gap)
            card_y = row * (kartu.winfo_reqheight() + gap)
            in_view = card_y < view_bot + margin and card_y + kartu.winfo_reqheight() > view_top - margin
            if in_view:
                try: canvas.coords(wid, card_x, card_y)
                except Exception: pass
                try: canvas.itemconfig(wid, state="normal")
                except Exception: pass
            else:
                try: canvas.coords(wid, -9999, -9999)
                except Exception: pass
                try: canvas.itemconfig(wid, state="hidden")
                except Exception: pass

    def _schedule_dash_visibility_check(self):
        if not self._scroll_canvas_dash:
            return
        canvas = self._scroll_canvas_dash
        cur_pos = (canvas.canvasx(0), canvas.canvasy(0))
        last_pos = getattr(self, '_dash_last_view_pos', (0, 0))
        if abs(cur_pos[0] - last_pos[0]) > 5 or abs(cur_pos[1] - last_pos[1]) > 5:
            self._dash_last_view_pos = cur_pos
            self._update_visible_dash_cards()
        if self._scroll_canvas_dash:
            self.after(300, self._schedule_dash_visibility_check)

    def _unlock_tambah(self):
        self._tambah_btn_enabled = True
        self.btn_tambah.configure(state="normal", text="➕  Tambah TV")

    def _unlock_tambah_warnet(self):
        self._tambah_warnet_btn_enabled = True
        self.btn_tambah_warnet.configure(state="normal", text="➕  Tambah Kursi")

    def _get_promo_add_tv_from_current(self) -> int:
        promo = getattr(self, "_promo_data", None)
        if not promo or not promo.get("promoAktif"):
            lic = LicenseManager.load()
            return lic.get("promo_add_tv", 0)
        add_map = promo.get("addTvOverride", {})
        if not add_map:
            lic = LicenseManager.load()
            return lic.get("promo_add_tv", 0)
        lic = LicenseManager.load()
        edition = lic.get("edition", "")
        EDITION_TO_KEY = {"BULANAN": "1 Bulan", "3BULAN": "3 Bulan", "TAHUNAN": "1 Tahun", "LIFETIME": "LIFETIME"}
        key = EDITION_TO_KEY.get(edition)
        if not key or key not in add_map:
            return lic.get("promo_add_tv", 0)
        return add_map[key]

    def _save_promo_override_to_license(self):
        promo = getattr(self, "_promo_data", None)
        if not promo or not promo.get("promoAktif"):
            return
        add_map = promo.get("addTvOverride", {})
        if not add_map:
            return
        lic = LicenseManager.load()
        if not lic.get("aktif"):
            return
        edition = lic.get("edition", "")
        EDITION_TO_KEY = {"BULANAN": "1 Bulan", "3BULAN": "3 Bulan", "TAHUNAN": "1 Tahun", "LIFETIME": "LIFETIME"}
        key = EDITION_TO_KEY.get(edition)
        if not key or key not in add_map:
            return
        new_val = add_map[key]
        existing = lic.get("promo_add_tv", 0)
        if new_val > existing:
            lic["promo_add_tv"] = new_val
            lic["promo_add_warnet"] = new_val
            LicenseManager.save(lic)

    def _promo_override_tv_limit(self, base_limit: int, is_warnet: bool = False) -> int:
        lic = LicenseManager.load()
        if lic.get("aktif") and lic.get("promo_add_tv"):
            override = lic["promo_add_tv"]
            if is_warnet and lic.get("promo_add_warnet"):
                override = lic["promo_add_warnet"]
            return max(override, base_limit)
        promo = getattr(self, "_promo_data", None)
        if not promo or not promo.get("promoAktif"):
            return base_limit
        add_map = promo.get("addTvOverride", {})
        if not add_map:
            return base_limit
        status = LicenseManager.get_status(current_user=self._resolve_license_user())
        edition = status.get("edition", "")
        EDITION_TO_KEY = {"BULANAN": "1 Bulan", "3BULAN": "3 Bulan", "TAHUNAN": "1 Tahun", "LIFETIME": "LIFETIME"}
        key = EDITION_TO_KEY.get(edition)
        if not key or key not in add_map:
            return base_limit
        override = add_map[key]
        return max(override, base_limit)

    def _buka_dialog_tambah(self):
        if not self._tambah_btn_enabled: return
        self._tambah_btn_enabled = False
        self.btn_tambah.configure(state="disabled", text="⏳ Menunggu...")
        DialogTambahTV(self, self.jumlah_tv + 1,
                       on_confirm=self._on_tv_confirmed,
                       on_close_cb=self._unlock_tambah,
                       daftar_grup=self.daftar_nama_grup())

    def _open_overlay_setting(self):
        DialogOverlaySetting(self,
                             mode=getattr(self, "tv_overlay_mode", "always"),
                             last_minutes=getattr(self, "tv_overlay_last_minutes", 5),
                             on_save=self._save_overlay_setting)

    def _save_overlay_setting(self, mode, minutes):
        self.tv_overlay_mode = mode
        self.tv_overlay_last_minutes = minutes
        ConfigManager.set("tv_overlay_mode", mode)
        ConfigManager.set("tv_overlay_last_minutes", minutes)
        messagebox.showinfo("✅ Pengaturan Disimpan",
                            f"Mode overlay: {mode}\n"
                            f"Tampil sejak sisa waktu {minutes} menit.",
                            parent=self)
 
    def _open_install_apk_dialog(self):
        DialogInstallAPK(self, ConfigManager)

    def _show_revenue_charts_dialog(self):
        if not MATPLOTLIB_AVAILABLE:
            messagebox.showerror("Error", "matplotlib tidak terinstall.\nInstall dengan: pip install matplotlib")
            return
        if not self.riwayat_transaksi:
            messagebox.showinfo("Info", "Belum ada data transaksi.")
            return

        dialog = ctk.CTkToplevel(self)
        dialog.title("📊 Revenue Charts")
        dialog.geometry("800x700")
        dialog.resizable(False, False)
        dialog.configure(fg_color=C_BG)
        dialog.transient(self)
        dialog.grab_set()

        container = ctk.CTkFrame(dialog, fg_color=C_BG)
        container.pack(fill="both", expand=True, padx=12, pady=12)

        top = ctk.CTkFrame(container, fg_color=C_PANEL, corner_radius=8)
        top.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(top, text="📊  Revenue Overview",
                     font=FONT_TITLE, text_color=C_ACCENT).pack(side="left", padx=14, pady=10)

        ctrl = ctk.CTkFrame(container, fg_color=C_PANEL, corner_radius=8)
        ctrl.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(ctrl, text="Periode:", font=FONT_BODY, text_color=C_MUTED).pack(side="left", padx=(14, 4), pady=10)

        period_var = ctk.StringVar(value="Hari Ini")
        period_menu = ctk.CTkOptionMenu(ctrl, variable=period_var,
                                        values=["Hari Ini", "7 Hari", "30 Hari", "Tahun Ini"],
                                        fg_color=C_BTN, button_color=C_ACCENT2,
                                        text_color=C_TEXT, font=FONT_BODY)
        period_menu.pack(side="left", padx=4, pady=10)

        chart_frame = ctk.CTkFrame(container, fg_color=C_PANEL, corner_radius=8)
        chart_frame.pack(fill="both", expand=True)

        def refresh_charts(*_):
            for w in chart_frame.winfo_children():
                w.destroy()

            now = datetime.now()
            period = period_var.get()
            if period == "Hari Ini":
                start = now.replace(hour=0, minute=0, second=0, microsecond=0)
            elif period == "7 Hari":
                start = now - timedelta(days=7)
            elif period == "30 Hari":
                start = now - timedelta(days=30)
            else:
                start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)

            filtered = []
            for i, row in enumerate(self.riwayat_transaksi):
                try:
                    t = datetime.strptime(row[0], "%Y-%m-%d %H:%M")
                except Exception:
                    continue
                if t >= start and i < len(self.riwayat_meta):
                    filtered.append((t, row, self.riwayat_meta[i]))

            if not filtered:
                ctk.CTkLabel(chart_frame, text="Belum ada data untuk periode ini.",
                             font=FONT_BODY, text_color=C_MUTED).pack(expand=True)
                return

            fig = Figure(figsize=(10, 6), dpi=90, facecolor="#1a1a2e")
            fig.subplots_adjust(left=0.1, right=0.95, top=0.92, bottom=0.15, hspace=0.35)

            daily = {}
            for t, row, meta in filtered:
                day_key = t.strftime("%Y-%m-%d")
                daily.setdefault(day_key, {"tv": 0, "warnet": 0, "food": 0, "total": 0})
                src = meta.get("source", "tv")
                paket_h = meta.get("paket_harga", 0)
                pesanan = meta.get("pesanan_total", 0)
                total = meta.get("total", 0)
                if src == "tv":
                    daily[day_key]["tv"] += paket_h
                else:
                    daily[day_key]["warnet"] += paket_h
                daily[day_key]["food"] += pesanan
                daily[day_key]["total"] += total

            sorted_days = sorted(daily.keys())
            tv_vals = [daily[d]["tv"] for d in sorted_days]
            warnet_vals = [daily[d]["warnet"] for d in sorted_days]
            food_vals = [daily[d]["food"] for d in sorted_days]

            ax1 = fig.add_subplot(211)
            x = range(len(sorted_days))
            bar_w = 0.25
            ax1.bar([i - bar_w for i in x], tv_vals, width=bar_w, label="TV Paket", color="#7c3aed", alpha=0.85)
            ax1.bar(x, warnet_vals, width=bar_w, label="Warnet Paket", color="#f59e0b", alpha=0.85)
            ax1.bar([i + bar_w for i in x], food_vals, width=bar_w, label="Makanan/Minuman", color="#10b981", alpha=0.85)
            ax1.set_xticks(list(x))
            ax1.set_xticklabels(sorted_days, rotation=30, ha="right", fontsize=7, color="#ccc")
            ax1.set_ylabel("Revenue (Rp)", fontsize=9, color="#ccc")
            ax1.legend(fontsize=7, loc="upper right", facecolor="#16213e", labelcolor="#ccc")
            ax1.set_title("Daily Revenue Breakdown", fontsize=11, color="#fff", fontweight="bold")
            ax1.tick_params(colors="#ccc", labelsize=7)
            ax1.set_facecolor("#16213e")
            for spine in ax1.spines.values():
                spine.set_color("#2a2a4a")

            total_tv = sum(tv_vals)
            total_warnet = sum(warnet_vals)
            total_food = sum(food_vals)
            pie_labels = []
            pie_sizes = []
            pie_colors = []
            if total_tv > 0:
                pie_labels.append("TV Paket"); pie_sizes.append(total_tv); pie_colors.append("#7c3aed")
            if total_warnet > 0:
                pie_labels.append("Warnet Paket"); pie_sizes.append(total_warnet); pie_colors.append("#f59e0b")
            if total_food > 0:
                pie_labels.append("Makanan/Minuman"); pie_sizes.append(total_food); pie_colors.append("#10b981")

            ax2 = fig.add_subplot(212)
            if pie_sizes:
                wedges, texts, autotexts = ax2.pie(
                    pie_sizes, labels=pie_labels, autopct="%1.1f%%",
                    colors=pie_colors, startangle=90,
                    textprops={"color": "#ccc", "fontsize": 8})
                for at in autotexts:
                    at.set_color("#fff")
                    at.set_fontsize(8)
            ax2.set_title("Revenue by Source", fontsize=11, color="#fff", fontweight="bold")
            ax2.set_facecolor("#1a1a2e")

            canvas = FigureCanvasTkAgg(fig, master=chart_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True)

        period_var.trace_add("write", refresh_charts)
        refresh_charts()

    def _on_tv_confirmed(self, ip, nama, port, nama_grup):
        self._unlock_tambah()
        self._tambah_tv(ip, nama, port, nama_grup)

    def _simpan_daftar_tv(self):
        """Simpan daftar kartu TV (ip/nama/port/grup) ke config (key daftar_tv)
        agar otomatis dimuat ulang saat login berikutnya."""
        try:
            daftar = []
            for kartu in list(getattr(self, "_semua_kartu_tv", [])):
                daftar.append({
                    "ip": getattr(kartu, "ip", ""),
                    "nama": getattr(kartu, "label_tv", ""),
                    "port": getattr(kartu, "port", 0),
                    "nama_grup": getattr(kartu, "nama_grup", NAMA_GRUP_DEFAULT),
                })
            try:
                ConfigManager.update(lambda cfg: cfg.__setitem__("daftar_tv", daftar) or cfg)
            except Exception as e:
                print(f"[TV] Gagal simpan daftar_tv: {e}", flush=True)
        except Exception as e:
            print(f"[TV] Gagal simpan daftar_tv: {e}", flush=True)
        self._jadwal_sync_tv_list()

    def _tv_list_cloud(self):
        """Daftar TV sesuai kontrak APK TV (field id/nama/ip/...).
        APK membaca billingps_users/{username}.tvList saat menghubungkan TV."""
        daftar = []
        for kartu in list(getattr(self, "_semua_kartu_tv", [])):
            daftar.append({
                "id": getattr(kartu, "label_tv", ""),
                "nama": getattr(kartu, "label_tv", ""),
                "ip": getattr(kartu, "ip", ""),
                "port": int(getattr(kartu, "port", 0) or 0),
                "nama_grup": getattr(kartu, "nama_grup", ""),
            })
        return daftar

    def _sync_tv_list_cloud(self):
        """Sinkronkan tvList ke Firestore agar APK TV bisa terhubung/memilih TV."""
        uname = getattr(self, "current_user", None) or ""
        if not uname:
            return
        try:
            from firestore_sync import get_firestore_client
            daftar = self._tv_list_cloud()
            ok = get_firestore_client().sync_tv_list(uname, daftar)
            print(f"[TV] tvList cloud tersinkron ({len(daftar)} TV) — {ok}", flush=True)
        except Exception as e:
            print(f"[TV] Sync tvList error: {e}", flush=True)

    def _jadwal_sync_tv_list(self):
        """Coalesce: 1,5 detik setelah perubahan terakhir, sync ke cloud."""
        try:
            if getattr(self, "_tv_sync_timer", None):
                self.after_cancel(self._tv_sync_timer)
        except Exception:
            pass
        try:
            self._tv_sync_timer = self.after(1500, self._flush_sync_tv_list)
        except Exception:
            pass

    def _flush_sync_tv_list(self):
        self._tv_sync_timer = None
        threading.Thread(target=self._sync_tv_list_cloud, daemon=True).start()

    def _muat_daftar_tv(self):
        """Buat ulang kartu TV dari config (daftar_tv) setelah login.

        Berlaku untuk admin maupun kasir (daftar disimpan admin); tanpa dialog,
        tanpa pemberitahuan admin — lewati diam-diam jika limit tercapai."""
        try:
            cfg = ConfigManager.load()
            daftar = cfg.get("daftar_tv", []) or []
            for item in daftar:
                try:
                    ip = str(item.get("ip", "")).strip()
                    nama = str(item.get("nama", "")).strip() or f"TV {self.jumlah_tv + 1}"
                    try:
                        port = int(item.get("port", 0) or 0)
                    except Exception:
                        port = 0
                    nama_grup = str(item.get("nama_grup", "")).strip() or NAMA_GRUP_DEFAULT
                    if not ip:
                        continue
                    if any(k.ip == ip and k.label_tv == nama
                           for k in getattr(self, "_semua_kartu_tv", [])):
                        continue
                    self._tambah_tv(ip, nama, port, nama_grup, restore=True)
                except Exception as e:
                    print(f"[TV] Gagal memuat kartu {item}: {e}", flush=True)
        except Exception as e:
            print(f"[TV] Gagal muat daftar_tv: {e}", flush=True)

    def _ui_call(self, cb, *args, **kwargs):
        """Jalankan callback di main thread Tk secara aman dari thread lain.
        Hanya after(0) + guard; callback harus sudah try/except sendiri."""
        try:
            self.after(0, lambda: cb(*args, **kwargs))
        except Exception:
            pass

    def _auto_restore_daftar_tv(self):
        """Pulihkan otomatis kartu TV yang hilang dari file sertifikat pairing
        (cert_<ip>.pem) + nama terakhir dari audit log. Dijalankan sekali per
        sesi saat login admin, hanya menambah kartu yang belum ada — tidak
        pernah menimpa daftar yang sudah tersimpan (biar bug limit lama tidak
        mengulang: daftar_tv tidak boleh terpotong oleh restore)."""
        if (self.current_role or "") != "admin":
            return
        try:
            if getattr(self, "_auto_restored_tv", False):
                return
            self._auto_restored_tv = True

            known = set()
            for k in getattr(self, "_semua_kartu_tv", []):
                known.add(str(getattr(k, "ip", "")).strip())

            ips = set()
            base_dirs = [APP_BASE_DIR, os.path.join(APP_BASE_DIR, "android_tv_certs")]
            for d in base_dirs:
                try:
                    if not os.path.isdir(d):
                        continue
                    for fn in os.listdir(d):
                        if fn.startswith("cert_") and fn.endswith(".pem"):
                            ip = fn[5:-4].replace("_", ".")
                            octets = ip.split(".")
                            if (len(octets) == 4
                                    and all(o.isdigit() and 0 <= int(o) <= 255 for o in octets)):
                                ips.add(ip)
                except OSError:
                    continue

            if not (ips - known):
                return

            nama_map = {}
            if os.path.exists(AUDIT_FILE):
                try:
                    with open(AUDIT_FILE, "r", encoding="utf-8") as f:
                        for line in f:
                            try:
                                rec = json.loads(line)
                                det = rec.get("details") or {}
                                lip, ltv = det.get("ip"), det.get("label_tv")
                                if rec.get("action") == "TV_AUTO_OFF" and lip and ltv:
                                    nama_map[str(lip)] = str(ltv)
                            except Exception:
                                continue
                except OSError:
                    pass

            candidates = sorted(ips - known,
                                key=lambda ip: [int(x) for x in ip.split(".")])
            tv_limit, _ = get_edition_limits(current_user=self._resolve_license_user())
            tv_limit = self._promo_override_tv_limit(tv_limit)
            current = len(getattr(self, "_semua_kartu_tv", []))
            room = max(0, tv_limit - current)
            if room <= 0:
                print(f"[TV] Auto-restore dilewati: {len(candidates)} kartu baru — limit {tv_limit} tercapai.", flush=True)
                return

            allowed = candidates[:room]

            # Pembuatan kartu harus di main thread (widget Tk). Scan sudah di
            # thread; UI disatukan lewat satu after(0) agar tidak blokir login.
            def _build_cards():
                try:
                    current = len(getattr(self, "_semua_kartu_tv", []))
                    for ip in allowed:
                        try:
                            nama = nama_map.get(ip) or f"TV {self.jumlah_tv + 1}"
                            self._tambah_tv(ip, nama, 0, None, restore=True)
                            print(f"[TV] Auto-restore: {ip} -> {nama}", flush=True)
                        except Exception as e:
                            print(f"[TV] Auto-restore gagal untuk {ip}: {e}", flush=True)
                    if len(getattr(self, "_semua_kartu_tv", [])) - current == len(allowed):
                        self._simpan_daftar_tv()
                    else:
                        print("[TV] Auto-restore sebagian: config tidak ditimpa agar kartu lain tak hilang.", flush=True)
                except Exception as e:
                    print(f"[TV] Auto-restore UI error: {e}", flush=True)

            self._ui_call(_build_cards)
        except Exception as e:
            print(f"[TV] Gagal auto-restore daftar_tv: {e}", flush=True)

    def _simpan_daftar_warnet(self):
        """Simpan daftar kursi warnet ke config ('daftar_warnet') supaya bisa
        dibangun ulang setelah login berikutnya (mirip daftar_tv)."""
        try:
            cfg = ConfigManager.load()
            daftar = []
            for kartu in getattr(self, '_semua_kartu_warnet', []):
                daftar.append({
                    "nama": getattr(kartu, 'label_kursi', ''),
                    "nama_grup": getattr(kartu, 'nama_grup', NAMA_GRUP_DEFAULT),
                    "client_id": getattr(kartu, '_client_id', None),
                    "pc_id": getattr(kartu, '_pc_id', None),
                    "pc_ip": getattr(kartu, '_pc_ip', None),
                    "pc_locked": bool(getattr(kartu, 'pc_locked', False)),
                    "pc_lock_reason": getattr(kartu, '_pc_lock_reason', '') or '',
                    "pc_lock_message": getattr(kartu, '_pc_lock_message', '') or '',
                })
            cfg["daftar_warnet"] = daftar
            ConfigManager.save(cfg)
        except Exception as e:
            print(f"[WARN] Gagal simpan daftar_warnet: {e}", flush=True)

    def _muat_daftar_warnet(self):
        """Buat ulang kartu warnet dari config (daftar_warnet) setelah login.

        Berlaku untuk admin maupun kasir; tanpa dialog, tanpa pemberitahuan —
        lewati diam-diam jika limit tercapai. Sesi aktif di-restore terpisah
        oleh TimerService.restore_timer_state (key = pc_id)."""
        try:
            cfg = ConfigManager.load()
            daftar = cfg.get("daftar_warnet", []) or []
            if not daftar:
                return
            if not getattr(self, '_scroll_canvas_warnet', None):
                return
            if not hasattr(self, '_semua_kartu_warnet'):
                self._semua_kartu_warnet = []
            if not hasattr(self, '_warnet_card_windows'):
                self._warnet_card_windows = []
            for item in daftar:
                try:
                    nama = str(item.get("nama", "")).strip()
                    if not nama:
                        continue
                    nama_grup = str(item.get("nama_grup", "")).strip() or NAMA_GRUP_DEFAULT
                    pc_id = item.get("pc_id")
                    if any(k.label_kursi == nama and getattr(k, '_pc_id', None) == pc_id
                           for k in self._semua_kartu_warnet):
                        continue
                    pc_info = None
                    if pc_id:
                        pc_info = {
                            "client_id": item.get("client_id"),
                            "pc_id": pc_id,
                            "ip": item.get("pc_ip"),
                        }
                    kartu = self._tambah_warnet(nama, nama_grup=nama_grup, pc_info=pc_info, restore=True)
                    if kartu is not None and item.get("pc_locked"):
                        # Restore state lock — LOCK akan di-requeue oleh GET_STATUS
                        # bila tidak ada sesi aktif (recovery lintas restart server).
                        kartu.pc_locked = True
                        kartu._pc_lock_reason = str(item.get("pc_lock_reason") or "") or "waktu_habis"
                        kartu._pc_lock_message = str(item.get("pc_lock_message") or "") or f"PC {nama} terkunci."
                except Exception as e:
                    print(f"[WARN] Gagal memuat kartu warnet {item}: {e}", flush=True)
        except Exception as e:
            print(f"[WARN] Gagal muat daftar_warnet: {e}", flush=True)

    def _tambah_tv(self, ip, nama, port, nama_grup=None, restore=False):
        if (self.current_role or "kasir") != "admin" and not restore:
            messagebox.showwarning("⚠ AKSES TERBATAS", "Hanya admin yang dapat menambah TV.")
            return
        tv_limit, _ = get_edition_limits(current_user=self._resolve_license_user())
        tv_limit = self._promo_override_tv_limit(tv_limit)
        if len(self._semua_kartu_tv) >= tv_limit:
            if restore:
                print(f"[TV] Restore dilewati '{nama}' ({ip}) — limit {tv_limit} TV tercapai", flush=True)
                return
            messagebox.showwarning("Limit Tercapai",
                f"Paket Anda hanya mengizinkan maksimal {tv_limit} TV.\n"
                f"Hapus beberapa TV atau upgrade paket.")
            return
        nama_grup = nama_grup or NAMA_GRUP_DEFAULT
        self.jumlah_tv += 1
        canvas = self._scroll_canvas_dash
        kartu = KartuTV(canvas, self.jumlah_tv,
                        ip=ip, port=port, label_tv=nama,
                        on_transaksi=self._catat_transaksi,
                        get_paket_data=lambda g=nama_grup: self.get_paket_data(g),
                        get_makanan_data=self.get_makanan_data,
                        get_minuman_data=self.get_minuman_data,
                        get_semua_kartu=self.get_semua_kartu,
                        get_daftar_grup=self.daftar_nama_grup,
                        on_ganti_grup=self._on_kartu_ganti_grup,
                        on_hapus=self._hapus_tv,
                        nama_grup=nama_grup,
                        is_first=(self.jumlah_tv == 1),
                        role=self.current_role)
        wid = canvas.create_window(0, 0, window=kartu, anchor="nw", tags=("_dash_card",))
        self._dash_card_windows.append((wid, kartu))
        self._semua_kartu_tv.append(kartu)
        self.lbl_total_tv.configure(text=f"Total TV: {self.jumlah_tv}")
        self._refresh_dashboard_total_pesanan()
        if not restore:
            self._simpan_daftar_tv()
        self.after_idle(self._debounced_layout_dash)

    def _bersihkan_konfig_tv(self, kartu):
        """Hapus seluruh jejak konfigurasi TV yang dihapus:
        - qr_call (kode QR tidak berlaku lagi)
        - timer_state (entry timer per nomor kartu)
        - sertifikat pairing cert_<ip>.pem / key_<ip>.pem
        Tanpa hapus sertifikat, auto-restore saat login berikutnya akan
        memunculkan lagi TV yang sudah dihapus."""
        ip = str(getattr(kartu, "ip", "")).strip()
        nama = str(getattr(kartu, "label_tv", "")).strip()
        nomor = str(getattr(kartu, "nomor", "")).strip()

        def _mut(c):
            qr = c.get("qr_call") or {}
            if isinstance(qr, dict) and nama and nama in qr:
                del qr[nama]
                c["qr_call"] = qr
            ts = c.get("timer_state") or {}
            if isinstance(ts, dict):
                tv_state = ts.get("tv")
                if isinstance(tv_state, dict) and nomor and nomor in tv_state:
                    del tv_state[nomor]
            return c

        try:
            ConfigManager.update(_mut)
            print(f"[TV] Konfigurasi '{nama}' dibersihkan (qr_call/timer_state).")
        except Exception as e:
            print(f"[TV] Bersihkan konfigurasi gagal: {e}")

        n_dihapus = 0
        if ip:
            try:
                from tv_mesin import _cert_paths_for_ip
                for p in _cert_paths_for_ip(ip):
                    try:
                        if os.path.exists(p):
                            os.remove(p)
                            n_dihapus += 1
                    except Exception:
                        pass
            except Exception as e:
                print(f"[TV] Gagal hapus sertifikat {ip}: {e}")
        if n_dihapus:
            print(f"[TV] Sertifikat pairing {ip} dihapus ({n_dihapus} file) — "
                  "TV tidak akan muncul lagi otomatis.")
    
    # ── PIN KEAMANAN (HAPUS TV & KURSI WANET) ──────────────────────────
    def _get_pin_hapus(self):
        """PIN tersimpan di config (lokal). Admin dapat melihatnya lagi via Profil."""
        cfg = ConfigManager.load()
        sec = cfg.get("security") or {}
        return str(sec.get("pin_hapus") or "").strip()

    def _set_pin_hapus(self, pin):
        cfg = ConfigManager.load()
        sec = dict(cfg.get("security") or {})
        if pin:
            sec["pin_hapus"] = str(pin)
            sec["pin_hapus_updated"] = datetime.now().isoformat(timespec="seconds")
        else:
            sec.pop("pin_hapus", None)
            sec.pop("pin_hapus_updated", None)
        cfg["security"] = sec
        ConfigManager.save(cfg)

    def _hapus_pin_hapus(self):
        if not self._get_pin_hapus():
            messagebox.showinfo("PIN Keamanan", "Belum ada PIN yang tersimpan.", parent=self)
            return
        if not messagebox.askyesno("Hapus PIN",
                                   "Hapus PIN Keamanan?\n\nHapus TV/kursi akan terkunci sampai PIN dibuat lagi.",
                                   parent=self):
            return
        self._set_pin_hapus("")
        messagebox.showinfo("✅ Selesai", "PIN Keamanan telah dihapus.", parent=self)
        self.after(200, self._setup_profil)

    def _dialog_set_pin_hapus(self):
        dlg = ctk.CTkToplevel(self.winfo_toplevel())
        dlg.title("🔑 PIN Keamanan — Hapus TV & Kursi")
        dlg.configure(fg_color=C_BG)
        dlg.transient(self.winfo_toplevel())
        dlg.resizable(False, False)
        dlg.grab_set()

        ctk.CTkLabel(dlg, text="Buat PIN (4–6 digit)", font=("Russo One", 13, "bold"),
                     text_color=C_ACCENT).pack(pady=(14, 2))
        ctk.CTkLabel(dlg, text="PIN wajib dimasukkan saat menghapus TV / kursi warnet.\n"
                               "Jika lupa, PIN bisa dilihat lagi lewat tombol 👁 Lihat PIN.",
                     font=FONT_SMALL, text_color=C_MUTED, wraplength=320).pack(padx=24)

        frm = ctk.CTkFrame(dlg, fg_color=C_PANEL, corner_radius=10)
        frm.pack(padx=24, pady=(10, 6))
        ctk.CTkLabel(frm, text="PIN Baru:", font=FONT_LABEL, text_color=C_MUTED).grid(
            row=0, column=0, padx=10, pady=8, sticky="w")
        ent_pin = ctk.CTkEntry(frm, show="●", width=180, font=("Consolas", 14, "bold"),
                               fg_color=C_BTN, text_color=C_ACCENT, border_color=C_BORDER)
        ent_pin.grid(row=0, column=1, padx=10, pady=8)
        ctk.CTkLabel(frm, text="Ulangi PIN:", font=FONT_LABEL, text_color=C_MUTED).grid(
            row=1, column=0, padx=10, pady=8, sticky="w")
        ent_pin2 = ctk.CTkEntry(frm, show="●", width=180, font=("Consolas", 14, "bold"),
                                fg_color=C_BTN, text_color=C_ACCENT, border_color=C_BORDER)
        ent_pin2.grid(row=1, column=1, padx=10, pady=8)
        show_var = tk.BooleanVar(value=False)

        def _toggle_lihat():
            show = show_var.get()
            ent_pin.configure(show="" if show else "●")
            ent_pin2.configure(show="" if show else "●")

        ctk.CTkCheckBox(frm, text="👁 Lihat PIN", variable=show_var, command=_toggle_lihat,
                        font=FONT_SMALL, text_color=C_MUTED).grid(
            row=2, column=1, padx=10, pady=(0, 10), sticky="w")
        lbl_err = ctk.CTkLabel(dlg, text="", font=FONT_SMALL, text_color=C_RED)
        lbl_err.pack()

        def _simpan():
            p1 = ent_pin.get().strip()
            p2 = ent_pin2.get().strip()
            if not (p1.isdigit() and 4 <= len(p1) <= 6):
                lbl_err.configure(text="PIN harus 4–6 digit angka.")
                return
            if p1 != p2:
                lbl_err.configure(text="PIN tidak sama.")
                return
            self._set_pin_hapus(p1)
            dlg.destroy()
            messagebox.showinfo("✅ Tersimpan",
                                "PIN Keamanan berhasil disimpan.\n\n"
                                "Simpan baik-baik — jika lupa, bisa dilihat lagi di Profil → 👁 Lihat PIN.",
                                parent=self)
            self.after(200, self._setup_profil)

        row = ctk.CTkFrame(dlg, fg_color="transparent")
        row.pack(pady=(4, 14))
        ctk.CTkButton(row, text="💾 Simpan", width=110, height=34, fg_color=C_ACCENT2,
                      font=("Russo One", 10, "bold"), command=_simpan).pack(side="left", padx=6)
        ctk.CTkButton(row, text="Batal", width=90, height=34, fg_color=C_BTN,
                      hover_color=C_RED, font=("Russo One", 10, "bold"),
                      command=dlg.destroy).pack(side="left", padx=6)
        dlg.bind("<Return>", lambda e: _simpan())
        ent_pin.focus_set()

    def _dialog_lihat_pin_hapus(self):
        pin = self._get_pin_hapus()
        if not pin:
            messagebox.showinfo("PIN Keamanan", "Belum ada PIN yang tersimpan.", parent=self)
            return
        dlg = ctk.CTkToplevel(self.winfo_toplevel())
        dlg.title("👁 Lihat PIN Keamanan")
        dlg.configure(fg_color=C_BG)
        dlg.transient(self.winfo_toplevel())
        dlg.resizable(False, False)
        dlg.grab_set()
        ctk.CTkLabel(dlg, text="PIN Keamanan Anda:", font=("Russo One", 12, "bold"),
                     text_color=C_ACCENT).pack(pady=(16, 6))
        ctk.CTkLabel(dlg, text=pin, font=("Consolas", 30, "bold"), text_color=C_YELLOW).pack(pady=(0, 6))
        ctk.CTkLabel(dlg, text="Dipakai saat menghapus TV / kursi warnet.",
                     font=FONT_SMALL, text_color=C_MUTED).pack(padx=24, pady=(0, 4))
        ctk.CTkLabel(dlg, text="🔒 Jangan bagikan PIN ini ke kasir.",
                     font=FONT_SMALL, text_color=C_GREEN).pack(pady=(0, 8))

        def _salin():
            dlg.clipboard_clear()
            dlg.clipboard_append(pin)
            messagebox.showinfo("📋 Disalin", "PIN telah disalin ke clipboard.", parent=dlg)

        row = ctk.CTkFrame(dlg, fg_color="transparent")
        row.pack(pady=(4, 14))
        ctk.CTkButton(row, text="📋 Salin", width=100, height=32, fg_color=C_ACCENT2,
                      font=("Russo One", 9, "bold"), command=_salin).pack(side="left", padx=6)
        ctk.CTkButton(row, text="Tutup", width=90, height=32, fg_color=C_BTN,
                      hover_color=C_ACCENT2, font=("Russo One", 9, "bold"),
                      command=dlg.destroy).pack(side="left", padx=6)

    def _minta_pin_hapus(self, on_ok):
        """Dialog PIN wajib sebelum hapus TV/kursi. on_ok() dipanggil jika PIN benar."""
        pin = self._get_pin_hapus()
        if not pin:
            messagebox.showwarning("🔒 PIN BELUM DIBUAT",
                                   "Admin belum membuat PIN Keamanan.\n\n"
                                   "Buat PIN di menu: Profil → PIN Keamanan → 🔑 Buat PIN.\n"
                                   "Hapus TV/kursi terkunci sampai PIN dibuat.",
                                   parent=self)
            return
        now = time.time()
        lock_until = getattr(self, "_pin_lock_until", 0.0)
        if now < lock_until:
            sisa = int(lock_until - now) + 1
            messagebox.showwarning("🔒 PIN TERKUNCI",
                                   f"Terlalu banyak PIN salah.\nCoba lagi dalam {sisa} detik.",
                                   parent=self)
            return

        dlg = ctk.CTkToplevel(self.winfo_toplevel())
        dlg.title("🔒 Masukkan PIN")
        dlg.configure(fg_color=C_BG)
        dlg.transient(self.winfo_toplevel())
        dlg.resizable(False, False)
        dlg.grab_set()
        ctk.CTkLabel(dlg, text="🔒 Konfirmasi PIN Keamanan", font=("Russo One", 13, "bold"),
                     text_color=C_ACCENT).pack(pady=(16, 2))
        ctk.CTkLabel(dlg, text="Masukkan PIN untuk melanjutkan penghapusan.",
                     font=FONT_SMALL, text_color=C_MUTED).pack(padx=24)
        frm = ctk.CTkFrame(dlg, fg_color=C_PANEL, corner_radius=10)
        frm.pack(padx=24, pady=(10, 6))
        ent = ctk.CTkEntry(frm, show="●", width=190, font=("Consolas", 16, "bold"),
                           fg_color=C_BTN, text_color=C_ACCENT, border_color=C_ACCENT)
        ent.pack(padx=12, pady=12)
        show_var = tk.BooleanVar(value=False)

        def _toggle_lihat():
            ent.configure(show="" if show_var.get() else "●")

        ctk.CTkCheckBox(frm, text="👁 Lihat PIN", variable=show_var, command=_toggle_lihat,
                        font=FONT_SMALL, text_color=C_MUTED).pack(pady=(0, 10))
        lbl_err = ctk.CTkLabel(dlg, text="", font=FONT_SMALL, text_color=C_RED)
        lbl_err.pack()
        state = {"salah": 0}

        def _cek():
            if ent.get().strip() == pin:
                dlg.destroy()
                on_ok()
                return
            state["salah"] += 1
            if state["salah"] >= 5:
                self._pin_lock_until = time.time() + 60
                dlg.destroy()
                messagebox.showwarning("🔒 PIN TERKUNCI",
                                       "Terlalu banyak PIN salah (5x).\n"
                                       "Fitur hapus TV/kursi dikunci selama 1 menit.",
                                       parent=self)
                return
            lbl_err.configure(text=f"PIN salah ({state['salah']}/5). Sisa {5 - state['salah']}x percobaan.")
            ent.delete(0, "end")
            ent.focus_set()

        row = ctk.CTkFrame(dlg, fg_color="transparent")
        row.pack(pady=(4, 14))
        ctk.CTkButton(row, text="✅ OK", width=100, height=34, fg_color=C_ACCENT2,
                      font=("Russo One", 10, "bold"), command=_cek).pack(side="left", padx=6)
        ctk.CTkButton(row, text="Batal", width=90, height=34, fg_color=C_BTN,
                      hover_color=C_RED, font=("Russo One", 10, "bold"),
                      command=dlg.destroy).pack(side="left", padx=6)
        dlg.bind("<Return>", lambda e: _cek())
        ent.focus_set()

    def _hapus_tv(self, kartu):
        if (self.current_role or "kasir") != "admin":
            messagebox.showwarning("⚠ AKSES TERBATAS", "Hanya admin yang dapat menghapus TV.")
            return
        if not messagebox.askyesno("Hapus TV", f"Yakin hapus '{kartu.label_tv}' dari dashboard?"):
            return
        self._minta_pin_hapus(lambda: self._do_hapus_tv(kartu))

    def _do_hapus_tv(self, kartu):
        wid_to_del = None
        for wid, k in self._dash_card_windows[:]:
            if k is kartu:
                wid_to_del = wid
                self._dash_card_windows.remove((wid, k))
                break
        if wid_to_del is not None:
            canvas = self._scroll_canvas_dash
            if canvas:
                try: canvas.delete(wid_to_del)
                except Exception: pass
        try:
            self._semua_kartu_tv.remove(kartu)
        except ValueError:
            pass
        try:
            kartu.destroy()
        except Exception:
            pass
        self.jumlah_tv = max(0, self.jumlah_tv - 1)
        self.lbl_total_tv.configure(text=f"Total TV: {self.jumlah_tv}")
        self._refresh_dashboard_total_pesanan()
        self._simpan_daftar_tv()
        self._bersihkan_konfig_tv(kartu)
        self.after_idle(self._debounced_layout_dash)

    def _on_kartu_ganti_grup(self, kartu, grup_baru):
        """Saat user mengganti grup tarif sebuah TV: refresh closure get_paket_data
        kartu itu supaya menunjuk ke grup yang baru."""
        kartu.get_paket_data = lambda g=grup_baru: self.get_paket_data(g)
        self._simpan_daftar_tv()

    # ══════════════════════════════════════════════════════════════════════════
    #  TAB 2: Kontrol Harga
    # ══════════════════════════════════════════════════════════════════════════
    def _buka_dialog_tambah_warnet(self):
        if not self._tambah_warnet_btn_enabled: return
        ok_client, msg_client = self._check_warnet_client_connection()
        if not ok_client:
            messagebox.showwarning(
                "Client Belum Siap",
                "Tambah Kursi hanya bisa dilakukan jika Config Client sudah benar\n"
                "dan PC client berhasil terhubung.\n\n"
                f"Detail: {msg_client}",
                parent=self
            )
            return

        self._tambah_warnet_btn_enabled = False
        self.btn_tambah_warnet.configure(state="disabled", text="⏳ Menunggu...")

        # Ambil daftar grup khusus warnet
        daftar_warnet = self.daftar_nama_grup(for_warnet=True)
        if not daftar_warnet:
            try:
                cfg = ConfigManager.load()
                shared = cfg.get('grup_tarif', {}) or {}
                fallback_group = next(iter(shared.keys())) if shared else NAMA_GRUP_DEFAULT
                source_map = shared.get(fallback_group, _PAKET_STANDAR)
                warnet_map = cfg.get('grup_tarif_warnet', {}) or {}
                warnet_map['Warnet'] = {k: dict(v) for k, v in source_map.items()} if isinstance(source_map, dict) else dict(source_map)
                cfg['grup_tarif_warnet'] = warnet_map
                ConfigManager.save(cfg)
                daftar_warnet = self.daftar_nama_grup(for_warnet=True)
            except Exception as e:
                print(f"[WARN] Gagal membuat grup_warnet default: {e}", flush=True)
                daftar_warnet = [NAMA_GRUP_DEFAULT]

        # Build PC list from connected sessions
        pc_options = []
        server = getattr(self, 'warnet_server', None)
        if server and getattr(server, 'running', False):
            with server.sessions_lock:
                active_sessions = list(server.sessions.values())
            cfg = ConfigManager.load()
            for sess in active_sessions:
                sid = sess.get("client_id", "")
                for c in cfg.get("warnet_clients", []):
                    if c.get("client_id") == sid:
                        for p in c.get("pcs", []):
                            label = f"{p.get('name', '?')} ({p.get('ip', '?')})"
                            pc_options.append({
                                "label": label,
                                "client_id": sid,
                                "pc_id": p.get("pc_id", ""),
                                "ip": p.get("ip", ""),
                                "name": p.get("name", ""),
                            })

        DialogTambahWarnet(self,
                            on_confirm=self._on_warnet_confirmed,
                            on_close_cb=self._unlock_tambah_warnet,
                            daftar_grup=daftar_warnet,
                            pc_options=pc_options)

    def _on_warnet_confirmed(self, nama, nama_grup, pc_info=None):
        self._unlock_tambah_warnet()
        self._tambah_warnet(nama=nama, nama_grup=nama_grup, pc_info=pc_info)

    def _tambah_warnet_demo(self):
        if (self.current_role or "kasir") != "admin":
            messagebox.showwarning("⚠ AKSES TERBATAS", "Hanya admin yang dapat menambah kursi.")
            return
        demo_count = sum(1 for k in getattr(self, '_semua_kartu_warnet', []) if k.label_kursi.startswith("Demo PC")) + 1
        demo_name = f"Demo PC {demo_count}"
        # Ensure warnet tab is visible so user can see the added demo card
        try:
            if hasattr(self, '_show_tab'):
                self._show_tab('warnet')
        except Exception:
            pass

        try:
            # Use first warnet group as demo default if available
            warnet_daftar = self.daftar_nama_grup(for_warnet=True)
            demo_group = warnet_daftar[0] if (warnet_daftar and len(warnet_daftar) > 0) else NAMA_GRUP_DEFAULT
            self._tambah_warnet(nama=demo_name, nama_grup=demo_group)
        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("❌ Demo Warnet Gagal", f"Gagal menambahkan Demo Warnet:\n{e}", parent=self)

    def _tambah_warnet(self, nama, nama_grup=None, pc_info=None, restore=False):
        if (self.current_role or "kasir") != "admin" and not restore:
            messagebox.showwarning("⚠ AKSES TERBATAS", "Hanya admin yang dapat menambah kursi.")
            return
        _, warnet_limit = get_edition_limits(current_user=self._resolve_license_user())
        warnet_limit = self._promo_override_tv_limit(warnet_limit, is_warnet=True)
        if len(getattr(self, '_semua_kartu_warnet', [])) >= warnet_limit:
            if restore:
                print(f"[WARNET] Restore dilewati '{nama}' — limit {warnet_limit} kursi tercapai", flush=True)
                return  # saat muat ulang: lewati diam-diam
            messagebox.showwarning("Limit Tercapai",
                f"Paket Anda hanya mengizinkan maksimal {warnet_limit} PC Warnet.\n"
                f"Hapus beberapa PC atau upgrade paket.")
            return
        # Update/register PC in config (read-only saat restore agar config
        # tidak ikut tertulis ulang dari jalur pemulihan daftar_warnet)
        if pc_info and not restore:
            try:
                cfg = ConfigManager.load()
                for c in cfg.get("warnet_clients", []):
                    if c.get("client_id") == pc_info["client_id"]:
                        for p in c.get("pcs", []):
                            if p.get("pc_id") == pc_info["pc_id"]:
                                p["name"] = nama
                                break
                        break
                ConfigManager.save(cfg)
            except Exception as e:
                print(f"[WARN] Gagal update config PC: {e}", flush=True)

        self.jumlah_warnet = getattr(self, 'jumlah_warnet', 0) + 1
        warnet_daftar = self.daftar_nama_grup(for_warnet=True)
        default_group = nama_grup or (warnet_daftar[0] if warnet_daftar else (self._grup_aktif if hasattr(self, '_grup_aktif') else NAMA_GRUP_DEFAULT))
        if not hasattr(self, '_semua_kartu_warnet'):
            self._semua_kartu_warnet = []
        canvas = self._scroll_canvas_warnet

        kartu = KartuWarnet(canvas, self.jumlah_warnet,
                            label_kursi=nama,
                            on_transaksi=self._catat_transaksi,
                            get_paket_data=lambda g=default_group: self.get_paket_data(g, for_warnet=True),
                            get_makanan_data=self.get_makanan_data,
                            get_minuman_data=self.get_minuman_data,
                            get_semua_kartu=lambda: self._semua_kartu_warnet,
                            get_daftar_grup=lambda: self.daftar_nama_grup(for_warnet=True),
                            on_ganti_grup=self._on_warnet_kartu_ganti_grup,
                            on_hapus=self._hapus_warnet,
                            nama_grup=default_group)
        wid = canvas.create_window(0, 0, window=kartu, anchor="nw", tags=("_warnet_card",))
        self._warnet_card_windows.append((wid, kartu))
        if pc_info:
            kartu._client_id = pc_info.get("client_id")
            kartu._pc_id = pc_info.get("pc_id")
            kartu._pc_ip = pc_info.get("ip")
            kartu._update_pc_status()
        self._semua_kartu_warnet.append(kartu)
        try:
            self.lbl_total_warnet.configure(text=f"Total Kursi: {self.jumlah_warnet}")
        except Exception:
            pass
        if hasattr(self, '_refresh_warnet_footer'):
            self._refresh_warnet_footer()
        if not restore:
            self._simpan_daftar_warnet()
        self.after_idle(self._debounced_layout_warnet)
        return kartu

    def _on_warnet_kartu_ganti_grup(self, kartu, grup_baru):
        # Binding grup -> harga paket kartu warnet ini (grup warnet atau shared)
        kartu.get_paket_data = lambda g=grup_baru: self.get_paket_data(g, for_warnet=True)
        if not kartu.sesi_kosong() and not kartu.is_bebas:
            total_pesanan = kartu.biaya_pesanan
            if hasattr(kartu, '_ids'):
                kartu.itemconfig(kartu._ids['lbl_paket'],
                    text=f"{fmt_durasi(int(kartu.sisa_waktu / 60))} | {fmt_rp(kartu.paket_harga_tetap + total_pesanan)}")
            else:
                kartu.lbl_paket.configure(
                    text=f"{fmt_durasi(int(kartu.sisa_waktu / 60))} | {fmt_rp(kartu.paket_harga_tetap + total_pesanan)}")
        self._simpan_daftar_warnet()

    def _hapus_warnet(self, kartu):
        """Hapus sebuah kartu Warnet dari dashboard.
        KartuWarnet._confirm_hapus sudah menanyakan konfirmasi, jadi di sini langsung lakukan penghapusan.
        """
        if (self.current_role or "kasir") != "admin":
            messagebox.showwarning("⚠ AKSES TERBATAS", "Hanya admin yang dapat menghapus kursi.")
            return
        self._minta_pin_hapus(lambda: self._do_hapus_warnet(kartu))

    def _do_hapus_warnet(self, kartu):
        try:
            # Remove from list if present
            if hasattr(self, '_semua_kartu_warnet') and kartu in self._semua_kartu_warnet:
                try:
                    self._semua_kartu_warnet.remove(kartu)
                except ValueError:
                    pass
            # Remove from canvas
            wid_to_del = None
            for wid, k in self._warnet_card_windows[:]:
                if k is kartu:
                    wid_to_del = wid
                    self._warnet_card_windows.remove((wid, k))
                    break
            if wid_to_del is not None:
                canvas = self._scroll_canvas_warnet
                if canvas:
                    try: canvas.delete(wid_to_del)
                    except Exception: pass
            # Destroy widget
            try:
                kartu.destroy()
            except Exception:
                pass
            # Update counters & footer
            self.jumlah_warnet = max(0, getattr(self, 'jumlah_warnet', 1) - 1)
            if hasattr(self, 'lbl_total_warnet'):
                try:
                    self.lbl_total_warnet.configure(text=f"Total Kursi: {self.jumlah_warnet}")
                except Exception:
                    pass
            if hasattr(self, '_refresh_warnet_footer'):
                try:
                    self._refresh_warnet_footer()
                except Exception:
                    pass
            self.after_idle(self._debounced_layout_warnet)
            self._simpan_daftar_warnet()
            AuditLogger.log(action="hapus_warnet", username=self.current_user or 'system', status='success', details={'label': getattr(kartu, 'label_kursi', 'unknown')})
        except Exception as e:
            AuditLogger.log(action="hapus_warnet", username=self.current_user or 'system', status='failed', details={'error': str(e)})
            try:
                messagebox.showerror("Error", f"Gagal menghapus kursi: {e}", parent=self)
            except Exception:
                pass

    def _refresh_warnet_footer(self):
        # Sum warnet paket totals from riwayat_meta (package sales only)
        total_warnet = sum(m.get('paket_harga', m.get('total', 0) - m.get('pesanan_total', 0)) for m in self.riwayat_meta if m.get('source') == 'warnet')
        if hasattr(self, 'lbl_warnet_total_pendapatan'):
            self.lbl_warnet_total_pendapatan.configure(text=f"Total Pendapatan Warnet: {fmt_rp(total_warnet)}")
 
    def _refresh_warnet_socket_status(self):
        if not hasattr(self, 'lbl_socket_warnet'):
            return
        ok_client, msg = self._check_warnet_client_connection()
        self.lbl_socket_warnet.configure(
            text=f"Client: {'Tersambung' if ok_client else 'Belum tersambung'}",
            text_color=C_GREEN if ok_client else C_YELLOW
        )

    def _schedule_warnet_status_refresh(self):
        self._refresh_warnet_socket_status()
        if hasattr(self, 'lbl_socket_warnet'):
            self.after(10000, self._schedule_warnet_status_refresh)

    def _check_warnet_client_connection(self):
        server = getattr(self, 'warnet_server', None)
        if server is None or not getattr(server, 'running', False):
            return False, "Socket server billing belum aktif."

        with server.sessions_lock:
            active_sessions = list(server.sessions.values())

        if not active_sessions:
            return False, "Belum ada PC client yang AUTH/login ke server."

        return True, f"{len(active_sessions)} client aktif terhubung."

    def _open_warnet_admin_code_generator(self):
        DialogWarnetAdminCode(self)

    def _open_deploy_client_dialog(self):
        """Buka dialog deploy client app ke PC warnet via SSH."""
        DialogDeployClient(self, warnet_server=self.warnet_server,
                           config_manager=ConfigManager)

    def _buka_upload_logo(self):
        """Upload logo lockscreen client → disimpan ke folder paket client.
        Logo ikut terkirim ke PC client saat tombol 🚀 Deploy Client dijalankan."""
        path = filedialog.askopenfilename(
            parent=self, title="Pilih Logo Lockscreen",
            filetypes=[("Gambar", "*.png *.jpg *.jpeg *.bmp"), ("Semua file", "*.*")])
        if not path:
            return
        try:
            img = Image.open(path)
            img.verify()
        except Exception as e:
            messagebox.showerror("Logo Tidak Valid",
                                 f"File bukan gambar yang valid:\n{e}", parent=self)
            return
        pkg_dir = None
        for p in DialogDeployClient.PKG_HINTS:
            if os.path.isdir(p):
                pkg_dir = p
                break
        if pkg_dir is None:
            messagebox.showerror("Folder Paket Tidak Ditemukan",
                                 "Folder paket client tidak ditemukan.\n"
                                 "Jalankan 🚀 Deploy Client sekali lalu coba lagi.", parent=self)
            return
        ext = os.path.splitext(path)[1].lower()
        if ext in (".jpg", ".jpeg", ".bmp"):
            dest = os.path.join(pkg_dir, "lockscreen_logo.jpg")
        else:
            dest = os.path.join(pkg_dir, "lockscreen_logo.png")
        try:
            shutil.copy2(path, dest)
        except Exception as e:
            messagebox.showerror("Gagal Simpan",
                                 f"Tidak bisa menyimpan logo:\n{e}", parent=self)
            return
        messagebox.showinfo("✅ Logo Tersimpan",
                            f"Logo disimpan ke:\n{dest}\n\nLogo akan ikut terkirim "
                            "ke PC client saat tombol 🚀 Deploy Client dijalankan.", parent=self)
 
    def _setup_harga(self):
        f = self.frames["harga"]
        hdr = ctk.CTkFrame(f, fg_color=C_PANEL, height=54, corner_radius=0)
        hdr.pack(fill="x")
        ctk.CTkLabel(hdr, text="⚙️  KONTROL HARGA",
                     font=FONT_TITLE, text_color=C_ACCENT).pack(side="left", padx=18, pady=14)
        ctk.CTkButton(hdr, text="💾 Simpan Semua", width=150, height=34,
                      fg_color="#1A4A1A", hover_color="#0A3A0A",
                      border_width=1, border_color=C_GREEN,
                      font=("Russo One", 10, "bold"), text_color=C_GREEN,
                      command=self._simpan_harga).pack(side="right", padx=18, pady=10)

        self.scroll_harga = ctk.CTkScrollableFrame(f, fg_color=C_BG)
        self.scroll_harga.pack(fill="both", expand=True, padx=14, pady=10)

        self._bersihkan_grup_warnet_bocor()
        self._pastikan_grup_warnet_ada()
        self._grup_aktif = self.daftar_nama_grup()[0]
        self._harga_entries = {}
        self._build_kotak_grup_dan_info()
        self._build_harga_section_editable("paket", self._judul_section_paket(),
                                            self.grup_tarif[self._grup_aktif])
        self._build_harga_section_editable("makanan", "🍔  Menu Makanan",     self.menu_makanan)
        self._build_harga_section_editable("minuman", "🥤  Menu Minuman",     self.menu_minuman)

    def _pastikan_grup_warnet_ada(self):
        """Jamin minimal ada satu grup warnet ('Warnet') di config, supaya Kontrol Harga
        selalu bisa mengedit paket warnet. Dibuat dari salinan grup shared pertama."""
        cfg = ConfigManager.load()
        warnet_map = cfg.get('grup_tarif_warnet', {}) or {}
        if warnet_map:
            return
        if getattr(self, 'grup_tarif', None):
            first = next(iter(self.grup_tarif.values()))
            warnet_map['Warnet'] = {k: dict(v) for k, v in first.items()}
        else:
            warnet_map['Warnet'] = {k: dict(v) for k, v in _PAKET_STANDAR.items()}
        cfg['grup_tarif_warnet'] = warnet_map
        cfg['warnet_only_groups'] = sorted(set(cfg.get('warnet_only_groups', []) or []) | {'Warnet'})
        ConfigManager.save(cfg)
        # Muat ke memori supaya dropdown/editing langsung melihatnya
        if 'Warnet' not in self.grup_tarif:
            self.grup_tarif['Warnet'] = {k: dict(v) for k, v in warnet_map['Warnet'].items()}

    def _judul_section_paket(self):
        if self._is_warnet_group(self._grup_aktif):
            return f"⏱  Paket Waktu Main — Grup WARNET: {self._grup_aktif} 🖥"
        return f"⏱  Paket Waktu Main — Grup: {self._grup_aktif}"

    def _bersihkan_grup_warnet_bocor(self):
        """Migrasi 1x: grup warnet yang terlanjur bocor ke grup_tarif (shared)
        dipindah ke grup_tarif_warnet + warnet_only_groups supaya tidak bisa dipakai kartu TV."""
        cfg = ConfigManager.load()
        warnet_map = dict(cfg.get('grup_tarif_warnet', {}) or {})
        warnet_only = set(cfg.get('warnet_only_groups', []) or [])
        shared = dict(cfg.get('grup_tarif', {}) or {})
        changed = False
        for g, paket in list(shared.items()):
            if g in warnet_only or g in warnet_map:
                warnet_map.setdefault(g, paket)
                del shared[g]
                changed = True
        if changed:
            cfg['grup_tarif'] = shared
            cfg['grup_tarif_warnet'] = warnet_map
            cfg['warnet_only_groups'] = sorted(set(warnet_only) | set(warnet_map.keys()))
            ConfigManager.save(cfg)

    def _refresh_grup_info(self):
        jumlah_tv = sum(1 for k in self._semua_kartu_tv if k.nama_grup == self._grup_aktif)
        jumlah_warnet = sum(1 for k in getattr(self, '_semua_kartu_warnet', []) if k.nama_grup == self._grup_aktif)
        total_pengguna = jumlah_tv + jumlah_warnet
        if self._is_warnet_group(self._grup_aktif):
            label_sumber = "PC Warnet (grup khusus warnet — tidak bisa dipakai TV)"
        else:
            label_sumber = "TV" if jumlah_warnet == 0 else "TV dan Warnet" if jumlah_tv and jumlah_warnet else "Warnet"
        self.lbl_grup_info.configure(
            text=f"Sedang mengedit harga untuk grup '{self._grup_aktif}'  ·  "
                 f"dipakai oleh {total_pengguna} {label_sumber} saat ini di Dashboard.")

    def _refresh_info_bebas(self):
        tarif_skrg = hitung_tarif_per_menit(self.grup_tarif[self._grup_aktif])
        self.lbl_info_bebas.configure(
            text=f"ℹ️  Tarif 'Main Bebas' grup '{self._grup_aktif}' dihitung otomatis dari paket "
                 f"'{PAKET_ACUAN_BEBAS}'  →  ≈ {fmt_rp(tarif_skrg)} / menit.\n"
                 f"    Ubah harga/durasi paket '{PAKET_ACUAN_BEBAS}' di bawah untuk mengubah tarif Main Bebas grup ini.")

    def _ganti_grup_aktif(self, grup_baru):
        self._simpan_paket_aktif_ke_memori()
        self._grup_aktif = grup_baru
        
        # Jika grup_baru tidak ada di self.grup_tarif (warnet group), muat dari config
        if self._grup_aktif not in self.grup_tarif:
            cfg = ConfigManager.load()
            warnet_map = cfg.get('grup_tarif_warnet', {}) or {}
            if self._grup_aktif in warnet_map:
                self.grup_tarif[self._grup_aktif] = warnet_map[self._grup_aktif]
        
        self._refresh_grup_info()
        self._refresh_info_bebas()
        self._rebuild_harga()

    def _tambah_grup_tarif(self):
        dlg = ctk.CTkInputDialog(text="Nama grup tarif baru (mis. PS5, Room VIP 2):",
                                  title="➕ Tambah Grup Tarif")
        nama_baru = dlg.get_input()
        if not nama_baru:
            return
        nama_baru = nama_baru.strip()
        if not nama_baru:
            return
        if nama_baru in self.grup_tarif:
            messagebox.showwarning("⚠ Sudah Ada", f"Grup '{nama_baru}' sudah ada.")
            return
        # Grup baru dimulai dari salinan paket grup yang sedang aktif, supaya
        # user tinggal sesuaikan harganya saja (tidak mulai dari kosong).
        # Tipe grup mengikuti grup aktif: kalau grup aktif adalah grup warnet,
        # grup baru ini juga menjadi grup warnet (tidak bisa dipakai TV).
        is_warnet_baru = self._is_warnet_group(self._grup_aktif)
        self.grup_tarif[nama_baru] = {k: dict(v) for k, v in self.grup_tarif[self._grup_aktif].items()}
        if is_warnet_baru:
            cfg = ConfigManager.load()
            cfg['warnet_only_groups'] = sorted(set(cfg.get('warnet_only_groups', []) or []) | {nama_baru})
            ConfigManager.save(cfg)
        self._simpan_paket_aktif_ke_memori()
        self._grup_aktif = nama_baru
        self._simpan_grup_tarif_ke_config()
        self._refresh_opt_grup_semua()
        self._refresh_grup_info()
        self._refresh_info_bebas()
        self._rebuild_harga()
        tipe = "grup warnet (khusus PC)" if is_warnet_baru else "grup (TV & PC warnet)"
        messagebox.showinfo("✅ Grup Ditambahkan",
                            f"Grup '{nama_baru}' dibuat sebagai {tipe} (salinan harga dari grup aktif).\n"
                            f"Atur harganya lalu klik Simpan Semua.\n\n"
                            f"Grup baru ini langsung tersedia di Dashboard TV/Warnet saat menambah/mengganti kartu.")

    def _tambah_grup_warnet(self):
        dlg = ctk.CTkInputDialog(text="Nama grup tarif warnet baru (mis. Warnet Reguler, Warnet VIP):",
                                  title="➕ Tambah Grup Warnet")
        nama_baru = dlg.get_input()
        if not nama_baru:
            return
        nama_baru = nama_baru.strip()
        if not nama_baru:
            return
        if nama_baru in self.grup_tarif or self._is_warnet_group(nama_baru):
            messagebox.showwarning("⚠ Sudah Ada", f"Grup '{nama_baru}' sudah ada.")
            return
        # Mulai dari salinan paket grup aktif (atau grup 'Warnet' bila ada)
        sumber_nama = self._grup_aktif if self._grup_aktif in self.grup_tarif else 'Warnet'
        sumber = self.grup_tarif.get(sumber_nama) or self.grup_tarif.get('Warnet') or _PAKET_STANDAR
        self.grup_tarif[nama_baru] = {k: dict(v) for k, v in sumber.items()}
        cfg = ConfigManager.load()
        cfg['warnet_only_groups'] = sorted(set(cfg.get('warnet_only_groups', []) or []) | {nama_baru})
        ConfigManager.save(cfg)
        self._simpan_paket_aktif_ke_memori()
        self._grup_aktif = nama_baru
        self._simpan_grup_tarif_ke_config()
        self._refresh_opt_grup_semua()
        self._refresh_grup_info()
        self._refresh_info_bebas()
        self._rebuild_harga()
        messagebox.showinfo("✅ Grup Warnet Ditambahkan",
                            f"Grup warnet '{nama_baru}' dibuat (salinan harga dari '{sumber_nama}').\n"
                            f"Atur harganya lalu klik Simpan Semua.\n\n"
                            f"Grup ini hanya bisa diikat ke kartu PC/Warnet — kartu TV tidak akan memakainya.")

    def _rename_grup_tarif(self):
        grup_lama = self._grup_aktif
        dlg = ctk.CTkInputDialog(text=f"Nama baru untuk grup '{grup_lama}':",
                                  title="✏️ Ganti Nama Grup")
        nama_baru = dlg.get_input()
        if not nama_baru:
            return
        nama_baru = nama_baru.strip()
        if not nama_baru or nama_baru == grup_lama:
            return
        if nama_baru in self.grup_tarif:
            messagebox.showwarning("⚠ Sudah Ada", f"Grup '{nama_baru}' sudah ada.")
            return
        self._simpan_paket_aktif_ke_memori()
        is_warnet_lama = self._is_warnet_group(grup_lama)
        self.grup_tarif[nama_baru] = self.grup_tarif.pop(grup_lama)
        if is_warnet_lama:
            # Update nama di warnet_only_groups supaya tetap terkunci untuk warnet
            cfg = ConfigManager.load()
            warnet_only = [g if g != grup_lama else nama_baru for g in (cfg.get('warnet_only_groups', []) or [])]
            cfg['warnet_only_groups'] = sorted(warnet_only)
            warnet_map = dict(cfg.get('grup_tarif_warnet', {}) or {})
            if grup_lama in warnet_map:
                warnet_map[nama_baru] = warnet_map.pop(grup_lama)
                cfg['grup_tarif_warnet'] = warnet_map
            ConfigManager.save(cfg)
        for kartu in self._semua_kartu_tv:
            if kartu.nama_grup == grup_lama:
                kartu.nama_grup = nama_baru
                kartu.itemconfig(kartu._ids['lbl_grup'], text=f"\U0001f3f7 {nama_baru}")
                kartu.get_paket_data = lambda g=nama_baru: self.get_paket_data(g)
        for kartu in getattr(self, '_semua_kartu_warnet', []):
            if kartu.nama_grup == grup_lama:
                kartu.nama_grup = nama_baru
                if hasattr(kartu, '_ids'):
                    kartu.itemconfig(kartu._ids['lbl_grup'], text=f"\u21bb {nama_baru}")
                else:
                    kartu.lbl_grup.configure(text=f"\u21bb {nama_baru}")
                kartu.get_paket_data = lambda g=nama_baru: self.get_paket_data(g, for_warnet=True)
        self._grup_aktif = nama_baru
        self._simpan_grup_tarif_ke_config()
        self._refresh_opt_grup_semua()
        self._refresh_grup_info()
        self._refresh_info_bebas()
        self._rebuild_harga()

    def _hapus_grup_tarif(self):
        grup = self._grup_aktif
        if grup in GRUP_TERKUNCI or len(self.grup_tarif) <= 1:
            messagebox.showwarning("⚠ Tidak Bisa Dihapus",
                                    f"Grup '{grup}' tidak bisa dihapus "
                                    f"(minimal harus ada 1 grup tersisa).")
            return
        is_warnet = self._is_warnet_group(grup)
        jumlah_tv = sum(1 for k in self._semua_kartu_tv if k.nama_grup == grup)
        jumlah_warnet = sum(1 for k in getattr(self, '_semua_kartu_warnet', []) if k.nama_grup == grup)
        pesan = f"Hapus grup tarif '{grup}'?"
        if is_warnet:
            pesan += (f"\n\n⚠ {jumlah_warnet} PC Warnet di Dashboard saat ini memakai grup ini.\n"
                      f"PC tersebut akan otomatis dipindah ke grup warnet lain.")
        elif jumlah_tv:
            pesan += (f"\n\n⚠ {jumlah_tv} TV di Dashboard saat ini memakai grup ini.\n"
                      f"TV tersebut akan otomatis dipindah ke grup '{NAMA_GRUP_DEFAULT}'.")
        if not messagebox.askyesno("🗑 Hapus Grup Tarif", pesan):
            return
        del self.grup_tarif[grup]
        if is_warnet:
            # Hapus dari warnet_map + warnet_only_groups
            cfg = ConfigManager.load()
            warnet_map = dict(cfg.get('grup_tarif_warnet', {}) or {})
            warnet_map.pop(grup, None)
            cfg['grup_tarif_warnet'] = warnet_map
            cfg['warnet_only_groups'] = [g for g in (cfg.get('warnet_only_groups', []) or []) if g != grup]
            ConfigManager.save(cfg)
            sisa_warnet = [g for g in self.daftar_nama_grup(for_warnet=True) if g != grup]
            fallback = ('Warnet' if 'Warnet' in sisa_warnet else sisa_warnet[0]) if sisa_warnet else NAMA_GRUP_DEFAULT
            for kartu in getattr(self, '_semua_kartu_warnet', []):
                if kartu.nama_grup == grup:
                    kartu.nama_grup = fallback
                    if hasattr(kartu, '_ids'):
                        kartu.itemconfig(kartu._ids['lbl_grup'], text=f"\u21bb {fallback}")
                    else:
                        kartu.lbl_grup.configure(text=f"\u21bb {fallback}")
                    kartu.get_paket_data = lambda g=fallback: self.get_paket_data(g, for_warnet=True)
        else:
            fallback = NAMA_GRUP_DEFAULT if NAMA_GRUP_DEFAULT in self.grup_tarif else next(iter(self.grup_tarif))
            for kartu in self._semua_kartu_tv:
                if kartu.nama_grup == grup:
                    kartu.nama_grup = fallback
                    if hasattr(kartu, '_ids') and 'lbl_grup' in kartu._ids:
                        kartu.itemconfig(kartu._ids['lbl_grup'], text=f"\U0001f3f7 {fallback}")
                    else:
                        kartu.lbl_grup.configure(text=f"\U0001f3f7 {fallback}")
                    kartu.get_paket_data = lambda g=fallback: self.get_paket_data(g)
            for kartu in getattr(self, '_semua_kartu_warnet', []):
                if kartu.nama_grup == grup:
                    kartu.nama_grup = fallback
                    if hasattr(kartu, '_ids'):
                        kartu.itemconfig(kartu._ids['lbl_grup'], text=f"\u21bb {fallback}")
                    else:
                        kartu.lbl_grup.configure(text=f"\u21bb {fallback}")
                    kartu.get_paket_data = lambda g=fallback: self.get_paket_data(g, for_warnet=True)
        self._grup_aktif = fallback
        self._simpan_grup_tarif_ke_config()
        self._refresh_opt_grup_semua()
        self._refresh_grup_info()
        self._refresh_info_bebas()
        self._rebuild_harga()

    def _refresh_opt_grup_semua(self):
        """Refresh dropdown grup di tab Kontrol Harga setelah daftar grup berubah."""
        daftar = self.daftar_semua_grup()
        self.opt_grup_aktif.configure(values=daftar)
        self.var_grup_aktif.set(self._grup_aktif)

    def _simpan_grup_tarif_ke_config(self, cfg=None):
        """Simpan grup tarif ke config dengan PEMISAHAN shared vs warnet.

        Grup warnet (warnet_only_groups ∪ grup_tarif_warnet) disimpan ke
        'grup_tarif_warnet' + 'warnet_only_groups', BUKAN ke 'grup_tarif' shared —
        supaya grup warnet tidak bocor ke dropdown kartu TV.
        """
        cfg = cfg or ConfigManager.load()
        warnet_map = dict(cfg.get('grup_tarif_warnet', {}) or {})
        warnet_only = set(cfg.get('warnet_only_groups', []) or [])
        shared = {}
        for g, paket in self.grup_tarif.items():
            if g in warnet_only or g in warnet_map:
                warnet_map[g] = paket
            else:
                shared[g] = paket
        cfg["grup_tarif"] = shared
        cfg["grup_tarif_warnet"] = warnet_map
        cfg["warnet_only_groups"] = sorted(set(warnet_only) | set(warnet_map.keys()))
        ConfigManager.save(cfg)

    def _simpan_paket_aktif_ke_memori(self):
        """Tulis perubahan yang sedang diketik user (belum ditekan Simpan Semua)
        kembali ke self.grup_tarif[grup_aktif] sebelum pindah grup/rebuild,
        supaya perubahan yang belum disimpan permanen tidak hilang begitu saja
        saat sekadar berpindah tampilan antar grup."""
        if not hasattr(self, "_harga_entries"):
            return
        hasil = {}
        for (kategori, nama_asli), entry in self._harga_entries.items():
            if kategori != "paket":
                continue
            nama_baru = entry["var_nama"].get().strip() or nama_asli
            if nama_asli == "Main Bebas":
                hasil["Main Bebas"] = {"harga": 0, "menit": 0}
                continue
            try:
                harga_baru = int(entry["var_harga"].get().strip() or 0)
            except ValueError:
                harga_baru = 0
            try:
                jam_baru = int(entry["var_jam"].get().strip() or 0)
            except ValueError:
                jam_baru = 0
            try:
                menit_baru = int(entry["var_menit"].get().strip() or 0)
            except ValueError:
                menit_baru = 0
            hasil[nama_baru] = {"harga": harga_baru, "menit": max(0, jam_baru) * 60 + max(0, menit_baru)}
        if "Main Bebas" not in hasil:
            hasil["Main Bebas"] = {"harga": 0, "menit": 0}
        if hasil and self._grup_aktif in self.grup_tarif:
            self.grup_tarif[self._grup_aktif] = hasil

    def _build_harga_section_editable(self, kategori, judul, menu_dict):
        sec = ctk.CTkFrame(self.scroll_harga, fg_color=C_PANEL, corner_radius=12)
        sec.pack(fill="x", pady=8)
        if kategori == "paket":
            self._frame_paket_section = sec

        hdr_row = ctk.CTkFrame(sec, fg_color="transparent")
        hdr_row.pack(fill="x", padx=16, pady=(10, 6))
        tk.Label(hdr_row, text=judul, font=FONT_SUB, fg=C_ACCENT2, bg=C_PANEL).pack(side="left")
        ctk.CTkButton(hdr_row, text="+ Tambah", width=150, height=34,
                      fg_color="#1A4A1A", hover_color="#0A3A0A",
                      border_width=1, border_color=C_GREEN,
                      font=("Russo One", 10, "bold"), text_color=C_GREEN,
                      command=lambda k=kategori: self._tambah_item_harga(k)).pack(side="right")

        if kategori == "paket":
            kol_row = tk.Frame(sec, bg=C_PANEL)
            kol_row.pack(fill="x", padx=16, pady=(0, 2))
            tk.Label(kol_row, text="Nama Paket", font=FONT_SMALL, fg=C_MUTED, bg=C_PANEL,
                     width=24, anchor="w").pack(side="left", padx=(0, 8))
            tk.Label(kol_row, text="Harga (Rp)", font=FONT_SMALL, fg=C_MUTED, bg=C_PANEL,
                     width=14, anchor="w").pack(side="left", padx=(0, 4))
            tk.Label(kol_row, text=" Rp  ", font=FONT_SMALL, fg=C_MUTED, bg=C_PANEL).pack(side="left", padx=(0, 12))
            tk.Label(kol_row, text="Jam", font=FONT_SMALL, fg=C_MUTED, bg=C_PANEL,
                     width=6, anchor="w").pack(side="left", padx=(0, 4))
            tk.Label(kol_row, text=" jam ", font=FONT_SMALL, fg=C_MUTED, bg=C_PANEL).pack(side="left", padx=(0, 10))
            tk.Label(kol_row, text="Menit", font=FONT_SMALL, fg=C_MUTED, bg=C_PANEL,
                     width=6, anchor="w").pack(side="left", padx=(0, 4))
            tk.Label(kol_row, text=" menit ", font=FONT_SMALL, fg=C_MUTED, bg=C_PANEL).pack(side="left", padx=(0, 12))
            tk.Label(kol_row, text=" ", font=FONT_SMALL, fg=C_MUTED, bg=C_PANEL,
                     width=4).pack(side="left")

        container = tk.Frame(sec, bg=C_PANEL)
        container.pack(fill="x", padx=0, pady=(0, 8))

        for nama, val in menu_dict.items():
            self._tambah_row_harga(container, kategori, nama, val, menu_dict)

        if not hasattr(self, '_harga_containers'):
            self._harga_containers = {}
        self._harga_containers[kategori] = (container, menu_dict)

    def _tambah_row_harga(self, container, kategori, nama, val, menu_dict):
        padx_left = 16
        row = tk.Frame(container, bg=C_PANEL)
        row.pack(fill="x", padx=padx_left, pady=2)

        if kategori == "paket":
            harga = val.get("harga", 0)
            menit_total = val.get("menit", 0)
            jam, menit = divmod(menit_total, 60)
        else:
            harga = val

        var_nama = tk.StringVar(value=nama)
        e_nama = tk.Entry(row, textvariable=var_nama, width=24,
                          bg=C_BTN, fg=C_TEXT, bd=1, font=FONT_SMALL)
        e_nama.pack(side="left", padx=(0, 8))

        var_harga = tk.StringVar(value=str(harga))
        e_harga = tk.Entry(row, textvariable=var_harga, width=14,
                           bg=C_BTN, fg=C_ACCENT, bd=1, font=FONT_SMALL)
        e_harga.pack(side="left", padx=(0, 4))

        var_jam = None
        var_menit = None
        if kategori == "paket":
            tk.Label(row, text="Rp", font=FONT_SMALL, fg=C_MUTED, bg=C_PANEL,
                     width=4, anchor="w").pack(side="left", padx=(0, 12))

            is_bebas_row = (nama == "Main Bebas")

            var_jam = tk.StringVar(value=str(jam))
            e_jam = tk.Entry(row, textvariable=var_jam, width=6,
                             bg=C_BTN, fg=C_YELLOW, bd=1, font=FONT_SMALL)
            e_jam.pack(side="left", padx=(0, 4))
            if is_bebas_row:
                e_jam.configure(state="disabled")
            tk.Label(row, text="jam", font=FONT_SMALL, fg=C_MUTED, bg=C_PANEL,
                     width=4, anchor="w").pack(side="left", padx=(0, 10))

            var_menit = tk.StringVar(value=str(menit))
            e_menit = tk.Entry(row, textvariable=var_menit, width=6,
                               bg=C_BTN, fg=C_YELLOW, bd=1, font=FONT_SMALL)
            e_menit.pack(side="left", padx=(0, 4))
            if is_bebas_row:
                e_menit.configure(state="disabled")
            tk.Label(row, text="menit", font=FONT_SMALL, fg=C_MUTED, bg=C_PANEL,
                     width=6, anchor="w").pack(side="left", padx=(0, 12))

            if is_bebas_row:
                e_harga.configure(state="disabled")
                tk.Label(row, text="(otomatis dari acuan)", font=FONT_SMALL,
                         fg=C_GREEN, bg=C_PANEL).pack(side="left", padx=(0, 8))
        else:
            tk.Label(row, text="Rp", font=FONT_SMALL, fg=C_MUTED, bg=C_PANEL,
                     width=4, anchor="w").pack(side="left", padx=(0, 12))

        if not (kategori == "paket" and nama == "Main Bebas"):
            tk.Button(row, text="X", width=3, height=1,
                      bg=C_BTN, fg=C_RED, font=FONT_SMALL,
                      relief="ridge", bd=1,
                      command=lambda r=row, k=kategori, n=nama, md=menu_dict:
                              self._hapus_item_harga(r, k, n, md)).pack(side="left")

        key = (kategori, nama)
        self._harga_entries[key] = {
            "var_nama": var_nama, "var_harga": var_harga,
            "var_jam": var_jam, "var_menit": var_menit, "nama_asli": nama,
        }

    def _tambah_item_harga(self, kategori):
        if kategori == "paket":
            paket_grup = self.grup_tarif[self._grup_aktif]
            nama_baru = f"Paket Baru {len(paket_grup) + 1}"
            val = {"harga": 0, "menit": 60}
            paket_grup[nama_baru] = val
        elif kategori == "makanan":
            nama_baru = f"Item Baru {len(self._harga_entries)}"
            val = 0
            self.menu_makanan[nama_baru] = val
        else:
            nama_baru = f"Item Baru {len(self._harga_entries)}"
            val = 0
            self.menu_minuman[nama_baru] = val

        container, menu_dict = self._harga_containers[kategori]
        self._tambah_row_harga(container, kategori, nama_baru, val, menu_dict)
        messagebox.showinfo("Tambah Item",
                            f"Item '{nama_baru}' ditambahkan.\nEdit nama, harga"
                            + (", jam & menit" if kategori == "paket" else "")
                            + " lalu klik Simpan Semua.")

    def _hapus_item_harga(self, row_widget, kategori, nama_asli, menu_dict):
        if kategori == "paket" and nama_asli == "Main Bebas":
            messagebox.showwarning("Tidak Bisa Dihapus",
                                    "Paket 'Main Bebas' adalah paket inti sistem dan tidak bisa dihapus.")
            return
        if messagebox.askyesno("Hapus Item", f"Hapus '{nama_asli}'?"):
            if kategori == "paket":
                paket_grup = self.grup_tarif[self._grup_aktif]
                if nama_asli in paket_grup:
                    del paket_grup[nama_asli]
            elif kategori == "makanan" and nama_asli in self.menu_makanan:
                del self.menu_makanan[nama_asli]
                self._hapus_stok_item("makanan", nama_asli)
            elif kategori == "minuman" and nama_asli in self.menu_minuman:
                del self.menu_minuman[nama_asli]
                self._hapus_stok_item("minuman", nama_asli)
            row_widget.destroy()
            self._harga_entries.pop((kategori, nama_asli), None)
            self._stok_save()
            self._stok_update_badge()

    def _rebuild_harga(self):
        for w in self.scroll_harga.winfo_children():
            w.destroy()
        self._harga_entries = {}
        # Kotak grup tarif & info Main Bebas dibangun ulang juga supaya konsisten
        self._build_kotak_grup_dan_info()
        self._build_harga_section_editable("paket", self._judul_section_paket(),
                                            self.grup_tarif[self._grup_aktif])
        self._build_harga_section_editable("makanan", "🍔  Menu Makanan",     self.menu_makanan)
        self._build_harga_section_editable("minuman", "🥤  Menu Minuman",     self.menu_minuman)

    def _build_kotak_grup_dan_info(self):
        grup_box = ctk.CTkFrame(self.scroll_harga, fg_color=C_PANEL, corner_radius=12,
                                 border_width=1, border_color=C_ACCENT2)
        grup_box.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(grup_box, text="🏷  GRUP TARIF (mis. PS3, PS4, Room VIP — masing-masing harga sendiri)",
                     font=FONT_SUB, text_color=C_ACCENT2).pack(anchor="w", padx=16, pady=(12, 6))

        grup_row = ctk.CTkFrame(grup_box, fg_color="transparent")
        grup_row.pack(fill="x", padx=16, pady=(0, 6))
        self.var_grup_aktif = ctk.StringVar(value=self._grup_aktif)
        self.opt_grup_aktif = ctk.CTkOptionMenu(
            grup_row, values=self.daftar_semua_grup(), variable=self.var_grup_aktif,
            fg_color=C_BTN, button_color=C_ACCENT2, button_hover_color="#5A0FCC",
            text_color=C_TEXT, font=("Consolas", 11, "bold"), dropdown_font=FONT_BODY,
            dropdown_fg_color=C_CARD, dropdown_text_color=C_TEXT, width=220,
            command=self._ganti_grup_aktif)
        self.opt_grup_aktif.pack(side="left", padx=(0, 8))

        ctk.CTkButton(grup_row, text="➕ Grup Baru", width=110, height=30,
                      fg_color=C_BTN, border_width=1, border_color=C_GREEN,
                      font=FONT_SMALL, text_color=C_GREEN,
                      command=self._tambah_grup_tarif).pack(side="left", padx=4)
        ctk.CTkButton(grup_row, text="🖥➕ Grup Warnet", width=110, height=30,
                      fg_color=C_BTN, border_width=1, border_color="#0FA0CE",
                      font=FONT_SMALL, text_color="#0FA0CE",
                      command=self._tambah_grup_warnet).pack(side="left", padx=4)
        ctk.CTkButton(grup_row, text="✏️ Ganti Nama", width=110, height=30,
                      fg_color=C_BTN, border_width=1, border_color=C_YELLOW,
                      font=FONT_SMALL, text_color=C_YELLOW,
                      command=self._rename_grup_tarif).pack(side="left", padx=4)
        ctk.CTkButton(grup_row, text="🗑 Hapus Grup", width=110, height=30,
                      fg_color=C_BTN, border_width=1, border_color=C_RED,
                      font=FONT_SMALL, text_color=C_RED,
                      command=self._hapus_grup_tarif).pack(side="left", padx=4)

        self.lbl_grup_info = ctk.CTkLabel(
            grup_box, text="", font=FONT_SMALL, text_color=C_MUTED, justify="left")
        self.lbl_grup_info.pack(anchor="w", padx=16, pady=(0, 12))
        self._refresh_grup_info()

        info_bebas = ctk.CTkFrame(self.scroll_harga, fg_color=C_PANEL, corner_radius=10,
                                   border_width=1, border_color=C_GREEN)
        info_bebas.pack(fill="x", pady=(0, 10))
        self.lbl_info_bebas = ctk.CTkLabel(
            info_bebas, text="", font=FONT_SMALL, text_color=C_GREEN, justify="left")
        self.lbl_info_bebas.pack(anchor="w", padx=14, pady=10)
        self._refresh_info_bebas()

    def _simpan_harga(self):
        new_makanan = {}
        new_minuman = {}
        new_paket_grup_aktif = {}

        for (kategori, nama_asli), entry in self._harga_entries.items():
            nama_baru  = entry["var_nama"].get().strip()
            harga_str  = entry["var_harga"].get().strip()
            if not nama_baru:
                continue

            # "Main Bebas" namanya dikunci, harga & durasi dihitung otomatis (bukan dari input)
            if kategori == "paket" and nama_asli == "Main Bebas":
                new_paket_grup_aktif["Main Bebas"] = {"harga": 0, "menit": 0}
                continue

            try:
                harga_baru = int(harga_str)
            except ValueError:
                harga_baru = 0

            if kategori == "paket":
                try:
                    jam_baru = int(entry["var_jam"].get().strip() or 0)
                except ValueError:
                    jam_baru = 0
                try:
                    menit_baru = int(entry["var_menit"].get().strip() or 0)
                except ValueError:
                    menit_baru = 0
                total_menit = max(0, jam_baru) * 60 + max(0, menit_baru)
                new_paket_grup_aktif[nama_baru] = {"harga": harga_baru, "menit": total_menit}
            elif kategori == "makanan":
                new_makanan[nama_baru] = harga_baru
            elif kategori == "minuman":
                new_minuman[nama_baru] = harga_baru
            # Ikutkan entri stok saat item di-rename
            if kategori in ("makanan", "minuman") and nama_asli != nama_baru:
                self._remap_stok_rename(kategori, nama_asli, nama_baru)

        # Jaga-jaga: pastikan "Main Bebas" selalu ada walau entry-nya hilang
        if "Main Bebas" not in new_paket_grup_aktif:
            new_paket_grup_aktif["Main Bebas"] = {"harga": 0, "menit": 0}

        if new_paket_grup_aktif:
            self.grup_tarif[self._grup_aktif] = new_paket_grup_aktif
        self.menu_makanan = new_makanan if new_makanan else self.menu_makanan
        self.menu_minuman = new_minuman if new_minuman else self.menu_minuman

        cfg = ConfigManager.load()
        cfg["menu_makanan"] = self.menu_makanan
        cfg["menu_minuman"] = self.menu_minuman
        cfg["stok"] = self.stok
        cfg["stok_min"] = self.stok_min
        # Grup tarif disimpan dengan pemisahan shared vs warnet (anti-bocor ke kartu TV)
        self._simpan_grup_tarif_ke_config(cfg)

        # Sinkronkan tampilan Dashboard TV: semua KartuTV yang sedang memakai
        # grup yang baru saja diedit otomatis pakai data harga terbaru
        # (karena get_paket_data closure-nya membaca self.grup_tarif langsung).

        messagebox.showinfo("✅ Tersimpan",
                            f"Harga & durasi untuk grup '{self._grup_aktif}' berhasil disimpan!\n"
                            f"Tarif Main Bebas grup ini sekarang: "
                            f"≈ {fmt_rp(hitung_tarif_per_menit(self.grup_tarif[self._grup_aktif]))}/menit")
        self._refresh_grup_info()
        self._refresh_info_bebas()
        self._rebuild_harga()
        # Update segera ke halaman web pelanggan (call.html Panggil Kasir):
        # menu makanan/minuman, paket, dan stok terbaru — tanpa menunggu
        # refresh berkala 5 menit.
        try:
            self._qr_push_menu_bg()
        except Exception:
            pass


    # ══════════════════════════════════════════════════════════════════════════
    #  TAB STOK: kelola stok makanan & minuman
    # ══════════════════════════════════════════════════════════════════════════
    def _setup_stok(self):
        f = self.frames["stok"]
        hdr = ctk.CTkFrame(f, fg_color=C_PANEL, height=54, corner_radius=0)
        hdr.pack(fill="x")
        ctk.CTkLabel(hdr, text="📦  MANAJEMEN STOK",
                     font=FONT_TITLE, text_color=C_ACCENT).pack(side="left", padx=18, pady=14)
        ctk.CTkButton(hdr, text="💾 Simpan Semua", width=150, height=34,
                      fg_color="#1A4A1A", hover_color="#0A3A0A",
                      border_width=1, border_color=C_GREEN,
                      font=("Russo One", 10, "bold"), text_color=C_GREEN,
                      command=self._stok_simpan_semua).pack(side="right", padx=18, pady=10)

        info = ctk.CTkFrame(f, fg_color=C_PANEL, corner_radius=10,
                            border_width=1, border_color=C_ACCENT2)
        info.pack(fill="x", padx=14, pady=(10, 4))
        ctk.CTkLabel(info, text=(
            "Isi stok & stok-minim per item. Item yang diisi stoknya otomatis "
            "DILACAK — pesanan melebihi stok akan ditolak.\n"
            "Kosongkan stok untuk menonaktifkan pelacakan item (tanpa blokir)."),
            font=FONT_SMALL, text_color=C_MUTED, justify="left", wraplength=1000).pack(anchor="w", padx=14, pady=8)

        self._stok_tracked_lbl = ctk.CTkLabel(f, text="", font=FONT_SMALL,
                                              text_color="#FFB300", anchor="w")
        self._stok_tracked_lbl.pack(fill="x", padx=16, pady=(2, 0))

        self.scroll_stok = ctk.CTkScrollableFrame(f, fg_color=C_BG)
        self.scroll_stok.pack(fill="both", expand=True, padx=14, pady=10)

        self._stok_entries = {}
        self._build_stok_section()

    def _build_stok_section(self):
        for w in self.scroll_stok.winfo_children():
            w.destroy()
        self._stok_entries = {}
        self._build_stok_section_kat("makanan", "🍔  STOK MAKANAN", self.menu_makanan)
        self._build_stok_section_kat("minuman", "🥤  STOK MINUMAN", self.menu_minuman)
        self._stok_update_badge()

    def _build_stok_section_kat(self, kategori, judul, menu_dict):
        sec = ctk.CTkFrame(self.scroll_stok, fg_color=C_PANEL, corner_radius=12)
        sec.pack(fill="x", pady=8)

        hdr_row = ctk.CTkFrame(sec, fg_color="transparent")
        hdr_row.pack(fill="x", padx=16, pady=(10, 6))
        tk.Label(hdr_row, text=judul, font=FONT_SUB, fg=C_ACCENT2, bg=C_PANEL).pack(side="left")

        kol_row = tk.Frame(sec, bg=C_PANEL)
        kol_row.pack(fill="x", padx=16, pady=(0, 2))
        tk.Label(kol_row, text="Item", font=FONT_SMALL, fg=C_MUTED, bg=C_PANEL,
                 width=24, anchor="w").pack(side="left", padx=(0, 8))
        tk.Label(kol_row, text="Stok Saat Ini", font=FONT_SMALL, fg=C_MUTED, bg=C_PANEL,
                 width=12, anchor="w").pack(side="left", padx=(0, 4))
        tk.Label(kol_row, text="Stok Minim", font=FONT_SMALL, fg=C_MUTED, bg=C_PANEL,
                 width=10, anchor="w").pack(side="left", padx=(0, 4))
        tk.Label(kol_row, text="Aksi", font=FONT_SMALL, fg=C_MUTED, bg=C_PANEL,
                 width=22, anchor="w").pack(side="left")

        container = tk.Frame(sec, bg=C_PANEL)
        container.pack(fill="x", padx=0, pady=(0, 8))

        for nama, harga in menu_dict.items():
            self._tambah_row_stok(container, kategori, nama, harga)

        self._stok_containers = getattr(self, "_stok_containers", {})
        self._stok_containers[kategori] = container

    def _tambah_row_stok(self, container, kategori, nama, harga):
        row = tk.Frame(container, bg=C_PANEL)
        row.pack(fill="x", padx=16, pady=2)

        stok = self._stok_get(nama) or 0
        mini = self._stok_min_get(nama) or 0
        dilacak = self._stok_get(nama) is not None

        var_stok = tk.StringVar(value=str(stok))
        var_min = tk.StringVar(value=str(mini))

        e_nama = tk.Label(row, text=nama, bg=C_PANEL, fg=C_TEXT, font=FONT_SMALL,
                          width=24, anchor="w")
        e_nama.pack(side="left", padx=(0, 8))

        e_stok = tk.Entry(row, textvariable=var_stok, width=10,
                          bg=C_BTN, fg=C_ACCENT, bd=1, font=FONT_SMALL,
                          justify="center")
        e_stok.pack(side="left", padx=(0, 4))
        if dilacak and stok <= 0:
            e_stok.configure(fg="#FFB300")

        e_min = tk.Entry(row, textvariable=var_min, width=8,
                         bg=C_BTN, fg=C_YELLOW, bd=1, font=FONT_SMALL,
                         justify="center")
        e_min.pack(side="left", padx=(0, 4))
        if not dilacak:
            e_stok.configure(fg=C_MUTED)
            e_min.configure(fg=C_MUTED)

        # Aksi: +5, +10, +50, Hapus pelacakan
        for jml in (5, 10, 50):
            tk.Button(row, text=f"+{jml}", width=4, height=1,
                      bg=C_BTN, fg=C_GREEN, font=FONT_SMALL,
                      relief="ridge", bd=1,
                      command=lambda v=var_stok, j=jml: self._stok_tambah_ke_entry(v, j)
                      ).pack(side="left", padx=2)
        tk.Button(row, text="✕ Hapus Stok", width=10, height=1,
                  bg=C_BTN, fg=C_RED, font=FONT_SMALL,
                  relief="ridge", bd=1,
                  command=lambda r=row, k=kategori, n=nama:
                      self._stok_hapus_pelacakan(r, k, n)).pack(side="left", padx=2)
        if dilacak and stok <= 0:
            tk.Label(row, text="⚠ HABIS", bg=C_PANEL, fg=C_RED,
                     font=("Consolas", 9, "bold")).pack(side="left", padx=6)

        self._stok_entries[(kategori, nama)] = {
            "var_stok": var_stok, "var_min": var_min, "nama": nama,
        }

    def _stok_tambah_ke_entry(self, var, jml):
        try:
            cur = int(var.get().strip() or 0)
        except ValueError:
            cur = 0
        var.set(str(cur + jml))

    def _stok_hapus_pelacakan(self, row_widget, kategori, nama):
        if not messagebox.askyesno("Hapus Pelacakan",
                                   f"Berhenti melacak stok '{nama}'?\n"
                                   "Pesanan item ini tidak akan diblokir lagi."):
            return
        self._hapus_stok_item(kategori, nama)
        row_widget.destroy()
        self._stok_entries.pop((kategori, nama), None)
        self._stok_save()
        self._stok_update_badge()
        self._stok_push_bg()

    def _stok_simpan_semua(self):
        """Simpan semua kolom stok/minim ke config + terapkan ke self.stok."""
        baru = 0
        for (kategori, nama), entry in self._stok_entries.items():
            stok_str = entry["var_stok"].get().strip()
            min_str = entry["var_min"].get().strip()
            try:
                stok_v = int(stok_str)
                if stok_v < 0:
                    stok_v = 0
                self._stok_set(nama, stok_v, kategori)
            except ValueError:
                self._hapus_stok_item(kategori, nama)  # kosong = tidak dilacak
            try:
                min_v = int(min_str)
                if min_v < 0:
                    min_v = 0
                self._stok_min_set(nama, min_v, kategori)
            except ValueError:
                self._stok_min_set(nama, 0, kategori)
            if stok_str.strip():
                baru += 1
        self._stok_save()
        self._stok_push_bg()
        self._stok_update_badge()
        self._build_stok_section()
        messagebox.showinfo("✅ Tersimpan",
                            f"Stok berhasil disimpan ({baru} item dilacak).\n"
                            "Halaman QR pelanggan diperbarui otomatis.")

    def _stok_penjualan_per_item(self):
        """Hitung total qty terjual per item dari riwayat (buat info tambahan)."""
        out = {}
        try:
            for m in getattr(self, "riwayat_meta", []) or []:
                pesanan = m.get("pesanan") or {}
                if not isinstance(pesanan, dict):
                    continue
                for nm, q in pesanan.items():
                    out[nm] = out.get(nm, 0) + int(q or 0)
        except Exception:
            pass
        return out


    # ══════════════════════════════════════════════════════════════════════════
    #  TAB 3: Riwayat
    # ══════════════════════════════════════════════════════════════════════════
    def _setup_riwayat(self):
        # (riwayat unchanged UI)
        f = self.frames["riwayat"]
        hdr = ctk.CTkFrame(f, fg_color=C_PANEL, height=54, corner_radius=0)
        hdr.pack(fill="x")
        ctk.CTkLabel(hdr, text="📊  RIWAYAT TRANSAKSI",
                     font=FONT_TITLE, text_color=C_ACCENT).pack(side="left", padx=18, pady=14)

        btn_row = ctk.CTkFrame(hdr, fg_color="transparent")
        btn_row.pack(side="right", padx=18, pady=10)
        ctk.CTkButton(btn_row, text="🔗 rrcctv.online/laporan", width=190, height=36,
                      fg_color="#1A1A4A", hover_color="#0A0A3A",
                      border_width=1, border_color=C_ACCENT,
                      font=("Russo One", 10, "bold"), text_color=C_ACCENT,
                      command=lambda: webbrowser.open("https://rrcctv.online/laporan")).pack(side="left", padx=4)
        ctk.CTkButton(btn_row, text="Export Excel", width=140, height=36,
                      fg_color="#1A4A1A", hover_color="#0A3A0A",
                      border_width=1, border_color=C_GREEN,
                      font=("Russo One", 10, "bold"), text_color=C_GREEN,
                      command=self._export_excel).pack(side="left", padx=4)

        filter_f = ctk.CTkFrame(f, fg_color=C_PANEL, height=64, corner_radius=0)
        filter_f.pack(fill="x")
        filter_f.pack_propagate(False)
        ctk.CTkLabel(filter_f, text="🗓 Tanggal:", font=FONT_SMALL, text_color=C_MUTED).pack(
            side="left", padx=(18, 4), pady=8)
        self.lbl_filter_tanggal = ctk.CTkLabel(filter_f, text="Hari Ini", font=FONT_SMALL,
                                               text_color=C_ACCENT)
        self.lbl_filter_tanggal.pack(side="left", padx=(0, 4))
        ctk.CTkButton(filter_f, text="📅 Kalender", width=96, height=30, fg_color=C_BTN,
                      hover_color=C_ACCENT2, border_width=1, border_color=C_ACCENT,
                      font=("Russo One", 9, "bold"), text_color=C_ACCENT,
                      command=self._buka_kalender).pack(side="left", padx=4)
        ctk.CTkButton(filter_f, text="Hari Ini", width=70, height=30, fg_color=C_BTN,
                      hover_color=C_ACCENT2, font=("Russo One", 9, "bold"), text_color=C_TEXT,
                      command=lambda: self._set_filter_tanggal("hari_ini")).pack(side="left", padx=2)
        ctk.CTkButton(filter_f, text="Kemarin", width=76, height=30, fg_color=C_BTN,
                      hover_color=C_ACCENT2, font=("Russo One", 9, "bold"), text_color=C_TEXT,
                      command=lambda: self._set_filter_tanggal("kemarin")).pack(side="left", padx=2)
        ctk.CTkButton(filter_f, text="7 Hari", width=64, height=30, fg_color=C_BTN,
                      hover_color=C_ACCENT2, font=("Russo One", 9, "bold"), text_color=C_TEXT,
                      command=lambda: self._set_filter_tanggal("7_hari")).pack(side="left", padx=2)
        ctk.CTkButton(filter_f, text="Bulan Ini", width=80, height=30, fg_color=C_BTN,
                      hover_color=C_ACCENT2, font=("Russo One", 9, "bold"), text_color=C_TEXT,
                      command=lambda: self._set_filter_tanggal("bulan_ini")).pack(side="left", padx=2)
        ctk.CTkButton(filter_f, text="Semua", width=66, height=30, fg_color=C_BTN,
                      hover_color=C_ACCENT2, font=("Russo One", 9, "bold"), text_color=C_TEXT,
                      command=lambda: self._set_filter_tanggal(None)).pack(side="left", padx=2)
        ctk.CTkLabel(filter_f, text="👤 Kasir:", font=FONT_SMALL, text_color=C_MUTED).pack(
            side="left", padx=(18, 4))
        kasir_vals = self._riwayat_kasir_choices()
        kasir_default = kasir_vals[0] if kasir_vals else "SEMUA"
        if (self.current_role or "") == "kasir":
            kasir_default = self.current_user or "SEMUA"
        self.opt_filter_kasir = ctk.CTkOptionMenu(filter_f, values=kasir_vals,
                                                  variable=ctk.StringVar(value=kasir_default),
                                                  fg_color=C_BTN, button_color=C_ACCENT2,
                                                  button_hover_color="#5A0FCC",
                                                  text_color=C_TEXT,
                                                  dropdown_fg_color=C_CARD,
                                                  dropdown_text_color=C_TEXT,
                                                  width=150, font=("Consolas", 10),
                                                  command=lambda _v: self._apply_riwayat_filter())
        self.opt_filter_kasir.pack(side="left", padx=(0, 8))
        if (self.current_role or "") == "kasir":
            self.opt_filter_kasir.configure(state="disabled")
        self._riwayat_filter_kasir = kasir_default

        self._riwayat_filter_status = getattr(self, "_riwayat_filter_status", "SEMUA") or "SEMUA"
        ctk.CTkLabel(filter_f, text="💳 Status:", font=FONT_SMALL, text_color=C_MUTED).pack(
            side="left", padx=(14, 4))
        self.opt_filter_status = ctk.CTkOptionMenu(filter_f, values=["SEMUA", "LUNAS", "BELUM"],
                                                   variable=ctk.StringVar(value=self._riwayat_filter_status),
                                                   fg_color=C_BTN, button_color=C_ACCENT2,
                                                   button_hover_color="#5A0FCC",
                                                   text_color=C_TEXT,
                                                   dropdown_fg_color=C_CARD,
                                                   dropdown_text_color=C_TEXT,
                                                   width=110, font=("Consolas", 10),
                                                   command=lambda _v: self._apply_riwayat_filter())
        self.opt_filter_status.pack(side="left", padx=(0, 8))
        self._riwayat_filter_cari = ""
        ctk.CTkLabel(filter_f, text="🔍 Cari:", font=FONT_SMALL, text_color=C_MUTED).pack(
            side="left", padx=(10, 4))
        self.ent_filter_cari = ctk.CTkEntry(filter_f, width=190, height=30,
                                            fg_color=C_BTN, text_color=C_TEXT,
                                            border_color=C_BORDER, font=("Consolas", 10),
                                            placeholder_text="TV/PC, kasir, paket…")
        self.ent_filter_cari.pack(side="left", padx=(0, 8))
        self.ent_filter_cari.bind("<KeyRelease>", self._on_cari_ketik)

        rekap_f = ctk.CTkFrame(f, fg_color=C_PANEL, height=44, corner_radius=0)
        rekap_f.pack(fill="x")
        rekap_f.pack_propagate(False)
        self.lbl_rekap = ctk.CTkLabel(rekap_f, text="Total Transaksi: 0  |  Total Pendapatan: Rp 0",
                                       font=FONT_SUB, text_color=C_YELLOW)
        self.lbl_rekap.pack(side="left", padx=18, pady=10)
        # Action buttons on the right of rekap_f
        act_f = ctk.CTkFrame(rekap_f, fg_color="transparent")
        act_f.pack(side="right", padx=10)
        self.btn_toggle_paid = ctk.CTkButton(act_f, text="✅ Tandai Lunas", width=130, height=32,
                                              fg_color=C_BTN, hover_color=C_ACCENT2,
                                              border_width=1, border_color=C_GREEN,
                                              font=("Russo One", 9, "bold"), text_color=C_GREEN,
                                              command=self._toggle_paid_status)
        self.btn_toggle_paid.pack(side="left", padx=4)
        self.btn_invoice = ctk.CTkButton(act_f, text="🧾 Invoice", width=110, height=32,
                                          fg_color=C_BTN, hover_color=C_ACCENT2,
                                          border_width=1, border_color=C_ACCENT2,
                                          font=("Russo One", 9, "bold"), text_color=C_ACCENT2,
                                          command=self._show_invoice_dialog)
        self.btn_invoice.pack(side="left", padx=4)
        self.btn_bukti = ctk.CTkButton(act_f, text="📎 Lihat Bukti", width=130, height=32,
                                        fg_color=C_BTN, hover_color=C_ACCENT2,
                                        border_width=1, border_color=C_YELLOW,
                                        font=("Russo One", 9, "bold"), text_color=C_YELLOW,
                                        command=self._show_bukti_pembayaran)
        self.btn_bukti.pack(side="left", padx=4)

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Game.Treeview",
                        background=C_CARD, fieldbackground=C_CARD,
                        foreground=C_TEXT, rowheight=38,
                        font=("Consolas", 14))
        style.configure("Game.Treeview.Heading",
                        background=C_PANEL, foreground=C_ACCENT,
                        font=("Russo One", 12, "bold"), relief="flat")
        style.map("Game.Treeview", background=[("selected", C_ACCENT2)])
        # Treeview tanpa style eksplisit (mis. dialog riwayat QR) — ikut
        # diperbesar supaya terbaca di layar PC resolusi tinggi.
        style.configure("Treeview",
                        background=C_CARD, fieldbackground=C_CARD,
                        foreground=C_TEXT, rowheight=32,
                        font=("Consolas", 12))
        style.configure("Treeview.Heading",
                        background=C_PANEL, foreground=C_ACCENT,
                        font=("Russo One", 11, "bold"), relief="flat")

        cols = ("Waktu", "Kasir", "TV/PC", "Paket", "Pesanan", "Diskon", "Total", "Status")
        self.tree = ttk.Treeview(f, columns=cols, show="headings", style="Game.Treeview")
        widths = [140, 80, 100, 140, 160, 80, 110, 110]
        for col, w in zip(cols, widths):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=w, anchor="center" if w < 150 else "w")
        self.tree.pack(fill="both", expand=True, padx=14, pady=(10, 4))
        # Tag for row coloring
        self.tree.tag_configure("paid", foreground=C_GREEN)
        self.tree.tag_configure("unpaid", foreground=C_RED)
        # Right-click: admin-only action menu (Hapus transaksi)
        self.tree.bind("<Button-3>", self._on_riwayat_right_click)
        self.tree.bind("<<TreeviewSelect>>", self._on_riwayat_select)

        footer_f = ctk.CTkFrame(f, fg_color=C_PANEL, height=40, corner_radius=0)
        footer_f.pack(fill="x", padx=14, pady=(0, 12))
        footer_f.pack_propagate(False)
        self.lbl_rekap_footer = ctk.CTkLabel(footer_f,
                                            text="Total Transaksi: 0  |  Total Pendapatan: Rp 0",
                                            font=FONT_SUB, text_color=C_YELLOW)
        self.lbl_rekap_footer.pack(side="left", padx=18, pady=8)

    def _on_riwayat_select(self, event):
        sel = self.tree.selection()
        if not sel:
            return
        iid = sel[0]
        idx = self._tree_item_to_index.get(iid, -1)
        paid = self.riwayat_meta[idx].get('paid', True) if 0 <= idx < len(self.riwayat_meta) else True
        self.btn_toggle_paid.configure(text="⏳ Tandai Belum Lunas" if paid else "✅ Tandai Lunas")
        has_bukti = False
        if 0 <= idx < len(self.riwayat_meta):
            has_bukti = bool(self.riwayat_meta[idx].get("bukti_pembayaran"))
        self.btn_bukti.configure(state="normal" if has_bukti else "disabled")

    def _toggle_paid_status(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Pilih Transaksi", "Silakan pilih transaksi dari daftar.", parent=self)
            return
        iid = sel[0]
        idx = self._tree_item_to_index.get(iid, -1)
        if idx < 0 or idx >= len(self.riwayat_meta):
            return
        current = self.riwayat_meta[idx].get('paid', True)
        self.riwayat_meta[idx]['paid'] = not current
        paid = not current
        # Update tree row
        status_str = "✅ Lunas" if paid else "⏳ Belum Lunas"
        try:
            row = list(self.riwayat_transaksi[idx])
            if len(row) >= 8:
                row[7] = status_str
            else:
                while len(row) < 8:
                    row.append("—")
                row[7] = status_str
            self.riwayat_transaksi[idx] = tuple(row)
        except Exception:
            pass
        vals = list(self.tree.item(iid, "values"))
        if len(vals) >= 8:
            vals[7] = status_str
        else:
            while len(vals) < 8:
                vals.append("—")
            vals[7] = status_str
        self.tree.item(iid, values=vals, tags=("paid" if paid else "unpaid",))
        self.btn_toggle_paid.configure(text="⏳ Tandai Belum Lunas" if paid else "✅ Tandai Lunas")
        self._refresh_riwayat_summary()
        self._save_riwayat()
        threading.Thread(target=self._upsert_tx_cloud_from_index, args=(idx,), daemon=True).start()
        AuditLogger.log(action="payment_toggle", username=self.current_user or "", status="success",
                        details={"transaction_index": idx, "new_status": "paid" if paid else "unpaid"})

    def _set_transaksi_paid(self, item_id, paid):
        """Sinkron status bayar dari kartu TV/PC ke baris riwayat terkait."""
        if not item_id:
            return
        idx = self._tree_item_to_index.get(item_id, -1)
        if idx < 0 or idx >= len(self.riwayat_meta):
            return
        self._set_transaksi_paid_idx(idx, paid)

    def _show_invoice_dialog(self, iid=None):
        if not iid:
            sel = self.tree.selection()
            if not sel:
                messagebox.showwarning("Pilih Transaksi", "Silakan pilih transaksi dari daftar.", parent=self)
                return
            iid = sel[0]
        idx = self._tree_item_to_index.get(iid, -1)
        if idx < 0 or idx >= len(self.riwayat_transaksi):
            return
        row = self.riwayat_transaksi[idx]
        meta = self.riwayat_meta[idx] if idx < len(self.riwayat_meta) else {}
        paid = meta.get('paid', True)
        status_str = "✅ LUNAS" if paid else "⏳ BELUM LUNAS"
        waktu = row[0] if len(row) > 0 else ""
        kasir = row[1] if len(row) > 1 else ""
        tv_pc = row[2] if len(row) > 2 else ""
        paket = row[3] if len(row) > 3 else ""
        pesanan = row[4] if len(row) > 4 else ""
        diskon = row[5] if len(row) > 5 else ""
        total = row[6] if len(row) > 6 else ""
        invoice_no = f"INV-{datetime.now().strftime('%Y%m%d')}-{idx+1:04d}"
        dlg = ctk.CTkToplevel(self)
        dlg.title(f"🧾 Invoice — {invoice_no}")
        dlg.geometry("520x620")
        dlg.resizable(False, False)
        dlg.transient(self)
        dlg.grab_set()
        ctk.CTkLabel(dlg, text=f"🧾  INVOICE", font=FONT_TITLE, text_color=C_ACCENT).pack(pady=(16, 4))
        ctk.CTkLabel(dlg, text=f"{invoice_no}", font=("Russo One", 10), text_color=C_MUTED).pack()
        sep = ctk.CTkFrame(dlg, height=1, fg_color=C_BORDER)
        sep.pack(fill="x", padx=30, pady=8)
        info = [
            ("Waktu", waktu),
            ("Kasir", kasir),
            ("TV/PC", tv_pc),
            ("Paket", paket),
            ("Pesanan", pesanan),
            ("Diskon", diskon),
            ("Total", total),
            ("Status", status_str),
        ]
        for label, val in info:
            r = ctk.CTkFrame(dlg, fg_color="transparent")
            r.pack(fill="x", padx=40, pady=2)
            ctk.CTkLabel(r, text=f"{label}:", font=FONT_LABEL, text_color=C_MUTED, width=90, anchor="w").pack(side="left")
            ctk.CTkLabel(r, text=val, font=FONT_BODY, text_color=C_TEXT, anchor="w").pack(side="left", padx=8)
        sep2 = ctk.CTkFrame(dlg, height=1, fg_color=C_BORDER)
        sep2.pack(fill="x", padx=30, pady=8)
        btn_f = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_f.pack(fill="x", padx=30, pady=(12, 16))
        ctk.CTkButton(btn_f, text="🖨 Cetak", width=140, height=36,
                       fg_color=C_ACCENT2, font=("Russo One", 10, "bold"),
                       command=lambda: self._cetak_invoice(row, meta, invoice_no)).pack(side="left", padx=4)
        ctk.CTkButton(btn_f, text="✖ Tutup", width=100, height=36,
                       fg_color=C_RED, font=("Russo One", 10, "bold"),
                       command=dlg.destroy).pack(side="right", padx=4)

    def _cetak_invoice(self, row, meta, invoice_no):
        waktu = row[0] if len(row) > 0 else ""
        kasir = row[1] if len(row) > 1 else ""
        tv_pc = row[2] if len(row) > 2 else ""
        paket = row[3] if len(row) > 3 else ""
        pesanan = row[4] if len(row) > 4 else ""
        diskon = row[5] if len(row) > 5 else ""
        total = row[6] if len(row) > 6 else ""
        paid = meta.get('paid', True)
        status_str = "LUNAS" if paid else "BELUM LUNAS"
        lines = [
            "=" * 40,
            "      RR BILLING PRO",
            f"      INVOICE {invoice_no}",
            "=" * 40,
            f"Waktu   : {waktu}",
            f"Kasir   : {kasir}",
            f"TV/PC   : {tv_pc}",
            f"Paket   : {paket}",
            f"Pesanan : {pesanan}",
            f"Diskon  : {diskon}",
            f"Total   : {total}",
            f"Status  : {status_str}",
            "-" * 40,
            f"Dicetak : {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
            "=" * 40,
        ]
        text = "\n".join(lines)
        app = self.winfo_toplevel()
        try:
            tersedia, ptype, addr = app._printer_tersedia()
        except Exception:
            tersedia, ptype, addr = False, "file", ""
        if not tersedia:
            if messagebox.askyesno(
                    "Printer Belum Terhubung",
                    "Belum ada printer terhubung.\n\n"
                    "Atur dulu di Pengaturan → PRINTER SETTINGS:\n"
                    "• Tipe: bluetooth (atau usb/network)\n"
                    "• Alamat printer, atau klik tombol '🔍 Cari Printer BT'\n"
                    "• Lalu Test Print untuk memastikan.\n\n"
                    "Buka Pengaturan Printer sekarang?",
                    parent=self):
                try:
                    app._show_tab("profil")
                except Exception:
                    pass
            return
        threading.Thread(target=app._print_via_escpos,
                         args=(text, ptype, addr), daemon=True).start()
        messagebox.showinfo("Cetak", "Invoice dikirim ke printer.")

    def _show_bukti_pembayaran(self, iid=None):
        if not iid:
            sel = self.tree.selection()
            if not sel:
                messagebox.showwarning("Pilih Transaksi", "Silakan pilih transaksi dari daftar.", parent=self)
                return
            iid = sel[0]
        idx = self._tree_item_to_index.get(iid, -1)
        if idx < 0 or idx >= len(self.riwayat_meta):
            return
        bukti_path = self.riwayat_meta[idx].get("bukti_pembayaran", "")
        if not bukti_path or not os.path.exists(bukti_path):
            messagebox.showinfo("Tidak Ada", "Tidak ada bukti pembayaran untuk transaksi ini.", parent=self)
            return
        try:
            os.startfile(bukti_path)
        except Exception as e:
            messagebox.showerror("Gagal", f"Gagal membuka bukti:\n{e}", parent=self)

    def _format_riwayat_row(self, waktu, tv_label, paket_nama, pesanan_dict, total_int, pesanan_total=None, diskoni=0, diskoni_mode="nominal", paid=True, kasir=None, booking_id=None):
        """Build a properly formatted riwayat row tuple with price annotations."""
        if pesanan_total is None:
            all_menu = {**self.menu_makanan, **self.menu_minuman}
            pesanan_total = sum(all_menu.get(nm, 0) * qty for nm, qty in (pesanan_dict.items() if isinstance(pesanan_dict, dict) else []))
        paket_harga = total_int - pesanan_total
        if paket_harga < 0:
            paket_harga = 0
        paket_tampil = f"{paket_nama} ({fmt_rp(paket_harga)})" if paket_harga > 0 else paket_nama
        if booking_id:
            paket_tampil = f"{paket_tampil.rstrip()} 📅{str(booking_id)[:8].upper()}"
        pesanan_str = ", ".join(f"{nm}×{qty}" for nm, qty in (pesanan_dict.items() if isinstance(pesanan_dict, dict) else [])) or "—"
        if pesanan_str != "—":
            pesanan_tampil = f"{pesanan_str} ({fmt_rp(pesanan_total)})"
        else:
            pesanan_tampil = "—"
        if diskoni > 0:
            if diskoni_mode == "persen":
                diskon_tampil = f"{diskoni}%"
            else:
                diskon_tampil = fmt_rp(diskoni)
        else:
            diskon_tampil = "—"
        status_str = "✅ Lunas" if paid else "⏳ Belum Lunas"
        return (waktu, kasir or self.current_user, tv_label, paket_tampil, pesanan_tampil, diskon_tampil, fmt_rp(total_int), status_str)

    def _save_riwayat(self):
        """Simpan riwayat transaksi ke file JSON agar tidak hilang saat restart."""
        try:
            data = {
                "riwayat_transaksi": [list(r) for r in self.riwayat_transaksi],
                "riwayat_meta": self.riwayat_meta,
            }
            with open(RIWAYAT_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _load_riwayat(self):
        """Load riwayat transaksi dari file JSON setelah login/restart."""
        # Default filter: pendapatan harian (hari ini); kasir hanya melihat transaksinya sendiri
        self._riwayat_filter_tanggal = datetime.now().strftime("%Y-%m-%d")
        if (self.current_role or "") == "admin":
            self._riwayat_filter_kasir = "SEMUA"
        else:
            self._riwayat_filter_kasir = self.current_user or "SEMUA"
        if hasattr(self, "lbl_filter_tanggal"):
            self.lbl_filter_tanggal.configure(
                text=f"Hari Ini ({self._riwayat_filter_tanggal})")
        if hasattr(self, "opt_filter_kasir"):
            try:
                self.opt_filter_kasir.set(self._riwayat_filter_kasir)
            except Exception:
                pass
        if not os.path.exists(RIWAYAT_FILE):
            return
        try:
            with open(RIWAYAT_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            rows = data.get("riwayat_transaksi", [])
            metas = data.get("riwayat_meta", [])
            if not rows:
                return
            # Pad rows to 8 columns for backward compat, sync status from meta
            padded = []
            for idx, r in enumerate(rows):
                rlist = list(r)
                while len(rlist) < 7:
                    rlist.append("—")
                paid = metas[idx].get('paid', True) if idx < len(metas) else True
                status_str = "✅ Lunas" if paid else "⏳ Belum Lunas"
                if len(rlist) >= 8:
                    rlist[7] = status_str
                else:
                    while len(rlist) < 8:
                        rlist.append("—")
                    rlist[7] = status_str
                padded.append(tuple(rlist))
            # Muat SEMUA data ke memori (utuh, untuk keperluan saving)
            self.riwayat_transaksi = padded
            self.riwayat_meta = metas
            # Beri cloud_id stabil ke baris lama agar bisa diupload tanpa duplikat
            try:
                self._backfill_cloud_ids()
            except Exception:
                pass
            # Tampilkan sesuai filter aktif (tanggal + kasir)
            self._render_riwayat_tree()
        except Exception:
            pass

    def _catat_transaksi(self, tv_label, paket, pesanan, total, source='tv', diskoni=0, diskoni_mode="nominal", paid=True, booking_meta=None):
        """Catat transaksi ke riwayat.
        pesanan: dict nama->qty
        tv_label: string label TV/PC
        source: 'tv' atau 'warnet'
        total: int (total rupiah)
        paid: status pembayaran (True=LUNAS, False=TAGIHAN)
        booking_meta: dict (opsional) keterangan booking online — booking_id,
            kode, metode, status_bayar, pelanggan, no_hp
        """
        waktu       = datetime.now().strftime("%Y-%m-%d %H:%M")
        pesanan = pesanan or {}
        src = source if source in ('tv', 'warnet') else 'tv'
        booking_meta = booking_meta or {}
        booking_id = str(booking_meta.get("booking_id") or "")

        # compute pesanan total (money) using menu prices
        all_menu = {**self.menu_makanan, **self.menu_minuman}
        pesanan_total = sum(all_menu.get(nm, 0) * qty for nm, qty in (pesanan.items() if isinstance(pesanan, dict) else []))

        # ensure total is int
        try:
            total_int = int(total)
        except Exception:
            try:
                total_int = int(str(total).replace('Rp ', '').replace('.', '').strip())
            except Exception:
                total_int = pesanan_total

        paket_harga = total_int - pesanan_total
        if paket_harga < 0:
            paket_harga = 0

        row = self._format_riwayat_row(waktu, tv_label, paket, pesanan, total_int, pesanan_total, diskoni, diskoni_mode, paid=paid, kasir=self.current_user, booking_id=booking_id)
        self.riwayat_transaksi.append(row)
        # maintain parallel meta + info upload cloud
        cloud_id = f"tx_{int(time.time() * 1000)}"
        self.riwayat_meta.append({'source': src, 'paket_harga': paket_harga, 'pesanan_total': pesanan_total, 'total': total_int, 'diskoni': diskoni, 'diskoni_mode': diskoni_mode, 'paid': paid, 'cloud_id': cloud_id, 'paket_raw': str(paket or ""), 'pesanan': {str(k): int(v) for k, v in (pesanan.items() if isinstance(pesanan, dict) else [])}})
        if booking_id:
            self.riwayat_meta[-1]['booking_id'] = booking_id
            self.riwayat_meta[-1]['booking_kode'] = str(booking_meta.get("kode") or "")
            self.riwayat_meta[-1]['booking_metode'] = str(booking_meta.get("metode") or "")
            self.riwayat_meta[-1]['booking_status_bayar'] = str(booking_meta.get("status_bayar") or "")
            self.riwayat_meta[-1]['pelanggan'] = str(booking_meta.get("pelanggan") or "")
            self.riwayat_meta[-1]['no_hp'] = str(booking_meta.get("no_hp") or "")

        # Tampilkan hanya bila lolos filter aktif
        item_id = None
        if self._row_matches_filter(row):
            item_id = self.tree.insert("", 0, values=row, tags=("paid" if paid else "unpaid",))
            self._tree_item_to_index[item_id] = len(self.riwayat_transaksi) - 1

        # Upload ke Firestore (billingps_users/{admin_utama}.transaksiList) agar
        # admin bisa memantau transaksi desktop dari HP (APTV2: RiwayatScreen).
        try:
            tx_cloud = self._build_tx_cloud(row, self.riwayat_meta[-1])
            if tx_cloud:
                self._pending_tx_uploads.append(tx_cloud)
                threading.Thread(target=self._flush_cloud_uploads, daemon=True).start()
        except Exception as e:
            _LOGGER.warning("Gagal menyiapkan upload transaksi: %s", e)

        # refresh summaries
        if hasattr(self, '_refresh_riwayat_summary'):
            self._refresh_riwayat_summary()
        if hasattr(self, '_refresh_warnet_footer'):
            self._refresh_warnet_footer()
        if hasattr(self, '_refresh_dashboard_total_pesanan'):
            self._refresh_dashboard_total_pesanan()

        self._save_riwayat()
        self._last_catat_idx = len(self.riwayat_transaksi) - 1
        self._last_catat_cloud_id = cloud_id
        return item_id

    def _bersihkan_riwayat(self):
        # Hanya admin yang boleh membersihkan seluruh riwayat
        if self.current_role != "admin":
            messagebox.showwarning("Akses Ditolak", "Hanya admin yang boleh membersihkan riwayat.")
            return
        if messagebox.askyesno("Bersihkan Riwayat", "Hapus semua data riwayat dari tampilan?"):
            self.riwayat_transaksi.clear()
            self.riwayat_meta.clear()
            self._tree_item_to_index.clear()
            self._pending_tx_uploads = []
            for item in self.tree.get_children():
                self.tree.delete(item)
            summary_text = "Total Transaksi: 0  |  Total Pendapatan: Rp 0"
            self.lbl_rekap.configure(text=summary_text)
            if hasattr(self, 'lbl_rekap_footer'):
                self.lbl_rekap_footer.configure(text=summary_text)
            self._save_riwayat()

    def _on_riwayat_right_click(self, event):
        # Dapatkan item di bawah kursor
        iid = self.tree.identify_row(event.y)
        if not iid:
            return
        values = self.tree.item(iid, "values")
        idx = self._tree_item_to_index.get(iid, -1)
        # Hanya tampilkan opsi hapus jika user adalah admin
        menu = tk.Menu(self.tree, tearoff=0)
        menu.add_command(label="🧾 Invoice", command=lambda i=iid: self._show_invoice_dialog(i))
        if idx >= 0 and idx < len(self.riwayat_meta) and self.riwayat_meta[idx].get("booking_id"):
            menu.add_command(label="🔗 Lihat Booking",
                             command=lambda ix=idx: self._buka_booking_dari_riwayat(ix))
        if idx >= 0 and idx < len(self.riwayat_meta) and self.riwayat_meta[idx].get("bukti_pembayaran"):
            menu.add_command(label="📎 Lihat Bukti", command=lambda i=iid: self._show_bukti_pembayaran(i))
        if self.current_role == "admin":
            menu.add_separator()
            menu.add_command(label="🗑 Hapus Transaksi (Admin)", command=lambda i=iid: self._ask_admin_and_delete(i))
        else:
            menu.add_command(label="🔒 Akses terbatas", command=lambda: messagebox.showinfo("Akses", "Hanya admin yang dapat menghapus transaksi."))
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    def _rebuild_tree_item_index(self):
        # Rebuild mapping item_id -> riwayat_transaksi index (best-effort by matching row values)
        self._tree_item_to_index.clear()
        items = list(self.tree.get_children())
        for item_id in items:
            vals = tuple(self.tree.item(item_id, 'values'))
            # find matching index in riwayat_transaksi
            idx = None
            for i, r in enumerate(self.riwayat_transaksi):
                if tuple(r) == vals:
                    idx = i
                    break
            if idx is None:
                # fallback: use position-based mapping (assume newest at 0)
                idx = max(0, len(self.riwayat_transaksi) - 1 - items.index(item_id))
            self._tree_item_to_index[item_id] = idx
        self._remap_kartu_transaction_items()

    def _riwayat_idx_by_booking(self, booking_id):
        """Index baris riwayat yang berasal dari booking online tertentu (-1
        bila tidak ada). Basis relasi baris kas ↔ booking."""
        if not booking_id:
            return -1
        bid = str(booking_id)
        try:
            for i, meta in enumerate(self.riwayat_meta):
                if isinstance(meta, dict) and str(meta.get("booking_id") or "") == bid:
                    return i
        except Exception:
            pass
        return -1

    def _resolve_session_idx(self, kartu):
        """Index riwayat milik kartu TV/PC: item-map -> idx tersimpan -> cloud_id."""
        try:
            item_id = getattr(kartu, '_last_transaction_item', None)
            if item_id is not None and hasattr(self.tree, 'exists') and self.tree.exists(item_id):
                idx = self._tree_item_to_index.get(item_id, -1)
                if idx >= 0:
                    return idx
            idx = getattr(kartu, '_last_riwayat_idx', None)
            if idx is not None and 0 <= idx < len(self.riwayat_meta):
                return idx
            cid = getattr(kartu, '_last_cloud_id', None)
            if cid:
                for i, meta in enumerate(self.riwayat_meta):
                    if isinstance(meta, dict) and meta.get("cloud_id") == cid:
                        return i
        except Exception:
            pass
        return -1

    def _tree_item_for_idx(self, idx):
        """Cari item id tree yang memetakan ke index riwayat (None jika tak tampil)."""
        for item_id, i in self._tree_item_to_index.items():
            if i == idx:
                return item_id
        return None

    def _remap_kartu_transaction_items(self):
        """Relink kartu ke item tree setelah tree di-render ulang (item id hangus)."""
        try:
            if not hasattr(self, "tree") or not hasattr(self.tree, "exists"):
                return
            item_by_cloud = {}
            for item_id, idx in self._tree_item_to_index.items():
                if 0 <= idx < len(self.riwayat_meta):
                    cid = self.riwayat_meta[idx].get("cloud_id")
                    if cid:
                        item_by_cloud[cid] = item_id
            for kartu in list(getattr(self, "_semua_kartu_tv", []) or []) + list(getattr(self, "_semua_kartu_warnet", []) or []):
                item_id = getattr(kartu, "_last_transaction_item", None)
                if item_id is not None and self.tree.exists(item_id) and self._tree_item_to_index.get(item_id) is not None:
                    continue
                cid = getattr(kartu, "_last_cloud_id", None)
                if cid and cid in item_by_cloud:
                    kartu._last_transaction_item = item_by_cloud[cid]
                else:
                    kartu._last_transaction_item = None
        except Exception:
            pass

    def _set_transaksi_paid_idx(self, idx, paid):
        """Sinkron status bayar ke baris riwayat terkait (berdasarkan index)."""
        if idx is None or not (0 <= idx < len(self.riwayat_meta)):
            return
        self.riwayat_meta[idx]['paid'] = bool(paid)
        status_str = "✅ Lunas" if paid else "⏳ Belum Lunas"
        try:
            row = list(self.riwayat_transaksi[idx])
            if len(row) >= 8:
                row[7] = status_str
            else:
                while len(row) < 8:
                    row.append("—")
                row[7] = status_str
            self.riwayat_transaksi[idx] = tuple(row)
        except Exception:
            pass
        item_id = self._tree_item_for_idx(idx)
        if item_id is not None and self.tree.exists(item_id):
            self.tree.item(item_id, values=self.riwayat_transaksi[idx],
                           tags=("paid" if paid else "unpaid",))
        if item_id is not None and item_id in self.tree.selection() and hasattr(self, 'btn_toggle_paid'):
            self.btn_toggle_paid.configure(
                text="⏳ Tandai Belum Lunas" if paid else "✅ Tandai Lunas")
        self._refresh_riwayat_summary()
        self._save_riwayat()
        threading.Thread(target=self._upsert_tx_cloud_from_index, args=(idx,), daemon=True).start()
        AuditLogger.log(action="payment_toggle_card", username=self.current_user or "", status="success",
                        details={"transaction_index": idx, "new_status": "paid" if paid else "unpaid"})

    def _ask_admin_and_delete(self, iid):
        # Dialog kecil untuk meminta username + password admin untuk otorisasi
        dlg = ctk.CTkToplevel(self)
        dlg.title("Otorisasi Admin")
        dlg.geometry("520x320")
        dlg.resizable(False, False)
        dlg.transient(self)
        dlg.grab_set()
        ctk.CTkLabel(dlg, text="Masukkan akun admin untuk menghapus transaksi", font=FONT_BODY, text_color=C_MUTED).pack(pady=(16,12), padx=16)
        entry_user = ctk.CTkEntry(dlg, placeholder_text="username admin", fg_color=C_BTN, text_color=C_TEXT)
        entry_user.pack(padx=24, pady=(4,10), fill="x")
        entry_pass = ctk.CTkEntry(dlg, placeholder_text="password admin", show="●", fg_color=C_BTN, text_color=C_TEXT)
        entry_pass.pack(padx=24, pady=(0,12), fill="x")
        status = ctk.CTkLabel(dlg, text="", text_color=C_RED)
        status.pack(padx=16, pady=(0,10))

        def submit():
            uname = entry_user.get().strip()
            pwd = entry_pass.get().strip()
            users = ConfigManager.get("users", LoginPage.DEFAULT_USERS)
            u = users.get(uname)
            if not u or u.get("role") != "admin":
                status.configure(text="✖ Akun tidak ditemukan atau bukan admin.")
                AuditLogger.log(action="authorize_delete_failed", username=uname, status="not_admin")
                return
            if not verify_password(pwd, u.get("password_enc", u.get("password", ""))):
                status.configure(text="✖ Password salah.")
                AuditLogger.log(action="authorize_delete_failed", username=uname, status="bad_password")
                return
            # authorized
            row_values = tuple(self.tree.item(iid, "values"))
            # find index in riwayat_transaksi
            idx = None
            for i, r in enumerate(self.riwayat_transaksi):
                if tuple(r) == row_values:
                    idx = i
                    break
            if idx is not None:
                try:
                    self.riwayat_transaksi.pop(idx)
                except Exception:
                    pass
                try:
                    self.riwayat_meta.pop(idx)
                except Exception:
                    pass
            else:
                # fallback: try to remove by value
                try:
                    self.riwayat_transaksi.remove(list(row_values))
                except Exception:
                    pass
            self.tree.delete(iid)
            AuditLogger.log(action="transaction_deleted", username=uname, status="success", details={"row": row_values})
            dlg.destroy()
            # rebuild index mapping and refresh rekap
            self._rebuild_tree_item_index()
            if hasattr(self, '_refresh_riwayat_summary'):
                self._refresh_riwayat_summary()
            if hasattr(self, 'lbl_rekap_footer'):
                self.lbl_rekap_footer.configure(text=self.lbl_rekap.cget('text'))
            self._save_riwayat()

        btn_frame = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_frame.pack(fill="x", padx=16, pady=(8, 16))
        ctk.CTkButton(btn_frame, text="✖ Batal", width=120, height=36,
                      fg_color=C_RED, hover_color="#7A1A1A", command=dlg.destroy).pack(side="left")
        ctk.CTkButton(btn_frame, text="🔒 Authorize & Hapus", width=180, height=36,
                      fg_color=C_ACCENT2, hover_color="#5A0FCC", command=submit).pack(side="right")


    def _import_riwayat_from_cloud(self, silent=False):
        if not self.current_user:
            if not silent:
                messagebox.showwarning("Login", "Silakan login terlebih dahulu.", parent=self)
            return
        try:
            fc = FirestoreClient()
            tx_list = fc.fetch_transactions(self._cloud_upload_target(), max_days=6)
        except Exception as e:
            if not silent:
                messagebox.showerror("Gagal", f"Gagal mengambil data dari cloud:\n{e}", parent=self)
            return
        if not tx_list:
            if not silent:
                messagebox.showinfo("Import Cloud", "Tidak ada data transaksi dari cloud (6 hari terakhir).", parent=self)
            return
        existing_ids = {m.get('cloud_id') for m in self.riwayat_meta if m.get('cloud_id')}
        existing_keys = {(r[0], r[2], r[3]) for r in self.riwayat_transaksi}
        imported = 0
        for tx in reversed(tx_list):
            if not isinstance(tx, dict):
                continue
            waktu = tx.get("waktu") or tx.get("timestamp") or ""
            label = tx.get("kota") or tx.get("tv_label") or tx.get("label") or ""
            paket_raw = tx.get("paket") or ""
            total = tx.get("total", 0)
            kasir = tx.get("kasir") or self.current_user or ""
            cloud_id = tx.get("id") or ""
            if not waktu or not label:
                continue
            key = (waktu, label, paket_raw)
            if cloud_id and cloud_id in existing_ids:
                continue
            if not cloud_id and key in existing_keys:
                continue
            try:
                total_int = int(total)
            except Exception:
                total_int = 0
            pesanan_raw = tx.get("pesanan") or {}
            pesanan = {}
            if isinstance(pesanan_raw, dict):
                for k, v in pesanan_raw.items():
                    try:
                        pesanan[str(k)] = int(v)
                    except Exception:
                        pesanan[str(k)] = 0
            pesanan_total = 0
            ph_raw = tx.get("pesananHarga") or {}
            if isinstance(ph_raw, dict):
                for v in ph_raw.values():
                    try:
                        pesanan_total += int(v)
                    except Exception:
                        pass
            paket_harga = tx.get("paketHarga", 0)
            try:
                paket_harga = int(paket_harga)
            except Exception:
                paket_harga = 0
            if paket_harga <= 0:
                paket_harga = max(0, total_int - pesanan_total)
            row = self._format_riwayat_row(waktu, label, paket_raw, pesanan, total_int,
                                           pesanan_total, kasir=kasir)
            self.riwayat_transaksi.append(row)
            self.riwayat_meta.append({
                'source': tx.get("source", "cloud"),
                'paket_harga': paket_harga,
                'pesanan_total': pesanan_total,
                'total': total_int,
                'paid': True,
                'cloud_id': cloud_id or "",
                'paket_raw': paket_raw,
            })
            imported += 1
        if imported > 0:
            self._save_riwayat()
            try:
                self.after(0, self._render_riwayat_tree)
            except Exception:
                pass
            if not silent:
                messagebox.showinfo("Import Cloud", f"{imported} transaksi berhasil diimport dari cloud (6 hr terakhir).", parent=self)
        elif not silent:
            messagebox.showinfo("Import Cloud", "Semua data dari cloud sudah ada di riwayat lokal.", parent=self)

    def _export_excel(self):
        rows_to_export = [self.riwayat_transaksi[i] for i in self._riwayat_filter_indices()]
        if not rows_to_export:
            messagebox.showwarning("Kosong", "Tidak ada data transaksi sesuai filter untuk diekspor.", parent=self)
            return

        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = "Riwayat Transaksi"

        header_fill = PatternFill("solid", fgColor="1A1A3A")
        header_font = Font(name="Consolas", bold=True, color="00FFCC", size=11)
        accent_fill = PatternFill("solid", fgColor="12122A")
        normal_font = Font(name="Consolas", size=10)
        center      = Alignment(horizontal="center", vertical="center")
        border      = Border(
            left=Side(style="thin", color="2A2A5A"),
            right=Side(style="thin", color="2A2A5A"),
            top=Side(style="thin", color="2A2A5A"),
            bottom=Side(style="thin", color="2A2A5A"),
        )

        ws.merge_cells("A1:G1")
        ws["A1"] = "LAPORAN TRANSAKSI — RR BILLING PRO"
        ws["A1"].font      = Font(name="Consolas", bold=True, color="00FFCC", size=14)
        ws["A1"].fill      = PatternFill("solid", fgColor="0D0D1A")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 32

        ws.merge_cells("A2:G2")
        _filter_info = ""
        if self._riwayat_filter_tanggal:
            _filter_info += f"  |  Tanggal: {self._riwayat_filter_tanggal}"
        if (self._riwayat_filter_kasir or "SEMUA") != "SEMUA":
            _filter_info += f"  |  Kasir: {self._riwayat_filter_kasir}"
        ws["A2"] = f"Dicetak: {datetime.now().strftime('%d %B %Y %H:%M')}  |  Kasir: {self.current_user}{_filter_info}"
        ws["A2"].font      = Font(name="Consolas", color="6060A0", size=9)
        ws["A2"].fill      = PatternFill("solid", fgColor="0D0D1A")
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 18

        headers = ["Waktu", "Kasir", "TV/PC", "Paket", "Pesanan Tambahan", "Diskon", "Total"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center
            cell.border    = border
        ws.row_dimensions[3].height = 24

        for r_idx, row in enumerate(rows_to_export, 4):
            fill = accent_fill if r_idx % 2 == 0 else PatternFill("solid", fgColor="161628")
            # Pad rows to 7 cols for backward compat (skip Status col)
            padded = list(row)[:7] + ["—"] * max(0, 7 - len(row))
            for c_idx, val in enumerate(padded[:7], 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font      = normal_font
                cell.fill      = fill
                cell.border    = border
                cell.alignment = center if c_idx in (1, 2, 3, 4, 6, 7) else Alignment(vertical="center")

        col_widths = [20, 12, 14, 14, 35, 10, 16]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        last_row = len(self.riwayat_transaksi) + 4
        ws.merge_cells(f"A{last_row}:F{last_row}")
        ws[f"A{last_row}"] = "TOTAL PENDAPATAN"
        ws[f"A{last_row}"].font      = Font(name="Consolas", bold=True, color="FFD700", size=11)
        ws[f"A{last_row}"].fill      = PatternFill("solid", fgColor="1A1A3A")
        ws[f"A{last_row}"].alignment = Alignment(horizontal="right")
        total_tv_paket = sum(m.get('paket_harga', m.get('total', 0) - m.get('pesanan_total', 0)) for m in self.riwayat_meta if m.get('source') == 'tv')
        total_warnet_paket = sum(m.get('paket_harga', m.get('total', 0) - m.get('pesanan_total', 0)) for m in self.riwayat_meta if m.get('source') == 'warnet')
        total_pesanan = sum(m.get('pesanan_total', 0) for m in self.riwayat_meta)
        total_all = sum(m.get('total', 0) for m in self.riwayat_meta)
        ws[f"G{last_row}"] = fmt_rp(total_all)
        ws[f"G{last_row}"].font      = Font(name="Consolas", bold=True, color="FFD700", size=11)
        ws[f"G{last_row}"].fill      = PatternFill("solid", fgColor="1A1A3A")
        ws[f"G{last_row}"].alignment = center

        summary_row = last_row + 1
        for label, amount in [
            ("TOTAL TV (Penjualan Paket)", total_tv_paket),
            ("TOTAL WARNET (Penjualan Paket)", total_warnet_paket),
            ("TOTAL MAKANAN & MINUMAN", total_pesanan),
            ("TOTAL KESELURUHAN", total_all),
        ]:
            ws.merge_cells(f"A{summary_row}:E{summary_row}")
            ws[f"A{summary_row}"] = label
            ws[f"A{summary_row}"].font      = Font(name="Consolas", bold=True, color="00FFCC" if label != "TOTAL KESELURUHAN" else "FFD700", size=10)
            ws[f"A{summary_row}"].fill      = PatternFill("solid", fgColor="12122A")
            ws[f"A{summary_row}"].alignment = Alignment(horizontal="right")
            ws[f"F{summary_row}"] = fmt_rp(amount)
            ws[f"F{summary_row}"].font      = Font(name="Consolas", bold=True, color="00FFCC" if label != "TOTAL KESELURUHAN" else "FFD700", size=10)
            ws[f"F{summary_row}"].fill      = PatternFill("solid", fgColor="12122A")
            ws[f"F{summary_row}"].alignment = center
            summary_row += 1

        tgl = datetime.now().strftime("%d-%m-%Y_%H-%M")
        default_name = f"laporan_rr_billing_{tgl}.xlsx"
        docs_dir = os.path.join(os.path.expanduser("~"), "Documents")
        initial_dir = docs_dir if os.path.isdir(docs_dir) else os.path.expanduser("~")

        save_path = filedialog.asksaveasfilename(
            parent=self,
            title="Simpan Laporan Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
            initialfile=default_name,
            initialdir=initial_dir
        )

        if not save_path:
            return

        try:
            wb.save(save_path)
        except Exception as e:
            messagebox.showerror("❌ Export Gagal",
                                 f"Gagal menyimpan file Excel:\n{str(e)}",
                                 parent=self)
            return

        messagebox.showinfo("✅ Export Berhasil",
                            f"File berhasil disimpan:\n{save_path}",
                            parent=self)

    # ══════════════════════════════════════════════════════════════════════════
    #  TAB 4.5: Booking Online (web /b/<user>)
    #  Pengaturan (logo, rekening x3, no DANA, alamat) + riwayat booking
    # ══════════════════════════════════════════════════════════════════════════
    def _setup_booking(self):
        f = self.frames["booking"]
        hdr = ctk.CTkFrame(f, fg_color=C_PANEL, height=54, corner_radius=0)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        ctk.CTkLabel(hdr, text="📅  BOOKING ONLINE",
                     font=FONT_TITLE, text_color=C_ACCENT).pack(side="left", padx=18, pady=14)

        role_skrg = self.current_role or "kasir"
        is_admin = role_skrg == "admin"
        owner = (self.current_user or "").strip()
        # Kasir melihat data pemilik (admin_utama) dalam mode lihat-saja
        if not is_admin:
            try:
                _users = ConfigManager.get("users", {})
                _rec = _users.get(self.current_user, {}) if isinstance(_users, dict) else {}
                if isinstance(_rec, dict) and _rec.get("admin_utama"):
                    owner = str(_rec["admin_utama"]).strip()
            except Exception:
                pass

        # Halaman pelanggan: link rrcctv.online/b/<username> — terisi otomatis
        # dengan username login dan bisa diklik (buka browser).
        owner_slug = owner.lower().strip()
        if owner_slug:
            ctk.CTkButton(hdr, text=f"🔗 rrcctv.online/b/{owner_slug}", height=30,
                          fg_color="transparent", hover_color="#1E1E4A",
                          border_width=1, border_color=C_ACCENT2,
                          font=FONT_SMALL, text_color=C_ACCENT2,
                          command=lambda: webbrowser.open(f"https://rrcctv.online/b/{owner_slug}")
                          ).pack(side="right", padx=18)
        else:
            ctk.CTkLabel(hdr, text="Halaman pelanggan: rrcctv.online/b/<username>",
                         font=FONT_SMALL, text_color=C_MUTED).pack(side="right", padx=18)

        body = ctk.CTkFrame(f, fg_color="transparent")
        body.pack(fill="both", expand=True, padx=14, pady=10)
        body.grid_columnconfigure(0, weight=5)
        body.grid_columnconfigure(1, weight=7)
        body.grid_rowconfigure(0, weight=1)

        # ── KIRI: PENGATURAN ────────────────────────────────────────────────
        set_f = ctk.CTkFrame(body, fg_color=C_PANEL, corner_radius=14)
        set_f.grid(row=0, column=0, sticky="nsw", padx=(0, 8), pady=(0, 8))
        ctk.CTkLabel(set_f, text="⚙️  PENGATURAN BOOKING WEB",
                     font=FONT_SUB, text_color=C_ACCENT2).pack(anchor="w", padx=16, pady=(14, 6))

        cfg = ConfigManager.load()
        psemua = cfg.get("profil_rental", {}) or {}
        puser = psemua.get(owner, {}) or {}
        if not isinstance(puser, dict):
            puser = {}

        def _fld(parent, label):
            ctk.CTkLabel(parent, text=label, font=FONT_LABEL,
                         text_color=C_MUTED, anchor="w").pack(anchor="w", padx=16, pady=(4, 2))
            e = ctk.CTkEntry(parent, fg_color=C_BTN, text_color=C_TEXT,
                             border_color=C_BORDER, font=FONT_BODY, height=32)
            e.pack(fill="x", padx=16, pady=(0, 4))
            return e

        e_nama_dana = _fld(set_f, "👤 Nama Akun DANA")
        e_nama_dana.insert(0, str(puser.get("nama_dana", "")))
        e_dana = _fld(set_f, "📱 No DANA (transfer)")
        e_dana.insert(0, str(puser.get("no_dana", "")))

        ctk.CTkLabel(set_f, text="🏪 Nama Rental PS", font=FONT_LABEL,
                     text_color=C_MUTED, anchor="w").pack(anchor="w", padx=16, pady=(4, 2))
        e_nama = ctk.CTkEntry(set_f, fg_color=C_BTN, text_color=C_TEXT,
                              border_color=C_BORDER, font=FONT_BODY, height=32)
        e_nama.insert(0, str(puser.get("nama_rental", "")))
        e_nama.pack(fill="x", padx=16, pady=(0, 6))
        if not is_admin:
            for _e in (e_nama_dana, e_dana, e_nama):
                _e.configure(state="disabled")

        logo_box = {"b64": str(puser.get("logo", "") or "")}
        logo_row = ctk.CTkFrame(set_f, fg_color="transparent")
        logo_row.pack(fill="x", padx=16, pady=(4, 2))
        ctk.CTkLabel(logo_row, text="🖼 Logo Header",
                     font=FONT_LABEL, text_color=C_MUTED, anchor="w").pack(anchor="w", pady=(0, 4))
        # Ikon status saja (tanpa teks) — di-pack sekali di paling kiri, tidak
        # pernah dipak/pack_forget lagi supaya tombol di sebelahnya tetap tampil.
        lbl_logo_icon = ctk.CTkLabel(logo_row,
                                     text="✔" if logo_box["b64"] else "✖",
                                     font=FONT_SUB, width=24,
                                     text_color=C_GREEN if logo_box["b64"] else C_MUTED,
                                     anchor="w")
        lbl_logo_icon.pack(side="left", padx=(0, 6))

        def _sync_logo_icon():
            ada = bool(logo_box["b64"])
            lbl_logo_icon.configure(text="✔" if ada else "✖",
                                    text_color=C_GREEN if ada else C_MUTED)

        if is_admin:
            ctk.CTkButton(logo_row, text="📁 Pilih", width=90, height=28,
                          fg_color=C_ACCENT2, font=("Russo One", 9, "bold"),
                          command=lambda: (self._profil_pilih_logo(logo_box, lbl_logo_icon), _sync_logo_icon())).pack(side="left", padx=(0, 5))
            ctk.CTkButton(logo_row, text="🗑 Hapus", width=80, height=28,
                          fg_color=C_RED, font=("Russo One", 9, "bold"),
                          command=lambda: (self._profil_hapus_logo(logo_box, lbl_logo_icon), _sync_logo_icon())).pack(side="left")

        qr_box = {"b64": str(puser.get("qr_pembayaran", "") or "")}
        qr_row = ctk.CTkFrame(set_f, fg_color="transparent")
        qr_row.pack(fill="x", padx=16, pady=(4, 2))
        ctk.CTkLabel(qr_row, text="📸 QR Kode Pembayaran (statis — tampil di web booking)",
                     font=FONT_LABEL, text_color=C_MUTED, anchor="w").pack(anchor="w", pady=(0, 4))
        lbl_qr_icon = ctk.CTkLabel(qr_row,
                                   text="✔" if qr_box["b64"] else "✖",
                                   font=FONT_SUB, width=24,
                                   text_color=C_GREEN if qr_box["b64"] else C_MUTED,
                                   anchor="w")
        lbl_qr_icon.pack(side="left", padx=(0, 6))

        def _sync_qr_icon():
            ada = bool(qr_box["b64"])
            lbl_qr_icon.configure(text="✔" if ada else "✖",
                                  text_color=C_GREEN if ada else C_MUTED)

        if is_admin:
            ctk.CTkButton(qr_row, text="📁 Pilih", width=90, height=28,
                          fg_color=C_ACCENT2, font=("Russo One", 9, "bold"),
                          command=lambda: (self._profil_pilih_qr(qr_box, lbl_qr_icon), _sync_qr_icon())).pack(side="left", padx=(0, 5))
            ctk.CTkButton(qr_row, text="🗑 Hapus", width=80, height=28,
                          fg_color=C_RED, font=("Russo One", 9, "bold"),
                          command=lambda: (self._profil_hapus_qr(qr_box, lbl_qr_icon), _sync_qr_icon())).pack(side="left")
        ctk.CTkLabel(set_f,
                     text="QR statis (QRIS / rekening) tampil di halaman web booking saat "
                          "pelanggan memilih metode Lunas / DP.",
                     font=FONT_SMALL, text_color=C_MUTED, justify="left", anchor="w",
                     wraplength=360).pack(anchor="w", padx=16, pady=(8, 4))

        ctk.CTkLabel(set_f,
                     text="No DANA + nama akun akan tampil di halaman booking saat "
                          "pelanggan memilih metode Transfer.\nData langsung dipush "
                          "ke call_meta (web) setelah disimpan.",
                     font=FONT_SMALL, text_color=C_MUTED, justify="left", anchor="w",
                     wraplength=360).pack(anchor="w", padx=16, pady=(8, 4))

        if is_admin:
            def _simpan_booking_set():
                data = {
                    "nama_dana": sanitize_text(e_nama_dana.get()),
                    "no_dana": sanitize_text(e_dana.get()),
                    "nama_rental": sanitize_text(e_nama.get()),
                    "logo": logo_box["b64"],
                    "qr_pembayaran": qr_box["b64"],
                }
                self._simpan_profil_rental(data)
                self._qr_push_menu_bg()

            ctk.CTkButton(set_f, text="💾  Simpan & Push ke Web", height=36,
                          fg_color=C_ACCENT2, font=("Russo One", 10, "bold"), text_color="white",
                          command=_simpan_booking_set).pack(anchor="w", padx=16, pady=(10, 14))
        else:
            ctk.CTkLabel(set_f, text="🔒  Mode Lihat — pengaturan hanya admin\n"
                                     "(kasir dapat melihat riwayat & player)",
                         font=FONT_LABEL, text_color=C_YELLOW, justify="left", anchor="w",
                         wraplength=360).pack(anchor="w", padx=16, pady=(10, 14))

        # ── KANAN: RIWAYAT BOOKING ──────────────────────────────────────────
        riw_f = ctk.CTkFrame(body, fg_color=C_PANEL, corner_radius=14)
        riw_f.grid(row=0, column=1, sticky="nsew", padx=(8, 0), pady=(0, 8))
        riw_hdr = ctk.CTkFrame(riw_f, fg_color="transparent")
        riw_hdr.pack(fill="x", padx=16, pady=(14, 6))
        ctk.CTkLabel(riw_hdr, text="📋  RIWAYAT BOOKING & PLAYER",
                     font=FONT_SUB, text_color=C_ACCENT2).pack(side="left")
        ctk.CTkButton(riw_hdr, text="⟳ Refresh", width=90, height=28,
                      fg_color=C_BTN, font=("Russo One", 9, "bold"),
                      command=self._booking_riwayat_refresh).pack(side="right")

        plant = ctk.CTkFrame(riw_f, fg_color="transparent")
        plant.pack(fill="x", padx=16, pady=(2, 2))
        cols = ["Kode", "Nama Player", "WhatsApp", "Perangkat", "Mulai", "Metode", "Status Bayar", "Booking", "Aksi"]
        wcols = [70, 100, 105, 85, 105, 60, 125, 80, 110]
        for c, w in zip(cols, wcols):
            ctk.CTkLabel(plant, text=c, width=w, font=("Russo One", 9, "bold"),
                         text_color=C_MUTED, anchor="w").pack(side="left", padx=(0, 4))
        self._booking_scroll = ctk.CTkScrollableFrame(riw_f, fg_color=C_BG)
        self._booking_scroll.pack(fill="both", expand=True, padx=16, pady=(2, 14))
        self._booking_riwayat_running = False
        self.after(1000, self._booking_riwayat_refresh)

    def _booking_riwayat_refresh(self):
        """Ambil booking dari Firestore (milik user ini) lalu render daftar."""
        if getattr(self, "_booking_riwayat_running", False):
            return
        self._booking_riwayat_running = True
        uname = (getattr(self, "current_user", None) or "").strip().lower()
        if not uname:
            self._booking_riwayat_running = False
            return
        # Kasir melihat booking milik pemiliknya (admin_utama)
        if (self.current_role or "kasir") != "admin":
            try:
                _users = ConfigManager.get("users", {})
                _rec = _users.get(self.current_user, {}) if isinstance(_users, dict) else {}
                if isinstance(_rec, dict) and _rec.get("admin_utama"):
                    uname = str(_rec["admin_utama"]).strip().lower()
            except Exception:
                pass

        def worker():
            docs = []
            try:
                from firestore_sync import get_firestore_client
                docs = get_firestore_client().query_all("bookings", limit=100,
                                                        order_field="createdAt")
            except Exception:
                docs = []
            rows = []
            for d in docs:
                if str(d.get("owner", "")).strip().lower() != uname:
                    continue
                rows.append(d)
            rows.sort(key=lambda d: str(d.get("createdAt", "")), reverse=True)
            try:
                self.after(0, lambda rs=rows: self._booking_riwayat_render(rs))
            except Exception:
                pass
            self._booking_riwayat_running = False

        try:
            threading.Thread(target=worker, daemon=True).start()
        except Exception:
            self._booking_riwayat_running = False
        self.after(15000, self._booking_riwayat_refresh)

    def _booking_status_bayar(self, d: dict) -> tuple:
        sb = str(d.get("statusBayar", "") or "")
        if sb == "lunas_transfer":
            return ("LUNAS VIA TRANSFER", C_GREEN)
        if sb == "dp":
            sisa = int(d.get("sisaBayar", 0) or 0)
            return (f"TAGIHAN SISA Rp {sisa:,}", C_ORANGE)
        if sb == "belum_bayar":
            return ("BELUM BAYAR", C_YELLOW)
        return ("—", C_MUTED)

    def _booking_buka_bukti(self, d: dict, parent=None):
        bukti = str(d.get("bukti", "") or "")
        if not bukti.startswith("data:image/"):
            messagebox.showinfo("Tidak Ada", "Tidak ada bukti transfer.", parent=parent or self)
            return
        try:
            import base64 as _b64
            folder = app_path("invoices_bukti")
            os.makedirs(folder, exist_ok=True)
            raw = bukti.split(",", 1)[1]
            ext = "png" if "image/png" in bukti else "jpg"
            did = str(d.get("_id", ""))
            path = os.path.join(folder, f"booking_bukti_{did}.{ext}")
            with open(path, "wb") as f:
                f.write(_b64.b64decode(raw))
            os.startfile(path)
        except Exception as e:
            messagebox.showerror("Gagal", f"Gagal membuka bukti:\n{e}", parent=parent or self)

    def _booking_update_status(self, did: str, status: str, alasan: str = ""):
        try:
            from firestore_sync import get_firestore_client
            patch = {"status": status, "kasir": str(getattr(self, "current_user", "")),
                     "alasan": alasan, "updatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
            get_firestore_client().set_document(f"bookings/{did}", patch, merge=True)
            self._qr_log(f"booking {did} -> {status} (dari tab Booking)")
            self._booking_riwayat_refresh()
        except Exception as e:
            self._qr_log(f"booking {did} update gagal: {e}")
            messagebox.showerror("Gagal", f"Gagal update status:\n{e}", parent=self)

    def _booking_riwayat_render(self, rows: list):
        sc = getattr(self, "_booking_scroll", None)
        if sc is None:
            return
        # Tanpa kedip: hanya bangun ulang daftar jika data benar-benar berubah.
        sig = tuple(sorted(
            (str(d.get("_id", "")), str(d.get("status", "")),
             str(d.get("statusBayar", "")), str(d.get("createdAt", "")),
             str(d.get("updatedAt", "")), str(d.get("namaPelanggan", "")),
             str(d.get("noHp", "")), str(d.get("perangkat", "")),
             str(d.get("tanggal", "")), str(d.get("jam", "")),
             str(d.get("metode", "")), str(d.get("totalHarga", "")),
             str(d.get("sisaBayar", "")), str(d.get("bukti", "") or ""))
            for d in rows))
        if getattr(self, "_booking_riwayat_sig", None) == sig:
            return
        self._booking_riwayat_sig = sig
        for w in sc.winfo_children():
            w.destroy()
        if not rows:
            ctk.CTkLabel(sc, text="Belum ada booking untuk rental ini.",
                         font=FONT_BODY, text_color=C_MUTED).pack(pady=20)
            return
        for d in rows:
            did = str(d.get("_id", ""))
            st = str(d.get("status", ""))
            sb_txt, sb_col = self._booking_status_bayar(d)
            st_txt = {"baru": "⏳ BARU", "dikonfirmasi": "✅ DITERIMA", "ditolak": "✖ DITOLAK"}.get(st, st)
            st_col = C_GREEN if st == "dikonfirmasi" else C_YELLOW if st == "baru" else C_RED
            mulai = f"{d.get('tanggal', '')} {str(d.get('jam', ''))[:5]}".strip()
            metode_txt = {"lunas": "Lunas", "dp": "DP"}.get(str(d.get("metode", "")), "Tempat")
            row = ctk.CTkFrame(sc, fg_color=C_PANEL, corner_radius=8)
            row.pack(fill="x", pady=3)
            vals = [did[:8].upper(), str(d.get("namaPelanggan", "-")),
                    str(d.get("noHp", "-")), str(d.get("perangkat", "-")),
                    mulai or "-", metode_txt]
            click_widgets = [row]
            for v, w in zip(vals, [70, 100, 105, 85, 105, 60]):
                lbl = ctk.CTkLabel(row, text=v, width=w, font=("Consolas" if len(v) < 20 else FONT_SMALL, 10),
                                   text_color=C_TEXT, anchor="w")
                lbl.pack(side="left", padx=(0, 4))
                click_widgets.append(lbl)
            lbl_sb = ctk.CTkLabel(row, text=sb_txt, width=125, font=("Russo One", 9, "bold"),
                                  text_color=sb_col, anchor="w")
            lbl_sb.pack(side="left", padx=(0, 4))
            click_widgets.append(lbl_sb)
            lbl_st = ctk.CTkLabel(row, text=st_txt, width=80, font=("Russo One", 9, "bold"),
                                  text_color=st_col, anchor="w")
            lbl_st.pack(side="left", padx=(0, 4))
            click_widgets.append(lbl_st)
            act = ctk.CTkFrame(row, fg_color="transparent")
            act.pack(side="left")
            wa_num = re.sub(r"[^0-9]", "", str(d.get("noHp", "") or ""))
            if wa_num.startswith("0"):
                wa_num = "62" + wa_num[1:]
            ctk.CTkButton(act, text="💬", width=30, height=24, fg_color=C_BTN,
                          font=("Segoe UI Emoji", 11),
                          command=lambda n=wa_num: self._booking_wa_chat(n),
                          state="normal" if wa_num else "disabled").pack(side="left", padx=2)
            ctk.CTkButton(act, text="🖼", width=30, height=24, fg_color=C_BTN,
                          font=("Segoe UI Emoji", 11),
                          command=lambda dd=dict(d): self._booking_buka_bukti(dd),
                          state="normal" if d.get("bukti") else "disabled").pack(side="left", padx=2)
            if str(d.get("statusBayar", "")) == "dp" and st != "ditolak":
                ctk.CTkButton(act, text="💰", width=30, height=24, fg_color=C_ORANGE,
                              text_color="#1A1200", font=("Segoe UI Emoji", 11),
                              command=lambda dd=dict(d): self._booking_lunas_sisa(dd)).pack(side="left", padx=2)
            if st == "baru":
                ctk.CTkButton(act, text="✅", width=30, height=24, fg_color=C_GREEN,
                              text_color="#06210F", font=("Segoe UI Emoji", 11),
                              command=lambda i=did: self._booking_update_status(i, "dikonfirmasi")).pack(side="left", padx=2)
                ctk.CTkButton(act, text="✖", width=30, height=24, fg_color=C_RED,
                              text_color="white", font=("Segoe UI Emoji", 11),
                              command=lambda i=did: self._booking_tolak(i)).pack(side="left", padx=2)
            for wdg in click_widgets:
                try:
                    wdg.bind("<Button-1>",
                             lambda e, dd=dict(d): self._qr_tampil_booking(dd, bunyi=False))
                except Exception:
                    pass

    def _booking_wa_chat(self, wa_num: str):
        if not wa_num:
            return
        try:
            webbrowser.open(f"https://wa.me/{wa_num}")
        except Exception as e:
            messagebox.showerror("Gagal", f"Gagal membuka WhatsApp:\n{e}", parent=self)

    def _booking_tolak(self, did: str):
        from tkinter import simpledialog
        alasan = simpledialog.askstring("Tolak Booking", "Alasan penolakan:", parent=self)
        if alasan is None:
            return
        self._booking_update_status(did, "ditolak", alasan.strip())

    def _booking_fetch_valid(self, tv_label, cb):
        """Ambil booking valid untuk kartu tertentu (thread, tanpa blokir UI):
        status dikonfirmasi, perangkat = kartu, tanggal+jam belum lewat, dan
        belum sesiDimulai. Hasil dipanggil di main thread via cb(rows)."""
        uname = (getattr(self, "current_user", None) or "").strip().lower()
        tv_s = str(tv_label or "").strip()

        def worker():
            rows = []
            try:
                from firestore_sync import get_firestore_client
                docs = get_firestore_client().query_all("bookings", limit=100,
                                                        order_field="createdAt")
            except Exception:
                docs = []
            now = datetime.now()
            for d in docs:
                try:
                    if str(d.get("owner", "")).strip().lower() != uname:
                        continue
                    if str(d.get("status", "")) != "dikonfirmasi":
                        continue
                    if str(d.get("perangkat", "") or "").strip() != tv_s:
                        continue
                    if d.get("sesiDimulai"):
                        continue
                    tgl = str(d.get("tanggal", "") or "").strip()
                    jam = str(d.get("jam", "") or "").strip()[:5]
                    try:
                        mulai = datetime.strptime(f"{tgl} {jam}", "%Y-%m-%d %H:%M")
                    except Exception:
                        mulai = None
                    if mulai is not None and mulai < now:
                        continue
                    rows.append(d)
                except Exception:
                    continue
            rows.sort(key=lambda x: str(x.get("jam", "")))
            try:
                self.after(0, lambda rs=rows: cb(rs))
            except Exception:
                pass

        try:
            threading.Thread(target=worker, daemon=True).start()
        except Exception:
            try:
                self.after(0, lambda: cb([]))
            except Exception:
                pass

    def _buka_booking_dari_riwayat(self, idx):
        """Klik kanan baris riwayat bertanda 📅 — buka popup booking terkait."""
        try:
            bid = str(self.riwayat_meta[idx].get("booking_id") or "")
        except Exception:
            return
        if not bid:
            return

        def worker():
            doc = None
            try:
                from firestore_sync import get_firestore_client
                doc = get_firestore_client().get_document(f"bookings/{bid}")
            except Exception:
                doc = None

            def done():
                if doc:
                    self._qr_tampil_booking(dict(doc), bunyi=False)
                else:
                    messagebox.showinfo("Booking", "Booking tidak ditemukan di Firestore.",
                                        parent=self)
            try:
                self.after(0, done)
            except Exception:
                pass

        try:
            threading.Thread(target=worker, daemon=True).start()
        except Exception as e:
            self._qr_log(f"buka booking {bid} gagal: {e}")

    def _booking_lunas_sisa(self, d: dict):
        """Tandai sisa DP booking sebagai LUNAS (dari tombol 💰 di riwayat)."""
        did = str(d.get("_id", ""))
        sisa = int(d.get("sisaBayar", 0) or 0)
        dpn = int(d.get("nominalDp", 0) or 0)
        total = int(d.get("totalHarga", 0) or 0)
        if not messagebox.askyesno(
                "Lunas Sisa",
                f"Tandai sisa Rp {sisa:,} dari total Rp {total:,} sebagai LUNAS?\n"
                f"(DP Rp {dpn:,} + sisa Rp {sisa:,})",
                parent=self):
            return
        try:
            from firestore_sync import get_firestore_client
            pat = {"statusBayar": "lunas_transfer",
                   "nominalTransfer": total,
                   "pelunasanSisa": sisa,
                   "lunasAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                   "updatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                   "kasir": str(getattr(self, "current_user", ""))}
            ok = get_firestore_client().set_document(f"bookings/{did}", pat, merge=True)
        except Exception as e:
            self._qr_log(f"booking {did} lunas sisa gagal: {e}")
            ok = False
        if not ok:
            messagebox.showerror("Gagal", "Gagal menyimpan — cek internet.", parent=self)
            return
        self._qr_log(f"booking {did} -> sisa dilunasi (dari riwayat)")
        idx = self._riwayat_idx_by_booking(did)
        if idx >= 0:
            try:
                self._set_transaksi_paid_idx(idx, True)
            except Exception:
                pass
        perangkat = str(d.get("perangkat", "") or "").strip()
        kursi = self._qr_cari_kartu(perangkat) if perangkat else None
        if kursi is not None:
            try:
                if hasattr(kursi, "_set_paid"):
                    kursi._set_paid(True)
            except Exception:
                pass
        messagebox.showinfo("Lunas",
                            f"✅ Sisa Rp {sisa:,} booking {did[:8].upper()} "
                            f"ditandai LUNAS via transfer.", parent=self)
        self._booking_riwayat_refresh()

    # ══════════════════════════════════════════════════════════════════════════
    #  TAB WiFi
    # ══════════════════════════════════════════════════════════════════════════
    def _setup_wifi(self):
        f = self.frames["wifi"]
        hdr = ctk.CTkFrame(f, fg_color=C_PANEL, height=54, corner_radius=0)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        ctk.CTkLabel(hdr, text="🔗  PANDUAN KONEKSI TV (ATPv2)",
                     font=FONT_TITLE, text_color=C_ACCENT).pack(side="left", padx=18, pady=14)

        scroll = ctk.CTkScrollableFrame(f, fg_color=C_BG)
        scroll.pack(fill="both", expand=True, padx=14, pady=10)

        # ── Frame 1: Langkah di TV ──────────────────────────────────────────
        tv_f = ctk.CTkFrame(scroll, fg_color=C_PANEL, corner_radius=12)
        tv_f.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(tv_f, text="📱  LANGKAH DI ANDROID TV",
                     font=FONT_SUB, text_color=C_ACCENT2).pack(anchor="w", padx=16, pady=(14, 8))

        langkah_tv = [
            ("1", "Pengaturan → Preferensi Perangkat → Tentang",
             "Klik 'Build' 7x sampai mode developer aktif"),
            ("2", "Pengaturan → Opsi Developer",
             "Aktifkan 'Wireless Debugging'"),
            ("3", "Opsi Developer → Stay awake",
             "Aktifkan 'Stay awake' (Tetap terjaga saat charging)"),
            ("4", "Tap 'Pair device with pairing code'",
             "Catat IP Address, Port Pairing (6466), dan Kode PIN 6 digit"),
        ]
        for no, judul, sub in langkah_tv:
            row = ctk.CTkFrame(tv_f, fg_color=C_CARD, corner_radius=8)
            row.pack(fill="x", padx=14, pady=4)
            ctk.CTkLabel(row, text=no, font=("Russo One", 12, "bold"),
                         text_color=C_BG, fg_color=C_ACCENT2, corner_radius=14,
                         width=28, height=28).pack(side="left", padx=(10, 12), pady=10)
            txt_f = ctk.CTkFrame(row, fg_color="transparent")
            txt_f.pack(side="left", fill="x", expand=True, pady=8)
            ctk.CTkLabel(txt_f, text=judul, font=("Consolas", 10, "bold"),
                         text_color=C_TEXT, anchor="w").pack(anchor="w")
            ctk.CTkLabel(txt_f, text=sub, font=FONT_SMALL,
                         text_color=C_MUTED, anchor="w").pack(anchor="w")
        ctk.CTkFrame(tv_f, fg_color="transparent", height=8).pack()

        # ── Frame 2: Langkah di Aplikasi ────────────────────────────────────
        app_f = ctk.CTkFrame(scroll, fg_color=C_PANEL, corner_radius=12)
        app_f.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(app_f, text="💻  LANGKAH DI APLIKASI (RR Billing PRO)",
                     font=FONT_SUB, text_color=C_ACCENT2).pack(anchor="w", padx=16, pady=(14, 8))

        langkah_app = [
            ("1", "Dashboard → Klik ➕ Tambah TV",
             "Isi IP TV dan Nama TV, klik Simpan & Pair"),
            ("2", "Masukkan PIN 6 angka dari TV",
             "Klik 'Pasangkan' — pairing akan dimulai"),
            ("3", "TV siap digunakan",
             "Badge ONLINE hijau muncul di KartuTV — TV siap digunakan"),
        ]
        for no, judul, sub in langkah_app:
            row = ctk.CTkFrame(app_f, fg_color=C_CARD, corner_radius=8)
            row.pack(fill="x", padx=14, pady=4)
            ctk.CTkLabel(row, text=no, font=("Russo One", 12, "bold"),
                         text_color=C_BG, fg_color=C_ACCENT2, corner_radius=14,
                         width=28, height=28).pack(side="left", padx=(10, 12), pady=10)
            txt_f = ctk.CTkFrame(row, fg_color="transparent")
            txt_f.pack(side="left", fill="x", expand=True, pady=8)
            ctk.CTkLabel(txt_f, text=judul, font=("Consolas", 10, "bold"),
                         text_color=C_TEXT, anchor="w").pack(anchor="w")
            ctk.CTkLabel(txt_f, text=sub, font=FONT_SMALL,
                         text_color=C_MUTED, anchor="w").pack(anchor="w")
        ctk.CTkFrame(app_f, fg_color="transparent", height=8).pack()

        # ── Frame 3: Ganti metode di TV existing ───────────────────────────
        ganti_f = ctk.CTkFrame(scroll, fg_color=C_PANEL, corner_radius=12)
        ganti_f.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(ganti_f, text="🔄  GANTI METODE DI TV SUDAH ADA",
                     font=FONT_SUB, text_color=C_ACCENT2).pack(anchor="w", padx=16, pady=(14, 8))

        langkah_ganti = [
            ("1", "Klik tombol IP di KartuTV (ucok/192.168.x.x)",
             "Dialog Ganti IP akan terbuka"),
            ("2", "Klik 🔗 Tes Koneksi",
             "Jika perlu pairing, masukkan PIN 6 angka dari TV"),
            ("3", "Jika berhasil",
             "Badge berubah jadi ONLINE hijau"),
        ]
        for no, judul, sub in langkah_ganti:
            row = ctk.CTkFrame(ganti_f, fg_color=C_CARD, corner_radius=8)
            row.pack(fill="x", padx=14, pady=4)
            ctk.CTkLabel(row, text=no, font=("Russo One", 12, "bold"),
                         text_color=C_BG, fg_color=C_ACCENT2, corner_radius=14,
                         width=28, height=28).pack(side="left", padx=(10, 12), pady=10)
            txt_f = ctk.CTkFrame(row, fg_color="transparent")
            txt_f.pack(side="left", fill="x", expand=True, pady=8)
            ctk.CTkLabel(txt_f, text=judul, font=("Consolas", 10, "bold"),
                         text_color=C_TEXT, anchor="w").pack(anchor="w")
            ctk.CTkLabel(txt_f, text=sub, font=FONT_SMALL,
                         text_color=C_MUTED, anchor="w").pack(anchor="w")
        ctk.CTkFrame(ganti_f, fg_color="transparent", height=8).pack()

        # ── Frame 4: Troubleshooting ────────────────────────────────────────
        trouble_f = ctk.CTkFrame(scroll, fg_color=C_PANEL, corner_radius=12)
        trouble_f.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(trouble_f, text="🔧  TROUBLESHOOTING",
                     font=FONT_SUB, text_color=C_ACCENT2).pack(anchor="w", padx=16, pady=(14, 12))

        tips = [
            ("🌐", "Satu jaringan", "PC dan TV harus terhubung ke jaringan LAN/Wi-Fi yang sama"),
            ("🧱", "Firewall", "Pastikan firewall tidak memblokir port 6466 (API) dan 6467 (Pairing)"),
            ("⏳", "Stay awake", "Aktifkan 'Stay awake' di Opsi Developer agar TV tidak tidur saat pairing"),
            ("🔄", "Pairing ulang", "Jika pairing gagal, tutup dialog dan ulangi dari awal — pastikan PIN benar"),
        ]
        for ico, judul, sub in tips:
            row = ctk.CTkFrame(trouble_f, fg_color=C_CARD, corner_radius=8)
            row.pack(fill="x", padx=14, pady=4)
            ctk.CTkLabel(row, text=ico, font=("Segoe UI Emoji", 14),
                         width=32).pack(side="left", padx=(8, 4), pady=10)
            txt_f = ctk.CTkFrame(row, fg_color="transparent")
            txt_f.pack(side="left", fill="x", expand=True, pady=8)
            ctk.CTkLabel(txt_f, text=judul, font=("Consolas", 10, "bold"),
                         text_color=C_TEXT, anchor="w").pack(anchor="w")
            ctk.CTkLabel(txt_f, text=sub, font=FONT_SMALL,
                         text_color=C_MUTED, anchor="w").pack(anchor="w")
        ctk.CTkFrame(trouble_f, fg_color="transparent", height=8).pack()



    # ══════════════════════════════════════════════════════════════════════════
    #  TAB 5: AKTIVASI
    # ══════════════════════════════════════════════════════════════════════════
    def _setup_aktivasi(self):
        f = self.frames["aktivasi"]
        is_admin = (self.current_role or "kasir") == "admin"
        hdr = ctk.CTkFrame(f, fg_color=C_PANEL, height=54, corner_radius=0)
        hdr.pack(fill="x")
        ctk.CTkLabel(hdr, text="🔑  AKTIVASI & BERLANGGANAN",
                     font=FONT_TITLE, text_color=C_ACCENT).pack(side="left", padx=18, pady=14)

        scroll = ctk.CTkScrollableFrame(f, fg_color=C_BG)
        scroll.pack(fill="both", expand=True, padx=20, pady=16)

        try:
            status_lic = LicenseManager.get_status(current_user=self._resolve_license_user())
        except Exception as e:
            _LOGGER.exception("Gagal get_status di _setup_aktivasi: %s", e)
            status_lic = {"status": "unknown", "sisa_hari": 0, "pesan": f"Error: {e}"}
        status_card = ctk.CTkFrame(scroll, fg_color=C_PANEL, corner_radius=14,
                                    border_width=2,
                                    border_color=C_GREEN if status_lic["status"] == "active"
                                                 else C_YELLOW if status_lic["status"] == "trial"
                                                 else C_RED)
        status_card.pack(fill="x", pady=(0, 16))

        ctk.CTkLabel(status_card, text="📋  STATUS LISENSI SAAT INI",
                     font=FONT_SUB, text_color=C_ACCENT2).pack(anchor="w", padx=20, pady=(16, 8))

        ico_status = "✅" if status_lic["status"] == "active" else "🕐" if status_lic["status"] == "trial" else "⛔"
        color_status = C_GREEN if status_lic["status"] == "active" else C_YELLOW if status_lic["status"] == "trial" else C_RED
        ctk.CTkLabel(status_card, text=f"{ico_status}  {status_lic['pesan']}",
                     font=("Russo One", 14, "bold"), text_color=color_status).pack(padx=20, pady=8)

        if status_lic["status"] == "active":
            lic = LicenseManager.load()
            ctk.CTkLabel(status_card, text=f"Kode: {lic.get('kode_aktivasi', '')}  |  Aktif sejak: {lic.get('tgl_aktivasi', '—')[:10]}",
                         font=FONT_SMALL, text_color=C_MUTED).pack(padx=20, pady=(0, 16))
            if is_admin:
                revoke_row = ctk.CTkFrame(status_card, fg_color="transparent")
                revoke_row.pack(padx=20, pady=(0, 12))
                ctk.CTkButton(revoke_row, text="🛑  Revoke License", width=160, height=34,
                              fg_color=C_RED, hover_color="#7A1A1A",
                              font=("Russo One", 10, "bold"), text_color="white",
                              command=self._revoke_license).pack(side="left")
        else:
            ctk.CTkLabel(status_card,
                         text="Aktifkan lisensi untuk menghilangkan batasan trial dan menggunakan semua fitur.",
                         font=FONT_BODY, text_color=C_MUTED, wraplength=700).pack(padx=20, pady=(0, 16))

        akt_card = ctk.CTkFrame(scroll, fg_color=C_PANEL, corner_radius=14)
        akt_card.pack(fill="x", pady=(0, 16))
        ctk.CTkLabel(akt_card, text="🔐  MASUKKAN KODE AKTIVASI",
                     font=FONT_SUB, text_color=C_ACCENT2).pack(anchor="w", padx=20, pady=(16, 6))
        ctk.CTkLabel(akt_card, text="Format kode:  RR-XXXX-XXXX-XXXX  (diperoleh setelah pembayaran)",
                     font=FONT_SMALL, text_color=C_MUTED).pack(anchor="w", padx=20, pady=(0, 10))
        akt_row = ctk.CTkFrame(akt_card, fg_color="transparent")
        akt_row.pack(fill="x", padx=20, pady=(0, 12))
        self.entry_kode = ctk.CTkEntry(akt_row, placeholder_text="RR-XXXX-XXXX-XXXX",
                                        fg_color=C_BTN, text_color=C_ACCENT,
                                        border_color=C_BORDER, font=("Consolas", 14, "bold"),
                                        height=42, width=340)
        self.entry_kode.pack(side="left", padx=(0, 12))
        ctk.CTkButton(akt_row, text="🔓  Aktifkan", width=140, height=42,
                      fg_color=C_ACCENT2, hover_color="#5A0FCC",
                      font=("Russo One", 11, "bold"), text_color="white",
                      command=self._lakukan_aktivasi).pack(side="left")
        self.lbl_akt_status = ctk.CTkLabel(akt_card, text="",
                                            font=FONT_BODY, text_color=C_MUTED)
        self.lbl_akt_status.pack(pady=(0, 16))

        # Fetch promo settings from Firestore
        promo_data = None
        try:
            fc = FirestoreClient()
            promo_data = fc.fetch_promo_settings()
        except Exception as e:
            _LOGGER.warning("Gagal fetch promo: %s", e)

        self._promo_data = promo_data

        promo_aktif = promo_data.get("promoAktif", False) if promo_data else False
        diskon_map = promo_data.get("diskonPerPaket", {}) if promo_data else {}
        add_tv_map = promo_data.get("addTvOverride", {}) if promo_data else {}

        PAKET_KEY_MAP = {"Bulanan": "1 Bulan", "3 Bulan": "3 Bulan", "Tahunan": "1 Tahun", "LIFETIME": "LIFETIME"}

        def _harga_after_diskon(nama, base_harga):
            if not promo_aktif:
                return base_harga
            key = PAKET_KEY_MAP.get(nama)
            if not key:
                return base_harga
            diskon = diskon_map.get(key, 0)
            if diskon <= 0:
                return base_harga
            return base_harga * (100 - diskon) // 100

        bayar_card = ctk.CTkFrame(scroll, fg_color=C_PANEL, corner_radius=14)
        bayar_card.pack(fill="x", pady=(0, 16))
        ctk.CTkLabel(bayar_card, text="💰  PAKET BERLANGGANAN",
                     font=FONT_SUB, text_color=C_ACCENT2).pack(anchor="w", padx=20, pady=(16, 4))
        ctk.CTkLabel(bayar_card, text="Klik 'Bayar Sekarang' untuk langsung diarahkan ke WhatsApp Admin",
                     font=("Courier New", 12), text_color=C_MUTED).pack(anchor="w", padx=20, pady=(0, 8))

        def fmt_rp(angka):
            s = f"{angka:,}".replace(",", ".")
            return f"Rp {s}"

        paket_base = [
            ("Bulanan",   "Rp 99.000 / bulan",   99_000,  "5 TV + 5 PC Warnet",          C_ACCENT,  "💎"),
            ("3 Bulan",   "Rp 299.000",           299_000, "10 TV + 10 PC Warnet",        C_GREEN,   "🚀"),
            ("Tahunan",   "Rp 999.000 / tahun",   999_000, "15 TV + 15 PC Warnet",        C_YELLOW,  "👑"),
            ("LIFETIME",  "Rp 2.000.000",         2_000_000, "UNLIMITED TV + PC Warnet 🏆", C_RED, "🏆"),
        ]

        paket_langganan = []
        for nama, harga_default, base_harga, deskripsi, color, ico in paket_base:
            harga_diskon = _harga_after_diskon(nama, base_harga)
            if promo_aktif and harga_diskon < base_harga:
                harga_tampil = f"~~{fmt_rp(base_harga)}~~ → {fmt_rp(harga_diskon)}"
            else:
                harga_tampil = harga_default

            if promo_aktif:
                key = PAKET_KEY_MAP.get(nama)
                if key:
                    tv_override = add_tv_map.get(key)
                    if tv_override is not None:
                        deskripsi = f"🔥 ADD TV {tv_override}"

            paket_langganan.append((nama, harga_tampil, deskripsi, color, ico, harga_diskon))

        paket_row = ctk.CTkFrame(bayar_card, fg_color="transparent")
        paket_row.pack(fill="x", padx=20, pady=(0, 16))

        for nama, harga, deskripsi, color, ico, harga_num in paket_langganan:
            card_p = ctk.CTkFrame(paket_row, fg_color=C_CARD, corner_radius=12,
                                   border_width=1, border_color=color)
            card_p.pack(side="left", fill="both", expand=True, padx=8)
            ctk.CTkLabel(card_p, text=ico, font=("Arial", 28)).pack(pady=(16, 4))
            ctk.CTkLabel(card_p, text=nama, font=("Russo One", 13, "bold"),
                         text_color=color).pack()
            ctk.CTkLabel(card_p, text=harga, font=("Consolas", 12, "bold"),
                         text_color=C_TEXT).pack(pady=4)
            ctk.CTkLabel(card_p, text=deskripsi, font=("Courier New", 12),
                         text_color=C_MUTED, wraplength=150).pack(pady=(0, 8))
            ctk.CTkButton(card_p, text="💳 Bayar Sekarang", width=130, height=32,
                          fg_color=color, hover_color=C_ACCENT2,
                          font=FONT_SUB, text_color=C_BG if color in (C_ACCENT, C_GREEN, C_YELLOW) else "white",
                          command=lambda n=nama, h=harga, hn=harga_num: self._pilih_paket_bayar(n, h, hn)
                          ).pack(pady=(0, 16))

        pay_card = ctk.CTkFrame(scroll, fg_color=C_PANEL, corner_radius=14)
        pay_card.pack(fill="x", pady=(0, 16))
        ctk.CTkLabel(pay_card, text="💳  METODE PEMBAYARAN",
                     font=FONT_SUB, text_color=C_ACCENT2).pack(anchor="w", padx=20, pady=(16, 8))

        metode = [
            ("🏦 Transfer Bank BCA",      "6145375553  a/n Rahmadani",           None),
            ("🏦 Transfer Bank BRI",      "0256 0109 2349 500  a/n Rahmadani",   None),
            ("💚 GoPay / OVO / Dana",     "0812-7064-7744 a/n Rahmadani",        None),
            ("🟦 QRIS",                   "Scan QR untuk pembayaran",             "qris"),
        ]
        for metode_nm, detail, aksi in metode:
            row_m = ctk.CTkFrame(pay_card, fg_color=C_CARD, corner_radius=8)
            row_m.pack(fill="x", padx=20, pady=4)
            ctk.CTkLabel(row_m, text=metode_nm, font=("Consolas", 12, "bold"),
                         text_color=C_TEXT, width=220, anchor="w").pack(side="left", padx=14, pady=10)
            ctk.CTkLabel(row_m, text=detail, font=("Courier New", 12),
                         text_color=C_MUTED).pack(side="left", padx=8)
            if aksi == "qris":
                ctk.CTkButton(row_m, text="🟦 Lihat QRIS", width=130, height=30,
                              fg_color="#0A2A4A", hover_color="#0A3A6A",
                              border_width=1, border_color="#3A8AFF",
                              font=("Russo One", 10, "bold"), text_color="#3A8AFF",
                              command=self._tampilkan_qris).pack(side="right", padx=14, pady=8)

        ctk.CTkLabel(pay_card,
                     text="💬  Setelah pembayaran, WhatsApp bukti transfer ke 0812-7064-7744\n"
                          "    Admin akan mengirim kode aktivasi dalam 1×24 jam.",
                     font=FONT_SMALL, text_color=C_MUTED, justify="left").pack(
                         anchor="w", padx=20, pady=(8, 4))

        ctk.CTkButton(pay_card, text="📲  Hubungi Admin via WhatsApp", height=36,
                      fg_color="#1A3A1A", hover_color="#0A2A0A",
                      border_width=1, border_color=C_GREEN,
                      font=FONT_SUB, text_color=C_GREEN,
                      command=lambda: webbrowser.open("https://wa.me/6281270647744")
                      ).pack(fill="x", padx=20, pady=(4, 16))

        if not is_admin:
            # Mode kasir: hanya lihat status lisensi (mengikuti admin utama),
            # semua aksi aktivasi/pembayaran disembunyikan.
            akt_card.pack_forget()
            bayar_card.pack_forget()
            pay_card.pack_forget()
            ctk.CTkLabel(scroll,
                         text="👤 Mode Kasir — status lisensi mengikuti akun admin utama.\n"
                              "Aktivasi, perpanjangan & pembayaran hanya dilakukan oleh admin.",
                         font=FONT_BODY, text_color=C_MUTED, justify="center").pack(pady=10)

        ctk.CTkLabel(scroll,
                     text="⚠  Aplikasi akan TERKUNCI otomatis setelah masa trial habis.\n"
                          "    Aktifkan lisensi untuk terus menggunakan semua fitur tanpa batasan.",
                     font=FONT_BODY, text_color=C_YELLOW,
                     justify="center").pack(pady=10)

    def _pilih_paket_bayar(self, nama, harga, harga_num=None):
        self._buka_dialog_pembayaran(nama, harga, harga_num)

    def _buka_dialog_pembayaran(self, nama_paket, harga_str, harga_num=None):
        dlg = ctk.CTkToplevel(self)
        dlg.title("💳 Pembayaran Paket")
        dlg.configure(fg_color=C_BG)
        dlg.resizable(False, False)
        dlg.transient(self)
        dlg.grab_set()

        ctk.CTkLabel(dlg, text="💳  PEMBAYARAN",
                     font=FONT_SUB, text_color=C_ACCENT2).pack(pady=(14, 2))
        ctk.CTkLabel(dlg, text=f"{nama_paket}  —  {harga_str}",
                     font=("Russo One", 14, "bold"), text_color=C_TEXT).pack(pady=(0, 6))

        qris_path = _qris_file()
        if qris_path:
            try:
                from PIL import Image
                img = Image.open(qris_path)
                img.thumbnail((280, 280), Image.LANCZOS)
                ctk_img = ctk.CTkImage(light_image=img, dark_image=img,
                                        size=(img.width, img.height))
                ctk.CTkLabel(dlg, image=ctk_img, text="").pack(pady=2)
            except Exception:
                ctk.CTkLabel(dlg, text="(Gagal memuat QRIS)",
                             font=FONT_SMALL, text_color=C_MUTED).pack(pady=2)
        else:
            ctk.CTkLabel(dlg, text="(QRIS tidak tersedia — letakkan qris.png di folder app)",
                         font=FONT_SMALL, text_color=C_MUTED).pack(pady=2)

        frame_metode = ctk.CTkFrame(dlg, fg_color=C_PANEL, corner_radius=10)
        frame_metode.pack(fill="x", padx=24, pady=4)
        ctk.CTkLabel(frame_metode, text="Metode Pembayaran:",
                     font=FONT_SUB, text_color=C_ACCENT2).pack(anchor="w", padx=14, pady=(8, 2))
        for m in [
            "🏦 BCA        : 6145375553  a/n Rahmadani",
            "🏦 BRI        : 0256 0109 2349 500  a/n Rahmadani",
            "💚 GoPay/OVO/Dana : 0812-7064-7744 a/n Rahmadani",
        ]:
            ctk.CTkLabel(frame_metode, text=m, font=("Consolas", 11),
                         text_color=C_TEXT, anchor="w").pack(fill="x", padx=14, pady=1)
        ctk.CTkLabel(frame_metode, text="", font=("", 2)).pack()

        frame_btn = ctk.CTkFrame(dlg, fg_color="transparent")
        frame_btn.pack(fill="x", padx=24, pady=(8, 4))
        btn_upload = ctk.CTkButton(frame_btn, text="📤  Upload Bukti Pembayaran", height=38,
                                   fg_color=C_ACCENT2, hover_color="#5A0FCC",
                                   font=("Russo One", 11, "bold"),
                                   command=lambda hn=harga_num: self._upload_bukti(dlg, nama_paket, harga_str, lbl_status, btn_upload, hn))
        btn_upload.pack(side="left", fill="x", expand=True, padx=(0, 6))
        ctk.CTkButton(frame_btn, text="⏳  Bayar Nanti", width=150, height=38,
                      fg_color=C_BTN, hover_color=C_ACCENT2,
                      border_width=1, border_color=C_MUTED,
                      font=FONT_SUB, text_color=C_MUTED,
                      command=dlg.destroy).pack(side="right")

        lbl_status = ctk.CTkLabel(dlg, text="", font=FONT_SMALL, text_color=C_GREEN,
                                   wraplength=460, justify="center")
        lbl_status.pack(pady=(4, 12))

        dlg.geometry("500x620")

    def _upload_bukti(self, dlg, nama_paket, harga_str, lbl_status, btn_upload, harga_num=None):
        import base64, string, shutil
        from tkinter import filedialog
        from firestore_sync import FirestoreClient
        path = filedialog.askopenfilename(
            parent=dlg,
            title="Pilih Bukti Pembayaran",
            filetypes=[("Image files", "*.png *.jpg *.jpeg"), ("All files", "*.*")]
        )
        if not path:
            return

        # Normalize package name for License Generator
        paket_map = {"Bulanan": "1 Bulan", "3 Bulan": "3 Bulan", "Tahunan": "1 Tahun", "LIFETIME": "LIFETIME"}
        paket_norm = paket_map.get(nama_paket, nama_paket)

        # Generate invoice ID
        from datetime import datetime
        ts = datetime.now()
        date_part = ts.strftime("%Y%m%d")
        rand_part = "".join(random.choices(string.ascii_uppercase + string.digits, k=6))
        inv_id = f"INV-{date_part}-{rand_part}"

        # Save image locally instead of inline base64 (Firestore 1MB limit)
        bukti_dir = app_path("invoices_bukti")
        os.makedirs(bukti_dir, exist_ok=True)
        ext = path.rsplit(".", 1)[-1].lower() if "." in path else "jpg"
        bukti_filename = f"{inv_id}.{ext}"
        bukti_path = os.path.join(bukti_dir, bukti_filename)
        try:
            shutil.copy2(path, bukti_path)
        except Exception as e:
            lbl_status.configure(text=f"✖ Gagal salin gambar: {e}", text_color=C_RED)
            return

        # Parse numeric price (if not already provided by promo system)
        if harga_num is None:
            harga_clean = harga_str.split("/")[0].replace("Rp", "").replace(" ", "").replace(".", "")
            try:
                harga_num = int(harga_clean)
            except ValueError:
                harga_num = 0

        uname = self.current_user or ""
        email = get_firebase_auth().get_email() or ""

        inv_data = {
            "id": inv_id,
            "username": uname,
            "email": email,
            "paket": paket_norm,
            "harga": harga_num,
            "status": "WAITING_CONFIRMATION",
            "dibuat": int(ts.timestamp() * 1000),
            "dibayar": 0,
            "confirmedBy": "",
            "kodeLisensi": "",
            "bukti_local": bukti_path,
        }

        try:
            fc = FirestoreClient()
            ok, err_msg = fc.create_invoice(inv_id, inv_data, username=uname)
            if not ok:
                lbl_status.configure(text=f"✖ Gagal simpan: {err_msg}", text_color=C_RED)
                return
        except Exception as e:
            lbl_status.configure(text=f"✖ Gagal simpan: {e}", text_color=C_RED)
            return

        lbl_status.configure(
            text=f"✅  Bukti terkirim!\n"
                 f"Invoice: {inv_id}\n"
                 f"Menunggu konfirmasi admin via License Generator...",
            text_color=C_GREEN)
        btn_upload.configure(state="disabled", text="✓  Terkirim")
        AuditLogger.log(
            action="upload_bukti_bayar",
            username=uname,
            status="success",
            details={"paket": nama_paket, "invoice_id": inv_id})

    def _tampilkan_qris(self):
        """Tampilkan dialog popup gambar QRIS pembayaran."""
        qris_path = _qris_file()
        if not qris_path:
            messagebox.showinfo(
                "🟦 QRIS Tidak Ditemukan",
                "File gambar QRIS belum diatur.\n\n"
                "Letakkan file gambar QRIS Anda dengan nama:\n"
                f"  qris.png\n\n"
                f"Di folder:\n  {APP_BASE_DIR}",
                parent=self)
            return

        dlg = ctk.CTkToplevel(self)
        dlg.title("🟦 Pembayaran via QRIS")
        dlg.configure(fg_color=C_BG)
        dlg.resizable(False, False)
        dlg.transient(self)
        dlg.grab_set()

        ctk.CTkLabel(dlg, text="🟦  SCAN QRIS UNTUK PEMBAYARAN",
                     font=FONT_SUB, text_color="#3A8AFF").pack(pady=(18, 4))
        ctk.CTkLabel(dlg, text="Scan kode QR di bawah menggunakan aplikasi e-wallet atau m-banking Anda.",
                     font=FONT_SMALL, text_color=C_MUTED).pack(pady=(0, 12))

        try:
            from PIL import Image
            img = Image.open(qris_path)
            # Resize proporsional agar pas di dialog (max 400x400)
            img.thumbnail((400, 400), Image.LANCZOS)
            ctk_img = ctk.CTkImage(light_image=img, dark_image=img, size=(img.width, img.height))
            lbl_img = ctk.CTkLabel(dlg, image=ctk_img, text="")
            lbl_img.pack(padx=24, pady=(0, 12))
            dlg.geometry(f"{img.width + 48}x{img.height + 160}")
        except Exception as e:
            ctk.CTkLabel(dlg, text=f"Gagal memuat gambar QRIS:\n{e}",
                         font=FONT_BODY, text_color=C_RED).pack(padx=24, pady=20)
            dlg.geometry("400x180")

        ctk.CTkLabel(dlg, text="GoPay / OVO / Dana / QRIS Bank — a/n Rahmadani",
                     font=("Consolas", 11, "bold"), text_color=C_YELLOW).pack(pady=(0, 4))
        ctk.CTkButton(dlg, text="✖  Tutup", width=120, height=34,
                      fg_color=C_BTN, hover_color=C_ACCENT2,
                      border_width=1, border_color=C_MUTED,
                      font=FONT_SUB, text_color=C_MUTED,
                      command=dlg.destroy).pack(pady=(4, 16))

    def _printer_tersedia(self):
        """Cek printer terkonfigurasi & benar-benar bisa dipakai.
        Return (tersedia, tipe, alamat). Untuk bluetooth alamat kosong tetap
        dicari otomatis (RPP02N dkk ter-pair) dan disimpan ke config."""
        try:
            cfg = ConfigManager.get("printer_settings", {}) or {}
        except Exception:
            cfg = {}
        ptype = str(cfg.get("type", "file") or "file")
        addr = str(cfg.get("address", "") or "").strip()
        if ptype == "bluetooth":
            if not addr:
                try:
                    mac = _ble_find_printer(timeout=4)
                    if mac:
                        try:
                            cfg2 = ConfigManager.load()
                            cfg2.setdefault("printer_settings", {})
                            cfg2["printer_settings"].update(
                                {"type": "bluetooth", "address": mac})
                            ConfigManager.save(cfg2)
                        except Exception:
                            pass
                        addr = mac
                except Exception:
                    addr = ""
            return bool(addr), ptype, addr
        if ptype in ("usb", "network"):
            return bool(addr) and ESCPOS_AVAILABLE, ptype, addr
        return False, ptype, ""

    def _print_receipt(self, trans: dict = None):
        """Show dialog with receipt formatted for thermal printer (40 columns)."""
        dlg = ctk.CTkToplevel(self)
        dlg.title("🖨 Cetak Struk")
        dlg.configure(fg_color=C_BG)
        dlg.geometry("480x600")
        dlg.transient(self)
        dlg.grab_set()
        txt = ctk.CTkTextbox(dlg, fg_color=C_CARD, text_color=C_TEXT,
                              font=("Courier New", 10), wrap="none",
                              width=440, height=400)
        txt.pack(padx=20, pady=(16, 8), fill="both", expand=True)
        # Build receipt content (40 cols)
        lines = []
        lines.append("=" * 40)
        lines.append("      RR BILLING PRO")
        lines.append("      STRUK TRANSAKSI")
        lines.append("=" * 40)
        now = datetime.now().strftime("%d/%m/%Y %H:%M")
        lines.append(f"Tanggal : {now}")
        lines.append(f"Kasir   : {self.current_user or '-'}")
        if trans:
            pc = trans.get("pc", trans.get("kursi", "-"))
            paket = trans.get("paket", "-")
            durasi = trans.get("durasi", trans.get("menit", 0))
            harga = trans.get("harga", trans.get("biaya", 0))
            lines.append("-" * 40)
            lines.append(f"PC/Kursi: {pc}")
            lines.append(f"Paket   : {paket}")
            lines.append(f"Durasi  : {durasi} menit")
            lines.append(f"Total   : Rp {harga:,.0f}".replace(",", "."))
        lines.append("-" * 40)
        lines.append("      TERIMA KASIH")
        lines.append("      Silakan datang kembali")
        lines.append("=" * 40)
        txt.insert("1.0", "\n".join(lines))
        txt.configure(state="disabled")
        btn_row = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_row.pack(fill="x", padx=20, pady=(0, 16))
        def _cetak():
            printer_cfg = ConfigManager.get("printer_settings", {})
            printer_type = printer_cfg.get("type", "file")
            address = printer_cfg.get("address", "")
            if printer_type == "bluetooth":
                threading.Thread(target=self._print_via_escpos,
                                 args=("\n".join(lines), printer_type, address),
                                 daemon=True).start()
            elif printer_type in ("usb", "network") and ESCPOS_AVAILABLE:
                threading.Thread(target=self._print_via_escpos,
                                 args=("\n".join(lines), printer_type, address),
                                 daemon=True).start()
            else:
                self._print_to_file("\n".join(lines))
            messagebox.showinfo("Cetak", "Struk dikirim ke printer / file.")
        ctk.CTkButton(btn_row, text="🖨  Cetak via Bluetooth", width=180, height=36,
                      fg_color=C_ACCENT2, hover_color="#5A0FCC",
                      font=("Russo One", 10, "bold"), text_color="white",
                      command=_cetak).pack(side="left", padx=4)
        ctk.CTkButton(btn_row, text="💾  Cetak ke File", width=140, height=36,
                      fg_color=C_BTN, hover_color=C_ACCENT2,
                      border_width=1, border_color=C_ACCENT2,
                      font=("Russo One", 10, "bold"), text_color=C_ACCENT2,
                      command=lambda: (self._print_to_file("\n".join(lines)),
                                       messagebox.showinfo("Cetak", "Struk disimpan di folder receipts/"))).pack(side="left", padx=4)
        ctk.CTkButton(btn_row, text="✖ Tutup", width=80, height=36,
                      fg_color=C_BTN, hover_color=C_RED,
                      border_width=1, border_color=C_MUTED,
                      font=("Russo One", 10, "bold"), text_color=C_MUTED,
                      command=dlg.destroy).pack(side="right", padx=4)

    def _print_via_escpos(self, text: str, ptype: str, address: str):
        """Print receipt via python-escpos / Bluetooth BLE (RPP02N)."""
        try:
            if ptype == "bluetooth":
                # Printer Bluetooth BLE (RPP02N dkk): ESC/POS langsung ditulis
                # ke karakteristik RX (ISSC Transparent UART), tanpa escpos.
                try:
                    mac = address.strip() if address.strip() else _ble_find_printer()
                    if not mac:
                        _LOGGER.warning("BLE print: printer Bluetooth tidak ditemukan")
                        self._print_to_file(text)
                        return
                    if mac != address.strip():
                        try:
                            cfg = ConfigManager.load()
                            cfg.setdefault("printer_settings", {})
                            cfg["printer_settings"].update({"type": "bluetooth", "address": mac})
                            ConfigManager.save(cfg)
                            print(f"[PRINT] Printer BLE disimpan: {mac}", flush=True)
                        except Exception:
                            pass
                    _ble_write_printer(mac, _ble_escpos_bytes(text))
                    print(f"[PRINT] BLE terkirim ke {mac}", flush=True)
                    return
                except ImportError:
                    _LOGGER.warning("BLE print: 'bleak' tidak terinstal")
                    self._print_to_file(text)
                    return
                except Exception as e:
                    _LOGGER.warning("BLE print error: %s", e)
                    self._print_to_file(text)
                    return
            if ptype == "network":
                from escpos.printer import Network
                p = Network(address)
            elif ptype == "usb":
                from escpos.printer import Usb
                parts = address.split(":")
                if len(parts) == 2:
                    p = Usb(int(parts[0], 16), int(parts[1], 16))
                else:
                    p = Usb(0x0416, 0x5011)
            else:
                return
            p.text(text + "\n\n")
            p.cut()
        except Exception as e:
            _LOGGER.warning("ESC/POS print error: %s", e)
            self._print_to_file(text)

    def _print_to_file(self, text: str):
        """Fallback: save receipt to text file."""
        try:
            receipt_dir = app_path("receipts")
            os.makedirs(receipt_dir, exist_ok=True)
            fname = f"struk_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            fpath = os.path.join(receipt_dir, fname)
            with open(fpath, "w", encoding="utf-8") as f:
                f.write(text)
            print(f"[PRINT] Receipt saved to {fpath}")
        except Exception as e:
            _LOGGER.error("Print to file error: %s", e)

    def _start_update_checker(self, interval_hours: int = 6):
        """Background loop: cek update via manifest URL atau Git setiap interval_hours."""
        manifest = ConfigManager.get('update_manifest_url') or ""
        # Delay awal agar UI tidak kena spam saat startup
        time.sleep(5)
        while True:
            try:
                if manifest and manifest.strip():
                    # Cek via manifest URL (hanya deteksi, tidak mengunduh otomatis)
                    try:
                        mf = self._detect_update()
                        if mf is not None:
                            msg = (f"Pembaruan v{mf.get('version')} tersedia "
                                   f"(versi Anda v{APP_VERSION}).\n\n"
                                   "Klik 'Download & Install' di sidebar untuk memasangnya.")
                            self.after(0, lambda m=msg: messagebox.showinfo("Pembaruan Tersedia", m))
                    except Exception:
                        pass
            except Exception:
                pass
            try:
                time.sleep(interval_hours * 3600)
            except Exception:
                break

    def _check_for_updates_background(self):
        """Periodic background update check with sidebar notification dot. Runs every 6 hours."""
        last_check_key = "_last_update_check_time"
        interval = 6 * 3600
        while True:
            try:
                now_ts = time.time()
                last_check = getattr(self, last_check_key, 0)
                if now_ts - last_check < interval:
                    time.sleep(300)
                    continue
                setattr(self, last_check_key, now_ts)
                # Also store in config
                cfg = ConfigManager.load()
                cfg["update_last_check_ts"] = now_ts
                ConfigManager.save(cfg)
                manifest = ConfigManager.get('update_manifest_url') or ""
                has_update = False
                if manifest and manifest.strip():
                    try:
                        mf = self._detect_update()
                        if mf is not None:
                            has_update = True
                    except Exception:
                        pass
                # Update sidebar notification dot
                self.after(0, lambda hu=has_update: self._set_update_notification(hu))
            except Exception:
                pass
            try:
                time.sleep(interval)
            except Exception:
                break

    def _set_update_notification(self, has_update: bool):
        """Show/hide yellow notification dot on the update button."""
        row_update = getattr(self, "_row_update", None)
        if row_update:
            for w in row_update.winfo_children():
                if isinstance(w, ctk.CTkLabel) and has_update:
                    w.configure(text="🟡")
                elif isinstance(w, ctk.CTkLabel) and not has_update:
                    w.configure(text="🔄")

    DEFAULT_MANIFEST_URL = ("https://github.com/dedekemoking-commits/rr_billing_pro_windows/"
                            "releases/latest/download/manifest.json")

    def _download_and_install_update(self):
        """Download & install update EXE RR Billing Pro dengan aman:
        manifest (tanda tangan RSA) -> konfirmasi -> unduh (cek sha256) -> ganti exe -> restart.
        URL manifest diisi otomatis (URL resmi) bila belum diatur.
        """
        manifest = str(ConfigManager.get('update_manifest_url') or "").strip()
        if not manifest:
            manifest = self.DEFAULT_MANIFEST_URL
            try:
                ConfigManager.set('update_manifest_url', manifest)
            except Exception:
                pass
        threading.Thread(target=self._download_install_thread,
                         args=(manifest,), daemon=True).start()

    def _download_install_thread(self, manifest_url: str):
        try:
            from scripts import check_update

            # 1) Fetch + verifikasi tanda tangan manifest
            mf = check_update.fetch_manifest(manifest_url)
            if not check_update.verify_manifest(mf, ConfigManager.get('update_pubkey_path') or None):
                self.after(0, lambda: messagebox.showerror(
                    "Gagal Verifikasi",
                    "Tanda tangan manifest TIDAK VALID. Update dibatalkan.\n\n"
                    "Kemungkinan file bukan dari developer RR Billing Pro."))
                return

            new_ver = str(mf.get('version', '?'))
            if new_ver == str(APP_VERSION):
                self.after(0, lambda: messagebox.showinfo(
                    "Sudah Terbaru", f"Versi {APP_VERSION} sudah terpasang."))
                return

            # 2) Konfirmasi ke user sebelum mengunduh
            confirmed = []
            ev = threading.Event()

            def _ask():
                r = messagebox.askyesno(
                    "⬇ Konfirmasi Update",
                    f"Pembaruan tersedia:\n\n"
                    f"  Versi baru  : v{new_ver}\n"
                    f"  Versi Anda  : v{APP_VERSION}\n\n"
                    f"Unduh dan pasang sekarang?\n"
                    f"(Aplikasi akan menutup lalu restart otomatis dengan versi baru)")
                confirmed.append(bool(r))
                ev.set()

            self.after(0, _ask)
            if not ev.wait(timeout=120):
                return
            if not confirmed or not confirmed[0]:
                return

            # 3) Dialog progress download
            dlg = ctk.CTkToplevel(self)
            dlg.title("⬇ Mengunduh Update")
            dlg.configure(fg_color=C_BG)
            dlg.geometry("440x150")
            dlg.transient(self)
            dlg.grab_set()
            ctk.CTkLabel(dlg, text="Mengunduh pembaruan…", font=FONT_SUB,
                         text_color=C_ACCENT2).pack(pady=(16, 8))
            bar = ctk.CTkProgressBar(dlg, width=380, height=18,
                                     fg_color=C_BTN, progress_color=C_ACCENT2)
            bar.pack(padx=20, pady=(0, 8))
            bar.set(0)
            st = ctk.CTkLabel(dlg, text="Memulai…", font=FONT_SMALL, text_color=C_MUTED)
            st.pack(pady=(0, 16))

            def _prog(done, total):
                try:
                    self.after(0, lambda d=done, t=total: (
                        bar.set((d / t) if t > 0 else 0),
                        st.configure(text=f"{d // 1024} KB / {max(t, 1) // 1024} KB")))
                except Exception:
                    pass

            # 4) Unduh + verifikasi sha256
            tmp = tempfile.mkdtemp(prefix='rr_update_')
            asset_path = os.path.join(tmp, "RRBILLINGPRO.exe")
            try:
                check_update.download_asset(mf['asset_url'], asset_path,
                                            mf.get('sha256'), _prog)
            except Exception as e:
                self.after(0, lambda err=str(e): (
                    dlg.destroy(),
                    messagebox.showerror("Unduh Gagal", str(err))))
                return

            # 5) Pasang: ganti exe & restart (batch mandiri saat EXE, helper saat source)
            app_exe = sys.executable if getattr(sys, "frozen", False) else os.path.abspath(__file__)
            try:
                pesan = check_update.launch_updater(app_exe, asset_path)
            except Exception as e:
                self.after(0, lambda err=str(e): (
                    dlg.destroy(),
                    messagebox.showerror("Gagal Pasang Update", str(err))))
                return

            self.after(0, lambda: (
                dlg.destroy(),
                messagebox.showinfo(
                    "✅ Update Siap",
                    f"v{new_ver} terunduh dan diverifikasi.\n\n{pesan}")))
            if getattr(sys, "frozen", False):
                # Tutup aplikasi; batch yang menunggu akan mengganti exe & restart
                self.after(3000, self.quit)
        except Exception as e:
            self.after(0, lambda err=str(e): messagebox.showerror(
                "Download & Install", str(err)))

    def _do_restart(self):
        """Restart aplikasi dengan benar."""
        try:
            if sys.platform == "win32":
                if getattr(sys, "frozen", False):
                    restart_cmd = [sys.executable]
                else:
                    restart_cmd = [sys.executable, os.path.abspath(__file__)]
                subprocess.Popen(restart_cmd, close_fds=True, **subprocess_no_window_kwargs())
            else:
                # Unix/Linux/Mac
                python_exe = sys.executable
                script_path = os.path.abspath(__file__)
                subprocess.Popen([python_exe, script_path], close_fds=True)
        except Exception as e:
            AuditLogger.log(
                action="restart_error",
                username=self.current_user or "system",
                status="error",
                details={"error": str(e)}
            )
        finally:
            # Keluar dari aplikasi lama
            time.sleep(0.5)
            self.quit()

    def _rebuild_sidebar_lic(self):
        self._refresh_license_ui()

    def _refresh_license_ui(self):
        """Refresh license status labels after activation."""
        status = LicenseManager.get_status(current_user=self._resolve_license_user())
        lic_color = C_GREEN if status["status"] == "active" else C_YELLOW if status["status"] == "trial" else C_RED
        if getattr(self, "lbl_sidebar_license_status", None) is not None:
            self.lbl_sidebar_license_status.configure(text=status["pesan"], text_color=lic_color)

        # Re-enable semua tab jika lisensi aktif/trial
        if status["status"] in ["active", "trial"]:
            for tab_key, btn in self.nav_btns.items():
                btn.configure(state="normal", text_color=C_TEXT if tab_key != self.current_tab else "white")
        
        if self.frames.get("aktivasi"):
            for child in self.frames["aktivasi"].winfo_children():
                child.destroy()
            self._setup_aktivasi()

    # ──────────────────────────────────────────────────────────────────────────
    #  APPLY THEME — reconfigure all frames with current theme colors
    # ──────────────────────────────────────────────────────────────────────────
    def _apply_theme(self, theme_name):
        ConfigManager.set("app_theme", theme_name)
        _load_theme()
        self.sidebar.configure(fg_color=C_PANEL)
        self.content.configure(fg_color=C_BG)
        for f in self.frames.values():
            f.configure(fg_color=C_BG)
        if self.current_tab == "profil":
            self._setup_profil()
            self._show_tab("profil")

    def _on_theme_change(self, choice):
        self._apply_theme(choice)
        messagebox.showinfo("✅ Tema Berubah", f"Tema diganti ke '{choice}'.")

    # ──────────────────────────────────────────────────────────────────────────
    #  BACKGROUND IMAGE — wallpaper di belakang Dashboard TV & Warnet
    # ──────────────────────────────────────────────────────────────────────────
    def _apply_bg_image(self, tab_key=None):
        import traceback
        from PIL import ImageTk

        try:
            target_tabs = [tab_key] if tab_key else ["dashboard", "warnet"]

            for t in target_tabs:
                f = self.frames.get(t)
                if f:
                    f.configure(fg_color=C_BG)

            path = ConfigManager.get("app_bg_image", "")
            self._bg_image_path = path
            if not path or not os.path.isfile(path):
                return

            Image.MAX_IMAGE_PIXELS = None

            for t in target_tabs:
                s_attr = "scroll_dash" if t == "dashboard" else "scroll_warnet"
                s = getattr(self, s_attr, None)
                if not s:
                    continue
                try:
                    canvas = s._parent_canvas
                except AttributeError:
                    continue

                self.update_idletasks()
                cw = canvas.winfo_width() or 1280
                bbox = canvas.bbox("all")
                ch = max(canvas.winfo_height() or 800, bbox[3] + 50 if bbox else 800)
                pil_img = Image.open(path)
                pil_img = pil_img.resize((cw, ch), Image.LANCZOS)
                if pil_img.mode not in ("RGB", "RGBA"):
                    pil_img = pil_img.convert("RGBA")
                photo = ImageTk.PhotoImage(pil_img)

                canvas.delete("_bg_img")
                canvas.create_image(0, 0, image=photo, anchor="nw",
                                    tags=("_bg_img",))
                canvas.tag_lower("_bg_img")
                self._bg_photo = photo

        except Exception as e:
            traceback.print_exc()
            print(f"Gagal load bg image: {e}")

    def _pilih_bg_image(self):
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            title="Pilih Gambar Background",
            filetypes=[("Image files", "*.png *.jpg *.jpeg *.bmp *.gif")]
        )
        if not path:
            return
        if not os.path.isfile(path):
            messagebox.showerror("Error", "File tidak ditemukan.")
            return
        ConfigManager.set("app_bg_image", path)
        self._bg_image_path = path
        tab = self.current_tab if self.current_tab in ("dashboard", "warnet") else "dashboard"
        self._apply_bg_image(tab)
        if tab != self.current_tab:
            messagebox.showinfo("Info", "Buka tab Dashboard TV atau Warnet untuk melihat background.")

    def _hapus_bg_image(self):
        ConfigManager.set("app_bg_image", "")
        self._bg_image_path = ""
        self._bg_photo = None
        for t in ["dashboard", "warnet"]:
            s_attr = "scroll_dash" if t == "dashboard" else "scroll_warnet"
            s = getattr(self, s_attr, None)
            if s:
                try:
                    s._parent_canvas.delete("_bg_img")
                except AttributeError:
                    pass
            f = self.frames.get(t)
            if f:
                f.configure(fg_color=C_BG)

    # ══════════════════════════════════════════════════════════════════════════
    #  TAB 6: Profil
    # ══════════════════════════════════════════════════════════════════════════
    def _setup_profil(self):
        f = self.frames["profil"]
        for w in f.winfo_children():
            w.destroy()
        hdr = ctk.CTkFrame(f, fg_color=C_PANEL, height=54, corner_radius=0)
        hdr.pack(fill="x")
        ctk.CTkLabel(hdr, text="👤  PROFIL & MANAJEMEN USER",
                     font=FONT_TITLE, text_color=C_ACCENT).pack(side="left", padx=18, pady=14)

        try:
            self._build_profil_content(f)
        except Exception as e:
            _LOGGER.exception("Gagal _setup_profil: %s", e)
            ctk.CTkLabel(f, text=f"⚠ Gagal memuat profil: {e}",
                         font=FONT_BODY, text_color=C_RED).pack(pady=40)

    def _build_profil_content(self, f):
        scroll = ctk.CTkScrollableFrame(f, fg_color=C_BG)
        scroll.pack(fill="both", expand=True, padx=20, pady=16)

        card = ctk.CTkFrame(scroll, fg_color=C_CARD, corner_radius=16)
        card.pack(fill="x", pady=(0, 16))

        # ── Logo profil: coba logo.png, fallback teks ─────────────────────────
        ctk_img_profil = load_ctk_image(size=(180, 68))
        if ctk_img_profil:
            logo_container = ctk.CTkFrame(card, fg_color=C_CARD, corner_radius=12,
                                           width=190, height=76)
            logo_container.pack(pady=(24, 8))
            logo_container.pack_propagate(False)
            ctk.CTkLabel(logo_container, text="", image=ctk_img_profil).place(
                relx=0.5, rely=0.5, anchor="center")
        else:
            ctk.CTkLabel(card, text="🎮  RR BILLING PRO",
                         font=("Russo One", 18, "bold"), text_color=C_ACCENT).pack(pady=(24, 4))

        ctk.CTkLabel(card, text="RR BILLING PRO" if ctk_img_profil else "",
                     font=("Russo One", 16, "bold"), text_color=C_ACCENT).pack(pady=(4 if ctk_img_profil else 0, 0))
        ctk.CTkLabel(card, text="Sistem Billing Rental PlayStation & TV",
                     font=FONT_BODY, text_color=C_MUTED).pack()
        sep = ctk.CTkFrame(card, height=1, fg_color=C_BORDER)
        sep.pack(fill="x", padx=30, pady=16)
        for label, val in [
            ("Versi",      APP_VERSION),
            ("Developer",  "RR CCTV"),
            ("Kontak",     "0812-7064-7744"),
            ("User Aktif", f"{self.current_user} [{self.current_role}]"),
            ("Lisensi",    LicenseManager.get_status(current_user=self._resolve_license_user())["pesan"]),
        ]:
            row = ctk.CTkFrame(card, fg_color="transparent")
            row.pack(fill="x", padx=30, pady=3)
            ctk.CTkLabel(row, text=f"{label}:", font=FONT_SUB,
                         text_color=C_MUTED, width=120, anchor="w").pack(side="left")
            ctk.CTkLabel(row, text=val, font=FONT_BODY, text_color=C_TEXT).pack(side="left")
        alamat_row = ctk.CTkFrame(card, fg_color="transparent")
        alamat_row.pack(fill="x", padx=30, pady=3)
        ctk.CTkLabel(alamat_row, text="Alamat:", font=FONT_SUB,
                     text_color=C_MUTED, width=120, anchor="w").pack(side="left")
        lbl_alamat_web = ctk.CTkLabel(alamat_row, text="🔗 rrcctv.online",
                                      font=FONT_BODY, text_color=C_ACCENT2, cursor="hand2")
        lbl_alamat_web.pack(side="left")
        lbl_alamat_web.bind("<Button-1>",
                            lambda e: webbrowser.open("https://rrcctv.online"))
        ctk.CTkLabel(card, text="© 2026 RR CCTV — All Rights Reserved",
                     font=FONT_SMALL, text_color=C_MUTED).pack(pady=(16, 24))

        # ── PROFIL RENTAL USER ───────────────────────────────────────────
        rental_card = ctk.CTkFrame(scroll, fg_color=C_PANEL, corner_radius=14)
        rental_card.pack(fill="x", pady=(0, 16))
        
        ctk.CTkLabel(rental_card, text="🏢  PROFIL RENTAL ANDA",
                     font=("Russo One", 12, "bold"), text_color=C_ACCENT).pack(anchor="w", padx=20, pady=(16, 12))
        
        # Get rental profile data
        cfg = ConfigManager.load()
        profil_semua = cfg.get("profil_rental", {})
        profil_user = profil_semua.get(self.current_user, {}) or {}

        # ── Form isi langsung (tanpa dialog) ─────────────────────────────
        def _field_profil(parent, label, value):
            ctk.CTkLabel(parent, text=label, font=FONT_LABEL,
                         text_color=C_MUTED, anchor="w").pack(anchor="w", padx=20, pady=(4, 2))
            e = ctk.CTkEntry(parent, fg_color=C_BTN, text_color=C_TEXT,
                             border_color=C_BORDER, font=FONT_BODY, height=34)
            e.insert(0, value)
            e.pack(fill="x", padx=20, pady=(0, 6))
            return e

        e_nama_rental = _field_profil(rental_card, "🏪 Nama Rental PS",
                                      profil_user.get("nama_rental", ""))
        e_nama_pemilik = _field_profil(rental_card, "👤 Nama Pemilik",
                                       profil_user.get("nama_pemilik", ""))
        e_hp = _field_profil(rental_card, "📱 No HP / WhatsApp",
                             str(profil_user.get("hp") or profil_user.get("no_hp") or ""))
        e_email = _field_profil(rental_card, "📧 Email / Gmail",
                                profil_user.get("email", ""))

        ctk.CTkLabel(rental_card, text="📍 Alamat Tempat", font=FONT_LABEL,
                     text_color=C_MUTED, anchor="w").pack(anchor="w", padx=20, pady=(4, 2))
        e_alamat = ctk.CTkTextbox(rental_card, height=64, fg_color=C_BTN,
                                  text_color=C_TEXT, border_color=C_BORDER,
                                  border_width=1, font=FONT_BODY)
        e_alamat.pack(fill="x", padx=20, pady=(0, 6))
        e_alamat.insert("1.0", profil_user.get("alamat", ""))

        def _simpan_profil_form():
            _hp = sanitize_text(e_hp.get())
            data = {
                "nama_rental": sanitize_text(e_nama_rental.get()),
                "nama_pemilik": sanitize_text(e_nama_pemilik.get()),
                "hp": _hp,
                "no_hp": _hp,
                "email": sanitize_text(e_email.get()).lower(),
                "alamat": sanitize_text(e_alamat.get("1.0", "end")),
            }
            self._simpan_profil_rental(data)
            self._qr_push_menu_bg()
            self._booking_push_tv_status()

        ctk.CTkButton(rental_card, text="💾  Simpan Profil", height=36,
                      fg_color=C_ACCENT2, font=("Russo One", 10, "bold"),
                      text_color="white",
                      command=_simpan_profil_form).pack(anchor="w", padx=20, pady=(8, 2))
        ctk.CTkLabel(rental_card,
                     text="Nama rental tampil di popup kanan atas TV "
                          "(berlaku setelah aktivasi lisensi LIFETIME).",
                     font=FONT_SMALL, text_color=C_MUTED).pack(anchor="w", padx=20, pady=(0, 10))

        # ── PIN KEAMANAN (Hapus TV & Kursi) ──────────────────────────
        pin_card = ctk.CTkFrame(scroll, fg_color=C_PANEL, corner_radius=14)
        pin_card.pack(fill="x", pady=(0, 16))

        ctk.CTkLabel(pin_card, text="🔒  PIN KEAMANAN (HAPUS TV & KURSI)",
                     font=("Russo One", 12, "bold"), text_color=C_ACCENT).pack(anchor="w", padx=20, pady=(16, 4))
        ctk.CTkLabel(pin_card, text="PIN wajib dimasukkan admin saat menghapus TV / kursi warnet.",
                     font=FONT_SMALL, text_color=C_MUTED).pack(anchor="w", padx=20, pady=(0, 8))

        pin_row = ctk.CTkFrame(pin_card, fg_color="transparent")
        pin_row.pack(fill="x", padx=20, pady=(0, 8))

        def _refresh_pin_status():
            ada = bool(self._get_pin_hapus())
            lbl_pin_status.configure(
                text="Status: AKTIF" if ada else "Status: BELUM DIBUAT",
                text_color=C_GREEN if ada else C_RED)
            btn_pin_set.configure(text="🔑 Ubah PIN" if ada else "🔑 Buat PIN")
            btn_pin_lihat.configure(state="normal" if ada else "disabled")

        lbl_pin_status = ctk.CTkLabel(pin_row, text="", font=FONT_SUB,
                                      text_color=C_GREEN, width=150, anchor="w")
        lbl_pin_status.pack(side="left")
        btn_pin_set = ctk.CTkButton(pin_row, text="🔑 Buat PIN", height=32, width=120,
                                    fg_color=C_ACCENT2, font=("Russo One", 9, "bold"),
                                    command=self._dialog_set_pin_hapus)
        btn_pin_set.pack(side="left", padx=4)
        btn_pin_lihat = ctk.CTkButton(pin_row, text="👁 Lihat PIN", height=32, width=110,
                                      fg_color=C_BTN, hover_color=C_ACCENT2,
                                      border_width=1, border_color=C_YELLOW,
                                      font=("Russo One", 9, "bold"), text_color=C_YELLOW,
                                      command=self._dialog_lihat_pin_hapus)
        btn_pin_lihat.pack(side="left", padx=4)
        ctk.CTkButton(pin_row, text="🗑 Hapus PIN", height=32, width=110,
                      fg_color=C_BTN, hover_color=C_RED,
                      border_width=1, border_color=C_RED,
                      font=("Russo One", 9, "bold"), text_color=C_RED,
                      command=self._hapus_pin_hapus).pack(side="left", padx=4)
        _refresh_pin_status()

        # ── TEMA APLIKASI ─────────────────────────────────────────────
        theme_card = ctk.CTkFrame(scroll, fg_color=C_PANEL, corner_radius=14)
        theme_card.pack(fill="x", pady=(0, 16))

        ctk.CTkLabel(theme_card, text="🎨  TEMA APLIKASI",
                     font=("Russo One", 12, "bold"), text_color=C_ACCENT).pack(anchor="w", padx=20, pady=(16, 12))

        theme_row = ctk.CTkFrame(theme_card, fg_color="transparent")
        theme_row.pack(fill="x", padx=20, pady=(0, 16))

        ctk.CTkLabel(theme_row, text="Pilih Tema:", font=FONT_LABEL,
                     text_color=C_MUTED, width=100, anchor="w").pack(side="left")

        self.theme_var = ctk.StringVar(value=_load_theme())
        theme_menu = ctk.CTkOptionMenu(theme_row, values=list(THEMES.keys()),
                                        variable=self.theme_var,
                                        fg_color=C_BTN, button_color=C_ACCENT2,
                                        button_hover_color=C_ACCENT,
                                        text_color=C_TEXT, font=FONT_BODY,
                                        dropdown_fg_color=C_CARD,
                                        dropdown_text_color=C_TEXT,
                                        command=self._on_theme_change)
        theme_menu.pack(side="left", padx=10)

        bg_row = ctk.CTkFrame(theme_card, fg_color="transparent")
        bg_row.pack(fill="x", padx=20, pady=(0, 8))

        ctk.CTkLabel(bg_row, text="Background:", font=FONT_LABEL,
                     text_color=C_MUTED, width=100, anchor="w").pack(side="left")

        ctk.CTkButton(bg_row, text="📁 Pilih Gambar", height=32,
                      fg_color=C_BTN, hover_color=C_ACCENT2,
                      border_width=1, border_color=C_ACCENT2,
                      font=("Russo One", 9, "bold"), text_color=C_ACCENT2,
                      command=self._pilih_bg_image).pack(side="left", padx=4)

        ctk.CTkButton(bg_row, text="✕ Hapus", height=32,
                      fg_color=C_BTN, hover_color=C_RED,
                      border_width=1, border_color=C_RED,
                      font=("Russo One", 9, "bold"), text_color=C_RED,
                      command=self._hapus_bg_image).pack(side="left", padx=4)

        # ── BACKUP & RESTORE ────────────────────────────────────────────
        backup_card = ctk.CTkFrame(scroll, fg_color=C_PANEL, corner_radius=14)
        backup_card.pack(fill="x", pady=(0, 16))

        ctk.CTkLabel(backup_card, text="💾  BACKUP & RESTORE",
                     font=("Russo One", 12, "bold"), text_color=C_ACCENT).pack(anchor="w", padx=20, pady=(16, 12))

        backup_row = ctk.CTkFrame(backup_card, fg_color="transparent")
        backup_row.pack(fill="x", padx=20, pady=(0, 8))
        ctk.CTkButton(backup_row, text="📤 Export Backup", height=34, width=150,
                      fg_color=C_ACCENT2, font=("Russo One", 9, "bold"),
                      command=self._export_backup).pack(side="left", padx=4)
        ctk.CTkButton(backup_row, text="📥 Import Backup", height=34, width=150,
                      fg_color=C_BTN, hover_color=C_ACCENT2,
                      border_width=1, border_color=C_ACCENT2,
                      font=("Russo One", 9, "bold"), text_color=C_ACCENT2,
                      command=self._import_backup).pack(side="left", padx=4)
        ctk.CTkButton(backup_row, text="🔄 Reset Data", height=34, width=150,
                      fg_color=C_RED, hover_color="#7A1A1A",
                      font=("Russo One", 9, "bold"),
                      command=self._reset_all_data).pack(side="left", padx=4)
        self.lbl_backup_status = ctk.CTkLabel(backup_card, text="", font=FONT_SMALL, text_color=C_GREEN)
        self.lbl_backup_status.pack(anchor="w", padx=20, pady=(0, 12))

        # ── PRINTER SETTINGS ──────────────────────────────────────────
        printer_card = ctk.CTkFrame(scroll, fg_color=C_PANEL, corner_radius=14)
        printer_card.pack(fill="x", pady=(0, 16))

        ctk.CTkLabel(printer_card, text="🖨  PRINTER SETTINGS",
                     font=("Russo One", 12, "bold"), text_color=C_ACCENT).pack(anchor="w", padx=20, pady=(16, 12))

        printer_cfg = ConfigManager.get("printer_settings", {})
        pr_type = printer_cfg.get("type", "file")
        pr_addr = printer_cfg.get("address", "")

        pr_row1 = ctk.CTkFrame(printer_card, fg_color="transparent")
        pr_row1.pack(fill="x", padx=20, pady=(0, 8))
        ctk.CTkLabel(pr_row1, text="Tipe Printer:", font=FONT_LABEL,
                     text_color=C_MUTED, width=110, anchor="w").pack(side="left")
        pr_type_var = ctk.StringVar(value=pr_type)
        pr_type_menu = ctk.CTkOptionMenu(pr_row1, values=["file", "usb", "network", "bluetooth"],
                                          variable=pr_type_var,
                                          fg_color=C_BTN, button_color=C_ACCENT2,
                                          button_hover_color=C_ACCENT,
                                          text_color=C_TEXT, font=FONT_BODY,
                                          dropdown_fg_color=C_CARD,
                                          dropdown_text_color=C_TEXT)
        pr_type_menu.pack(side="left", padx=10)

        pr_row2 = ctk.CTkFrame(printer_card, fg_color="transparent")
        pr_row2.pack(fill="x", padx=20, pady=(0, 8))
        ctk.CTkLabel(pr_row2, text="Alamat/Port:", font=FONT_LABEL,
                     text_color=C_MUTED, width=110, anchor="w").pack(side="left")
        pr_addr_entry = ctk.CTkEntry(pr_row2, placeholder_text="IP:port atau USB vid:pid",
                                      fg_color=C_BTN, text_color=C_ACCENT,
                                      border_color=C_BORDER, font=FONT_BODY,
                                      height=30, width=200)
        pr_addr_entry.pack(side="left", padx=10)
        if pr_addr:
            pr_addr_entry.insert(0, pr_addr)

        def _save_printer_settings():
            cfg_save = ConfigManager.load()
            cfg_save["printer_settings"] = {
                "type": pr_type_var.get(),
                "address": pr_addr_entry.get().strip(),
            }
            ConfigManager.save(cfg_save)
            messagebox.showinfo("✅ Tersimpan", "Pengaturan printer berhasil disimpan.")

        def _test_print():
            cfg_test = ConfigManager.load()
            ps = cfg_test.get("printer_settings", {})
            ptype = ps.get("type", "file")
            addr = ps.get("address", "")
            test_lines = []
            test_lines.append("=" * 40)
            test_lines.append("   TEST PRINT RR BILLING PRO")
            test_lines.append("=" * 40)
            test_lines.append(f"Tipe  : {ptype}")
            test_lines.append(f"Alamat: {addr or '-'}")
            test_lines.append(f"Waktu : {datetime.now().strftime('%d/%m/%Y %H:%M')}")
            test_lines.append("-" * 40)
            test_lines.append("Jika terbaca, printer berfungsi!")
            test_lines.append("=" * 40)
            test_text = "\n".join(test_lines)
            if ptype == "bluetooth":
                threading.Thread(target=self._print_via_escpos,
                                 args=(test_text, ptype, addr),
                                 daemon=True).start()
            elif ptype in ("usb", "network") and ESCPOS_AVAILABLE:
                threading.Thread(target=self._print_via_escpos,
                                 args=(test_text, ptype, addr),
                                 daemon=True).start()
            else:
                self._print_to_file(test_text)
            messagebox.showinfo("Test Print", "Test print dikirim. Periksa printer / folder receipts/.")

        def _cari_printer_bt():
            def worker():
                try:
                    mac = _ble_find_printer()
                    self.after(0, lambda m=mac: _hasil_cari(m))
                except Exception as e:
                    self.after(0, lambda msg=str(e): _hasil_cari("", msg))

            def _hasil_cari(mac, err=""):
                if mac:
                    pr_addr_entry.delete(0, "end")
                    pr_addr_entry.insert(0, mac)
                    pr_type_var.set("bluetooth")
                    messagebox.showinfo("Printer Ditemukan",
                                        f"Printer Bluetooth ditemukan:\n{mac}\n\n"
                                        "Klik Simpan lalu Test Print.", parent=printer_card)
                else:
                    messagebox.showwarning("Tidak Ditemukan",
                                           "Tidak ada printer Bluetooth terdeteksi.\n"
                                           "Pastikan printer sudah ter-pair dengan PC ini."
                                           + (f"\n\n({err})" if err else ""),
                                           parent=printer_card)

            threading.Thread(target=worker, daemon=True).start()

        pr_row3 = ctk.CTkFrame(printer_card, fg_color="transparent")
        pr_row3.pack(fill="x", padx=20, pady=(0, 16))
        ctk.CTkButton(pr_row3, text="💾 Simpan", height=32, width=100,
                      fg_color=C_ACCENT2, font=("Russo One", 9, "bold"),
                      command=_save_printer_settings).pack(side="left", padx=4)
        ctk.CTkButton(pr_row3, text="🔍 Cari Printer BT", height=32, width=150,
                      fg_color=C_BTN, hover_color=C_ACCENT2,
                      border_width=1, border_color=C_ACCENT2,
                      font=("Russo One", 9, "bold"), text_color=C_ACCENT2,
                      command=_cari_printer_bt).pack(side="left", padx=4)
        ctk.CTkButton(pr_row3, text="🖨 Test Print", height=32, width=120,
                      fg_color=C_BTN, hover_color=C_ACCENT2,
                      border_width=1, border_color=C_ACCENT2,
                      font=("Russo One", 9, "bold"), text_color=C_ACCENT2,
                      command=_test_print).pack(side="left", padx=4)

        if self.current_role == "admin":
            user_card = ctk.CTkFrame(scroll, fg_color=C_PANEL, corner_radius=14)
            user_card.pack(fill="x", pady=(0, 16))

            ctk.CTkLabel(user_card, text="Ganti Password Akun Saya:",
                         font=FONT_LABEL, text_color=C_MUTED).pack(anchor="w", padx=20, pady=(16, 4))
            ctk.CTkLabel(user_card, text="Username (dibuat otomatis saat login — tidak bisa diubah):",
                         font=FONT_SMALL, text_color=C_MUTED).pack(anchor="w", padx=20, pady=(0, 2))
            self.entry_username_fix = ctk.CTkEntry(user_card, fg_color=C_BTN, text_color=C_ACCENT,
                                                    border_color=C_BORDER, font=FONT_BODY,
                                                    height=34, width=240)
            self.entry_username_fix.insert(0, self.current_user or "")
            self.entry_username_fix.configure(state="readonly")
            self.entry_username_fix.pack(anchor="w", padx=20, pady=(0, 8))
            self.entry_new_pass = ctk.CTkEntry(user_card, placeholder_text="Password baru",
                                                fg_color=C_BTN, text_color=C_ACCENT,
                                                border_color=C_BORDER, font=FONT_BODY,
                                                height=34, show="●", width=240)
            self.entry_new_pass.pack(anchor="w", padx=20, pady=(0, 8))
            ctk.CTkButton(user_card, text="🔒 Simpan Password Baru", width=200, height=34,
                          fg_color=C_ACCENT2, font=FONT_SUB, text_color="white",
                          command=self._ganti_password).pack(anchor="w", padx=20, pady=(0, 16))

    def _ganti_password(self):
        new_pass = self.entry_new_pass.get().strip()
        if len(new_pass) < 6:
            messagebox.showwarning("⚠ Terlalu Pendek", "Password minimal 6 karakter.")
            return
        user = self.current_user or ""
        if not user:
            messagebox.showerror("✖ Error", "Tidak ada akun yang sedang login.")
            return
        # Atomic load → mutate → save dalam satu lock. Jika akun (mis. dari login
        # Google) belum ada di 'users', dibuat otomatis (role admin).
        def mutator(cfg):
            users = cfg.get("users")
            if not isinstance(users, dict):
                users = {}
            if user not in users:
                users[user] = {"password_enc": hash_password(new_pass),
                               "role": "admin",
                               "email": getattr(self, "current_user_email", "") or ""}
            else:
                users[user]["password_enc"] = hash_password(new_pass)
            cfg["users"] = users
            return cfg
        try:
            ConfigManager.update(mutator)
        except Exception as e:
            AuditLogger.log(
                action="password_change",
                username=user,
                status="failed",
                details={"reason": f"save_error: {e}"}
            )
            messagebox.showerror("✖ Error",
                                 f"Gagal menyimpan password:\n{e}\n\n"
                                 "Pastikan aplikasi punya izin tulis di folder config.")
            return
        AuditLogger.log(
            action="password_change",
            username=user,
            status="success",
            details={"initiated_by": "self"}
        )
        messagebox.showinfo("✅ Berhasil", "Password berhasil diubah!")

    # ── BACKUP / RESTORE ──────────────────────────────────────────────────
    def _export_backup(self):
        try:
            cfg = ConfigManager.load()
            export_cfg = {k: v for k, v in cfg.items() if k != 'users'}
            export_cfg['users_safe'] = {u: {kk: vv for kk, vv in d.items() if kk != 'password'} for u, d in cfg.get('users', {}).items()}
            riwayat_data = {"riwayat_transaksi": [list(r) for r in self.riwayat_transaksi], "riwayat_meta": self.riwayat_meta}
            audit_logs = []
            if os.path.exists(AUDIT_FILE):
                with open(AUDIT_FILE, "r", encoding="utf-8") as f:
                    for line in f:
                        try:
                            audit_logs.append(json.loads(line))
                        except Exception:
                            pass
            backup = {
                "app_version": APP_VERSION,
                "exported_at": datetime.now().isoformat(),
                "exported_by": self.current_user or "unknown",
                "config": export_cfg,
                "riwayat": riwayat_data,
                "audit_log": audit_logs[-5000:] if len(audit_logs) > 5000 else audit_logs,
            }
            default_name = f"rr_billing_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            save_path = filedialog.asksaveasfilename(parent=self, title="Export Backup",
                                                      defaultextension=".json",
                                                      filetypes=[("JSON Backup", "*.json")],
                                                      initialfile=default_name)
            if not save_path:
                return
            with open(save_path, "w", encoding="utf-8") as f:
                json.dump(backup, f, indent=2, ensure_ascii=False)
            AuditLogger.log(action="backup_export", username=self.current_user or "", status="success")
            if hasattr(self, 'lbl_backup_status'):
                self.lbl_backup_status.configure(text=f"✅ Backup berhasil: {os.path.basename(save_path)}")
            messagebox.showinfo("✅ Export Berhasil", f"Backup berhasil disimpan:\n{save_path}", parent=self)
        except Exception as e:
            messagebox.showerror("❌ Export Gagal", f"Gagal export backup:\n{str(e)}", parent=self)

    def _import_backup(self):
        file_path = filedialog.askopenfilename(parent=self, title="Import Backup",
                                                filetypes=[("JSON Backup", "*.json")])
        if not file_path:
            return
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                backup = json.load(f)
        except Exception as e:
            messagebox.showerror("❌ Load Gagal", f"Gagal membaca file backup:\n{str(e)}", parent=self)
            return
        # Show summary
        config_keys = list(backup.get("config", {}).keys())
        riwayat_count = len(backup.get("riwayat", {}).get("riwayat_transaksi", []))
        audit_count = len(backup.get("audit_log", []))
        ver = backup.get("app_version", "?")
        exported = backup.get("exported_at", "?")
        summary = (
            f"Versi Aplikasi: {ver}\n"
            f"Diekspor: {exported}\n"
            f"Data Config: {len(config_keys)} kategori\n"
            f"Data Riwayat: {riwayat_count} transaksi\n"
            f"Data Audit Log: {audit_count} entri\n\n"
            f"⚠ Catatan: User existing TIDAK akan ditimpa.\n"
            f"Config riwayat & audit akan ditambahkan."
        )
        if not messagebox.askyesno("📥 Import Backup", f"Yakin ingin mengimpor backup?\n\n{summary}", parent=self):
            return
        # Ask about overwriting users
        overwrite_users = False
        if backup.get("config", {}).get("users_safe"):
            ow = messagebox.askyesno("Import Users",
                                      "Ikut mengimpor data users (tanpa password)?\n'Ya' untuk timpa, 'Tidak' untuk lewati.",
                                      parent=self)
            overwrite_users = ow
        try:
            cfg = ConfigManager.load()
            imported_cfg = backup.get("config", {})
            for k, v in imported_cfg.items():
                if k == 'users_safe':
                    continue
                if k == 'users' and not overwrite_users:
                    continue
                cfg[k] = v
            ConfigManager.save(cfg)
            # Riwayat
            imported_rows = backup.get("riwayat", {}).get("riwayat_transaksi", [])
            imported_metas = backup.get("riwayat", {}).get("riwayat_meta", [])
            if imported_rows:
                for r in imported_rows:
                    self.riwayat_transaksi.append(tuple(r))
                for m in imported_metas:
                    self.riwayat_meta.append(m)
                self._save_riwayat()
                self._refresh_riwayat_summary()
            # Audit log append
            imported_audit = backup.get("audit_log", [])
            if imported_audit:
                for entry in imported_audit:
                    AuditLogger._append_line(json.dumps(entry, ensure_ascii=False))
            AuditLogger.log(action="backup_import", username=self.current_user or "", status="success",
                            details={"riwayat_count": len(imported_rows), "audit_count": len(imported_audit)})
            if hasattr(self, 'lbl_backup_status'):
                self.lbl_backup_status.configure(text=f"✅ Import berhasil dari: {os.path.basename(file_path)}")
            messagebox.showinfo("✅ Import Berhasil",
                                f"Backup berhasil diimpor.\nRiwayat: {len(imported_rows)} transaksi\nAudit: {len(imported_audit)} entri",
                                parent=self)
        except Exception as e:
            messagebox.showerror("❌ Import Gagal", f"Gagal import backup:\n{str(e)}", parent=self)

    def _reset_all_data(self):
        # Triple confirmation
        if not messagebox.askyesno("🔄 Reset Data — Peringatan!",
                                    "Ini akan MENGHAPUS SEMUA data:\n"
                                    "• Seluruh riwayat transaksi\n"
                                    "• Semua log audit\n"
                                    "• Data yang sedang ditampilkan\n\n"
                                    "Lanjutkan?",
                                    parent=self):
            return
        if not messagebox.askyesno("🔄 Konfirmasi Kedua",
                                    "Data akan dihapus PERMANEN.\n"
                                    "Tidak bisa dikembalikan.\n\n"
                                    "Yakin ingin melanjutkan?",
                                    parent=self):
            return
        # Third: type confirmation
        dlg = ctk.CTkToplevel(self)
        dlg.title("Konfirmasi Akhir")
        dlg.geometry("420x200")
        dlg.transient(self)
        dlg.grab_set()
        ctk.CTkLabel(dlg, text="Ketik 'RESET' untuk konfirmasi:", font=FONT_LABEL, text_color=C_RED).pack(pady=(20, 8))
        entry_confirm = ctk.CTkEntry(dlg, fg_color=C_BTN, text_color=C_TEXT)
        entry_confirm.pack(padx=20, fill="x")
        status_lbl = ctk.CTkLabel(dlg, text="", font=FONT_SMALL, text_color=C_RED)
        status_lbl.pack(pady=8)

        def do_reset():
            if entry_confirm.get().strip() != "RESET":
                status_lbl.configure(text="✖ Ketik 'RESET' dengan benar.")
                return
            # Clear riwayat
            self.riwayat_transaksi.clear()
            self.riwayat_meta.clear()
            self._tree_item_to_index.clear()
            for item in self.tree.get_children():
                self.tree.delete(item)
            self._save_riwayat()
            self._refresh_riwayat_summary()
            # Clear audit log
            try:
                if os.path.exists(AUDIT_FILE):
                    with open(AUDIT_FILE, "w", encoding="utf-8") as f:
                        f.write("")
                        f.flush()
            except Exception:
                pass
            AuditLogger.log(action="data_reset", username=self.current_user or "", status="success",
                            details={"initiated_by": self.current_user})
            dlg.destroy()
            if hasattr(self, 'lbl_backup_status'):
                self.lbl_backup_status.configure(text="✅ Semua data telah direset.")
            messagebox.showinfo("✅ Reset Selesai", "Semua data riwayat dan log telah dihapus.", parent=self)

        btn_f = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_f.pack(fill="x", padx=20, pady=(8, 16))
        ctk.CTkButton(btn_f, text="✖ Batal", width=100, fg_color=C_BTN,
                       command=dlg.destroy).pack(side="left", padx=4)
        ctk.CTkButton(btn_f, text="🔄 Reset Semua", width=140, fg_color=C_RED,
                       command=do_reset).pack(side="right", padx=4)

    # ══════════════════════════════════════════════════════════════════════════
    #  TAB: LOG APLIKASI — Viewer lengkap untuk audit log
    # ══════════════════════════════════════════════════════════════════════════
    def _setup_log_aplikasi(self):
        f = self.frames.get("log_aplikasi")
        if not f:
            return
        for w in f.winfo_children():
            w.destroy()
        
        # Header
        hdr = ctk.CTkFrame(f, fg_color=C_PANEL, height=54, corner_radius=0)
        hdr.pack(fill="x")
        ctk.CTkLabel(hdr, text="📋  LOG APLIKASI LENGKAP", 
                     font=FONT_TITLE, text_color=C_ACCENT).pack(side="left", padx=18, pady=14)
        
        # Filter bar
        filter_frame = ctk.CTkFrame(f, fg_color=C_PANEL, height=44)
        filter_frame.pack(fill="x", padx=0, pady=0)
        filter_frame.pack_propagate(False)
        
        ctk.CTkLabel(filter_frame, text="Filter:", font=FONT_LABEL, 
                     text_color=C_MUTED).pack(side="left", padx=16)
        
        self.log_filter_var = ctk.StringVar(value="all")
        for val, lbl in [("all", "Semua"), ("login", "Login"), ("transaction", "Transaksi"), 
                         ("rental", "Rental"), ("update", "Update"), ("error", "Error")]:
            ctk.CTkRadioButton(filter_frame, text=lbl, variable=self.log_filter_var, 
                              value=val, font=FONT_LABEL, text_color=C_TEXT,
                              command=self._refresh_log_view).pack(side="left", padx=8)
        
        # Main log viewer
        self.log_textbox = ctk.CTkTextbox(f, fg_color=C_BTN, text_color=C_TEXT,
                                          border_color=C_BORDER, border_width=1,
                                          font=("Courier New", 13))
        self.log_textbox.pack(fill="both", expand=True, padx=6, pady=6)
        self.log_textbox.configure(state="disabled")
        
        # Load logs
        self._refresh_log_view()
    
    def _refresh_log_view(self):
        """Refresh dan tampilkan logs berdasarkan filter."""
        self.log_textbox.configure(state="normal")
        self.log_textbox.delete("1.0", "end")
        
        try:
            # Baca audit log
            audit_file = "rr_billing_audit.jsonl"
            if not os.path.exists(audit_file):
                self.log_textbox.insert("end", "❌ File audit log tidak ditemukan.\n")
                self.log_textbox.configure(state="disabled")
                return
            
            logs = []
            filter_type = self.log_filter_var.get()
            
            with open(audit_file, "r", encoding="utf-8") as f:
                for line in f:
                    try:
                        entry = json.loads(line)
                        # Filter berdasarkan action
                        if filter_type == "all":
                            logs.append(entry)
                        elif filter_type == "login" and "login" in entry.get("action", ""):
                            logs.append(entry)
                        elif filter_type == "transaction" and "transaksi" in entry.get("action", ""):
                            logs.append(entry)
                        elif filter_type == "rental" and "rental" in entry.get("action", ""):
                            logs.append(entry)
                        elif filter_type == "update" and "update" in entry.get("action", ""):
                            logs.append(entry)
                        elif filter_type == "error" and entry.get("status") == "failed":
                            logs.append(entry)
                    except json.JSONDecodeError:
                        continue
            
            # Sort by timestamp (terbaru di atas)
            logs.sort(key=lambda x: x.get("timestamp", ""), reverse=True)
            
            # Display logs
            if not logs:
                self.log_textbox.insert("end", f"✓ Tidak ada log untuk filter '{filter_type}'.\n")
            else:
                self.log_textbox.insert("end", f"📊 Total: {len(logs)} entri\n" + "="*80 + "\n\n")
                
                for entry in logs:
                    timestamp = entry.get("timestamp", "N/A")
                    action = entry.get("action", "unknown")
                    username = entry.get("username", "system")
                    status = entry.get("status", "unknown")
                    details = entry.get("details", {})
                    
                    # Color based on status
                    if status == "success":
                        status_icon = "✅"
                        status_color = C_GREEN
                    elif status == "failed":
                        status_icon = "❌"
                        status_color = C_RED
                    else:
                        status_icon = "⚠️ "
                        status_color = C_YELLOW
                    
                    log_line = f"{status_icon} [{timestamp}] {action.upper()}\n"
                    log_line += f"   User: {username} | Status: {status}\n"
                    
                    if details:
                        log_line += f"   Details: {json.dumps(details, ensure_ascii=False)[:120]}\n"
                    
                    log_line += "\n"
                    
                    self.log_textbox.insert("end", log_line)
            
            # Footer
            self.log_textbox.insert("end", "\n" + "="*80 + "\n")
            self.log_textbox.insert("end", f"✓ Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            
        except Exception as e:
            self.log_textbox.insert("end", f"❌ Error loading logs: {str(e)}\n")
        
        finally:
            self.log_textbox.configure(state="disabled")

    # ══════════════════════════════════════════════════════════════════════════
    #  TAB: MANAJEMEN KASIR (Admin-only) — CRUD akun kasir/admin (APTV2-style)
    # ══════════════════════════════════════════════════════════════════════════
    def _setup_users(self):
        f = self.frames.get("users")
        if not f:
            return
        for w in f.winfo_children():
            w.destroy()
        hdr = ctk.CTkFrame(f, fg_color=C_PANEL, height=54, corner_radius=0)
        hdr.pack(fill="x")
        ctk.CTkLabel(hdr, text="👥  MANAJEMEN KASIR (Admin)", font=FONT_TITLE, text_color=C_ACCENT).pack(side="left", padx=18, pady=14)

        content = ctk.CTkScrollableFrame(f, fg_color=C_BG)
        content.pack(fill="both", expand=True, padx=16, pady=12)

        form = ctk.CTkFrame(content, fg_color=C_CARD, corner_radius=12)
        form.pack(fill="x", pady=(0,12))
        ctk.CTkLabel(form, text="Tambah Akun Kasir (Sub-Akun Admin)", font=FONT_SUB, text_color=C_ACCENT2).pack(anchor="w", padx=12, pady=(12,6))
        self.u_username = ctk.CTkEntry(form, placeholder_text="username (a-z0-9_.)", fg_color=C_BTN, text_color=C_TEXT)
        self.u_username.pack(fill="x", padx=12, pady=(0,6))
        self.u_password = ctk.CTkEntry(form, placeholder_text="password (min 6, huruf besar & angka)", show="●", fg_color=C_BTN, text_color=C_TEXT)
        self.u_password.pack(fill="x", padx=12, pady=(0,6))
        self.u_show_pw = ctk.CTkCheckBox(form, text="👁 Lihat Password", fg_color=C_ACCENT2,
                                         hover_color=C_ACCENT, font=FONT_SMALL, text_color=C_TEXT,
                                         command=self._toggle_u_show_pw)
        self.u_show_pw.pack(anchor="w", padx=12, pady=(0,6))
        ctk.CTkLabel(form, text=f"Role: kasir  •  Sub-akun dari: {self.current_user or '—'}\n"
                                "Lisensi akun kasir mengikuti admin utama.",
                     font=FONT_SMALL, text_color=C_MUTED, justify="left").pack(anchor="w", padx=12, pady=(0,6))
        ctk.CTkButton(form, text="➕ Daftarkan Kasir", fg_color=C_ACCENT2, command=self._create_user).pack(padx=12, pady=(0,12))

        list_card = ctk.CTkFrame(content, fg_color=C_PANEL, corner_radius=12)
        list_card.pack(fill="both", expand=True)
        ctk.CTkLabel(list_card, text="Daftar Akun Kasir", font=FONT_SUB, text_color=C_ACCENT2).pack(anchor="w", padx=12, pady=(12,6))

        self.user_list_box = ctk.CTkScrollableFrame(list_card, fg_color="transparent")
        self.user_list_box.pack(fill="both", expand=True, padx=12, pady=(0,12))

        self._refresh_user_list()

    def _toggle_u_show_pw(self):
        if hasattr(self, "u_password"):
            show = bool(getattr(self, "u_show_pw", None) and self.u_show_pw.get())
            self.u_password.configure(show="" if show else "●")

    def _refresh_user_list(self):
        for w in self.user_list_box.winfo_children():
            w.destroy()
        users = ConfigManager.get("users", LoginPage.DEFAULT_USERS)
        for uname, u in users.items():
            role = u.get("role", "kasir") if isinstance(u, dict) else "kasir"
            # Hanya akun kasir yang ditampilkan (akun admin tidak tampil)
            if role != "kasir":
                continue
            dibuat = (u.get("dibuat", "") or "") if isinstance(u, dict) else ""
            admin_utama = (u.get("admin_utama", "") or "") if isinstance(u, dict) else ""
            row = ctk.CTkFrame(self.user_list_box, fg_color=C_CARD, corner_radius=8)
            row.pack(fill="x", pady=6)
            info = ctk.CTkFrame(row, fg_color="transparent")
            info.pack(side="left", fill="x", expand=True, padx=12, pady=6)
            ctk.CTkLabel(info, text=f"{uname}", font=FONT_BODY, text_color=C_TEXT).pack(anchor="w")
            sub = f"Role: {role}"
            if admin_utama:
                sub += f"  •  Sub-akun: {admin_utama}"
            if dibuat:
                sub += f"  •  Dibuat: {dibuat}"
            if uname == self.current_user:
                sub += "  •  Akun Anda"
            ctk.CTkLabel(info, text=sub, font=FONT_SMALL, text_color=C_MUTED).pack(anchor="w")
            if uname != self.current_user:
                ctk.CTkButton(row, text="🔁 Reset PW", fg_color=C_BTN, width=100, command=lambda n=uname: self._reset_user_pw(n)).pack(side="right", padx=4)
                ctk.CTkButton(row, text="🗑 Hapus", fg_color=C_RED, width=90, command=lambda n=uname: self._delete_user(n)).pack(side="right", padx=8)

    def _create_user(self):
        username = self.u_username.get().strip().lower()
        password = self.u_password.get().strip()
        if not is_valid_username(username):
            messagebox.showwarning("⚠ Username tidak valid", "Gunakan 4-20 karakter: huruf kecil, angka, ., _")
            return
        if not is_valid_kasir_password(password):
            messagebox.showwarning("⚠ Password tidak valid", "Password kasir: minimal 6 karakter, wajib huruf besar & angka")
            return
        cfg = ConfigManager.load()
        users = cfg.get("users", dict(LoginPage.DEFAULT_USERS))
        if username in users:
            messagebox.showwarning("⚠ Sudah ada", "Username sudah dipakai")
            return
        users[username] = {
            "password_enc": hash_password(password),
            "role": "kasir",
            "admin_utama": self.current_user or "",
            "dibuat": datetime.now().strftime("%Y-%m-%d %H:%M"),
        }
        cfg["users"] = users
        ConfigManager.save(cfg)
        AuditLogger.log(action="user_created", username=username, status="success",
                        details={"created_by": self.current_user, "role": "kasir",
                                 "admin_utama": self.current_user or ""})
        messagebox.showinfo("✅ Berhasil", f"Akun kasir '{username}' dibuat sebagai sub-akun dari '{self.current_user}'.")
        self.u_username.delete(0, 'end')
        self.u_password.delete(0, 'end')
        self._refresh_user_list()

    def _delete_user(self, username):
        if not messagebox.askyesno("Hapus User", f"Hapus user '{username}'? Akun tidak bisa login lagi.\n\nIni tidak bisa dibatalkan."):
            return
        cfg = ConfigManager.load()
        users = cfg.get("users", {})
        if username in users:
            users.pop(username)
            cfg["users"] = users
            ConfigManager.save(cfg)
            AuditLogger.log(action="user_deleted", username=username, status="success", details={"deleted_by": self.current_user})
            self._refresh_user_list()

    def _reset_user_pw(self, username):
        # Admin memasukkan password baru secara langsung (pendekatan APTV2),
        # bukan password temporary fix.
        cfg = ConfigManager.load()
        users = cfg.get("users", {})
        if username not in users:
            messagebox.showwarning("⚠ Tidak ada", f"User '{username}' tidak ditemukan.")
            return
        role = users[username].get("role", "kasir")

        dlg = ctk.CTkToplevel(self.winfo_toplevel())
        dlg.title("🔁 Reset Password")
        dlg.geometry("440x330")
        dlg.configure(fg_color=C_BG)
        dlg.transient(self.winfo_toplevel())
        dlg.resizable(False, False)
        dlg.grab_set()

        ctk.CTkLabel(dlg, text=f"RESET PASSWORD — {username}",
                     font=("Russo One", 14, "bold"), text_color=C_YELLOW).pack(pady=(16, 2))
        ctk.CTkLabel(dlg, text=f"Role: {role}",
                     font=FONT_SMALL, text_color=C_MUTED).pack(pady=(0, 10))

        entry = ctk.CTkEntry(dlg, placeholder_text="Password baru", show="●",
                             fg_color=C_BTN, text_color=C_TEXT, font=("Consolas", 12),
                             height=36, width=320, justify="center")
        entry.pack(pady=(0, 4), padx=30)

        show_pw_var = ctk.CTkCheckBox(dlg, text="👁 Lihat Password", fg_color=C_ACCENT2,
                                      hover_color=C_ACCENT, font=FONT_SMALL, text_color=C_TEXT,
                                      command=lambda: entry.configure(
                                          show="" if show_pw_var.get() else "●"))
        show_pw_var.pack(pady=(0, 4))

        lbl = ctk.CTkLabel(dlg, text="", font=("Consolas", 10), text_color=C_RED)
        lbl.pack(pady=(0, 6))

        def _save():
            pw = entry.get().strip()
            valid = is_valid_kasir_password(pw) if role == "kasir" else is_valid_password(pw)
            hint = ("Kasir: min 6 karakter, huruf besar & angka"
                    if role == "kasir" else "Admin: min 8 karakter, huruf & angka")
            if not pw or not valid:
                lbl.configure(text=hint)
                return
            cfg2 = ConfigManager.load()
            users2 = cfg2.get("users", {})
            if username in users2:
                users2[username]["password_enc"] = hash_password(pw)
                cfg2["users"] = users2
                ConfigManager.save(cfg2)
                AuditLogger.log(action="user_pw_reset", username=username, status="success",
                                details={"reset_by": self.current_user})
                dlg.destroy()
                messagebox.showinfo("✅ Reset Pw", f"Password akun '{username}' berhasil diganti.")

        def _cancel():
            dlg.destroy()

        ctk.CTkButton(dlg, text="💾 SIMPAN", fg_color=C_YELLOW, hover_color=C_ACCENT,
                      font=("Russo One", 12, "bold"), command=_save).pack(pady=(8, 4), padx=30, fill="x")
        ctk.CTkButton(dlg, text="Batal", fg_color="transparent", hover_color=C_BTN,
                      border_width=1, border_color=C_BORDER, font=("Russo One", 11), text_color=C_MUTED,
                      command=_cancel).pack(pady=(0, 12), padx=30, fill="x")

# ═══════════════════════════════════════════════════════════════════════════════
# ── LOG APLIKASI (app.log — selalu ada di setiap run) ─────────────────────────
_APP_LOG_HANDLER = None  # logging.handlers.RotatingFileHandler


def app_log_dir() -> str:
    """Folder log aplikasi: folder exe bila writable; fallback %LOCALAPPDATA%
    (folder exe terkunci, mis. Program Files) supaya log tetap tercipta."""
    base = APP_BASE_DIR
    try:
        probe = os.path.join(base, ".log_write_test")
        with open(probe, "a", encoding="utf-8"):
            pass
        os.remove(probe)
        return base
    except Exception:
        fallback = os.path.join(os.environ.get("LOCALAPPDATA", ""), "RRBillingPro")
        try:
            os.makedirs(fallback, exist_ok=True)
            return fallback
        except Exception:
            return base


class _TeeOut:
    """Redirect stdout/stderr: tulis ke stream asli (console dev) DAN ke
    app.log — semua print() aplikasi terekam walau EXE tanpa window."""

    def __init__(self, stream, level=logging.INFO):
        self._stream = stream
        self._level = level

    def write(self, data):
        try:
            if _APP_LOG_HANDLER is not None and data and data.strip() != "":
                _APP_LOG_HANDLER.emit(logging.LogRecord(
                    "app.out", self._level, "", 0, str(data).rstrip("\n"),
                    None, None))
        except Exception:
            pass
        try:
            self._stream.write(data)
            self._stream.flush()
        except Exception:
            pass

    def flush(self):
        try:
            self._stream.flush()
        except Exception:
            pass

    def isatty(self):
        return False

    @property
    def buffer(self):
        return getattr(self._stream, "buffer", None)

    @staticmethod
    def install():
        if not getattr(_TeeOut, "_saved", None):
            _TeeOut._saved = (sys.stdout, sys.stderr)
        sys.stdout = _TeeOut(_TeeOut._saved[0], logging.INFO)
        sys.stderr = _TeeOut(_TeeOut._saved[1], logging.WARNING)


def setup_app_logging() -> str:
    """Init log aplikasi: RotatingFileHandler app.log (2MB x 3, utf-8) di
    root logger + redirect stdout/stderr. Return path file log (selalu ada)."""
    global _APP_LOG_HANDLER
    try:
        from logging.handlers import RotatingFileHandler
        log_dir = app_log_dir()
        path = os.path.join(log_dir, "app.log")
        _APP_LOG_HANDLER = RotatingFileHandler(
            path, maxBytes=2 * 1024 * 1024, backupCount=3, encoding="utf-8")
        _APP_LOG_HANDLER.setFormatter(logging.Formatter(
            "%(asctime)s %(levelname)s %(message)s"))
        logging.getLogger().addHandler(_APP_LOG_HANDLER)
        _LOGGER.setLevel(logging.INFO)
        _LOGGER.info("=== RRBILLINGPRO v%s — mulai %s (log: %s) ===",
                     APP_VERSION, datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                     path)
        _TeeOut.install()
        return path
    except Exception as e:
        print(f"[LOG] setup gagal: {e}")
        return ""


# ═══════════════════════════════════════════════════════════════════════════════
def _dump_exc(exc_type, exc_value, exc_tb):
    """logcat: semua traceback (termasuk callback Tk) ditulis ke app_exceptions.log
    + ikut terekam di app.log."""
    try:
        import traceback as _tb
        with open(os.path.join(app_log_dir(), "app_exceptions.log"),
                  "a", encoding="utf-8") as _f:
            _f.write(f"\n===== {datetime.now()} {exc_type.__name__}: {exc_value} =====\n")
            _tb.print_exception(exc_type, exc_value, exc_tb, file=_f)
    except Exception:
        pass
    try:
        logging.getLogger("app.exc").error(
            "EXCEPTION %s: %s", exc_type.__name__, exc_value,
            exc_info=(exc_type, exc_value, exc_tb))
    except Exception:
        pass


sys.excepthook = _dump_exc


if __name__ == "__main__":
    setup_app_logging()
    app = AutoRentApp()
    app.report_callback_exception = lambda exc, val, tb: _dump_exc(exc, val, tb)
    app.mainloop()
